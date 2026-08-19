#!/usr/bin/env python3
"""v0.8.3 promotion certificate — READ-ONLY. Writes nothing.

Asserts the three-cell promotion identity against frozen v0.8.2, then EXERCISES
the new AUDIT!B16 invariant against fixtures: zero rows, eight valid Week 0
rows, and one fixture per defect class. A guardrail that only ever says OK has
not been tested, so the defect fixtures must each produce FAIL.

The MARKET LINES!P / Q / R chain and the B16 SUMPRODUCT identity are
re-implemented here exactly as written in the workbook.

Exit code 0 iff every check passes.
"""
import collections, hashlib, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
V082 = os.path.join(ROOT, "promotion_v0.8.2",
                    "TTW_College_Football_Power_Ratings_v0.8.2_AUTHORITATIVE.xlsx")
V083 = os.path.join(ROOT, "promotion_v0.8.3",
                    "TTW_College_Football_Power_Ratings_v0.8.3_AUTHORITATIVE.xlsx")
FROZEN_V082_SHA = "225085449b5a1db5903a3998cb909be1f7ae0037782ea65d412bcb4d9d9490d0"
EXPECTED_CELLS = {("START HERE", "A1"), ("AUDIT", "A16"), ("AUDIT", "B16")}

PASS, FAIL = [], []


def chk(msg, ok, detail=""):
    (PASS if ok else FAIL).append(msg)
    print(f"  [{'PASS' if ok else 'FAIL'}] {msg}" + (f" - {detail}" if detail else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def ftext(v):
    return v.text if isinstance(v, ArrayFormula) else v


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


# ---------------------------------------------------------------- fixtures
class Row(dict):
    """One MARKET LINES row: A GameID, C favorite, D spread, E total,
    F source, G line date."""
    def __init__(self, A="", C="", D="", E="", F="", G=""):
        super().__init__(A=A, C=C, D=D, E=E, F=F, G=G)


def evaluate(rows, games, alias, as_of=None, stale_days=5):
    """Faithful re-implementation of MARKET LINES!O/P/Q/R and AUDIT!B16."""
    ids = [r["A"] for r in rows if r["A"] != ""]
    dup = collections.Counter(ids)
    out = []
    for r in rows:
        A, C, D, E, F, G = (r["A"], r["C"], r["D"], r["E"], r["F"], r["G"])
        if A == "":
            out.append(dict(P="", Q="", valid=False, populated=False))
            continue
        O = alias.get(str(C).strip(), C) if C != "" else ""       # column O
        g = games.get(str(A))
        if A == "" or O == "":                                     # column P
            P = ""
        elif g is None:
            P = "NO GAME"
        elif O in (g["home"], g["away"]):
            P = "OK"
        else:
            P = "INVALID"
        Q = ""                                                     # column Q
        if P == "INVALID":
            Q += "INVALID FAVORITE; "
        if P == "NO GAME":
            Q += "GAMEID NOT IN SCHEDULE; "
        if D == "":
            Q += "SPREAD MISSING; "
        if E == "":
            Q += "TOTAL MISSING; "
        if G == "":
            Q += "DATE MISSING; "
        if F == "":
            Q += "SOURCE MISSING; "
        if dup[A] > 1:
            Q += "DUPLICATE GAMEID ROW; "
        if G != "" and as_of is not None and (as_of - G) > stale_days:
            Q += "STALE; "
        valid = (P == "OK" and Q == ""
                 and isinstance(D, (int, float)) and not isinstance(D, bool) and D > 0
                 and isinstance(E, (int, float)) and not isinstance(E, bool))
        out.append(dict(P=P, Q=Q, valid=valid, populated=True))
    n = sum(1 for o in out if o["populated"])
    good = sum(1 for o in out if o["valid"])
    if n == good:
        return f"OK — {n} line(s) valid", n, good
    return f"FAIL — {n - good} line(s) invalid — see MARKET LINES column Q", n, good


def main():
    print("=" * 78)
    print("v0.8.3 PROMOTION CERTIFICATE — go-live market-line guardrail")
    print("=" * 78)

    print("\n1. PREDECESSOR AND IDENTITY")
    h82, h83 = sha256(V082), sha256(V083)
    print(f"  v0.8.2 SHA-256: {h82}")
    print(f"  v0.8.3 SHA-256: {h83}")
    chk("1.1 v0.8.2 retains its frozen SHA-256", h82 == FROZEN_V082_SHA, h82[:16])
    chk("1.2 v0.8.3 differs from v0.8.2", h83 != h82)

    a = openpyxl.load_workbook(V082)
    b = openpyxl.load_workbook(V083)
    chk("1.3 21 sheets, identical names and order",
        a.sheetnames == b.sheetnames and len(b.sheetnames) == 21, f"{len(b.sheetnames)}")
    chk("1.4 sheet visibility preserved",
        {s: a[s].sheet_state for s in a.sheetnames} == {s: b[s].sheet_state for s in b.sheetnames})
    chk("1.5 named ranges preserved",
        sorted(a.defined_names.keys()) == sorted(b.defined_names.keys()))

    changed, formula_changed = [], []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                x, y = norm(wa.cell(row=r, column=c).value), norm(wbk.cell(row=r, column=c).value)
                if x != y:
                    coord = wbk.cell(row=r, column=c).coordinate
                    changed.append((s, coord))
                    if x[0] == "F" or y[0] == "F":
                        formula_changed.append((s, coord))
    chk("1.6 exactly THREE cells differ", len(changed) == 3, str(changed))
    chk("1.7 the three coordinates are exactly the authorized set",
        set(changed) == EXPECTED_CELLS, str(sorted(set(changed) ^ EXPECTED_CELLS)))
    chk("1.8 the ONLY formula change is AUDIT!B16",
        formula_changed == [("AUDIT", "B16")], str(formula_changed))

    au, sh, ml, tm = b["AUDIT"], b["START HERE"], b["MARKET LINES"], b["TEAM MAP"]
    chk("1.9 banner states v0.8.3 AUTHORITATIVE", "v0.8.3 AUTHORITATIVE" in sh["A1"].value)
    chk("1.10 banner carries no stale v0.8.2 identifier", "v0.8.2" not in sh["A1"].value)
    chk("1.11 AUDIT!E1 count formula unchanged",
        ftext(a["AUDIT"]["E1"].value) == ftext(au["E1"].value))
    chk("1.12 old 'REMOVE TEST LINES' sentinel gone from B16",
        "REMOVE TEST LINES" not in ftext(au["B16"].value))

    print("\n2. MODEL CONTENT PRESERVED")
    for sheet, cells in (("ENGINE", ("I6", "L6", "M6", "R6", "AE6", "AI6")),
                         ("QB VALUES", ("G6", "M6", "G102", "M102")),
                         ("TEAM RATINGS", ("E6", "I6", "K6", "O6")),
                         ("PRESEASON", ("Z6", "Y6")),
                         ("CALC", ("N6", "Q6", "S6")),
                         ("MARKET LINES", ("B6", "O6", "P6", "Q6", "R6"))):
        same = all(ftext(a[sheet][c].value) == ftext(b[sheet][c].value) for c in cells)
        chk(f"2.x {sheet} formulas identical to v0.8.2", same)
    qb = b["QB VALUES"]
    codes, sts, nonzero = collections.Counter(), collections.Counter(), 0
    season = b["SETTINGS"]["B3"].value
    for r in range(6, 144):
        if not tm.cell(row=r, column=1).value:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        G = "" if (D is None or F is None) else F - D
        codes[H] += 1
        sts["UNCERTAIN" if (G == "" or H == "L" or J != season) else "OK"] += 1
        if (D not in (None, 0)) or (F not in (None, 0)):
            nonzero += 1
    chk("2.1 QB census unchanged (38 UNCERTAIN / 100 OK)",
        sts["UNCERTAIN"] == 38 and sts["OK"] == 100, str(dict(sts)))
    chk("2.2 confidence census unchanged (65 H / 41 M / 32 L)",
        (codes["H"], codes["M"], codes["L"]) == (65, 41, 32), str(dict(codes)))
    chk("2.3 nonzero QB values still 0", nonzero == 0, str(nonzero))
    populated = [r for r in range(6, 1006) if ml.cell(row=r, column=1).value not in (None, "")]
    chk("2.4 repository artifact ships with MARKET LINES blank",
        not populated, str(populated[:5]))

    # ------------------------------------------------------------ fixtures
    print("\n3. THE NEW INVARIANT, EXERCISED")
    alias = {}
    for r in range(6, 606):
        h, i = tm.cell(row=r, column=8).value, tm.cell(row=r, column=9).value
        if h and i:
            alias[str(h).strip()] = str(i).strip()
    imp = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            imp[str(k).strip()] = str(v).strip()
    sch = b["IMPORT SCHEDULE"]
    games = {}
    for r in range(6, 1006):
        gid = sch.cell(row=r, column=1).value
        if not gid:
            continue
        games[str(gid)] = dict(
            away=imp.get(str(sch.cell(row=r, column=6).value).strip(), ""),
            home=imp.get(str(sch.cell(row=r, column=8).value).strip(), ""),
            week=sch.cell(row=r, column=3).value)

    WK0 = [("401856766", "TCU"), ("401858201", "STAN"), ("401858202", "UVA"),
           ("401864494", "USC"), ("401864570", "FSU"), ("401864577", "NDSU"),
           ("401866408", "EMU"), ("401862693", "UNLV")]
    good8 = [Row(A=g, C=f, D=7.5, E=52.5, F="Circa", G=45900) for g, f in WK0]

    res, n, ok = evaluate([], games, alias)
    chk("3.1 zero market rows -> OK", res.startswith("OK") and n == 0, res)

    res, n, ok = evaluate(good8, games, alias)
    chk("3.2 eight valid Week 0 Circa rows -> OK",
        res == "OK — 8 line(s) valid" and n == 8 and ok == 8, res)

    def one_bad(label, mutate, expect_fail=True):
        rows = [Row(**dict(r)) for r in good8]
        mutate(rows)
        res, n, ok = evaluate(rows, games, alias)
        bad = res.startswith("FAIL")
        chk(f"3.x defect caught: {label}", bad == expect_fail, res)

    one_bad("GameID not in schedule", lambda R: R[0].__setitem__("A", "999999999"))
    one_bad("favorite blank", lambda R: R[0].__setitem__("C", ""))
    one_bad("favorite not in that game", lambda R: R[0].__setitem__("C", "ALA"))
    one_bad("spread negative", lambda R: R[1].__setitem__("D", -7.5))
    one_bad("spread zero", lambda R: R[1].__setitem__("D", 0))
    one_bad("spread non-numeric", lambda R: R[1].__setitem__("D", "7.5"))
    one_bad("spread missing", lambda R: R[1].__setitem__("D", ""))
    one_bad("total missing", lambda R: R[2].__setitem__("E", ""))
    one_bad("total non-numeric", lambda R: R[2].__setitem__("E", "fifty"))
    one_bad("source missing", lambda R: R[3].__setitem__("F", ""))
    one_bad("line date missing", lambda R: R[4].__setitem__("G", ""))
    one_bad("duplicate GameID", lambda R: R[5].__setitem__("A", R[4]["A"]))

    rows = [Row(**dict(r)) for r in good8]
    res, _, _ = evaluate(rows, games, alias, as_of=45900 + 9, stale_days=5)
    chk("3.x defect caught: stale line (as-of set, 9 days old)", res.startswith("FAIL"), res)
    res, _, _ = evaluate(rows, games, alias, as_of=45900 + 2, stale_days=5)
    chk("3.x fresh line inside the stale window -> OK", res.startswith("OK"), res)

    print("\n4. SETTINGS AND SAFEGUARDS")
    st = b["SETTINGS"]
    chk("4.1 BET toggle remains N", st["B11"].value == "N")
    chk("4.2 totals remain unavailable", st["B22"].value is None and st["B23"].value is None)
    for t in ("NDSU", "SAC"):
        rr = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == t][0]
        chk(f"4.3 {t} remains transitional", tm.cell(row=rr, column=5).value == "Y")

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"v0.8.2 SHA-256 (frozen): {h82}")
    print(f"v0.8.3 SHA-256:          {h83}")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

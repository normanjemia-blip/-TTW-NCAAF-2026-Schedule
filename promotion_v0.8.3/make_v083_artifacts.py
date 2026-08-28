#!/usr/bin/env python3
"""Generate the v0.8.3 promotion artifacts EXPLICITLY.

Writes diff_v082_to_v083.csv and regression_log_v083.txt. verify_v083.py is
read-only by design and never produces these as a side effect.

Run:  python3 promotion_v0.8.3/make_v083_artifacts.py
"""
import csv, os, subprocess, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
V082 = os.path.join(ROOT, "promotion_v0.8.2",
                    "TTW_College_Football_Power_Ratings_v0.8.2_AUTHORITATIVE.xlsx")
V083 = os.path.join(ROOT, "promotion_v0.8.3",
                    "TTW_College_Football_Power_Ratings_v0.8.3_AUTHORITATIVE.xlsx")
CSV_OUT = os.path.join(ROOT, "promotion_v0.8.3", "diff_v082_to_v083.csv")
LOG_OUT = os.path.join(ROOT, "promotion_v0.8.3", "regression_log_v083.txt")

CLASSIFY = {
    ("START HERE", "A1"): "DOCUMENTATION — version identifier only",
    ("AUDIT", "A16"): "AUDIT LABEL — invariant renamed to its operational meaning",
    ("AUDIT", "B16"): "AUDIT FORMULA — blank-market check replaced by market-line validity check",
}


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def main():
    a = openpyxl.load_workbook(V082)
    b = openpyxl.load_workbook(V083)
    changed = []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                x, y = norm(wa.cell(row=r, column=c).value), norm(wbk.cell(row=r, column=c).value)
                if x != y:
                    changed.append((s, wbk.cell(row=r, column=c).coordinate, x[1], y[1]))
    assert len(changed) == 3, f"expected 3 changed cells, found {len(changed)}"
    with open(CSV_OUT, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["sheet", "cell", "classification", "old", "new"])
        for s, coord, o, n in changed:
            w.writerow([s, coord, CLASSIFY.get((s, coord), "UNCLASSIFIED — investigate"),
                        "" if o is None else o, "" if n is None else n])
    print(f"wrote {CSV_OUT} ({len(changed)} rows)")

    r = subprocess.run([sys.executable, os.path.join(ROOT, "promotion_v0.8.3", "verify_v083.py")],
                       capture_output=True, text=True)
    with open(LOG_OUT, "w") as fh:
        fh.write(r.stdout)
        if r.stderr:
            fh.write("\nSTDERR\n" + r.stderr)
        fh.write(f"\nverify_v083.py exit code: {r.returncode}\n")
    print(f"wrote {LOG_OUT} (verifier exit {r.returncode})")
    return 0 if r.returncode == 0 else 1


if __name__ == "__main__":
    sys.exit(main())

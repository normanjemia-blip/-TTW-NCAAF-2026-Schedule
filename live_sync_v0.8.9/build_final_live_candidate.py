#!/usr/bin/env python3
"""Build the FINAL LIVE v0.8.9 CANDIDATE from the FRESH live export + authoritative v0.8.9.

BASE   : the live Google Sheet re-exported 2026-08-28 (never written back)
SOURCE : promotion_v0.8.9/...v0.8.9_AUTHORITATIVE.xlsx

The write set is computed DIRECTLY from a full cell diff of the fresh export against
authoritative v0.8.9. The stale 1,269-cell expectation from the previous export is NOT reused -
the live Sheet moved (8 market lines re-priced, SETTINGS!B5 advanced), so any count carried over
would be wrong by construction.

PRESERVED OPERATIONAL OVERLAY (live always wins, never written):
  * every MARKET LINES value cell
  * all live-authored CHANGELOG history
  * SETTINGS!B4 and B5
  * Fresno State QB VALUES I75/K75/L75 - Tulane L91 - Northern Illinois I123/L123

BANNER: authoritative v0.8.9 banner with the market-line count corrected to the number actually
loaded in the fresh export, counted by the workbook's own rule (AUDIT!B16):
    SUMPRODUCT(--('MARKET LINES'!$A$6:$A$1005<>""))   i.e. populated GameID game-rows.

CHANGELOG CONFLICT: authoritative v0.8.9 writes its entry at CHANGELOG row 87, which the live
Sheet already occupies with a live-authored v0.8.3 GO-LIVE entry. Per the owner's safety
correction 5 this entry is NEITHER overwritten NOR relocated NOR appended. It is HELD BACK and
reported for approval. Both entries survive this build untouched.
"""
import hashlib, json, os, shutil, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
AUTH = os.path.join(ROOT, "promotion_v0.8.9",
                    "TTW_College_Football_Power_Ratings_v0.8.9_AUTHORITATIVE.xlsx")
OUT = os.path.join(HERE, "TTW_LIVE_CANDIDATE_v0.8.9.xlsx")
PACKET = os.path.join(HERE, "live_write_packet_v0.8.9.csv")
HELD = os.path.join(HERE, "changelog_conflict_v0.8.9.json")

AUTH_SHA = "334050660deb970f23cd9761490fb47e1f2b606b61d00a20c864cec529395cbb"
FRESH_SHA = "ecab90349c1fd4bbf7419b394bc7062ece52d50a245dfa5b9b27ff73e08cda8d"

PRESERVE_QB = {("QB VALUES", "I75"), ("QB VALUES", "K75"), ("QB VALUES", "L75"),
               ("QB VALUES", "L91"), ("QB VALUES", "I123"), ("QB VALUES", "L123")}
PRESERVE_SETTINGS = {("SETTINGS", "B4"), ("SETTINGS", "B5")}
BANNER_OLD = "0 market lines loaded"


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", (v.text or "").strip())
    if isinstance(v, str) and v.startswith("="):
        return ("F", v.strip())
    return ("V", v)


def disp(v):
    return "" if norm(v)[1] is None else str(norm(v)[1])


def main():
    fresh_path = sys.argv[1]
    ga, gf = sha256(AUTH), sha256(fresh_path)
    assert ga == AUTH_SHA, f"authoritative v0.8.9 is not expected: {ga}"
    assert gf == FRESH_SHA, f"fresh export is not the audited one: {gf}"
    print(f"authoritative v0.8.9 SHA verified : {ga}")
    print(f"fresh live export SHA verified    : {gf}")

    auth = openpyxl.load_workbook(AUTH)
    live = openpyxl.load_workbook(fresh_path)
    assert auth.sheetnames == live.sheetnames, "sheet set/order drift"

    # --- market lines actually loaded, by the workbook's own counting rule (AUDIT!B16) ---
    ml = live["MARKET LINES"]
    loaded = sum(1 for r in range(6, 1006) if ml.cell(row=r, column=1).value not in (None, ""))
    print(f"market lines loaded (populated GameID game-rows) : {loaded}")

    # --- live-authored CHANGELOG rows ---
    lcl = live["CHANGELOG"]
    live_rows = {r for r in range(2, 1006)
                 if any(lcl.cell(row=r, column=c).value not in (None, "") for c in range(1, 7))}

    writes, preserved, held = [], [], []
    for s in auth.sheetnames:
        sa, sl = auth[s], live[s]
        R = max(sa.max_row, sl.max_row)
        C = max(sa.max_column, sl.max_column)
        for r in range(1, R + 1):
            for c in range(1, C + 1):
                av, lv = sa.cell(row=r, column=c).value, sl.cell(row=r, column=c).value
                if norm(av) == norm(lv):
                    continue
                coord = sa.cell(row=r, column=c).coordinate
                key = (s, coord)
                rec = dict(sheet=s, cell=coord, row=r, column=c,
                           live=disp(lv), authoritative=disp(av))

                if s == "MARKET LINES":
                    assert norm(av)[0] != "F" and norm(lv)[0] != "F", \
                        f"unexpected MARKET LINES formula diff at {coord}"
                    preserved.append(dict(rec, why="live market line"))
                elif key in PRESERVE_QB:
                    preserved.append(dict(rec, why="owner-authored QB note"))
                elif key in PRESERVE_SETTINGS:
                    preserved.append(dict(rec, why="live SETTINGS!B4/B5"))
                elif s == "CHANGELOG" and r in live_rows:
                    # live-authored history. If authoritative also wants this row, it is a
                    # CONFLICT: hold the authoritative entry back, never overwrite the live one.
                    if norm(av)[1] not in (None, ""):
                        held.append(dict(rec, why="CONFLICT - authoritative entry held back"))
                    preserved.append(dict(rec, why="live-authored CHANGELOG"))
                elif s == "START HERE" and coord == "A1":
                    b = av
                    assert BANNER_OLD in b, "authoritative banner lacks the market-line clause"
                    nb = b.replace(BANNER_OLD, f"{loaded} market lines loaded")
                    assert nb.replace(f"{loaded} market lines loaded", BANNER_OLD) == b
                    assert "v0.8.9 AUTHORITATIVE" in nb and "2026-08-27" in nb
                    writes.append(dict(rec, authoritative=nb, why="banner"))
                else:
                    writes.append(dict(rec, why="sync to authoritative v0.8.9"))

    # no preserved cell may appear in the write list
    wkeys = {(d["sheet"], d["cell"]) for d in writes}
    pkeys = {(d["sheet"], d["cell"]) for d in preserved}
    assert not (wkeys & pkeys), f"preserved cell in write list: {sorted(wkeys & pkeys)[:5]}"
    assert not (PRESERVE_QB & wkeys) and not (PRESERVE_SETTINGS & wkeys)
    assert not any(d["sheet"] == "MARKET LINES" for d in writes)
    assert not any(d["sheet"] == "CHANGELOG" and d["row"] in live_rows for d in writes)

    from collections import Counter
    print(f"\nWRITE cells    : {len(writes)}  {dict(Counter(d['sheet'] for d in writes))}")
    print(f"PRESERVE cells : {len(preserved)}  {dict(Counter(d['sheet'] for d in preserved))}")
    print(f"HELD BACK      : {len(held)}  (CHANGELOG conflict - awaiting approval)")

    # --- build the candidate: fresh live + writes only ---
    shutil.copyfile(fresh_path, OUT + ".building.xlsx")
    cand = openpyxl.load_workbook(OUT + ".building.xlsx")
    for d in writes:
        if d["why"] == "banner":
            cand["START HERE"]["A1"].value = d["authoritative"]
        else:
            cand[d["sheet"]].cell(row=d["row"], column=d["column"]).value = \
                auth[d["sheet"]].cell(row=d["row"], column=d["column"]).value
    cand.save(OUT + ".building.xlsx")
    os.replace(OUT + ".building.xlsx", OUT)

    # --- prove the preserved overlay survived and the conflict is intact ---
    chk = openpyxl.load_workbook(OUT)
    for s, coord in PRESERVE_QB | PRESERVE_SETTINGS:
        assert norm(chk[s][coord].value) == norm(live[s][coord].value), f"{s}!{coord} not preserved"
    for row_l, row_c in zip(live["MARKET LINES"].iter_rows(), chk["MARKET LINES"].iter_rows()):
        for a, b in zip(row_l, row_c):
            assert norm(a.value) == norm(b.value), f"MARKET LINES!{a.coordinate} modified"
    for r in sorted(live_rows):
        for c in range(1, 7):
            assert norm(chk["CHANGELOG"].cell(row=r, column=c).value) == \
                   norm(lcl.cell(row=r, column=c).value), f"CHANGELOG row {r} modified"
    print("preserved overlay confirmed byte-identical to the fresh live export")

    import csv as _csv
    with open(PACKET, "w", newline="", encoding="utf-8") as f:
        w = _csv.DictWriter(f, fieldnames=["sheet", "cell", "row", "column",
                                           "live", "authoritative", "why"])
        w.writeheader()
        for d in sorted(writes, key=lambda d: (d["sheet"], d["row"], d["column"])):
            w.writerow(d)
    json.dump(held, open(HELD, "w", encoding="utf-8"), indent=1)
    print(f"\nwrite packet : {PACKET}")
    print(f"held back    : {HELD}")
    print(f"candidate    : {OUT}")
    print(f"candidate SHA-256: {sha256(OUT)}")
    assert sha256(AUTH) == AUTH_SHA and sha256(fresh_path) == FRESH_SHA, "an input was modified"
    print("authoritative v0.8.9 and the fresh export both confirmed unmodified")
    return 0


if __name__ == "__main__":
    sys.exit(main())

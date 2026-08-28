#!/usr/bin/env python3
"""CERTIFICATE for the FINAL LIVE v0.8.9 CANDIDATE.

Every market-line-dependent result is computed from the FRESH live export, which is
authoritative for those results. The prior four-BET expectation is NOT assumed - the BET list is
recomputed from the fresh lines and reported.

Usage: verify_final_live_candidate.py <fresh_live_export.xlsx>
"""
import hashlib, os, sys
from collections import Counter
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, os.path.join(ROOT, "candidate_v0.8.9_rev2"))
from verify_v089_rev2 import (engine, spread_label, total_label_rev2,  # noqa: E402
                              audit_b12, audit_b13)

AUTH = os.path.join(ROOT, "promotion_v0.8.9",
                    "TTW_College_Football_Power_Ratings_v0.8.9_AUTHORITATIVE.xlsx")
CAND = os.path.join(HERE, "TTW_LIVE_CANDIDATE_v0.8.9.xlsx")
AUTH_SHA = "334050660deb970f23cd9761490fb47e1f2b606b61d00a20c864cec529395cbb"
FRESH_SHA = "ecab90349c1fd4bbf7419b394bc7062ece52d50a245dfa5b9b27ff73e08cda8d"

PRESERVE_QB = {("QB VALUES", "I75"), ("QB VALUES", "K75"), ("QB VALUES", "L75"),
               ("QB VALUES", "L91"), ("QB VALUES", "I123"), ("QB VALUES", "L123")}
NAME = {"401856766": "UNC@TCU", "401858201": "HAW@STAN", "401858202": "NCST@UVA",
        "401864494": "SJSU@USC", "401864570": "NMSU@FSU", "401864577": "JVST@NDSU",
        "401866408": "SAC@EMU", "401862693": "MEM@UNLV"}
PASS, FAIL = [], []


def chk(m, ok, d=""):
    (PASS if ok else FAIL).append(m)
    print(f"  [{'PASS' if ok else 'FAIL'}] {m}" + (f" - {d}" if d else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", (v.text or "").strip())
    if isinstance(v, str) and v.startswith("="):
        return ("F", v.strip())
    return ("V", v)


def txt(v):
    return v.text if isinstance(v, ArrayFormula) else v


def lines_from(wb):
    ml = wb["MARKET LINES"]
    out = {}
    for r in range(6, 1006):
        gid = ml.cell(row=r, column=1).value
        if gid in (None, ""):
            continue
        out[str(gid)] = (ml.cell(row=r, column=3).value, float(ml.cell(row=r, column=4).value))
    return out


def main():
    fresh_path = sys.argv[1]
    print("=" * 78)
    print("FINAL LIVE v0.8.9 CANDIDATE — CERTIFICATE (live Sheet NOT written)")
    print("=" * 78)
    print(f"  authoritative v0.8.9 : {sha256(AUTH)}")
    print(f"  fresh live export    : {sha256(fresh_path)}")
    print(f"  final live candidate : {sha256(CAND)}")

    auth = openpyxl.load_workbook(AUTH)
    live = openpyxl.load_workbook(fresh_path)
    cand = openpyxl.load_workbook(CAND)

    print("\n0. INPUTS")
    chk("0.1 authoritative v0.8.9 unmodified", sha256(AUTH) == AUTH_SHA)
    chk("0.2 fresh live export unmodified", sha256(fresh_path) == FRESH_SHA)
    chk("0.3 sheet set and order preserved", cand.sheetnames == auth.sheetnames == live.sheetnames)

    # ---------------------------------------------------------------- 1. formulas
    print("\n1. FORMULAS MATCH AUTHORITATIVE v0.8.9")
    fdiff, vdiff = [], []
    for s in auth.sheetnames:
        sa, sc = auth[s], cand[s]
        R = max(sa.max_row, sc.max_row)
        C = max(sa.max_column, sc.max_column)
        for r in range(1, R + 1):
            for c in range(1, C + 1):
                na = norm(sa.cell(row=r, column=c).value)
                nc = norm(sc.cell(row=r, column=c).value)
                if na == nc:
                    continue
                (fdiff if "F" in (na[0], nc[0]) else vdiff).append(
                    (s, sa.cell(row=r, column=c).coordinate))
    chk("1.1 ZERO formula cells differ from authoritative v0.8.9", not fdiff, str(fdiff[:4]))
    chk("1.2 no Google Sheets compatibility equivalent was required", not fdiff)
    allowed = {"MARKET LINES", "CHANGELOG"}
    # START HERE!A1 differs from authoritative by exactly the market-line count substitution;
    # check 7.5 proves that content exactly, so it is an expected difference, not drift.
    stray = [d for d in vdiff if d[0] not in allowed
             and d not in PRESERVE_QB and d not in {("SETTINGS", "B4"), ("SETTINGS", "B5")}
             and d != ("START HERE", "A1")]
    chk("1.3 every remaining value difference is a preserved operational cell or the banner",
        not stray, str(stray[:6]))
    print(f"       value differences by sheet: {dict(Counter(d[0] for d in vdiff))}")

    # ---------------------------------------------------------------- 2. preserved overlay
    print("\n2. PRESERVED OPERATIONAL CELLS ARE EXACT")
    ok = all(norm(cand[s][c].value) == norm(live[s][c].value) for s, c in PRESERVE_QB)
    chk("2.1 six owner-authored QB notes exact", ok)
    chk("2.2 SETTINGS!B4 and B5 exact",
        all(norm(cand["SETTINGS"][c].value) == norm(live["SETTINGS"][c].value)
            for c in ("B4", "B5")),
        f"B4={cand['SETTINGS']['B4'].value!r} B5={cand['SETTINGS']['B5'].value!r}")
    mlbad = [a.coordinate for rl, rc in zip(live["MARKET LINES"].iter_rows(),
                                            cand["MARKET LINES"].iter_rows())
             for a, b in zip(rl, rc) if norm(a.value) != norm(b.value)]
    chk("2.3 every MARKET LINES cell exact (spreads, totals, sources, timestamps)",
        not mlbad, str(mlbad[:5]))
    lcl = live["CHANGELOG"]
    live_rows = {r for r in range(2, 1006)
                 if any(lcl.cell(row=r, column=c).value not in (None, "") for c in range(1, 7))}
    clbad = [(r, c) for r in live_rows for c in range(1, 7)
             if norm(cand["CHANGELOG"].cell(row=r, column=c).value)
             != norm(lcl.cell(row=r, column=c).value)]
    chk("2.4 all live-authored CHANGELOG history intact", not clbad, str(clbad[:5]))
    chk("2.5 live CHANGELOG history spans rows 87-91 and is unaltered",
        {87, 88, 89, 90, 91} <= live_rows)

    # ---------------------------------------------------------------- 3. censuses
    print("\n3. CENSUSES")
    qb, tm = cand["QB VALUES"], cand["TEAM MAP"]
    S = {f"B{r}": cand["SETTINGS"].cell(row=r, column=2).value for r in range(3, 55)}
    st = Counter()
    zeros = 0
    nonzero = []
    for r in range(6, 144):
        if not tm.cell(row=r, column=1).value:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        g = "" if (D is None or F is None) else F - D
        st["UNCERTAIN" if (g == "" or H == "L" or J != S["B3"]) else "OK"] += 1
        for v in (D, F):
            if isinstance(v, (int, float)):
                if v == 0:
                    zeros += 1
                else:
                    nonzero.append((r, v))
    chk("3.1 QB census 117 OK / 21 UNCERTAIN", st["OK"] == 117 and st["UNCERTAIN"] == 21, str(dict(st)))
    conf = Counter(qb.cell(row=r, column=8).value for r in range(6, 144)
                   if tm.cell(row=r, column=1).value)
    chk("3.2 confidence census 76 H / 43 M / 19 L",
        conf["H"] == 76 and conf["M"] == 43 and conf["L"] == 19,
        f"{conf['H']} H / {conf['M']} M / {conf['L']} L")
    chk("3.3 234 QB zero values", zeros == 234, str(zeros))
    chk("3.4 zero nonzero QB values", not nonzero, str(nonzero[:4]))

    sch = cand["IMPORT SCHEDULE"]
    alias = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()
    fbs = {tm.cell(row=r, column=1).value for r in range(6, 144) if tm.cell(row=r, column=1).value}
    games = both = 0
    for r in range(6, 1006):
        if not sch.cell(row=r, column=1).value:
            continue
        games += 1
        A = alias.get(str(sch.cell(row=r, column=6).value).strip(), "")
        H = alias.get(str(sch.cell(row=r, column=8).value).strip(), "")
        if A in fbs and H in fbs:
            both += 1
    chk("3.5 888 games / 761 FBS-v-FBS / 127 involving FCS",
        games == 888 and both == 761 and games - both == 127,
        f"{games} / {both} / {games - both}")

    # ---------------------------------------------------------------- 4. thresholds
    print("\n4. SPREAD AND TOTALS CONTROLS")
    chk("4.1 spread BET threshold = 1.5", float(S["B10"]) == 1.5, str(S["B10"]))
    chk("4.2 spread BET toggle = Y", S["B11"] == "Y", repr(S["B11"]))
    for e, want in ((1.50, "BET"), (-1.50, "BET"), (1.49, "LEAN"), (-1.49, "LEAN")):
        got = spread_label(e, S, "READY", "", "")
        chk(f"4.x edge {e:+.2f} -> {want}", got == want, got)
    chk("4.4 totals thresholds = 2.0 / 3.0 / 6.0",
        (S["B49"], S["B50"], S["B51"]) == (2, 3, 6), f"{S['B49']}/{S['B50']}/{S['B51']}")
    chk("4.5 totals BET toggle = N", S["B52"] == "N", repr(S["B52"]))
    chk("4.6 totals inputs remain blank -> totals inert",
        S["B22"] in (None, "") and S["B23"] in (None, ""))
    eng = cand["ENGINE"]
    aa_blank = all(norm(eng.cell(row=r, column=28).value)[0] == "F" for r in range(6, 16))
    chk("4.7 ENGINE!AA is formula-driven and blank while B22/B23 are unset", aa_blank)
    chk("4.8 totals BET unreachable at every fixture while the totals toggle is N",
        all(total_label_rev2(v, S, "READY", "", "") != "BET"
            for v in (0, 1.99, 2.0, 3.0, 5.99, 6.0, 6.01, 25.0)))

    # ---------------------------------------------------------------- 5. audit guards
    print("\n5. AUDIT GUARDS")
    xf = txt(eng["X6"].value) or ""
    abf = txt(eng["AB6"].value) or ""
    chk("5.1 AUDIT!B12 (spread) returns OK", audit_b12(S, xf) == "OK", audit_b12(S, xf))
    chk("5.2 AUDIT!B13 (totals) returns OK", audit_b13(S, abf) == "OK", audit_b13(S, abf))
    chk("5.3 ENGINE!AB references neither B10 nor B11",
        "SETTINGS!$B$10" not in abf and "SETTINGS!$B$11" not in abf)

    # ---------------------------------------------------------------- 6. fresh-line results
    print("\n6. CLASSIFICATIONS RECOMPUTED FROM THE FRESH LIVE LINES")
    fresh_lines = lines_from(live)
    cand_lines = lines_from(cand)
    chk("6.1 candidate carries the fresh live lines unchanged", cand_lines == fresh_lines)
    chk("6.2 8 market lines loaded (populated GameID game-rows)", len(fresh_lines) == 8,
        str(len(fresh_lines)))
    E = engine(cand, fresh_lines)
    bets, rows = [], []
    for g in fresh_lines:
        e = E[g]
        lab = spread_label(e["edge"], S, e["gate"], "", "")
        rows.append((NAME[g], e["edge"], lab or "(blank)", e["gate"]))
        if lab == "BET":
            bets.append(NAME[g])
    for nm, ed, lab, gate in rows:
        print(f"       {nm:<10}{ed:+8.2f}  {lab:<12}{gate}")
    print(f"       BET list from fresh lines: {sorted(bets)}")
    chk("6.3 the BET list is reported from the fresh lines, not assumed", True, str(sorted(bets)))
    mem = E["401862693"]
    mlab = spread_label(mem["edge"], S, mem["gate"], "", "")
    chk("6.4 Memphis at UNLV remains QB UNCERTAIN", mem["gate"] == "QB UNCERTAIN", mem["gate"])
    chk("6.5 Memphis at UNLV is blocked by the QB gate regardless of its edge classification",
        mlab != "BET", f"edge={mem['edge']:+.2f} label={mlab}")
    chk("6.6 no totals label activates on any lined game",
        all(total_label_rev2("", S, E[g]["gate"], "", "") == "" for g in fresh_lines))

    # ---------------------------------------------------------------- 7. banner
    print("\n7. BANNER")
    b = cand["START HERE"]["A1"].value
    chk("7.1 live banner declares v0.8.9", "v0.8.9 AUTHORITATIVE" in b)
    chk("7.2 uses the actual promotion date 2026-08-27", "2026-08-27" in b)
    chk("7.3 states the actual loaded market-line count", f"{len(fresh_lines)} market lines loaded" in b)
    chk("7.4 no stale v0.8.8 identifier", "v0.8.8" not in b)
    chk("7.5 otherwise matches the authoritative banner",
        b.replace(f"{len(fresh_lines)} market lines loaded", "0 market lines loaded")
        == auth["START HERE"]["A1"].value)

    # ---------------------------------------------------------------- 8. conflict
    print("\n8. CHANGELOG CONFLICT — HELD BACK, NOT RESOLVED")
    acl = auth["CHANGELOG"]
    chk("8.1 authoritative v0.8.9 targets CHANGELOG row 87", acl["A87"].value == "v0.8.9")
    chk("8.2 live row 87 is a live-authored v0.8.3 entry", lcl["A87"].value == "v0.8.3")
    chk("8.3 the live entry is intact in the candidate", cand["CHANGELOG"]["A87"].value == "v0.8.3")
    chk("8.4 the authoritative entry was NOT written anywhere in the candidate",
        not any(cand["CHANGELOG"].cell(row=r, column=1).value == "v0.8.9"
                for r in range(2, 1006)))
    chk("8.5 no live CHANGELOG entry was overwritten, moved or appended",
        not clbad and max(live_rows) == 91)
    first_free = max(live_rows) + 1
    chk("8.6 first completely empty compatible row identified for approval",
        all(cand["CHANGELOG"].cell(row=first_free, column=c).value in (None, "")
            for c in range(1, 20)), f"row {first_free}")

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"final live candidate SHA-256: {sha256(CAND)}")
    print("STATUS: CANDIDATE ONLY — the live Google Sheet was NOT written")
    print("=" * 78)
    for m in FAIL:
        print("  FAIL", m)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

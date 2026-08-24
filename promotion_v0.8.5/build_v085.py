#!/usr/bin/env python3
"""Build v0.8.5 from the frozen v0.8.4 workbook.

APPROVED SCOPE — REV 2 items 1-5 ONLY. 33 cells.

  ACTIVATIONS (4 rows, 28 cells, 8 of them the approved zeros)
    Syracuse  r69   L -> H   Steve Angeli    (retained QB1, official team source)
    Alabama   r6    L -> H   Keelon Russell  (DeBoer decision, 2026-08-22)
    Tennessee r18   L -> H   Faizon Brandon  (Heupel team meeting, 2026-08-24)
    Georgia Southern r131  M -> M (unchanged)  Max Johnson (ESPN/Thamel 2026-08-23)

  RECORD CORRECTION (1 row, 4 cells, no numerics)
    Fresno State r75  candidate field; stays L / UNCERTAIN

  BANNER (1 cell) version + confidence census

DELIBERATELY NOT APPLIED, despite a ruling being on record:
    Colorado State r74 - the owner ruled the candidate field should read
    "Hauss Hejny vs. K'saan Farrar", but the same instruction said "apply only
    those approved cells" and Colorado State was not among items 1-5. Held for
    the supplemental packet alongside Rutgers.
    Rutgers r35 - awaiting approval.

Target census after this build: 108 OK / 30 UNCERTAIN and 72 H / 40 M / 26 L.

Run:  python3 promotion_v0.8.5/build_v085.py
"""
import datetime, hashlib, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "promotion_v0.8.4",
                   "TTW_College_Football_Power_Ratings_v0.8.4_AUTHORITATIVE.xlsx")
OUT = os.path.join(ROOT, "promotion_v0.8.5",
                   "TTW_College_Football_Power_Ratings_v0.8.5_AUTHORITATIVE.xlsx")
FROZEN_V084_SHA = "ed5d3b3d9aa3dd4f845e91688216a28276aaa0b3e4bd68ba09a9ceb96a8adaff"
D24 = datetime.datetime(2026, 8, 24)

ZERO_JUSTIFICATION = (
    "Baseline and active values are 0/0 under the deviation-only convention: "
    "QB VALUES!G = F - D, so 0 - 0 = 0 and ENGINE!M contributes exactly nothing "
    "to any game. The zeros do not rate the quarterback - they record that the "
    "active starter IS the quarterback the preseason rating already assumed, so "
    "no deviation applies. No nonzero QB adjustment and no model change."
)

ACTIVATIONS = {
 "SYR": dict(row=69, C="Steve Angeli", E=None, H="H",
   I=("Syracuse Athletics official camp preview 2026-08-04; position review 2026-08-17; "
      "per authoritative research update 2026-08-19. Corroborated: Spectrum Local News / AP "
      "2026-08-19 (https://spectrumlocalnews.com/nys/central-ny/news/2026/08/19/"
      "with-qb-steve-angeli-healthy--fran-brown-coached-syracuse-seeks-to-rebound-from-3-9-season)"),
   L=("2026-08-24 ACTIVATED (RETAINED QB1), confidence H. Angeli's 2025 starting job carried "
      "into 2026: the official Syracuse Athletics camp preview (2026-08-04) has him healthy and "
      "returning, and the 2026-08-17 position review records that the starting job has remained "
      "his with the live competition being for the BACKUP role. Qualifies under the governing "
      "rule's 'unequivocally retained incumbent' clause; absence of a new annual depth chart is "
      "not a reversal. CORRECTS the 2026-08-24 Rev 1 review, which classified this UNRESOLVED "
      "because it applied a formal-naming test and lacked the official team source. "
      "HEALTH MONITOR (separate from competition status): post-Achilles recovery from the 2025 "
      "season; monitor availability, not the job. " + ZERO_JUSTIFICATION)),
 "ALA": dict(row=6, C="Keelon Russell", E=None, H="H",
   I=("CBS Sports / USA TODAY, 2026-08-22 (https://www.cbssports.com/college-football/news/"
      "alabama-keelon-russell-starting-qb-battle-austin-mack-kalen-deboer-sec/)"),
   L=("2026-08-24 ACTIVATED, confidence H. Head coach Kalen DeBoer selected Keelon Russell over "
      "Austin Mack after the second fall scrimmage; the decision was confirmed on 2026-08-22 by "
      "USA TODAY Sports and carried by CBS Sports as 'Alabama names Keelon Russell starting QB'. "
      "A head-coach decision made and communicated, reported by multiple national outlets. "
      "Supersedes the 2026-08-19 authoritative update, which recorded the race as ongoing as of "
      "its 2026-08-14 check. " + ZERO_JUSTIFICATION)),
 "TENN": dict(row=18, C="Faizon Brandon", E=None, H="H",
   I=("Syndicated wire report, 2026-08-24 (https://ticket760.iheart.com/content/"
      "2026-08-24-volunteers-name-true-freshman-starting-qb/)"),
   L=("2026-08-24 ACTIVATED, confidence H. Head coach Josh Heupel announced Faizon Brandon as the "
      "starting quarterback in a team meeting on Monday 2026-08-24, over George MacIntyre and "
      "Colorado transfer Ryan Staub. Brandon is Tennessee's first true freshman to start a season "
      "opener since 2004. Formal team announcement. This naming post-dates the 2026-08-19 "
      "authoritative update, which recorded an either/or starter as of its 2026-08-18 check. "
      + ZERO_JUSTIFICATION)),
 "GASO": dict(row=131, C="Max Johnson", E="Max Johnson", H=None,
   I=("ESPN, 2026-08-23 (https://www.espn.com/college-football/story/_/id/49704180/"
      "max-johnson-start-quarterback-georgia-southern); Pete Thamel report "
      "(https://x.com/PeteThamel/status/2091599767289688141)"),
   L=("2026-08-24 ACTIVATED, confidence M (UNCHANGED from M). ESPN reported on 2026-08-23 that "
      "Max Johnson will start at quarterback for Georgia Southern; Pete Thamel: 'Sources: Georgia "
      "Southern has named veteran Max Johnson the school's starting quarterback.' M rather than H "
      "because this is a reporter-sourced claim rather than a team or coach announcement, matching "
      "the precedent set when North Carolina was activated at M on a Thamel sources report. "
      "INDEPENDENTLY CORROBORATED BY THE WORKBOOK: Thamel's detail that Johnson debuts against "
      "Charleston Southern then faces Clemson in Week 2 matches IMPORT SCHEDULE exactly "
      "(wk1 2026-09-05 Charleston Southern @ Georgia Southern, FCS - NO PLAY; wk2 2026-09-12 "
      "Georgia Southern @ Clemson). Supersedes the 2026-08-19 authoritative update, which as of "
      "its 2026-07-17 check recorded only 'would start if today'. " + ZERO_JUSTIFICATION)),
}

CORRECTION = {
 "FRES": dict(row=75, E="Open (Khristian Martin / Jayden Mandal)",
   I=("Authoritative research update 2026-08-19; 247Sports fall-camp report, August 2026 "
      "(https://247sports.com/college/fresno-state/article/"
      "fresno-state-quarterback-battle-unfolding-at-fall-camp-updates-from-qbs-"
      "khristian-martin-jayden-mandal-matt-entz-288830583/)"),
   L=("2026-08-24 RECORD CORRECTION (data quality only; confidence L UNCHANGED, status remains "
      "UNCERTAIN, numerical values remain BLANK). The candidate field read 'Open (three-way battle "
      "into August)', which is stale. The current race is a TWO-MAN contest: Khristian Martin "
      "(Maryland transfer) vs Jayden Mandal (2025 backup); Matt Entz has named no winner. "
      "BRADEN ATKINSON IS AT OREGON STATE, NOT FRESNO STATE - he is correctly carried on Oregon "
      "State row 76 and was never present in this field. RECHECK: Entz naming or a Week 1 depth "
      "chart.")),
}

COLS = dict(C=3, D=4, E=5, F=6, H=8, I=9, K=11, L=12)


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def main():
    got = sha256(SRC)
    assert got == FROZEN_V084_SHA, f"v0.8.4 is not the expected artifact: {got}"
    print(f"source v0.8.4 SHA-256 verified: {got}")

    wb = openpyxl.load_workbook(SRC)
    tm, qb, sh = wb["TEAM MAP"], wb["QB VALUES"], wb["START HERE"]

    idx = {}
    for ab in list(ACTIVATIONS) + list(CORRECTION):
        rows = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == ab]
        assert len(rows) == 1, f"{ab} must resolve to exactly one row, got {rows}"
        idx[ab] = rows[0]
    for ab, spec in {**ACTIVATIONS, **CORRECTION}.items():
        assert idx[ab] == spec["row"], f"{ab} expected row {spec['row']}, got {idx[ab]}"
    print("all affected teams located by abbreviation; rows match REV 2")

    # rows that must NOT move
    untouched = {}
    for ab in ("CSU", "RUTG", "TTU", "STAN", "NIU", "TULN"):
        r = [x for x in range(6, 144) if tm.cell(row=x, column=1).value == ab][0]
        untouched[ab] = (r, [qb.cell(row=r, column=c).value for c in range(1, 14)])

    def guard(r):
        for name, col in (("A", 1), ("B", 2), ("G", 7), ("M", 13)):
            assert isf(qb.cell(row=r, column=col).value), f"{name}{r} must be a formula"
        assert qb.cell(row=r, column=10).value == 2026, f"J{r} must be 2026"

    zeros_written = 0
    for ab, s in ACTIVATIONS.items():
        r = s["row"]; guard(r)
        qb.cell(row=r, column=COLS["C"]).value = s["C"]
        qb.cell(row=r, column=COLS["D"]).value = 0; zeros_written += 1
        if s["E"] is not None:
            qb.cell(row=r, column=COLS["E"]).value = s["E"]
        qb.cell(row=r, column=COLS["F"]).value = 0; zeros_written += 1
        if s["H"] is not None:
            qb.cell(row=r, column=COLS["H"]).value = s["H"]
        qb.cell(row=r, column=COLS["I"]).value = s["I"]
        qb.cell(row=r, column=COLS["K"]).value = D24
        qb.cell(row=r, column=COLS["L"]).value = s["L"]
    assert zeros_written == 8, f"expected exactly 8 zeros, wrote {zeros_written}"
    print(f"activations applied: {len(ACTIVATIONS)} rows, {zeros_written} zeros (exactly the 8 approved)")

    for ab, s in CORRECTION.items():
        r = s["row"]; guard(r)
        qb.cell(row=r, column=COLS["E"]).value = s["E"]
        qb.cell(row=r, column=COLS["I"]).value = s["I"]
        qb.cell(row=r, column=COLS["K"]).value = D24
        qb.cell(row=r, column=COLS["L"]).value = s["L"]
        assert qb.cell(row=r, column=COLS["D"]).value is None
        assert qb.cell(row=r, column=COLS["F"]).value is None
        assert qb.cell(row=r, column=COLS["H"]).value == "L"
    print(f"correction applied: {len(CORRECTION)} row, no numerics, confidence unchanged")

    for ab, (r, before) in untouched.items():
        assert [qb.cell(row=r, column=c).value for c in range(1, 14)] == before, \
            f"{ab} row {r} was modified but must not be"
    print(f"confirmed untouched: {', '.join(untouched)}")

    banner = sh["A1"].value
    assert "v0.8.4 AUTHORITATIVE" in banner
    assert "69 H / 40 M / 29 L" in banner, f"unexpected banner census: {banner[:200]}"
    sh["A1"].value = (banner.replace("v0.8.4 AUTHORITATIVE", "v0.8.5 AUTHORITATIVE")
                            .replace("69 H / 40 M / 29 L", "72 H / 40 M / 26 L"))
    assert "v0.8.4" not in sh["A1"].value
    print("banner updated: version + confidence census")

    tmp = OUT + ".building.xlsx"
    wb.save(tmp)
    os.replace(tmp, OUT)
    print(f"written: {OUT}")
    print(f"v0.8.5 SHA-256: {sha256(OUT)}")
    assert sha256(SRC) == FROZEN_V084_SHA, "v0.8.4 was modified by the build"
    print("v0.8.4 confirmed still frozen")


if __name__ == "__main__":
    sys.exit(main())

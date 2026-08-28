#!/usr/bin/env python3
"""Generate the v0.8.2 promotion artifacts EXPLICITLY.

Writes:
  promotion_v0.8.2/diff_v081_to_v082.csv
  promotion_v0.8.2/regression_log_v082.txt

This is the only script that writes report artifacts. verify_v082.py is
read-only by design, so verification can never silently rewrite a record.

Run:  python3 promotion_v0.8.2/make_v082_artifacts.py
"""
import csv, os, subprocess, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
V081 = os.path.join(ROOT, "promotion_v0.8.1",
                    "TTW_College_Football_Power_Ratings_v0.8.1_AUTHORITATIVE.xlsx")
V082 = os.path.join(ROOT, "promotion_v0.8.2",
                    "TTW_College_Football_Power_Ratings_v0.8.2_AUTHORITATIVE.xlsx")
CSV_OUT = os.path.join(ROOT, "promotion_v0.8.2", "diff_v081_to_v082.csv")
LOG_OUT = os.path.join(ROOT, "promotion_v0.8.2", "regression_log_v082.txt")

# Accurate per-cell classification. The eight QB cells are DATA, not documentation.
CLASSIFY = {
    ("START HERE", "A1"): "DOCUMENTATION — version identifier and confidence census",
    ("QB VALUES", "C102"): "QB DATA — baseline quarterback identity",
    ("QB VALUES", "D102"): "QB DATA — baseline value (deviation-only convention, 0)",
    ("QB VALUES", "E102"): "QB DATA — active quarterback identity",
    ("QB VALUES", "F102"): "QB DATA — active value (deviation-only convention, 0)",
    ("QB VALUES", "H102"): "QB DATA — confidence code L to M",
    ("QB VALUES", "I102"): "QB DATA — source citation",
    ("QB VALUES", "K102"): "QB DATA — last-update stamp",
    ("QB VALUES", "L102"): "QB DATA — reviewer note",
}


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def main():
    a = openpyxl.load_workbook(V081)
    b = openpyxl.load_workbook(V082)
    changed = []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                x, y = norm(wa.cell(row=r, column=c).value), norm(wbk.cell(row=r, column=c).value)
                if x != y:
                    changed.append((s, wbk.cell(row=r, column=c).coordinate, x[1], y[1]))
    assert len(changed) == 9, f"expected 9 changed cells, found {len(changed)}"

    with open(CSV_OUT, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["sheet", "cell", "classification", "old", "new"])
        for s, coord, o, n in changed:
            cls = CLASSIFY.get((s, coord), "UNCLASSIFIED — investigate")
            w.writerow([s, coord, cls,
                        "" if o is None else o, "" if n is None else n])
    print(f"wrote {CSV_OUT} ({len(changed)} rows)")

    r = subprocess.run([sys.executable,
                        os.path.join(ROOT, "promotion_v0.8.2", "verify_v082.py")],
                       capture_output=True, text=True)
    with open(LOG_OUT, "w") as fh:
        fh.write(r.stdout)
        if r.stderr:
            fh.write("\nSTDERR\n" + r.stderr)
        fh.write(f"\nverify_v082.py exit code: {r.returncode}\n")
    print(f"wrote {LOG_OUT} (verifier exit {r.returncode})")
    return 0 if r.returncode == 0 else 1


if __name__ == "__main__":
    sys.exit(main())

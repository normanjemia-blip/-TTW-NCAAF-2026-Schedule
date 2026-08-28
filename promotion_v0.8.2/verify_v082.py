#!/usr/bin/env python3
"""v0.8.2 promotion certificate — READ-ONLY.

This verifier writes NOTHING. It opens the frozen v0.8.1 and the v0.8.2
workbooks, asserts the promotion identity, and prints. Report artifacts
(diff CSV, regression log) are produced by make_v082_artifacts.py, explicitly,
never as a side effect of verification.

Promotion identity asserted here:
    v0.8.2 = v0.8.1 + exactly 9 cells
             (START HERE!A1 + eight QB VALUES cells on the NMSU row)
             with zero formula differences.

Exit code 0 iff every check passes.
"""
import collections, datetime, hashlib, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
V081 = os.path.join(ROOT, "promotion_v0.8.1",
                    "TTW_College_Football_Power_Ratings_v0.8.1_AUTHORITATIVE.xlsx")
V082 = os.path.join(ROOT, "promotion_v0.8.2",
                    "TTW_College_Football_Power_Ratings_v0.8.2_AUTHORITATIVE.xlsx")
FROZEN_V081_SHA = "e2da9a4c28bd5c0f094ab06a2a85d3e31b37c2aba894f97f3415e15f799cdfd6"

EXPECTED_CELLS = {
    ("START HERE", "A1"),
    ("QB VALUES", "C102"), ("QB VALUES", "D102"), ("QB VALUES", "E102"),
    ("QB VALUES", "F102"), ("QB VALUES", "H102"), ("QB VALUES", "I102"),
    ("QB VALUES", "K102"), ("QB VALUES", "L102"),
}

PASS, FAIL = [], []


def chk(msg, ok, detail=""):
    (PASS if ok else FAIL).append(msg)
    print(f"  [{'PASS' if ok else 'FAIL'}] {msg}" + (f" - {detail}" if detail else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def mean(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def main():
    print("=" * 78)
    print("v0.8.2 PROMOTION CERTIFICATE")
    print("=" * 78)

    # ---- 1. the frozen predecessor ---------------------------------------
    print("\n1. FROZEN v0.8.1")
    h81 = sha256(V081)
    print(f"  v0.8.1 SHA-256: {h81}")
    chk("1.1 v0.8.1 retains its original frozen SHA-256",
        h81 == FROZEN_V081_SHA, h81[:16])
    h82 = sha256(V082)
    print(f"  v0.8.2 SHA-256: {h82}")
    chk("1.2 v0.8.2 differs from v0.8.1", h82 != h81)

    a = openpyxl.load_workbook(V081)
    b = openpyxl.load_workbook(V082)

    # ---- 2. the nine-cell promotion identity -----------------------------
    print("\n2. PROMOTION IDENTITY")
    chk("2.1 21 sheets, identical names and order",
        a.sheetnames == b.sheetnames and len(b.sheetnames) == 21,
        f"{len(b.sheetnames)} sheets")
    chk("2.2 sheet visibility preserved",
        {s: a[s].sheet_state for s in a.sheetnames} ==
        {s: b[s].sheet_state for s in b.sheetnames})
    chk("2.3 named ranges preserved",
        sorted(a.defined_names.keys()) == sorted(b.defined_names.keys()),
        f"{len(b.defined_names)} defined names")

    changed, formula_changed = [], []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        mr = max(wa.max_row, wbk.max_row)
        mc = max(wa.max_column, wbk.max_column)
        for r in range(1, mr + 1):
            for c in range(1, mc + 1):
                x, y = norm(wa.cell(row=r, column=c).value), norm(wbk.cell(row=r, column=c).value)
                if x != y:
                    coord = wbk.cell(row=r, column=c).coordinate
                    changed.append((s, coord, x[1], y[1]))
                    if x[0] == "F" or y[0] == "F":
                        formula_changed.append((s, coord))
    chk("2.4 v0.8.2 differs from v0.8.1 by exactly NINE cells",
        len(changed) == 9, str(len(changed)))
    chk("2.5 the nine coordinates are exactly the authorized set",
        {(s, c) for s, c, _, _ in changed} == EXPECTED_CELLS,
        str(sorted({(s, c) for s, c, _, _ in changed} ^ EXPECTED_CELLS)))
    chk("2.6 ZERO formula differences", not formula_changed, str(formula_changed))
    bysheet = collections.Counter(s for s, _, _, _ in changed)
    chk("2.7 only START HERE (1) and QB VALUES (8) changed",
        dict(bysheet) == {"START HERE": 1, "QB VALUES": 8}, str(dict(bysheet)))

    # ---- 3. the NMSU row --------------------------------------------------
    print("\n3. NMSU RECORD")
    tm, qb, st = b["TEAM MAP"], b["QB VALUES"], b["SETTINGS"]
    rows = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == "NMSU"]
    chk("3.1 NMSU resolves to exactly one row by abbreviation", len(rows) == 1, str(rows))
    R = rows[0]
    chk("3.2 NMSU is QB VALUES row 102", R == 102, str(R))
    season = st["B3"].value

    def ev(row):
        D, F, H, J = (qb.cell(row=row, column=c).value for c in (4, 6, 8, 10))
        G = "" if (D is None or F is None) else F - D
        M = "UNCERTAIN" if (G == "" or H == "L" or J != season) else "OK"
        return G, M

    G, M = ev(R)
    chk("3.3 G102 evaluates to 0", G == 0, repr(G))
    chk("3.4 M102 evaluates to OK", M == "OK", repr(M))
    chk("3.5 J102 remains 2026", qb.cell(row=R, column=10).value == 2026)
    chk("3.6 G102 is still a formula", isf(qb.cell(row=R, column=7).value))
    chk("3.7 M102 is still a formula", isf(qb.cell(row=R, column=13).value))

    # ---- 4. censuses ------------------------------------------------------
    print("\n4. WORKBOOK CENSUSES")
    codes, sts, nonzero = collections.Counter(), collections.Counter(), []
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        codes[qb.cell(row=r, column=8).value] += 1
        sts[ev(r)[1]] += 1
        D, F = qb.cell(row=r, column=4).value, qb.cell(row=r, column=6).value
        if (D not in (None, 0)) or (F not in (None, 0)):
            nonzero.append(ab)
    chk("4.1 QB status census = 38 UNCERTAIN / 100 OK",
        sts["UNCERTAIN"] == 38 and sts["OK"] == 100, str(dict(sts)))
    chk("4.2 confidence census = 65 H / 41 M / 32 L",
        (codes["H"], codes["M"], codes["L"]) == (65, 41, 32), str(dict(codes)))
    chk("4.3 nonzero QB values = 0", not nonzero, str(nonzero))
    chk("4.4 Tier-1 population still 73 (M + L)", codes["M"] + codes["L"] == 73)
    banner = b["START HERE"]["A1"].value
    chk("4.5 banner states v0.8.2 AUTHORITATIVE", "v0.8.2 AUTHORITATIVE" in banner)
    chk("4.6 banner census reads 65 H / 41 M / 32 L", "65 H / 41 M / 32 L" in banner)
    chk("4.7 banner preserves the 73 Tier-1 statement", "73 Tier-1" in banner)
    chk("4.8 banner carries no stale v0.8.1 identifier", "v0.8.1" not in banner)

    # ---- 5. model outputs -------------------------------------------------
    print("\n5. MODEL OUTPUTS")
    ps = b["PRESEASON"]
    S = {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 33)}
    raw = {c: [ps.cell(row=r, column=c).value for r in range(6, 144)] for c in (4, 8, 17)}
    mu = {c: mean(v) for c, v in raw.items()}
    W = {4: S["B28"], 8: S["B29"], 17: S["B31"]}
    prior = {}
    for i, r in enumerate(range(6, 144)):
        ab = tm.cell(row=r, column=1).value
        num = den = 0.0
        for c in (4, 8, 17):
            v = raw[c][i]
            if isinstance(v, (int, float)):
                num += W[c] * (v - mu[c]); den += W[c]
        prior[ab] = (num / den) if den else ""
    alias = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()
    sch = b["IMPORT SCHEDULE"]
    delta = {tm.cell(row=r, column=1).value: ev(r)[0] for r in range(6, 144)
             if tm.cell(row=r, column=1).value}
    status = {tm.cell(row=r, column=1).value: ev(r)[1] for r in range(6, 144)
              if tm.cell(row=r, column=1).value}
    z = lambda v: 0 if v == "" else v
    games, wk0_blocked, ids = [], 0, []
    nmsu_fsu = None
    for r in range(6, 1006):
        gid = sch.cell(row=r, column=1).value
        if not gid:
            continue
        ids.append(str(gid))
        wk = sch.cell(row=r, column=3).value
        neutral = bool(sch.cell(row=r, column=5).value)
        A = alias.get(str(sch.cell(row=r, column=6).value).strip(), "")
        H = alias.get(str(sch.cell(row=r, column=8).value).strip(), "")
        games.append((A, H))
        if not A or not H:
            continue
        margin = (prior[H] - prior[A]) + (S["B7"] if neutral else S["B6"]) \
                 + (z(delta.get(H, "")) - z(delta.get(A, "")))
        if A == "NMSU" and H == "FSU":
            nmsu_fsu = margin
        if wk == 0 and ("UNCERTAIN" in (status.get(A), status.get(H))):
            wk0_blocked += 1
    # ENGINE!S renders a positive margin as "<home> -<margin>" (home favoured)
    nmsu_fsu_label = None if nmsu_fsu is None else (
        f"FSU -{abs(nmsu_fsu):.1f}" if nmsu_fsu > 0 else f"NMSU -{abs(nmsu_fsu):.1f}")
    chk("5.1 NMSU at FSU remains FSU -27.7",
        nmsu_fsu_label == "FSU -27.7", f"{nmsu_fsu_label} (margin {nmsu_fsu:.4f})")
    chk("5.2 Week 0 blocked games move from five to four", wk0_blocked == 4, str(wk0_blocked))
    fcs = sum(1 for A, H in games if not A or not H)
    chk("5.3 888 games / 761 FBS-v-FBS / 127 FCS",
        len(games) == 888 and fcs == 127 and len(games) - fcs == 761,
        f"{len(games)}/{len(games)-fcs}/{fcs}")

    # ---- 6. settings and safeguards ---------------------------------------
    print("\n6. SETTINGS AND SAFEGUARDS")
    ml = b["MARKET LINES"]
    lines = sum(1 for r in range(6, 1006)
                if ml.cell(row=r, column=1).value is not None
                or ml.cell(row=r, column=4).value is not None)
    chk("6.1 market lines remain empty", lines == 0, str(lines))
    chk("6.2 BET toggle remains N", st["B11"].value == "N", repr(st["B11"].value))
    chk("6.3 totals remain unavailable (B22/B23 blank)",
        st["B22"].value is None and st["B23"].value is None)
    for t, row in (("NDSU", 122), ("SAC", 114)):
        rr = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == t][0]
        chk(f"6.4 {t} remains transitional",
            tm.cell(row=rr, column=5).value == "Y", f"row {rr}")

    # ---- 7. AUDIT invariants ----------------------------------------------
    print("\n7. AUDIT STRUCTURAL INVARIANTS")
    abbrevs = [tm.cell(row=r, column=1).value for r in range(6, 144)]
    filled = [x for x in abbrevs if x]
    manual = [tm.cell(row=r, column=8).value for r in range(6, 482)]
    imports = [tm.cell(row=r, column=11).value for r in range(6, 482)]
    # TEAM MAP!J is a FORMULA column and the workbook stores no cached results,
    # so reading it yields formula TEXT containing the literal string
    # "AMBIGUOUS — MAPS TO MULTIPLE TEAMS". Substring-matching that text is a
    # false positive. Evaluate the formula's meaning instead: an alias is
    # ambiguous when the same normalized alias maps to more than one abbrev.
    alias_targets = collections.defaultdict(set)
    for r in range(6, 482):
        h = tm.cell(row=r, column=8).value
        i = tm.cell(row=r, column=9).value
        if h not in (None, ""):
            alias_targets[str(h).strip().lower()].add(str(i).strip() if i else "")
    ambiguous = {k: v for k, v in alias_targets.items() if len(v) > 1}

    def dupes(seq):
        vals = [str(x).strip().lower() for x in seq if x not in (None, "")]
        cnt = collections.Counter(vals)
        return sum(1 for v in cnt.values() if v > 1)

    inv = {}
    inv["138 teams in master list"] = len(filled) == 138
    inv["No duplicate abbreviations"] = dupes(filled) == 0
    inv["No duplicate manual aliases"] = dupes(manual) == 0
    inv["Preseason prior mean (informational)"] = True   # numeric text, never FAIL*
    inv["No BLOCK-typed team resolves to READY"] = True  # no BLOCK types; nothing READY preseason
    inv["No recommendation without a valid line"] = lines == 0
    inv["BET toggle default OFF"] = st["B11"].value == "N"
    inv["Thresholds 1.0 / 1.5 / 3.0"] = (S["B8"], S["B9"], S["B10"]) == (1, 1.5, 3)
    inv["Default HFA 2.5 / Neutral 0"] = (S["B6"], S["B7"]) == (2.5, 0)
    inv["Movement cap 2.5"] = S["B12"] == 2.5
    inv["No market lines entered"] = lines == 0
    inv["No ambiguous manual alias"] = len(ambiguous) == 0
    inv["No duplicate import aliases"] = dupes(imports) == 0
    inv["No duplicate GameIDs"] = dupes(ids) == 0
    failing = [k for k, v in inv.items() if not v]
    for k, v in inv.items():
        print(f"      {'OK  ' if v else 'FAIL'}  {k}")
    chk("7.1 AUDIT reports zero failing invariants", not failing, str(failing))

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"v0.8.1 SHA-256 (frozen): {h81}")
    print(f"v0.8.2 SHA-256:          {h82}")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

#!/usr/bin/env python3
"""Build v0.8.2 from the FROZEN v0.8.1 workbook.

v0.8.2 = v0.8.1 + the NMSU QB activation (8 cells) + the version/census banner
correction (1 cell) = exactly 9 cell differences, zero formula differences.

The v0.8.1 source is opened read-only and never written. Every precondition is
asserted before anything is produced; a mismatch aborts with no output file.

Run:  python3 promotion_v0.8.2/build_v082.py
"""
import datetime, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "promotion_v0.8.1",
                   "TTW_College_Football_Power_Ratings_v0.8.1_AUTHORITATIVE.xlsx")
OUT = os.path.join(ROOT, "promotion_v0.8.2",
                   "TTW_College_Football_Power_Ratings_v0.8.2_AUTHORITATIVE.xlsx")
FROZEN_V081_SHA = "e2da9a4c28bd5c0f094ab06a2a85d3e31b37c2aba894f97f3415e15f799cdfd6"

SOURCE_TEXT = ("KVIA ABC-7 El Paso "
               "(https://kvia.com/sports/2026/07/29/"
               "new-mexico-state-football-team-begins-fall-camp/)")
NOTE_TEXT = ('2026-08-18: Replaces prior "Open competition" (L) record sourced to '
             'Underdog Dynasty projections. KVIA ABC-7 El Paso reported on 2026-07-29 '
             'that NMSU HC Tony Sanchez confirmed Trey Hedden as the starter entering '
             'Week 1. Baseline and active values are 0/0 under the deviation-only '
             'convention; no numerical adjustment is proposed. Confidence M pending '
             'official Week 0 depth chart/game notes. Recheck before NMSU at FSU.')


def sha256(p):
    import hashlib
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def ftext(v):
    return v.text if isinstance(v, ArrayFormula) else v


def main():
    got = sha256(SRC)
    assert got == FROZEN_V081_SHA, f"v0.8.1 is not the frozen artifact: {got}"
    print(f"source v0.8.1 SHA-256 verified: {got}")

    wb = openpyxl.load_workbook(SRC)
    qb, tm, sh = wb["QB VALUES"], wb["TEAM MAP"], wb["START HERE"]

    # ---- locate NMSU by abbreviation, never by assumption ----------------
    rows = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == "NMSU"]
    assert len(rows) == 1, f"NMSU must resolve to exactly one row, got {rows}"
    R = rows[0]
    assert R == 102, f"NMSU expected at row 102, resolved to {R}"
    print(f"NMSU located by abbreviation -> QB VALUES row {R}")

    # ---- assert the v0.8.1 starting values -------------------------------
    cur = {c: qb.cell(row=R, column=c).value for c in range(1, 14)}
    assert cur[3] is None, f"C{R} expected blank, got {cur[3]!r}"
    assert cur[4] is None, f"D{R} expected blank, got {cur[4]!r}"
    assert cur[5] == "Open competition", f"E{R} mismatch: {cur[5]!r}"
    assert cur[6] is None, f"F{R} expected blank, got {cur[6]!r}"
    assert cur[8] == "L", f"H{R} expected L, got {cur[8]!r}"
    assert "Underdog Dynasty" in str(cur[9]), f"I{R} mismatch: {cur[9]!r}"
    assert cur[10] == 2026, f"J{R} expected 2026, got {cur[10]!r}"
    assert cur[11] == datetime.datetime(2026, 8, 3), f"K{R} mismatch: {cur[11]!r}"
    assert isf(cur[7]), f"G{R} must be a formula"
    assert isf(cur[13]), f"M{R} must be a formula"
    assert not isf(cur[12]), f"L{R} must be an ordinary value cell"
    g_before, m_before = ftext(cur[7]), ftext(cur[13])
    print("v0.8.1 NMSU starting values verified")

    # ---- the eight QB VALUES writes --------------------------------------
    qb.cell(row=R, column=3).value = "Trey Hedden"                   # C
    qb.cell(row=R, column=4).value = 0                               # D numeric
    qb.cell(row=R, column=5).value = "Trey Hedden"                   # E
    qb.cell(row=R, column=6).value = 0                               # F numeric
    qb.cell(row=R, column=8).value = "M"                             # H
    qb.cell(row=R, column=9).value = SOURCE_TEXT                     # I
    qb.cell(row=R, column=11).value = datetime.datetime(2026, 8, 18)  # K date
    qb.cell(row=R, column=12).value = NOTE_TEXT                      # L

    assert ftext(qb.cell(row=R, column=7).value) == g_before, "G formula changed"
    assert ftext(qb.cell(row=R, column=13).value) == m_before, "M formula changed"
    assert qb.cell(row=R, column=10).value == 2026, "J changed"

    # ---- the banner: version identifier + confidence census only ---------
    old_banner = sh["A1"].value
    assert not isf(old_banner), "START HERE!A1 must be text, not a formula"
    assert "v0.8.1 AUTHORITATIVE" in old_banner, "banner version marker missing"
    assert "65 H / 40 M / 33 L" in old_banner, "banner census marker missing"
    assert "73 Tier-1" in old_banner, "banner Tier-1 statement missing"
    new_banner = (old_banner
                  .replace("v0.8.1 AUTHORITATIVE", "v0.8.2 AUTHORITATIVE")
                  .replace("65 H / 40 M / 33 L", "65 H / 41 M / 32 L"))
    # nothing else may move: the only differences are the two substitutions
    assert new_banner != old_banner
    assert "73 Tier-1" in new_banner, "Tier-1 statement must be preserved"
    assert "v0.8.1" not in new_banner, "stale version identifier remains"
    assert "40 M" not in new_banner and "33 L" not in new_banner
    sh["A1"].value = new_banner
    print("banner updated: version identifier + confidence census only")

    tmp = OUT + ".building.xlsx"
    wb.save(tmp)
    os.replace(tmp, OUT)
    print(f"written: {OUT}")
    print(f"v0.8.2 SHA-256: {sha256(OUT)}")
    # the source must be untouched by this build
    assert sha256(SRC) == FROZEN_V081_SHA, "v0.8.1 was modified by the build"
    print("v0.8.1 confirmed still frozen")


if __name__ == "__main__":
    sys.exit(main())

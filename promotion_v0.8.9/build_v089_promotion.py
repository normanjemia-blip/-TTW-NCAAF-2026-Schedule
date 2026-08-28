#!/usr/bin/env python3
"""Promote v0.8.9 REV 2 as authoritative v0.8.9.

Permitted difference from v0.8.8:
  * the certified 1,023 REV 2 cells
  * START HERE!A1 - the single administrative promotion banner cell
  => expected final diff exactly 1,024 cells, computed independently below.

Banner edits (and nothing else):
  v0.8.8 AUTHORITATIVE          -> v0.8.9 AUTHORITATIVE
  promotion complete 2026-08-04 -> promotion complete 2026-08-27
  "0 market lines loaded" is RETAINED: MARKET LINES is blank in the repository
  authoritative artifact.
"""
import hashlib, os, shutil, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
V088 = os.path.join(ROOT, "promotion_v0.8.8",
                    "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
REV2 = os.path.join(ROOT, "candidate_v0.8.9_rev2",
                    "TTW_College_Football_Power_Ratings_v0.8.9_REV2_CANDIDATE.xlsx")
OUT = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.9_AUTHORITATIVE.xlsx")

V088_SHA = "b2a920feddc0f49f0647957334db0ecd0e922fe6a3933fc6a11af31587b56450"
REV2_SHA = "fcb4d6e63c7ab260b17ffbc47081a14def59bdbd81b4f9cff2194ea1fca18298"

OLD_VER, NEW_VER = "v0.8.8 AUTHORITATIVE", "v0.8.9 AUTHORITATIVE"
OLD_DATE, NEW_DATE = "promotion complete 2026-08-04", "promotion complete 2026-08-27"
KEEP = "0 market lines loaded"


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", (v.text or "").strip())
    if isinstance(v, str) and v.startswith("="):
        return ("F", v.strip())
    return ("V", v)


def main():
    a_sha, r_sha = sha256(V088), sha256(REV2)
    assert a_sha == V088_SHA, f"v0.8.8 mismatch: {a_sha}"
    assert r_sha == REV2_SHA, f"REV 2 mismatch: {r_sha}"
    print(f"v0.8.8 SHA-256 verified : {a_sha}")
    print(f"REV 2  SHA-256 verified : {r_sha}")

    shutil.copyfile(REV2, OUT + ".building.xlsx")
    wb = openpyxl.load_workbook(OUT + ".building.xlsx")
    sh = wb["START HERE"]
    b = sh["A1"].value
    assert OLD_VER in b and OLD_DATE in b and KEEP in b, f"unexpected banner: {b[:200]}"
    nb = b.replace(OLD_VER, NEW_VER).replace(OLD_DATE, NEW_DATE)
    # exactly the two substitutions, nothing else
    assert nb.replace(NEW_VER, OLD_VER).replace(NEW_DATE, OLD_DATE) == b
    assert KEEP in nb and "v0.8.8" not in nb and "2026-08-04" not in nb
    sh["A1"].value = nb
    print(f"banner: version -> {NEW_VER}; date -> {NEW_DATE}; {KEEP!r} retained")

    wb.save(OUT + ".building.xlsx")
    os.replace(OUT + ".building.xlsx", OUT)

    # only the banner may differ from REV 2
    r = openpyxl.load_workbook(REV2)
    p = openpyxl.load_workbook(OUT)
    d = [(s, c.coordinate) for s in r.sheetnames
         for rowr, rowp in zip(r[s].iter_rows(), p[s].iter_rows())
         for c, e in zip(rowr, rowp) if norm(c.value) != norm(e.value)]
    assert d == [("START HERE", "A1")], f"expected only the banner to differ, got {d[:5]}"
    print("confirmed: REV 2 -> v0.8.9 differs by exactly START HERE!A1")

    # independent full diff against v0.8.8
    av = openpyxl.load_workbook(V088)
    changed = []
    for s in av.sheetnames:
        wa, wp = av[s], p[s]
        R = max(wa.max_row, wp.max_row); C = max(wa.max_column, wp.max_column)
        for rr in range(1, R + 1):
            for cc in range(1, C + 1):
                if norm(wa.cell(row=rr, column=cc).value) != norm(wp.cell(row=rr, column=cc).value):
                    changed.append((s, wp.cell(row=rr, column=cc).coordinate))
    print(f"independent diff v0.8.8 -> v0.8.9: {len(changed)} cells")
    assert len(changed) == 1024, f"EXPECTED 1024, COMPUTED {len(changed)} - STOPPING"
    print("count reconciles: 1023 REV 2 cells + 1 banner = 1024")

    print(f"\nwritten: {OUT}")
    print(f"v0.8.9 SHA-256: {sha256(OUT)}")
    assert sha256(V088) == V088_SHA and sha256(REV2) == REV2_SHA, "an input was modified"
    print("v0.8.8 and REV 2 confirmed unmodified")
    return 0


if __name__ == "__main__":
    sys.exit(main())

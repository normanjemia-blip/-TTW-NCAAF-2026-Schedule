#!/usr/bin/env python3
"""v0.8.9 PROMOTION CERTIFICATE — READ-ONLY. Writes nothing.

v0.8.9 = v0.8.8 + the certified 1,023 REV 2 cells + 1 promotion banner cell.
Every fixture is evaluated AT EACH BUILD'S OWN PRODUCTION SETTINGS.
"""
import collections, hashlib, io, json, os, re, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, os.path.join(ROOT, "candidate_v0.8.9_rev2"))
from verify_v089_rev2 import (spread_label, total_label_088, total_label_rev2,  # noqa: E402
                              audit_b12, audit_b13, engine, LIVE_LINES, EXPECTED_CHANGES)

V088 = os.path.join(ROOT, "promotion_v0.8.8",
                    "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
V089 = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.9_AUTHORITATIVE.xlsx")
REV2 = os.path.join(ROOT, "candidate_v0.8.9_rev2",
                    "TTW_College_Football_Power_Ratings_v0.8.9_REV2_CANDIDATE.xlsx")
V088_SHA = "b2a920feddc0f49f0647957334db0ecd0e922fe6a3933fc6a11af31587b56450"
REV2_SHA = "fcb4d6e63c7ab260b17ffbc47081a14def59bdbd81b4f9cff2194ea1fca18298"

PASS, FAIL = [], []


def chk(m, ok, d=""):
    (PASS if ok else FAIL).append(m)
    print(f"  [{'PASS' if ok else 'FAIL'}] {m}" + (f" - {d}" if d else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", (v.text or "").strip())
    if isinstance(v, str) and v.startswith("="):
        return ("F", v.strip())
    return ("V", v)


def txt(v):
    return v.text if isinstance(v, ArrayFormula) else v


def main():
    print("=" * 78)
    print("v0.8.9 PROMOTION CERTIFICATE")
    print("=" * 78)
    h8, h9 = sha256(V088), sha256(V089)
    print(f"  v0.8.8 SHA-256: {h8}")
    print(f"  v0.8.9 SHA-256: {h9}")

    print("\n0. PREDECESSOR AND PROVENANCE")
    chk("0.1 v0.8.8 retains its frozen SHA-256", h8 == V088_SHA, h8[:16])
    chk("0.2 REV 2 candidate unmodified", sha256(REV2) == REV2_SHA)
    chk("0.3 v0.8.9 differs from v0.8.8", h9 != h8)

    a = openpyxl.load_workbook(V088)
    b = openpyxl.load_workbook(V089)
    r = openpyxl.load_workbook(REV2)
    chk("0.4 21 sheets, names and order preserved",
        a.sheetnames == b.sheetnames and len(b.sheetnames) == 21)

    print("\n1. SCOPE — 1,023 CERTIFIED CELLS + 1 BANNER")
    changed = []
    for s in a.sheetnames:
        wa, wb_ = a[s], b[s]
        R = max(wa.max_row, wb_.max_row); C = max(wa.max_column, wb_.max_column)
        for rr in range(1, R + 1):
            for cc in range(1, C + 1):
                if norm(wa.cell(row=rr, column=cc).value) != norm(wb_.cell(row=rr, column=cc).value):
                    changed.append((s, wb_.cell(row=rr, column=cc).coordinate))
    by = collections.Counter(s for s, _ in changed)
    chk("1.1 exactly 1024 cells changed", len(changed) == 1024, str(len(changed)))
    chk("1.2 sheets: SETTINGS 15, ENGINE 1000, AUDIT 4, CHANGELOG 4, START HERE 1",
        dict(by) == {"SETTINGS": 15, "ENGINE": 1000, "AUDIT": 4, "CHANGELOG": 4,
                     "START HERE": 1}, str(dict(by)))
    d = [(s, c.coordinate) for s in r.sheetnames
         for rowr, rowp in zip(r[s].iter_rows(), b[s].iter_rows())
         for c, e in zip(rowr, rowp) if norm(c.value) != norm(e.value)]
    chk("1.3 v0.8.9 differs from certified REV 2 by exactly START HERE!A1",
        d == [("START HERE", "A1")], str(d[:5]))

    print("\n2. BANNER")
    ba, bb = a["START HERE"]["A1"].value, b["START HERE"]["A1"].value
    chk("2.1 declares v0.8.9 AUTHORITATIVE", "v0.8.9 AUTHORITATIVE" in bb)
    chk("2.2 no stale v0.8.8 identifier", "v0.8.8" not in bb)
    chk("2.3 promotion date is the actual America/New_York date 2026-08-27",
        "promotion complete 2026-08-27" in bb)
    chk("2.4 stale 2026-08-04 clause removed", "2026-08-04" not in bb)
    chk("2.5 retains '0 market lines loaded' (MARKET LINES is blank in the repo artifact)",
        "0 market lines loaded" in bb)
    chk("2.6 no other banner change — reversing both tokens reproduces v0.8.8 exactly",
        bb.replace("v0.8.9 AUTHORITATIVE", "v0.8.8 AUTHORITATIVE")
          .replace("promotion complete 2026-08-27", "promotion complete 2026-08-04") == ba)
    ml = b["MARKET LINES"]
    chk("2.7 MARKET LINES is in fact blank, so the banner is truthful",
        sum(1 for rr in range(6, 1006) for cc in (1, 4)
            if ml.cell(row=rr, column=cc).value is not None) == 0)

    Sb = {f"B{i}": b["SETTINGS"].cell(row=i, column=2).value for i in range(3, 55)}
    Sa = {f"B{i}": a["SETTINGS"].cell(row=i, column=2).value for i in range(3, 55)}
    xf, abf = txt(b["ENGINE"]["X6"].value), txt(b["ENGINE"]["AB6"].value)

    print("\n3. SPREAD RULE")
    chk("3.1 spread BET threshold is exactly 1.5", Sb["B10"] == 1.5, str(Sb["B10"]))
    chk("3.2 spread BET toggle is Y", Sb["B11"] == "Y", repr(Sb["B11"]))
    for v in (1.50, -1.50):
        chk(f"3.x edge {v:+.2f} qualifies as BET",
            spread_label(v, Sb, "READY", "", "") == "BET",
            spread_label(v, Sb, "READY", "", ""))
    for v in (1.49, -1.49):
        chk(f"3.x edge {v:+.2f} is NOT BET",
            spread_label(v, Sb, "READY", "", "") != "BET",
            spread_label(v, Sb, "READY", "", ""))
    chk("3.5 QB-gated game with |edge|=4.0 is not BET",
        spread_label(4.0, Sb, "QB UNCERTAIN", "", "") == "INVESTIGATE")

    print("\n4. TOTALS")
    chk("4.1 totals thresholds remain 2.0 / 3.0 / 6.0",
        (Sb["B49"], Sb["B50"], Sb["B51"]) == (2, 3, 6),
        f"{Sb['B49']}/{Sb['B50']}/{Sb['B51']}")
    chk("4.2 totals BET toggle remains N", Sb["B52"] == "N", repr(Sb["B52"]))
    chk("4.3 totals remain disabled (B22/B23 blank)",
        Sb["B22"] is None and Sb["B23"] is None)
    mism = []
    for aa in (1.99, -1.99, 2.00, -2.00, 2.99, -2.99, 3.00, -3.00,
               5.99, -5.99, 6.00, -6.00, 6.01, -6.01, ""):
        if total_label_088(aa, Sa, "READY", "", "") != total_label_rev2(aa, Sb, "READY", "", ""):
            mism.append(aa)
    chk("4.4 totals classifications identical to v0.8.8 at every fixture, "
        "each build at its OWN production config", not mism, str(mism))

    print("\n5. INDEPENDENCE OF THE SPREAD AND TOTALS CONTROLS")
    chk("5.1 ENGINE!X uses B10/B11 and neither B51 nor B52",
        "SETTINGS!$B$10" in xf and "SETTINGS!$B$11" in xf
        and "SETTINGS!$B$51" not in xf and "SETTINGS!$B$52" not in xf)
    chk("5.2 ENGINE!AB uses B49/B50/B51/B52 and neither B10 nor B11",
        all(t in abf for t in ("SETTINGS!$B$49", "SETTINGS!$B$50",
                               "SETTINGS!$B$51", "SETTINGS!$B$52"))
        and "SETTINGS!$B$10" not in abf and "SETTINGS!$B$11" not in abf)
    fl11 = dict(Sb); fl11["B11"] = "N"
    chk("5.3 flipping B11 produces zero totals-label changes",
        not [aa for aa in (2.0, 3.0, 6.0, 6.01, -6.0)
             if total_label_rev2(aa, Sb, "READY", "", "")
             != total_label_rev2(aa, fl11, "READY", "", "")])
    fl52 = dict(Sb); fl52["B52"] = "Y"
    chk("5.4 flipping the totals toggle produces zero spread-label changes",
        not [v for v in (1.5, 3.0, 6.0, -1.5, -6.0)
             if spread_label(v, Sb, "READY", "", "")
             != spread_label(v, fl52, "READY", "", "")])

    print("\n6. AUDIT GUARDS")
    chk("6.1 AUDIT!B12 (spread) returns OK", audit_b12(Sb, xf) == "OK", audit_b12(Sb, xf))
    chk("6.2 AUDIT!B13 (totals) returns OK", audit_b13(Sb, abf) == "OK", audit_b13(Sb, abf))
    chk("6.3 B12 CHECKs on threshold drift", audit_b12(dict(Sb, B10=3), xf) == "CHECK")
    chk("6.4 B13 CHECKs on totals-toggle drift", audit_b13(dict(Sb, B52="Y"), abf) == "CHECK")
    chk("6.5 no intentional non-OK audit result remains in production",
        audit_b12(Sb, xf) == "OK" and audit_b13(Sb, abf) == "OK")

    print("\n7. MODEL OUTPUTS AND THE EIGHT LINED GAMES")
    E8, E9 = engine(a, LIVE_LINES), engine(b, LIVE_LINES)
    proj = [g for g in E8 if (round(E8[g]["model"], 9), E8[g]["edge"], E8[g]["side"], E8[g]["gate"])
            != (round(E9[g]["model"], 9), E9[g]["edge"], E9[g]["side"], E9[g]["gate"])]
    chk("7.1 model spread, edge, side and gate identical across all 761 FBS-v-FBS games",
        not proj, str(proj[:5]))
    moved = set()
    for g in LIVE_LINES:
        if spread_label(E8[g]["edge"], Sa, E8[g]["gate"], "", "") \
           != spread_label(E9[g]["edge"], Sb, E9[g]["gate"], "", ""):
            moved.add(g)
    chk("7.2 exactly four spread labels change", len(moved) == 4, str(len(moved)))
    chk("7.3 they are SAC@EMU, UNC@TCU, SJSU@USC, NMSU@FSU", moved == EXPECTED_CHANGES)
    mem = E9["401862693"]
    chk("7.4 Memphis at UNLV remains LEAN and QB UNCERTAIN",
        mem["gate"] == "QB UNCERTAIN"
        and spread_label(mem["edge"], Sb, mem["gate"], "", "") == "LEAN")

    print("\n8. NOTHING ELSE MOVED")
    for sheet in ("QB VALUES", "IMPORT SCHEDULE", "TEAM MAP", "PRESEASON", "TEAM RATINGS",
                  "MARKET LINES", "ADJUSTMENTS", "CLEAN", "CALC", "DASHBOARD"):
        chk(f"8.x {sheet} byte-identical to v0.8.8", by.get(sheet, 0) == 0, str(by.get(sheet, 0)))
    tm, qb, st = b["TEAM MAP"], b["QB VALUES"], b["SETTINGS"]
    codes, sts, zeros = collections.Counter(), collections.Counter(), 0
    for rr in range(6, 144):
        ab = tm.cell(row=rr, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=rr, column=c).value for c in (4, 6, 8, 10))
        for v in (D, F):
            if v == 0:
                zeros += 1
        G = "" if (D is None or F is None) else F - D
        codes[H] += 1
        sts["UNCERTAIN" if (G == "" or H == "L" or J != st["B3"].value) else "OK"] += 1
    chk("8.11 QB census unchanged at 117 OK / 21 UNCERTAIN",
        (sts["OK"], sts["UNCERTAIN"]) == (117, 21), str(dict(sts)))
    chk("8.12 confidence census unchanged at 76 H / 43 M / 19 L",
        (codes["H"], codes["M"], codes["L"]) == (76, 43, 19), str(dict(codes)))
    chk("8.13 QB zero count unchanged at 234", zeros == 234, str(zeros))
    chk("8.14 rating inputs untouched",
        (Sb["B6"], Sb["B7"], Sb["B12"], Sb["B8"], Sb["B9"]) == (2.5, 0, 2.5, 1, 1.5))

    print("\n9. PRODUCTION POINTERS")
    man = json.load(io.open(os.path.join(ROOT, "PROJECT_MANIFEST.json"), encoding="utf-8"))
    chk("9.1 manifest current_version is v0.8.9", man["current_version"]["version"] == "v0.8.9")
    chk("9.2 manifest sha256 is the final v0.8.9 workbook", man["current_version"]["sha256"] == h9)
    chk("9.3 manifest current_authoritative points at v0.8.9",
        man["current_authoritative"]["source_sha256"] == h9)
    chk("9.4 v0.8.8 preserved as the immediate rollback with its exact SHA",
        V088_SHA in man["current_version"]["supersedes"])
    chk("9.5 older rollback hashes preserved",
        man["rollback"]["source_sha256"].startswith("bbb17b50")
        and man["intermediate_rollback"]["source_sha256"].startswith("661f8ab0"))
    readme = io.open(os.path.join(ROOT, "README.md"), encoding="utf-8").read()
    chk("9.6 README names v0.8.9 and its SHA", "v0.8.9 AUTHORITATIVE" in readme and h9 in readme)
    pol = io.open(os.path.join(ROOT, "phase9a_production_config",
                               "MASTER_AND_WORKING_COPY_POLICY.md"), encoding="utf-8").read()
    chk("9.7 policy names v0.8.9 and preserves v0.8.8 frozen",
        h9 in pol and V088_SHA in pol)
    dry = io.open(os.path.join(ROOT, "phase11_week0_dryrun", "week0_dryrun.py"),
                  encoding="utf-8").read()
    chk("9.8 Week 0 dry run targets v0.8.9", "promotion_v0.8.9" in dry and h9 in dry)

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"v0.8.8 SHA-256 (frozen): {h8}")
    print(f"v0.8.9 SHA-256:          {h9}")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

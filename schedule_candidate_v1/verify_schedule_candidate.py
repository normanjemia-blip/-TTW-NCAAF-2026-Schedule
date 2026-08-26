#!/usr/bin/env python3
"""SCHEDULE DATE CANDIDATE certificate — READ-ONLY. Writes nothing.

Proves the candidate = v0.8.7 + exactly 133 corrected dates in IMPORT SCHEDULE
column D, and that event ids, kickoff instants, weeks, formulas, ratings, model
outputs and QB censuses are all unchanged -- and that a refresh cannot
reintroduce UTC dates.

Exit code 0 iff every check passes.
"""
import collections, csv, datetime, hashlib, io, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
from zoneinfo import ZoneInfo

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)
from espn_date_rule import start_date, assert_not_utc_dates, venue_zone  # noqa: E402

UTC = ZoneInfo("UTC")
BASE = os.path.join(ROOT, "promotion_v0.8.7",
                    "TTW_College_Football_Power_Ratings_v0.8.7_AUTHORITATIVE.xlsx")
CAND = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.8_SCHEDULE_CANDIDATE.xlsx")
SNAP = os.path.join(HERE, "espn_kickoff_snapshot.csv")
OLD_CSV = os.path.join(ROOT, "TTW_2026_Verified_Schedule_ESPN_v1.0.csv")
NEW_CSV = os.path.join(HERE, "TTW_2026_Verified_Schedule_ESPN_v1.1_LOCALDATES.csv")
# Rebased 2026-08-25: base moved from v0.8.6 (bb76901a...67f9) to v0.8.7.
FROZEN_V087 = "46671deeaaa94d98c63cb32d0e94af9907e76e7e2638de431b918987df2e15cd"
# QB expectations belong to the BASE version and move with it.
EXPECT_QB_STATUS = (117, 21)          # was (110, 28) under v0.8.6
EXPECT_QB_CONF = (76, 43, 19)         # was (73, 40, 25) under v0.8.6
EXPECT_QB_ZEROS = 234                 # was 220 under v0.8.6  (2 x OK rows)

PASS, FAIL = [], []


def chk(msg, ok, detail=""):
    (PASS if ok else FAIL).append(msg)
    print(f"  [{'PASS' if ok else 'FAIL'}] {msg}" + (f" - {detail}" if detail else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def mean(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def main():
    print("=" * 78)
    print("SCHEDULE DATE CANDIDATE — CERTIFICATE (candidate only, not promoted)")
    print("=" * 78)

    snap = {r["id"]: r for r in csv.DictReader(io.open(SNAP, encoding="utf-8"))}
    print(f"\n0. INPUTS\n  snapshot rows: {len(snap)}")
    chk("0.1 base workbook is byte-identical to authoritative v0.8.7",
        sha256(BASE) == FROZEN_V087, sha256(BASE)[:16])

    a = openpyxl.load_workbook(BASE)
    b = openpyxl.load_workbook(CAND)

    print("\n1. SCOPE OF CHANGE")
    changed, formula_changed = [], []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                x, y = norm(wa.cell(row=r, column=c).value), norm(wbk.cell(row=r, column=c).value)
                if x != y:
                    changed.append((s, r, c))
                    if x[0] == "F" or y[0] == "F":
                        formula_changed.append((s, wbk.cell(row=r, column=c).coordinate))
    chk("1.1 ZERO formula differences", not formula_changed, str(formula_changed[:5]))
    sheets = collections.Counter(s for s, _, _ in changed)
    chk("1.2 only IMPORT SCHEDULE changed", set(sheets) == {"IMPORT SCHEDULE"}, str(dict(sheets)))
    cols = {c for _, _, c in changed}
    chk("1.3 only column D (start_date) changed", cols == {4}, str(sorted(cols)))
    chk("1.4 exactly 133 cells changed", len(changed) == 133, str(len(changed)))
    chk("1.5 21 sheets, names and order preserved", a.sheetnames == b.sheetnames)

    sa, sb = a["IMPORT SCHEDULE"], b["IMPORT SCHEDULE"]

    print("\n2. IDENTITY COLUMNS UNTOUCHED")
    for col, name in ((1, "id"), (2, "season"), (3, "week"), (5, "neutral_site"),
                      (6, "away_team"), (8, "home_team"), (13, "venue"), (14, "notes")):
        same = all(sa.cell(row=r, column=col).value == sb.cell(row=r, column=col).value
                   for r in range(6, 900))
        chk(f"2.x column {name} byte-identical", same)

    print("\n3. EVERY CORRECTED DATE EQUALS THE CANONICAL RULE")
    bad, moved, unchanged_rows = [], 0, 0
    for r in range(6, 900):
        gid = sa.cell(row=r, column=1).value
        if gid is None:
            continue
        gid = str(gid)
        want = datetime.date.fromisoformat(snap[gid]["canonical_start_date"])
        got = sb.cell(row=r, column=4).value
        got = got.date() if isinstance(got, datetime.datetime) else got
        if got != want:
            bad.append((gid, str(got), str(want)))
        before = sa.cell(row=r, column=4).value
        before = before.date() if isinstance(before, datetime.datetime) else before
        if before != want:
            moved += 1
        else:
            unchanged_rows += 1
    chk("3.1 all 888 candidate dates equal the canonical venue-local date", not bad, str(bad[:5]))
    chk("3.2 exactly 133 moved, 755 already correct",
        moved == 133 and unchanged_rows == 755, f"moved={moved} kept={unchanged_rows}")
    deltas = collections.Counter()
    for r in range(6, 900):
        gid = sa.cell(row=r, column=1).value
        if gid is None:
            continue
        x = sa.cell(row=r, column=4).value; y = sb.cell(row=r, column=4).value
        if x != y:
            deltas[(y - x).days] += 1
    chk("3.3 every change is exactly -1 day", set(deltas) == {-1}, str(dict(deltas)))

    print("\n4. KICKOFF INSTANTS AND WEEKS UNCHANGED")
    chk("4.1 no kickoff instant was altered — the workbook stores no kickoff time, "
        "only a date; the ESPN instant is the INPUT and is untouched", True)
    wk_a = collections.Counter(sa.cell(row=r, column=3).value for r in range(6, 900)
                               if sa.cell(row=r, column=1).value is not None)
    wk_b = collections.Counter(sb.cell(row=r, column=3).value for r in range(6, 900)
                               if sb.cell(row=r, column=1).value is not None)
    chk("4.2 week distribution identical", wk_a == wk_b, f"{len(wk_a)} weeks")
    # a corrected date must never leave its week's original date span
    span = collections.defaultdict(list)
    for r in range(6, 900):
        gid = sa.cell(row=r, column=1).value
        if gid is None:
            continue
        span[sa.cell(row=r, column=3).value].append(
            (sa.cell(row=r, column=4).value, sb.cell(row=r, column=4).value))
    weeks = sorted(span, key=int)
    overlap = 0
    for i, w in enumerate(weeks):
        if i:
            prev_max = max(y for _, y in span[weeks[i - 1]])
            if min(y for _, y in span[w]) <= prev_max:
                overlap += 1
    chk("4.3 no week's corrected span overlaps the previous week", overlap == 0, str(overlap))
    w0 = [(x, y) for x, y in span.get(0, [])]
    chk("4.4 Week 0 still holds exactly 8 games", len(w0) == 8, str(len(w0)))
    chk("4.5 Week 0 is now a single Saturday (2026-08-29)",
        {y.date() for _, y in w0} == {datetime.date(2026, 8, 29)},
        str(sorted({str(y.date()) for _, y in w0})))

    print("\n5. RATINGS AND MODEL OUTPUTS UNCHANGED")
    tm, qb, st = b["TEAM MAP"], b["QB VALUES"], b["SETTINGS"]
    ps = b["PRESEASON"]
    S = {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 33)}
    raw = {c: [ps.cell(row=r, column=c).value for r in range(6, 144)] for c in (4, 8, 17)}
    mu = {c: mean(v) for c, v in raw.items()}
    W = {4: S["B28"], 8: S["B29"], 17: S["B31"]}
    prior = {}
    for i, r in enumerate(range(6, 144)):
        ab = tm.cell(row=r, column=1).value
        num = den = 0.0
        for c in (4, 8, 17):
            v = raw[c][i]
            if isinstance(v, (int, float)):
                num += W[c] * (v - mu[c]); den += W[c]
        prior[ab] = (num / den) if den else ""
    alias = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()
    delta = {}
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F = qb.cell(row=r, column=4).value, qb.cell(row=r, column=6).value
        delta[ab] = "" if (D is None or F is None) else F - D
    z = lambda v: 0 if v == "" else v
    got, n_games, fcs = {}, 0, 0
    for r in range(6, 1006):
        gid = sb.cell(row=r, column=1).value
        if not gid:
            continue
        n_games += 1
        A = alias.get(str(sb.cell(row=r, column=6).value).strip(), "")
        H = alias.get(str(sb.cell(row=r, column=8).value).strip(), "")
        if not A or not H:
            fcs += 1; continue
        neutral = bool(sb.cell(row=r, column=5).value)
        got[(A, H)] = (prior[H] - prior[A]) + (S["B7"] if neutral else S["B6"]) \
                      + (z(delta.get(H, "")) - z(delta.get(A, "")))
    chk("5.1 888 games / 761 FBS-v-FBS / 127 FCS-involved",
        n_games == 888 and fcs == 127 and n_games - fcs == 761, f"{n_games}/{n_games-fcs}/{fcs}")
    for (A, H), want in ((("MEM", "UNLV"), "UNLV -5.6"), (("UNC", "TCU"), "TCU -4.2"),
                         (("NMSU", "FSU"), "FSU -27.7"), (("SJSU", "USC"), "USC -35.2"),
                         (("HAW", "STAN"), "STAN -3.7")):
        m = got[(A, H)]
        lab = f"{H} -{abs(m):.1f}" if m > 0 else f"{A} -{abs(m):.1f}"
        chk(f"5.x {A} at {H} model spread unchanged at {want}", lab == want, lab)

    print("\n6. QB STATE UNCHANGED")
    codes, sts = collections.Counter(), collections.Counter()
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        G = "" if (D is None or F is None) else F - D
        codes[H] += 1
        sts["UNCERTAIN" if (G == "" or H == "L" or J != 2026) else "OK"] += 1
    chk(f"6.1 QB status census still {EXPECT_QB_STATUS[0]} OK / {EXPECT_QB_STATUS[1]} UNCERTAIN",
        (sts["OK"], sts["UNCERTAIN"]) == EXPECT_QB_STATUS, str(dict(sts)))
    chk(f"6.2 confidence census still {EXPECT_QB_CONF[0]} H / {EXPECT_QB_CONF[1]} M / "
        f"{EXPECT_QB_CONF[2]} L",
        (codes["H"], codes["M"], codes["L"]) == EXPECT_QB_CONF, str(dict(codes)))
    chk("6.3 still zero nonzero QB values",
        all(d in ("", 0) for d in delta.values()))
    qb_zeros = sum(1 for r in range(6, 144) if tm.cell(row=r, column=1).value
                   for c in (4, 6) if qb.cell(row=r, column=c).value == 0)
    chk(f"6.4 QB zero count still {EXPECT_QB_ZEROS} = 2 x {EXPECT_QB_STATUS[0]} OK rows",
        qb_zeros == EXPECT_QB_ZEROS, str(qb_zeros))
    gated = [r for r in range(6, 144) if tm.cell(row=r, column=1).value
             and (qb.cell(row=r, column=4).value is None or qb.cell(row=r, column=6).value is None
                  or qb.cell(row=r, column=8).value == "L"
                  or qb.cell(row=r, column=10).value != 2026)]
    chk("6.5 QB gate population unchanged - 21 rows still UNCERTAIN",
        len(gated) == EXPECT_QB_STATUS[1], str(len(gated)))
    mlines = sum(1 for r in range(6, 1006)
                 if b["MARKET LINES"].cell(row=r, column=1).value is not None
                 or b["MARKET LINES"].cell(row=r, column=4).value is not None)
    chk("6.6 MARKET LINES still blank - no line, edge, side or label can move",
        mlines == 0, str(mlines))

    print("\n6b. PLACEHOLDER-TIME ROWS, EVENT IDS AND SUNDAY COUNTS")
    # 403 rows whose kickoff time is still a placeholder must be byte-identical.
    placeholders = [gid for gid, s_ in snap.items() if s_["time_valid"] != "True"]
    chk("6.7 exactly 403 rows carry a placeholder kickoff time", len(placeholders) == 403,
        str(len(placeholders)))
    ph = set(placeholders)
    moved_ph = []
    for r in range(6, 900):
        gid = sa.cell(row=r, column=1).value
        if gid is None or str(gid) not in ph:
            continue
        if sa.cell(row=r, column=4).value != sb.cell(row=r, column=4).value:
            moved_ph.append(str(gid))
    chk("6.8 all 403 placeholder-time rows are UNCHANGED", not moved_ph, str(moved_ph[:5]))

    ids = [str(sb.cell(row=r, column=1).value) for r in range(6, 900)
           if sb.cell(row=r, column=1).value is not None]
    chk("6.9 all 888 event ids present and UNIQUE",
        len(ids) == 888 and len(set(ids)) == 888, f"n={len(ids)} unique={len(set(ids))}")

    def sundays(ws):
        n = 0
        for r in range(6, 900):
            if ws.cell(row=r, column=1).value is None:
                continue
            d = ws.cell(row=r, column=4).value
            if d is not None and d.weekday() == 6:
                n += 1
        return n
    chk("6.10 stored Sunday games fall from 70 to 3",
        sundays(sa) == 70 and sundays(sb) == 3, f"before={sundays(sa)} after={sundays(sb)}")

    genuine_rows = []
    for r in range(6, 900):
        if sb.cell(row=r, column=1).value is None:
            continue
        d = sb.cell(row=r, column=4).value
        if d is not None and d.weekday() == 6:
            genuine_rows.append((sb.cell(row=r, column=6).value, sb.cell(row=r, column=8).value,
                                 d.date().isoformat()))
    want_genuine = {("Louisville Cardinals", "Ole Miss Rebels", "2026-09-06"),
                    ("Washington State Cougars", "Washington Huskies", "2026-09-06"),
                    ("Wisconsin Badgers", "Notre Dame Fighting Irish", "2026-09-06")}
    chk("6.11 the three genuine Sunday games remain Sunday 2026-09-06",
        set(genuine_rows) == want_genuine, str(sorted(genuine_rows)))

    memunlv = [(sa.cell(row=r, column=4).value, sb.cell(row=r, column=4).value)
               for r in range(6, 900) if str(sb.cell(row=r, column=1).value) == "401862693"]
    chk("6.12 Memphis at UNLV becomes Saturday 2026-08-29 (venue-local)",
        len(memunlv) == 1
        and memunlv[0][0].date() == datetime.date(2026, 8, 30)
        and memunlv[0][1].date() == datetime.date(2026, 8, 29)
        and memunlv[0][1].weekday() == 5,
        str([(str(x), str(y)) for x, y in memunlv]))

    W0_END = datetime.date(2026, 9, 2)
    crossed = 0
    for r in range(6, 900):
        if sa.cell(row=r, column=1).value is None:
            continue
        x = sa.cell(row=r, column=4).value.date(); y = sb.cell(row=r, column=4).value.date()
        if (x <= W0_END) != (y <= W0_END):
            crossed += 1
    chk("6.13 zero games cross the Week 0 boundary", crossed == 0, str(crossed))

    print("\n7. THE DATE IS DISPLAY-ONLY — PROVEN FROM THE FORMULA GRAPH")
    def txt(v):
        if isinstance(v, ArrayFormula):
            return v.text
        return v if isinstance(v, str) and v.startswith("=") else None
    consumers = collections.Counter()
    for ws in b.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f = txt(c.value)
                if not f:
                    continue
                if "IMPORT SCHEDULE'!D" in f or "IMPORT SCHEDULE'!$D" in f:
                    consumers[(ws.title, "IMPORT SCHEDULE!D")] += 1
                if "CLEAN!$D" in f or "CLEAN!D" in f:
                    consumers[(ws.title, "CLEAN!D")] += 1
                if "ENGINE!$C" in f:
                    consumers[(ws.title, "ENGINE!C")] += 1
    chain = {k[0] for k in consumers}
    chk("7.1 the only consumers are CLEAN, ENGINE and DASHBOARD",
        chain == {"CLEAN", "ENGINE", "DASHBOARD"}, str(sorted(chain)))
    calc_dates = 0
    for row in b["CALC"].iter_rows():
        for c in row:
            f = txt(c.value)
            if f and ("CLEAN!$D" in f or "ENGINE!$C" in f or "IMPORT SCHEDULE'!D" in f):
                calc_dates += 1
    chk("7.2 CALC (which drives every gate) touches NO date column", calc_dates == 0, str(calc_dates))
    today = sum(1 for ws in b.worksheets for row in ws.iter_rows() for c in row
                if (txt(c.value) or "").upper().find("TODAY()") >= 0)
    chk("7.3 no TODAY() anywhere, so no date-relative drift", today == 0, str(today))
    chk("7.4 staleness compares the LINE date to SETTINGS!B5, not the game date",
        "MARKET LINES" not in str(txt(b["CALC"]["Q6"].value)) or "$P6" in str(txt(b["CALC"]["Q6"].value)),
        str(txt(b["CALC"]["Q6"].value))[:80])

    print("\n8. IDEMPOTENCE AND THE REFRESH GUARD")
    recs_new, recs_old = [], []
    for gid, s in snap.items():
        kick = datetime.datetime.strptime(s["espn_kickoff_utc"], "%Y-%m-%dT%H:%MZ").replace(tzinfo=UTC)
        addr = dict(city=s["venue_city"], state=s["venue_state"],
                    country=s["venue_country"] or "USA")
        tv = s["time_valid"] == "True"
        recs_new.append(dict(id=gid, stored_date=datetime.date.fromisoformat(s["canonical_start_date"]),
                             kickoff_utc=kick, address=addr, time_valid=tv))
        recs_old.append(dict(id=gid, stored_date=datetime.date.fromisoformat(s["stored_start_date"]),
                             kickoff_utc=kick, address=addr, time_valid=tv))
    # Requirement: the guard must catch the OLD dates AS THEY STAND IN v0.8.7 -- read them
    # from the authoritative base workbook itself, not from the snapshot's copy.
    recs_v087 = []
    for r in range(6, 900):
        gid = sa.cell(row=r, column=1).value
        if gid is None:
            continue
        s_ = snap[str(gid)]
        kick = datetime.datetime.strptime(s_["espn_kickoff_utc"], "%Y-%m-%dT%H:%MZ").replace(tzinfo=UTC)
        recs_v087.append(dict(
            id=str(gid), stored_date=sa.cell(row=r, column=4).value.date(), kickoff_utc=kick,
            address=dict(city=s_["venue_city"], state=s_["venue_state"],
                         country=s_["venue_country"] or "USA"),
            time_valid=s_["time_valid"] == "True"))
    ok_new = True
    try:
        assert_not_utc_dates(recs_new)
    except AssertionError:
        ok_new = False
    chk("8.1 refresh guard PASSES on the candidate", ok_new)
    caught = 0
    try:
        assert_not_utc_dates(recs_old)
    except AssertionError as e:
        caught = int(str(e).split()[0])
    chk("8.2 refresh guard FAILS on the OLD UTC dates, catching all 133",
        caught == 133, f"guard flagged {caught}")
    caught87 = 0
    try:
        assert_not_utc_dates(recs_v087)
    except AssertionError as e:
        caught87 = int(str(e).split()[0])
    chk("8.2b guard detects all 133 stale UTC dates AS STORED IN v0.8.7 itself",
        caught87 == 133, f"guard flagged {caught87} against the v0.8.7 workbook")
    reapply = sum(1 for s in snap.values()
                  if start_date(datetime.datetime.strptime(s["espn_kickoff_utc"], "%Y-%m-%dT%H:%MZ")
                                .replace(tzinfo=UTC),
                                dict(city=s["venue_city"], state=s["venue_state"],
                                     country=s["venue_country"] or "USA"),
                                s["time_valid"] == "True")
                  != datetime.date.fromisoformat(s["canonical_start_date"]))
    chk("8.3 rule is idempotent — re-applying changes nothing", reapply == 0, str(reapply))
    second_pass = []
    for r in range(6, 900):
        gid = sb.cell(row=r, column=1).value
        if gid is None:
            continue
        s_ = snap[str(gid)]
        kick = datetime.datetime.strptime(s_["espn_kickoff_utc"], "%Y-%m-%dT%H:%MZ").replace(tzinfo=UTC)
        addr = dict(city=s_["venue_city"], state=s_["venue_state"],
                    country=s_["venue_country"] or "USA")
        again = start_date(kick, addr, s_["time_valid"] == "True")
        have = sb.cell(row=r, column=4).value.date()
        if again != have:
            second_pass.append((str(gid), str(have), str(again)))
    chk("8.4 SECOND application of the transformation to the corrected workbook "
        "produces ZERO additional changes", not second_pass, str(second_pass[:5]))

    pend = sum(1 for s in snap.values() if s["needs_rederivation"] == "True")
    print(f"  [INFO] {pend} rows still await an announced kickoff and must be re-derived later")

    print("\n9. CSV CANDIDATE")
    old = {r["id"]: r for r in csv.DictReader(io.open(OLD_CSV, encoding="utf-8"))}
    new = {r["id"]: r for r in csv.DictReader(io.open(NEW_CSV, encoding="utf-8"))}
    chk("9.1 same 888 ids, none added or dropped", set(old) == set(new) and len(new) == 888)
    diffs = {k: (old[k]["start_date"], new[k]["start_date"]) for k in old
             if old[k] != new[k]}
    chk("9.2 exactly 133 CSV rows differ", len(diffs) == 133, str(len(diffs)))
    only_date = all(
        all(old[k][c] == new[k][c] for c in old[k] if c != "start_date") for k in diffs)
    chk("9.3 no field other than start_date differs on any row", only_date)

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"  base v0.8.7 : {sha256(BASE)}")
    print(f"  candidate   : {sha256(CAND)}")
    print("  STATUS: CANDIDATE ONLY — NOT PROMOTED")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

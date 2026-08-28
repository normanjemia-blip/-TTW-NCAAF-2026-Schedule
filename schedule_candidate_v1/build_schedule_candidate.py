#!/usr/bin/env python3
"""Build the SCHEDULE DATE CANDIDATE. Nothing is promoted.

Produces, from the frozen v0.8.6 production workbook and the ESPN event record:

  espn_kickoff_snapshot.csv   888 rows: id, kickoff UTC, timeValid, venue, zone,
                              canonical local date -- so the candidate is
                              reproducible without re-fetching ESPN
  TTW_2026_Verified_Schedule_ESPN_v1.1_LOCALDATES.csv
                              the corrected schedule file
  TTW_College_Football_Power_Ratings_v0.8.8_SCHEDULE_CANDIDATE.xlsx
                              v0.8.7 with IMPORT SCHEDULE column D corrected
                              and NOTHING else touched

THIS CANDIDATE IS DELIBERATELY INDEPENDENT OF THE QB PROMOTION CHAIN.
It is now rebased onto v0.8.7. When a further QB version is promoted, rebase by
re-running this script with --source pointing at that workbook; the date
correction is orthogonal and applies unchanged.

REBASE HISTORY
  built on v0.8.6 (bb76901a...67f9)  2026-08-25
  rebased  on v0.8.7 (46671dee...15cd) 2026-08-25 -- same 133 cells, unchanged

Run:  python3 schedule_candidate_v1/build_schedule_candidate.py
"""
import argparse, csv, datetime, glob, io, json, os, shutil, sys
from zoneinfo import ZoneInfo

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)
from espn_date_rule import start_date, venue_zone, needs_rederivation  # noqa: E402

UTC = ZoneInfo("UTC")
SRC_CSV = os.path.join(ROOT, "TTW_2026_Verified_Schedule_ESPN_v1.0.csv")
OUT_CSV = os.path.join(HERE, "TTW_2026_Verified_Schedule_ESPN_v1.1_LOCALDATES.csv")
SNAP = os.path.join(HERE, "espn_kickoff_snapshot.csv")
DEFAULT_SRC_XLSX = os.path.join(
    ROOT, "promotion_v0.8.7", "TTW_College_Football_Power_Ratings_v0.8.7_AUTHORITATIVE.xlsx")
OUT_XLSX = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.8_SCHEDULE_CANDIDATE.xlsx")
CACHE = os.environ.get("ESPN_CACHE", os.path.join(
    "/tmp/claude-0/-home-user--TTW-NCAAF-2026-Schedule",
    "93f60580-cc70-5bb0-a25b-86aea5198243/scratchpad/espn"))

DATE_COL = 4          # IMPORT SCHEDULE column D = start_date
FIRST_ROW = 6


def load_espn_events():
    ev = {}
    for f in glob.glob(os.path.join(CACHE, "*.json")):
        for e in json.load(io.open(f, encoding="utf-8")).get("events", []):
            ev[str(e["id"])] = e
    for f in glob.glob(os.path.join(os.path.dirname(CACHE), "ev", "*.json")):
        e = json.load(io.open(f, encoding="utf-8"))
        if e.get("id") and e.get("date") and isinstance(e.get("competitions"), list):
            ev.setdefault(str(e["id"]), e)
    return ev


def build_snapshot(rows):
    """Derive the canonical date for every game and persist the evidence."""
    ev = load_espn_events()
    snap = {}
    out = []
    for r in rows:
        e = ev.get(r["id"])
        if not e:
            raise SystemExit(f"ESPN event {r['id']} not in cache; cannot build candidate")
        comp = e["competitions"][0]
        kick = datetime.datetime.strptime(e["date"], "%Y-%m-%dT%H:%MZ").replace(tzinfo=UTC)
        addr = (comp.get("venue") or {}).get("address") or {}
        tv = bool(comp.get("timeValid"))
        tz = venue_zone(addr)
        canon = start_date(kick, addr, tv)
        snap[r["id"]] = canon
        out.append(dict(
            id=r["id"], espn_kickoff_utc=e["date"], time_valid=tv,
            venue=r["venue"], venue_city=addr.get("city", ""),
            venue_state=addr.get("state", ""), venue_country=addr.get("country", ""),
            venue_tz=str(tz), local_kickoff=kick.astimezone(tz).strftime("%Y-%m-%d %H:%M"),
            canonical_start_date=canon.isoformat(),
            stored_start_date=r["start_date"],
            needs_rederivation=needs_rederivation(tv)))
    with io.open(SNAP, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(out[0]))
        w.writeheader(); w.writerows(out)
    print(f"  snapshot written: {SNAP} ({len(out)} rows)")
    return snap


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", default=DEFAULT_SRC_XLSX,
                    help="workbook to rebase onto (default: frozen v0.8.6)")
    args = ap.parse_args()

    print("=" * 78)
    print("SCHEDULE DATE CANDIDATE — BUILD (nothing promoted)")
    print("=" * 78)
    rows = list(csv.DictReader(io.open(SRC_CSV, encoding="utf-8")))
    print(f"  source CSV rows: {len(rows)}")
    snap = build_snapshot(rows)

    # ---------- corrected CSV ----------
    changed = 0
    for r in rows:
        canon = snap[r["id"]].isoformat()
        if r["start_date"] != canon:
            r["start_date"] = canon
            changed += 1
    with io.open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0]))
        w.writeheader(); w.writerows(rows)
    print(f"  corrected CSV written: {OUT_CSV}")
    print(f"  CSV rows changed: {changed}")

    # ---------- corrected workbook: IMPORT SCHEDULE column D ONLY ----------
    import openpyxl
    wb = openpyxl.load_workbook(args.source)
    sch = wb["IMPORT SCHEDULE"]
    wrote = 0
    seen = set()
    for r in range(FIRST_ROW, FIRST_ROW + 1200):
        gid = sch.cell(row=r, column=1).value
        if gid is None:
            continue
        gid = str(gid)
        if gid not in snap:
            raise SystemExit(f"workbook row {r} has id {gid} not present in the snapshot")
        seen.add(gid)
        cell = sch.cell(row=r, column=DATE_COL)
        want = datetime.datetime.combine(snap[gid], datetime.time())
        if cell.value != want:
            cell.value = want
            wrote += 1
    missing = set(snap) - seen
    assert not missing, f"{len(missing)} scheduled ids absent from the workbook: {list(missing)[:5]}"
    assert wrote == changed, f"workbook changed {wrote} cells but CSV changed {changed}"
    tmp = OUT_XLSX + ".building.xlsx"
    wb.save(tmp); os.replace(tmp, OUT_XLSX)
    print(f"  candidate workbook written: {OUT_XLSX}")
    print(f"  IMPORT SCHEDULE!D cells changed: {wrote}")
    print("\n  NOT PROMOTED. Run verify_schedule_candidate.py before any approval.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

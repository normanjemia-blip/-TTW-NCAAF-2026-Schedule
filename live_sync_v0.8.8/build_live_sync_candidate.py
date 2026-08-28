#!/usr/bin/env python3
"""Build the LIVE-SYNC CANDIDATE from the exported live Sheet + authoritative v0.8.8.

BASE   : the live Google Sheet, exported as XLSX (never written back)
SOURCE : promotion_v0.8.8/...AUTHORITATIVE.xlsx

APPROVED RULINGS APPLIED
  1-3  MARKET LINES (72), CHANGELOG (20) and SETTINGS!B4/B5 are PRESERVED - never written.
  4    Six owner-authored QB note cells are PRESERVED:
         Fresno State I75 K75 L75 · Tulane L91 · Northern Illinois I123 L123
  5    Every other approved synchronization cell is applied.
  6    Banner = the authoritative v0.8.8 banner with its market-line statement
       corrected from "0 market lines loaded" to "8 market lines loaded".
  7    No other banner or metadata correction is made.

EXPECTED WRITE COUNT: 246 = 252 proposed - 6 preserved owner notes.
The build STOPS if the independently recomputed count is not exactly 246.

Writes only the local candidate file. Never touches the live Sheet or v0.8.8.
"""
import csv, hashlib, io, os, shutil, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
AUTH = os.path.join(ROOT, "promotion_v0.8.8",
                    "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
CSV_IN = os.path.join(HERE, "live_sync_cells_v0.8.8.csv")
OUT = os.path.join(HERE, "TTW_LIVE_SYNC_CANDIDATE_v0.8.8.xlsx")

AUTH_SHA = "b2a920feddc0f49f0647957334db0ecd0e922fe6a3933fc6a11af31587b56450"
LIVE_SHA = "78d7151c20052535455bac200db0eae55976816040a9cea6eaf2179f38aca3b3"

PRESERVE_CELLS = {("QB VALUES", "I75"), ("QB VALUES", "K75"), ("QB VALUES", "L75"),
                  ("QB VALUES", "L91"), ("QB VALUES", "I123"), ("QB VALUES", "L123")}
BANNER_OLD = "0 market lines loaded"
BANNER_NEW = "8 market lines loaded"


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def main():
    live_path = sys.argv[1]
    got_live, got_auth = sha256(live_path), sha256(AUTH)
    assert got_auth == AUTH_SHA, f"authoritative v0.8.8 is not expected: {got_auth}"
    assert got_live == LIVE_SHA, f"live export is not the audited one: {got_live}"
    print(f"live export SHA-256 verified : {got_live}")
    print(f"authoritative SHA-256 verified: {got_auth}")

    rows = list(csv.DictReader(io.open(CSV_IN, encoding="utf-8")))
    sync = [d for d in rows if d["expected_action"].startswith("OVERWRITE")]
    assert len(sync) == 252, f"expected 252 proposed sync cells, got {len(sync)}"
    present = {(d["sheet"], d["cell"]) for d in sync}
    assert PRESERVE_CELLS <= present, "a preserved owner cell is not in the sync set"
    writes = [d for d in sync if (d["sheet"], d["cell"]) not in PRESERVE_CELLS]
    assert len(writes) == 246, f"EXPECTED 246 writes, computed {len(writes)} - STOPPING"
    print(f"write count independently recomputed: {len(writes)} (252 - 6 preserved)")

    shutil.copyfile(live_path, OUT + ".building.xlsx")
    cand = openpyxl.load_workbook(OUT + ".building.xlsx")
    auth = openpyxl.load_workbook(AUTH)
    live = openpyxl.load_workbook(live_path)

    banner_written = 0
    applied = 0
    for d in writes:
        s, r, c = d["sheet"], int(d["row"]), int(d["column"])
        if s == "START HERE" and d["cell"] == "A1":
            b = auth["START HERE"]["A1"].value
            assert BANNER_OLD in b, f"authoritative banner lacks {BANNER_OLD!r}"
            nb = b.replace(BANNER_OLD, BANNER_NEW)
            # exactly the one substitution, nothing else
            assert nb.replace(BANNER_NEW, BANNER_OLD) == b
            assert "v0.8.8 AUTHORITATIVE" in nb and "76 H / 43 M / 19 L" in nb
            cand["START HERE"]["A1"].value = nb
            banner_written += 1
        else:
            cand[s].cell(row=r, column=c).value = auth[s].cell(row=r, column=c).value
        applied += 1
    assert applied == 246 and banner_written == 1
    print(f"applied {applied} cells (banner: {banner_written})")

    # preserved cells must still equal the LIVE originals
    for s, coord in PRESERVE_CELLS:
        assert cand[s][coord].value == live[s][coord].value, f"{s}!{coord} was not preserved"
    for s in ("MARKET LINES", "CHANGELOG"):
        for row_l, row_c in zip(live[s].iter_rows(), cand[s].iter_rows()):
            for a, b in zip(row_l, row_c):
                assert a.value == b.value or isinstance(a.value, ArrayFormula), \
                    f"{s}!{a.coordinate} was modified"
    for coord in ("B4", "B5"):
        assert cand["SETTINGS"][coord].value == live["SETTINGS"][coord].value
    print("preserved regions confirmed identical to the live export")

    cand.save(OUT + ".building.xlsx")
    os.replace(OUT + ".building.xlsx", OUT)
    print(f"written: {OUT}")
    print(f"candidate SHA-256: {sha256(OUT)}")
    assert sha256(AUTH) == AUTH_SHA, "authoritative v0.8.8 was modified"
    assert sha256(live_path) == LIVE_SHA, "the live export was modified"
    print("authoritative v0.8.8 and the live export both confirmed unmodified")
    return 0


if __name__ == "__main__":
    sys.exit(main())

#!/usr/bin/env python3
"""LIVE-SYNC CANDIDATE certificate — READ-ONLY. Writes nothing.

Proves the candidate = live export + exactly 246 approved cells, with every
preserved region byte-identical, and reports the operational dashboard/gate/
label impact of the approved QB activations while the 8 live market lines
remain in place.
"""
import collections, csv, hashlib, io, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
AUTH = os.path.join(ROOT, "promotion_v0.8.8",
                    "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
CAND = os.path.join(HERE, "TTW_LIVE_SYNC_CANDIDATE_v0.8.8.xlsx")
CSV_IN = os.path.join(HERE, "live_sync_cells_v0.8.8.csv")
AUTH_SHA = "b2a920feddc0f49f0647957334db0ecd0e922fe6a3933fc6a11af31587b56450"
LIVE_SHA = "78d7151c20052535455bac200db0eae55976816040a9cea6eaf2179f38aca3b3"
PRESERVE_CELLS = {("QB VALUES", "I75"), ("QB VALUES", "K75"), ("QB VALUES", "L75"),
                  ("QB VALUES", "L91"), ("QB VALUES", "I123"), ("QB VALUES", "L123")}

PASS, FAIL = [], []


def chk(m, ok, d=""):
    (PASS if ok else FAIL).append(m)
    print(f"  [{'PASS' if ok else 'FAIL'}] {m}" + (f" - {d}" if d else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", (v.text or "").strip())
    if isinstance(v, str) and v.startswith("="):
        return ("F", v.strip())
    return ("V", v)


def isf(v):
    return norm(v)[0] == "F"


def mean(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def engine_rows(wb):
    """Re-derive model spread, market edge, side and label for every game."""
    tm, qb, st, ps = wb["TEAM MAP"], wb["QB VALUES"], wb["SETTINGS"], wb["PRESEASON"]
    S = {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 33)}
    raw = {c: [ps.cell(row=r, column=c).value for r in range(6, 144)] for c in (4, 8, 17)}
    mu = {c: mean(v) for c, v in raw.items()}
    W = {4: S["B28"], 8: S["B29"], 17: S["B31"]}
    prior = {}
    for i, r in enumerate(range(6, 144)):
        ab = tm.cell(row=r, column=1).value
        num = den = 0.0
        for c in (4, 8, 17):
            v = raw[c][i]
            if isinstance(v, (int, float)):
                num += W[c] * (v - mu[c]); den += W[c]
        prior[ab] = (num / den) if den else ""
    alias = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()
    season = S["B3"]
    delta, status = {}, {}
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        g = "" if (D is None or F is None) else F - D
        delta[ab] = g
        status[ab] = "UNCERTAIN" if (g == "" or H == "L" or J != season) else "OK"
    lines = {}
    ml = wb["MARKET LINES"]
    for r in range(6, 1006):
        gid = ml.cell(row=r, column=1).value
        if gid is None:
            continue
        lines[str(gid)] = dict(fav=ml.cell(row=r, column=3).value,
                               spread=ml.cell(row=r, column=4).value,
                               total=ml.cell(row=r, column=5).value)
    z = lambda v: 0 if v == "" else v
    sch = wb["IMPORT SCHEDULE"]
    out = {}
    n = fcs = 0
    for r in range(6, 1006):
        gid = sch.cell(row=r, column=1).value
        if not gid:
            continue
        n += 1
        A = alias.get(str(sch.cell(row=r, column=6).value).strip(), "")
        H = alias.get(str(sch.cell(row=r, column=8).value).strip(), "")
        if not A or not H:
            fcs += 1
            continue
        neutral = bool(sch.cell(row=r, column=5).value)
        R = (prior[H] - prior[A]) + (S["B7"] if neutral else S["B6"]) \
            + (z(delta.get(H, "")) - z(delta.get(A, "")))
        lab = f"{H} -{abs(R):.1f}" if R > 0 else (f"{A} -{abs(R):.1f}" if R < 0 else "PICK")
        ln = lines.get(str(gid))
        edge = side = mlabel = ""
        if ln and isinstance(ln["spread"], (int, float)) and ln["fav"]:
            T = -abs(ln["spread"]) if ln["fav"] == H else (abs(ln["spread"]) if ln["fav"] == A else "")
            if T != "":
                V = R + T
                edge = V
                side = H if V > 0 else A
                a = abs(V)
                mlabel = ("BET" if a >= S["B10"] else
                          "INVESTIGATE" if a >= S["B9"] else
                          "LEAN" if a >= S["B8"] else "PASS")
        gate = "QB UNCERTAIN" if (status.get(A) == "UNCERTAIN" or status.get(H) == "UNCERTAIN") else "READY"
        out[str(gid)] = dict(away=A, home=H, model=R, label=lab, edge=edge, side=side,
                             mlabel=mlabel, gate=gate, market=ln, S=S)
    out["_census"] = (n, n - fcs, fcs)
    return out


def main():
    live_path = sys.argv[1]
    print("=" * 78)
    print("LIVE-SYNC CANDIDATE CERTIFICATE (candidate only - live Sheet NOT written)")
    print("=" * 78)
    hc = sha256(CAND)
    print(f"  live export SHA-256 : {sha256(live_path)}")
    print(f"  authoritative SHA   : {sha256(AUTH)}")
    print(f"  candidate SHA-256   : {hc}")

    print("\n0. INPUTS UNMODIFIED")
    chk("0.1 live export unmodified", sha256(live_path) == LIVE_SHA)
    chk("0.2 authoritative v0.8.8 unmodified", sha256(AUTH) == AUTH_SHA)

    live = openpyxl.load_workbook(live_path)
    cand = openpyxl.load_workbook(CAND)
    auth = openpyxl.load_workbook(AUTH)

    print("\n1. SCOPE: EXACTLY 246 CELLS CHANGED FROM THE LIVE BASE")
    changed, fdiff = [], []
    for s in live.sheetnames:
        a, b = live[s], cand[s]
        R = max(a.max_row, b.max_row); C = max(a.max_column, b.max_column)
        for r in range(1, R + 1):
            for c in range(1, C + 1):
                x, y = norm(a.cell(row=r, column=c).value), norm(b.cell(row=r, column=c).value)
                if x == y:
                    continue
                coord = b.cell(row=r, column=c).coordinate
                (fdiff if (x[0] == "F" or y[0] == "F") else changed).append((s, coord))
    chk("1.1 exactly 246 cells changed", len(changed) == 246, str(len(changed)))
    chk("1.2 ZERO formula changes", not fdiff, str(fdiff[:5]))
    by = collections.Counter(s for s, _ in changed)
    chk("1.3 changes confined to IMPORT SCHEDULE (133), QB VALUES (112), START HERE (1)",
        dict(by) == {"IMPORT SCHEDULE": 133, "QB VALUES": 112, "START HERE": 1}, str(dict(by)))

    print("\n2. FORMULA POPULATION")
    nf = sum(1 for ws in cand.worksheets for row in ws.iter_rows() for c in row if isf(c.value))
    nl = sum(1 for ws in live.worksheets for row in ws.iter_rows() for c in row if isf(c.value))
    chk("2.1 123,011 formulas present and byte-identical to the live base",
        nf == 123011 and nl == 123011 and not fdiff, f"cand={nf} live={nl}")

    print("\n3. PRESERVED REGIONS BYTE-IDENTICAL TO LIVE")
    ml_diff = [c.coordinate for rl, rc in zip(live["MARKET LINES"].iter_rows(),
                                              cand["MARKET LINES"].iter_rows())
               for c, d in zip(rl, rc) if norm(c.value) != norm(d.value)]
    chk("3.1 all MARKET LINES cells preserved exactly", not ml_diff, str(ml_diff[:5]))
    mlrows = [r for r in range(6, 1006) if cand["MARKET LINES"].cell(row=r, column=1).value is not None]
    chk("3.2 all eight market lines still present", len(mlrows) == 8, str(len(mlrows)))
    # The authoritative artifact ships MARKET LINES blank, so every populated live
    # value cell shows up as a difference against it. That count must be exactly 72,
    # across GameID/Favorite/Spread/Total/Source/Line date/Notes/Circa spread/Circa total.
    vals = sum(1 for r in range(1, 1006) for c in range(1, 16)
               if norm(cand["MARKET LINES"].cell(row=r, column=c).value)
               != norm(auth["MARKET LINES"].cell(row=r, column=c).value))
    chk("3.3 the 72 live market-line value cells intact (vs the blank authoritative sheet)",
        vals == 72, str(vals))
    cl_diff = [c.coordinate for rl, rc in zip(live["CHANGELOG"].iter_rows(),
                                              cand["CHANGELOG"].iter_rows())
               for c, d in zip(rl, rc) if norm(c.value) != norm(d.value)]
    chk("3.4 all 20 live CHANGELOG cells preserved exactly", not cl_diff, str(cl_diff[:5]))
    chk("3.5 SETTINGS!B4 preserved",
        cand["SETTINGS"]["B4"].value == live["SETTINGS"]["B4"].value,
        repr(cand["SETTINGS"]["B4"].value))
    chk("3.6 SETTINGS!B5 preserved",
        cand["SETTINGS"]["B5"].value == live["SETTINGS"]["B5"].value,
        str(cand["SETTINGS"]["B5"].value))
    for s, coord in sorted(PRESERVE_CELLS):
        chk(f"3.x owner note {s}!{coord} preserved",
            cand[s][coord].value == live[s][coord].value)

    print("\n4. BANNER")
    b = cand["START HERE"]["A1"].value
    chk("4.1 banner declares v0.8.8 AUTHORITATIVE", "v0.8.8 AUTHORITATIVE" in b)
    chk("4.2 banner census reads 76 H / 43 M / 19 L", "76 H / 43 M / 19 L" in b)
    chk("4.3 banner market-line statement corrected to 8 market lines loaded",
        "8 market lines loaded" in b and "0 market lines loaded" not in b)
    chk("4.4 banner is otherwise the authoritative v0.8.8 text",
        b.replace("8 market lines loaded", "0 market lines loaded")
        == auth["START HERE"]["A1"].value)

    print("\n5. QB STATE")
    tm, qb, st = cand["TEAM MAP"], cand["QB VALUES"], cand["SETTINGS"]
    season = st["B3"].value
    codes, sts, zeros, nonzero = collections.Counter(), collections.Counter(), 0, []
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        for v in (D, F):
            if v == 0:
                zeros += 1
            elif v is not None:
                nonzero.append((ab, v))
        G = "" if (D is None or F is None) else F - D
        codes[H] += 1
        sts["UNCERTAIN" if (G == "" or H == "L" or J != season) else "OK"] += 1
    chk("5.1 QB status census 117 OK / 21 UNCERTAIN",
        (sts["OK"], sts["UNCERTAIN"]) == (117, 21), str(dict(sts)))
    chk("5.2 confidence census 76 H / 43 M / 19 L",
        (codes["H"], codes["M"], codes["L"]) == (76, 43, 19), str(dict(codes)))
    chk("5.3 QB zeros 234", zeros == 234, str(zeros))
    chk("5.4 zero nonzero QB values", not nonzero, str(nonzero[:5]))

    print("\n6. SCHEDULE")
    sc, sa = cand["IMPORT SCHEDULE"], auth["IMPORT SCHEDULE"]
    mism = [str(sc.cell(row=r, column=1).value) for r in range(6, 900)
            if sc.cell(row=r, column=1).value is not None
            and sc.cell(row=r, column=4).value != sa.cell(row=r, column=4).value]
    chk("6.1 all 888 schedule dates equal authoritative v0.8.8", not mism, str(mism[:5]))
    sl = live["IMPORT SCHEDULE"]
    moved = [r for r in range(6, 900) if sc.cell(row=r, column=1).value is not None
             and sl.cell(row=r, column=4).value != sc.cell(row=r, column=4).value]
    chk("6.2 exactly the 133 approved schedule dates were applied", len(moved) == 133, str(len(moved)))
    deltas = collections.Counter((sc.cell(row=r, column=4).value
                                  - sl.cell(row=r, column=4).value).days for r in moved)
    chk("6.3 every applied date moved exactly -1 day", set(deltas) == {-1}, str(dict(deltas)))

    print("\n7. MODEL OUTPUTS")
    E = engine_rows(cand)
    n, fbs, fcs = E.pop("_census")
    chk("7.1 888 games / 761 FBS-v-FBS / 127 FCS", (n, fbs, fcs) == (888, 761, 127),
        f"{n}/{fbs}/{fcs}")
    ref = {("MEM", "UNLV"): "UNLV -5.6", ("UNC", "TCU"): "TCU -4.2",
           ("NMSU", "FSU"): "FSU -27.7", ("SJSU", "USC"): "USC -35.2",
           ("HAW", "STAN"): "STAN -3.7"}
    bykey = {(v["away"], v["home"]): v for v in E.values()}
    for k, want in ref.items():
        chk(f"7.x {k[0]} at {k[1]} spread {want}", bykey[k]["label"] == want, bykey[k]["label"])

    print("\n8. NO UNEXPLAINED DRIFT")
    rows = list(csv.DictReader(io.open(CSV_IN, encoding="utf-8")))
    approved = {(d["sheet"], d["cell"]) for d in rows
                if d["expected_action"].startswith("OVERWRITE")} - PRESERVE_CELLS
    chk("8.1 every changed cell is an approved synchronization cell",
        set(changed) == approved, str(sorted(set(changed) - approved)[:5]))

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"candidate SHA-256: {hc}")
    print("STATUS: CANDIDATE ONLY - the live Google Sheet was NOT written")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

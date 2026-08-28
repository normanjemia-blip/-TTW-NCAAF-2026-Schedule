#!/usr/bin/env python3
"""READ-ONLY live-Google-Sheet synchronization packet for authoritative v0.8.8.

Compares the live production master, exported from Drive as XLSX, against
promotion_v0.8.8/...AUTHORITATIVE.xlsx cell by cell.

WRITES NOTHING to the live Sheet, the authoritative workbook, or any pointer.
It emits a CSV and summary statistics for approval.

Usage:  python3 live_sync_v0.8.8/build_live_sync_packet.py <live_export.xlsx>
"""
import collections, csv, io, os, sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
AUTH = os.path.join(ROOT, "promotion_v0.8.8",
                    "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
V084 = os.path.join(ROOT, "promotion_v0.8.4",
                    "TTW_College_Football_Power_Ratings_v0.8.4_AUTHORITATIVE.xlsx")
OUT_CSV = os.path.join(HERE, "live_sync_cells_v0.8.8.csv")

# Live-only regions: data the live Sheet legitimately holds that the repo
# artifact does not ship. Overwriting these would DESTROY live data.
LIVE_ONLY = {
    "MARKET LINES": "operational Week 0 market lines (repo ships MARKET LINES blank by design)",
    "CHANGELOG": "live-authored changelog history rows",
}
LIVE_ONLY_CELLS = {("SETTINGS", "B4"): "results-through-week operational setting",
                   ("SETTINGS", "B5"): "run date used by line-staleness"}


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", (v.text or "").strip())
    if isinstance(v, str) and v.startswith("="):
        return ("F", v.strip())
    return ("V", v)


def show(v):
    k, x = norm(v)
    if x is None:
        return ""
    return str(x)


def main():
    live_path = sys.argv[1]
    lw = openpyxl.load_workbook(live_path)
    aw = openpyxl.load_workbook(AUTH)
    rw = openpyxl.load_workbook(V084)

    # which QB rows moved between v0.8.4 and v0.8.8 in the repo, and how
    qb_a, qb_r = aw["QB VALUES"], rw["QB VALUES"]
    tmap = aw["TEAM MAP"]
    qb_kind = {}
    for r in range(6, 144):
        D_a, F_a = qb_a.cell(row=r, column=4).value, qb_a.cell(row=r, column=6).value
        D_r, F_r = qb_r.cell(row=r, column=4).value, qb_r.cell(row=r, column=6).value
        activated = (D_r is None and D_a == 0) or (F_r is None and F_a == 0)
        qb_kind[r] = "QB ACTIVATION" if activated else "QB RECORD CORRECTION"

    team_by_row = {r: tmap.cell(row=r, column=2).value for r in range(6, 144)}
    sched_a = aw["IMPORT SCHEDULE"]
    sched_team = {r: (sched_a.cell(row=r, column=1).value,
                      sched_a.cell(row=r, column=6).value,
                      sched_a.cell(row=r, column=8).value) for r in range(6, 900)}

    rows, formula_diffs = [], []
    for s in aw.sheetnames:
        la, aa = lw[s], aw[s]
        R = max(la.max_row, aa.max_row)
        C = max(la.max_column, aa.max_column)
        for r in range(1, R + 1):
            for c in range(1, C + 1):
                lv, av = la.cell(row=r, column=c).value, aa.cell(row=r, column=c).value
                x, y = norm(lv), norm(av)
                if x == y:
                    continue
                coord = f"{get_column_letter(c)}{r}"
                if x[0] == "F" or y[0] == "F":
                    formula_diffs.append((s, coord, str(x[1])[:120], str(y[1])[:120]))
                    continue

                # ---- classify ----
                context = ""
                if (s, coord) in LIVE_ONLY_CELLS:
                    cls, action = "LIVE-ONLY DATA", "HOLD - ruling required"
                    context = LIVE_ONLY_CELLS[(s, coord)]
                elif s in LIVE_ONLY:
                    cls, action = "LIVE-ONLY DATA", "HOLD - ruling required"
                    context = LIVE_ONLY[s]
                elif s == "START HERE" and coord == "A1":
                    cls, action = "BANNER", "OVERWRITE with v0.8.8"
                elif s == "IMPORT SCHEDULE" and c == 4:
                    cls, action = "SCHEDULE DATE", "OVERWRITE with v0.8.8"
                    gid, aw_t, hm_t = sched_team.get(r, ("", "", ""))
                    context = f"{gid} {aw_t} @ {hm_t}"
                elif s == "QB VALUES":
                    cls = qb_kind.get(r, "QB RECORD CORRECTION")
                    action = "OVERWRITE with v0.8.8"
                    context = team_by_row.get(r, "")
                else:
                    cls, action = "UNEXPECTED DRIFT", "STOP - investigate"
                rows.append(dict(
                    sheet=s, cell=coord, row=r, column=c,
                    column_letter=get_column_letter(c),
                    context=context,
                    live_value=show(lv), authoritative_value=show(av),
                    expected_action=action, classification=cls))

    rows.sort(key=lambda d: (d["sheet"], d["row"], d["column"]))
    with io.open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0]))
        w.writeheader()
        w.writerows(rows)

    print("=" * 78)
    print("LIVE-SHEET SYNCHRONIZATION PACKET — v0.8.8 (READ-ONLY)")
    print("=" * 78)
    print(f"  live export : {os.path.basename(live_path)}")
    print(f"  authoritative: {os.path.basename(AUTH)}")
    print()
    print(f"  FORMULA DIFFERENCES: {len(formula_diffs)}")
    for s, coord, a, b in formula_diffs[:10]:
        print(f"    {s}!{coord}\n      LIVE: {a}\n      AUTH: {b}")
    print()
    byc = collections.Counter(d["classification"] for d in rows)
    print("  CELLS BY CLASSIFICATION")
    for k, v in byc.most_common():
        print(f"    {k:<24}{v}")
    print(f"    {'TOTAL DIFFERING CELLS':<24}{len(rows)}")
    print()
    sync = [d for d in rows if d["expected_action"].startswith("OVERWRITE")]
    hold = [d for d in rows if not d["expected_action"].startswith("OVERWRITE")]
    print(f"  CELLS REQUIRING SYNCHRONIZATION (overwrite): {len(sync)}")
    print(f"  CELLS HELD PENDING RULING (live-only/drift) : {len(hold)}")
    print()
    print("  BY SHEET")
    for s, n in collections.Counter(d["sheet"] for d in rows).most_common():
        so = sum(1 for d in rows if d["sheet"] == s and d["expected_action"].startswith("OVERWRITE"))
        print(f"    {s:<18} total {n:<5} sync {so:<5} hold {n - so}")
    print()
    qbrows = sorted({d["row"] for d in rows if d["sheet"] == "QB VALUES"})
    print(f"  QB VALUES rows touched ({len(qbrows)}): "
          + ", ".join(f"{r}:{team_by_row.get(r,'')}" for r in qbrows))
    print(f"\n  CSV written: {OUT_CSV}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

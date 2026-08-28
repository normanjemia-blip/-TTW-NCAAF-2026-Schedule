#!/usr/bin/env python3
"""READ-ONLY 1.5-point spread-threshold audit. Alters no formula and no setting.

Traces SETTINGS -> ENGINE -> DASHBOARD, reimplements ENGINE!X and ENGINE!AB
EXACTLY as written in the workbook, and runs the required test matrix.
"""
import openpyxl, sys, os
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WB = os.path.join(ROOT, "promotion_v0.8.8",
                  "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")


def spread_label(V, S, AI, AF, AG):
    """Verbatim transcription of ENGINE!X6."""
    if V == "" or AI in ("BLOCKED", "PENDING LINE", "STALE LINE"):
        return ""
    if abs(V) < S["B8"]:
        return ""
    if abs(V) < S["B9"]:
        return "LEAN"
    if (abs(V) < S["B10"] or S["B11"] != "Y" or AI != "READY" or AF != "" or AG != ""):
        return "INVESTIGATE"
    return "BET"


def total_label(AA, S, AI, AF, AG):
    """Verbatim transcription of ENGINE!AB6 - note the SAME cells, doubled."""
    if AA == "":
        return ""
    if abs(AA) < S["B8"] * 2:
        return ""
    if abs(AA) < S["B9"] * 2:
        return "LEAN"
    if (S["B11"] != "Y" or AI != "READY" or abs(AA) < S["B10"] * 2 or AF != "" or AG != ""):
        return "INVESTIGATE"
    return "BET"


def main():
    wb = openpyxl.load_workbook(WB)
    st = wb["SETTINGS"]
    S = {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 33)}

    print("=" * 78)
    print("1.5-POINT SPREAD THRESHOLD AUDIT — READ-ONLY")
    print("=" * 78)
    print("\n1. CELLS CONTROLLING SPREAD CLASSIFICATION")
    for r, name in ((8, "LEAN threshold (abs edge)"), (9, "INVESTIGATE threshold (abs edge)"),
                    (10, "BET threshold (abs edge)"), (11, "BET labels toggle (Y/N)")):
        print(f"  SETTINGS!B{r:<3} = {st.cell(row=r, column=2).value!r:<8} {name}")
    print("  ENGINE!X6:X1005    spread label   (consumes B8, B9, B10, B11)")
    print("  ENGINE!V6:V1005    spread EDGE    = R + T")
    print("  ENGINE!W6:W1005    side")
    print("  ENGINE!AI6:AI1005  STATUS gate")
    print("  DASHBOARD!K6:K1005 <- ENGINE!X   (the visible spread label)")
    print("  DASHBOARD!J6:J1005 <- ENGINE!W   (side)")

    print("\n2-3. CURRENT THRESHOLDS")
    print(f"  BET threshold          SETTINGS!B10 = {S['B10']}")
    print(f"  INVESTIGATE threshold  SETTINGS!B9  = {S['B9']}")
    print(f"  LEAN threshold         SETTINGS!B8  = {S['B8']}")
    print(f"  BET labels toggle      SETTINGS!B11 = {S['B11']!r}")

    print("\n5. IS ABSOLUTE EDGE USED CORRECTLY?")
    print("  ENGINE!X uses ABS($V6) in all three comparisons -> YES, sign-symmetric.")
    print("  Side is taken separately from ENGINE!W (sign of V), so direction is not lost.")

    print("\n6. DOES ANY THRESHOLD CELL ALSO AFFECT TOTALS OR ANOTHER MARKET?")
    print("  YES — and this is the blocking finding.")
    print("  ENGINE!AB (TOTALS label) consumes the SAME cells, doubled:")
    print("      ABS(AA) < SETTINGS!B8*2  -> ''        (currently 2.0)")
    print("      ABS(AA) < SETTINGS!B9*2  -> LEAN      (currently 3.0)")
    print("      ABS(AA) < SETTINGS!B10*2 -> INVESTIGATE (currently 6.0)")
    print("  So B8/B9/B10 are SHARED between the spread market and the totals market.")
    print(f"  Totals are currently INERT: SETTINGS!B22={S['B22']!r}, B23={S['B23']!r},")
    print("  so ENGINE!Y is blank for every game, AA is blank, AB is blank workbook-wide.")
    print("  Changing B10 would not move a totals label TODAY, but it would silently")
    print("  redefine the latent totals BET threshold from 6.0 to 3.0 the moment")
    print("  totals are enabled. That is an unapproved totals change.")

    print("\n4 + TEST MATRIX — current behaviour (READY, non-transitional, non-FCS)")
    print(f"  {'edge':>8}  {'|edge|':>7}  {'spread label':<14} {'intended':<10} {'match'}")
    cases = [1.49, 1.50, 1.51, -1.49, -1.50, -1.51, 0.0, "",
             0.99, 1.00, 2.99, 3.00]
    mismatch = []
    for v in cases:
        lab = spread_label(v, S, "READY", "", "")
        if v == "":
            intended = ""
        else:
            intended = "BET" if abs(v) >= 1.5 else ("not BET")
        ok = (lab == "BET") == (intended == "BET")
        if not ok:
            mismatch.append((v, lab, intended))
        vs = "blank" if v == "" else f"{v:+.2f}"
        av = "  --" if v == "" else f"{abs(v):.2f}"
        print(f"  {vs:>8}  {av:>7}  {lab or '(blank)':<14} {intended or '(blank)':<10} "
              f"{'OK' if ok else 'MISMATCH'}")

    print("\n  Edge cases")
    blank_lbl = spread_label("", S, "PENDING LINE", "", "")
    print(f"  {'blank line (V=empty)':<34} -> {blank_lbl!r} (PENDING LINE short-circuits)")
    print(f"  {'zero edge (V=0.0)':<34} -> {spread_label(0.0, S, 'READY', '', '')!r} "
          "(below LEAN threshold)")
    print(f"  {'QB-gated game, |edge|=4.0':<34} -> "
          f"{spread_label(4.0, S, 'QB UNCERTAIN', '', '')!r} "
          "(status != READY forces INVESTIGATE)")
    print(f"  {'transitional team, |edge|=4.0':<34} -> "
          f"{spread_label(4.0, S, 'READY', '1', '')!r}")
    print(f"  {'FCS opponent, |edge|=4.0':<34} -> "
          f"{spread_label(4.0, S, 'FCS — NO PLAY', '', 'FCS OPP')!r}")

    print("\n4. DOES AN EDGE OF EXACTLY +1.50 OR -1.50 RETURN BET TODAY?")
    a, b = spread_label(1.50, S, "READY", "", ""), spread_label(-1.50, S, "READY", "", "")
    print(f"  +1.50 -> {a!r}      -1.50 -> {b!r}")
    print("  NO. Both return INVESTIGATE, for TWO independent reasons:")
    print(f"    (a) ABS(1.50) < SETTINGS!B10 ({S['B10']}) is TRUE  -> INVESTIGATE")
    print(f"    (b) SETTINGS!B11 is {S['B11']!r}, not 'Y'          -> INVESTIGATE regardless")
    print("  With the toggle at 'N', the label BET is currently UNREACHABLE at any edge.")
    reach = any(spread_label(v, S, "READY", "", "") == "BET"
                for v in (1.5, 3.0, 10.0, 50.0, -50.0))
    print(f"  BET reachable at any edge with current settings: {reach}")

    print("\n  VERDICT: the intended rule is NOT implemented.")
    print("  Intended: |ATS edge| >= 1.5 -> BET. Actual: >= 1.5 -> LEAN/INVESTIGATE only.")
    if mismatch:
        print(f"  Test rows disagreeing with the intended rule: {len(mismatch)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

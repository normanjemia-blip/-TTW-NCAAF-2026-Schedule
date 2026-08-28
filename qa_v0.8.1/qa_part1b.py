"""QA Part 1b — adjudicate the three Part-1 findings. Assume my detector is wrong first."""
import openpyxl, re, collections
from openpyxl.worksheet.formula import ArrayFormula
WB="/home/user/-TTW-NCAAF-2026-Schedule/promotion_v0.8.1/TTW_College_Football_Power_Ratings_v0.8.1_AUTHORITATIVE.xlsx"
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
def ft(v): return v.text if isinstance(v,ArrayFormula) else v
wb=openpyxl.load_workbook(WB)
OUT=[]
def say(s): print(s); OUT.append(s)

say("="*78); say("ADJUDICATION OF PART-1 FINDINGS"); say("="*78)

# ---------- F1: literal #VALUE! ----------
say("\nF1 — literal error strings (3 hits)")
for s in wb.worksheets:
    for row in s.iter_rows():
        for c in row:
            v=ft(c.value)
            if isinstance(v,str) and not v.startswith("=") and any(e in v for e in ("#VALUE!","#REF!","#N/A")):
                say(f"  {s.title}!{c.coordinate}:")
                say(f"    {v[:400]}")
say("  VERDICT: these are CHANGELOG narrative entries describing a historical defect.")
say("           They are prose, not error values. FALSE POSITIVE in my detector.")

# ---------- F2: self-reference ----------
say("\nF2 — self-referencing cells")
say("  My detector stripped '$' then searched for the coordinate anywhere in the")
say("  formula, so DASHBOARD!A6 ='...ENGINE!$A6...' matched its own name. Re-testing")
say("  with sheet-qualified references removed first.")
strip_qualified=re.compile(r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_ .]*)!\$?[A-Z]{1,3}\$?\d+")
real_self=[]
for s in wb.worksheets:
    for row in s.iter_rows():
        for c in row:
            v=ft(c.value)
            if not (isinstance(v,str) and v.startswith("=")): continue
            bare=strip_qualified.sub("", v).replace("$","")
            if re.search(rf"(?<![A-Z0-9]){re.escape(c.coordinate)}(?![0-9])", bare):
                real_self.append((s.title,c.coordinate,v[:120]))
say(f"  genuine self-references after correction: {len(real_self)}")
for x in real_self[:10]: say(f"    {x}")
say("  VERDICT: " + ("CLEAN — false positive confirmed." if not real_self else "REAL DEFECT."))

# ---------- F3: TEAM RATINGS <-> CALC ----------
say("\nF3 — TEAM RATINGS <-> CALC sheet-level cycle: is it circular AT CELL LEVEL?")
ref=re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ .]*))!\$?([A-Z]{1,3})\$?(\d+)")
def col_map(src, dst):
    """which columns of `src` read which columns of `dst`"""
    m=collections.defaultdict(set)
    for row in wb[src].iter_rows():
        for c in row:
            v=ft(c.value)
            if not (isinstance(v,str) and v.startswith("=")): continue
            for g in ref.finditer(v):
                tgt=(g.group(1) or g.group(2)).strip()
                if tgt==dst:
                    m[c.column_letter].add(g.group(3))
    return m
tr_reads_calc=col_map("TEAM RATINGS","CALC")
calc_reads_tr=col_map("CALC","TEAM RATINGS")
say("  TEAM RATINGS columns that read CALC:")
for k in sorted(tr_reads_calc): say(f"    TEAM RATINGS!{k} -> CALC!{sorted(tr_reads_calc[k])}")
say("  CALC columns that read TEAM RATINGS:")
for k in sorted(calc_reads_tr): say(f"    CALC!{k} -> TEAM RATINGS!{sorted(calc_reads_tr[k])}")
tr_src=set().union(*calc_reads_tr.values()) if calc_reads_tr else set()   # TR cols CALC depends on
tr_dst=set(tr_reads_calc)                                                 # TR cols that depend on CALC
calc_src=set().union(*tr_reads_calc.values()) if tr_reads_calc else set() # CALC cols TR depends on
calc_dst=set(calc_reads_tr)                                               # CALC cols that depend on TR
say(f"\n  TEAM RATINGS columns CALC depends on : {sorted(tr_src)}")
say(f"  TEAM RATINGS columns that depend on CALC: {sorted(tr_dst)}")
overlap_tr=tr_src & tr_dst
say(f"  OVERLAP on TEAM RATINGS: {sorted(overlap_tr)}")
say(f"\n  CALC columns TEAM RATINGS depends on : {sorted(calc_src)}")
say(f"  CALC columns that depend on TEAM RATINGS: {sorted(calc_dst)}")
overlap_calc=calc_src & calc_dst
say(f"  OVERLAP on CALC: {sorted(overlap_calc)}")
if not overlap_tr and not overlap_calc:
    say("\n  VERDICT: NO CELL-LEVEL CIRCULARITY. The two sheets reference each other")
    say("           through DISJOINT column sets, which is legal and normal. A")
    say("           sheet-level cycle is not a circular reference. FALSE POSITIVE.")
else:
    say("\n  VERDICT: POTENTIAL REAL CIRCULARITY — overlapping columns; investigate rows.")
    # row-level check on the overlapping columns
    for col in sorted(overlap_tr):
        for r in range(1, wb["TEAM RATINGS"].max_row+1):
            v=ft(wb["TEAM RATINGS"].cell(row=r,column=openpyxl.utils.column_index_from_string(col)).value)
            if isinstance(v,str) and "CALC" in v:
                say(f"    TEAM RATINGS!{col}{r} = {v[:150]}")
                break
open("/tmp/qa/part1b_log.txt","w").write("\n".join(OUT)+"\n")

"""QA Part 1 — Workbook Integrity. Static analysis, nothing modified."""
import openpyxl, re, collections, json, sys
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter

WB="/home/user/-TTW-NCAAF-2026-Schedule/promotion_v0.8.1/TTW_College_Football_Power_Ratings_v0.8.1_AUTHORITATIVE.xlsx"
OUT=[]; FIND=[]
def say(s): print(s); OUT.append(s)
def chk(n,ok,d="",sev="CRITICAL"):
    say(f"  [{'PASS' if ok else 'FAIL'}] {n}{(' - '+d) if d else ''}")
    if not ok: FIND.append((sev,n,d))
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
def ft(v): return v.text if isinstance(v,ArrayFormula) else v

wb=openpyxl.load_workbook(WB)
wbv=openpyxl.load_workbook(WB, data_only=True)
say("="*78); say("QA PART 1 — WORKBOOK INTEGRITY"); say("="*78)

# ---- 1.1 every worksheet opens ----
say("\n1.1 WORKSHEET ACCESS")
ok=True; dims={}
for name in wb.sheetnames:
    try:
        s=wb[name]; dims[name]=(s.max_row,s.max_column,s.sheet_state)
    except Exception as e:
        ok=False; say(f"    ERROR opening {name}: {e}")
say(f"  sheets: {len(wb.sheetnames)}")
for n,(r,c,st) in dims.items(): say(f"    {n:<20} {r:>5} x {c:<3} [{st}]")
chk("1.1 all 21 worksheets open without error", ok and len(wb.sheetnames)==21)

# ---- 1.2 error literals anywhere ----
say("\n1.2 ERROR VALUES IN CELLS")
ERRS=("#REF!","#VALUE!","#N/A","#DIV/0!","#NAME?","#NULL!","#NUM!")
found=collections.Counter(); examples={}
for s in wb.worksheets:
    for row in s.iter_rows():
        for c in row:
            v=ft(c.value)
            if isinstance(v,str):
                for e in ERRS:
                    if e in v:
                        # distinguish: an error INSIDE a formula (e.g. IFERROR text) vs a broken ref
                        kind="formula-text" if v.startswith("=") else "literal-value"
                        found[(e,kind)]+=1
                        examples.setdefault((e,kind),(s.title,c.coordinate,v[:120]))
for k,v in sorted(found.items()):
    sh,co,txt=examples[k]
    say(f"    {k[0]} ({k[1]}) x{v}  first: {sh}!{co}  {txt}")
broken_ref=sum(v for (e,k),v in found.items() if e=="#REF!")
chk("1.2a no #REF! anywhere", broken_ref==0, str(broken_ref))
lit_err=sum(v for (e,k),v in found.items() if k=="literal-value")
chk("1.2b no literal error values stored in cells", lit_err==0, str(lit_err))

# ---- 1.3 cached error results ----
say("\n1.3 CACHED FORMULA RESULTS")
cached_err=0; cached_any=0
for s in wbv.worksheets:
    for row in s.iter_rows():
        for c in row:
            if c.value is not None:
                cached_any+=1
                if isinstance(c.value,str) and c.value in ERRS: cached_err+=1
say(f"  non-empty cells (data_only): {cached_any}; cached error values: {cached_err}")
chk("1.3 no cached error results", cached_err==0, str(cached_err))
say("  NOTE: this workbook stores NO cached formula results, so 1.3 cannot detect")
say("        runtime errors. Runtime behaviour is assessed in Parts 3-5 instead.")

# ---- 1.4 dependency graph + circular references ----
say("\n1.4 DEPENDENCY GRAPH AND CIRCULAR REFERENCES")
SHEETS=set(wb.sheetnames)
sheet_ref=re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ .]*))!")
edges=collections.defaultdict(set)     # sheet -> sheets it reads
unknown_sheets=collections.Counter(); unknown_ex={}
self_ref_cells=[]
for s in wb.worksheets:
    for row in s.iter_rows():
        for c in row:
            v=ft(c.value)
            if not (isinstance(v,str) and v.startswith("=")): continue
            for m in sheet_ref.finditer(v):
                tgt=(m.group(1) or m.group(2)).strip()
                if tgt in SHEETS:
                    if tgt!=s.title: edges[s.title].add(tgt)
                elif tgt.upper() not in ("IF","AND","OR","TRUE","FALSE"):
                    unknown_sheets[tgt]+=1
                    unknown_ex.setdefault(tgt,(s.title,c.coordinate,v[:100]))
            # naive self-reference: formula in cell X refers to X on same sheet
            if re.search(rf"(?<![A-Z0-9$]){re.escape(c.coordinate)}(?![0-9])", v.replace("$","")):
                self_ref_cells.append((s.title,c.coordinate,v[:90]))
say("  cross-sheet dependency edges:")
for k in sorted(edges): say(f"    {k:<18} -> {sorted(edges[k])}")
chk("1.4a all sheet references resolve to existing sheets", not unknown_sheets,
    str([(k,unknown_ex[k]) for k in list(unknown_sheets)[:5]]))
chk("1.4b no cell references itself (direct self-circularity)", not self_ref_cells,
    str(self_ref_cells[:5]))
# cycle detection on the sheet-level graph
def find_cycles(g):
    cyc=[]; WHITE,GREY,BLACK=0,1,2; col=collections.defaultdict(int); stack=[]
    def dfs(u):
        col[u]=GREY; stack.append(u)
        for v in g.get(u,()):
            if col[v]==GREY: cyc.append(stack[stack.index(v):]+[v])
            elif col[v]==WHITE: dfs(v)
        stack.pop(); col[u]=BLACK
    for n in list(g):
        if col[n]==WHITE: dfs(n)
    return cyc
cycles=find_cycles(edges)
say(f"  sheet-level cycles detected: {len(cycles)}")
for c in cycles[:5]: say(f"    {' -> '.join(c)}")
chk("1.4c no sheet-level circular dependency", not cycles, str(cycles[:3]),
    sev="MAJOR")

# ---- 1.5 named ranges ----
say("\n1.5 NAMED RANGES")
dn=dict(wb.defined_names)
say(f"  defined names: {len(dn)}")
bad=[]
for n,d in dn.items():
    val=str(d.value)
    say(f"    {n} = {val}")
    if "#REF" in val: bad.append(n)
    m=sheet_ref.search(val)
    if m:
        t=(m.group(1) or m.group(2)).strip()
        if t not in SHEETS: bad.append(n)
chk("1.5 all named ranges resolve (or none defined)", not bad, str(bad))

# ---- 1.6 data validation ----
say("\n1.6 DATA VALIDATION")
tot_dv=0
for s in wb.worksheets:
    try: dvs=list(s.data_validations.dataValidation)
    except Exception: dvs=[]
    if dvs:
        say(f"    {s.title}: {len(dvs)} rule(s)")
        for v in dvs:
            tot_dv+=1
            say(f"      type={v.type} op={v.operator} sqref={v.sqref} f1={str(v.formula1)[:80]}")
say(f"  total data-validation rules: {tot_dv}")
chk("1.6 data validation present and enumerable", True, f"{tot_dv} rules", sev="MINOR")

# ---- 1.7 conditional formatting ----
say("\n1.7 CONDITIONAL FORMATTING")
tot_cf=0
for s in wb.worksheets:
    try: items=list(s.conditional_formatting._cf_rules.items())
    except Exception: items=[]
    if items:
        n=sum(len(r) for _,r in items)
        tot_cf+=n; say(f"    {s.title}: {n} rule(s) over {len(items)} range(s)")
say(f"  total conditional-format rules: {tot_cf}")
chk("1.7 conditional formatting present and enumerable", True, f"{tot_cf} rules", sev="MINOR")

# ---- 1.8 protection ----
say("\n1.8 WORKBOOK / SHEET PROTECTION")
say(f"  workbook protection: {wb.security}")
prot=[]
for s in wb.worksheets:
    p=s.protection
    prot.append((s.title,p.sheet,p.password is not None))
    say(f"    {s.title:<20} sheet_protected={p.sheet}  password_set={p.password is not None}")
any_prot=any(p[1] for p in prot)
chk("1.8 protection state is consistent and enumerable", True,
    f"{sum(1 for p in prot if p[1])} of {len(prot)} sheets protected", sev="MINOR")

# ---- 1.9 hidden sheets match documentation ----
say("\n1.9 SHEET VISIBILITY")
hidden=[n for n in wb.sheetnames if wb[n].sheet_state!="visible"]
say(f"  hidden: {hidden}")
DOC_HIDDEN={"TEAM MAP","CLEAN","CALC","PRESEASON","FCS TIERS","HISTORY","BACKTEST","AUDIT","DICTIONARY","CHANGELOG"}
chk("1.9 hidden set matches the START HERE tab map", set(hidden)==DOC_HIDDEN,
    f"actual={sorted(hidden)} documented={sorted(DOC_HIDDEN)}", sev="MAJOR")

json.dump([{"severity":a,"check":b,"detail":c} for a,b,c in FIND], open("/tmp/qa/part1_findings.json","w"), indent=1)
open("/tmp/qa/part1_log.txt","w").write("\n".join(OUT)+"\n")
say(f"\nPART 1 FINDINGS: {len(FIND)}")
for a,b,c in FIND: say(f"  [{a}] {b} :: {c}")

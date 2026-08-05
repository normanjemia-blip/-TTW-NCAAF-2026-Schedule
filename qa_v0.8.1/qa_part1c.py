"""QA Part 1c — DEFINITIVE circular-reference and self-reference test.

Node = (sheet, column). Edge = dependent -> dependency, extracted from every
formula in the workbook. Cycle detection over the full graph.
"""
import openpyxl, re, collections
from openpyxl.worksheet.formula import ArrayFormula
WB="/home/user/-TTW-NCAAF-2026-Schedule/promotion_v0.8.1/TTW_College_Football_Power_Ratings_v0.8.1_AUTHORITATIVE.xlsx"
def ft(v): return v.text if isinstance(v,ArrayFormula) else v
wb=openpyxl.load_workbook(WB)
SHEETS=set(wb.sheetnames)
OUT=[]
def say(s): print(s); OUT.append(s)

say("="*78); say("DEFINITIVE CIRCULAR-REFERENCE TEST (column-level graph)"); say("="*78)

# strip qualified RANGES first, then qualified single refs
QRANGE=re.compile(r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_ .]*)!\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+")
QREF  =re.compile(r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_ .]*)!\$?[A-Z]{1,3}\$?\d+")
ANYREF=re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ .]*))!\$?([A-Z]{1,3})\$?(\d+)(?::\$?([A-Z]{1,3})\$?(\d+))?")
LOCAL =re.compile(r"(?<![A-Z0-9_!])\$?([A-Z]{1,3})\$?(\d+)(?![0-9(])")

# ---------- self-reference, done correctly ----------
say("\nA. SELF-REFERENCE (cell refers to itself)")
selfrefs=[]
for s in wb.worksheets:
    for row in s.iter_rows():
        for c in row:
            v=ft(c.value)
            if not (isinstance(v,str) and v.startswith("=")): continue
            bare=QREF.sub(" ", QRANGE.sub(" ", v))     # remove ALL sheet-qualified refs/ranges
            for m in LOCAL.finditer(bare):
                if f"{m.group(1)}{m.group(2)}"==c.coordinate:
                    selfrefs.append((s.title,c.coordinate,v[:140])); break
say(f"  genuine self-references: {len(selfrefs)}")
for x in selfrefs[:10]: say(f"    {x}")
say("  VERDICT: " + ("CLEAN — no cell references itself." if not selfrefs else "REAL DEFECT"))

# ---------- column-level graph ----------
say("\nB. COLUMN-LEVEL DEPENDENCY GRAPH")
g=collections.defaultdict(set); ex={}
for s in wb.worksheets:
    for row in s.iter_rows():
        for c in row:
            v=ft(c.value)
            if not (isinstance(v,str) and v.startswith("=")): continue
            src=(s.title, c.column_letter)
            # cross-sheet refs
            for m in ANYREF.finditer(v):
                tgt=(m.group(1) or m.group(2)).strip()
                if tgt not in SHEETS: continue
                c1=m.group(3); c2=m.group(5)
                cols={c1} if not c2 else set()
                if c2:
                    a=openpyxl.utils.column_index_from_string(c1); b=openpyxl.utils.column_index_from_string(c2)
                    cols={openpyxl.utils.get_column_letter(i) for i in range(min(a,b),max(a,b)+1)}
                for cc in cols:
                    d=(tgt,cc)
                    if d!=src:
                        g[src].add(d); ex.setdefault((src,d), f"{s.title}!{c.coordinate}: {v[:90]}")
            # same-sheet refs
            bare=QREF.sub(" ", QRANGE.sub(" ", v))
            for m in LOCAL.finditer(bare):
                d=(s.title, m.group(1))
                if d!=src:
                    g[src].add(d); ex.setdefault((src,d), f"{s.title}!{c.coordinate}: {v[:90]}")
say(f"  nodes: {len(set(list(g)+[d for v in g.values() for d in v]))}   edges: {sum(len(v) for v in g.values())}")

WHITE,GREY,BLACK=0,1,2
col=collections.defaultdict(int); stack=[]; cycles=[]
def dfs(u):
    col[u]=GREY; stack.append(u)
    for v in sorted(g.get(u,())):
        if col[v]==GREY:
            cycles.append(stack[stack.index(v):]+[v])
        elif col[v]==WHITE:
            dfs(v)
    stack.pop(); col[u]=BLACK
import sys; sys.setrecursionlimit(10000)
for n in sorted(set(list(g)+[d for v in g.values() for d in v])):
    if col[n]==WHITE: dfs(n)
say(f"  column-level cycles: {len(cycles)}")
seen=set()
for cyc in cycles:
    key=tuple(sorted(set(cyc)))
    if key in seen: continue
    seen.add(key)
    say("    CYCLE: " + " -> ".join(f"{a}!{b}" for a,b in cyc))
    for i in range(len(cyc)-1):
        e=ex.get((cyc[i],cyc[i+1]))
        if e: say(f"        {cyc[i][0]}!{cyc[i][1]} <- {e}")
    if len(seen)>=6: break
say("  VERDICT: " + ("CLEAN — no circular reference at column level."
                     if not cycles else f"{len(seen)} distinct cycle(s) — investigate row level."))

# ---------- the specific TEAM RATINGS <-> CALC question ----------
say("\nC. TEAM RATINGS <-> CALC, resolved explicitly")
say("  TEAM RATINGS!H reads CALC!A and CALC!H")
say("  CALC!A/X/Y read TEAM RATINGS!A and TEAM RATINGS!F")
tr_h_deps = g.get(("TEAM RATINGS","H"), set())
say(f"  ('TEAM RATINGS','H') depends on: {sorted(tr_h_deps)}")
for probe in (("TEAM RATINGS","A"),("TEAM RATINGS","F"),("CALC","A"),("CALC","H")):
    say(f"  {probe} depends on: {sorted(g.get(probe,set()))}")
say("  If TEAM RATINGS!A and TEAM RATINGS!F do NOT depend on CALC, the sheet-level")
say("  'cycle' is disjoint at column level and is NOT a circular reference.")
open("/tmp/qa/part1c_log.txt","w").write("\n".join(OUT)+"\n")

"""
v0.4.1 validation battery:
 A. Source revalidation: independently re-parse the archived FPI and
    TeamRankings raw files and compare every loaded value against
    PRESEASON!H6:H143 and Q6:Q143 (276 comparisons).
 B. Full cell/formula diff v0.4.1 vs approved v0.3.1 (every sheet, every
    cell, ArrayFormula compared by ref+text) -> written to a report file.
 C. Formula-integrity check: assert ZERO cells that contain a formula in
    v0.3.1 differ in v0.4.1.
 D. Blank-input assertions: SP+/TTW/VSiN raw+date+cite input columns and
    all QB VALUES manual-input columns still fully blank.
 E. Structural counts: CHANGELOG rows, IMPORT SCHEDULE row count intact.
"""
import json
import re
import html as htmllib
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

P4 = "../phase4"
failures = []

def check(label, ok, detail=""):
    print(("PASS  " if ok else "FAIL  ") + label + (f"  {detail}" if detail else ""))
    if not ok:
        failures.append((label, detail))

wb = openpyxl.load_workbook("v0.4.1_working.xlsx", data_only=False)
ps = wb["PRESEASON"]
master = json.load(open(f"{P4}/master_list.json"))

# ---------- A1. FPI revalidation (independent re-parse of archived JSON) ----------
fpi_raw = json.load(open(f"{P4}/raw_sources/espn_fpi_2026preseason_20260719_parsed.json"))
stats = fpi_raw["page"]["content"]["table"]["stats"]
k_to_abbrev = json.load(open(f"{P4}/k_to_abbrev.json"))
fpi_by_abbrev = {}
for s in stats:
    ab = k_to_abbrev.get(s["team"]["displayName"])
    val = next(st["value"] for st in s["stats"] if st["name"] == "fpi")
    fpi_by_abbrev[ab] = float(val)
check("FPI re-parse team count == 138", len(fpi_by_abbrev) == 138, str(len(fpi_by_abbrev)))

mismatch = []
for row_str, (abbrev, name, conf) in master.items():
    cell_val = ps.cell(row=int(row_str), column=8).value
    if cell_val != fpi_by_abbrev.get(abbrev):
        mismatch.append((abbrev, cell_val, fpi_by_abbrev.get(abbrev)))
check("FPI 138/138 workbook values match archived raw source", not mismatch, str(mismatch[:5]))

# ---------- A2. TeamRankings revalidation (independent re-parse of archived HTML) ----------
tr_html = open(f"{P4}/raw_sources/teamrankings_predictive_2026preseason_20260719.html").read()
rows = re.findall(
    r'<td class="rank[^"]*"[^>]*>(\d+)</td>\s*'
    r'<td class="nowrap"[^>]*><a href="[^"]*">([^<]+)</a>.*?'
    r'<td class="text-right"[^>]*>(-?\d+\.\d+)</td>',
    tr_html, re.S)
check("TR re-parse row count == 138", len(rows) == 138, str(len(rows)))

# reuse the explicit, manually-verified mapping table (the audited artifact)
import os
src = open(f"{P4}/match_teamrankings.py").read()
ns = {}
_cwd = os.getcwd()
os.chdir(P4)  # the script prefix opens master_list.json relatively
try:
    exec(src.split("rows = []")[0], ns)  # defs + mapping table only
finally:
    os.chdir(_cwd)
TR_TO_ABBREV = ns["TR_TO_ABBREV"]

tr_by_abbrev = {}
for rank, name, rating in rows:
    name = htmllib.unescape(name).strip()
    ab = TR_TO_ABBREV.get(name)
    if ab is None:
        failures.append(("TR unmatched name in revalidation", name))
        continue
    tr_by_abbrev[ab] = float(rating)
check("TR re-parse matched 138 unique abbrevs", len(tr_by_abbrev) == 138, str(len(tr_by_abbrev)))

mismatch = []
for row_str, (abbrev, name, conf) in master.items():
    cell_val = ps.cell(row=int(row_str), column=17).value
    if cell_val != tr_by_abbrev.get(abbrev):
        mismatch.append((abbrev, cell_val, tr_by_abbrev.get(abbrev)))
check("TR 138/138 workbook values match archived raw source", not mismatch, str(mismatch[:5]))

# ---------- B. Full diff v0.4.1 vs v0.3.1 ----------
def normval(v):
    if isinstance(v, ArrayFormula):
        return ("ARRAYFORMULA", v.ref, v.text)
    return v

wb_old = openpyxl.load_workbook("v0.3.1_with_schedule.xlsx", data_only=False)
diff_lines = []
diff_counts = {}
formula_diffs = []
for name in wb_old.sheetnames:
    ws_old, ws_new = wb_old[name], wb[name]
    mr = max(ws_old.max_row, ws_new.max_row)
    mc = max(ws_old.max_column, ws_new.max_column)
    cnt = 0
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            vo = normval(ws_old.cell(row=r, column=c).value)
            vn = normval(ws_new.cell(row=r, column=c).value)
            if vo != vn:
                cnt += 1
                coord = ws_new.cell(row=r, column=c).coordinate
                diff_lines.append(f"{name}\t{coord}\t{repr(vo)}\t{repr(vn)}")
                is_formula_old = (isinstance(vo, tuple) or (isinstance(vo, str) and vo.startswith("=")))
                if is_formula_old:
                    formula_diffs.append((name, coord))
    if cnt:
        diff_counts[name] = cnt

with open("v0.4.1_vs_v0.3.1_full_diff.tsv", "w") as f:
    f.write("sheet\tcell\tv0.3.1_value\tv0.4.1_value\n")
    f.write("\n".join(diff_lines) + "\n")
print("Diff counts by sheet:", diff_counts)

expected_sheets = {"START HERE", "PRESEASON", "CHANGELOG", "SETTINGS", "FCS TIERS"}
check("Diff touches only expected sheets", set(diff_counts) == expected_sheets, str(set(diff_counts) ^ expected_sheets))
check("START HERE diff == 2 cells (A1 banner + A3 state)", diff_counts.get("START HERE") == 2, str(diff_counts.get("START HERE")))
check("PRESEASON diff == 829 cells (828 data + A1 header)", diff_counts.get("PRESEASON") == 829, str(diff_counts.get("PRESEASON")))
check("SETTINGS diff == 2 cells (C22, C23)", diff_counts.get("SETTINGS") == 2, str(diff_counts.get("SETTINGS")))
check("FCS TIERS diff == 1 cell (A2)", diff_counts.get("FCS TIERS") == 1, str(diff_counts.get("FCS TIERS")))
check("CHANGELOG diff == 48 cells (12 rows x 4)", diff_counts.get("CHANGELOG") == 48, str(diff_counts.get("CHANGELOG")))

# ---------- C. Formula integrity ----------
check("ZERO formula cells changed (v0.3.1 formula cells all identical)", not formula_diffs, str(formula_diffs[:8]))

# ---------- D. Blank-input assertions ----------
blank_cols = {"SP+ raw/date/cite": [4, 5, 6], "TTW raw/date/cite": [12, 14, 15], "VSiN raw/date/cite": [21, 22, 23],
              "Win-total AD/AE": [30, 31]}
for label, cols in blank_cols.items():
    nonblank = [(r, c) for r in range(6, 144) for c in cols if ps.cell(row=r, column=c).value is not None]
    check(f"PRESEASON {label} columns fully blank", not nonblank, str(nonblank[:5]))

qb = wb["QB VALUES"]
qb_manual_cols = [3, 4, 5, 6, 8, 9, 10, 11, 12]  # C,D,E,F,H,I,J,K,L
nonblank = [(r, c) for r in range(6, 144) for c in qb_manual_cols if qb.cell(row=r, column=c).value is not None]
check("QB VALUES manual-input columns fully blank", not nonblank, str(nonblank[:5]))

# ---------- E. Structural counts ----------
cl = wb["CHANGELOG"]
n = 0
r = 7
while cl.cell(row=r, column=1).value is not None:
    n += 1; r += 1
print(f"CHANGELOG total entries from row 7: {n}")
v041 = sum(1 for rr in range(7, r) if cl.cell(row=rr, column=1).value == "v0.4.1")
v04 = sum(1 for rr in range(7, r) if cl.cell(row=rr, column=1).value == "v0.4")
check("CHANGELOG has 9 v0.4 entries", v04 == 9, str(v04))
check("CHANGELOG has 3 v0.4.1 entries", v041 == 3, str(v041))

sched = wb["IMPORT SCHEDULE"]
ids = [sched.cell(row=r, column=1).value for r in range(6, 894)]
check("IMPORT SCHEDULE 888 game ids intact", sum(1 for i in ids if i) == 888, "")
check("IMPORT SCHEDULE dims A1:N1005", sched.dimensions == "A1:N1005", sched.dimensions)

print()
if failures:
    print(f"*** {len(failures)} FAILURES ***")
    for f_ in failures: print("   ", f_)
    raise SystemExit(1)
print("ALL CHECKS PASSED")

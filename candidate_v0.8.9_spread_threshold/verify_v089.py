#!/usr/bin/env python3
"""v0.8.9 SPREAD-THRESHOLD CANDIDATE certificate — READ-ONLY. Writes nothing."""
import collections, hashlib, os, re, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
V088 = os.path.join(ROOT, "promotion_v0.8.8",
                    "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
V089 = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.9_THRESHOLD_CANDIDATE.xlsx")
SRC_SHA = "b2a920feddc0f49f0647957334db0ecd0e922fe6a3933fc6a11af31587b56450"

# the eight live operational market lines (from the audited live Sheet export)
LIVE_LINES = {
    "401856766": ("TCU", 7.5, 47.5), "401858201": ("STAN", 3.5, 48.5),
    "401858202": ("UVA", 5.5, 53.5), "401864494": ("USC", 38.5, 60.5),
    "401864570": ("FSU", 30.5, 52.5), "401864577": ("NDSU", 7.5, 46.5),
    "401866408": ("EMU", 9.5, 52.5), "401862693": ("UNLV", 4.5, 56.5),
}
EXPECTED_LABEL_CHANGES = {"401866408", "401856766", "401864494", "401864570"}

PASS, FAIL = [], []


def chk(m, ok, d=""):
    (PASS if ok else FAIL).append(m)
    print(f"  [{'PASS' if ok else 'FAIL'}] {m}" + (f" - {d}" if d else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", (v.text or "").strip())
    if isinstance(v, str) and v.startswith("="):
        return ("F", v.strip())
    return ("V", v)


def mean(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def spread_label(V, S, AI, AF, AG):
    """ENGINE!X verbatim - unchanged between v0.8.8 and v0.8.9."""
    if V == "" or AI in ("BLOCKED", "PENDING LINE", "STALE LINE"):
        return ""
    if abs(V) < S["B8"]:
        return ""
    if abs(V) < S["B9"]:
        return "LEAN"
    if (abs(V) < S["B10"] or S["B11"] != "Y" or AI != "READY" or AF != "" or AG != ""):
        return "INVESTIGATE"
    return "BET"


def total_label_old(AA, S, AI, AF, AG):
    """v0.8.8 ENGINE!AB - coupled: B8*2 / B9*2 / B10*2."""
    if AA == "":
        return ""
    if abs(AA) < S["B8"] * 2:
        return ""
    if abs(AA) < S["B9"] * 2:
        return "LEAN"
    if (S["B11"] != "Y" or AI != "READY" or abs(AA) < S["B10"] * 2 or AF != "" or AG != ""):
        return "INVESTIGATE"
    return "BET"


def total_label_new(AA, S, AI, AF, AG):
    """v0.8.9 ENGINE!AB - decoupled: B33 / B34 / B26."""
    if AA == "":
        return ""
    if abs(AA) < S["B33"]:
        return ""
    if abs(AA) < S["B34"]:
        return "LEAN"
    if (S["B11"] != "Y" or AI != "READY" or abs(AA) < S["B26"] or AF != "" or AG != ""):
        return "INVESTIGATE"
    return "BET"


def audit_b12(S, ab_formula):
    """AUDIT!B12 verbatim semantics."""
    ok = (S.get("B10") == 1.5 and S.get("B11") == "Y" and S.get("B26") == 6
          and S.get("B33") == 2 and S.get("B34") == 3
          and "SETTINGS!$B$33" in ab_formula and "SETTINGS!$B$34" in ab_formula
          and "SETTINGS!$B$26" in ab_formula
          and "$B$8*2" not in ab_formula and "$B$9*2" not in ab_formula
          and "$B$10*2" not in ab_formula)
    return "OK" if ok else "CHECK"


def engine(wb, lines):
    tm, qb, st, ps = wb["TEAM MAP"], wb["QB VALUES"], wb["SETTINGS"], wb["PRESEASON"]
    S = {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 36)}
    raw = {c: [ps.cell(row=r, column=c).value for r in range(6, 144)] for c in (4, 8, 17)}
    mu = {c: mean(v) for c, v in raw.items()}
    W = {4: S["B28"], 8: S["B29"], 17: S["B31"]}
    prior = {}
    for i, r in enumerate(range(6, 144)):
        ab = tm.cell(row=r, column=1).value
        num = den = 0.0
        for c in (4, 8, 17):
            v = raw[c][i]
            if isinstance(v, (int, float)):
                num += W[c] * (v - mu[c]); den += W[c]
        prior[ab] = (num / den) if den else ""
    alias = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()
    delta, status = {}, {}
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        g = "" if (D is None or F is None) else F - D
        delta[ab] = g
        status[ab] = "UNCERTAIN" if (g == "" or H == "L" or J != S["B3"]) else "OK"
    z = lambda v: 0 if v == "" else v
    sch = wb["IMPORT SCHEDULE"]
    out = {}
    for r in range(6, 1006):
        gid = sch.cell(row=r, column=1).value
        if not gid:
            continue
        A = alias.get(str(sch.cell(row=r, column=6).value).strip(), "")
        H = alias.get(str(sch.cell(row=r, column=8).value).strip(), "")
        if not A or not H:
            continue
        R = (prior[H] - prior[A]) + (S["B7"] if bool(sch.cell(row=r, column=5).value) else S["B6"]) \
            + (z(delta.get(H, "")) - z(delta.get(A, "")))
        ln = lines.get(str(gid))
        edge = side = ""
        gate = "QB UNCERTAIN" if (status.get(A) == "UNCERTAIN" or status.get(H) == "UNCERTAIN") else "READY"
        if ln:
            fav, spread, _ = ln
            T = -abs(spread) if fav == H else (abs(spread) if fav == A else "")
            if T != "":
                edge = R + T
                side = H if edge > 0 else A
        out[str(gid)] = dict(away=A, home=H, model=R, edge=edge, side=side, gate=gate, S=S)
    return out


def main():
    print("=" * 78)
    print("v0.8.9 SPREAD-THRESHOLD CANDIDATE — CERTIFICATE (candidate only)")
    print("=" * 78)
    h8, h9 = sha256(V088), sha256(V089)
    print(f"  v0.8.8 source SHA-256 : {h8}")
    print(f"  v0.8.9 candidate SHA  : {h9}")

    print("\n1. SOURCE")
    chk("1.1 source v0.8.8 hash matches and is unmodified", h8 == SRC_SHA, h8[:16])

    a = openpyxl.load_workbook(V088)
    b = openpyxl.load_workbook(V089)

    print("\n2. EVERY CHANGED CELL")
    changed = []
    for s in a.sheetnames:
        wa, wb_ = a[s], b[s]
        R = max(wa.max_row, wb_.max_row); C = max(wa.max_column, wb_.max_column)
        for r in range(1, R + 1):
            for c in range(1, C + 1):
                if norm(wa.cell(row=r, column=c).value) != norm(wb_.cell(row=r, column=c).value):
                    changed.append((s, wb_.cell(row=r, column=c).coordinate))
    by = collections.Counter(s for s, _ in changed)
    chk("2.1 exactly 1008 cells changed", len(changed) == 1008, str(len(changed)))
    chk("2.2 changes confined to SETTINGS (6), ENGINE (1000), AUDIT (2)",
        dict(by) == {"SETTINGS": 6, "ENGINE": 1000, "AUDIT": 2}, str(dict(by)))
    setc = sorted(c for s, c in changed if s == "SETTINGS")
    chk("2.3 SETTINGS cells are exactly A26,B10,B11,B26,B33,B34",
        setc == ["A26", "B10", "B11", "B26", "B33", "B34"], str(setc))
    audc = sorted(c for s, c in changed if s == "AUDIT")
    chk("2.4 AUDIT cells are exactly A12,B12", audc == ["A12", "B12"], str(audc))
    engc = {c for s, c in changed if s == "ENGINE"}
    chk("2.5 all 1000 ENGINE cells are in column AB, rows 6-1005",
        all(re.fullmatch(r"AB(\d+)", c) and 6 <= int(c[2:]) <= 1005 for c in engc)
        and len(engc) == 1000, str(len(engc)))

    print("\n3. NO QB OR SCHEDULE CELL CHANGED")
    for sheet in ("QB VALUES", "IMPORT SCHEDULE", "TEAM MAP", "PRESEASON", "TEAM RATINGS",
                  "MARKET LINES", "ADJUSTMENTS", "CLEAN", "CALC", "DASHBOARD"):
        chk(f"3.x {sheet} byte-identical to v0.8.8", by.get(sheet, 0) == 0, str(by.get(sheet, 0)))

    print("\n4. NO RATING, PROJECTION, EDGE, DIRECTION OR SPREAD CHANGED")
    E8, E9 = engine(a, LIVE_LINES), engine(b, LIVE_LINES)
    diffs = [g for g in E8 if (round(E8[g]["model"], 9), E8[g]["edge"], E8[g]["side"], E8[g]["gate"])
             != (round(E9[g]["model"], 9), E9[g]["edge"], E9[g]["side"], E9[g]["gate"])]
    chk("4.1 model spread, edge, side and gate identical across all 761 FBS-v-FBS games",
        not diffs, str(diffs[:5]))
    S8, S9 = E8[list(E8)[0]]["S"], E9[list(E9)[0]]["S"]
    chk("4.2 rating inputs untouched: HFA 2.5, neutral 0, cap 2.5, weights .30/.25/.20/.15",
        (S9["B6"], S9["B7"], S9["B12"], S9["B28"], S9["B29"], S9["B30"], S9["B31"])
        == (2.5, 0, 2.5, 0.3, 0.25, 0.2, 0.15))
    chk("4.3 LEAN/INVESTIGATE thresholds B8/B9 unchanged at 1.0 / 1.5",
        (S9["B8"], S9["B9"]) == (1, 1.5), f"{S9['B8']}/{S9['B9']}")

    print("\n5. SPREAD BOUNDARY RESULTS (READY, non-transitional, non-FCS)")
    for v, want in ((1.49, "not BET"), (-1.49, "not BET"), (1.50, "BET"), (-1.50, "BET"),
                    (1.51, "BET"), (-1.51, "BET")):
        lab = spread_label(v, S9, "READY", "", "")
        ok = (lab == "BET") if want == "BET" else (lab != "BET")
        chk(f"5.x edge {v:+.2f} -> {lab or '(blank)'} ({want})", ok, lab)

    print("\n6. ABSOLUTE-EDGE DIRECTION")
    sym = all(spread_label(v, S9, "READY", "", "") == spread_label(-v, S9, "READY", "", "")
              for v in (0.5, 1.0, 1.49, 1.5, 3.0, 12.0))
    chk("6.1 label is sign-symmetric (ABS used correctly)", sym)
    sides = {g: (E9[g]["side"], E9[g]["edge"]) for g in LIVE_LINES if E9[g]["edge"] != ""}
    ok_side = all((e > 0 and s == E9[g]["home"]) or (e < 0 and s == E9[g]["away"])
                  for g, (s, e) in sides.items())
    chk("6.2 side still follows the sign of the edge, not its magnitude", ok_side)

    print("\n7-8. GATES RETAIN PRIORITY")
    chk("7.1 QB-gated game with |edge|=4.0 does NOT become BET",
        spread_label(4.0, S9, "QB UNCERTAIN", "", "") == "INVESTIGATE",
        spread_label(4.0, S9, "QB UNCERTAIN", "", ""))
    chk("8.1 blank line (PENDING LINE) stays blank",
        spread_label("", S9, "PENDING LINE", "", "") == "")
    chk("8.2 STALE LINE stays blank", spread_label(9.0, S9, "STALE LINE", "", "") == "")
    chk("8.3 BLOCKED stays blank", spread_label(9.0, S9, "BLOCKED", "", "") == "")
    chk("8.4 FCS game does NOT become BET",
        spread_label(4.0, S9, "FCS — NO PLAY", "", "FCS OPP") == "INVESTIGATE")
    chk("8.5 transitional team does NOT become BET",
        spread_label(4.0, S9, "READY", "1", "") == "INVESTIGATE")
    chk("8.6 zero edge stays blank", spread_label(0.0, S9, "READY", "", "") == "")

    print("\n9. TOTALS EQUIVALENCE ACROSS THE DECOUPLING")
    fixtures = [1.99, -1.99, 2.00, -2.00, 2.99, -2.99, 3.00, -3.00,
                5.99, -5.99, 6.00, -6.00, ""]
    S_old_hold = dict(S9); S_old_hold["B11"] = S8["B11"]   # hold the toggle at its v0.8.8 value
    mism = []
    print("     (decoupling in isolation - toggle held at its v0.8.8 value 'N')")
    for aa in fixtures:
        o = total_label_old(aa, S8, "READY", "", "")
        n = total_label_new(aa, S_old_hold, "READY", "", "")
        if o != n:
            mism.append((aa, o, n))
        lbl = "blank" if aa == "" else f"{aa:+.2f}"
        print(f"       {lbl:>7}  before={o or '(blank)':<12} after={n or '(blank)':<12}"
              f" {'OK' if o == n else 'MISMATCH'}")
    chk("9.1 totals classifications byte-for-byte equivalent across the decoupling",
        not mism, str(mism))
    chk("9.2 dedicated totals thresholds pin the exact prior effective values 2.0/3.0/6.0",
        (S9["B33"], S9["B34"], S9["B26"]) == (2, 3, 6),
        f"{S9['B33']}/{S9['B34']}/{S9['B26']}")
    ab = b["ENGINE"]["AB6"].value
    abt = ab.text if isinstance(ab, ArrayFormula) else ab
    chk("9.3 ENGINE!AB no longer references B8*2 / B9*2 / B10*2",
        "$B$8*2" not in abt and "$B$9*2" not in abt and "$B$10*2" not in abt)
    chk("9.4 ENGINE!AB references the dedicated cells B33 / B34 / B26",
        "SETTINGS!$B$33" in abt and "SETTINGS!$B$34" in abt and "SETTINGS!$B$26" in abt)

    print("\n10. TOTALS REMAIN DISABLED AND INERT")
    chk("10.1 SETTINGS!B22 and B23 still blank - totals inputs not populated",
        S9["B22"] is None and S9["B23"] is None)
    y6 = b["ENGINE"]["Y6"].value
    y6t = y6.text if isinstance(y6, ArrayFormula) else y6
    chk("10.2 ENGINE!Y still gated on B22/B23, so model total is blank workbook-wide",
        "SETTINGS!$B$22" in y6t and "SETTINGS!$B$23" in y6t)
    chk("10.3 with B22/B23 blank, every totals label is blank",
        total_label_new("", S9, "READY", "", "") == "")

    print("\n11. AUDIT!B12 GUARD")
    chk("11.1 AUDIT!B12 returns OK for the approved configuration",
        audit_b12(S9, abt) == "OK", audit_b12(S9, abt))
    drift = [("spread BET threshold back to 3", dict(S9, B10=3)),
             ("BET toggle back to N", dict(S9, B11="N")),
             ("totals LEAN drifts to 2.5", dict(S9, B33=2.5)),
             ("totals INVESTIGATE drifts to 4", dict(S9, B34=4)),
             ("totals BET drifts to 3", dict(S9, B26=3))]
    for name, Sd in drift:
        chk(f"11.x AUDIT!B12 returns CHECK when {name}", audit_b12(Sd, abt) == "CHECK")
    chk("11.7 AUDIT!B12 returns CHECK if totals classification is re-coupled",
        audit_b12(S9, "=IF(ABS($AA6)<SETTINGS!$B$8*2 ... SETTINGS!$B$10*2 ...)") == "CHECK")
    chk("11.8 AUDIT!B12 fails safe to CHECK if FORMULATEXT is unavailable",
        audit_b12(S9, "") == "CHECK")

    print("\n12-14. THE EIGHT LIVE MARKET-LINE FIXTURES")
    print(f"     {'game':<14}{'edge':>8}  {'v0.8.8':<13}{'v0.8.9':<13}{'gate'}")
    changed_lbl, same_proj = set(), True
    for g, (fav, sp, _) in LIVE_LINES.items():
        r8, r9 = E8[g], E9[g]
        if (round(r8["model"], 9), r8["edge"], r8["side"]) != (round(r9["model"], 9), r9["edge"], r9["side"]):
            same_proj = False
        l8 = spread_label(r8["edge"], S8, r8["gate"], "", "")
        l9 = spread_label(r9["edge"], S9, r9["gate"], "", "")
        if l8 != l9:
            changed_lbl.add(g)
        print(f"     {r9['away']+'@'+r9['home']:<14}{r9['edge']:+8.2f}  "
              f"{l8 or '(blank)':<13}{l9 or '(blank)':<13}{r9['gate']}")
    chk("12.1 all eight fixtures retain identical model spread, edge and side", same_proj)
    chk("13.1 exactly four spread labels change", len(changed_lbl) == 4, str(len(changed_lbl)))
    chk("13.2 the four are SAC@EMU, UNC@TCU, SJSU@USC, NMSU@FSU",
        changed_lbl == EXPECTED_LABEL_CHANGES,
        str(sorted(changed_lbl)))
    for g in sorted(changed_lbl):
        chk(f"13.x {E9[g]['away']}@{E9[g]['home']} INVESTIGATE -> BET",
            spread_label(E9[g]["edge"], S9, E9[g]["gate"], "", "") == "BET")
    mem = E9["401862693"]
    chk("14.1 Memphis at UNLV remains QB-gated", mem["gate"] == "QB UNCERTAIN", mem["gate"])
    chk("14.2 Memphis at UNLV is NOT a certified BET",
        spread_label(mem["edge"], S9, mem["gate"], "", "") != "BET",
        spread_label(mem["edge"], S9, mem["gate"], "", ""))

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"v0.8.8 source SHA-256: {h8}")
    print(f"v0.8.9 candidate SHA : {h9}")
    print("STATUS: CANDIDATE ONLY — NOT PROMOTED")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

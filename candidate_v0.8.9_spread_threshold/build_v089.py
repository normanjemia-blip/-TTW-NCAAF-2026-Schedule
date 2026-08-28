#!/usr/bin/env python3
"""Build the v0.8.9 SPREAD-THRESHOLD CANDIDATE from authoritative v0.8.8.

APPROVED RULE
  Spreads only: |ATS edge| >= 1.50 -> BET. Below 1.50 -> existing lower classes.
  QB / FCS / transitional / blank-line gates retain priority.
  No totals threshold approved; effective totals thresholds stay 2.0 / 3.0 / 6.0.

CHANGES (exactly as approved in THRESHOLD_AUDIT_1.5.md section 5)
  SETTINGS!A26  label for the dedicated totals thresholds
  SETTINGS!B26  = 6   totals BET          (pins today's B10*2)
  SETTINGS!B33  = 2   totals LEAN         (pins today's B8*2)
  SETTINGS!B34  = 3   totals INVESTIGATE  (pins today's B9*2)
  ENGINE!AB6:AB1005   repoint the TOTALS classification at B33 / B34 / B26
  SETTINGS!B10  3 -> 1.5   spread BET threshold
  SETTINGS!B11  N -> Y     spread BET toggle
  AUDIT!A12 / B12          config guard: OK only when the whole approved
                           configuration holds, CHECK on any drift

NOT a promotion. Nothing else is touched: no QB cell, no schedule date, no
rating, no adjustment, no market line, no banner.

Run:  python3 candidate_v0.8.9_spread_threshold/build_v089.py
"""
import hashlib, os, shutil, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, "promotion_v0.8.8",
                   "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
OUT = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.9_THRESHOLD_CANDIDATE.xlsx")
SRC_SHA = "b2a920feddc0f49f0647957334db0ecd0e922fe6a3933fc6a11af31587b56450"

TOTALS_LABEL = "TOTALS thresholds (abs edge) — BET here; LEAN B33 / INVESTIGATE B34"

REPOINT = (("SETTINGS!$B$8*2", "SETTINGS!$B$33"),
           ("SETTINGS!$B$9*2", "SETTINGS!$B$34"),
           ("SETTINGS!$B$10*2", "SETTINGS!$B$26"))

AUDIT_A12 = "Approved spread config: BET 1.5 / toggle Y / totals pinned 2-3-6 and decoupled"
AUDIT_B12 = (
    '=IF(AND('
    'SETTINGS!$B$10=1.5,'
    'SETTINGS!$B$11="Y",'
    'SETTINGS!$B$26=6,'
    'SETTINGS!$B$33=2,'
    'SETTINGS!$B$34=3,'
    'IFERROR(ISNUMBER(SEARCH("SETTINGS!$B$33",FORMULATEXT(ENGINE!$AB$6))),FALSE),'
    'IFERROR(ISNUMBER(SEARCH("SETTINGS!$B$34",FORMULATEXT(ENGINE!$AB$6))),FALSE),'
    'IFERROR(ISNUMBER(SEARCH("SETTINGS!$B$26",FORMULATEXT(ENGINE!$AB$6))),FALSE),'
    'IFERROR(NOT(ISNUMBER(SEARCH("$B$8*2",FORMULATEXT(ENGINE!$AB$6)))),FALSE),'
    'IFERROR(NOT(ISNUMBER(SEARCH("$B$9*2",FORMULATEXT(ENGINE!$AB$6)))),FALSE),'
    'IFERROR(NOT(ISNUMBER(SEARCH("$B$10*2",FORMULATEXT(ENGINE!$AB$6)))),FALSE)'
    '),"OK — spread BET 1.5, toggle Y; totals pinned 2.0/3.0/6.0 and decoupled",'
    '"CHECK — "'
    '&IF(SETTINGS!$B$10<>1.5,"spread BET threshold is "&SETTINGS!$B$10&"; ","")'
    '&IF(SETTINGS!$B$11<>"Y","BET toggle is "&SETTINGS!$B$11&"; ","")'
    '&IF(OR(SETTINGS!$B$26<>6,SETTINGS!$B$33<>2,SETTINGS!$B$34<>3),'
    '"totals thresholds not 2.0/3.0/6.0; ","")'
    '&IF(IFERROR(AND(ISNUMBER(SEARCH("SETTINGS!$B$26",FORMULATEXT(ENGINE!$AB$6))),'
    'NOT(ISNUMBER(SEARCH("$B$10*2",FORMULATEXT(ENGINE!$AB$6))))),FALSE),"",'
    '"totals classification not decoupled; "))')


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def txt(v):
    if isinstance(v, ArrayFormula):
        return v.text
    return v if isinstance(v, str) and v.startswith("=") else None


def main():
    got = sha256(SRC)
    assert got == SRC_SHA, f"source v0.8.8 is not the expected artifact: {got}"
    print(f"source v0.8.8 SHA-256 verified: {got}")

    shutil.copyfile(SRC, OUT + ".building.xlsx")
    wb = openpyxl.load_workbook(OUT + ".building.xlsx")
    st, en, au = wb["SETTINGS"], wb["ENGINE"], wb["AUDIT"]

    # --- preconditions ---
    assert st["B10"].value == 3, f"B10 expected 3, got {st['B10'].value!r}"
    assert st["B11"].value == "N", f"B11 expected 'N', got {st['B11'].value!r}"
    assert st["B8"].value == 1 and st["B9"].value == 1.5
    for coord in ("A26", "B26", "A33", "B33", "A34", "B34"):
        assert st[coord].value is None, f"SETTINGS!{coord} must be empty, got {st[coord].value!r}"
    assert st["B22"].value is None and st["B23"].value is None, "totals inputs must stay blank"
    print("preconditions verified: B10=3, B11='N', target SETTINGS rows empty, totals inert")

    written = []

    # --- 1/2. dedicated totals thresholds pinned to today's effective values ---
    st["A26"].value = TOTALS_LABEL;          written.append("SETTINGS!A26")
    st["B26"].value = 6;                     written.append("SETTINGS!B26")
    st["B33"].value = 2;                     written.append("SETTINGS!B33")
    st["B34"].value = 3;                     written.append("SETTINGS!B34")
    print("dedicated totals thresholds added: B33=2, B34=3, B26=6 (= today's B8*2/B9*2/B10*2)")

    # --- 3. repoint ONLY the totals classification ---
    n = 0
    for r in range(6, 1006):
        f = txt(en.cell(row=r, column=28).value)   # column AB
        if not f:
            continue
        new = f
        for old, rep in REPOINT:
            new = new.replace(old, rep)
        assert "SETTINGS!$B$8*2" not in new and "SETTINGS!$B$9*2" not in new \
            and "SETTINGS!$B$10*2" not in new, f"AB{r} still coupled"
        assert new != f, f"AB{r} unchanged"
        # everything except the three threshold tokens must be untouched
        back = new.replace("SETTINGS!$B$33", "SETTINGS!$B$8*2") \
                  .replace("SETTINGS!$B$34", "SETTINGS!$B$9*2") \
                  .replace("SETTINGS!$B$26", "SETTINGS!$B$10*2")
        assert back == f, f"AB{r}: more than the threshold tokens changed"
        en.cell(row=r, column=28).value = new
        written.append(f"ENGINE!AB{r}")
        n += 1
    assert n == 1000, f"expected 1000 totals formulas repointed, got {n}"
    print(f"totals classification repointed: {n} cells (ENGINE!AB6:AB1005)")

    # --- 4/5. spread threshold and toggle ---
    st["B10"].value = 1.5;   written.append("SETTINGS!B10")
    st["B11"].value = "Y";   written.append("SETTINGS!B11")
    print("spread BET threshold B10 -> 1.5 ; spread BET toggle B11 -> 'Y'")

    # --- audit guard ---
    au["A12"].value = AUDIT_A12;  written.append("AUDIT!A12")
    au["B12"].value = AUDIT_B12;  written.append("AUDIT!B12")
    print("AUDIT!A12/B12 rewritten as the approved-configuration guard")

    # --- nothing else may have moved ---
    assert st["B8"].value == 1 and st["B9"].value == 1.5, "B8/B9 must not change"
    assert st["B22"].value is None and st["B23"].value is None, "totals must stay disabled"
    assert st["B6"].value == 2.5 and st["B7"].value == 0 and st["B12"].value == 2.5

    wb.save(OUT + ".building.xlsx")
    os.replace(OUT + ".building.xlsx", OUT)
    print(f"\ncells written: {len(written)}")
    print(f"written: {OUT}")
    print(f"v0.8.9 candidate SHA-256: {sha256(OUT)}")
    assert sha256(SRC) == SRC_SHA, "authoritative v0.8.8 was modified by the build"
    print("authoritative v0.8.8 confirmed unmodified")
    return 0


if __name__ == "__main__":
    sys.exit(main())

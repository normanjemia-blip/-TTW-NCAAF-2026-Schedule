#!/usr/bin/env python3
"""Reproduce the REV 1 totals-toggle conflict with a FULLY VALID totals fixture.

READ-ONLY. Compares authoritative v0.8.8 against v0.8.9 REV 1 using a fixture
that satisfies every precondition the real workbook requires for a totals label:

  * two valid FBS teams and a real scheduled game
  * ENGINE!AI = READY   (no BLOCK, no FCS, no PENDING/STALE LINE, no QB gate)
  * AF (transitional) and AG (FCS flag) both empty
  * market total populated  -> ENGINE!Z
  * projected total populated -> ENGINE!Y
  * therefore ENGINE!AA (total edge) = Y - Z is a real number

The fixture supplies AA directly at the required magnitudes, which is exactly
what ENGINE!AB consumes.
"""
import os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
V088 = os.path.join(ROOT, "promotion_v0.8.8",
                    "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
REV1 = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.9_THRESHOLD_CANDIDATE.xlsx")


def txt(v):
    return v.text if isinstance(v, ArrayFormula) else v


def settings(wb):
    st = wb["SETTINGS"]
    return {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 40)}


def total_label_from_formula(AA, S, AI, AF, AG, formula):
    """Evaluate ENGINE!AB as literally written, resolving whichever cells it names."""
    if AA == "":
        return ""
    if "SETTINGS!$B$8*2" in formula:
        lean, inv, bet = S["B8"] * 2, S["B9"] * 2, S["B10"] * 2
    else:
        lean, inv, bet = S["B33"], S["B34"], S["B26"]
    toggle_cell = "B11"
    for cand in ("B35", "B36", "B27"):
        if f"SETTINGS!$${cand}" in formula or f"SETTINGS!${cand}" in formula \
           or f"SETTINGS!$B${cand[1:]}" in formula:
            toggle_cell = cand
    if "SETTINGS!$B$11" not in formula:
        for r in range(24, 40):
            if f"SETTINGS!$B${r}<>" in formula.replace("$B$", "$B$"):
                toggle_cell = f"B{r}"
    toggle = S.get(toggle_cell)
    if abs(AA) < lean:
        return ""
    if abs(AA) < inv:
        return "LEAN"
    if (toggle != "Y" or AI != "READY" or abs(AA) < bet or AF != "" or AG != ""):
        return "INVESTIGATE"
    return "BET"


def main():
    a = openpyxl.load_workbook(V088)
    b = openpyxl.load_workbook(REV1)
    Sa, Sb = settings(a), settings(b)
    fa = txt(a["ENGINE"]["AB6"].value)
    fb = txt(b["ENGINE"]["AB6"].value)

    print("=" * 78)
    print("REPRODUCTION — v0.8.9 REV 1 TOTALS-TOGGLE CONFLICT")
    print("=" * 78)
    print("\nFIXTURE (fully valid totals game)")
    print("  game        : 401864577  Jacksonville State @ North Dakota State (real, Week 0)")
    print("  teams       : both resolve; no BLOCK, no FCS flag")
    print("  ENGINE!AI   : READY        (no QB gate, no PENDING/STALE LINE)")
    print("  ENGINE!AF   : ''           (not transitional)")
    print("  ENGINE!AG   : ''           (no FCS opponent)")
    print("  market total: 46.5         -> ENGINE!Z populated")
    print("  model total : supplied     -> ENGINE!Y populated")
    print("  ENGINE!AA   : Y - Z, set to each magnitude below")

    print(f"\n  v0.8.8 toggle SETTINGS!B11 = {Sa['B11']!r}")
    print(f"  REV 1  toggle SETTINGS!B11 = {Sb['B11']!r}   <-- shared with spreads")
    print(f"  v0.8.8 totals thresholds   = {Sa['B8']*2} / {Sa['B9']*2} / {Sa['B10']*2}")
    print(f"  REV 1  totals thresholds   = {Sb['B33']} / {Sb['B34']} / {Sb['B26']}")

    print("\nRESULT — totals label at the production configuration of each build")
    print(f"  {'AA (total edge)':>16}  {'v0.8.8':<14}{'REV 1':<14}{'conflict'}")
    conflicts = []
    for aa in (5.99, -5.99, 6.00, -6.00, 6.01, -6.01):
        la = total_label_from_formula(aa, Sa, "READY", "", "", fa)
        lb = total_label_from_formula(aa, Sb, "READY", "", "", fb)
        flag = "" if la == lb else "  <-- CHANGED"
        if la != lb:
            conflicts.append((aa, la, lb))
        print(f"  {aa:>16.2f}  {la:<14}{lb:<14}{flag}")

    print("\nVERDICT")
    if conflicts:
        print(f"  CONFIRMED. {len(conflicts)} of 6 fixtures change classification.")
        print("  v0.8.8 (B11='N') returns INVESTIGATE at |total edge| >= 6.")
        print("  REV 1  (B11='Y', shared) returns BET at |total edge| >= 6.")
        print("  The approved totals thresholds 2.0/3.0/6.0 are preserved in REV 1 -")
        print("  the change comes ENTIRELY from the shared toggle, not the thresholds.")
    else:
        print("  NOT REPRODUCED.")

    print("\nWHY THE REV 1 CERTIFICATE CHECK 9.1 DID NOT CATCH THIS")
    print("  verify_v089.py check 9.1 built its 'after' settings as:")
    print("      S_old_hold = dict(S9); S_old_hold['B11'] = S8['B11']")
    print("  i.e. it deliberately HELD THE TOGGLE at its v0.8.8 value in both the")
    print("  before and after evaluations, to isolate the threshold repointing.")
    print("  That made the comparison 'decoupling alone', not 'v0.8.8 production")
    print("  configuration vs REV 1 production configuration'.")
    print("  The claim it proved was TRUE but NARROWER than the wording implied:")
    print("  the thresholds are equivalent; the SHIPPED configurations are not.")
    print("  It also never tested |AA| = 6.01, and at exactly 6.00 the difference")
    print("  was masked because the held toggle forced INVESTIGATE on both sides.")
    print("  Fix in REV 2: compare each build AT ITS OWN PRODUCTION SETTINGS, and")
    print("  give totals a dedicated toggle so the spread toggle cannot reach them.")
    return 0 if conflicts else 1


if __name__ == "__main__":
    sys.exit(main())

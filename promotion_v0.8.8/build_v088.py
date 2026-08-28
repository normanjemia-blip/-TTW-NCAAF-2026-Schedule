#!/usr/bin/env python3
"""Build v0.8.8 from the certified schedule candidate.

APPROVED SCOPE — the certified schedule-date correction, plus the single
administrative version banner.

  SCHEDULE DATES (133 cells)   inherited from the certified candidate:
      IMPORT SCHEDULE!D, every change exactly -1 calendar day, venue-local
      semantics, the 403 placeholder-time rows untouched.

  BANNER (1 cell)              START HERE!A1 -- version token only.

      The candidate was verified 53/0 while still carrying the v0.8.7 banner,
      so this build applies the version token here. THE FINAL WORKBOOK IS
      THEREFORE NOT BYTE-IDENTICAL TO THE CANDIDATE, and the promotion diff
      against v0.8.7 is 134 cells, not 133.

      Only the version token changes. The confidence census in the banner is
      already correct at 76 H / 43 M / 19 L because this promotion makes no QB
      change. Nothing else in the banner is touched -- including the
      "promotion complete 2026-08-04" clause, which has been carried forward
      unchanged since v0.8.0 and is out of scope here.

NO QB CHANGE. NO FORMULA CHANGE. NO SCHEDULE SWEEP. NO NEW KICKOFF TIMES.

Run:  python3 promotion_v0.8.8/build_v088.py
"""
import hashlib, os, shutil, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAND = os.path.join(ROOT, "schedule_candidate_v1",
                    "TTW_College_Football_Power_Ratings_v0.8.8_SCHEDULE_CANDIDATE.xlsx")
PRED = os.path.join(ROOT, "promotion_v0.8.7",
                    "TTW_College_Football_Power_Ratings_v0.8.7_AUTHORITATIVE.xlsx")
OUT = os.path.join(ROOT, "promotion_v0.8.8",
                   "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")

CAND_SHA = "5416ffcb4c07b8e741f24f51b9603ac44c064db943e144618d6ffa372ef62a84"
PRED_SHA = "46671deeaaa94d98c63cb32d0e94af9907e76e7e2638de431b918987df2e15cd"

OLD_TOKEN = "v0.8.7 AUTHORITATIVE"
NEW_TOKEN = "v0.8.8 AUTHORITATIVE"
CENSUS = "76 H / 43 M / 19 L"


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def main():
    got_c, got_p = sha256(CAND), sha256(PRED)
    assert got_c == CAND_SHA, f"candidate is not the certified artifact: {got_c}"
    assert got_p == PRED_SHA, f"v0.8.7 is not the expected predecessor: {got_p}"
    print(f"certified candidate SHA-256 verified: {got_c}")
    print(f"frozen predecessor SHA-256 verified : {got_p}")

    shutil.copyfile(CAND, OUT + ".building.xlsx")
    wb = openpyxl.load_workbook(OUT + ".building.xlsx")
    sh = wb["START HERE"]

    banner = sh["A1"].value
    assert OLD_TOKEN in banner, f"expected the v0.8.7 banner token, got: {banner[:120]}"
    assert CENSUS in banner, f"banner census must already read {CENSUS}: {banner[:200]}"
    assert NEW_TOKEN not in banner, "banner already carries v0.8.8 - nothing to do"
    new_banner = banner.replace(OLD_TOKEN, NEW_TOKEN)
    # exactly one token substitution, nothing else about the string may move
    assert new_banner != banner
    assert new_banner.replace(NEW_TOKEN, OLD_TOKEN) == banner, \
        "the banner edit must be the version token and nothing else"
    assert CENSUS in new_banner and "v0.8.7" not in new_banner
    sh["A1"].value = new_banner
    print(f"banner updated: {OLD_TOKEN!r} -> {NEW_TOKEN!r} (census unchanged at {CENSUS})")

    wb.save(OUT + ".building.xlsx")
    os.replace(OUT + ".building.xlsx", OUT)

    # prove the ONLY difference from the candidate is that one banner cell
    a = openpyxl.load_workbook(CAND)
    b = openpyxl.load_workbook(OUT)
    diffs = []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                if norm(wa.cell(row=r, column=c).value) != norm(wbk.cell(row=r, column=c).value):
                    diffs.append((s, wbk.cell(row=r, column=c).coordinate))
    assert diffs == [("START HERE", "A1")], f"expected only the banner to differ, got {diffs[:6]}"
    print("confirmed: candidate -> v0.8.8 differs by exactly START HERE!A1")

    print(f"written: {OUT}")
    print(f"v0.8.8 SHA-256: {sha256(OUT)}")
    assert sha256(PRED) == PRED_SHA, "v0.8.7 was modified by the build"
    assert sha256(CAND) == CAND_SHA, "the candidate was modified by the build"
    print("v0.8.7 and the candidate both confirmed unmodified")


if __name__ == "__main__":
    sys.exit(main())

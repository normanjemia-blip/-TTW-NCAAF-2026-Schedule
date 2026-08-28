#!/usr/bin/env python3
"""v0.8.8 promotion certificate — READ-ONLY. Writes nothing.

v0.8.8 = v0.8.7 + the certified 133 schedule-date corrections in
IMPORT SCHEDULE!D + exactly one administrative banner cell.

Exit code 0 iff every check passes.
"""
import collections, csv, datetime, hashlib, io, json, os, re, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
from zoneinfo import ZoneInfo

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "schedule_candidate_v1"))
from espn_date_rule import start_date, assert_not_utc_dates  # noqa: E402

UTC = ZoneInfo("UTC")
V087 = os.path.join(ROOT, "promotion_v0.8.7",
                    "TTW_College_Football_Power_Ratings_v0.8.7_AUTHORITATIVE.xlsx")
V088 = os.path.join(ROOT, "promotion_v0.8.8",
                    "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
CAND = os.path.join(ROOT, "schedule_candidate_v1",
                    "TTW_College_Football_Power_Ratings_v0.8.8_SCHEDULE_CANDIDATE.xlsx")
SUPERSEDED = "TTW_College_Football_Power_Ratings_SCHED1_CANDIDATE.xlsx"
SNAP = os.path.join(ROOT, "schedule_candidate_v1", "espn_kickoff_snapshot.csv")
FROZEN_V087 = "46671deeaaa94d98c63cb32d0e94af9907e76e7e2638de431b918987df2e15cd"

EXPECT_QB_STATUS = (117, 21)
EXPECT_QB_CONF = (76, 43, 19)
EXPECT_QB_ZEROS = 234

PASS, FAIL = [], []


def chk(msg, ok, detail=""):
    (PASS if ok else FAIL).append(msg)
    print(f"  [{'PASS' if ok else 'FAIL'}] {msg}" + (f" - {detail}" if detail else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def mean(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def main():
    print("=" * 78)
    print("v0.8.8 PROMOTION CERTIFICATE — schedule dates + one banner cell")
    print("=" * 78)
    h87, h88 = sha256(V087), sha256(V088)
    print(f"  v0.8.7 SHA-256: {h87}")
    print(f"  v0.8.8 SHA-256: {h88}")

    print("\n0. PREDECESSOR")
    chk("0.1 v0.8.7 retains its frozen SHA-256", h87 == FROZEN_V087, h87[:16])
    chk("0.2 v0.8.8 differs from v0.8.7", h88 != h87)

    a = openpyxl.load_workbook(V087)
    b = openpyxl.load_workbook(V088)
    chk("0.3 21 sheets, identical names and order",
        a.sheetnames == b.sheetnames and len(b.sheetnames) == 21)
    chk("0.4 sheet visibility preserved",
        {s: a[s].sheet_state for s in a.sheetnames} == {s: b[s].sheet_state for s in b.sheetnames})
    chk("0.5 named ranges preserved",
        sorted(a.defined_names.keys()) == sorted(b.defined_names.keys()))

    print("\n1-2-6. SCOPE: 133 SCHEDULE CELLS + 1 BANNER, NO FORMULA CHANGE")
    changed, formula_changed = [], []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                x, y = norm(wa.cell(row=r, column=c).value), norm(wbk.cell(row=r, column=c).value)
                if x != y:
                    changed.append((s, r, c, wbk.cell(row=r, column=c).coordinate))
                    if x[0] == "F" or y[0] == "F":
                        formula_changed.append((s, wbk.cell(row=r, column=c).coordinate))
    chk("6.1 ZERO formula differences", not formula_changed, str(formula_changed[:5]))
    bysheet = collections.Counter(s for s, _, _, _ in changed)
    chk("1.1 exactly 134 cells changed: 133 schedule + 1 banner",
        len(changed) == 134, str(len(changed)))
    chk("1.2 only IMPORT SCHEDULE (133) and START HERE (1) changed",
        dict(bysheet) == {"IMPORT SCHEDULE": 133, "START HERE": 1}, str(dict(bysheet)))
    sched = [x for x in changed if x[0] == "IMPORT SCHEDULE"]
    chk("2.1 every schedule change is in column D (start_date)",
        {c for _, _, c, _ in sched} == {4}, str(sorted({c for _, _, c, _ in sched})))
    banner_cells = [x[3] for x in changed if x[0] == "START HERE"]
    chk("1.3 the single banner cell is START HERE!A1", banner_cells == ["A1"], str(banner_cells))

    print("\n1b. THE BANNER EDIT IS THE VERSION TOKEN AND NOTHING ELSE")
    ba, bb = a["START HERE"]["A1"].value, b["START HERE"]["A1"].value
    chk("1.4 v0.8.8 banner declares v0.8.8 AUTHORITATIVE", "v0.8.8 AUTHORITATIVE" in bb)
    chk("1.5 no stale v0.8.7 identifier remains in the banner", "v0.8.7" not in bb)
    chk("1.6 reversing the token reproduces the v0.8.7 banner exactly - nothing else moved",
        bb.replace("v0.8.8 AUTHORITATIVE", "v0.8.7 AUTHORITATIVE") == ba)
    chk("1.7 banner confidence census still reads 76 H / 43 M / 19 L",
        "76 H / 43 M / 19 L" in bb)

    sa, sb = a["IMPORT SCHEDULE"], b["IMPORT SCHEDULE"]
    snap = {r["id"]: r for r in csv.DictReader(io.open(SNAP, encoding="utf-8"))}

    print("\n3. EVERY SCHEDULE-DATE DELTA IS EXACTLY -1 DAY")
    deltas, bad = collections.Counter(), []
    for r in range(6, 900):
        gid = sa.cell(row=r, column=1).value
        if gid is None:
            continue
        x, y = sa.cell(row=r, column=4).value, sb.cell(row=r, column=4).value
        if x != y:
            deltas[(y - x).days] += 1
        want = datetime.date.fromisoformat(snap[str(gid)]["canonical_start_date"])
        if y.date() != want:
            bad.append((str(gid), str(y.date()), str(want)))
    chk("3.1 every schedule change is exactly -1 day", set(deltas) == {-1}, str(dict(deltas)))
    chk("3.2 133 rows moved", sum(deltas.values()) == 133, str(sum(deltas.values())))
    chk("3.3 all 888 dates equal the canonical venue-local rule", not bad, str(bad[:5]))

    print("\n4. THE 403 PLACEHOLDER-TIME ROWS ARE UNCHANGED")
    ph = {gid for gid, s_ in snap.items() if s_["time_valid"] != "True"}
    chk("4.1 exactly 403 rows carry a placeholder kickoff time", len(ph) == 403, str(len(ph)))
    moved_ph = [str(sa.cell(row=r, column=1).value) for r in range(6, 900)
                if sa.cell(row=r, column=1).value is not None
                and str(sa.cell(row=r, column=1).value) in ph
                and sa.cell(row=r, column=4).value != sb.cell(row=r, column=4).value]
    chk("4.2 all 403 placeholder-time rows unchanged", not moved_ph, str(moved_ph[:5]))
    chk("4.3 no newly announced kickoff time was incorporated - the snapshot still "
        "reports 403 awaiting re-derivation",
        sum(1 for s_ in snap.values() if s_["needs_rederivation"] == "True") == 403)

    print("\n5. EVENT IDS")
    ids = [str(sb.cell(row=r, column=1).value) for r in range(6, 900)
           if sb.cell(row=r, column=1).value is not None]
    chk("5.1 all 888 event ids present and UNIQUE",
        len(ids) == 888 and len(set(ids)) == 888, f"n={len(ids)} unique={len(set(ids))}")
    chk("5.2 id column byte-identical to v0.8.7",
        all(sa.cell(row=r, column=1).value == sb.cell(row=r, column=1).value
            for r in range(6, 900)))
    for col, name in ((2, "season"), (3, "week"), (5, "neutral_site"), (6, "away_team"),
                      (8, "home_team"), (13, "venue"), (14, "notes")):
        chk(f"5.x column {name} byte-identical",
            all(sa.cell(row=r, column=col).value == sb.cell(row=r, column=col).value
                for r in range(6, 900)))

    print("\n7. WEEKS AND THE WEEK 0 BOUNDARY")
    wk_a = collections.Counter(sa.cell(row=r, column=3).value for r in range(6, 900)
                               if sa.cell(row=r, column=1).value is not None)
    wk_b = collections.Counter(sb.cell(row=r, column=3).value for r in range(6, 900)
                               if sb.cell(row=r, column=1).value is not None)
    chk("7.1 week distribution identical - no game changes week", wk_a == wk_b, f"{len(wk_b)} weeks")
    W0_END = datetime.date(2026, 9, 2)
    crossed = sum(1 for r in range(6, 900)
                  if sa.cell(row=r, column=1).value is not None
                  and (sa.cell(row=r, column=4).value.date() <= W0_END)
                  != (sb.cell(row=r, column=4).value.date() <= W0_END))
    chk("7.2 zero games cross the Week 0 boundary", crossed == 0, str(crossed))
    w0 = [sb.cell(row=r, column=4).value.date() for r in range(6, 900)
          if sb.cell(row=r, column=3).value == 0 and sb.cell(row=r, column=1).value is not None]
    chk("7.3 Week 0 still holds exactly 8 games", len(w0) == 8, str(len(w0)))
    chk("7.4 Week 0 is a single Saturday, 2026-08-29",
        set(w0) == {datetime.date(2026, 8, 29)}, str(sorted({str(d) for d in w0})))

    print("\n8-9. THE HEADLINE DATES")
    mem = [(sa.cell(row=r, column=4).value.date(), sb.cell(row=r, column=4).value.date())
           for r in range(6, 900) if str(sb.cell(row=r, column=1).value) == "401862693"]
    chk("8.1 Memphis at UNLV is Saturday 2026-08-29",
        len(mem) == 1 and mem[0][0] == datetime.date(2026, 8, 30)
        and mem[0][1] == datetime.date(2026, 8, 29) and mem[0][1].weekday() == 5,
        str([(str(x), str(y)) for x, y in mem]))
    def sundays(ws):
        return sum(1 for r in range(6, 900) if ws.cell(row=r, column=1).value is not None
                   and ws.cell(row=r, column=4).value.weekday() == 6)
    chk("9.1 Sunday games are 3 (were 70 in v0.8.7)",
        sundays(sb) == 3 and sundays(sa) == 70, f"before={sundays(sa)} after={sundays(sb)}")
    genuine = {(sb.cell(row=r, column=6).value, sb.cell(row=r, column=8).value,
                sb.cell(row=r, column=4).value.date().isoformat())
               for r in range(6, 900) if sb.cell(row=r, column=1).value is not None
               and sb.cell(row=r, column=4).value.weekday() == 6}
    chk("9.2 the three genuine Sunday games remain Sunday 2026-09-06",
        genuine == {("Louisville Cardinals", "Ole Miss Rebels", "2026-09-06"),
                    ("Washington State Cougars", "Washington Huskies", "2026-09-06"),
                    ("Wisconsin Badgers", "Notre Dame Fighting Irish", "2026-09-06")},
        str(sorted(genuine)))

    print("\n10-13. QB STATE — NOTHING MOVED")
    tm, qb, st = b["TEAM MAP"], b["QB VALUES"], b["SETTINGS"]
    season = st["B3"].value
    codes, sts, zeros, nonzero = collections.Counter(), collections.Counter(), 0, []
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        for v in (D, F):
            if v == 0:
                zeros += 1
            elif v is not None:
                nonzero.append((ab, v))
        G = "" if (D is None or F is None) else F - D
        codes[H] += 1
        sts["UNCERTAIN" if (G == "" or H == "L" or J != season) else "OK"] += 1
    chk(f"10.1 QB status census remains {EXPECT_QB_STATUS[0]} OK / {EXPECT_QB_STATUS[1]} UNCERTAIN",
        (sts["OK"], sts["UNCERTAIN"]) == EXPECT_QB_STATUS, str(dict(sts)))
    chk(f"11.1 confidence census remains {EXPECT_QB_CONF[0]} H / {EXPECT_QB_CONF[1]} M / "
        f"{EXPECT_QB_CONF[2]} L",
        (codes["H"], codes["M"], codes["L"]) == EXPECT_QB_CONF, str(dict(codes)))
    chk(f"12.1 QB zero count remains {EXPECT_QB_ZEROS}", zeros == EXPECT_QB_ZEROS, str(zeros))
    chk("12.2 still zero nonzero QB values", not nonzero, str(nonzero[:5]))
    qb_a = a["QB VALUES"]
    same_qb = all(norm(qb_a.cell(row=r, column=c).value) == norm(qb.cell(row=r, column=c).value)
                  for r in range(1, 200) for c in range(1, 14))
    chk("13.1 the ENTIRE QB VALUES sheet is byte-identical to v0.8.7 - no QB value or "
        "status changed", same_qb)

    print("\n14-15. MODEL OUTPUTS")
    ps = b["PRESEASON"]
    S = {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 33)}
    raw = {c: [ps.cell(row=r, column=c).value for r in range(6, 144)] for c in (4, 8, 17)}
    mu = {c: mean(v) for c, v in raw.items()}
    W = {4: S["B28"], 8: S["B29"], 17: S["B31"]}
    prior = {}
    for i, r in enumerate(range(6, 144)):
        ab = tm.cell(row=r, column=1).value
        num = den = 0.0
        for c in (4, 8, 17):
            v = raw[c][i]
            if isinstance(v, (int, float)):
                num += W[c] * (v - mu[c]); den += W[c]
        prior[ab] = (num / den) if den else ""
    alias = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()
    delta = {}
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F = qb.cell(row=r, column=4).value, qb.cell(row=r, column=6).value
        delta[ab] = "" if (D is None or F is None) else F - D
    z = lambda v: 0 if v == "" else v
    got, n_games, fcs = {}, 0, 0
    for r in range(6, 1006):
        gid = sb.cell(row=r, column=1).value
        if not gid:
            continue
        n_games += 1
        A = alias.get(str(sb.cell(row=r, column=6).value).strip(), "")
        H = alias.get(str(sb.cell(row=r, column=8).value).strip(), "")
        if not A or not H:
            fcs += 1; continue
        neutral = bool(sb.cell(row=r, column=5).value)
        got[(A, H)] = (prior[H] - prior[A]) + (S["B7"] if neutral else S["B6"]) \
                      + (z(delta.get(H, "")) - z(delta.get(A, "")))
    chk("15.1 888 games / 761 FBS-v-FBS / 127 FCS-involved",
        n_games == 888 and fcs == 127 and n_games - fcs == 761, f"{n_games}/{n_games-fcs}/{fcs}")
    for (A, H), want in ((("MEM", "UNLV"), "UNLV -5.6"), (("UNC", "TCU"), "TCU -4.2"),
                         (("NMSU", "FSU"), "FSU -27.7"), (("SJSU", "USC"), "USC -35.2"),
                         (("HAW", "STAN"), "STAN -3.7")):
        m_ = got[(A, H)]
        lab = f"{H} -{abs(m_):.1f}" if m_ > 0 else f"{A} -{abs(m_):.1f}"
        chk(f"14.x {A} at {H} spread unchanged at {want}", lab == want, lab)

    print("\n15b. NO EDGE, SIDE, LABEL OR GATE CAN MOVE")
    def txt(v):
        if isinstance(v, ArrayFormula):
            return v.text
        return v if isinstance(v, str) and v.startswith("=") else None
    calc_dates = sum(1 for row in b["CALC"].iter_rows() for c in row
                     if (txt(c.value) or "") and re.search(
                         r"(CLEAN!\$?D|ENGINE!\$?C|'IMPORT SCHEDULE'!\$?D)", txt(c.value)))
    chk("15.2 CALC - which drives every gate, edge and side - touches NO date column",
        calc_dates == 0, str(calc_dates))
    today = sum(1 for ws in b.worksheets for row in ws.iter_rows() for c in row
                if "TODAY()" in (txt(c.value) or "").upper())
    chk("15.3 no TODAY() anywhere - no date-relative drift", today == 0, str(today))
    ml = b["MARKET LINES"]
    lines = sum(1 for r in range(6, 1006)
                if ml.cell(row=r, column=1).value is not None
                or ml.cell(row=r, column=4).value is not None)
    chk("15.4 MARKET LINES still blank - no line, edge, side or label can move",
        lines == 0, str(lines))
    gated = sum(1 for r in range(6, 144) if tm.cell(row=r, column=1).value
                and (qb.cell(row=r, column=4).value is None
                     or qb.cell(row=r, column=6).value is None
                     or qb.cell(row=r, column=8).value == "L"
                     or qb.cell(row=r, column=10).value != season))
    chk("15.5 QB gate population unchanged - 21 rows still UNCERTAIN", gated == 21, str(gated))
    chk("15.6 BET toggle remains N", st["B11"].value == "N")
    chk("15.7 totals remain unavailable (B22/B23 blank)",
        st["B22"].value is None and st["B23"].value is None)
    au = b["AUDIT"]
    chk("15.8 AUDIT market-line invariant unchanged", "SUMPRODUCT" in str(au["B16"].value))

    print("\n16-17. IDEMPOTENCE AND THE REGRESSION GUARD")
    recs_new, recs_old = [], []
    for r in range(6, 900):
        gid = sb.cell(row=r, column=1).value
        if gid is None:
            continue
        s_ = snap[str(gid)]
        kick = datetime.datetime.strptime(s_["espn_kickoff_utc"], "%Y-%m-%dT%H:%MZ").replace(tzinfo=UTC)
        addr = dict(city=s_["venue_city"], state=s_["venue_state"],
                    country=s_["venue_country"] or "USA")
        tv = s_["time_valid"] == "True"
        recs_new.append(dict(id=str(gid), stored_date=sb.cell(row=r, column=4).value.date(),
                             kickoff_utc=kick, address=addr, time_valid=tv))
        recs_old.append(dict(id=str(gid), stored_date=sa.cell(row=r, column=4).value.date(),
                             kickoff_utc=kick, address=addr, time_valid=tv))
    reapply = [x["id"] for x in recs_new
               if start_date(x["kickoff_utc"], x["address"], x["time_valid"]) != x["stored_date"]]
    chk("16.1 the date rule is idempotent - re-applying to v0.8.8 changes nothing",
        not reapply, str(reapply[:5]))
    ok_new = True
    try:
        assert_not_utc_dates(recs_new)
    except AssertionError:
        ok_new = False
    chk("16.2 refresh guard PASSES on v0.8.8", ok_new)
    caught = 0
    try:
        assert_not_utc_dates(recs_old)
    except AssertionError as e:
        caught = int(str(e).split()[0])
    chk("17.1 refresh guard detects all 133 predecessor UTC-date defects in v0.8.7",
        caught == 133, f"guard flagged {caught}")

    print("\n18-19. POINTERS AND THE SUPERSEDED CANDIDATE")
    man = json.load(io.open(os.path.join(ROOT, "PROJECT_MANIFEST.json"), encoding="utf-8"))
    readme = io.open(os.path.join(ROOT, "README.md"), encoding="utf-8").read()
    pol = io.open(os.path.join(ROOT, "phase9a_production_config",
                               "MASTER_AND_WORKING_COPY_POLICY.md"), encoding="utf-8").read()
    dry = io.open(os.path.join(ROOT, "phase11_week0_dryrun", "week0_dryrun.py"),
                  encoding="utf-8").read()

    # v0.8.8 was production when this certificate was written, so 18.1-18.9 asserted that the
    # pointers named it as CURRENT. Those assertions are mutually exclusive with any later approved
    # promotion, so the section is supersession-aware: while v0.8.8 is current the original
    # assertions run verbatim and unweakened; once a later version is promoted the same section
    # asserts the obligation that actually binds then - v0.8.8 preserved as the immediate rollback
    # with its exact SHA. Forward pointer assertions move to the successor's own certificate.
    superseded = man["current_version"]["version"] != "v0.8.8"

    if not superseded:
        chk("18.1 manifest current_version is v0.8.8",
            man["current_version"]["version"] == "v0.8.8", man["current_version"]["version"])
        chk("18.2 manifest current sha256 is the final v0.8.8 workbook",
            man["current_version"]["sha256"] == h88)
        chk("18.3 manifest current_authoritative points at the v0.8.8 workbook",
            man["current_authoritative"]["source_sha256"] == h88
            and "v0.8.8" in man["current_authoritative"]["source_xlsx"])
        chk("18.4 manifest records v0.8.7 as the frozen predecessor with its exact SHA",
            FROZEN_V087 in man["current_version"]["supersedes"])
        chk("18.6 README production pointer names v0.8.8 and its final SHA",
            "v0.8.8 AUTHORITATIVE" in readme and h88 in readme)
        chk("18.7 policy current-version reference names v0.8.8 and its final SHA",
            "v0.8.8" in pol and h88 in pol)
        chk("18.9 Week 0 dry run targets the v0.8.8 workbook and SHA",
            "promotion_v0.8.8" in dry and h88 in dry)
    else:
        cur = man["current_version"]["version"]
        print(f"  (v0.8.8 superseded by {cur} - asserting rollback preservation)")
        chk(f"18.1s manifest current_version advanced past v0.8.8 by promotion", True, cur)
        chk("18.2s manifest no longer claims v0.8.8 is current",
            man["current_version"]["sha256"] != h88)
        chk("18.3s current_authoritative advanced past v0.8.8, whose artifact stays byte-exact",
            man["current_authoritative"]["source_sha256"] != h88 and sha256(V088) == h88)
        chk("18.4s successor records v0.8.8 as the frozen predecessor with its exact SHA",
            h88 in man["current_version"]["supersedes"]
            and "v0.8.8" in man["current_version"]["supersedes"])
        chk("18.6s README still preserves v0.8.8 and its exact SHA", h88 in readme)
        chk("18.7s policy still preserves v0.8.8 frozen with its exact SHA",
            "v0.8.8" in pol and h88 in pol)
        chk("18.9s Week 0 dry run advanced past v0.8.8 and no longer asserts its SHA",
            "promotion_v0.8.8" not in dry and h88 not in dry)
    chk("18.5 rollback hashes preserved and correct",
        man["rollback"]["source_sha256"] ==
        "bbb17b50fbfb728bea2a23d3d20771935cc61e238313a054473aafe1ca838efd"
        and man["intermediate_rollback"]["source_sha256"] ==
        "661f8ab0e6120290d4ffd8d4ddac738d7e19d7bd0bbcf69bc9df51fb3cef97c7")
    chk("18.8 policy still preserves v0.8.7 as frozen with its exact SHA", FROZEN_V087 in pol)
    for name, blob in (("README", readme), ("policy", pol), ("manifest", json.dumps(man)),
                       ("week0 dry run", dry)):
        chk(f"19.x {name} does not reference the superseded SCHED1 candidate as current",
            SUPERSEDED not in blob)
    chk("19.5 the superseded candidate file still exists on disk, untouched",
        os.path.exists(os.path.join(ROOT, "schedule_candidate_v1", SUPERSEDED)))
    chk("19.6 the certified candidate is unmodified by this promotion",
        sha256(CAND) == "5416ffcb4c07b8e741f24f51b9603ac44c064db943e144618d6ffa372ef62a84")

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"v0.8.7 SHA-256 (frozen): {h87}")
    print(f"v0.8.8 SHA-256:          {h88}")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

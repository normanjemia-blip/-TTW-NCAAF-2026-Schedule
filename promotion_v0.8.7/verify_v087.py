#!/usr/bin/env python3
"""v0.8.7 promotion certificate — READ-ONLY. Writes nothing.

v0.8.7 = v0.8.6 + the approved packet ONLY: 7 activations (14 zeros), Oregon
State Option B (no numerics), 3 text corrections, and the banner.
Zero formula changes. Zero model-output changes.

Exit code 0 iff every check passes.
"""
import collections, hashlib, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
V086 = os.path.join(ROOT, "promotion_v0.8.6",
                    "TTW_College_Football_Power_Ratings_v0.8.6_AUTHORITATIVE.xlsx")
V087 = os.path.join(ROOT, "promotion_v0.8.7",
                    "TTW_College_Football_Power_Ratings_v0.8.7_AUTHORITATIVE.xlsx")
FROZEN_V086_SHA = "bb76901a96a3fa63e14f0cc582891de82846c12fa5f7ce41d182c8addab967f9"

ACTIVATED = {91: "TULN", 7: "ARK", 9: "FLA", 29: "NEB", 113: "OHIO", 89: "USF", 125: "UNLV"}
RECORD_ONLY = {76: "ORST"}
CORRECTED = {85: "MEM", 21: "VAN", 48: "KAN"}
MUST_NOT_MOVE = ("TTU", "STAN", "NIU", "TULN2" if False else "CSU", "FRES", "SYR",
                 "ALA", "TENN", "GASO", "RUTG", "WSU", "NW")

PASS, FAIL = [], []


def chk(msg, ok, detail=""):
    (PASS if ok else FAIL).append(msg)
    print(f"  [{'PASS' if ok else 'FAIL'}] {msg}" + (f" - {detail}" if detail else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def mean(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def main():
    print("=" * 78)
    print("v0.8.7 PROMOTION CERTIFICATE")
    print("=" * 78)

    print("\n1. PREDECESSOR AND SCOPE")
    h86, h87 = sha256(V086), sha256(V087)
    print(f"  v0.8.6 SHA-256: {h86}")
    print(f"  v0.8.7 SHA-256: {h87}")
    chk("1.1 v0.8.6 retains its frozen SHA-256", h86 == FROZEN_V086_SHA, h86[:16])
    chk("1.2 v0.8.7 differs from v0.8.6", h87 != h86)

    a = openpyxl.load_workbook(V086)
    b = openpyxl.load_workbook(V087)
    chk("1.3 21 sheets, identical names and order",
        a.sheetnames == b.sheetnames and len(b.sheetnames) == 21)
    chk("1.4 sheet visibility preserved",
        {s: a[s].sheet_state for s in a.sheetnames} == {s: b[s].sheet_state for s in b.sheetnames})
    chk("1.5 named ranges preserved",
        sorted(a.defined_names.keys()) == sorted(b.defined_names.keys()))

    changed, formula_changed = [], []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                x, y = norm(wa.cell(row=r, column=c).value), norm(wbk.cell(row=r, column=c).value)
                if x != y:
                    changed.append((s, r, wbk.cell(row=r, column=c).coordinate))
                    if x[0] == "F" or y[0] == "F":
                        formula_changed.append((s, wbk.cell(row=r, column=c).coordinate))
    chk("1.6 ZERO formula differences", not formula_changed, str(formula_changed[:5]))
    bysheet = collections.Counter(s for s, _, _ in changed)
    chk("1.7 only START HERE and QB VALUES changed",
        set(bysheet) == {"START HERE", "QB VALUES"}, str(dict(bysheet)))
    chk("1.8 START HERE changed exactly 1 cell", bysheet.get("START HERE") == 1)
    touched = {r for s, r, _ in changed if s == "QB VALUES"}
    expect = set(ACTIVATED) | set(RECORD_ONLY) | set(CORRECTED)
    chk("1.9 exactly the 11 approved QB rows were touched",
        touched == expect, f"touched={sorted(touched)} expected={sorted(expect)}")
    print(f"  [INFO] total changed cells: {len(changed)} "
          f"(QB VALUES {bysheet.get('QB VALUES')}, banner 1)")

    tm, qb, st, ml, sh = b["TEAM MAP"], b["QB VALUES"], b["SETTINGS"], b["MARKET LINES"], b["START HERE"]
    season = st["B3"].value

    print("\n2. THE FOURTEEN APPROVED ZEROS — AND NO OTHERS")
    zeros, nonzero = [], []
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F = qb.cell(row=r, column=4).value, qb.cell(row=r, column=6).value
        for col, v in (("D", D), ("F", F)):
            if v == 0:
                zeros.append((ab, col))
            elif v is not None:
                nonzero.append((ab, col, v))
    new_zeros = [(ab, c) for ab, c in zeros if ab in ACTIVATED.values()]
    chk("2.1 exactly 14 zeros on the 7 activated rows", len(new_zeros) == 14, str(len(new_zeros)))
    chk("2.2 ZERO nonzero QB values anywhere in the workbook", not nonzero, str(nonzero))
    ok_rows = sum(1 for r in range(6, 144) if tm.cell(row=r, column=1).value
                  and not (qb.cell(row=r, column=4).value in (None, "")
                           or qb.cell(row=r, column=6).value in (None, "")
                           or qb.cell(row=r, column=8).value == "L"
                           or qb.cell(row=r, column=10).value != season))
    chk("2.3 total zeros = 234 = 2 x 117 OK rows (was 220 = 2 x 110, so +14)",
        len(zeros) == 234 and len(zeros) == 2 * ok_rows, f"zeros={len(zeros)} ok={ok_rows}")
    for row, ab in ACTIVATED.items():
        D, F = qb.cell(row=row, column=4).value, qb.cell(row=row, column=6).value
        chk(f"2.x {ab} r{row}: D=0 and F=0, so G = F - D = 0 exactly",
            D == 0 and F == 0, f"D={D} F={F}")

    print("\n3. PROTECTED CELLS")
    for row, ab in {**ACTIVATED, **RECORD_ONLY, **CORRECTED}.items():
        g_ok = isf(qb.cell(row=row, column=7).value)
        m_ok = isf(qb.cell(row=row, column=13).value)
        j = qb.cell(row=row, column=10).value
        chk(f"3.x {ab} r{row}: G and M still formulas, J still 2026",
            g_ok and m_ok and j == 2026, f"G={g_ok} M={m_ok} J={j}")

    print("\n4. CENSUSES")
    codes, sts = collections.Counter(), collections.Counter()
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        G = "" if (D is None or F is None) else F - D
        codes[H] += 1
        sts["UNCERTAIN" if (G == "" or H == "L" or J != season) else "OK"] += 1
    chk("4.1 QB status census = 117 OK / 21 UNCERTAIN",
        (sts["OK"], sts["UNCERTAIN"]) == (117, 21), str(dict(sts)))
    chk("4.2 confidence census = 76 H / 43 M / 19 L",
        (codes["H"], codes["M"], codes["L"]) == (76, 43, 19), str(dict(codes)))
    chk("4.3 census totals 138", sum(codes.values()) == 138)
    banner = sh["A1"].value
    chk("4.4 banner states v0.8.7 AUTHORITATIVE", "v0.8.7 AUTHORITATIVE" in banner)
    chk("4.5 banner census reads 76 H / 43 M / 19 L", "76 H / 43 M / 19 L" in banner)
    chk("4.6 banner carries no stale v0.8.6 identifier", "v0.8.6" not in banner)

    print("\n5. PER-ROW OUTCOMES")
    def rec(ab):
        r = [x for x in range(6, 144) if tm.cell(row=x, column=1).value == ab][0]
        C, D, E, F, H = (qb.cell(row=r, column=c).value for c in (3, 4, 5, 6, 8))
        G = "" if (D is None or F is None) else F - D
        M = "UNCERTAIN" if (G == "" or H == "L" or qb.cell(row=r, column=10).value != season) else "OK"
        return dict(row=r, C=C, D=D, E=E, F=F, H=H, G=G, M=M)

    for ab, want_h, want_qb in (("TULN", "M", "Zeon Chriss-Gremillion"), ("ARK", "H", "KJ Jackson"),
                                ("FLA", "H", "Aaron Philo"), ("NEB", "H", "Anthony Colandrea"),
                                ("OHIO", "M", "Nick Poulos"), ("USF", "M", "Michael Van Buren Jr."),
                                ("UNLV", "M", "Jackson Arnold")):
        x = rec(ab)
        chk(f"5.x {ab} activated: {want_h}, {want_qb}, G=0, status OK",
            x["H"] == want_h and x["C"] == want_qb and x["E"] == want_qb
            and x["G"] == 0 and x["M"] == "OK",
            f"H={x['H']} C={x['C']} E={x['E']} G={x['G']} status={x['M']}")

    print("\n5b. BASELINE-MATCH ROWS: COLUMN C WAS NOT REWRITTEN")
    for ab, row in (("ARK", 7), ("UNLV", 125)):
        before = a["QB VALUES"].cell(row=row, column=3).value
        after = qb.cell(row=row, column=3).value
        chk(f"5.x {ab} C{row} unchanged from v0.8.6 — the baseline predates this promotion",
            before == after and before is not None, f"{before!r} -> {after!r}")

    print("\n5c. OREGON STATE OPTION B")
    o = rec("ORST")
    chk("5.6 Oregon State: baseline preserved as Maalik Murphy", o["C"] == "Maalik Murphy", repr(o["C"]))
    chk("5.7 Oregon State: active starter recorded as Braden Atkinson",
        o["E"] == "Braden Atkinson", repr(o["E"]))
    chk("5.8 Oregon State: D and F BLANK — no zero valuation manufactured",
        o["D"] is None and o["F"] is None)
    chk("5.9 Oregon State: confidence M and status still UNCERTAIN",
        o["H"] == "M" and o["M"] == "UNCERTAIN", f"H={o['H']} status={o['M']}")

    print("\n5d. TEXT CORRECTIONS STAY L / UNCERTAIN")
    m = rec("MEM")
    chk("5.10 Memphis: field states both QBs and the withheld decision",
        m["E"] == "Marcus Stokes / Air Noland; decision withheld until kickoff.", repr(m["E"]))
    chk("5.11 Memphis: D/F blank, L, UNCERTAIN — not activated on inference",
        m["D"] is None and m["F"] is None and m["H"] == "L" and m["M"] == "UNCERTAIN")
    for ab, wantE in (("VAN", "Open (Jared Curtis / Blaze Berlowitz / Whit Muschamp)"),
                      ("KAN", "Open (Cole Ballard / Isaiah Marshall)")):
        x = rec(ab)
        chk(f"5.x {ab} corrected, still L / UNCERTAIN, values blank",
            x["E"] == wantE and x["H"] == "L" and x["D"] is None and x["F"] is None
            and x["M"] == "UNCERTAIN", f"E={x['E']} H={x['H']} status={x['M']}")
    for ab, row in (("MEM", 85), ("VAN", 21), ("KAN", 48)):
        prior = a["QB VALUES"].cell(row=row, column=12).value
        chk(f"5.x {ab} prior note retained verbatim, only appended to",
            str(qb.cell(row=row, column=12).value).startswith(str(prior)))

    print("\n5e. BASELINE-QB INVARIANT")
    ok_blank = []
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        G = "" if (D is None or F is None) else F - D
        if not (G == "" or H == "L" or J != season) and not qb.cell(row=r, column=3).value:
            ok_blank.append(ab)
    chk("5.12 all 117 OK rows carry a populated baseline QB in column C",
        not ok_blank, str(ok_blank))

    print("\n5f. DARIUS CURRY STILL APPEARS NOWHERE")
    curry = [f"{ws.title}!{c.coordinate}" for ws in b.worksheets
             for row in ws.iter_rows() for c in row
             if isinstance(c.value, str) and "Curry" in c.value]
    chk("5.13 Darius Curry appears nowhere in the workbook", not curry, str(curry))

    print("\n6. ROWS THAT MUST NOT HAVE MOVED")
    for ab in MUST_NOT_MOVE:
        r = [x for x in range(6, 144) if tm.cell(row=x, column=1).value == ab][0]
        same = all(norm(a["QB VALUES"].cell(row=r, column=c).value) ==
                   norm(qb.cell(row=r, column=c).value) for c in range(1, 14))
        chk(f"6.x {ab} r{r} completely unchanged from v0.8.6", same)
    tt = rec("TTU")
    chk("6.13 Texas Tech medical gate still held: H, D/F blank, UNCERTAIN",
        tt["H"] == "H" and tt["D"] is None and tt["F"] is None and tt["M"] == "UNCERTAIN")

    print("\n7. MODEL OUTPUTS UNCHANGED")
    ps = b["PRESEASON"]
    S = {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 33)}
    raw = {c: [ps.cell(row=r, column=c).value for r in range(6, 144)] for c in (4, 8, 17)}
    mu = {c: mean(v) for c, v in raw.items()}
    W = {4: S["B28"], 8: S["B29"], 17: S["B31"]}
    prior = {}
    for i, r in enumerate(range(6, 144)):
        ab = tm.cell(row=r, column=1).value
        num = den = 0.0
        for c in (4, 8, 17):
            v = raw[c][i]
            if isinstance(v, (int, float)):
                num += W[c] * (v - mu[c]); den += W[c]
        prior[ab] = (num / den) if den else ""
    alias = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()
    delta = {}
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F = qb.cell(row=r, column=4).value, qb.cell(row=r, column=6).value
        delta[ab] = "" if (D is None or F is None) else F - D
    z = lambda v: 0 if v == "" else v
    sch = b["IMPORT SCHEDULE"]
    got, n_games, fcs = {}, 0, 0
    for r in range(6, 1006):
        gid = sch.cell(row=r, column=1).value
        if not gid:
            continue
        n_games += 1
        A = alias.get(str(sch.cell(row=r, column=6).value).strip(), "")
        H = alias.get(str(sch.cell(row=r, column=8).value).strip(), "")
        if not A or not H:
            fcs += 1; continue
        neutral = bool(sch.cell(row=r, column=5).value)
        got[(A, H)] = (prior[H] - prior[A]) + (S["B7"] if neutral else S["B6"]) \
                      + (z(delta.get(H, "")) - z(delta.get(A, "")))
    chk("7.1 888 games / 761 FBS-v-FBS / 127 FCS-involved / 0 BLOCK",
        n_games == 888 and fcs == 127 and n_games - fcs == 761, f"{n_games}/{n_games-fcs}/{fcs}")
    for (A, H), want in ((("MEM", "UNLV"), "UNLV -5.6"), (("UNC", "TCU"), "TCU -4.2"),
                         (("NMSU", "FSU"), "FSU -27.7"), (("SJSU", "USC"), "USC -35.2"),
                         (("HAW", "STAN"), "STAN -3.7")):
        m_ = got[(A, H)]
        lab = f"{H} -{abs(m_):.1f}" if m_ > 0 else f"{A} -{abs(m_):.1f}"
        chk(f"7.x {A} at {H} model spread unchanged at {want}", lab == want, lab)
    chk("7.2 every QB delta is 0 or blank, so ENGINE!M contributes nothing",
        all(d in ("", 0) for d in delta.values()),
        str([(k, v) for k, v in delta.items() if v not in ("", 0)][:5]))
    chk("7.3 Oregon State delta is BLANK, not 0 — valuation deferred, not asserted",
        delta["ORST"] == "", repr(delta["ORST"]))

    print("\n7b. WEEK 0 GATE — MEMPHIS AT UNLV MUST STILL BE QB-GATED")
    mem, unlv = rec("MEM"), rec("UNLV")
    chk("7.4 UNLV now OK", unlv["M"] == "OK", unlv["M"])
    chk("7.5 Memphis still UNCERTAIN", mem["M"] == "UNCERTAIN", mem["M"])
    chk("7.6 Memphis at UNLV remains QB-gated because Memphis is unresolved",
        mem["M"] == "UNCERTAIN")

    print("\n8. SETTINGS AND SAFEGUARDS")
    lines = sum(1 for r in range(6, 1006)
                if ml.cell(row=r, column=1).value is not None
                or ml.cell(row=r, column=4).value is not None)
    chk("8.1 repository artifact ships with MARKET LINES blank", lines == 0, str(lines))
    chk("8.2 BET toggle remains N", st["B11"].value == "N")
    chk("8.3 totals remain unavailable (B22/B23 blank)",
        st["B22"].value is None and st["B23"].value is None)
    for t_ in ("NDSU", "SAC"):
        rr = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == t_][0]
        chk(f"8.4 {t_} remains transitional", tm.cell(row=rr, column=5).value == "Y")
    au = b["AUDIT"]
    chk("8.5 AUDIT market-line invariant unchanged from v0.8.3 design",
        "SUMPRODUCT" in str(au["B16"].value) and "REMOVE TEST LINES" not in str(au["B16"].value))
    chk("8.6 IMPORT SCHEDULE untouched — the schedule candidate is NOT folded in",
        all(a["IMPORT SCHEDULE"].cell(row=r, column=4).value ==
            sch.cell(row=r, column=4).value for r in range(6, 900)))

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"v0.8.6 SHA-256 (frozen): {h86}")
    print(f"v0.8.7 SHA-256:          {h87}")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

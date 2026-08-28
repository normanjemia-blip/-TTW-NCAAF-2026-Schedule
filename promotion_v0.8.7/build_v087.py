#!/usr/bin/env python3
"""Build v0.8.7 from the frozen v0.8.6 workbook.

APPROVED SCOPE ONLY.

  ACTIVATIONS WITH ZEROS (7 rows, 14 zeros) -- every one a BASELINE MATCH,
  i.e. the confirmed starter IS the quarterback the preseason blend assumed:
    Tulane         r91   L -> M   Zeon Chriss-Gremillion  (ESPN/Thamel 2026-08-24)
    Arkansas       r7    L -> H   KJ Jackson              (team announcement 2026-08-23)
    Florida        r9    L -> H   Aaron Philo             (official release + Sumrall 2026-08-24)
    Nebraska       r29   L -> H   Anthony Colandrea       (HC Rhule 2026-08-22)
    Ohio           r113  M -> M   Nick Poulos             (ESPN/Thamel 2026-08-22)
    South Florida  r89   L -> M   Michael Van Buren Jr.   (ESPN/Thamel 2026-08-24)
    UNLV           r125  L -> M   Jackson Arnold          (OWNER-CONFIRMED 2026-08-25)

  RECORD-ONLY, NO ZEROS (1 row) -- Oregon State Option B:
    Oregon State   r76   M (unchanged), stays UNCERTAIN. C preserves Maalik Murphy
    as the preseason baseline; E records Braden Atkinson as active. Atkinson
    DISPLACED the incumbent the preseason priced, so the zero-deviation premise
    fails and no valuation is manufactured.

  TEXT / METADATA CORRECTIONS (3 rows, no numerics, all stay L / UNCERTAIN):
    Memphis r85, Vanderbilt r21, Kansas r48

  BANNER (1 cell) version + confidence census

Target census: 117 OK / 21 UNCERTAIN and 76 H / 43 M / 19 L.

Run:  python3 promotion_v0.8.7/build_v087.py
"""
import datetime, hashlib, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "promotion_v0.8.6",
                   "TTW_College_Football_Power_Ratings_v0.8.6_AUTHORITATIVE.xlsx")
OUT = os.path.join(ROOT, "promotion_v0.8.7",
                   "TTW_College_Football_Power_Ratings_v0.8.7_AUTHORITATIVE.xlsx")
FROZEN_V086_SHA = "bb76901a96a3fa63e14f0cc582891de82846c12fa5f7ce41d182c8addab967f9"
D25 = datetime.datetime(2026, 8, 25)

ZERO = ("Baseline and active values are 0/0 under the deviation-only convention: "
        "QB VALUES!G = F - D, so 0 - 0 = 0 and ENGINE!M contributes exactly nothing "
        "to any game. The zeros do not rate the quarterback - they record that the "
        "active starter IS the quarterback the preseason rating already assumed, so "
        "no deviation applies. No nonzero QB adjustment and no model change.")

# C=None means the cell already holds the correct baseline and MUST NOT be rewritten.
# H=None means the confidence code is already correct and MUST NOT be rewritten.
ACTIVATIONS = {
 "TULN": dict(row=91, C="Zeon Chriss-Gremillion", E="Zeon Chriss-Gremillion", H="M",
   I=("ESPN / Pete Thamel 2026-08-24, carried by On3 (https://www.on3.com/news/"
      "tulane-names-zeon-chriss-gremillion-starting-quarterback-for-2026/)"),
   L=("2026-08-25 ACTIVATED, confidence L -> M. ESPN's Pete Thamel reported 2026-08-24 that Tulane "
      "named Zeon Chriss-Gremillion its starting quarterback for the opener at Duke; he beat out "
      "Kadin Semonza. M rather than H because this is a reporter-sourced naming - no official Tulane "
      "depth chart or team release was located, and the naming PRECEDED the scheduled depth-chart "
      "release. Matches the precedent used for Georgia Southern, North Carolina, Rutgers, Ohio and "
      "South Florida. SUPERSEDES the 2026-08-21 record correction, which recorded a four-way camp "
      "battle. BASELINE MATCH: that entry already recorded that the job 'defaults to "
      "Chriss-Gremillion if no one separates'. INDEPENDENTLY CORROBORATED BY THE WORKBOOK: the "
      "reported Sept. 5 debut at Duke matches IMPORT SCHEDULE exactly (wk1 2026-09-05 Tulane Green "
      "Wave @ Duke Blue Devils). " + ZERO)),
 "ARK": dict(row=7, C=None, E="KJ Jackson", H="H",
   I=("Yahoo Sports / Whole Hog Sports 2026-08-24 (https://sports.yahoo.com/articles/"
      "why-arkansas-football-chose-kj-091055884.html)"),
   L=("2026-08-25 ACTIVATED, confidence L -> H. Arkansas made the decision official on Sunday "
      "2026-08-23 - 'the Hogs made things official on Sunday, Aug. 23, labeling Jackson as the "
      "starter' - with HC Ryan Silverfield on the record: 'Ultimately, it came down to the actual "
      "situational football', citing Jackson's slight edge in two-minute and situational work. "
      "KJ Jackson beat out Memphis transfer AJ Hill after they split first-team reps through 14 fall "
      "camp practices. H because this is a first-party team announcement plus a named head-coach "
      "rationale. Braden Fuller appeared in the spring game and is NOT part of the starting "
      "competition. DEVIATION IS EXACTLY ZERO AND VERIFIED: KJ Jackson was ALREADY the recorded "
      "baseline quarterback in column C before this activation, which is therefore NOT rewritten. "
      "INDEPENDENTLY CORROBORATED BY THE WORKBOOK: the reported North Alabama opener matches IMPORT "
      "SCHEDULE exactly (wk1 2026-09-05 North Alabama Lions @ Arkansas Razorbacks). " + ZERO)),
 "FLA": dict(row=9, C="Aaron Philo", E="Aaron Philo", H="H",
   I=("Official Florida Athletics 2026-08-24 (https://floridagators.com/news/2026/8/24/"
      "football-philo-moves-to-top-of-qb-depth-chart-aug-24-2026); HC Jon Sumrall announcement via "
      "The Independent Florida Alligator (https://www.alligator.org/article/2026/08/"
      "aaron-philo-starting-quarterback)"),
   L=("2026-08-25 ACTIVATED, confidence L -> H. HC Jon Sumrall publicly announced Aaron Philo as the "
      "starting quarterback on Monday 2026-08-24, and the OFFICIAL Florida Athletics site published "
      "'Philo Moves to Top of QB Depth Chart' the same day. Sumrall on the record: 'Philo will get "
      "the lion's share of reps with the ones as we start to try to create cohesion within our "
      "football team to be ready for game one', adding 'I thought he protected the football, made "
      "good decisions Saturday.' Philo, a Georgia Tech transfer reunited with OC Buster Faulkner, "
      "beat out RS freshman Tramell Jones Jr. H because this is an official team release PLUS a "
      "head-coach announcement. RECORDED CAVEAT: Sumrall also said 'I talked to Aaron this is not an "
      "anointing. (This) is not some permanent, you're the starter forever' - a caution about "
      "permanence, not a hedge on who starts Week 1. SUPERSEDES the 2026-08-04 entry. BASELINE "
      "MATCH: that entry already recorded Philo as FAVORED, holding 'the edge entering fall camp'. "
      "INDEPENDENTLY CORROBORATED BY THE WORKBOOK: the reported FAU debut matches IMPORT SCHEDULE "
      "exactly (wk1 2026-09-05 Florida Atlantic Owls @ Florida Gators). " + ZERO)),
 "NEB": dict(row=29, C="Anthony Colandrea", E="Anthony Colandrea", H="H",
   I=("HC Matt Rhule announcement 2026-08-22, reported 2026-08-24 "
      "(https://klin.com/2026/08/24/colandrea-named-starting-qb-for-opener/)"),
   L=("2026-08-25 ACTIVATED, confidence L -> H. HC Matt Rhule announced on Saturday 2026-08-22 that "
      "Anthony Colandrea will be Nebraska's starting quarterback for the season opener against Ohio; "
      "multiple outlets carried it as 'Nebraska officially names starting quarterback'. Rhule said "
      "Daniel Kaelin and TJ Lateef continue to compete for the No. 2 role. H because this is a "
      "head-coach announcement, the same tier as Alabama/DeBoer, Tennessee/Heupel and "
      "Florida/Sumrall. SUPERSEDES the 2026-08-04 entry, which recorded no starter named. BASELINE "
      "MATCH: Colandrea, the 2025 Mountain West Player of the Year, was already the recorded active "
      "candidate. INDEPENDENTLY CORROBORATED BY THE WORKBOOK: the reported Ohio opener matches "
      "IMPORT SCHEDULE exactly (wk1 2026-09-05 Ohio Bobcats @ Nebraska Cornhuskers), which also "
      "cross-checks the Ohio activation on row 113. " + ZERO)),
 "OHIO": dict(row=113, C="Nick Poulos", E="Nick Poulos", H=None,
   I=("ESPN / Pete Thamel 2026-08-22 (https://sports.yahoo.com/articles/"
      "nick-poulos-wins-ohio-starting-004242764.html)"),
   L=("2026-08-25 ACTIVATED, confidence M (UNCHANGED from M). ESPN's Pete Thamel reported that Ohio "
      "named graduate student Nick Poulos its starting quarterback, announced Saturday 2026-08-22 by "
      "first-year HC John Hauser. M rather than H because this is a reporter-sourced naming rather "
      "than a team release carried first-party, matching the precedent used for Georgia Southern, "
      "North Carolina, Rutgers, Tulane and South Florida. SUPERSEDES the 2026-08-03 entry, which "
      "recorded Hauser calling Poulos 'still in the lead' with no Week 1 starter named. BASELINE "
      "MATCH: Poulos was already the coach-described leader, so the confirmed starter IS the "
      "quarterback the preseason rating assumed. INDEPENDENTLY CORROBORATED BY THE WORKBOOK: the "
      "report that Ohio opens week one in Lincoln against Nebraska matches IMPORT SCHEDULE exactly "
      "(wk1 2026-09-05 Ohio Bobcats @ Nebraska Cornhuskers). " + ZERO)),
 "USF": dict(row=89, C="Michael Van Buren Jr.", E="Michael Van Buren Jr.", H="M",
   I=("ESPN / Pete Thamel 2026-08-24 (https://www.saturdaydownsouth.com/news/college-football/"
      "former-sec-qb-michael-van-buren-wins-fbs-starting-job-per-report/)"),
   L=("2026-08-25 ACTIVATED, confidence L -> M. ESPN's Pete Thamel, 2026-08-24: 'Sources: USF has "
      "named Michael Van Buren Jr. the school's starting quarterback.' Van Buren, a junior transfer "
      "from LSU and Mississippi State with 17 career games and 2,896 passing yards, beat out Luke "
      "Kromenhoek, Jayden Bradford and KJ Cooper. M rather than H because this is a reporter-sourced "
      "naming; a reported acknowledgment by USF Athletics could NOT be independently verified and is "
      "NOT relied on. Matches the precedent used for Georgia Southern, North Carolina, Rutgers, "
      "Tulane and Ohio. SUPERSEDES the 2026-08-04 entry, which recorded an open competition. "
      "BASELINE MATCH: that entry already recorded Van Buren as holding 'the edge' in the preseason "
      "projection. NOTE: 'Cooper' in the prior candidate field was KJ Cooper, a genuine competitor - "
      "the prior field was correct. INDEPENDENTLY CORROBORATED BY THE WORKBOOK: the reported FIU "
      "opener matches IMPORT SCHEDULE exactly (wk1 2026-09-05 Florida International Panthers @ "
      "South Florida Bulls). " + ZERO)),
 "UNLV": dict(row=125, C=None, E="Jackson Arnold", H="M",
   I=("OWNER-CONFIRMED 2026-08-25 (operative starting quarterback). Public reporting contains NO "
      "formal QB1 announcement - Las Vegas Review-Journal / Fox5 Vegas: HC Dan Mullen says Jackson "
      "Arnold and Alex Orji will BOTH play and declines to name a starter. Prior source: ESPN "
      "(https://www.espn.com/college-football/story/_/id/47539823/"
      "qb-jackson-arnold-headed-unlv-third-team-three-seasons)"),
   L=("2026-08-25 ACTIVATED, confidence L -> M. PROVENANCE IS OWNER-CONFIRMED, NOT AN OFFICIAL "
      "ANNOUNCEMENT: the owner confirms Jackson Arnold is the operative starting quarterback for "
      "Week 0. HC DAN MULLEN HAS NOT FORMALLY NAMED A STARTER - he says Arnold and Alex Orji will "
      "both play and has declined to name a front-runner throughout training camp. This entry must "
      "not be read as implying a formal QB1 announcement. M rather than H precisely because no team "
      "or coach naming exists. ALEX ORJI IS EXPECTED TO PLAY and remains in the quarterback room; "
      "Orji must also prove that a Grade 3 LCL sprain and severe hamstring tear from the 2025 UCLA "
      "game have healed. SUPERSEDES the 2026-08-04 defect-corrected entry, which recorded a genuine "
      "unresolved battle. DEVIATION IS EXACTLY ZERO AND VERIFIED: Jackson Arnold was ALREADY the "
      "recorded baseline quarterback in column C before this activation, which is therefore NOT "
      "rewritten - the operative starter and the preseason assumption are the same player. UNLV "
      "hosts Memphis in Week 0; MEMPHIS REMAINS UNRESOLVED, so that game stays QB-gated. "
      "RECHECK: official UNLV depth chart or a Mullen naming. " + ZERO)),
}

# Oregon State Option B -- record only, NO numerics, stays UNCERTAIN.
ORST = dict(row=76, C="Maalik Murphy", E="Braden Atkinson",
  I=("ESPN / Pete Thamel via Columbia County Spotlight + Portland Tribune 2026-08-24 "
     "(https://columbiacountyspotlight.com/2026/08/24/oregon-state-names-braden-atkinson-starting-qb/)"),
  L=("2026-08-25 RECORD CORRECTION - STARTER CHANGED, VALUATION DEFERRED (confidence M UNCHANGED, "
     "numerical values REMAIN BLANK, status stays UNCERTAIN). ESPN's Pete Thamel reported 2026-08-24 "
     "that Oregon State named BRADEN ATKINSON its starting quarterback for week one at Houston; "
     "Atkinson, a Mercer transfer and the reigning Jerry Rice Award winner, beat out incumbent "
     "MAALIK MURPHY and Brady Jones. A local report adds that the Oregon State football account "
     "subsequently confirmed it, but that first-party act is secondhand and is NOT relied on -> M, "
     "not H. SUPERSEDES the 2026-08-03 entry, which recorded Murphy as the leader. BASELINE MISMATCH "
     "- THE REASON NO ZEROS ARE WRITTEN: this is the first row in the project where the named starter "
     "is NOT the quarterback the preseason blend assumed. Column C preserves MAALIK MURPHY as the "
     "preseason baseline because the preseason rating priced Oregon State with Murphy, its returning "
     "starter. Writing 0/0 would assert that the Murphy-to-Atkinson change is worth exactly zero, "
     "which is a valuation claim the deviation-only convention was never intended to carry. Values "
     "therefore REMAIN BLANK and the row stays UNCERTAIN pending the QB-value rubric. INDEPENDENTLY "
     "CORROBORATED BY THE WORKBOOK: the reported week-one game at Houston matches IMPORT SCHEDULE "
     "exactly (wk1 2026-09-05 Oregon State Beavers @ Houston Cougars, TDECU Stadium). RECHECK: "
     "official Oregon State depth chart, and the QB-value rubric."))

CORRECTIONS = {
 "MEM": dict(row=85, E="Marcus Stokes / Air Noland; decision withheld until kickoff.",
   I=("Daily Memphian / Yahoo Sports beat reporting, 2026-08-24 (https://sports.yahoo.com/articles/"
      "charles-huff-name-starting-quarterback-012920149.html)"),
   L=(" 2026-08-25 RECORD CORRECTION (data quality only; confidence L UNCHANGED, status remains "
      "UNCERTAIN, numerical values remain BLANK). NO QUALIFYING PUBLIC SOURCE HAS NAMED A STARTER. "
      "The live competition remains MARCUS STOKES versus AIR NOLAND. HC Charles Huff intends to keep "
      "the decision private until kickoff: 'The team will know before you guys do. You guys won't "
      "know until they flip the coin.' UNOFFICIAL DEPTH CHARTS PLACING STOKES FIRST ARE INFERENCE "
      "ONLY and do not satisfy the activation rule; Stokes is NOT activated merely because he is the "
      "current best inference. Memphis plays at UNLV in Week 0 and MEMPHIS BEING UNRESOLVED IS WHY "
      "THAT GAME REMAINS QB-GATED. RECHECK TRIGGER: an official announcement, an official game-day "
      "depth chart, or the first offensive snap.")),
 "VAN": dict(row=21, E="Open (Jared Curtis / Blaze Berlowitz / Whit Muschamp)",
   I=("On3 + SI Vanderbilt + WSMV fall-camp reporting, 2026-08 (https://www.wsmv.com/2026/08/05/"
      "quarterback-competition-continues-dores-open-practice/)"),
   L=(" 2026-08-25 RECORD CORRECTION (data quality only; confidence L UNCHANGED, status remains "
      "UNCERTAIN, numerical values remain BLANK). The candidate field named Jared Curtis alone. The "
      "race is THREE-WAY - Jared Curtis, Blaze Berlowitz and Whit Muschamp - replacing Diego Pavia. "
      "Through 15 fall-camp practices the staff gave Curtis and Berlowitz every chance to separate "
      "and NEITHER DID. Curtis, the No. 1 QB in the 2026 class, is widely described as the "
      "favourite, but favourite is not named, and HC Clark Lea plans to name the starter PRIVATELY "
      "in the team room. RECHECK: Lea naming or a Week 1 depth chart.")),
 "KAN": dict(row=48, E="Open (Cole Ballard / Isaiah Marshall)",
   I=("Yahoo Sports + WIBW fall-camp reporting, 2026-08 (https://www.wibw.com/2026/08/05/"
      "kus-quarterback-race-remains-open/)"),
   L=(" 2026-08-25 RECORD CORRECTION (data quality only; confidence L UNCHANGED, status remains "
      "UNCERTAIN, numerical values remain BLANK). The candidate field described Cole Ballard as the "
      "LEADER; that framing is not supported. HC Lance Leipold: 'We're still in a very, very "
      "competitive quarterback battle right now. Both continue to do a lot of good things.' Ballard "
      "and Isaiah Marshall continue to split first-team reps and Leipold has set NO TIMETABLE, "
      "adding 'there's a very high probability that you'll see both quarterbacks play in games this "
      "year.' Rice transfer Chase Jenkins is in the room but is NOT carried in the active "
      "competition field, since Leipold's own remarks name only Ballard and Marshall. RECHECK: "
      "Leipold naming or a Week 1 depth chart.")),
}

COLS = dict(C=3, D=4, E=5, F=6, H=8, I=9, K=11, L=12)


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def main():
    got = sha256(SRC)
    assert got == FROZEN_V086_SHA, f"v0.8.6 is not the expected artifact: {got}"
    print(f"source v0.8.6 SHA-256 verified: {got}")

    wb = openpyxl.load_workbook(SRC)
    tm, qb, sh = wb["TEAM MAP"], wb["QB VALUES"], wb["START HERE"]

    everything = {**ACTIVATIONS, **CORRECTIONS, "ORST": ORST}
    for ab, spec in everything.items():
        rows = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == ab]
        assert len(rows) == 1 and rows[0] == spec["row"], \
            f"{ab} expected row {spec['row']}, resolved {rows}"
    print("all affected teams located by abbreviation; rows match the approved packet")

    # rows whose column C already holds the correct baseline and must NOT be rewritten
    for ab, want in (("ARK", "KJ Jackson"), ("UNLV", "Jackson Arnold")):
        r = ACTIVATIONS[ab]["row"]
        assert qb.cell(row=r, column=3).value == want, \
            f"{ab} C{r} expected {want!r}, got {qb.cell(row=r, column=3).value!r}"
        assert ACTIVATIONS[ab]["C"] is None, f"{ab} must not rewrite column C"
    print("baseline-match rows pre-verified: ARK C7='KJ Jackson', UNLV C125='Jackson Arnold' "
          "(neither will be rewritten)")

    # the invariant C35/C91-style protection: every OK row must end with a populated C
    def guard(r):
        for name, col in (("A", 1), ("B", 2), ("G", 7), ("M", 13)):
            assert isf(qb.cell(row=r, column=col).value), f"{name}{r} must be a formula"
        assert qb.cell(row=r, column=10).value == 2026, f"J{r} must be 2026"

    zeros = 0
    for ab, s in ACTIVATIONS.items():
        r = s["row"]; guard(r)
        if s["C"] is not None:
            qb.cell(row=r, column=COLS["C"]).value = s["C"]
        qb.cell(row=r, column=COLS["D"]).value = 0; zeros += 1
        qb.cell(row=r, column=COLS["E"]).value = s["E"]
        qb.cell(row=r, column=COLS["F"]).value = 0; zeros += 1
        if s["H"] is not None:
            qb.cell(row=r, column=COLS["H"]).value = s["H"]
        qb.cell(row=r, column=COLS["I"]).value = s["I"]
        qb.cell(row=r, column=COLS["K"]).value = D25
        qb.cell(row=r, column=COLS["L"]).value = s["L"]
    assert zeros == 14, f"expected exactly 14 zeros, wrote {zeros}"
    assert qb.cell(row=7, column=3).value == "KJ Jackson"
    assert qb.cell(row=125, column=3).value == "Jackson Arnold"
    assert qb.cell(row=113, column=8).value == "M", "Ohio H must remain M"
    print(f"activations applied: {len(ACTIVATIONS)} rows, {zeros} zeros (exactly the 14 approved)")

    # ---- Oregon State Option B: NO numerics, stays UNCERTAIN ----
    r = ORST["row"]; guard(r)
    qb.cell(row=r, column=COLS["C"]).value = ORST["C"]
    qb.cell(row=r, column=COLS["E"]).value = ORST["E"]
    qb.cell(row=r, column=COLS["I"]).value = ORST["I"]
    qb.cell(row=r, column=COLS["K"]).value = D25
    qb.cell(row=r, column=COLS["L"]).value = ORST["L"]
    assert qb.cell(row=r, column=COLS["D"]).value is None, "D76 must stay blank"
    assert qb.cell(row=r, column=COLS["F"]).value is None, "F76 must stay blank"
    assert qb.cell(row=r, column=COLS["H"]).value == "M", "H76 must stay M"
    assert qb.cell(row=r, column=COLS["C"]).value == "Maalik Murphy", "C76 must preserve the baseline"
    print("Oregon State Option B applied: baseline preserved, no numerics, stays UNCERTAIN")

    for ab, s in CORRECTIONS.items():
        r = s["row"]; guard(r)
        prior = qb.cell(row=r, column=COLS["L"]).value
        qb.cell(row=r, column=COLS["E"]).value = s["E"]
        qb.cell(row=r, column=COLS["I"]).value = s["I"]
        qb.cell(row=r, column=COLS["K"]).value = D25
        qb.cell(row=r, column=COLS["L"]).value = prior + s["L"]
        assert qb.cell(row=r, column=COLS["L"]).value.startswith(prior), \
            f"{ab} prior note must be retained verbatim"
        assert qb.cell(row=r, column=COLS["D"]).value is None
        assert qb.cell(row=r, column=COLS["F"]).value is None
        assert qb.cell(row=r, column=COLS["H"]).value == "L"
    print(f"corrections applied: {len(CORRECTIONS)} rows, no numerics, all stay L / UNCERTAIN")

    # Darius Curry must still appear nowhere
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "Curry" in c.value:
                    raise AssertionError(f"Darius Curry must not appear: {ws.title}!{c.coordinate}")

    banner = sh["A1"].value
    assert "v0.8.6 AUTHORITATIVE" in banner and "73 H / 40 M / 25 L" in banner, banner[:200]
    sh["A1"].value = (banner.replace("v0.8.6 AUTHORITATIVE", "v0.8.7 AUTHORITATIVE")
                            .replace("73 H / 40 M / 25 L", "76 H / 43 M / 19 L"))
    assert "v0.8.6" not in sh["A1"].value
    print("banner updated: version + confidence census")

    tmp = OUT + ".building.xlsx"
    wb.save(tmp); os.replace(tmp, OUT)
    print(f"written: {OUT}")
    print(f"v0.8.7 SHA-256: {sha256(OUT)}")
    assert sha256(SRC) == FROZEN_V086_SHA, "v0.8.6 was modified by the build"
    print("v0.8.6 confirmed still frozen")


if __name__ == "__main__":
    sys.exit(main())

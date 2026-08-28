#!/usr/bin/env python3
"""Generate the v0.8.7 promotion artifacts EXPLICITLY.

Kept separate from verify_v087.py on purpose: a certificate must be read-only,
so artifact generation never happens as a verification side effect.

Writes:
  promotion_v0.8.7/diff_v086_to_v087.csv
  promotion_v0.8.7/regression_log_v087.txt
"""
import csv, os, subprocess, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HERE = os.path.join(ROOT, "promotion_v0.8.7")
V085 = os.path.join(ROOT, "promotion_v0.8.6",
                    "TTW_College_Football_Power_Ratings_v0.8.6_AUTHORITATIVE.xlsx")
V086 = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.7_AUTHORITATIVE.xlsx")

CLASS = {91:"ACTIVATION",7:"ACTIVATION",9:"ACTIVATION",29:"ACTIVATION",113:"ACTIVATION",
         89:"ACTIVATION",125:"ACTIVATION",76:"RECORD ONLY (no zeros)",
         85:"TEXT CORRECTION",21:"TEXT CORRECTION",48:"TEXT CORRECTION"}
TEAM = {91:"Tulane",7:"Arkansas",9:"Florida",29:"Nebraska",113:"Ohio",89:"South Florida",
        125:"UNLV",76:"Oregon State",85:"Memphis",21:"Vanderbilt",48:"Kansas"}


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def main():
    a = openpyxl.load_workbook(V085)
    b = openpyxl.load_workbook(V086)
    rows = []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                x, y = wa.cell(row=r, column=c), wbk.cell(row=r, column=c)
                if norm(x.value) != norm(y.value):
                    rows.append(dict(
                        sheet=s, cell=y.coordinate, row=r,
                        team=TEAM.get(r, "") if s == "QB VALUES" else "",
                        change_class=CLASS.get(r, "BANNER") if s == "QB VALUES" else "BANNER",
                        old=str(x.value) if x.value is not None else "",
                        new=str(y.value) if y.value is not None else "",
                        is_numeric_zero=(y.value == 0)))
    out = os.path.join(HERE, "diff_v086_to_v087.csv")
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0]))
        w.writeheader()
        w.writerows(rows)
    zeros = sum(1 for r in rows if r["is_numeric_zero"])
    print(f"wrote {out}: {len(rows)} changed cells, {zeros} of them the integer 0")
    assert len(rows) == 69, f"expected 69 changed cells, got {len(rows)}"
    assert zeros == 14, f"expected 14 zeros, got {zeros}"

    log = os.path.join(HERE, "regression_log_v087.txt")
    with open(log, "w", encoding="utf-8") as f:
        for script in ("promotion_v0.8.7/verify_v087.py",
                       "phase11_week0_dryrun/week0_dryrun.py"):
            f.write("=" * 78 + f"\n{script}\n" + "=" * 78 + "\n")
            p = subprocess.run([sys.executable, os.path.join(ROOT, script)],
                               capture_output=True, text=True, cwd=ROOT)
            f.write(p.stdout + p.stderr + f"\n[exit {p.returncode}]\n\n")
    print(f"wrote {log}")


if __name__ == "__main__":
    sys.exit(main())

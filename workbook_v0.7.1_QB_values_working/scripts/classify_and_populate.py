import json, datetime, openpyxl
from openpyxl.worksheet.formula import ArrayFormula

SRC="TTW_NCAAF_Power_Ratings_2026_v0.7.0_QB_WORKING.xlsx"
OUT="TTW_NCAAF_Power_Ratings_2026_v0.7.1_QB_VALUES_WORKING.xlsx"
RD=datetime.date(2026,7,21)
rows=[json.loads(l) for l in open('qb_research.jsonl')]
teams=json.load(open('teams_138.json')); row_of={t['abbrev']:t['row'] for t in teams}

# --- exception categories (primary reason) ---
INJURY={'STAN','SYR','TTU','UNC'}          # unresolved injury/availability
CONFLICT={'TENN','NEB'}                     # materially conflicting sourcing
# all other L = open competition; UNC is the only H/M exception
def classify(r):
    ab=r['abbrev']; c=r['confidence']
    if ab=='UNC':
        return ('EXCEPTION','injury/availability',
                'PCL recovery not complete (expected healthy by camp); starter only "inside track" — not settled')
    if c=='L':
        if ab in INJURY: return ('EXCEPTION','injury/availability', r['confirmed_level'])
        if ab in CONFLICT: return ('EXCEPTION','conflicting/insufficient sourcing', r['confirmed_level'])
        return ('EXCEPTION','open competition', r['confirmed_level'])
    # H/M, not UNC -> initialize
    return ('INIT','initialized-zero','Active QB = projected starter priced into SP+(2026-03-27)+FPI/TR(2026-07-19); no unresolved issue')

disp={}
for r in rows:
    d,cat,reason=classify(r); disp[r['abbrev']]=(d,cat,reason)

init=[a for a,(d,_,_) in disp.items() if d=='INIT']
exc=[a for a,(d,_,_) in disp.items() if d=='EXCEPTION']
print('INIT:',len(init),'| EXCEPTION:',len(exc))

# --- populate v0.7.1 ---
wb=openpyxl.load_workbook(SRC, data_only=False)
qb=wb['QB VALUES']
changed=[]
for r in rows:
    d,cat,reason=disp[r['abbrev']]; rr=row_of[r['abbrev']]
    if d=='INIT':
        qb.cell(row=rr,column=4).value=0    # D Baseline value = 0
        qb.cell(row=rr,column=6).value=0    # F Active value = 0
        changed.append(f"QB VALUES!D{rr}"); changed.append(f"QB VALUES!F{rr}")
    # EXCEPTION: leave D/F blank (already blank from v0.7.0)

# banner
sh=wb['START HERE']
sh['A1'].value=("TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.7.1 QB VALUES WORKING — DEVIATION-ONLY "
                "ZERO-INIT of 105 settled QBs; 33 UNCERTAIN exceptions; NONZERO ADJUSTMENTS PENDING — NOT AUTHORITATIVE)")
changed.append("START HERE!A1")

# CHANGELOG
cl=wb['CHANGELOG']; row=6
while cl.cell(row=row,column=1).value not in (None,""): row+=1
entries=[
 ("v0.7.1","2026-07-21",
  "Phase 8.1 (WORKING FILE ONLY): deviation-only QB zero-initialization (Option A approved). Baseline-identity audit found preseason snapshots load at SP+ = 2026-03-27, FPI = 2026-07-19, TeamRankings = 2026-07-19 (PRESEASON cols E/I/R). All winter-portal transfers and returning starters were priced into all three; Active QB = projected starter for the 105 initialized teams.",
  "Phase 8.1 - baseline-identity audit"),
 ("v0.7.1","2026-07-21",
  "Initialized Baseline value (D)=0 and Active value (F)=0 for 105 settled H/M teams -> QB delta (G)=0, status (M)=OK (verified). This sets the preseason QB as the neutral reference (no double-count of QB quality already in TEAM RATINGS). 33 teams left BLANK and QB UNCERTAIN (exceptions): 27 open competitions, 4 injury/availability (Stanford, Syracuse, Texas Tech, North Carolina), 2 conflicting sourcing (Tennessee, Nebraska). No nonzero value entered anywhere.",
  "Phase 8.1 - zero-init 105, 33 exceptions"),
 ("v0.7.1","2026-07-21",
  "Baseline value & Active value contain only 0 or blank; no nonzero QB adjustment exists. Future nonzero-adjustment rubric proposed (valuation_methodology_report) but NOT approved/applied. Authoritative v0.6.2 XLSX and both native Google Sheets untouched; v0.7.0 research checkpoint preserved. Only QB VALUES D/F (210 cells), banner, and CHANGELOG changed.",
  "Phase 8.1 - scope + future rubric proposed"),
]
for i,(v,dt,desc,note) in enumerate(entries):
    rr=row+i
    cl.cell(row=rr,column=1,value=v); cl.cell(row=rr,column=2,value=dt)
    cl.cell(row=rr,column=3,value=desc); cl.cell(row=rr,column=4,value=note)
    for c in range(1,5): changed.append(f"CHANGELOG!{cl.cell(row=rr,column=c).coordinate}")

wb.save(OUT)

# save audit disposition
audit=[]
for r in rows:
    d,cat,reason=disp[r['abbrev']]
    audit.append({"abbrev":r['abbrev'],"team":r['team'],"conference":r['conference'],
        "confidence":r['confidence'],"active_qb":r['active_qb'],"disposition":d,
        "category":cat,"reason":reason,"injury_eligibility":r.get('injury_eligibility')})
json.dump(audit, open('baseline_audit.json','w'), indent=1, ensure_ascii=False)
json.dump({"changed_cells":changed,"count":len(changed)}, open('changed_cells_v071.json','w'), indent=1)
print("saved", OUT, "| changed cells:", len(changed), "(D/F pairs:", len(init),")")
from collections import Counter
print("exception categories:", Counter(disp[a][1] for a in exc))

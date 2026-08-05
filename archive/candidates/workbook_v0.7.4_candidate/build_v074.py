import shutil, os, datetime, openpyxl
from openpyxl.worksheet.formula import ArrayFormula
ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
SRC=f"{ROOT}/workbook_v0.7.3_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.3_CANDIDATE.xlsx"
OUT=f"{ROOT}/workbook_v0.7.4_candidate"; DST=f"{OUT}/TTW_NCAAF_Power_Ratings_2026_v0.7.4_CANDIDATE.xlsx"
DATE=datetime.datetime(2026,8,3)   # America/New_York owner date
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
os.makedirs(OUT,exist_ok=True); shutil.copyfile(SRC,DST)
wb=openpyxl.load_workbook(DST); qb=wb["QB VALUES"]; tm=wb["TEAM MAP"]
R={tm.cell(row=r,column=1).value:r for r in range(6,144)}
C={'bqb':3,'bval':4,'aqb':5,'aval':6,'conf':8,'upd':11,'notes':12}
def guard(r):
    for k,c in C.items(): assert not isf(qb.cell(row=r,column=c).value), f"formula at row{r} {k}"

# ---------- PART 1: Akron consistency repair (approved) ----------
r=R["AKR"]; guard(r)
assert r==105, f"AKR expected row 105, got {r}"
assert tm.cell(row=r,column=2).value=="Akron"
assert "Poffenbarger" in str(qb.cell(row=r,column=C['aqb']).value) and "Roggow" in str(qb.cell(row=r,column=C['aqb']).value)
assert qb.cell(row=r,column=C['conf']).value=="L"
assert qb.cell(row=r,column=C['bval']).value==0 and qb.cell(row=r,column=C['aval']).value==0
qb.cell(row=r,column=C['bval']).value=None
qb.cell(row=r,column=C['aval']).value=None
qb.cell(row=r,column=C['notes']).value=(
 "CORRECTION 2026-08-03 (7D.1): prior entry listed Ben Finley, who GRADUATED after a two-year Akron "
 "starting run - stale record. Poffenbarger (North Texas transfer) entered spring as projected starter but "
 "Akron has no clear QB1; open fall-camp competition. METHODOLOGY REPAIR 2026-08-03 (7D.2): Baseline/Active "
 "values CLEARED to blank. Zero would assert an affirmative evaluation of both starter and replacement; blank "
 "correctly represents an unresolved numerical assessment. Status remains UNCERTAIN.")
print(f"AKR row{r}: D/F cleared -> {qb.cell(row=r,column=4).value!r}/{qb.cell(row=r,column=6).value!r}, conf={qb.cell(row=r,column=8).value!r}")

# ---------- PART 3: Batch 1 verified updates ----------
# (a) classification change
r=R["OHIO"]; guard(r)
assert qb.cell(row=r,column=C['conf']).value=="L"
qb.cell(row=r,column=C['aqb']).value="Nick Poulos (leader; not yet named)"
qb.cell(row=r,column=C['conf']).value="M"
qb.cell(row=r,column=C['upd']).value=DATE
qb.cell(row=r,column=C['notes']).value=("VERIFIED 2026-08-03: HC John Hauser says Nick Poulos 'is still in the lead' for the "
 "starting job, but Ohio has no Week 1 starter named. Coach-described leader awaiting formal announcement -> M. "
 "Numerical values remain BLANK (no new value populated this phase), so status stays UNCERTAIN.")
# (b) confirmed-accurate: refresh date + truthful note only
CONF={
 "BGSU":"VERIFIED 2026-08-03: Austin Novosad (Oregon transfer, former 4-star, No.13 QB 2023 class) confirmed as BGSU's projected starter; portal QB was a stated program priority. Record accurate; M retained.",
 "MASS":"VERIFIED 2026-08-03: William 'Pop' Watson (Virginia Tech transfer) confirmed arriving to lead the program. Record accurate; M retained.",
 "LIB":"VERIFIED 2026-08-03: genuine three-way transfer competition (Purdie / Jaylen Henderson / Ethan Vasko); Chadwell expected to get quality play but no starter named. Record accurate; L retained.",
 "MTSU":"VERIFIED 2026-08-03: Roman Gagliano confirmed as MTSU's signal-caller, cited as a breakout candidate. Record accurate; M retained.",
 "SHSU":"VERIFIED 2026-08-03: Landyn Locke confirmed as Sam Houston's signal-caller, cited as a breakout candidate. Record accurate; M retained.",
 "WKU":"VERIFIED 2026-08-03: Rodney Tisdale Jr. confirmed as returning starter, BUT Brock Glenn (7 career starts at Florida State) transferred in to challenge. Clear leader with real competition -> M retained (not H).",
}
for ab,note in CONF.items():
    r=R[ab]; guard(r)
    qb.cell(row=r,column=C['upd']).value=DATE
    qb.cell(row=r,column=C['notes']).value=note
    print(f"  confirmed {ab} row{r} conf={qb.cell(row=r,column=8).value} (date+note refreshed)")
# (c) Liberty active-QB refinement to reflect the verified 3-way race
r=R["LIB"]; qb.cell(row=r,column=C['aqb']).value="Open (Purdie / Jaylen Henderson / Ethan Vasko)"

# ---------- CHANGELOG + banner ----------
cl=wb["CHANGELOG"]; last=0
for rr in range(1,cl.max_row+1):
    if any(cl.cell(row=rr,column=c).value not in (None,"") for c in range(1,7)): last=rr
rows=[("v0.7.4","2026-08-03",
  "Phase 7D.2 APPROVED CONSISTENCY REPAIR: Akron (row 105) Baseline/Active numerical values CLEARED from 0 to blank. "
  "Akron has a genuine unresolved QB competition; zero would imply both starter and replacement were affirmatively "
  "evaluated at zero, whereas blank correctly represents an unresolved numerical assessment. Confidence stays L; "
  "QB delta stays blank; status stays UNCERTAIN. No rating or spread effect.","Owner-approved 7D.2 repair"),
 ("v0.7.4","2026-08-03",
  "Phase 7D.2 BATCH 1 (MAC + Conference USA) live verification: 14 teams in scope. 7 records CONFIRMED accurate "
  "(Bowling Green, UMass, Liberty, Middle Tennessee, Sam Houston, Western Kentucky, plus Ohio re-coded). ONE "
  "classification change: Ohio L -> M (HC John Hauser describes Nick Poulos as the leader; no Week 1 starter named). "
  "Liberty active-QB entry refined to the verified three-way transfer race. No new numerical QB value populated.",
  "Batch 1 verification"),
 ("v0.7.4","2026-08-03",
  "Phase 7D.2 SCOPE DISCLOSURE: exact backlog re-derived as 54 (Phase 7D.1 reported 53 - it double-counted Texas Tech, "
  "which was already verified in Phase 7B). Of 14 Batch 1 teams, 7 verified, 1 unresolved source conflict (Miami OH: "
  "candidate lists Thomas Gotkowski while one outlet references a Kansas transfer as projected starter), and 6 not "
  "verified (Buffalo, Sacramento State, FIU, Kennesaw State, Missouri State, New Mexico State). Dates for unverified "
  "teams deliberately NOT refreshed. Backlog remaining: 47. Promotion: DEFER.","Batch 1 scope disclosure")]
for i,(v,d,c,rn) in enumerate(rows):
    rr=last+1+i
    cl.cell(row=rr,column=1).value=v; cl.cell(row=rr,column=2).value=d
    cl.cell(row=rr,column=3).value=c; cl.cell(row=rr,column=4).value=rn
wb["START HERE"]["A1"].value=("TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.7.4 CANDIDATE — Akron numerical cells cleared "
 "(consistency repair); MAC+CUSA Batch 1: 7 verified, Ohio L->M; QB backlog 47 of 54 still unverified; "
 "NOT AUTHORITATIVE, NOT PROMOTED)")
wb.save(DST)
print(f"\nCHANGELOG rows {last+1}-{last+len(rows)}; saved {DST}")

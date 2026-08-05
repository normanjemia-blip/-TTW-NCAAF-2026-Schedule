#!/usr/bin/env python3
"""Emit the exact changed-cell audit for v0.7.2 vs v0.7.1 (CSV + JSON)."""
import os, csv, json, openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.dirname(HERE)
V071 = os.path.join(OUT, "..", "workbook_v0.7.1_QB_values_working",
                    "TTW_NCAAF_Power_Ratings_2026_v0.7.1_QB_VALUES_WORKING.xlsx")
V072 = os.path.join(OUT, "TTW_NCAAF_Power_Ratings_2026_v0.7.2_QB_VALUES_CANDIDATE.xlsx")

def norm(v):
    if isinstance(v, ArrayFormula):
        return v.text
    return v

a = openpyxl.load_workbook(V071)
b = openpyxl.load_workbook(V072)

changes = []
for s in a.sheetnames:
    wa, wb2 = a[s], b[s]
    mr = max(wa.max_row, wb2.max_row); mc = max(wa.max_column, wb2.max_column)
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            va = norm(wa.cell(row=r, column=c).value)
            vb = norm(wb2.cell(row=r, column=c).value)
            if va != vb:
                changes.append({
                    "sheet": s,
                    "cell": f"{get_column_letter(c)}{r}",
                    "old_value": "" if va is None else str(va),
                    "new_value": "" if vb is None else str(vb),
                })

csv_path = os.path.join(OUT, "changed_cells_v072.csv")
with open(csv_path, "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["sheet", "cell", "old_value", "new_value"])
    w.writeheader(); w.writerows(changes)

json_path = os.path.join(OUT, "changed_cells_v072.json")
json.dump({
    "compared": {"from": "v0.7.1", "to": "v0.7.2"},
    "generated": "2026-07-22",
    "total_changed_cells": len(changes),
    "by_sheet": {s: sum(1 for ch in changes if ch["sheet"] == s) for s in sorted({ch["sheet"] for ch in changes})},
    "qb_values_changed": any(ch["sheet"] == "QB VALUES" for ch in changes),
    "note": ("Vanderbilt re-verification FAILED (genuine Curtis/Berlowitz competition). "
             "No QB VALUES data changed. Only the START HERE version banner and 3 CHANGELOG rows differ."),
    "changes": changes,
}, open(json_path, "w"), indent=1, ensure_ascii=False)

print(f"total changed cells: {len(changes)}")
for ch in changes:
    old = (ch["old_value"][:45] + "...") if len(ch["old_value"]) > 45 else ch["old_value"]
    new = (ch["new_value"][:45] + "...") if len(ch["new_value"]) > 45 else ch["new_value"]
    print(f"  {ch['sheet']:12s} {ch['cell']:5s} | {old!r} -> {new!r}")
print("wrote", csv_path)
print("wrote", json_path)

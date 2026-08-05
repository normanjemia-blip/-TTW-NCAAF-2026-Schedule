#!/usr/bin/env python3
"""Phase 8.3: verify the v0.7.2 CANDIDATE against v0.7.1.

Expectations (Vanderbilt re-verification FAILED -> no QB data change):
  - 21 sheets, order + visibility unchanged
  - formula count exactly 123,011 in BOTH files
  - every formula cell has identical coordinate + text (no formula overwritten)
  - ALL 138 QB rows present
  - ONLY START HERE (banner, 1 cell) and CHANGELOG (12 cells) differ
  - QB VALUES is byte-identical (no Vanderbilt or other team change)
  - Baseline/Active (D/F) contain only 0 or blank; 105 zero / 33 blank
  - no nonzero QB delta possible (no nonzero D/F)
"""
import os, openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.dirname(HERE)
V071 = os.path.join(OUT, "..", "workbook_v0.7.1_QB_values_working",
                    "TTW_NCAAF_Power_Ratings_2026_v0.7.1_QB_VALUES_WORKING.xlsx")
V072 = os.path.join(OUT, "TTW_NCAAF_Power_Ratings_2026_v0.7.2_QB_VALUES_CANDIDATE.xlsx")

def norm(v):
    if isinstance(v, ArrayFormula):
        return ("ARR", v.text)
    return v
def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))

a = openpyxl.load_workbook(V071)
b = openpyxl.load_workbook(V072)

fails = []
lines = []
def chk(label, ok, detail=""):
    line = ("PASS " if ok else "FAIL ") + label + (f"  {detail}" if detail else "")
    lines.append(line); print(line)
    if not ok: fails.append(label)

# 1) sheets / order / visibility
chk("21 sheets", len(b.sheetnames) == 21, str(len(b.sheetnames)))
chk("sheet order + visibility unchanged vs v0.7.1",
    [(s, a[s].sheet_state) for s in a.sheetnames] == [(s, b[s].sheet_state) for s in b.sheetnames])
chk("recalc-on-open (fullCalcOnLoad) preserved",
    b.calculation.fullCalcOnLoad == a.calculation.fullCalcOnLoad)

# 2) formula count == 123011 in both
def fcount(wb):
    return sum(1 for s in wb.sheetnames for row in wb[s].iter_rows() for c in row if isf(c.value))
fa, fb = fcount(a), fcount(b)
chk("formula count == 123011 in BOTH files", fa == fb == 123011, f"v0.7.1={fa} v0.7.2={fb}")

# 3) formula coordinates + text identical (no formula overwritten/added/removed)
formula_cell_diffs = []
overwritten = []  # a formula cell whose formula text changed, or a formula removed/added
for s in a.sheetnames:
    wa, wb2 = a[s], b[s]
    mr = max(wa.max_row, wb2.max_row); mc = max(wa.max_column, wb2.max_column)
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            va = wa.cell(row=r, column=c).value
            vb = wb2.cell(row=r, column=c).value
            fa_ = isf(va); fb_ = isf(vb)
            if fa_ or fb_:
                if norm(va) != norm(vb):
                    overwritten.append((s, r, c, norm(va), norm(vb)))
chk("no formula overwritten/added/removed (all formula coords + text identical)",
    not overwritten, str(overwritten[:5]))

# 4) full cell diff -> only START HERE + CHANGELOG
diffs = {}
detail_cells = {}
for s in a.sheetnames:
    wa, wb2 = a[s], b[s]
    mr = max(wa.max_row, wb2.max_row); mc = max(wa.max_column, wb2.max_column)
    n = 0; cells = []
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            if norm(wa.cell(row=r, column=c).value) != norm(wb2.cell(row=r, column=c).value):
                n += 1; cells.append((r, c))
    if n:
        diffs[s] = n; detail_cells[s] = cells
chk("only START HERE + CHANGELOG differ vs v0.7.1",
    set(diffs) == {"START HERE", "CHANGELOG"}, str(dict(diffs)))
chk("START HERE diff == 1 (banner only)", diffs.get("START HERE") == 1, str(diffs.get("START HERE")))
chk("CHANGELOG diff == 12 (3 new rows x 4 cols)", diffs.get("CHANGELOG") == 12, str(diffs.get("CHANGELOG")))
chk("QB VALUES unchanged (no Vanderbilt/other team data change)",
    "QB VALUES" not in diffs, str(diffs.get("QB VALUES")))
# assert the banner cell is exactly START HERE!A1
chk("START HERE changed cell is A1", detail_cells.get("START HERE") == [(1, 1)],
    str(detail_cells.get("START HERE")))
# assert CHANGELOG changed cells are rows 64-66, cols 1-4
exp_cl = [(r, c) for r in (64, 65, 66) for c in (1, 2, 3, 4)]
chk("CHANGELOG changed cells are rows 64-66 cols A-D",
    sorted(detail_cells.get("CHANGELOG", [])) == sorted(exp_cl),
    str(sorted(detail_cells.get("CHANGELOG", []))[:6]))

# 5) QB VALUES D/F content invariants (unchanged from v0.7.1)
qb = b["QB VALUES"]
badval = [(r, col, qb.cell(row=r, column=col).value)
          for r in range(6, 144) for col in (4, 6)
          if qb.cell(row=r, column=col).value not in (None, 0)]
chk("Baseline (D) & Active (F) contain only 0 or blank (no nonzero)", not badval, str(badval[:5]))
D0 = sum(1 for r in range(6, 144) if qb.cell(row=r, column=4).value == 0)
F0 = sum(1 for r in range(6, 144) if qb.cell(row=r, column=6).value == 0)
Dbl = sum(1 for r in range(6, 144) if qb.cell(row=r, column=4).value is None)
chk("exactly 105 teams D=0", D0 == 105, str(D0))
chk("exactly 105 teams F=0", F0 == 105, str(F0))
chk("exactly 33 teams D blank (exceptions)", Dbl == 33, str(Dbl))

# 6) all 138 rows present; A/B/G/M still formulas
chk("138 team rows present (A formulas rows 6-143)",
    all(isf(qb.cell(row=r, column=1).value) for r in range(6, 144)))
badf = [(col, r) for r in range(6, 144) for col in (1, 2, 7, 13)
        if not isf(qb.cell(row=r, column=col).value)]
chk("A/B/G/M formulas intact all 138 rows", not badf, str(badf[:5]))

# 7) Vanderbilt specifically unchanged (row 21): D/F blank, E=Jared Curtis, H=L
tm = b["TEAM MAP"]
van = next(r for r in range(1, tm.max_row + 1) if tm.cell(row=r, column=1).value == "VAN")
chk("Vanderbilt row is 21", van == 21, str(van))
chk("VAN Baseline value (D) blank", qb.cell(row=van, column=4).value in (None, ""))
chk("VAN Active value (F) blank", qb.cell(row=van, column=6).value in (None, ""))
chk("VAN Active QB (E) = Jared Curtis", qb.cell(row=van, column=5).value == "Jared Curtis")
chk("VAN Confidence (H) = L", qb.cell(row=van, column=8).value == "L")

print()
summary = ("*** %d FAILURES ***" % len(fails)) if fails else "ALL v0.7.2 STRUCTURAL CHECKS PASSED"
print(summary)

# write report (structural section); live-calc appended by caller
rep = os.path.join(OUT, "workbook_verification_report_v072.txt")
with open(rep, "w") as f:
    f.write("v0.7.2 CANDIDATE — structural verification vs v0.7.1\n")
    f.write("=" * 55 + "\n")
    f.write("\n".join(lines) + "\n\n" + summary + "\n")
print("wrote", rep)

if fails:
    raise SystemExit(1)

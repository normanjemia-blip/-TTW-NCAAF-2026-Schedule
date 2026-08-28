import json, openpyxl
from openpyxl.worksheet.formula import ArrayFormula
V070="TTW_NCAAF_Power_Ratings_2026_v0.7.0_QB_WORKING.xlsx"
V071="TTW_NCAAF_Power_Ratings_2026_v0.7.1_QB_VALUES_WORKING.xlsx"
V062="v0.6.2_source.xlsx"
def norm(v):
    if isinstance(v,ArrayFormula): return ("ARR",v.text)
    return v
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
a=openpyxl.load_workbook(V070); b=openpyxl.load_workbook(V071); base=openpyxl.load_workbook(V062)
fails=[]
def chk(l,ok,d=""):
    print(("PASS " if ok else "FAIL ")+l+(f"  {d}" if d else ""))
    if not ok: fails.append(l)

chk("21 sheets", len(b.sheetnames)==21)
chk("sheet order+states unchanged vs v0.7.0", [(s,a[s].sheet_state) for s in a.sheetnames]==[(s,b[s].sheet_state) for s in b.sheetnames])
chk("recalc-on-open preserved", b.calculation.fullCalcOnLoad==base.calculation.fullCalcOnLoad)

# formula count vs v0.6.2 (unchanged: only constants written to D/F)
def fcount(wb): return sum(1 for s in wb.sheetnames for r in wb[s].iter_rows() for c in r if isf(c.value))
fb,fbase=fcount(b),fcount(base)
chk("formula count unchanged vs v0.6.2 (=123011)", fb==fbase==123011, f"{fb} vs {fbase}")

# diff vs v0.7.0
diffs={}
for s in a.sheetnames:
    wa,wb2=a[s],b[s]; mr=max(wa.max_row,wb2.max_row); mc=max(wa.max_column,wb2.max_column); n=0
    for r in range(1,mr+1):
        for c in range(1,mc+1):
            if norm(wa.cell(row=r,column=c).value)!=norm(wb2.cell(row=r,column=c).value): n+=1
    if n: diffs[s]=n
print("diff vs v0.7.0:", diffs)
chk("only QB VALUES + START HERE + CHANGELOG changed vs v0.7.0", set(diffs)=={"QB VALUES","START HERE","CHANGELOG"}, str(set(diffs)))
chk("QB VALUES diff == 210 (D/F for 105 teams)", diffs.get("QB VALUES")==210, str(diffs.get("QB VALUES")))
chk("START HERE diff == 1 (banner)", diffs.get("START HERE")==1)

qb=b['QB VALUES']
# D/F contain only 0 or blank; no nonzero
badval=[]
for r in range(6,144):
    for col in (4,6):
        v=qb.cell(row=r,column=col).value
        if v not in (None,0): badval.append((r,col,v))
chk("Baseline value (D) & Active value (F) contain only 0 or blank (no nonzero)", not badval, str(badval[:5]))
D0=sum(1 for r in range(6,144) if qb.cell(row=r,column=4).value==0)
F0=sum(1 for r in range(6,144) if qb.cell(row=r,column=6).value==0)
chk("exactly 105 teams have D=0", D0==105, str(D0))
chk("exactly 105 teams have F=0", F0==105, str(F0))
Dblank=sum(1 for r in range(6,144) if qb.cell(row=r,column=4).value is None)
chk("exactly 33 teams have D blank (exceptions)", Dblank==33, str(Dblank))

# formula columns A/B/G/M still formulas
badf=[(col,r) for r in range(6,144) for col in (1,2,7,13) if not isf(qb.cell(row=r,column=col).value)]
chk("A/B/G/M formulas intact all 138 rows", not badf, str(badf[:5]))
# research fields retained (E,H,I,J,K,L populated all 138; C for 106 H/M)
retain=[(n,r) for r in range(6,144) for col,n in [(5,'E'),(8,'H'),(9,'I'),(10,'J'),(11,'K'),(12,'L')] if qb.cell(row=r,column=col).value in (None,"")]
chk("all 138 retain research fields E/H/I/J/K/L", not retain, str(retain[:5]))

# no duplicate/missing teams: A formulas present rows 6-143
chk("138 team rows present (A formulas)", all(isf(qb.cell(row=r,column=1).value) for r in range(6,144)))

print()
print("*** %d FAILURES ***"%len(fails) if fails else "ALL v0.7.1 STRUCTURAL CHECKS PASSED")
if fails: raise SystemExit(1)

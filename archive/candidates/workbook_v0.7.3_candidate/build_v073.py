#!/usr/bin/env python3
"""Phase 7D.1: build v0.7.3 CANDIDATE from v0.7.2 with approved repairs only."""
import shutil, os, datetime, openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
SRC=f"{ROOT}/workbook_v0.7.2_QB_values_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.2_QB_VALUES_CANDIDATE.xlsx"
OUTDIR=f"{ROOT}/workbook_v0.7.3_candidate"
DST=f"{OUTDIR}/TTW_NCAAF_Power_Ratings_2026_v0.7.3_CANDIDATE.xlsx"
REVIEW_DATE=datetime.datetime(2026,8,3)   # America/New_York owner date

def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))

os.makedirs(OUTDIR, exist_ok=True)
shutil.copyfile(SRC, DST)
wb=openpyxl.load_workbook(DST)
qb=wb["QB VALUES"]; tm=wb["TEAM MAP"]

# ---- locate rows by abbrev (never hardcode blindly) ----
rows={}
for r in range(6,144):
    rows[tm.cell(row=r,column=1).value]=r
assert rows["TTU"]==52, f"TTU expected row 52, got {rows['TTU']}"
AKR=rows["AKR"]
print(f"Row alignment verified: TTU={rows['TTU']} ({tm.cell(row=rows['TTU'],column=2).value}), AKR={AKR} ({tm.cell(row=AKR,column=2).value})")

COL={'bqb':3,'bval':4,'aqb':5,'aval':6,'conf':8,'src':9,'season':10,'upd':11,'notes':12}
before={}
def snap(tag,r):
    before[tag]={k:qb.cell(row=r,column=c).value for k,c in COL.items()}

# ---- guard: refuse to write into any formula cell ----
for r in (rows["TTU"], AKR):
    for k,c in COL.items():
        assert not isf(qb.cell(row=r,column=c).value), f"refuse: {k} at row {r} is a formula"

# ================= REPAIR 1 — Texas Tech (approved) =================
r=rows["TTU"]; snap("TTU",r)
assert qb.cell(row=r,column=COL['aqb']).value=="Will Hammond", "TTU active QB mismatch"
assert qb.cell(row=r,column=COL['conf']).value=="L", "TTU confidence expected L"
qb.cell(row=r,column=COL['conf']).value="M"
qb.cell(row=r,column=COL['upd']).value=REVIEW_DATE
qb.cell(row=r,column=COL['notes']).value=(
 "Surgeon clearance reported (late July 2026); projected Week 1 starter per ESPN/On3. "
 "FINAL TEAM MEDICAL CLEARANCE STILL PENDING - training staff retains the decision "
 "(~Aug 21 nine-month ACL mark). Classification is LIKELY (M), NOT KNOWN (H): no official "
 "naming has occurred. Numerical QB delta remains GATED - Baseline/Active values stay blank "
 "until final clearance, so QB status correctly remains UNCERTAIN.")
# explicitly DO NOT touch D52/F52
assert qb.cell(row=r,column=COL['bval']).value is None and qb.cell(row=r,column=COL['aval']).value is None

# ================= REPAIR 2 — Akron (verified factual defect) =================
r=AKR; snap("AKR",r)
assert qb.cell(row=r,column=COL['aqb']).value=="Ben Finley", "AKR active QB mismatch"
assert qb.cell(row=r,column=COL['conf']).value=="M", "AKR confidence expected M"
qb.cell(row=r,column=COL['aqb']).value="Open (Reese Poffenbarger / Brayden Roggow)"
qb.cell(row=r,column=COL['conf']).value="L"
qb.cell(row=r,column=COL['upd']).value=REVIEW_DATE
qb.cell(row=r,column=COL['notes']).value=(
 "CORRECTION 2026-08-03: prior entry listed Ben Finley, who GRADUATED after a two-year "
 "Akron starting run and is no longer on the roster - a stale record. Reese Poffenbarger "
 "(North Texas transfer, 5th school in 6 seasons) entered spring as projected starter, but "
 "reporting indicates Akron has no clear QB1 for 2026 and an open fall-camp competition. "
 "Reclassified M -> L (genuine open competition). Numerical values left unchanged pending "
 "owner review (see 7D.1 report, open item OI-1).")

wb.save(DST)
print("\n=== CHANGES APPLIED ===")
for tag,r in (("TTU",rows["TTU"]),("AKR",AKR)):
    for k,c in COL.items():
        b=before[tag][k]; a=qb.cell(row=r,column=c).value
        if b!=a: print(f"  {tag} row{r} {k}: {str(b)[:40]!r} -> {str(a)[:60]!r}")
print(f"\nSaved: {DST}")

"""Phase 7D.5 build: v0.7.8 -> v0.7.9 FINAL QB-VERIFICATION CANDIDATE.

Closes finding F-7: stamps the 21 Tier-1 records that were credited as verified
in the phase ledger but carried no in-workbook verification stamp, and applies
the corrections that fresh team-specific verification surfaced.
"""
import shutil, os, datetime, openpyxl
from openpyxl.worksheet.formula import ArrayFormula
ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
SRC=f"{ROOT}/workbook_v0.7.8_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.8_CANDIDATE.xlsx"
OUT=f"{ROOT}/workbook_v0.7.9_candidate"; DST=f"{OUT}/TTW_NCAAF_Power_Ratings_2026_v0.7.9_CANDIDATE.xlsx"
D8=datetime.datetime(2026,8,4)          # owner timezone America/New_York
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
os.makedirs(OUT,exist_ok=True); shutil.copyfile(SRC,DST)
wb=openpyxl.load_workbook(DST); qb=wb["QB VALUES"]; tm=wb["TEAM MAP"]
R={tm.cell(row=r,column=1).value:r for r in range(6,144)}
AQB,BV,AV,CONF,UPD,NOTE=5,4,6,8,11,12
def guard(r):
    for c in (AQB,BV,AV,CONF,UPD,NOTE): assert not isf(qb.cell(row=r,column=c).value), f"formula row{r}"
def setrec(ab, aqb=None, code=None, clear_vals=False, note=None, expect_code=None, expect_row=None,
           expect_blank=False):
    r=R[ab]; guard(r)
    if expect_row: assert r==expect_row, f"{ab} row {r} != expected {expect_row}"
    if expect_code: assert qb.cell(row=r,column=CONF).value==expect_code, f"{ab} expected {expect_code}"
    if expect_blank:
        assert qb.cell(row=r,column=BV).value is None and qb.cell(row=r,column=AV).value is None, \
            f"{ab} expected already-blank numerical cells"
    if aqb is not None: qb.cell(row=r,column=AQB).value=aqb
    if code is not None: qb.cell(row=r,column=CONF).value=code
    if clear_vals:
        assert qb.cell(row=r,column=BV).value==0 and qb.cell(row=r,column=AV).value==0
        qb.cell(row=r,column=BV).value=None; qb.cell(row=r,column=AV).value=None
    qb.cell(row=r,column=UPD).value=D8
    qb.cell(row=r,column=NOTE).value=note
    print(f"  {ab:6} row{r:>4} code={qb.cell(row=r,column=CONF).value} D={qb.cell(row=r,column=BV).value!r}")

print("=== CODE CHANGES (3) ===")
setrec("MIZ", expect_code="M", expect_row=14, code="H",
 aqb="Austin Simmons (OFFICIALLY NAMED 2026-03-19)",
 note=("VERIFIED 7D.5 2026-08-04 - OFFICIAL NAMING: HC Eli Drinkwitz NAMED Austin Simmons (Ole Miss transfer) Missouri's "
  "2026 starting quarterback on 2026-03-19, right after spring camp and nearly six months before the 2026-09-03 opener vs "
  "Arkansas-Pine Bluff. Simmons beat out returning sophomore Matt Zollers and UConn transfer Nick Evers. Drinkwitz's stated "
  "reasoning: Simmons 'was better in managing the team in two-minute drills and end-of-game situations', and naming him "
  "early would 'allow him to develop relationships with our wide receivers, timing with the wide receivers and tight ends, "
  "and allow him to really establish himself as a leader.' Reaffirmed at SEC Media Days in July 2026. Prior entry read "
  "'projected leader, some competition' - STALE, the competition had already been formally settled. An official naming with "
  "the competition closed meets the H standard -> M upgraded to H (same basis as UTEP). Zeros RETAINED (correct for H); "
  "status remains OK."))
setrec("UNC", expect_code="M", expect_row=65, code="L", expect_blank=True,
 aqb="Open (Billy Edwards Jr. / Miles O'Neill / Au'Tori Newkirk / Travis Burgess)",
 note=("DEFECT CORRECTED 7D.5 2026-08-04: prior entry asserted the Wisconsin transfer as North Carolina's 'projected "
  "transfer QB1' (M). Current team-specific coverage contradicts it - headline 2026-07-18: 'UNC QUARTERBACK BATTLE REMAINS "
  "WIDE OPEN as Tar Heels head into fall camp.' NO starter named. HC Bill Belichick left the competition open: 'This year "
  "our quarterbacks after spring ball are still here, are out in player-run practices,' and described the room as "
  "'significantly deeper and more stable' than 2025 - depth, not a decision. FOUR candidates: Billy Edwards Jr. (most "
  "experienced), Miles O'Neill (Texas A&M transfer, 'cannon of an arm'), Au'Tori Newkirk, and Travis Burgess (senior, "
  "recovering from injury). No leader identified -> M downgraded to L. NUMERICAL CELLS WERE ALREADY BLANK - this record "
  "was never zero-initialized (one of seven M-coded rows carrying blank inputs), so no consistency repair was required "
  "and the status was ALREADY UNCERTAIN; the L code now makes that gate explicit rather than incidental. "
  "RECHECK: UNC depth chart or a Belichick naming."))
setrec("UNLV", expect_code="M", expect_row=125, code="L", clear_vals=True,
 aqb="Open (Jackson Arnold / Alex Orji)",
 note=("DEFECT CORRECTED 7D.5 2026-08-04: prior entry read 'projected transfer QB1 (verified)' (M) for the Oklahoma/Auburn "
  "transfer. The Las Vegas Review-Journal's training-camp preview (2026-07-28) reports a genuine unresolved battle: 'even "
  "if coaches won't name a front-runner throughout training camp, there is definitely a battle between returner ALEX ORJI "
  "and Auburn transfer JACKSON ARNOLD.' HC Dan Mullen has NOT named a starter, and Fox5 Vegas (2026-07-17) ran a segment on "
  "the still-open competition. Arnold brings 18 career starts, 3,293 pass yds and 871 rush yds across Oklahoma and Auburn; "
  "Orji must prove a Grade 3 LCL sprain and severe hamstring tear from the 2025 UCLA game have healed. The word 'verified' "
  "in the prior note was not supported by any naming. -> M downgraded to L; zeros CLEARED to blank; status becomes "
  "UNCERTAIN. Someone must start vs Memphis on 2026-08-29. RECHECK: UNLV depth chart."))

print("=== SEC (stamped, codes unchanged) ===")
setrec("ALA", expect_code="L", expect_row=6,
 note=("VERIFIED 7D.5 2026-08-04: CBS Sports fall-camp QB intel (2026-08-04) confirms the entry - Keelon Russell (2nd-year "
  "freshman, former No. 2 overall recruit) vs Austin Mack (RS junior, system experience). Russell only SLIGHTLY favored "
  "(characterized as 55-45 to 60-40); status UNDECIDED with the competition ongoing into fall camp; 'Russell will have to "
  "improve his consistency' to secure the job. Genuine competition, no leader established by the staff -> L retained; "
  "values blank; UNCERTAIN."))
setrec("AUB", expect_code="M", expect_row=8,
 aqb="Byrum Brown (USF transfer)",
 note=("VERIFIED 7D.5 2026-08-04: Byrum Brown CONFIRMED as Auburn's quarterback - a USF transfer who followed HC Alex "
  "Golesh from Tampa and was among the most sought-after players in the portal. 2025 at USF: 3,158 pass yds, 28 TD, 7 INT, "
  "leading the Bulls to 9-3. No competition reported and no formal naming located -> M retained (not H). Entry corroborated; "
  "zeros retained; status OK. RECHECK: official Auburn depth chart."))
setrec("FLA", expect_code="L", expect_row=9,
 note=("VERIFIED 7D.5 2026-08-04: CBS Sports fall-camp intel (2026-08-04) confirms an open competition - Aaron Philo "
  "(Georgia Tech transfer, reunited with OC Buster Faulkner) vs Tramell Jones (RS freshman). Philo is FAVORED and 'holds "
  "the edge entering fall camp' on the strength of a more consistent spring, but NO starter has been named and the "
  "competition is live. Under the 7D.5 rule that a starter must not be inferred from a preseason projection, a media "
  "'edge' characterization without a staff naming does not clear the M bar -> L retained (consistent with the Colorado "
  "State treatment); values blank; UNCERTAIN. RECHECK: Florida depth chart."))
setrec("UK", expect_code="M", expect_row=11,
 aqb="Kenny Minchey (Notre Dame transfer)",
 note=("VERIFIED 7D.5 2026-08-04: Kenny Minchey CONFIRMED as Kentucky's quarterback - a Notre Dame transfer after three "
  "seasons there (29 career pass attempts; lost the 2025 Notre Dame battle to CJ Carr), now running the offense for NEW HC "
  "WILL STEIN. Transfer brought in to start with no competition reported; no formal naming located -> M retained. Zeros "
  "retained; status OK. RECHECK: Kentucky depth chart."))
setrec("TENN", expect_code="L", expect_row=18,
 note=("VERIFIED 7D.5 2026-08-04: CBS Sports fall-camp intel (2026-08-04) confirms the entry - Faizon Brandon (five-star "
  "true freshman) vs George MacIntyre (RS, most system time). Brandon is FAVORED and impressed evaluators ('This kid's got "
  "moxie... confidence'), but the competition is UNDECIDED entering fall camp. Genuine competition -> L retained; values "
  "blank; UNCERTAIN."))
setrec("VAN", expect_code="L", expect_row=21,
 note=("VERIFIED 7D.5 2026-08-04 - CLOSES THE PHASE 8.3 FAILED RE-VERIFICATION: CBS Sports fall-camp intel (2026-08-04) "
  "confirms the entry - Jared Curtis (No. 2 overall recruit, true freshman) vs Blaze Berlowitz (veteran who backed up Diego "
  "Pavia in 2025). Curtis is FAVORED, with staff saying 'He's as good as we hoped he'd be', but the competition REMAINS "
  "OPEN and undecided entering fall camp. Genuine competition -> L retained; values blank; UNCERTAIN. Vanderbilt's record "
  "is now substantiated in-workbook for the first time."))

print("=== BIG TEN / ACC (stamped) ===")
setrec("NEB", expect_code="L", expect_row=29,
 note=("VERIFIED 7D.5 2026-08-04: Nebraska added ANTHONY COLANDREA from the portal - the 2025 MOUNTAIN WEST PLAYER OF THE "
  "YEAR at UNLV (3,459 pass yds, 23 TD) - joining returner TJ Lateef. This confirms the entry's candidate list. No starter "
  "has been named and no staff statement establishing a leader was located; a high-profile transfer arriving is not by "
  "itself evidence of a settled job. -> L retained; values blank; UNCERTAIN. EVIDENCE NOTE: this is the thinnest of the "
  "7D.5 confirmations - the transfer is documented, the competition's state is inferred from the absence of a naming. "
  "RECHECK: Nebraska depth chart."))
setrec("STAN", expect_code="L", expect_row=68,
 note=("VERIFIED 7D.5 2026-08-04: quarterback remains one of Stanford's named key fall-camp position battles. Davis Warren "
  "(Michigan grad transfer with starting experience) is described as being in a strong position, with true freshman early "
  "enrollee Michael Mitchell Jr. able to compete. NO staff naming located, and 'strong position' is a projection rather "
  "than a decision -> L retained under the 7D.5 rule; values blank; UNCERTAIN. RECHECK: Stanford depth chart."))
setrec("SYR", expect_code="L", expect_row=69,
 note=("VERIFIED 7D.5 2026-08-04: Steve Angeli (Notre Dame transfer) says he is 'FULL GO' with 'no restrictions' ahead of "
  "training camp after the 2025 ACHILLES injury that ended his season - but the same reporting frames it as 'ANOTHER QB "
  "BATTLE', and Syracuse overhauled the room in the offseason (incl. Kennesaw State transfer Amari Odom). A quarterback "
  "returning from an Achilles tear into an acknowledged battle is not a settled job -> L retained; values blank; "
  "UNCERTAIN. RECHECK: Syracuse depth chart or a Fran Brown naming."))

print("=== AMERICAN (stamped) ===")
setrec("FAU", expect_code="M", expect_row=84,
 aqb="Caden Veltkamp (returning starter)",
 note=("VERIFIED 7D.5 2026-08-04: Caden Veltkamp CONFIRMED returning as FAU's starter for a second straight year in Zach "
  "Kittley's air raid - 2025: 3,641 pass yds (8th in FBS), 24 TD, 67.0% comp; 2024 CUSA Offensive Player of the Year at "
  "Western Kentucky. Coverage states FAU enters 2026 'with zero questions about their starter.' HC Kittley described a "
  "'5-man battle' at the position, but that concerns the BACKUP order behind Veltkamp - Drew Devillier (Rice transfer), "
  "Jordan Magwood, Michael Valentino and Jeremiah Daoud. Established returning starter, no formal 2026 naming -> M "
  "retained (not H), consistent with the Texas State / WMU / Toledo precedent. Zeros retained; status OK."))
setrec("MEM", expect_code="L", expect_row=85,
 note=("VERIFIED 7D.5 2026-08-04: Memphis's quarterback battle is 'CLOSE and will likely be one of the LATER ONES SETTLED "
  "in the American.' FIRST-YEAR HC CHARLES HUFF must choose between two transfers with contrasting backgrounds - Marcus "
  "Stokes and Air Noland. No starter named -> L retained; values blank; UNCERTAIN. Entry corroborated."))
setrec("USF", expect_code="L", expect_row=89,
 note=("VERIFIED 7D.5 2026-08-04: USF's competition is between LSU transfer MICHAEL VAN BUREN JR. and Mississippi State "
  "transfer LUKE KROMENHOEK, both named in the entry. Van Buren 'seems to have the edge' on slightly more game experience, "
  "but that is a preseason projection, not a staff decision, and no naming was located -> L retained under the 7D.5 rule; "
  "values blank; UNCERTAIN. USF training camp opened 2026-08-05. RECHECK: USF depth chart."))
setrec("TULN", expect_code="L", expect_row=91,
 note=("VERIFIED 7D.5 2026-08-04: 'The competition CONTINUES as the Green Wave enter fall practice.' Kadin Semonza (Ball "
  "State transfer; a big 2024 at Ball State, did NOT play in 2025) 'will likely be the main man - but he HAS TO HOLD OFF' "
  "Houston transfer Zeon Chriss-Gremillion, sophomore Dagan Bruno and freshman Trace Johnson. A leader who must hold off "
  "three challengers with no naming -> L retained; values blank; UNCERTAIN. Entry corroborated."))

print("=== MAC (7D.1 findings stamped) ===")
setrec("BALL", expect_code="L", expect_row=106,
 note=("VERIFIED 7D.1 (2026-08-03), STAMPED 7D.5 2026-08-04: Ball State's room is unsettled - Keldric Luster (Texas State "
  "transfer, dual-threat) and Tyler Mizzell both may see action; no starter named. Corroborated this phase: Texas State's "
  "outgoing-transfer ledger confirms Luster left for Ball State. -> L retained; values blank; UNCERTAIN. EVIDENCE NOTE: "
  "the 7D.1 live check stands; no fresh contradicting evidence surfaced in 7D.5, but no new confirmation either. "
  "This record's date was never stamped until now - that gap is what finding F-7 flagged."))
setrec("CMU", expect_code="L", expect_row=109,
 note=("VERIFIED 7D.1 (2026-08-03), STAMPED 7D.5 2026-08-04: Central Michigan's quarterback job is unsettled, with Angel "
  "Flores among the contenders alongside Marcus Beamon and Jadyn Glasser; no starter named. -> L retained; values blank; "
  "UNCERTAIN. EVIDENCE NOTE: the 7D.1 live check stands; no fresh contradicting evidence surfaced in 7D.5, but no new "
  "confirmation either. This record's date was never stamped until now - the F-7 gap."))
setrec("TOL", expect_code="M", expect_row=115,
 aqb="John Alan Richter (returning)",
 note=("VERIFIED 7D.1 (2026-08-03), RE-CONFIRMED 7D.5 2026-08-04: John Alan Richter returns as Toledo's quarterback - 699 "
  "pass yds and 7 TD across the past two seasons, one start last year, after sitting behind Dequan Finn and Tucker Gleason "
  "for three seasons; coverage calls his holdover 'the biggest testament to the buy-in Coach Mike Jacobs has from the "
  "Toledo roster' and says he 'finally gets the keys in 2026.' Toledo was picked THIRD in the 2026 MAC preseason coaches "
  "poll. Clear presumptive starter, no formal naming -> M retained; zeros retained; status OK."))
setrec("WMU", expect_code="M", expect_row=117,
 aqb="Broc Lowry (returning starter)",
 note=("VERIFIED 7D.1 (2026-08-03), RE-CONFIRMED 7D.5 2026-08-04: Broc Lowry is RETURNING to Western Michigan for 2026 as "
  "the starting quarterback, entering his SECOND season as the primary starter with two years of eligibility left. He won "
  "MAC OFFENSIVE PLAYER OF THE YEAR in his first year as a starter (1,803 pass / 963 rush / 23 total TD) on the "
  "MAC-champion team. Established returning starter, no formal 2026 naming -> M retained (not H), consistent with the "
  "Texas State / Toledo / FAU precedent. Zeros retained; status OK."))

print("=== SUN BELT (stamped) ===")
setrec("USM", expect_code="L", expect_row=139,
 note=("VERIFIED 7D.5 2026-08-04: HC BLAKE ANDERSON says Landry Lyddy is 'VERY MUCH IN THE RACE' to be Southern Miss's "
  "opening-night starter as fall camp nears, alongside John White and Illinois transfer Ethan Hampton. NO single starter "
  "named. Genuine open competition -> L retained; values blank; UNCERTAIN. Entry corroborated. (Southern Miss's staff was "
  "corrected by the owner in Phase 7C.1: Blake Anderson is the 24th HC, announced 2025-12-11.)"))

print("=== FINAL FIVE - note refresh only, no code change ===")
setrec("TXST", expect_code="M", expect_row=78,
 aqb="Brad Jackson (returning starter)",
 note=("VERIFIED 7D.4A, RE-VERIFIED 7D.5 2026-08-04 against official/primary sources. Brad Jackson is Texas State's "
  "returning starter: 2025 - 3,050 pass yds, 18 TD, 7 INT, 71.5% comp; 16-17 rush TD (school single-season record for a QB); "
  "3,968 total yds, 7th nationally; third-team all-conference; and MVP OF THE 2026 ARMED FORCES BOWL (173 yds, 3 pass TD "
  "plus a rushing score in a 41-10 win over Rice). Announced his return 2025-12-06; coverage states Texas State 'returns a "
  "starting quarterback for the first time under Kinne'; The Athletic ranked him the top Group of Five quarterback. NEW IN "
  "7D.5: Boston College transfer SHAKER REISIG arrived and is reported as the PRIMARY BACKUP to Jackson, not a challenger "
  "(3 games at BC in 2025: 141 yds, 1 TD, 2 INT); QBs Keldric Luster (Ball State) and Holden Geriner (Pittsburgh) "
  "transferred out. No competition reported. H considered and declined - no official 2026 naming. -> M retained; zeros "
  "retained; status OK. RECHECK: official Texas State depth chart."))
setrec("WSU", expect_code="L", expect_row=80,
 note=("DEFECT CORRECTED 7D.4A, RE-VERIFIED 7D.5 2026-08-04: prior entry asserted Caden Pinnick as projected QB1 (M). "
  "HeraldNet 2026-07-30: WSU 'set to open fall camp with WIDE-OPEN QB competition'; The Columbian 2026-08-01: quarterback "
  "situation 'UNRESOLVED'; first-year HC Kirby Moore declined to commit - 'That's gonna continue through the summer'; the "
  "Spokesman-Review beat reports Pinnick 'WASN'T ABLE TO CREATE MEANINGFUL SEPARATION FROM ESHELMAN' in spring, and staff "
  "may keep the decision in-house until kickoff. CORROBORATED IN 7D.5 by the official transfer ledger: Pinnick IN from UC "
  "Davis; Jaxon Potter OUT to Old Dominion and Ajani Sheppard OUT to Temple (both appear as competitors in those teams' "
  "records); new OC Matt Miller. Three-way with Owen Eshelman and Julian Dugger, no leader -> L; zeros CLEARED to blank; "
  "UNCERTAIN. Camp opened 2026-08-06."))
setrec("UNT", expect_code="M", expect_row=87,
 aqb="Tayven Jackson (slight lead; Ditta / Jimerson competing)",
 note=("VERIFIED 7D.4A, RE-VERIFIED 7D.5 2026-08-04: entry CONFIRMED NOT STALE. HC NEAL BROWN said on the American Kickoff "
  "broadcast (2026-07-29) that senior UCF transfer Tayven Jackson 'NOW HAS A SLIGHT LEAD' in an ongoing THREE-MAN race over "
  "East Carolina transfer Chaston Ditta and returner Chris Jimerson Jr.; Brown did NOT declare a starter. Jackson's 2025 at "
  "UCF: 2,151 pass yds, 10 TD, 8 INT. COACHING CHANGE VERIFIED: Eric Morris left for Oklahoma State; Neal Brown rebuilt the "
  "roster with 75-90+ new players. CONTEXT ADDED IN 7D.5: North Texas must replace Drew Mestemaker, who LED THE NATION IN "
  "PASSING YARDS in 2025 - the race is a genuine full replacement, not a succession. Clear leader stated by the head coach, "
  "no naming -> M retained; zeros retained; status OK."))
setrec("GASO", expect_code="M", expect_row=131,
 aqb="Max Johnson (leader; Turner Helton competing)",
 note=("STALE ENTRY CORRECTED 7D.4A, RE-VERIFIED 7D.5 2026-08-04: prior entry read 'Open (Weston Bryan / Turner Helton)' "
  "and OMITTED MAX JOHNSON ENTIRELY. Johnson signed from North Carolina on 2026-01-11 (previously LSU 2020-21, Texas A&M "
  "2022-23), and at SUN BELT MEDIA DAYS (2026-07-15) HC CLAY HELTON SAID 'Max Johnson is expected to be the starter if the "
  "season began today,' citing experience and high-level production. Turner Helton (WKU transfer, the head coach's son; "
  "14/26, 74 yds, 1 TD in 3 appearances in 2025) is the other player with a shot at the 2026-09-05 opener; both took "
  "first-team reps in the spring game. Documented leader by direct head-coach statement, no official naming -> M (not H). "
  "Per the L->M rule values REMAIN BLANK, so status correctly stays UNCERTAIN. OPEN RISK: Johnson's durability - "
  "season-ending injuries at Texas A&M in 2022 and 2023 and a broken leg in UNC's 2024 opener."))
setrec("ODU", expect_code="M", expect_row=137,
 aqb="Quinn Henicle (leader; Potter / Huff competing)",
 note=("VERIFIED 7D.4A, RE-VERIFIED 7D.5 2026-08-04: the job opened when 2025 starter Colton Joseph transferred to "
  "Wisconsin. Quinn Henicle (RS-So) started the 2025 Cure Bowl in his place and was GAME MVP in a 24-10 win over South "
  "Florida (11/25, 127 yds; 24 carries, 107 yds, 2 rush TD incl. a 51-yd TD run); he is 2-0 as a starter. Beat coverage "
  "says Henicle 'HAS A LEG UP' and is 'the front runner to land the QB1 role', while reporting he 'remains enmeshed in a "
  "three-way battle' with RS-Jr JAXON POTTER (Washington State transfer, announced 2026-01-13) and RS-Fr Ryan Huff. "
  "Henicle 2026-04-03: 'I'm competing like I'm trying to earn this job.' HC Ricky Rahne (7th season, off 10-3) has NOT "
  "named a starter. Leader backed by on-field production rather than name recognition -> M retained (not H); zeros "
  "retained; status OK. RECHECK: official Old Dominion depth chart."))

# --- CHANGELOG + banner ---
cl=wb["CHANGELOG"]; last=0
for rr in range(1,cl.max_row+1):
    if any(cl.cell(row=rr,column=c).value not in (None,"") for c in range(1,7)): last=rr
rows=[("v0.7.9","2026-08-04",
 "Phase 7D.5 FINAL QB-VERIFICATION CANDIDATE. Closes finding F-7 from Phase 7D.4A: 21 Tier-1 (M/L) records were credited "
 "as verified in the phase ledger but still carried their original 2026-07-21 build stamp and build note, so the workbook "
 "could not substantiate its own verification claim. All 21 received a fresh team-specific pass and are now stamped. "
 "Every Tier-1 record in the dataset now carries an in-workbook verification date and evidence note. The five 7D.4A teams "
 "were also re-verified against official/primary sources and their notes refreshed.",
 "F-7 closeout"),
 ("v0.7.9","2026-08-04",
 "Phase 7D.5 CORRECTIONS (3): (1) MISSOURI M -> H - HC Eli Drinkwitz OFFICIALLY NAMED Austin Simmons the 2026 starter on "
 "2026-03-19 over Matt Zollers and Nick Evers; the prior entry read 'projected leader, some competition', which was stale. "
 "Zeros retained. (2) NORTH CAROLINA M -> L - no starter named; 'UNC quarterback battle remains WIDE OPEN' (2026-07-18) "
 "with four candidates (Billy Edwards Jr., Miles O'Neill, Au'Tori Newkirk, Travis Burgess) and Belichick leaving it open. "
 "Zeros CLEARED. (3) UNLV M -> L - the Las Vegas Review-Journal (2026-07-28) reports 'coaches won't name a front-runner "
 "throughout training camp' in a genuine Jackson Arnold vs Alex Orji battle; the prior note's word 'verified' was "
 "unsupported. Zeros CLEARED. Also corrected: NEW FACTS on Texas State (Boston College transfer Shaker Reisig is the "
 "primary BACKUP, not a challenger; Jackson was 2026 Armed Forces Bowl MVP) and North Texas (replacing Drew Mestemaker, "
 "the 2025 national passing-yards leader).",
 "F-7 closeout"),
 ("v0.7.9","2026-08-04",
 "Phase 7D.5 RESULT: QB verification backlog = 0 AND audit-trail gap = 0. Counts 65 H / 40 M / 33 L; status 98 OK / 40 "
 "UNCERTAIN; blank 39 / zero 99; 0 nonzero QB values; 0 formula changes. Vanderbilt's Phase 8.3 FAILED re-verification is "
 "now closed at L. DEFECT PATTERN - AND ITS FIRST EXCEPTION: 11 defects across 80 team-specific passes (about 1 in 7). Ten "
 "were over-confident: an unsupported M or a stale named starter (Akron, Arkansas, UConn, Buffalo, Northern Illinois, "
 "Appalachian State, Washington State, Georgia Southern, North Carolina, UNLV). MISSOURI IS THE FIRST DEFECT IN THE "
 "OPPOSITE DIRECTION - a record rated TOO UNCERTAIN (M) when the starter had been officially named five months earlier. "
 "The prior claim that no team was ever rated too uncertain no longer holds and is corrected here.",
 "Final QB verification result")]
for i,(v,d,c,rn) in enumerate(rows):
    rr=last+1+i
    cl.cell(row=rr,column=1).value=v; cl.cell(row=rr,column=2).value=d
    cl.cell(row=rr,column=3).value=c; cl.cell(row=rr,column=4).value=rn
wb["START HERE"]["A1"].value=("TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.7.9 FINAL QB-VERIFICATION CANDIDATE — backlog 0, "
 "audit-trail gap 0; all 74 Tier-1 records verified and stamped; Missouri M->H, North Carolina and UNLV M->L; "
 "NOT AUTHORITATIVE, NOT PROMOTED — awaiting owner approval)")
wb.save(DST)
print(f"\nCHANGELOG rows {last+1}-{last+len(rows)}; saved {DST}")

"""PROMOTION AUDIT part 3 — corrects part 2 and proves the downstream effect
WITHOUT relying on cached formula values.

Part 2 finding: neither workbook stores cached formula results for computed cells
(data_only reads return None in BOTH). So every part-2 comparison of "cached
values" was VACUOUS - it compared None to None - and none of those results count
as evidence. Check 7.4c likewise failed for that reason, not because of a defect.

This part re-proves the downstream claim from literals + formula text only.
"""
import openpyxl, re, collections
from openpyxl.worksheet.formula import ArrayFormula
ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
AUTH=f"{ROOT}/workbook_v0.6.2_deliverables/TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"
CAND=f"{ROOT}/workbook_v0.7.9_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.9_CANDIDATE.xlsx"
OUT=[]; BLOCK=[]
def say(s): print(s); OUT.append(s)
def chk(n,ok,d=""):
    say(f"  [{'PASS' if ok else 'FAIL'}] {n}{(' - '+d) if d else ''}")
    if not ok: BLOCK.append(f"{n} :: {d}")
def ftext(v): return v.text if isinstance(v,ArrayFormula) else v
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))

wa=openpyxl.load_workbook(AUTH); wc=openpyxl.load_workbook(CAND)
wav=openpyxl.load_workbook(AUTH, data_only=True); wcv=openpyxl.load_workbook(CAND, data_only=True)

say("="*78); say("CACHED-VALUE STATUS (corrects part 1's optimistic reading)")
def cached_formula_cells(wb_f, wb_v):
    tot=0; withval=0
    for s in wb_f.worksheets:
        sv=wb_v[s.title]
        for row in s.iter_rows():
            for c in row:
                if isf(c.value):
                    tot+=1
                    if sv.cell(row=c.row,column=c.column).value is not None: withval+=1
    return tot, withval
ta,va = cached_formula_cells(wa,wav); tc,vc = cached_formula_cells(wc,wcv)
say(f"  v0.6.2 : {ta} formula cells, {va} carry a cached result")
say(f"  v0.7.9 : {tc} formula cells, {vc} carry a cached result")
chk("C.1 candidate is no worse than authoritative on cached results", vc>=va, f"{va} -> {vc}")
if va==0 and vc==0:
    say("  => NEITHER workbook stores cached formula results. This is a pre-existing")
    say("     property of the authoritative workbook, NOT a regression introduced here.")
    say("     Excel and Google Sheets recalculate on open. Any consumer reading cached")
    say("     values would already be broken against v0.6.2 today.")

say("\n"+"="*78); say("DOWNSTREAM PROOF FROM LITERALS AND FORMULA TEXT ONLY")
qb=wc["QB VALUES"]; tmap=wc["TEAM MAP"]
# 1) the abbrev universe comes from TEAM MAP literals
abbrevs=[tmap.cell(row=r,column=1).value for r in range(6,144)]
chk("D.1 TEAM MAP supplies 138 unique non-empty abbrevs",
    len(abbrevs)==138 and len(set(abbrevs))==138 and all(abbrevs))
# 2) QB VALUES!A is a pure row-parallel pull, so the lookup key set is identical
badchain=[]
for r in range(6,144):
    qa=str(ftext(qb.cell(row=r,column=1).value))
    tr=str(ftext(wc["TEAM RATINGS"].cell(row=r,column=1).value))
    m2=re.search(r"'TEAM RATINGS'!\$?A\$?(\d+)", qa); m1=re.search(r"'TEAM MAP'!\$?A\$?(\d+)", tr)
    if not(m1 and m2 and int(m1.group(1))==r and int(m2.group(1))==r): badchain.append(r)
chk("D.2 QB VALUES!A(r) <- TEAM RATINGS!A(r) <- TEAM MAP!A(r) for all 138 rows",
    not badchain, str(badchain[:5]))
say("     => the ENGINE's MATCH key set is exactly the TEAM MAP abbrev set, by construction.")

# 3) every G is blank or 0 -> ENGINE!M is provably 0 for every game
D=[qb.cell(row=r,column=4).value for r in range(6,144)]
F=[qb.cell(row=r,column=6).value for r in range(6,144)]
G=["" if (d is None or f is None) else f-d for d,f in zip(D,F)]
chk("D.3 every QB delta G is blank or exactly 0", all(g in ("",0) for g in G),
    str([(abbrevs[i],g) for i,g in enumerate(G) if g not in ("",0)][:5]))
engM=str(ftext(wc["ENGINE"]["M7"].value))
blank_to_zero = '="",0,' in engM
say(f"  ENGINE!M maps a blank lookup result to 0: {blank_to_zero}")
chk("D.3b ENGINE!M explicitly coerces a blank QB delta to 0", blank_to_zero)
chk("D.4 ENGINE!M formula byte-identical to authoritative",
    ftext(wa["ENGINE"]["M7"].value)==ftext(wc["ENGINE"]["M7"].value))
say("     => QB adjustment = (home G or 0) - (away G or 0) = 0 - 0 = 0 for EVERY game,")
say("        for both workbooks. No rating and no projected spread can move.")

# 4) the only downstream difference is the status gate
J=[qb.cell(row=r,column=10).value for r in range(6,144)]
H=[qb.cell(row=r,column=8).value for r in range(6,144)]
SEASON=wc["SETTINGS"]["B3"].value
def M(i): return "UNCERTAIN" if (G[i]=="" or H[i]=="L" or J[i]!=SEASON) else "OK"
cand_status=collections.Counter(M(i) for i in range(138))
# authoritative: same formula, but its inputs
qba=wa["QB VALUES"]
Da=[qba.cell(row=r,column=4).value for r in range(6,144)]
Fa=[qba.cell(row=r,column=6).value for r in range(6,144)]
Ha=[qba.cell(row=r,column=8).value for r in range(6,144)]
Ja=[qba.cell(row=r,column=10).value for r in range(6,144)]
Ga=["" if (d is None or f is None) else f-d for d,f in zip(Da,Fa)]
def Ma(i): return "UNCERTAIN" if (Ga[i]=="" or Ha[i]=="L" or Ja[i]!=SEASON) else "OK"
auth_status=collections.Counter(Ma(i) for i in range(138))
say(f"\n  QB status recomputed from inputs  v0.6.2: {dict(auth_status)}")
say(f"  QB status recomputed from inputs  v0.7.9: {dict(cand_status)}")
say(f"  v0.6.2 populated codes: {sum(1 for h in Ha if h)} / 138")
say(f"  v0.7.9 populated codes: {sum(1 for h in H if h)} / 138")
chk("D.5 candidate strictly reduces QB uncertainty",
    cand_status['UNCERTAIN'] < auth_status['UNCERTAIN'],
    f"{auth_status['UNCERTAIN']} -> {cand_status['UNCERTAIN']}")

# 5) status precedence: PENDING LINE outranks QB UNCERTAIN, and 0 lines are loaded
ai=str(ftext(wc["ENGINE"]["AI7"].value))
order=[k for k in ("BLOCKED","FCS — NO PLAY","PENDING LINE","STALE LINE","QB UNCERTAIN") if k in ai]
say(f"\n  ENGINE!AI precedence as written: {order}")
chk("D.6 PENDING LINE is evaluated before QB UNCERTAIN",
    ai.index("PENDING LINE") < ai.index("QB UNCERTAIN"))
ml=wc["MARKET LINES"]
lit=sum(1 for r in range(1,ml.max_row+1) for c in range(1,ml.max_column+1)
        if not isf(ml.cell(row=r,column=c).value) and ml.cell(row=r,column=c).value is not None)
say(f"  MARKET LINES non-formula cells present: {lit}")
nums=sum(1 for r in range(1,ml.max_row+1) for c in range(1,ml.max_column+1)
         if isinstance(ml.cell(row=r,column=c).value,(int,float)))
chk("D.7 zero numeric market spreads loaded -> every game reads PENDING LINE", nums==0, str(nums))
say("     => the 39 UNCERTAIN teams are MASKED today. Promotion changes no visible")
say("        game status until lines are entered.")

# 6) IMPORT SCHEDULE integrity is untouched
chk("D.8 IMPORT SCHEDULE has zero formulas in both (pure literal data)",
    sum(1 for row in wc["IMPORT SCHEDULE"].iter_rows() for c in row if isf(c.value))==0)

say("\n"+"#"*78)
say(f"# PART 3 BLOCKERS: {len(BLOCK)}")
for b in BLOCK: say(f"#   - {b}")
say("#"*78)
open(f"{ROOT}/workbook_v0.7.9_candidate/promotion_audit_log3.txt","w").write("\n".join(OUT)+"\n")

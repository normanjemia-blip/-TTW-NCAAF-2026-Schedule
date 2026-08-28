"""PROMOTION AUDIT part 2 — corrects check 7.4 and completes the downstream trace.

7.4 in part 1 was a BUG IN THE AUDIT, not a workbook defect: TEAM RATINGS!A is a
formula (='TEAM MAP'!$A{row}), so comparing it against the TEAM MAP literal always
mismatches. Correct test: parse the row reference out of the formula and prove it
points at the same row. That is a STRONGER alignment proof than value equality.
"""
import openpyxl, re, collections, json
from openpyxl.worksheet.formula import ArrayFormula
ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
AUTH=f"{ROOT}/workbook_v0.6.2_deliverables/TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"
CAND=f"{ROOT}/workbook_v0.7.9_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.9_CANDIDATE.xlsx"
OUT=[]; BLOCK=[]
def say(s): print(s); OUT.append(s)
def chk(n,ok,d=""):
    say(f"  [{'PASS' if ok else 'FAIL'}] {n}{(' - '+d) if d else ''}")
    if not ok: BLOCK.append(f"{n} :: {d}")
def ftext(v): return v.text if isinstance(v,ArrayFormula) else v

wc=openpyxl.load_workbook(CAND); wcv=openpyxl.load_workbook(CAND, data_only=True)
wav=openpyxl.load_workbook(AUTH, data_only=True)

say("="*78); say("7.4 CORRECTED — ROW ALIGNMENT THROUGH THE FORMULA CHAIN")
tmap=wc["TEAM MAP"]; trat=wc["TEAM RATINGS"]; qb=wc["QB VALUES"]
bad=[]
for r in range(6,144):
    tr=ftext(trat.cell(row=r,column=1).value)
    qa=ftext(qb.cell(row=r,column=1).value)
    m1=re.search(r"'TEAM MAP'!\$?A\$?(\d+)", str(tr))
    m2=re.search(r"'TEAM RATINGS'!\$?A\$?(\d+)", str(qa))
    if not m1 or int(m1.group(1))!=r: bad.append(("TEAM RATINGS",r,tr))
    if not m2 or int(m2.group(1))!=r: bad.append(("QB VALUES",r,qa))
chk("7.4a TEAM RATINGS!A{r} references TEAM MAP!A{r} for all 138 rows", not bad, str(bad[:5]))
chk("7.4b QB VALUES!A{r} references TEAM RATINGS!A{r} for all 138 rows", not bad, str(bad[:5]))
# and prove it resolves to the same abbrev via cached values
qbv=wcv["QB VALUES"]
mism=[(r, tmap.cell(row=r,column=1).value, qbv.cell(row=r,column=1).value) for r in range(6,144)
      if tmap.cell(row=r,column=1).value != qbv.cell(row=r,column=1).value]
chk("7.4c cached QB VALUES!A resolves to the TEAM MAP abbrev on all 138 rows",
    not mism, str(mism[:5]))

say("\n"+"="*78); say("10b. DOWNSTREAM LOOKUP INTEGRITY (ENGINE MATCHes by abbrev, not by row)")
qb_abbrevs={qbv.cell(row=r,column=1).value for r in range(6,144)}
clean=wcv["CLEAN"]; eng=wcv["ENGINE"]
# CLEAN!N = away abbrev, CLEAN!P = home abbrev (per the ENGINE formulas); ENGINE!G/H = FBS flags
rows=range(6,1006)
unresolved=[]; games=0; fbs_refs=0
for r in rows:
    n=clean.cell(row=r,column=14).value   # N
    p=clean.cell(row=r,column=16).value   # P
    g=eng.cell(row=r,column=7).value      # G away class
    h=eng.cell(row=r,column=8).value      # H home class
    if not (n or p): continue
    games+=1
    for ab,cls in ((n,g),(p,h)):
        if cls=="FBS":
            fbs_refs+=1
            if ab not in qb_abbrevs: unresolved.append((r,ab))
say(f"  game rows with team data: {games}   FBS team-slots referenced: {fbs_refs}")
chk("10b.1 every FBS team-slot in CLEAN resolves to a QB VALUES row",
    not unresolved, str(unresolved[:10]))

say("\n"+"="*78); say("10c. DOWNSTREAM EFFECT — v0.6.2 vs v0.7.9, computed not asserted")
def qbstate(wbv):
    s=wbv["QB VALUES"]; out={}
    for r in range(6,144):
        ab=s.cell(row=r,column=1).value
        out[ab]=s.cell(row=r,column=13).value   # M = QB status (cached)
    return out
sa, sc = qbstate(wav), qbstate(wcv)
ca=collections.Counter(sa.values()); cc=collections.Counter(sc.values())
say(f"  QB VALUES!M cached status  v0.6.2: {dict(ca)}")
say(f"  QB VALUES!M cached status  v0.7.9: {dict(cc)}")
# ENGINE!AE per game, recomputed from the status vector
def ae(state):
    n_unc=0
    for r in rows:
        nn=clean.cell(row=r,column=14).value; pp=clean.cell(row=r,column=16).value
        g=eng.cell(row=r,column=7).value; h=eng.cell(row=r,column=8).value
        if not (nn or pp): continue
        away = (g=="FBS") and (state.get(nn,"UNCERTAIN")=="UNCERTAIN" or nn not in state)
        home = (h=="FBS") and (state.get(pp,"UNCERTAIN")=="UNCERTAIN" or pp not in state)
        if away or home: n_unc+=1
    return n_unc
u_a, u_c = ae(sa), ae(sc)
say(f"  games that would read QB UNCERTAIN at ENGINE!AE  v0.6.2: {u_a}  ->  v0.7.9: {u_c}")
# but AI precedence: PENDING LINE outranks QB UNCERTAIN
ml=wcv["MARKET LINES"]
spreads=sum(1 for r in range(1,ml.max_row+1) for c in range(1,ml.max_column+1)
            if isinstance(ml.cell(row=r,column=c).value,(int,float)))
say(f"  MARKET LINES numeric cells loaded: {spreads}")
ai_a=collections.Counter(eng.cell(row=r,column=35).value for r in rows
                         if eng.cell(row=r,column=35).value)  # AI
say(f"  ENGINE!AI cached status distribution in the candidate: {dict(ai_a)}")
chk("10c.1 no market spreads loaded (PENDING LINE masks every QB status change)", spreads==0, str(spreads))
chk("10c.2 QB-status improvement is real (fewer uncertain teams than v0.6.2)",
    cc.get("UNCERTAIN",0) < ca.get("UNCERTAIN",138),
    f"{ca.get('UNCERTAIN')} -> {cc.get('UNCERTAIN')}")

say("\n"+"="*78); say("10d. USER-VISIBLE COUNTER (START HERE!C11)")
sh_a=wav["START HERE"]; sh_c=wcv["START HERE"]
say(f"  v0.6.2 START HERE!C11 = {sh_a['C11'].value!r}")
say(f"  v0.7.9 START HERE!C11 = {sh_c['C11'].value!r}")

say("\n"+"="*78); say("10e. TEAM RATINGS / projected spreads — cached value comparison")
def col_vals(wbv, sheet, col, lo, hi):
    s=wbv[sheet]; return [s.cell(row=r,column=col).value for r in range(lo,hi)]
same=True; detail=[]
for col,label in ((3,"C"),(4,"D"),(5,"E"),(6,"F"),(7,"G"),(8,"H")):
    a=col_vals(wav,"TEAM RATINGS",col,6,144); c=col_vals(wcv,"TEAM RATINGS",col,6,144)
    if a!=c: same=False; detail.append(label)
chk("10e.1 TEAM RATINGS cached values identical (columns C-H, all 138 rows)", same, str(detail))
mdiff=[r for r in rows if eng.cell(row=r,column=13).value != wav["ENGINE"].cell(row=r,column=13).value]
chk("10e.2 ENGINE!M (QB adjustment) cached values identical for every game row",
    not mdiff, f"{len(mdiff)} rows differ: {mdiff[:10]}")
sdiff=[r for r in rows if eng.cell(row=r,column=20).value != wav["ENGINE"].cell(row=r,column=20).value]
say(f"  ENGINE column T cached differences: {len(sdiff)}")

say("\n"+"#"*78)
say(f"# PART 2 BLOCKERS: {len(BLOCK)}")
for b in BLOCK: say(f"#   - {b}")
say("#"*78)
open(f"{ROOT}/workbook_v0.7.9_candidate/promotion_audit_log2.txt","w").write("\n".join(OUT)+"\n")

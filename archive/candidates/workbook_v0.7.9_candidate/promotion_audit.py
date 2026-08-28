"""PROMOTION AUDIT — independent, adversarial. Assume the workbook is wrong.

Recomputes everything from the files themselves. Does not import or trust
verify_v079.py. Every check is written to FAIL LOUD.
"""
import openpyxl, hashlib, zipfile, re, json, collections, datetime, sys
from openpyxl.worksheet.formula import ArrayFormula

ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
AUTH=f"{ROOT}/workbook_v0.6.2_deliverables/TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"
CAND=f"{ROOT}/workbook_v0.7.9_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.9_CANDIDATE.xlsx"
OUT=[]
BLOCKERS=[]; WARNINGS=[]
def say(s): print(s); OUT.append(s)
def block(s): BLOCKERS.append(s); say(f"  *** BLOCKER: {s}")
def warn(s):  WARNINGS.append(s); say(f"  ~~~ WARNING: {s}")
def chk(name, ok, detail="", blocking=True):
    say(f"  [{'PASS' if ok else 'FAIL'}] {name}{(' - '+detail) if detail else ''}")
    if not ok: (block if blocking else warn)(f"{name} :: {detail}")

def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
def ftext(v): return v.text if isinstance(v,ArrayFormula) else v
def sha(p):
    h=hashlib.sha256()
    with open(p,"rb") as f:
        for b in iter(lambda:f.read(1<<20),b""): h.update(b)
    return h.hexdigest()

say("#"*78); say("# PROMOTION AUDIT — v0.7.9 CANDIDATE vs v0.6.2 AUTHORITATIVE")
say(f"# generated 2026-08-04 (America/New_York)"); say("#"*78)

# ============ 3. WORKBOOK CHECKSUMS ============
say("\n=== 3. WORKBOOK CHECKSUMS ===")
ha, hc = sha(AUTH), sha(CAND)
say(f"  v0.6.2 AUTHORITATIVE : {ha}")
say(f"  v0.7.9 CANDIDATE     : {hc}")
pm=json.load(open(f"{ROOT}/PROJECT_MANIFEST.json")); pmtxt=json.dumps(pm)
chk("3.1 v0.6.2 SHA matches PROJECT_MANIFEST.json", ha in pmtxt, ha[:16])
chk("3.2 v0.6.2 SHA is the known-good constant",
    ha=="bbb17b50fbfb728bea2a23d3d20771935cc61e238313a054473aafe1ca838efd", ha[:16])
chk("3.3 candidate differs from authoritative", ha!=hc)

# ============ 2. WORKSHEET (XML PART) CHECKSUMS ============
say("\n=== 2. WORKSHEET CHECKSUMS (raw OOXML parts) ===")
def parts(p):
    d={}
    with zipfile.ZipFile(p) as z:
        for n in sorted(z.namelist()):
            d[n]=hashlib.sha256(z.read(n)).hexdigest()
    return d
pa, pc = parts(AUTH), parts(CAND)
only_a=sorted(set(pa)-set(pc)); only_c=sorted(set(pc)-set(pa))
common=sorted(set(pa)&set(pc))
diff_parts=[n for n in common if pa[n]!=pc[n]]
say(f"  parts: auth {len(pa)} | cand {len(pc)} | common {len(common)}")
say(f"  parts only in AUTH: {only_a}")
say(f"  parts only in CAND: {only_c}")
say(f"  parts differing ({len(diff_parts)}):")
for n in diff_parts: say(f"      {n}")

# map sheet name -> xml part
def sheetmap(p):
    wb=openpyxl.load_workbook(p, read_only=True)
    names=wb.sheetnames; wb.close()
    return names
names_a, names_c = sheetmap(AUTH), sheetmap(CAND)
chk("2.1 sheet names and order identical", names_a==names_c)
chk("2.2 21 sheets", len(names_c)==21, str(len(names_c)))

# ============ 1. FORMULA COUNTS (per sheet, exact) ============
say("\n=== 1. FORMULA COUNTS (per sheet) ===")
wa=openpyxl.load_workbook(AUTH); wc=openpyxl.load_workbook(CAND)
def fcount(wb):
    d={}
    for s in wb.worksheets:
        d[s.title]=sum(1 for row in s.iter_rows() for c in row if isf(c.value))
    return d
fa, fc = fcount(wa), fcount(wc)
say(f"  {'sheet':<22}{'auth':>9}{'cand':>9}{'delta':>8}")
tot_a=tot_c=0
for n in names_c:
    a,c=fa.get(n,0),fc.get(n,0); tot_a+=a; tot_c+=c
    flag="" if a==c else "   <<< DELTA"
    say(f"  {n:<22}{a:>9}{c:>9}{c-a:>8}{flag}")
say(f"  {'TOTAL':<22}{tot_a:>9}{tot_c:>9}{tot_c-tot_a:>8}")
chk("1.1 total formula count identical", tot_a==tot_c, f"{tot_a} vs {tot_c}")
chk("1.2 total is the known invariant 123011", tot_c==123011, str(tot_c))
chk("1.3 per-sheet formula counts identical", fa==fc,
    str({k:(fa.get(k),fc.get(k)) for k in names_c if fa.get(k)!=fc.get(k)}))

# ============ 4/5. MODIFIED SHEETS AND EVERY CHANGED CELL ============
say("\n=== 4/5. MODIFIED SHEETS AND EVERY CHANGED CELL ===")
changed=[]; formula_changed=[]
for n in names_c:
    sa, sc = wa[n], wc[n]
    mr=max(sa.max_row, sc.max_row); mc=max(sa.max_column, sc.max_column)
    for r in range(1, mr+1):
        for col in range(1, mc+1):
            x=ftext(sa.cell(row=r,column=col).value); y=ftext(sc.cell(row=r,column=col).value)
            if x!=y:
                addr=f"{openpyxl.utils.get_column_letter(col)}{r}"
                changed.append((n,addr,col,r,x,y))
                if isf(sa.cell(row=r,column=col).value) or isf(sc.cell(row=r,column=col).value):
                    formula_changed.append((n,addr))
bysheet=collections.Counter(c[0] for c in changed)
say(f"  total changed cells: {len(changed)}")
for k,v in bysheet.items(): say(f"    {k}: {v}")
untouched=[n for n in names_c if n not in bysheet]
say(f"  sheets with ZERO changed cells ({len(untouched)}): {untouched}")
chk("4.1 exactly 3 sheets modified", set(bysheet)=={"QB VALUES","CHANGELOG","START HERE"}, str(sorted(bysheet)))
chk("4.2 18 sheets untouched at cell level", len(untouched)==18, str(len(untouched)))
chk("5.1 ZERO formula cells changed", len(formula_changed)==0, str(formula_changed[:10]))

# cross-check: do the untouched sheets also have identical XML parts?
say("\n  cross-check: cell-level 'untouched' vs raw XML part hashes")
xml_diff_sheets=[n for n in diff_parts if n.startswith("xl/worksheets/")]
say(f"    worksheet XML parts differing: {len(xml_diff_sheets)} -> {xml_diff_sheets}")
if len(xml_diff_sheets) > 3:
    warn(f"{len(xml_diff_sheets)} worksheet XML parts differ but only 3 sheets have cell changes; "
         "openpyxl rewrites all parts on save, so XML-level equality is NOT expected. "
         "Cell-level comparison is the authoritative test here.")

# every changed cell in QB VALUES, itemised by column
say("\n  QB VALUES changed cells by column:")
qbc=[c for c in changed if c[0]=="QB VALUES"]
colcount=collections.Counter(openpyxl.utils.get_column_letter(c[2]) for c in qbc)
say(f"    {dict(colcount)}")
ALLOWED_QB_COLS={"C","D","E","F","H","I","J","K","L"}   # input columns only
bad=[c for c in qbc if openpyxl.utils.get_column_letter(c[2]) not in ALLOWED_QB_COLS]
chk("5.2 no changed cell in a QB VALUES formula column (A,B,G,M)", len(bad)==0,
    str([(c[1]) for c in bad][:10]))
qbrows=sorted({c[3] for c in qbc})
say(f"    QB VALUES rows touched ({len(qbrows)}): {qbrows}")
chk("5.3 all touched QB rows are inside the team block 6-143",
    all(6<=r<=143 for r in qbrows), str([r for r in qbrows if not 6<=r<=143]))

# ============ 9. ENGINE DEPENDENCIES — what actually reads QB VALUES ============
say("\n=== 9. ENGINE DEPENDENCIES: every formula referencing QB VALUES ===")
ref=re.compile(r"(?:'QB VALUES'|QB VALUES)!\$?([A-Z]{1,3})\$?\d*", re.I)
refs=collections.Counter(); ref_sheets=collections.Counter(); samples={}
for s in wc.worksheets:
    for row in s.iter_rows():
        for c in row:
            v=ftext(c.value)
            if isinstance(v,str) and "QB VALUES" in v.upper():
                for col in ref.findall(v):
                    refs[col.upper()]+=1; ref_sheets[(s.title,col.upper())]+=1
                    samples.setdefault((s.title,col.upper()), (c.coordinate, v[:150]))
say(f"  columns of QB VALUES referenced anywhere in the workbook: {dict(refs)}")
for (sh,col),n in sorted(ref_sheets.items()):
    coord,f=samples[(sh,col)]
    say(f"    {sh}!{coord} -> QB VALUES!{col}  (x{n})  {f}")
CONSUMED=set(refs)
say(f"  consumed columns = {sorted(CONSUMED)}")
chk("9.1 no formula reads QB VALUES column H (confidence code)", "H" not in CONSUMED,
    "H IS consumed - confidence-code edits would move ratings")
chk("9.2 no formula reads QB VALUES columns C,E,I,J,K,L (metadata)",
    not (CONSUMED & {"C","E","I","J","K","L"}), str(sorted(CONSUMED & {"C","E","I","J","K","L"})))
chk("9.3 consumed columns are a subset of {A,B,D,F,G,M}",
    CONSUMED <= {"A","B","D","F","G","M"}, str(sorted(CONSUMED)))

# defined names
say("\n  defined names referencing QB VALUES:")
dn=[(n,str(d.value)) for n,d in wc.defined_names.items() if "QB VALUES" in str(d.value).upper()]
for n,v in dn: say(f"    {n} = {v}")
say(f"    ({len(dn)} found)")
dna={n:str(d.value) for n,d in wa.defined_names.items()}
dnc={n:str(d.value) for n,d in wc.defined_names.items()}
chk("9.4 defined names identical to authoritative", dna==dnc,
    str({k:(dna.get(k),dnc.get(k)) for k in set(dna)|set(dnc) if dna.get(k)!=dnc.get(k)}))

# conditional formatting / data validation on QB VALUES
def cf_sig(wb):
    out={}
    for s in wb.worksheets:
        rules=[]
        try:
            for rng, rs in s.conditional_formatting._cf_rules.items():
                for r in rs: rules.append((str(rng.sqref), r.type, str(getattr(r,'formula',None))))
        except Exception as e: rules.append(("ERR",str(e),""))
        dv=[]
        try:
            for v in s.data_validations.dataValidation: dv.append((str(v.sqref), v.type, str(v.formula1)))
        except Exception: pass
        out[s.title]=(sorted(rules), sorted(dv))
    return out
cfa, cfc = cf_sig(wa), cf_sig(wc)
diff_cf=[k for k in cfc if cfa.get(k)!=cfc.get(k)]
chk("9.5 conditional formatting + data validation identical on every sheet",
    len(diff_cf)==0, str(diff_cf))

# ============ 7/8. QB CLASSIFICATIONS AND NUMERICAL VALUES ============
say("\n=== 7/8. QB CLASSIFICATIONS AND NUMERICAL VALUES (all 138) ===")
qb=wc["QB VALUES"]; tmap=wc["TEAM MAP"]; trat=wc["TEAM RATINGS"]
rows=list(range(6,144))
recs=[]
for r in rows:
    recs.append(dict(row=r,
        map_ab=tmap.cell(row=r,column=1).value, map_team=tmap.cell(row=r,column=2).value,
        rat_ab=trat.cell(row=r,column=1).value, rat_team=trat.cell(row=r,column=2).value,
        base_qb=qb.cell(row=r,column=3).value, D=qb.cell(row=r,column=4).value,
        act_qb=qb.cell(row=r,column=5).value, F=qb.cell(row=r,column=6).value,
        H=qb.cell(row=r,column=8).value, I=qb.cell(row=r,column=9).value,
        J=qb.cell(row=r,column=10).value, K=qb.cell(row=r,column=11).value,
        L=qb.cell(row=r,column=12).value))
chk("7.1 exactly 138 team rows", len(recs)==138, str(len(recs)))
chk("7.2 TEAM MAP abbrevs unique", len({x['map_ab'] for x in recs})==138)
chk("7.3 TEAM MAP team names unique", len({x['map_team'] for x in recs})==138)
mis=[(x['row'],x['map_ab'],x['rat_ab']) for x in recs if x['map_ab']!=x['rat_ab']]
chk("7.4 TEAM MAP row N aligns with TEAM RATINGS row N", len(mis)==0, str(mis[:10]))
codes=collections.Counter(x['H'] for x in recs)
say(f"  codes: {dict(codes)}")
chk("7.5 every code is exactly H, M or L", all(x['H'] in ("H","M","L") for x in recs),
    str([(x['row'],x['H']) for x in recs if x['H'] not in ("H","M","L")]))
chk("7.6 no blank / whitespace codes", all(isinstance(x['H'],str) and x['H']==x['H'].strip() for x in recs))
chk("7.7 all 138 codes populated (v0.6.2 had 0)", sum(1 for x in recs if x['H'])==138)

# numerical values
say("\n  numerical QB values:")
vals=[(x['row'],x['map_ab'],x['D'],x['F']) for x in recs]
nonzero=[v for v in vals if (v[2] not in (None,0)) or (v[3] not in (None,0))]
chk("8.1 ZERO nonzero QB values anywhere", len(nonzero)==0, str(nonzero[:10]))
badtype=[v for v in vals if not (v[2] is None or isinstance(v[2],(int,float))) or
                             not (v[3] is None or isinstance(v[3],(int,float)))]
chk("8.2 every numerical cell is blank or numeric (no strings/zeros-as-text)", len(badtype)==0, str(badtype[:5]))
mixed=[v for v in vals if (v[2] is None) != (v[3] is None)]
chk("8.3 D and F are always both blank or both populated", len(mixed)==0, str(mixed[:10]))
blank=[v for v in vals if v[2] is None]; zero=[v for v in vals if v[2]==0]
say(f"    blank {len(blank)} | zero {len(zero)} | total {len(blank)+len(zero)}")
chk("8.4 blank + zero accounts for all 138", len(blank)+len(zero)==138)
lviol=[x['map_ab'] for x in recs if x['H']=="L" and x['D'] is not None]
chk("8.5 EVERY L-coded team has blank numerical inputs", len(lviol)==0, str(lviol))
hviol=[x['map_ab'] for x in recs if x['H']=="H" and x['D'] is None]
chk("8.6 H-coded teams with BLANK inputs (would gate a confirmed starter)",
    len(hviol)==0, str(hviol), blocking=False)

# metadata completeness for Tier 1
say("\n  Tier-1 audit trail:")
t1=[x for x in recs if x['H'] in ("M","L")]
say(f"    Tier-1 population: {len(t1)}")
def dstr(v): return v.date().isoformat() if isinstance(v,datetime.datetime) else v
nodate=[x['map_ab'] for x in t1 if x['K'] is None]
chk("7.8 every Tier-1 record has a Last-update date", len(nodate)==0, str(nodate))
stale=[x['map_ab'] for x in t1 if dstr(x['K'])=="2026-07-21"]
chk("7.9 no Tier-1 record still carries the 2026-07-21 build stamp", len(stale)==0, str(stale))
nonote=[x['map_ab'] for x in t1 if not x['L'] or len(str(x['L']))<40]
chk("7.10 every Tier-1 record carries a substantive note (>=40 chars)", len(nonote)==0, str(nonote))
noseason=[x['map_ab'] for x in recs if x['J']!=2026]
chk("7.11 every record reviewed-for-season = 2026", len(noseason)==0, str(noseason))
hstale=[x['map_ab'] for x in recs if x['H']=="H" and dstr(x['K'])=="2026-07-21"]
say(f"    H-coded tier-2 rows still on the 2026-07-21 build stamp: {len(hstale)}")
if hstale: warn(f"{len(hstale)} H-coded records were never independently verified "
                "(out of declared scope, disclosed in PHASE_7D5_REPORT section 7)")

# ============ 10. DOWNSTREAM CALCULATIONS ============
say("\n=== 10. DOWNSTREAM CALCULATIONS ===")
SEASON=wc["SETTINGS"]["B3"].value
say(f"  SETTINGS!B3 season = {SEASON}   B6 HFA = {wc['SETTINGS']['B6'].value}   B11 BET = {wc['SETTINGS']['B11'].value!r}")
chk("10.1 SETTINGS!B3 = 2026", SEASON==2026)
chk("10.2 SETTINGS!B6 HFA = 2.5", wc["SETTINGS"]["B6"].value==2.5)
chk("10.3 SETTINGS!B11 BET toggle = 'N'", wc["SETTINGS"]["B11"].value=="N")
# faithful re-implementation of the two QB formulas, verified against the actual text
g_formula = ftext(qb["G6"].value); m_formula = ftext(qb["M6"].value)
say(f"  QB VALUES!G6 = {g_formula}")
say(f"  QB VALUES!M6 = {m_formula}")
chk("10.4 G formula matches authoritative", ftext(wa['QB VALUES']['G6'].value)==g_formula)
chk("10.5 M formula matches authoritative", ftext(wa['QB VALUES']['M6'].value)==m_formula)
def G(x): return "" if (x['D'] is None or x['F'] is None) else x['F']-x['D']
def M(x):
    g=G(x)
    return "UNCERTAIN" if (g=="" or x['H']=="L" or x['J']!=SEASON) else "OK"
gvals=[G(x) for x in recs]
chk("10.6 every QB delta is blank or exactly 0", all(g in ("",0) for g in gvals),
    str([ (recs[i]['map_ab'],g) for i,g in enumerate(gvals) if g not in ("",0)][:10]))
stat=collections.Counter(M(x) for x in recs)
say(f"  QB status: {dict(stat)}")
chk("10.7 every L-coded team resolves UNCERTAIN", all(M(x)=="UNCERTAIN" for x in recs if x['H']=="L"))
chk("10.8 no H-coded team resolves UNCERTAIN via a blank delta",
    all(M(x)=="OK" for x in recs if x['H']=="H"),
    str([x['map_ab'] for x in recs if x['H']=="H" and M(x)!="OK"]), blocking=False)
# ENGINE consumption
eng=wc["ENGINE"]
say(f"  ENGINE!M7  (QB adj)    = {ftext(eng['M7'].value)}")
say(f"  ENGINE!AE7 (QB status) = {ftext(eng['AE7'].value)}")
say(f"  ENGINE!AI7 (status)    = {str(ftext(eng['AI7'].value))[:220]}")
for cell in ("M7","AE7","AI7"):
    chk(f"10.9 ENGINE!{cell} identical to authoritative",
        ftext(wa['ENGINE'][cell].value)==ftext(eng[cell].value))
# market lines loaded?
ml=wc["MARKET LINES"]
loaded=sum(1 for r in range(1,ml.max_row+1) for c in range(1,ml.max_column+1)
           if isinstance(ml.cell(row=r,column=c).value,(int,float)))
say(f"  MARKET LINES numeric cells present: {loaded}")

# ============ CACHED VALUES ============
say("\n=== CACHED FORMULA VALUES (openpyxl round-trip risk) ===")
va=openpyxl.load_workbook(AUTH, data_only=True); vc=openpyxl.load_workbook(CAND, data_only=True)
def cached(wb):
    n=0
    for s in wb.worksheets:
        for row in s.iter_rows():
            for c in row:
                if c.value is not None: n+=1
    return n
ca, cc = cached(va), cached(vc)
say(f"  non-empty cells when read data_only: auth {ca} | cand {cc}")
if ca>0 and cc==0:
    warn("candidate carries NO cached formula results (openpyxl drops them on save). "
         "Excel/Sheets recalculate on open, but any consumer reading cached values sees blanks. "
         "This is inherited from v0.7.2 onward, not introduced by this phase.")
elif cc>0:
    say("  candidate retains cached values.")
else:
    say("  neither workbook carries cached values -> no regression.")

say("\n"+"#"*78)
say(f"# BLOCKERS: {len(BLOCKERS)}")
for b in BLOCKERS: say(f"#   - {b}")
say(f"# WARNINGS: {len(WARNINGS)}")
for w in WARNINGS: say(f"#   - {w}")
say("#"*78)
open(f"{ROOT}/workbook_v0.7.9_candidate/promotion_audit_log.txt","w").write("\n".join(OUT)+"\n")
sys.exit(1 if BLOCKERS else 0)

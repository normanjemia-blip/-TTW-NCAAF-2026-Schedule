"""Phase 7D.5 verification: v0.7.8 -> v0.7.9 diff, cumulative v0.6.2 -> v0.7.9 diff,
full regression battery, final inventory, backlog + audit-trail ledger, manifest."""
import openpyxl, hashlib, json, csv, datetime, collections
from openpyxl.worksheet.formula import ArrayFormula

ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
OLD=f"{ROOT}/workbook_v0.7.8_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.8_CANDIDATE.xlsx"
NEW=f"{ROOT}/workbook_v0.7.9_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.9_CANDIDATE.xlsx"
AUTH=f"{ROOT}/workbook_v0.6.2_deliverables/TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"
OUT=f"{ROOT}/workbook_v0.7.9_candidate"

def sha(p):
    h=hashlib.sha256()
    with open(p,"rb") as f:
        for b in iter(lambda:f.read(1<<20),b""): h.update(b)
    return h.hexdigest()
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
def norm(v): return v.text if isinstance(v,ArrayFormula) else v

wo=openpyxl.load_workbook(OLD); wn=openpyxl.load_workbook(NEW); wa=openpyxl.load_workbook(AUTH)
R=[]
def say(s): print(s); R.append(s)

def diff(a,b):
    d=[]; fd=[]
    for name in b.sheetnames:
        if name not in a.sheetnames: continue
        sa,sb=a[name],b[name]
        for r in range(1,max(sa.max_row,sb.max_row)+1):
            for c in range(1,max(sa.max_column,sb.max_column)+1):
                x=norm(sa.cell(row=r,column=c).value); y=norm(sb.cell(row=r,column=c).value)
                if x!=y:
                    addr=f"{openpyxl.utils.get_column_letter(c)}{r}"
                    d.append((name,addr,x,y))
                    if isf(sa.cell(row=r,column=c).value) or isf(sb.cell(row=r,column=c).value): fd.append((name,addr))
    return d,fd

# ---------- 1. structure ----------
say("="*78); say("1. STRUCTURE")
say(f"  sheets old/new: {len(wo.sheetnames)} / {len(wn.sheetnames)}  identical order: {wo.sheetnames==wn.sheetnames}")
vis_o={s.title:s.sheet_state for s in wo.worksheets}; vis_n={s.title:s.sheet_state for s in wn.worksheets}
say(f"  visibility identical: {vis_o==vis_n}")
fo=sum(1 for s in wo.worksheets for row in s.iter_rows() for c in row if isf(c.value))
fn=sum(1 for s in wn.worksheets for row in s.iter_rows() for c in row if isf(c.value))
fa=sum(1 for s in wa.worksheets for row in s.iter_rows() for c in row if isf(c.value))
say(f"  formula cells: v0.6.2 {fa} | v0.7.8 {fo} -> v0.7.9 {fn}  (delta this phase {fn-fo}; vs authoritative {fn-fa})")

# ---------- 2. diff ----------
say("="*78); say("2. CELL DIFF v0.7.8 -> v0.7.9")
diffs,formula_diffs=diff(wo,wn)
say(f"  changed cells: {len(diffs)}   FORMULA changes: {len(formula_diffs)}")
by_sheet=collections.Counter(d[0] for d in diffs)
for k,v in by_sheet.items(): say(f"    {k}: {v}")

CODE_CHANGE={14:"MIZ",65:"UNC",125:"UNLV"}
STAMP={6:"ALA",8:"AUB",9:"FLA",11:"UK",18:"TENN",21:"VAN",29:"NEB",68:"STAN",69:"SYR",
       84:"FAU",85:"MEM",89:"USF",91:"TULN",106:"BALL",109:"CMU",115:"TOL",117:"WMU",139:"USM"}
FIVE={78:"TXST",80:"WSU",87:"UNT",131:"GASO",137:"ODU"}
classified=[]
for name,addr,a,b in diffs:
    col=''.join(ch for ch in addr if ch.isalpha()); row=int(''.join(ch for ch in addr if ch.isdigit()))
    if name=="QB VALUES" and row in CODE_CHANGE:
        lab="NUMERICAL-CELL CONSISTENCY REPAIR" if col in ("D","F") else "VERIFIED QB CLASSIFICATION CHANGE"
    elif name=="QB VALUES" and row in STAMP: lab="F-7 AUDIT-TRAIL STAMP (verification date + evidence note)"
    elif name=="QB VALUES" and row in FIVE:  lab="FINAL-FIVE RE-VERIFICATION NOTE REFRESH"
    elif name in ("CHANGELOG","START HERE"): lab="VERSION OR CHANGELOG"
    else: lab="UNRELATED / UNAUTHORIZED / UNKNOWN"
    classified.append((name,addr,lab,a,b))
cnt=collections.Counter(x[2] for x in classified)
say("  classification:")
for k,v in sorted(cnt.items(), key=lambda kv:-kv[1]): say(f"    {k}: {v}")
unknown=[x for x in classified if x[2]=="UNRELATED / UNAUTHORIZED / UNKNOWN"]
say(f"  UNRELATED/UNAUTHORIZED/UNKNOWN: {len(unknown)}")
for x in unknown[:20]: say(f"      !! {x[0]}!{x[1]}  {x[3]!r} -> {x[4]!r}")

say("");say("  CUMULATIVE DIFF v0.6.2 AUTHORITATIVE -> v0.7.9 CANDIDATE")
cdiffs,cfd=diff(wa,wn)
csheets=collections.Counter(d[0] for d in cdiffs)
say(f"    changed cells: {len(cdiffs)}   FORMULA changes: {len(cfd)}")
for k,v in csheets.items(): say(f"      {k}: {v}")
say(f"    sheets touched vs authoritative: {sorted(csheets)}")

# ---------- 3. inventory ----------
say("="*78); say("3. QB VALUES INVENTORY (v0.7.9 FINAL)")
qb=wn["QB VALUES"]; tm=wn["TEAM MAP"]
inv=[]
for r in range(6,144):
    inv.append(dict(row=r,abbrev=tm.cell(row=r,column=1).value,team=tm.cell(row=r,column=2).value,
        baseline_qb=qb.cell(row=r,column=3).value, baseline_value=qb.cell(row=r,column=4).value,
        active_qb=qb.cell(row=r,column=5).value, active_value=qb.cell(row=r,column=6).value,
        confidence=qb.cell(row=r,column=8).value, source=qb.cell(row=r,column=9).value,
        reviewed_for_season=qb.cell(row=r,column=10).value,
        last_update=(qb.cell(row=r,column=11).value.date().isoformat()
                     if isinstance(qb.cell(row=r,column=11).value,datetime.datetime) else qb.cell(row=r,column=11).value),
        note=qb.cell(row=r,column=12).value))
say(f"  rows: {len(inv)}  unique abbrevs: {len(set(x['abbrev'] for x in inv))}  unique teams: {len(set(x['team'] for x in inv))}")
codes=collections.Counter(x["confidence"] for x in inv); say(f"  codes: {dict(codes)}")
say(f"  invalid codes: {[x['abbrev'] for x in inv if x['confidence'] not in ('H','M','L')]}")
blank=[x for x in inv if x["baseline_value"] is None and x["active_value"] is None]
zero =[x for x in inv if x["baseline_value"]==0 and x["active_value"]==0]
say(f"  blank numerical: {len(blank)}   zero numerical: {len(zero)}   other: {len(inv)-len(blank)-len(zero)}")
nonzero=[x for x in inv if (x["baseline_value"] not in (None,0)) or (x["active_value"] not in (None,0))]
say(f"  NONZERO QB VALUES: {len(nonzero)}")
lviol=[x["abbrev"] for x in inv if x["confidence"]=="L" and not (x["baseline_value"] is None and x["active_value"] is None)]
say(f"  L-coded with non-blank numerical cells: {len(lviol)} {lviol}")
SEASON=wn["SETTINGS"]["B3"].value
def G(x): return "" if (x["baseline_value"] in (None,"") or x["active_value"] in (None,"")) else x["active_value"]-x["baseline_value"]
def M(x):
    g=G(x); return "UNCERTAIN" if (g=="" or x["confidence"]=="L" or x["reviewed_for_season"]!=SEASON) else "OK"
st=collections.Counter(M(x) for x in inv)
say(f"  SETTINGS!B3={SEASON}   QB status: {dict(st)}   nonzero deltas: {sum(1 for x in inv if G(x) not in ('',0))}")

# ---------- 4. backlog + audit trail ----------
say("="*78); say("4. BACKLOG AND AUDIT TRAIL")
prev=json.load(open(f"{ROOT}/workbook_v0.7.8_candidate/qb_inventory_v078.json"))
LEDGER={r["abbrev"]:r["verification_status"] for r in prev["records"]}
for ab in list(CODE_CHANGE.values())+list(STAMP.values()): LEDGER[ab]="VERIFIED 7D.5"
for x in inv: x["verification_status"]=LEDGER[x["abbrev"]]
tier1=[x for x in inv if x["confidence"] in ("M","L")]
backlog=[x for x in inv if x["verification_status"].startswith("NOT VERIFIED")]
say(f"  Tier 1 population (M+L): {len(tier1)}")
say(f"  ledger: {dict(collections.Counter(x['verification_status'] for x in inv))}")
say(f"  >>> QB VERIFICATION BACKLOG: {len(backlog)} {[x['abbrev'] for x in backlog]}")
stale_t1=[x for x in tier1 if x["last_update"]=="2026-07-21"]
stale_h =[x for x in inv if x["confidence"]=="H" and x["last_update"]=="2026-07-21"]
say(f"  >>> AUDIT-TRAIL GAP (Tier-1 rows unstamped): {len(stale_t1)} {[x['abbrev'] for x in stale_t1]}")
say(f"  H-coded tier-2 rows still on the 2026-07-21 build stamp: {len(stale_h)} (never in Tier-1 scope; expected)")

# ---------- 5. regression battery ----------
say("="*78); say("5. REGRESSION BATTERY")
res=[]
def t(n,ok,detail=""):
    res.append(ok); say(f"  [{'PASS' if ok else 'FAIL'}] {n}{(' - '+detail) if detail else ''}")
t("1 no formula changes this phase", len(formula_diffs)==0, str(len(formula_diffs)))
t("2 no formula changes vs v0.6.2 AUTHORITATIVE", len(cfd)==0, str(len(cfd)))
t("3 formula count unchanged", fo==fn==fa, f"{fa} / {fo} / {fn}")
t("4 sheet count/order/visibility unchanged", wo.sheetnames==wn.sheetnames and vis_o==vis_n)
t("5 138 unique teams, no shifted rows", len(inv)==138 and len(set(x['team'] for x in inv))==138)
t("6 no invalid H/M/L codes", all(x["confidence"] in ("H","M","L") for x in inv))
t("7 zero nonzero QB values", len(nonzero)==0)
t("8 zero nonzero QB deltas", sum(1 for x in inv if G(x) not in ("",0))==0)
t("9 every L-coded team blank-gated", len(lviol)==0)
t("10 every L-coded team UNCERTAIN", all(M(x)=="UNCERTAIN" for x in inv if x["confidence"]=="L"))
t("11 TEAM RATINGS untouched (this phase and vs v0.6.2)", "TEAM RATINGS" not in by_sheet and "TEAM RATINGS" not in csheets)
t("12 ENGINE untouched (no spread movement, vs v0.6.2 too)", "ENGINE" not in by_sheet and "ENGINE" not in csheets)
t("13 SETTINGS untouched vs v0.6.2", "SETTINGS" not in csheets)
t("14 ADJUSTMENTS untouched vs v0.6.2", "ADJUSTMENTS" not in csheets)
t("15 MARKET LINES untouched vs v0.6.2", "MARKET LINES" not in csheets)
t("16 PRESEASON untouched vs v0.6.2 (source weights)", "PRESEASON" not in csheets)
t("17 IMPORT SCHEDULE untouched vs v0.6.2", "IMPORT SCHEDULE" not in csheets)
t("18 only QB VALUES + CHANGELOG + START HERE in diff", set(by_sheet)<={"QB VALUES","CHANGELOG","START HERE"}, str(sorted(by_sheet)))
t("19 cumulative diff confined to the same three sheets", set(csheets)<={"QB VALUES","CHANGELOG","START HERE"}, str(sorted(csheets)))
t("20 no unrelated/unauthorized/unknown cells", len(unknown)==0)
t("21 QB VALUES diff confined to the 26 intended rows",
  all(int(''.join(c for c in a if c.isdigit())) in {**CODE_CHANGE,**STAMP,**FIVE} for s,a,_,_ in diffs if s=="QB VALUES"))
t("22 formula columns A,B,G,M intact on all 138 rows",
  all(isf(qb.cell(row=r,column=c).value) for r in range(6,144) for c in (1,2,7,13)))
t("23 input columns hold constants only",
  all(not isf(qb.cell(row=r,column=c).value) for r in range(6,144) for c in (3,4,5,6,8,9,10,11,12)))
t("24 SETTINGS!B3 season = 2026", SEASON==2026)
t("25 SETTINGS!B6 HFA = 2.5", wn["SETTINGS"]["B6"].value==2.5)
t("26 SETTINGS!B11 BET toggle = N", wn["SETTINGS"]["B11"].value=="N")
t("27 QB VERIFICATION BACKLOG = 0", len(backlog)==0, str(len(backlog)))
t("28 AUDIT-TRAIL GAP = 0 (finding F-7 closed)", len(stale_t1)==0, str(len(stale_t1)))
t("29 every Tier-1 record carries a note", all(x["note"] for x in tier1))
miz=[x for x in inv if x["abbrev"]=="MIZ"][0]
t("30 MIZ M->H, zeros retained, status OK",
  miz["confidence"]=="H" and miz["baseline_value"]==0 and miz["active_value"]==0 and M(miz)=="OK")
unc=[x for x in inv if x["abbrev"]=="UNC"][0]
t("31 UNC M->L, values blank (were already blank), UNCERTAIN",
  unc["confidence"]=="L" and unc["baseline_value"] is None and M(unc)=="UNCERTAIN")
unlv=[x for x in inv if x["abbrev"]=="UNLV"][0]
t("32 UNLV M->L, zeros cleared, UNCERTAIN",
  unlv["confidence"]=="L" and unlv["baseline_value"] is None and unlv["active_value"] is None and M(unlv)=="UNCERTAIN")
t("33 ENGINE!AI status precedence unchanged vs v0.6.2", norm(wa["ENGINE"]["AI7"].value)==norm(wn["ENGINE"]["AI7"].value))
t("34 ENGINE!M (QB adj) unchanged vs v0.6.2", norm(wa["ENGINE"]["M7"].value)==norm(wn["ENGINE"]["M7"].value))
t("35 ENGINE!AE (QB status) unchanged vs v0.6.2", norm(wa["ENGINE"]["AE7"].value)==norm(wn["ENGINE"]["AE7"].value))
t("36 QB VALUES G formula unchanged vs v0.6.2", norm(wa["QB VALUES"]["G6"].value)==norm(qb["G6"].value))
t("37 QB VALUES M formula unchanged vs v0.6.2", norm(wa["QB VALUES"]["M6"].value)==norm(qb["M6"].value))
t("38 no market spreads loaded (pristine preseason state preserved)",
  "MARKET LINES" not in csheets)
say(f"\n  RESULT: {sum(res)}/{len(res)} PASS, {len(res)-sum(res)} FAIL")

# ---------- 6. manifest ----------
say("="*78); say("6. ZERO-CHANGE VERIFICATION / MANIFEST")
h_auth=sha(AUTH); h_old=sha(OLD); h_new=sha(NEW)
say(f"  v0.6.2 AUTHORITATIVE {h_auth}  UNCHANGED={h_auth=='bbb17b50fbfb728bea2a23d3d20771935cc61e238313a054473aafe1ca838efd'}")
say(f"  v0.7.8 CANDIDATE     {h_old}  UNCHANGED={h_old=='8f655e5e369a6a8c12fdb34f3309cff13a92c9310af6186b77081be4b3c389cb'}")
say(f"  v0.7.9 CANDIDATE     {h_new}  (new, FINAL)")
say(f"  PROJECT_MANIFEST.json match for v0.6.2: {h_auth in open(f'{ROOT}/PROJECT_MANIFEST.json').read()}")

summary=dict(generated="2026-08-04", phase="7D.5", counts=dict(codes), status=dict(st),
  blank_D=len(blank), zero_D=len(zero), nonzero_qb_values=len(nonzero),
  backlog_remaining=len(backlog), audit_trail_gap_tier1=len(stale_t1),
  h_coded_tier2_unstamped=len(stale_h), sha256=h_new, records=inv)
with open(f"{OUT}/qb_inventory_v079.json","w") as f: json.dump(summary,f,indent=1,default=str)
with open(f"{OUT}/qb_inventory_v079.csv","w",newline="") as f:
    w=csv.DictWriter(f,fieldnames=list(inv[0].keys())); w.writeheader()
    for x in inv: w.writerow(x)
for fn_,rows_ in ((f"{OUT}/diff_v078_to_v079.csv",classified),
                  (f"{OUT}/diff_v062_to_v079.csv",[(s,a,"CUMULATIVE",o,n) for s,a,o,n in cdiffs])):
    with open(fn_,"w",newline="") as f:
        w=csv.writer(f); w.writerow(["sheet","cell","classification","old","new"]); w.writerows(rows_)
with open(f"{OUT}/verification_log_v079.txt","w") as f: f.write("\n".join(R)+"\n")
say(f"\n  artifacts written to {OUT}")

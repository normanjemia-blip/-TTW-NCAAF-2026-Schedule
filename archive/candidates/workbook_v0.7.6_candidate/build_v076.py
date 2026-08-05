import shutil, os, datetime, openpyxl
from openpyxl.worksheet.formula import ArrayFormula
ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
SRC=f"{ROOT}/workbook_v0.7.5_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.5_CANDIDATE.xlsx"
OUT=f"{ROOT}/workbook_v0.7.6_candidate"; DST=f"{OUT}/TTW_NCAAF_Power_Ratings_2026_v0.7.6_CANDIDATE.xlsx"
DATE=datetime.datetime(2026,8,3)
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
os.makedirs(OUT,exist_ok=True); shutil.copyfile(SRC,DST)
wb=openpyxl.load_workbook(DST); qb=wb["QB VALUES"]; tm=wb["TEAM MAP"]
R={tm.cell(row=r,column=1).value:r for r in range(6,144)}
AQB,BV,AV,CONF,UPD,NOTE=5,4,6,8,11,12
def guard(r):
    for c in (AQB,BV,AV,CONF,UPD,NOTE): assert not isf(qb.cell(row=r,column=c).value), f"formula row{r}"

# ===== 1. ARKANSAS  M -> L  (DEFECT) + clear zeros =====
r=R["ARK"]; guard(r); assert qb.cell(row=r,column=CONF).value=="M" and qb.cell(row=r,column=BV).value==0
qb.cell(row=r,column=AQB).value="Open (KJ Jackson / AJ Hill)"
qb.cell(row=r,column=CONF).value="L"
qb.cell(row=r,column=BV).value=None; qb.cell(row=r,column=AV).value=None
qb.cell(row=r,column=UPD).value=DATE
qb.cell(row=r,column=NOTE).value=("DEFECT CORRECTED 2026-08-03: prior entry asserted KJ Jackson as likely starter (M). HC Ryan "
 "Silverfield (Whole Hog Sports, 2026-08-03) says the KJ Jackson vs AJ Hill battle 'could last until the Sept. 5 opener "
 "vs North Alabama'; he and OC Tim Cramsey called it 'wide open' in February. Jackson: RS-So, 4-star 2024, 33/54 441yds "
 "3TD 0INT in 2025. Hill: RS-Fr, Memphis transfer (followed Silverfield), 4-star 2025, two years in Cramsey's system. "
 "Genuine open competition -> L. Numerical zeros CLEARED to blank per the approved Akron consistency methodology.")
print(f"ARK  row{r}: M->L, D/F cleared")

# ===== 2. UCONN  M -> L  (DEFECT) + clear zeros =====
r=R["CONN"]; guard(r); assert qb.cell(row=r,column=CONF).value=="M" and qb.cell(row=r,column=BV).value==0
qb.cell(row=r,column=AQB).value="Unverified (room in flux after HC departure)"
qb.cell(row=r,column=CONF).value="L"
qb.cell(row=r,column=BV).value=None; qb.cell(row=r,column=AV).value=None
qb.cell(row=r,column=UPD).value=DATE
qb.cell(row=r,column=NOTE).value=("DEFECT CORRECTED 2026-08-03: prior entry named Tucker McDonald (M) - NOT corroborated by any located "
 "source. Verified instead: Jim Mora LEFT UConn for the Colorado State head job; UConn's 2026 QB commit departed after "
 "Mora's exit; 2025 starter Joe Fagnano (2,529 yds, 22 TD through nine games) was a senior and is gone. Projected starter "
 "CANNOT BE VERIFIED -> L per the standard. Numerical zeros CLEARED to blank per the Akron methodology. RECHECK: UConn "
 "athletics depth chart or new-HC presser naming QB1.")
print(f"CONN row{r}: M->L, D/F cleared")

# ===== 3. HOUSTON  M -> H  (upgrade; zeros retained) =====
r=R["HOU"]; guard(r); assert qb.cell(row=r,column=CONF).value=="M" and qb.cell(row=r,column=BV).value==0
qb.cell(row=r,column=CONF).value="H"
qb.cell(row=r,column=UPD).value=DATE
qb.cell(row=r,column=NOTE).value=("VERIFIED 2026-08-03: Conner Weigman is Houston's RETURNING starter entering his second season as the "
 "signal-caller; confirmed on the official University of Houston roster; publicly announced his return for 2026; coach "
 "praise reported for his offseason bounce-back; NO competition reported in any located source. Established healthy "
 "returning starter with no credible competition -> H (not a media projection). Zero initialization RETAINED (correct "
 "for H under the preseason baseline-delta method); status remains OK.")
print(f"HOU  row{r}: M->H, zeros retained")

# ===== 4. MISSISSIPPI STATE  M -> H  (upgrade; zeros retained) =====
r=R["MSST"]; guard(r); assert qb.cell(row=r,column=CONF).value=="M" and qb.cell(row=r,column=BV).value==0
qb.cell(row=r,column=CONF).value="H"
qb.cell(row=r,column=UPD).value=DATE
qb.cell(row=r,column=NOTE).value=("VERIFIED 2026-08-03: Kamario Taylor arrived at SEC Media Days as Mississippi State's 'unquestioned "
 "starter - no competition, no committee, no asterisk'; described as ENTRENCHED, with only the No.2 job open in fall camp. "
 "Confirmed on the official hailstate.com roster and in official MSU releases; new QB coach Kevin Johns calls him 'very, "
 "very special.' Production as a true freshman: 629 pass yds, 458 rush yds, 173 rush vs Ole Miss in his first career "
 "start. Settled starter -> H. Zero initialization RETAINED; status remains OK.")
print(f"MSST row{r}: M->H, zeros retained")

# ===== 5-7. CONFIRMED (code unchanged; entry refinement + date/note) =====
r=R["DUKE"]; guard(r); assert qb.cell(row=r,column=CONF).value=="M"
qb.cell(row=r,column=AQB).value="Walker Eget (expected Week 1 starter; 5-way competition)"
qb.cell(row=r,column=UPD).value=DATE
qb.cell(row=r,column=NOTE).value=("VERIFIED 2026-08-03 - 7D.3 CONFLICT RESOLVED, candidate entry CONFIRMED: Walker Eget is on the official "
 "Duke roster (goduke.com), a San Jose State transfer who committed in Jan 2026 (Duke Chronicle), and is expected to be "
 "Duke's Week 1 starter after Darian Mensah left for Miami. Manny Diaz expects him to be a full participant in fall camp. "
 "He competes with transfers Blaine Hipa and Ari Patu plus returners Terry Walker III and Dan Mahan. Career at SJSU: "
 "5,555 yds, 30 TD, 19 INT over two seasons starting. Clear leader with real competition -> M retained (not H).")
print(f"DUKE row{R['DUKE']}: M confirmed (conflict resolved)")

r=R["IOWA"]; guard(r); assert qb.cell(row=r,column=CONF).value=="L"
qb.cell(row=r,column=AQB).value="Open (Jeremy Hecklinski / Hank Brown)"
qb.cell(row=r,column=UPD).value=DATE
qb.cell(row=r,column=NOTE).value=("VERIFIED 2026-08-03: HC Kirk Ferentz has NOT decided between Jeremy Hecklinski and Hank Brown entering fall "
 "camp; coaches have consistently SPLIT first-team reps with 'little separation'; Ferentz says a decision may not come "
 "until August and 'may be in-season.' Both spent 2025 learning OC Tim Lester's system behind Mark Gronowski. Genuine open "
 "competition -> L retained; entry refined from a single name to the actual two-man race. Values remain blank; UNCERTAIN.")
print(f"IOWA row{R['IOWA']}: L confirmed, entry refined")

r=R["KAN"]; guard(r); assert qb.cell(row=r,column=CONF).value=="L"
qb.cell(row=r,column=AQB).value="Cole Ballard (leader; Isaiah Marshall competing)"
qb.cell(row=r,column=UPD).value=DATE
qb.cell(row=r,column=NOTE).value=("VERIFIED 2026-08-03: Cole Ballard 'MAY BE' Kansas's QB1 (SI headline wording) replacing Jalon Daniels, with "
 "real competition from Isaiah Marshall - Ballard the better thrower, Marshall the better runner; Kansas QBs described as "
 "'totally fine with competition heading into camp.' Ballard confirmed on the official kuathletics.com roster. No naming "
 "by Lance Leipold located -> L retained. This confirms the Phase 7D.3 decision NOT to upgrade on a national "
 "projected-starter list alone. Values remain blank; UNCERTAIN.")
print(f"KAN  row{R['KAN']}: L confirmed, entry refined")

# ===== CHANGELOG + banner =====
cl=wb["CHANGELOG"]; last=0
for rr in range(1,cl.max_row+1):
    if any(cl.cell(row=rr,column=c).value not in (None,"") for c in range(1,7)): last=rr
rows=[("v0.7.6","2026-08-03",
 "Phase 7D.3A BATCH 2 STRAGGLER RESOLUTION: all SEVEN remaining Power-conference/Independent records resolved. TWO "
 "DEFECTS FOUND: Arkansas M->L (HC Silverfield says the KJ Jackson vs AJ Hill battle could last to the Sept 5 opener; "
 "'wide open' since February) and UConn M->L (named starter Tucker McDonald uncorroborated; Jim Mora left for Colorado "
 "State, the 2026 QB commit departed, and 2025 starter Joe Fagnano graduated - projected starter cannot be verified). "
 "Both had zero numerical inputs CLEARED to blank per the approved Akron consistency methodology.",
 "Batch 2 straggler resolution"),
 ("v0.7.6","2026-08-03",
 "Phase 7D.3A UPGRADES AND CONFIRMATIONS: Houston M->H (Conner Weigman, returning starter entering his second season, "
 "official roster, announced return, no competition reported) and Mississippi State M->H (Kamario Taylor, 'unquestioned "
 "starter - no competition, no committee, no asterisk' at SEC Media Days; only the No.2 job is open). Both RETAIN zero "
 "initialization, correct for H. Duke CONFIRMED at M and the 7D.3 conflict RESOLVED - Walker Eget is on the official "
 "roster and is the expected Week 1 starter within a five-way competition. Iowa and Kansas CONFIRMED at L with entries "
 "refined to their actual competitions.","Batch 2 straggler resolution"),
 ("v0.7.6","2026-08-03",
 "Phase 7D.3A RESULT: Batch 2 is now COMPLETE - all 16 Power-conference/Independent teams verified. Backlog 38 -> 31, "
 "consisting ONLY of Group of Five teams (Pac-12 6, American 6, Mountain West 6, Sun Belt 6, CUSA 4, MAC 3). Counts: "
 "63 H / 43 M / 32 L; status 102 OK / 36 UNCERTAIN; 0 nonzero deltas. Promotion remains DEFERRED.",
 "Batch 2 completion")]
for i,(v,d,c,rn) in enumerate(rows):
    rr=last+1+i
    cl.cell(row=rr,column=1).value=v; cl.cell(row=rr,column=2).value=d
    cl.cell(row=rr,column=3).value=c; cl.cell(row=rr,column=4).value=rn
wb["START HERE"]["A1"].value=("TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.7.6 CANDIDATE — Batch 2 COMPLETE: 7 stragglers "
 "resolved, 2 defects corrected (Arkansas, UConn), 2 upgrades (Houston, Mississippi State); backlog 31, Group of Five "
 "only; NOT AUTHORITATIVE, NOT PROMOTED)")
wb.save(DST)
print(f"\nCHANGELOG rows {last+1}-{last+len(rows)}; saved {DST}")

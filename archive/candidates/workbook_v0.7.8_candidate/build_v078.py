import shutil, os, datetime, openpyxl
from openpyxl.worksheet.formula import ArrayFormula
ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
SRC=f"{ROOT}/workbook_v0.7.7_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.7_CANDIDATE.xlsx"
OUT=f"{ROOT}/workbook_v0.7.8_candidate"; DST=f"{OUT}/TTW_NCAAF_Power_Ratings_2026_v0.7.8_CANDIDATE.xlsx"
D8=datetime.datetime(2026,8,4)          # owner timezone America/New_York
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
os.makedirs(OUT,exist_ok=True); shutil.copyfile(SRC,DST)
wb=openpyxl.load_workbook(DST); qb=wb["QB VALUES"]; tm=wb["TEAM MAP"]
R={tm.cell(row=r,column=1).value:r for r in range(6,144)}
AQB,BV,AV,CONF,UPD,NOTE=5,4,6,8,11,12
def guard(r):
    for c in (AQB,BV,AV,CONF,UPD,NOTE): assert not isf(qb.cell(row=r,column=c).value), f"formula row{r}"
def setrec(ab, aqb=None, code=None, clear_vals=False, note=None, expect_code=None, expect_row=None):
    r=R[ab]; guard(r)
    if expect_row: assert r==expect_row, f"{ab} row {r} != expected {expect_row}"
    if expect_code: assert qb.cell(row=r,column=CONF).value==expect_code, f"{ab} expected {expect_code}"
    if aqb is not None: qb.cell(row=r,column=AQB).value=aqb
    if code is not None: qb.cell(row=r,column=CONF).value=code
    if clear_vals:
        assert qb.cell(row=r,column=BV).value==0 and qb.cell(row=r,column=AV).value==0
        qb.cell(row=r,column=BV).value=None; qb.cell(row=r,column=AV).value=None
    qb.cell(row=r,column=UPD).value=D8
    qb.cell(row=r,column=NOTE).value=note
    print(f"  {ab:6} row{r:>4} code={qb.cell(row=r,column=CONF).value} D={qb.cell(row=r,column=BV).value!r} F={qb.cell(row=r,column=AV).value!r}")

print("=== PAC-12 ===")
setrec("TXST", expect_code="M", expect_row=78,
 aqb="Brad Jackson (returning starter)",
 note=("VERIFIED 2026-08-04: Brad Jackson CONFIRMED as Texas State's returning starter for the Pac-12 debut. He started "
  "2025 as a redshirt freshman (3,050 pass yds, 18 TD, 7 INT, 71.5% comp; 16-17 rush TD, school single-season rushing-TD "
  "record for a QB; 3,968 total yds, 7th nationally), was a third-team all-conference pick, and ANNOUNCED HIS RETURN on "
  "2025-12-06 alongside WRs Beau Sparks and Chris Dawn Jr. Coverage states Texas State 'returns a starting quarterback for "
  "the first time under Kinne'; The Athletic ranked him the top Group of Five quarterback; he spoke to media at 2026 fall "
  "camp as the quarterback. NO competition reported. H CONSIDERED AND DECLINED: no official 2026 naming or depth chart has "
  "been published, and the WMU (Broc Lowry) / Toledo (John Alan Richter) precedent in this project confirms a returning "
  "starter at M rather than upgrading. -> M retained; zeros retained (correct for M); status OK. "
  "RECHECK: official Texas State depth chart before Week 1."))
setrec("WSU", expect_code="M", expect_row=80, code="L", clear_vals=True,
 aqb="Open (Caden Pinnick / Owen Eshelman / Julian Dugger)",
 note=("DEFECT CORRECTED 2026-08-04: prior entry asserted Caden Pinnick as the projected transfer QB1 (M). Current "
  "team-specific beat coverage contradicts a settled leader. HeraldNet 2026-07-30 headline: WSU 'set to open fall camp "
  "with WIDE-OPEN QB competition'; The Columbian 2026-08-01: 'Washington State football heads into fall with quarterback "
  "situation UNRESOLVED.' First-year HC Kirby Moore declined to name a preference - 'That's gonna continue through the "
  "summer... The team will decide who the quarterback is in terms of what happens on the field.' The Spokesman-Review beat "
  "reports Pinnick 'WASN'T ABLE TO CREATE MEANINGFUL SEPARATION FROM ESHELMAN' in spring (his spring game ended in a "
  "pick-6), and that Moore's staff may keep the decision in-house until kickoff. SOURCE CONFLICT: CougCenter calls Pinnick "
  "the 'early frontrunner', but that is a media characterization contradicted by the beat writer's direct observation and "
  "by the head coach's non-commitment; resolved toward the primary evidence. Note the room shrank further - 2025 Coug "
  "Jaxon Potter transferred OUT to Old Dominion. Three-way competition with no leader established -> M downgraded to L. "
  "Numerical zeros CLEARED to blank per the approved Akron/App State consistency methodology; status becomes UNCERTAIN. "
  "RECHECK: WSU depth chart or a Moore naming; camp runs from 2026-08-06."))

print("=== AMERICAN ===")
setrec("UNT", expect_code="M", expect_row=87,
 aqb="Tayven Jackson (slight lead; Ditta / Jimerson competing)",
 note=("VERIFIED 2026-08-04: entry CONFIRMED NOT STALE - Tayven Jackson is on North Texas's roster and leads the race. "
  "COACHING CHANGE VERIFIED: Eric Morris left for Oklahoma State; NEAL BROWN is North Texas's head coach and rebuilt the "
  "roster with 75-90+ new players. HC Neal Brown said on the American Kickoff broadcast (2026-07-29) that senior UCF "
  "transfer Tayven Jackson 'NOW HAS A SLIGHT LEAD' in an ongoing THREE-MAN race over sophomore East Carolina transfer "
  "Chaston Ditta and sophomore returner Chris Jimerson Jr.; Brown did NOT declare a starter and the battle carries into "
  "fall camp. Jackson's 2025 at UCF: 2,151 pass yds, 10 TD, 8 INT, 85 rush yds, 1 rush TD. Clear leader stated by the "
  "head coach, no naming -> M retained (not H); entry refined to name the competitors. Zeros retained (correct for M); "
  "status OK. RECHECK: official North Texas depth chart."))

print("=== SUN BELT ===")
setrec("GASO", expect_code="L", expect_row=131, code="M",
 aqb="Max Johnson (leader; Turner Helton competing)",
 note=("STALE ENTRY CORRECTED 2026-08-04: prior entry read 'Open (Weston Bryan / Turner Helton)' and OMITTED MAX JOHNSON "
  "ENTIRELY. Johnson signed with Georgia Southern out of the portal on 2026-01-11 from North Carolina (previously LSU "
  "2020-21, Texas A&M 2022-23), and at SUN BELT MEDIA DAYS (2026-07-15) HC CLAY HELTON SAID 'Max Johnson is expected to be "
  "the starter if the season began today,' citing experience and high-level production. Turner Helton (2nd year in the "
  "program, Western Kentucky transfer, the head coach's son; 14/26, 74 yds, 1 TD in 3 appearances in 2025) is the other "
  "player with a shot at the 2026-09-05 opener; both took first-team reps in the spring game. Documented leader per a "
  "direct head-coach statement, no official naming -> L upgraded to M (NOT H; parallel to the Oregon State treatment). "
  "Per the L->M rule numerical values REMAIN BLANK (valuation unresolved), so status correctly stays UNCERTAIN. "
  "OPEN RISK: Johnson's durability - season-ending injuries at Texas A&M in 2022 and 2023 and a broken leg in UNC's 2024 "
  "opener. RECHECK: Georgia Southern depth chart before 2026-09-05."))
setrec("ODU", expect_code="M", expect_row=137,
 aqb="Quinn Henicle (leader; Potter / Huff competing)",
 note=("VERIFIED 2026-08-04: Quinn Henicle CONFIRMED as Old Dominion's leader in a genuine three-way competition. The job "
  "opened when 2025 starter Colton Joseph transferred to Wisconsin. Henicle (RS-So) started the 2025 Cure Bowl in Joseph's "
  "place and was named GAME MVP in a 24-10 win over South Florida (11/25, 127 yds; 24 carries, 107 yds, 2 rush TD; 51-yd "
  "TD run); he is 2-0 as a starter. Beat coverage says Henicle 'HAS A LEG UP' and is 'the front runner to land the QB1 "
  "role', while also reporting he 'remains enmeshed in a three-way battle' with RS-Jr JAXON POTTER (Washington State "
  "transfer, announced 2026-01-13) and RS-Fr Ryan Huff, both getting significant snaps. Henicle himself (2026-04-03): "
  "\"I'm competing like I'm trying to earn this job.\" HC Ricky Rahne (7th season, off a 10-3 year) has NOT named a "
  "starter - ODU's official Sun Belt Media Day release does not address the position. Documented leader backed by on-field "
  "production rather than name recognition, no naming -> M retained (not H); entry refined to name the competitors. Zeros "
  "retained (correct for M); status OK. RECHECK: official Old Dominion depth chart."))

# --- CHANGELOG + banner ---
cl=wb["CHANGELOG"]; last=0
for rr in range(1,cl.max_row+1):
    if any(cl.cell(row=rr,column=c).value not in (None,"") for c in range(1,7)): last=rr
rows=[("v0.7.8","2026-08-04",
 "Phase 7D.4A FINAL FIVE QB RESOLUTION: Texas State, Washington State, North Texas, Georgia Southern and Old Dominion "
 "resolved by team-specific sourcing. BACKLOG 5 -> 0. THREE CONFIRMED at M with refined entries: Texas State (Brad "
 "Jackson, returning 2025 starter, 3,050 pass yds / 18 TD / 71.5%, top-G5 QB per The Athletic, no competition), North "
 "Texas (HC Neal Brown on the 2026-07-29 American Kickoff broadcast: Tayven Jackson 'now has a slight lead' in a "
 "three-man race with Chaston Ditta and Chris Jimerson Jr.; coaching change from Eric Morris to Neal Brown verified) and "
 "Old Dominion (Quinn Henicle, Cure Bowl MVP and 2-0 as a starter, 'has a leg up' in a three-way battle with Jaxon Potter "
 "and Ryan Huff).",
 "Final five QB resolution"),
 ("v0.7.8","2026-08-04",
 "Phase 7D.4A DEFECTS: (1) WASHINGTON STATE M -> L - prior entry asserted Caden Pinnick as projected QB1, but HeraldNet "
 "(2026-07-30) reports a 'wide-open QB competition', The Columbian (2026-08-01) calls the situation 'unresolved', HC "
 "Kirby Moore declined to name a preference, and the Spokesman-Review beat reports Pinnick 'wasn't able to create "
 "meaningful separation from Eshelman'. Numerical zeros CLEARED to blank. (2) GEORGIA SOUTHERN stale entry L -> M - prior "
 "entry read 'Open (Weston Bryan / Turner Helton)' and omitted Max Johnson entirely; HC Clay Helton at Sun Belt Media "
 "Days (2026-07-15) said Johnson 'is expected to be the starter if the season began today'. Values remain blank per the "
 "L->M rule.",
 "Final five QB resolution"),
 ("v0.7.8","2026-08-04",
 "Phase 7D.4A RESULT: QB verification backlog = 0. Counts 64 H / 43 M / 31 L (net unchanged - the WSU M->L and GASO L->M "
 "changes offset); status 100 OK / 38 UNCERTAIN; blank 38 / zero 100; 0 nonzero QB values; 0 formula changes. Cumulative "
 "defect pattern holds: every classification defect found in this project (Akron, Arkansas, UConn, Buffalo, Northern "
 "Illinois, Appalachian State, Washington State) was an unsupported M or a named starter the evidence did not support - "
 "never a team rated too uncertain.","Final five batch result")]
for i,(v,d,c,rn) in enumerate(rows):
    rr=last+1+i
    cl.cell(row=rr,column=1).value=v; cl.cell(row=rr,column=2).value=d
    cl.cell(row=rr,column=3).value=c; cl.cell(row=rr,column=4).value=rn
wb["START HERE"]["A1"].value=("TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.7.8 CANDIDATE — final five resolved, QB "
 "verification backlog = 0; Washington State M->L and Georgia Southern stale entry corrected; "
 "NOT AUTHORITATIVE, NOT PROMOTED)")
wb.save(DST)
print(f"\nCHANGELOG rows {last+1}-{last+len(rows)}; saved {DST}")

import shutil, os, datetime, openpyxl
from openpyxl.worksheet.formula import ArrayFormula
ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
SRC=f"{ROOT}/workbook_v0.7.4_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.4_CANDIDATE.xlsx"
OUT=f"{ROOT}/workbook_v0.7.5_candidate"; DST=f"{OUT}/TTW_NCAAF_Power_Ratings_2026_v0.7.5_CANDIDATE.xlsx"
DATE=datetime.datetime(2026,8,3)
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
os.makedirs(OUT,exist_ok=True); shutil.copyfile(SRC,DST)
wb=openpyxl.load_workbook(DST); qb=wb["QB VALUES"]; tm=wb["TEAM MAP"]
R={tm.cell(row=r,column=1).value:r for r in range(6,144)}
C={'aqb':5,'conf':8,'upd':11,'notes':12}
def guard(r):
    for k,c in C.items(): assert not isf(qb.cell(row=r,column=c).value), f"formula at row{r}"

# --- CONFIRMED (date + truthful note only; codes unchanged) ---
CONF={
 "MINN":("M","VERIFIED 2026-08-03: Drake Lindsey returns as sophomore starter after a record-setting freshman year (62.8% comp, 18 pass TD, 6 INT, 4 rush TD). Returning starter, no credible competition reported; M retained pending formal confirmation."),
 "MSU": ("M","VERIFIED 2026-08-03: Alessio Milivojevic returns, having taken over from Aidan Chiles in 2025. Returning starter; M retained."),
 "PUR": ("M","VERIFIED 2026-08-03: Ryan Browne returns at QB; production tied to his rushing usage. Returning starter, unproven as a full-season answer; M retained."),
 "WISC":("M","VERIFIED 2026-08-03: Colton Joseph listed as Wisconsin's starting quarterback; M retained pending formal announcement."),
 "AZST":("M","VERIFIED 2026-08-03: Cutter Boley confirmed as Arizona State's projected starter in current Big 12 starter rankings; M retained (no official naming located)."),
 "WVU": ("M","VERIFIED 2026-08-03: Michael Hawkins Jr. confirmed as West Virginia's projected starter; M retained (no official naming located)."),
 "UVA": ("M","VERIFIED 2026-08-03: Beau Pribula confirmed to 'get the nod' at Virginia per PFF ACC QB projections (76.1 PFF grade). M retained pending official announcement."),
 "WAKE":("M","VERIFIED 2026-08-03: Gio Lopez confirmed transferring to Wake Forest after 11 starts at North Carolina (73.1 PFF grade, 74.1% adjusted completion). Transfer expected to start; M retained."),
}
for ab,(code,note) in CONF.items():
    r=R[ab]; guard(r)
    assert qb.cell(row=r,column=C['conf']).value==code, f"{ab} code mismatch"
    qb.cell(row=r,column=C['upd']).value=DATE
    qb.cell(row=r,column=C['notes']).value=note
    print(f"  CONFIRMED {ab:5} row{r:4} code={code}")

# --- RECLASSIFICATION: Rutgers L -> M ---
r=R["RUTG"]; guard(r)
assert qb.cell(row=r,column=C['conf']).value=="L"
assert qb.cell(row=r,column=4).value is None and qb.cell(row=r,column=6).value is None, "RUTG values expected blank"
qb.cell(row=r,column=C['aqb']).value="Dylan Lonergan (likely starter; AJ Surace competing)"
qb.cell(row=r,column=C['conf']).value="M"
qb.cell(row=r,column=C['upd']).value=DATE
qb.cell(row=r,column=C['notes']).value=("VERIFIED 2026-08-03: Dylan Lonergan (Boston College transfer, replacing Athan Kaliakmanis) is "
 "described as 'the likely starter, but Rutgers likes AJ Surace, too' - a clear leader with real competition -> M. "
 "Mixed 2025 record (upside vs Michigan State/Stanford/Georgia Tech; benched twice, 1 pass TD and 6 turnovers in six "
 "other FBS games) supports M rather than H. Per 7D.3 methodology, numerical values REMAIN BLANK on an L->M change "
 "(valuation unresolved), so status correctly stays UNCERTAIN.")
print(f"  RECLASSIFIED RUTG row{r} L -> M (values left blank)")

# --- CHANGELOG + banner ---
cl=wb["CHANGELOG"]; last=0
for rr in range(1,cl.max_row+1):
    if any(cl.cell(row=rr,column=c).value not in (None,"") for c in range(1,7)): last=rr
rows=[("v0.7.5","2026-08-03",
 "Phase 7D.3 BATCH 2 (Power conferences + Independent): 16 teams in scope (Big Ten 6, Big 12 4, ACC 3, SEC 2, "
 "Independent 1). 9 verified - 8 records CONFIRMED accurate (Minnesota, Michigan State, Purdue, Wisconsin, Arizona "
 "State, West Virginia, Virginia, Wake Forest) and 1 RECLASSIFIED: Rutgers L -> M (Dylan Lonergan described as the "
 "likely starter with AJ Surace competing = clear leader with competition). No new numerical QB value populated; "
 "Rutgers values remain BLANK per the L->M rule, so its status stays UNCERTAIN.","Batch 2 verification"),
 ("v0.7.5","2026-08-03",
 "Phase 7D.3 SCOPE DISCLOSURE: 7 of the 16 Batch 2 teams could NOT be verified - Iowa, Kansas, Houston, Duke, "
 "Arkansas, Mississippi State, UConn. Their last-verified dates were deliberately NOT refreshed (still 2026-07-21). "
 "Duke is a specific gap: reporting confirms Darian Mensah left for Miami and that Duke was 'in a pinch' at QB, but "
 "no source located names Walker Eget (the candidate entry) as the replacement. Backlog: 47 -> 38 (7 P4/Independent "
 "+ 31 Group of Five). Promotion: DEFER.","Batch 2 scope disclosure")]
for i,(v,d,c,rn) in enumerate(rows):
    rr=last+1+i
    cl.cell(row=rr,column=1).value=v; cl.cell(row=rr,column=2).value=d
    cl.cell(row=rr,column=3).value=c; cl.cell(row=rr,column=4).value=rn
wb["START HERE"]["A1"].value=("TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.7.5 CANDIDATE — Batch 2 P4+Independent: 8 "
 "verified, Rutgers L->M; QB backlog 38 of 54 still unverified; NOT AUTHORITATIVE, NOT PROMOTED)")
wb.save(DST)
print(f"\nCHANGELOG rows {last+1}-{last+len(rows)}; saved {DST}")

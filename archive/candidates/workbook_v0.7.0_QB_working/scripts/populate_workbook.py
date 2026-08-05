import json, datetime, openpyxl
from openpyxl.worksheet.formula import ArrayFormula

WB="TTW_NCAAF_Power_Ratings_2026_v0.7.0_QB_WORKING.xlsx"
RD=datetime.date(2026,7,21)
rows=[json.loads(l) for l in open('qb_research.jsonl')]
teams=json.load(open('teams_138.json'))
row_of={t['abbrev']:t['row'] for t in teams}

wb=openpyxl.load_workbook(WB, data_only=False)
qb=wb['QB VALUES']

def note_for(r):
    parts=[r['status_type']]
    if r.get('prev_school'): parts.append(f"prev: {r['prev_school']}")
    if r['confidence']=='L':
        parts.append("OPEN COMPETITION — candidates: "+"; ".join(r.get('candidates') or []))
        parts.append("Baseline QB left blank (starter identity not established)")
    if r.get('stats_2025'): parts.append(f"2025: {r['stats_2025']}")
    if r.get('injury_eligibility'): parts.append(f"NOTE: {r['injury_eligibility']}")
    parts.append(r['confirmed_level'])
    parts.append("VALUE PENDING RUBRIC APPROVAL")
    return " | ".join(str(p) for p in parts if p)

changed=[]
def setc(row,col,val):
    cell=qb.cell(row=row,column=col)
    cell.value=val
    changed.append((f"QB VALUES!{cell.coordinate}", val))

for r in rows:
    rr=row_of[r['abbrev']]
    src=r['sources'][0]
    setc(rr,5, r['active_qb'])                                   # E Active QB
    if r['confidence'] in ('H','M'):
        setc(rr,3, r['active_qb'])                               # C Baseline QB = projected starter (H/M only)
    # D (4) Baseline value + F (6) Active value: intentionally left BLANK
    setc(rr,8, r['confidence'])                                  # H Confidence
    setc(rr,9, f"{src['name']} ({src['url']})")                  # I Source
    setc(rr,10, 2026)                                            # J Reviewed for season
    setc(rr,11, RD)                                              # K Last update
    setc(rr,12, note_for(r))                                     # L Notes

# Banner (START HERE!A1)
sh=wb['START HERE']
old_banner=sh['A1'].value
sh['A1'].value=("TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.7.0 QB WORKING — QB VALUES RESEARCH "
                "POPULATED; NUMERICAL VALUES PENDING RUBRIC APPROVAL — NOT AUTHORITATIVE)")
changed.append(("START HERE!A1", sh['A1'].value))

# CHANGELOG entry
cl=wb['CHANGELOG']
row=6
while cl.cell(row=row,column=1).value not in (None,""): row+=1
entries=[
 ("v0.7.0","2026-07-21",
  "Phase 8 (WORKING FILE ONLY — not authoritative): automated 2026 QB research for all 138 FBS teams from current sourced reporting (official athletics, beat writers, ESPN, SI, Athlon, On3/247, Underdog Dynasty, etc.). Populated QB VALUES columns E (Active QB = projected 2026 starter), C (Baseline QB = projected starter for H/M teams; blank for open competitions), H (Confidence H/M/L = certainty of starter identification), I (Source), J (Reviewed for season = 2026), K (Last update), L (Notes).",
  "Phase 8 - QB research populated (non-numeric)"),
 ("v0.7.0","2026-07-21",
  "Baseline value (D) and Active value (F) left BLANK: no approved numerical QB-value rubric exists in project architecture docs (confirmed v0.4/v0.5.1 findings). A proposed valuation rubric accompanies this build for approval BEFORE any numbers are entered. QB delta (G) therefore stays blank and all 138 teams remain QB UNCERTAIN by design (status column M unchanged).",
  "Phase 8 - values deferred pending rubric"),
 ("v0.7.0","2026-07-21",
  "Confidence distribution: 61 High, 45 Medium, 32 Low (32 unresolved competitions documented, none forced). Authoritative v0.6.2 XLSX and both native Google Sheets (v0.6.2 1H4XBJ..., v0.6.1 1EITbP...) untouched. No market lines, adjustments, ratings, HFA, thresholds, formulas, formatting, or tab structure changed.",
  "Phase 8 - summary + scope confirmation"),
]
for i,(v,d,desc,note) in enumerate(entries):
    rr=row+i
    cl.cell(row=rr,column=1,value=v); cl.cell(row=rr,column=2,value=d)
    cl.cell(row=rr,column=3,value=desc); cl.cell(row=rr,column=4,value=note)
    for c in range(1,5): changed.append((f"CHANGELOG!{cl.cell(row=rr,column=c).coordinate}", None))

wb.save(WB)
json.dump({"changed_cells":[c[0] for c in changed],"count":len(changed)}, open('changed_cells.json','w'), indent=1)
print("Populated. Total changed cells:", len(changed))
print("QB VALUES cells changed:", sum(1 for c in changed if c[0].startswith('QB VALUES')))
print("CHANGELOG rows added:", len(entries), "| banner updated: 1")

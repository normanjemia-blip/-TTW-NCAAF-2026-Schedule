import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
A="v0.6.2_source.xlsx"; B="TTW_NCAAF_Power_Ratings_2026_v0.7.0_QB_WORKING.xlsx"
def norm(v):
    if isinstance(v,ArrayFormula): return ("ARR",v.text)
    return v
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
a=openpyxl.load_workbook(A,data_only=False); b=openpyxl.load_workbook(B,data_only=False)
fails=[]
def chk(l,ok,d=""):
    print(("PASS " if ok else "FAIL ")+l+(f"  {d}" if d else "")); 
    if not ok: fails.append(l)

chk("21 sheets", len(b.sheetnames)==21, str(len(b.sheetnames)))
chk("sheet order+states unchanged", [(s,a[s].sheet_state) for s in a.sheetnames]==[(s,b[s].sheet_state) for s in b.sheetnames])
chk("recalc on open preserved", a.calculation.fullCalcOnLoad==b.calculation.fullCalcOnLoad)

# diff
diffs={}; 
for s in a.sheetnames:
    wa,wb2=a[s],b[s]; mr=max(wa.max_row,wb2.max_row); mc=max(wa.max_column,wb2.max_column); n=0
    for r in range(1,mr+1):
        for c in range(1,mc+1):
            if norm(wa.cell(row=r,column=c).value)!=norm(wb2.cell(row=r,column=c).value): n+=1
    if n: diffs[s]=n
print("diff by sheet:", diffs)
chk("only QB VALUES + START HERE + CHANGELOG changed", set(diffs)=={"QB VALUES","START HERE","CHANGELOG"}, str(set(diffs)))
chk("START HERE diff == 1 (banner)", diffs.get("START HERE")==1)
chk("QB VALUES diff == 934", diffs.get("QB VALUES")==934, str(diffs.get("QB VALUES")))

qb=b['QB VALUES']
# formula columns intact (A,B,G,M) for all 138 rows
badf=[]
for r in range(6,144):
    if not isf(qb.cell(row=r,column=1).value): badf.append(('A',r))
    if not isf(qb.cell(row=r,column=2).value): badf.append(('B',r))
    if not isf(qb.cell(row=r,column=7).value): badf.append(('G',r))
    if not isf(qb.cell(row=r,column=13).value): badf.append(('M',r))
chk("A/B/G/M formulas intact for all 138 rows", not badf, str(badf[:5]))

# D (baseline value) and F (active value) BLANK for all -> no numeric values written
nb_val=[(r,col) for r in range(6,144) for col in (4,6) if qb.cell(row=r,column=col).value is not None]
chk("Baseline value (D) & Active value (F) blank for all 138 (no numbers written)", not nb_val, str(nb_val[:5]))
# G (delta) has no cached value (formula only) -> stays blank
chk("QB delta (G) left as formula (no manual value)", all(isf(qb.cell(row=r,column=7).value) for r in range(6,144)))

# E (active qb) populated for all 138; H,I,J,K,L populated
miss=[]
for r in range(6,144):
    for col,name in [(5,'E'),(8,'H'),(9,'I'),(10,'J'),(11,'K'),(12,'L')]:
        if qb.cell(row=r,column=col).value in (None,""): miss.append((name,r))
chk("E/H/I/J/K/L populated for all 138 rows", not miss, str(miss[:5]))
# J == 2026 all
chk("Reviewed for season (J) == 2026 for all 138", all(qb.cell(row=r,column=10).value==2026 for r in range(6,144)))
# K populated (dates)
import datetime
chk("Last update (K) populated for all 138", all(isinstance(qb.cell(row=r,column=11).value,(datetime.date,datetime.datetime)) for r in range(6,144)))
# Confidence only H/M/L
confs=set(qb.cell(row=r,column=8).value for r in range(6,144))
chk("Confidence values are only H/M/L", confs<= {"H","M","L"}, str(confs))
# Baseline QB (C): populated for H/M, blank for L
Cpop=[r for r in range(6,144) if qb.cell(row=r,column=3).value not in (None,"")]
Lrows=[r for r in range(6,144) if qb.cell(row=r,column=8).value=="L"]
Cblank_L=[r for r in Lrows if qb.cell(row=r,column=3).value in (None,"")]
chk("Baseline QB (C) blank for all L-confidence (open competition) teams", len(Cblank_L)==len(Lrows), f"{len(Cblank_L)}/{len(Lrows)}")
chk("Baseline QB (C) populated for all H/M teams (106)", len(Cpop)==106, str(len(Cpop)))

# no duplicate/missing teams: A column formulas -> 138 distinct handled by TEAM MAP (unchanged). Confirm A/B formula rows 6-143 present and 144 blank pattern
chk("QB VALUES row 144 A formula still present (structure intact)", isf(qb.cell(row=144,column=1).value) or qb.cell(row=144,column=1).value is None)

print()
print("*** %d FAILURES ***"%len(fails) if fails else "ALL WORKBOOK VERIFICATION CHECKS PASSED")
if fails: raise SystemExit(1)

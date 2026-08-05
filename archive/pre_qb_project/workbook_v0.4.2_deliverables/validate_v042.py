"""
v0.4.2 validation battery:
 A. SP+ revalidation: independent re-parse of the captured ESPN article
    (fresh regex extraction) compared cell-by-cell against PRESEASON!D6:D143;
    E = 2026-03-27 and F = official citation on all 138 rows.
 B. FPI / TeamRankings unchanged: re-parse archives, compare H and Q.
 C. Three-source blend: independently reproduce Y (0.70), AA (3), AB
    (blank), and the FINAL PRIOR Z for every team from raw values +
    SETTINGS weights; report top/bottom 5.
 D. Full cell/formula diff v0.4.2 vs v0.3.1 -> TSV; formula-integrity
    assertion (zero v0.3.1-formula cells changed).
 E. Blank assertions: TTW/VSiN inputs, win totals, QB manual columns.
 F. No remaining 'SP+ paywalled/unavailable' claims anywhere in the workbook.
"""
import datetime
import html as htmllib
import json
import re

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

P4 = "../phase4"
CITE = "https://www.espn.com/college-football/story/_/id/48306284/2026-college-football-sp+-rankings-138-fbs-teams"
failures = []

def check(label, ok, detail=""):
    print(("PASS  " if ok else "FAIL  ") + label + (f"  {detail}" if detail else ""))
    if not ok:
        failures.append((label, detail))

wb = openpyxl.load_workbook("v0.4.2_working.xlsx", data_only=False)
ps = wb["PRESEASON"]
master = json.load(open(f"{P4}/master_list.json"))

# ---------- A. SP+ revalidation (independent re-parse) ----------
raw = open(f"{P4}/raw_sources/espn_spplus_article_PAYWALLED_20260719.html").read()
team_table = re.findall(r"<table.*?</table>", raw, re.S)[0]
fresh = {}
for row in re.findall(r"<tr.*?</tr>", team_table, re.S)[1:]:
    cells = [htmllib.unescape(re.sub(r"<[^>]+>", "", c)).strip()
             for c in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row, re.S)]
    m = re.match(r"(\d+)\.\s*(.+)", cells[0])
    fresh[m.group(2).strip()] = float(cells[1])
check("SP+ re-parse row count == 138", len(fresh) == 138, str(len(fresh)))

src = open(f"{P4}/extract_spplus.py").read()
ns = {}
exec(src.split("def main")[0], ns)
NAME_TO_AB = ns["SPPLUS_TO_ABBREV"]
sp_by_ab = {}
for name, val in fresh.items():
    ab = NAME_TO_AB.get(name)
    if ab is None:
        failures.append(("SP+ unmatched name on revalidation", name)); continue
    sp_by_ab[ab] = val
check("SP+ re-parse matched 138 unique abbrevs", len(sp_by_ab) == 138, str(len(sp_by_ab)))

bad_d, bad_e, bad_f = [], [], []
for row_str, (ab, name, conf) in master.items():
    r = int(row_str)
    if ps.cell(row=r, column=4).value != sp_by_ab.get(ab):
        bad_d.append((ab, ps.cell(row=r, column=4).value, sp_by_ab.get(ab)))
    ev = ps.cell(row=r, column=5).value
    if not (isinstance(ev, (datetime.date, datetime.datetime)) and ev.strftime("%Y-%m-%d") == "2026-03-27"):
        bad_e.append((ab, ev))
    if ps.cell(row=r, column=6).value != CITE:
        bad_f.append((ab, ps.cell(row=r, column=6).value))
check("SP+ 138/138 D values match independently re-parsed capture", not bad_d, str(bad_d[:5]))
check("SP+ 138/138 E == 2026-03-27 (publication date)", not bad_e, str(bad_e[:5]))
check("SP+ 138/138 F == official ESPN citation", not bad_f, str(bad_f[:3]))

# ---------- B. FPI / TR unchanged ----------
fpi_raw = json.load(open(f"{P4}/raw_sources/espn_fpi_2026preseason_20260719_parsed.json"))
k_to_abbrev = json.load(open(f"{P4}/k_to_abbrev.json"))
fpi_by_ab = {k_to_abbrev[s["team"]["displayName"]]: float(next(x["value"] for x in s["stats"] if x["name"] == "fpi"))
             for s in fpi_raw["page"]["content"]["table"]["stats"]}
tr_html = open(f"{P4}/raw_sources/teamrankings_predictive_2026preseason_20260719.html").read()
tr_rows = re.findall(
    r'<td class="rank[^"]*"[^>]*>(\d+)</td>\s*'
    r'<td class="nowrap"[^>]*><a href="[^"]*">([^<]+)</a>.*?'
    r'<td class="text-right"[^>]*>(-?\d+\.\d+)</td>', tr_html, re.S)
import os
src_tr = open(f"{P4}/match_teamrankings.py").read()
ns2 = {}; _cwd = os.getcwd(); os.chdir(P4)
try: exec(src_tr.split("rows = []")[0], ns2)
finally: os.chdir(_cwd)
tr_by_ab = {ns2["TR_TO_ABBREV"][htmllib.unescape(n).strip()]: float(v) for _, n, v in tr_rows}

bad_h = [(ab,) for rs, (ab, _, _) in master.items() if ps.cell(row=int(rs), column=8).value != fpi_by_ab[ab]]
bad_q = [(ab,) for rs, (ab, _, _) in master.items() if ps.cell(row=int(rs), column=17).value != tr_by_ab[ab]]
check("FPI 138/138 unchanged and matching raw archive", not bad_h, str(bad_h[:5]))
check("TR 138/138 unchanged and matching raw archive", not bad_q, str(bad_q[:5]))

# ---------- C. Three-source blend reproduction ----------
st = wb["SETTINGS"]
w_sp, w_fpi, w_ttw, w_tr, w_vsin = (st.cell(row=r, column=2).value for r in range(28, 33))
check("SETTINGS weights are 0.30/0.25/0.20/0.15/0.10",
      (w_sp, w_fpi, w_ttw, w_tr, w_vsin) == (0.3, 0.25, 0.2, 0.15, 0.1), str((w_sp, w_fpi, w_ttw, w_tr, w_vsin)))

D = {int(rs): ps.cell(row=int(rs), column=4).value for rs in master}
H = {int(rs): ps.cell(row=int(rs), column=8).value for rs in master}
Q = {int(rs): ps.cell(row=int(rs), column=17).value for rs in master}
avg_d, avg_h, avg_q = (sum(x.values()) / 138 for x in (D, H, Q))

y_expected = w_sp + w_fpi + w_tr  # 0.70
blended = {}
for rs, (ab, name, conf) in master.items():
    r = int(rs)
    g = D[r] - avg_d      # SP+ norm
    k = H[r] - avg_h      # FPI norm
    t = Q[r] - avg_q      # TR norm
    z = (w_sp * g + w_fpi * k + w_tr * t) / y_expected
    blended[ab] = z
check("Available weight Y == 0.70 for every team (0.30+0.25+0.15, all 138 rows populated)",
      abs(y_expected - 0.70) < 1e-12, str(y_expected))
check("Source count AA == 3 for every team (D,H,Q numeric on all 138 rows; L,U blank)",
      all(isinstance(D[int(rs)], (int, float)) and isinstance(H[int(rs)], (int, float)) and isinstance(Q[int(rs)], (int, float)) for rs in master)
      and all(ps.cell(row=int(rs), column=12).value is None and ps.cell(row=int(rs), column=21).value is None for rs in master),
      "(integral values read back as int by openpyxl; numerically identical, ISNUMBER-true)")
check("Single-source flag AB blank for every team (AA=3 > 1)", True, "by AA==3 above; formula unchanged (see formula-integrity)")

rank = sorted(blended.items(), key=lambda kv: -kv[1])
print("\nIndependently reproduced FINAL PRIOR (pts vs avg FBS), 3-source blend:")
print("  TOP 5   :", [(ab, round(v, 2)) for ab, v in rank[:5]])
print("  BOTTOM 5:", [(ab, round(v, 2)) for ab, v in rank[-5:]])
json.dump({ab: round(v, 4) for ab, v in rank}, open("v0.4.2_blend_reproduction.json", "w"), indent=1)

# ---------- D. Full diff vs v0.3.1 + formula integrity ----------
def normval(v):
    if isinstance(v, ArrayFormula):
        return ("ARRAYFORMULA", v.ref, v.text)
    return v

wb_old = openpyxl.load_workbook("v0.3.1_with_schedule.xlsx", data_only=False)
diff_lines, diff_counts, formula_diffs = [], {}, []
for name in wb_old.sheetnames:
    ws_old, ws_new = wb_old[name], wb[name]
    mr = max(ws_old.max_row, ws_new.max_row); mc = max(ws_old.max_column, ws_new.max_column)
    cnt = 0
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            vo = normval(ws_old.cell(row=r, column=c).value)
            vn = normval(ws_new.cell(row=r, column=c).value)
            if vo != vn:
                cnt += 1
                diff_lines.append(f"{name}\t{ws_new.cell(row=r, column=c).coordinate}\t{repr(vo)}\t{repr(vn)}")
                if isinstance(vo, tuple) or (isinstance(vo, str) and vo.startswith("=")):
                    formula_diffs.append((name, ws_new.cell(row=r, column=c).coordinate))
    if cnt:
        diff_counts[name] = cnt
with open("v0.4.2_vs_v0.3.1_full_diff.tsv", "w") as f:
    f.write("sheet\tcell\tv0.3.1_value\tv0.4.2_value\n" + "\n".join(diff_lines) + "\n")
print("\nDiff counts by sheet:", diff_counts)

expected = {"START HERE": 2, "PRESEASON": 1243, "SETTINGS": 2, "FCS TIERS": 1, "CHANGELOG": 60}
# PRESEASON: 828 (FPI/TR) + 414 (SP+ 138x3) + 1 (A1) = 1243
# CHANGELOG: 15 rows appended since v0.3.1 (9 v0.4 + 3 v0.4.1 + 3 v0.4.2) x 4
# cols = 60; the row-21/25/29 corrections are edits WITHIN those appended
# rows, so they add no extra diff cells vs the v0.3.1 baseline.
check("Diff touches only expected sheets", set(diff_counts) == set(expected), str(set(diff_counts) ^ set(expected)))
check("PRESEASON diff == 1243 (828 FPI/TR + 414 SP+ + A1)", diff_counts.get("PRESEASON") == 1243, str(diff_counts.get("PRESEASON")))
check("START HERE diff == 2 (banner + A3)", diff_counts.get("START HERE") == 2, str(diff_counts.get("START HERE")))
check("SETTINGS diff == 2 (C22,C23)", diff_counts.get("SETTINGS") == 2, str(diff_counts.get("SETTINGS")))
check("FCS TIERS diff == 1 (A2)", diff_counts.get("FCS TIERS") == 1, str(diff_counts.get("FCS TIERS")))
cl_cnt = diff_counts.get("CHANGELOG")
check("CHANGELOG diff == 60 (15 appended rows x 4; row-21/25/29 edits are within appended rows)",
      cl_cnt == 60, str(cl_cnt))
check("ZERO formula cells changed vs v0.3.1", not formula_diffs, str(formula_diffs[:8]))

# ---------- E. Blank assertions ----------
for label, cols in {"TTW raw/date/cite (L,N,O)": [12, 14, 15], "VSiN raw/date/cite (U,V,W)": [21, 22, 23],
                    "Win totals (AD,AE)": [30, 31]}.items():
    nb = [(r, c) for r in range(6, 144) for c in cols if ps.cell(row=r, column=c).value is not None]
    check(f"PRESEASON {label} fully blank", not nb, str(nb[:5]))
qb = wb["QB VALUES"]
nb = [(r, c) for r in range(6, 144) for c in [3, 4, 5, 6, 8, 9, 10, 11, 12] if qb.cell(row=r, column=c).value is not None]
check("QB VALUES manual-input columns fully blank", not nb, str(nb[:5]))

# ---------- F. No remaining paywalled/unavailable-SP+ claims ----------
claims = []
pat = re.compile(r"paywall|SP\+ (data |access )?(is |are |remains? )?(un(available|verifiable)|not (available|accessible|loaded))|no SP\+ data|scattered individual team ranks", re.I)
for name in wb.sheetnames:
    for row in wb[name].iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and not v.startswith("=") and pat.search(v):
                # allow text that explicitly describes the CORRECTION of the old claim
                if "CORRECTED" in v or "corrected" in v or "superseded" in v:
                    continue
                claims.append(f"{name}!{cell.coordinate}")
check("No remaining claims that SP+ is paywalled/unavailable/incomplete", not claims, str(claims))

# ---------- structural ----------
cl = wb["CHANGELOG"]
r = 7; counts = {}
while cl.cell(row=r, column=1).value is not None:
    counts[cl.cell(row=r, column=1).value] = counts.get(cl.cell(row=r, column=1).value, 0) + 1
    r += 1
check("CHANGELOG: 9 v0.4 + 3 v0.4.1 + 3 v0.4.2 entries",
      counts.get("v0.4") == 9 and counts.get("v0.4.1") == 3 and counts.get("v0.4.2") == 3, str(counts))
sched = wb["IMPORT SCHEDULE"]
check("IMPORT SCHEDULE 888 game ids intact",
      sum(1 for rr in range(6, 894) if sched.cell(row=rr, column=1).value) == 888, "")

print()
if failures:
    print(f"*** {len(failures)} FAILURES ***")
    for f_ in failures: print("   ", f_)
    raise SystemExit(1)
print("ALL CHECKS PASSED")

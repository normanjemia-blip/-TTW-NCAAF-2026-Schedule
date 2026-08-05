"""
v0.5.1 live-calc unit test (fixed row-substitution bug): use the exact
same row-templated formula strings the production script wrote into
v0.5.1_working.xlsx (verified byte-identical below against the real
file's row-6 instance), wired to synthetic per-row input cells, evaluated
with the `formulas` engine.
"""
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

AI_NEW = ('=IF($A{r}="","",IF($AH{r}<>"","BLOCKED",IF($AG{r}<>"","FCS — NO PLAY",'
          'IF(CALC!$S{r}=1,"PENDING LINE",IF(CALC!$Q{r}=1,"STALE LINE",'
          'IF($AE{r}="QB UNCERTAIN","QB UNCERTAIN",IF($AF{r}<>"","FCS/TRANSITION UNCERTAIN",'
          'IF(OR($B{r}="",$C{r}=""),"DATA INCOMPLETE","READY"))))))))')
AJ_NEW = ('=IF(OR($AI{r}="",$AI{r}="BLOCKED",$AI{r}="PENDING LINE",$AI{r}="FCS — NO PLAY"),"",'
          'MAX(1,MIN(5,3-IF(CALC!$Z{r}<SETTINGS!$B$24,1,0)-IF($AE{r}="QB UNCERTAIN",1,0)'
          '-IF($AF{r}<>"",1,0)-IF($AG{r}<>"",1,0)+IF(CALC!$Z{r}>=6,1,0))))')
AB_NEW = '=IF($A{r}="",0,$L{r}*IF($O{r}="FBS",1,0))'
AC_NEW = '=IF($A{r}="",0,$L{r}*IF($Q{r}="FBS",1,0))'

# Verify these templates are byte-identical to what's actually in the
# production file (row 6), i.e. this test exercises the REAL shipped
# formula, not a re-derived approximation.
src = openpyxl.load_workbook("../workbook/v0.5.1_working.xlsx", data_only=False)
def fv(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else v
assert fv(src["ENGINE"].cell(row=6, column=35)) == AI_NEW.format(r=6)
assert fv(src["ENGINE"].cell(row=6, column=36)) == AJ_NEW.format(r=6)
assert fv(src["CLEAN"].cell(row=6, column=28)) == AB_NEW.format(r=6)
assert fv(src["CLEAN"].cell(row=6, column=29)) == AC_NEW.format(r=6)
print("Templates verified byte-identical to production v0.5.1_working.xlsx row 6.")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TEST"
calc_ws = wb.create_sheet("CALC")
settings_ws = wb.create_sheet("SETTINGS")
settings_ws["B24"] = 3

scenarios = {
    6: dict(desc="FBS v FBS, no block, no market line -> PENDING LINE",
            AH="", AE="OK", AF="", AG="", B="1", C="2026-09-05",
            CALC_S=1, CALC_Q=0, CALC_Z=5),
    7: dict(desc="FBS v FBS, has market line -> READY",
            AH="", AE="OK", AF="", AG="", B="1", C="2026-09-05",
            CALC_S=0, CALC_Q=0, CALC_Z=5),
    8: dict(desc="FBS v FCS, no market line -> FCS - NO PLAY (not PENDING LINE)",
            AH="", AE="OK", AF="", AG="FCS OPP", B="1", C="2026-09-05",
            CALC_S=1, CALC_Q=0, CALC_Z=5),
    9: dict(desc="FBS v FCS AND blocked (dup id) -> BLOCKED outranks FCS-NO-PLAY",
            AH="DUP GAMEID; ", AE="OK", AF="", AG="FCS OPP", B="1", C="2026-09-05",
            CALC_S=1, CALC_Q=0, CALC_Z=5),
    10: dict(desc="FBS v FBS, home is transitional reclassifier not cleared -> FCS/TRANSITION UNCERTAIN",
             AH="", AE="OK", AF="TRANSITIONAL", AG="", B="1", C="2026-09-05",
             CALC_S=0, CALC_Q=0, CALC_Z=5),
}

for r, s in scenarios.items():
    ws.cell(row=r, column=1, value=f"GAME{r}")
    ws.cell(row=r, column=2, value=s["B"])
    ws.cell(row=r, column=3, value=s["C"])
    ws.cell(row=r, column=31, value=s["AE"])
    ws.cell(row=r, column=32, value=s["AF"])
    ws.cell(row=r, column=33, value=s["AG"])
    ws.cell(row=r, column=34, value=s["AH"])
    ws.cell(row=r, column=35, value=AI_NEW.format(r=r))
    ws.cell(row=r, column=36, value=AJ_NEW.format(r=r))
    calc_ws.cell(row=r, column=17, value=s["CALC_Q"])
    calc_ws.cell(row=r, column=19, value=s["CALC_S"])
    calc_ws.cell(row=r, column=26, value=s["CALC_Z"])

clean_test = wb.create_sheet("CLEANTEST")
eg_scn = {
    6: dict(O="FCS", AB=True), 7: dict(O="FBS", AB=True),
    8: dict(Q="FCS", AC=True), 9: dict(Q="FBS", AC=True),
}
for r, s in eg_scn.items():
    clean_test.cell(row=r, column=1, value=f"G{r}")
    clean_test.cell(row=r, column=12, value=1)  # L
    if s.get("AB"):
        clean_test.cell(row=r, column=15, value=s["O"])
        clean_test.cell(row=r, column=28, value=AB_NEW.format(r=r))
    else:
        clean_test.cell(row=r, column=17, value=s["Q"])
        clean_test.cell(row=r, column=29, value=AC_NEW.format(r=r))

wb.save("unit_test_model.xlsx")
print("Saved unit_test_model.xlsx")

"""
v0.5.1 validation battery vs approved v0.4.2 (v0.5 is superseded and not
used as a base).
 A. Formula-pattern audit: confirm ENGINE!AI/AJ and CLEAN!AB/AC changed
    on exactly the expected 1000 rows each, with exact expected text, and
    NO other formula cell anywhere in the workbook changed.
 B. Full cell diff vs v0.4.2 -> TSV.
 C. FCS TIERS confirmed empty (no ratings loaded/maintained).
 D. Blank assertions: QB VALUES, MARKET LINES manual cols, PRESEASON
    TTW/VSiN cols.
 E. NDSU/Sacramento State unaffected: still FBS-RECLASSIFYING, not FCS;
    transition-review-cleared still blank.
 F. Structural: 888 schedule ids intact; CHANGELOG entries present.
"""
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

failures = []
def check(label, ok, detail=""):
    print(("PASS  " if ok else "FAIL  ") + label + (f"  {detail}" if detail else ""))
    if not ok:
        failures.append((label, detail))

def fv(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else v

def normval(v):
    if isinstance(v, ArrayFormula):
        return ("ARRAYFORMULA", v.ref, v.text)
    return v

wb = openpyxl.load_workbook("v0.5.1_working.xlsx", data_only=False)
wb_base = openpyxl.load_workbook("v0.4.2_working.xlsx", data_only=False)

AI_NEW = ('=IF($A{r}="","",IF($AH{r}<>"","BLOCKED",IF($AG{r}<>"","FCS — NO PLAY",'
          'IF(CALC!$S{r}=1,"PENDING LINE",IF(CALC!$Q{r}=1,"STALE LINE",'
          'IF($AE{r}="QB UNCERTAIN","QB UNCERTAIN",IF($AF{r}<>"","FCS/TRANSITION UNCERTAIN",'
          'IF(OR($B{r}="",$C{r}=""),"DATA INCOMPLETE","READY"))))))))')
AJ_NEW = ('=IF(OR($AI{r}="",$AI{r}="BLOCKED",$AI{r}="PENDING LINE",$AI{r}="FCS — NO PLAY"),"",'
          'MAX(1,MIN(5,3-IF(CALC!$Z{r}<SETTINGS!$B$24,1,0)-IF($AE{r}="QB UNCERTAIN",1,0)'
          '-IF($AF{r}<>"",1,0)-IF($AG{r}<>"",1,0)+IF(CALC!$Z{r}>=6,1,0))))')
AB_NEW = '=IF($A{r}="",0,$L{r}*IF($O{r}="FBS",1,0))'
AC_NEW = '=IF($A{r}="",0,$L{r}*IF($Q{r}="FBS",1,0))'

# ---------- A. Formula-pattern audit ----------
engine, clean = wb["ENGINE"], wb["CLEAN"]
bad = {"AI": [], "AJ": [], "AB": [], "AC": []}
for r in range(6, 1006):
    if fv(engine.cell(row=r, column=35)) != AI_NEW.format(r=r): bad["AI"].append(r)
    if fv(engine.cell(row=r, column=36)) != AJ_NEW.format(r=r): bad["AJ"].append(r)
    if fv(clean.cell(row=r, column=28)) != AB_NEW.format(r=r): bad["AB"].append(r)
    if fv(clean.cell(row=r, column=29)) != AC_NEW.format(r=r): bad["AC"].append(r)
for k, v in bad.items():
    check(f"{k}: all 1000 rows (6:1005) match the new template exactly", not v, str(v[:5]))

# ---------- B. Full diff vs v0.4.2 ----------
diff_lines, diff_counts, formula_diffs = [], {}, []
for name in wb_base.sheetnames:
    ws_old, ws_new = wb_base[name], wb[name]
    mr = max(ws_old.max_row, ws_new.max_row); mc = max(ws_old.max_column, ws_new.max_column)
    cnt = 0
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            vo = normval(ws_old.cell(row=r, column=c).value)
            vn = normval(ws_new.cell(row=r, column=c).value)
            if vo != vn:
                cnt += 1
                diff_lines.append(f"{name}\t{ws_new.cell(row=r, column=c).coordinate}\t{repr(vo)}\t{repr(vn)}")
                is_formula = isinstance(vo, tuple) or (isinstance(vo, str) and vo.startswith("="))
                if is_formula:
                    formula_diffs.append((name, ws_new.cell(row=r, column=c).coordinate))
    if cnt: diff_counts[name] = cnt
with open("v0.5.1_vs_v0.4.2_full_diff.tsv", "w") as f:
    f.write("sheet\tcell\tv0.4.2_value\tv0.5.1_value\n" + "\n".join(diff_lines) + "\n")
print("\nDiff counts (v0.5.1 vs v0.4.2):", diff_counts)

expected_formula_cells = 4000  # 1000 rows x 4 columns (AI,AJ in ENGINE; AB,AC in CLEAN)
check("Exactly 4000 formula cells changed (1000 rows x AI/AJ/AB/AC)",
      len(formula_diffs) == expected_formula_cells, str(len(formula_diffs)))
non_engine_clean_formula = [d for d in formula_diffs if d[0] not in ("ENGINE", "CLEAN")]
check("All formula diffs confined to ENGINE and CLEAN", not non_engine_clean_formula, str(non_engine_clean_formula[:5]))

expected_sheets = {"START HERE", "FCS TIERS", "DICTIONARY", "SETTINGS", "DATA QUALITY", "ENGINE", "CLEAN", "CHANGELOG"}
check("Diff touches only expected sheets", set(diff_counts) == expected_sheets, str(set(diff_counts) ^ expected_sheets))
check("ENGINE diff == 2000 (AI+AJ x 1000 rows)", diff_counts.get("ENGINE") == 2000, str(diff_counts.get("ENGINE")))
check("CLEAN diff == 2000 (AB+AC x 1000 rows)", diff_counts.get("CLEAN") == 2000, str(diff_counts.get("CLEAN")))
check("FCS TIERS diff == 2 (A1 + A2 banner text only, no data)", diff_counts.get("FCS TIERS") == 2, str(diff_counts.get("FCS TIERS")))
check("SETTINGS diff == 1 (C21 comment only)", diff_counts.get("SETTINGS") == 1, str(diff_counts.get("SETTINGS")))

# ---------- C. FCS TIERS confirmed empty ----------
ft = wb["FCS TIERS"]
nonblank = [(r, c) for r in range(6, 206) for c in range(1, 14) if c != 1 and ft.cell(row=r, column=c).value is not None]
check("FCS TIERS carries zero rating/tier data (fully inactive)", not nonblank, str(nonblank[:5]))
check("FCS TIERS!A1 marked INACTIVE", "INACTIVE" in (ft["A1"].value or ""), ft["A1"].value)

# ---------- D. Blank assertions ----------
ps = wb["PRESEASON"]
for label, cols in {"TTW (L,N,O)": [12, 14, 15], "VSiN (U,V,W)": [21, 22, 23]}.items():
    nb = [(r, c) for r in range(6, 144) for c in cols if ps.cell(row=r, column=c).value is not None]
    check(f"PRESEASON {label} fully blank", not nb, str(nb[:5]))
qb = wb["QB VALUES"]
nb = [(r, c) for r in range(6, 144) for c in [3,4,5,6,8,9,10,11,12] if qb.cell(row=r, column=c).value is not None]
check("QB VALUES manual-input columns fully blank (0 of 138 loaded)", not nb, str(nb[:5]))
ml = wb["MARKET LINES"]
manual_cols = [1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
nb_ml = [(r, c) for r in range(6, 1006) for c in manual_cols if ml.cell(row=r, column=c).value is not None]
check("MARKET LINES manual-input columns fully blank", not nb_ml, str(nb_ml[:5]))

# ---------- E. NDSU / Sacramento State unaffected ----------
tm = wb["TEAM MAP"]
check("SAC still FBS-RECLASSIFYING (TEAM MAP D114)", tm.cell(row=114, column=4).value == "FBS-RECLASSIFYING", "")
check("NDSU still FBS-RECLASSIFYING (TEAM MAP D122)", tm.cell(row=122, column=4).value == "FBS-RECLASSIFYING", "")
check("SAC still Transitional=Y (TEAM MAP E114)", tm.cell(row=114, column=5).value == "Y", "")
check("NDSU still Transitional=Y (TEAM MAP E122)", tm.cell(row=122, column=5).value == "Y", "")
tr = wb["TEAM RATINGS"]
check("TEAM RATINGS!X114 (SAC review-cleared) still blank", tr.cell(row=114, column=24).value is None, "")
check("TEAM RATINGS!X122 (NDSU review-cleared) still blank", tr.cell(row=122, column=24).value is None, "")

# ---------- F. Structural ----------
sched = wb["IMPORT SCHEDULE"]
check("IMPORT SCHEDULE 888 game ids intact", sum(1 for r in range(6,894) if sched.cell(row=r, column=1).value) == 888, "")
cl = wb["CHANGELOG"]
r = 7; counts = {}
while cl.cell(row=r, column=1).value is not None:
    counts[cl.cell(row=r, column=1).value] = counts.get(cl.cell(row=r, column=1).value, 0) + 1
    r += 1
check("CHANGELOG has >=1 v0.5.1 entry", counts.get("v0.5.1", 0) >= 1, str(counts.get("v0.5.1")))
check("CHANGELOG v0.5 entries NOT present (v0.5.1 built from v0.4.2, not v0.5)", "v0.5" not in counts, str(list(counts)))

print()
if failures:
    print(f"*** {len(failures)} FAILURES ***")
    for f_ in failures: print("   ", f_)
    raise SystemExit(1)
print("ALL CHECKS PASSED")

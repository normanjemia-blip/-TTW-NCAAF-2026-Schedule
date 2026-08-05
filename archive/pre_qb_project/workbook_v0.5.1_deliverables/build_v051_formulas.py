"""
v0.5.1: FBS-vs-FCS scope-change formula edits, applied to rows 6:1005
across ENGINE!AI, ENGINE!AJ, CLEAN!AB, CLEAN!AC. Every row's exact
pre-edit formula is asserted against the expected v0.4.2 template before
being overwritten, so any unexpected drift aborts the run instead of
silently overwriting something unforeseen.
"""
import openpyxl

FILE = "v0.5.1_working.xlsx"
wb = openpyxl.load_workbook(FILE, data_only=False)
engine = wb["ENGINE"]
clean = wb["CLEAN"]

AI_OLD = ('=IF($A{r}="","",IF($AH{r}<>"","BLOCKED",IF(CALC!$S{r}=1,"PENDING LINE",'
          'IF(CALC!$Q{r}=1,"STALE LINE",IF($AE{r}="QB UNCERTAIN","QB UNCERTAIN",'
          'IF(OR($AF{r}<>"",AND($G{r}="FCS",$I{r}=""),AND($H{r}="FCS",$J{r}="")),'
          '"FCS/TRANSITION UNCERTAIN",IF(OR($B{r}="",$C{r}=""),"DATA INCOMPLETE","READY")))))))')
AI_NEW = ('=IF($A{r}="","",IF($AH{r}<>"","BLOCKED",IF($AG{r}<>"","FCS — NO PLAY",'
          'IF(CALC!$S{r}=1,"PENDING LINE",IF(CALC!$Q{r}=1,"STALE LINE",'
          'IF($AE{r}="QB UNCERTAIN","QB UNCERTAIN",IF($AF{r}<>"","FCS/TRANSITION UNCERTAIN",'
          'IF(OR($B{r}="",$C{r}=""),"DATA INCOMPLETE","READY"))))))))')

AJ_OLD = ('=IF(OR($AI{r}="",$AI{r}="BLOCKED",$AI{r}="PENDING LINE"),"",'
          'MAX(1,MIN(5,3-IF(CALC!$Z{r}<SETTINGS!$B$24,1,0)-IF($AE{r}="QB UNCERTAIN",1,0)'
          '-IF($AF{r}<>"",1,0)-IF($AG{r}<>"",1,0)+IF(CALC!$Z{r}>=6,1,0))))')
AJ_NEW = ('=IF(OR($AI{r}="",$AI{r}="BLOCKED",$AI{r}="PENDING LINE",$AI{r}="FCS — NO PLAY"),"",'
          'MAX(1,MIN(5,3-IF(CALC!$Z{r}<SETTINGS!$B$24,1,0)-IF($AE{r}="QB UNCERTAIN",1,0)'
          '-IF($AF{r}<>"",1,0)-IF($AG{r}<>"",1,0)+IF(CALC!$Z{r}>=6,1,0))))')

AB_OLD = '=IF($A{r}="",0,$L{r}*IF($O{r}="FCS",SETTINGS!$B$21,IF($O{r}="FBS",1,0)))'
AB_NEW = '=IF($A{r}="",0,$L{r}*IF($O{r}="FBS",1,0))'

AC_OLD = '=IF($A{r}="",0,$L{r}*IF($Q{r}="FCS",SETTINGS!$B$21,IF($Q{r}="FBS",1,0)))'
AC_NEW = '=IF($A{r}="",0,$L{r}*IF($Q{r}="FBS",1,0))'

counts = {"AI": 0, "AJ": 0, "AB": 0, "AC": 0}
mismatches = []

for r in range(6, 1006):
    # ENGINE!AI (col 35)
    c = engine.cell(row=r, column=35)
    expected = AI_OLD.format(r=r)
    if c.value != expected:
        mismatches.append(("ENGINE", "AI", r, c.value, expected)); continue
    c.value = AI_NEW.format(r=r)
    counts["AI"] += 1

    # ENGINE!AJ (col 36)
    c = engine.cell(row=r, column=36)
    expected = AJ_OLD.format(r=r)
    if c.value != expected:
        mismatches.append(("ENGINE", "AJ", r, c.value, expected)); continue
    c.value = AJ_NEW.format(r=r)
    counts["AJ"] += 1

    # CLEAN!AB (col 28)
    c = clean.cell(row=r, column=28)
    expected = AB_OLD.format(r=r)
    if c.value != expected:
        mismatches.append(("CLEAN", "AB", r, c.value, expected)); continue
    c.value = AB_NEW.format(r=r)
    counts["AB"] += 1

    # CLEAN!AC (col 29)
    c = clean.cell(row=r, column=29)
    expected = AC_OLD.format(r=r)
    if c.value != expected:
        mismatches.append(("CLEAN", "AC", r, c.value, expected)); continue
    c.value = AC_NEW.format(r=r)
    counts["AC"] += 1

print("Rows changed per column:", counts)
if mismatches:
    print(f"*** {len(mismatches)} MISMATCHES (aborting save) ***")
    for m in mismatches[:10]:
        print("  ", m[0], m[1], "row", m[2])
        print("     actual:  ", m[3])
        print("     expected:", m[4])
    raise SystemExit(1)

assert all(v == 1000 for v in counts.values()), counts

# SETTINGS!C21 - note the constant is now orphaned (comment only, no formula change)
settings = wb["SETTINGS"]
old_c21 = settings["C21"].value
settings["C21"] = ("Per approved fade amendment. NO LONGER REFERENCED as of v0.5.1: "
                    "FBS-vs-FCS games are fully excluded from effective-game "
                    "calculations (CLEAN!AB/AC), not partially weighted. Constant "
                    "retained, unused, for historical/audit reference.")
print("SETTINGS!C21 updated (constant orphaned, not deleted). Old:", old_c21)

wb.save(FILE)
print("Saved", FILE)

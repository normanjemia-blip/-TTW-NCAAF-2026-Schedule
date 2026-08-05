import copy
import openpyxl

def copy_style(src_cell, dst_cell):
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.number_format = copy.copy(src_cell.number_format)
    dst_cell.protection = copy.copy(src_cell.protection)
    dst_cell.alignment = copy.copy(src_cell.alignment)

FILE = "v0.5.2_working.xlsx"
wb = openpyxl.load_workbook(FILE, data_only=False)
cl = wb["CHANGELOG"]

# --- Fix row 35 (ENGINE!AI change description): two stale string variants ---
old_35 = cl["C35"].value
assert "FCS-TRANSITION UNCERTAIN" in old_35 and "FCS/TRANSITION UNCERTAIN" in old_35
new_35 = (old_35
          .replace("QB UNCERTAIN/FCS-TRANSITION UNCERTAIN/DATA INCOMPLETE",
                   "QB UNCERTAIN/TRANSITION UNCERTAIN/DATA INCOMPLETE")
          .replace('not cleared -> FCS/TRANSITION UNCERTAIN (unaffected)',
                    'not cleared -> TRANSITION UNCERTAIN (unaffected)'))
assert "FCS-TRANSITION UNCERTAIN" not in new_35 and "FCS/TRANSITION UNCERTAIN" not in new_35
cl["C35"] = new_35
print("CHANGELOG!C35 corrected (2 stale string variants)")

# --- Fix row 37 (CLEAN!AB/AC change description): range typo ---
old_37 = cl["C37"].value
assert "CLEAN!AC1005" in old_37 and "CLEAN!AC6:AC1005" not in old_37
new_37 = old_37.replace("CLEAN!AB6:AB1005 and CLEAN!AC1005",
                         "CLEAN!AB6:AB1005 and CLEAN!AC6:AC1005")
assert "CLEAN!AC1005" not in new_37.replace("CLEAN!AC6:AC1005", "")
cl["C37"] = new_37
print("CHANGELOG!C37 corrected (range typo CLEAN!AC1005 -> CLEAN!AC6:AC1005)")

# --- Banner ---
sh = wb["START HERE"]
old_banner = sh["A1"].value
sh["A1"] = "TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.5.2 CORRECTION BUILD — PENDING APPROVAL)"
print("Banner -> v0.5.2. Old:", old_banner)

# --- New v0.5.2 CHANGELOG entry ---
next_row = 7
while cl.cell(row=next_row, column=1).value is not None:
    next_row += 1
TODAY = "2026-07-20"
entries = [
    ("v0.5.2", TODAY,
     "CORRECTION: v0.5.1 left a documentation/implementation mismatch - "
     "DICTIONARY!B12:B14 and DATA QUALITY!A2/A11 already described the "
     "transitional-reclassifier status as \"TRANSITION UNCERTAIN\", but "
     "the actual ENGINE!AI6:AI1005 formula (1000 cells) and the "
     "DATA QUALITY!B11 COUNTIF (1 cell) still returned/counted the old "
     "string \"FCS/TRANSITION UNCERTAIN\". Corrected both to return/count "
     "\"TRANSITION UNCERTAIN\", matching the docs. STRING-LITERAL ONLY: "
     "the status priority order and every other branch of the formula are "
     "byte-identical to v0.5.1 - BLOCKED still outranks FCS — NO PLAY, "
     "which still outranks PENDING LINE/STALE LINE/QB UNCERTAIN/"
     "TRANSITION UNCERTAIN/DATA INCOMPLETE/READY, unchanged. Also "
     "corrected two documentation-only errors: CHANGELOG row 35's "
     "priority-list and scenario text still said \"FCS-TRANSITION "
     "UNCERTAIN\"/\"FCS/TRANSITION UNCERTAIN\" (now \"TRANSITION "
     "UNCERTAIN\"); CHANGELOG row 37 had a range typo, \"CLEAN!AC1005\" "
     "(now \"CLEAN!AC6:AC1005\", matching the actual 1000-row range that "
     "was edited). The FCS — NO PLAY logic, effective-game exclusions, "
     "NDSU/Sacramento State safeguards, QB state, TTW-prior state, and "
     "every other v0.5.1 decision are unchanged.",
     "v0.5.2 - TRANSITION UNCERTAIN string correction"),
]
for i, (ver, date, change, reason) in enumerate(entries):
    r = next_row + i
    cl.cell(row=r, column=1, value=ver)
    cl.cell(row=r, column=2, value=date)
    cl.cell(row=r, column=3, value=change)
    cl.cell(row=r, column=4, value=reason)
    for c in range(1, 5):
        copy_style(cl.cell(row=next_row - 1, column=c), cl.cell(row=r, column=c))
print(f"CHANGELOG: added {len(entries)} v0.5.2 entries starting at row {next_row}")

wb.save(FILE)
print("Saved", FILE)

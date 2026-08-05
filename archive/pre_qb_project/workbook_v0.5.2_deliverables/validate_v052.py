"""
v0.5.2 validation battery.
 A. Formula-pattern audit: ENGINE!AI matches the corrected template on
    all 1000 rows; DATA QUALITY!B11 corrected; confirm NO other formula
    cell anywhere differs from v0.5.1.
 B. Full diff vs v0.5.1 (immediate base) -> TSV.
 C. Full diff vs v0.4.2 (lineage) -> TSV, confirming the v0.5.1 footprint
    plus this correction is the ONLY change from the approved base.
 D. Zero remaining "FCS/TRANSITION UNCERTAIN" or "FCS-TRANSITION
    UNCERTAIN" anywhere in the workbook.
 E. Re-run every v0.5.1 standing check (blanks, NDSU/SAC, structural)
    to confirm nothing else moved.
"""
import re
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

failures = []
def check(label, ok, detail=""):
    print(("PASS  " if ok else "FAIL  ") + label + (f"  {detail}" if detail else ""))
    if not ok:
        failures.append((label, detail))

def fv(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else v

def normval(v):
    if isinstance(v, ArrayFormula):
        return ("ARRAYFORMULA", v.ref, v.text)
    return v

wb = openpyxl.load_workbook("v0.5.2_working.xlsx", data_only=False)
wb_v051 = openpyxl.load_workbook("v0.5.1_working.xlsx", data_only=False)
wb_v042 = openpyxl.load_workbook("v0.4.2_working.xlsx", data_only=False)

AI_NEW = ('=IF($A{r}="","",IF($AH{r}<>"","BLOCKED",IF($AG{r}<>"","FCS — NO PLAY",'
          'IF(CALC!$S{r}=1,"PENDING LINE",IF(CALC!$Q{r}=1,"STALE LINE",'
          'IF($AE{r}="QB UNCERTAIN","QB UNCERTAIN",IF($AF{r}<>"","TRANSITION UNCERTAIN",'
          'IF(OR($B{r}="",$C{r}=""),"DATA INCOMPLETE","READY"))))))))')

# ---------- A. Formula-pattern audit ----------
engine = wb["ENGINE"]
bad = [r for r in range(6, 1006) if fv(engine.cell(row=r, column=35)) != AI_NEW.format(r=r)]
check("ENGINE!AI: all 1000 rows match the corrected template exactly", not bad, str(bad[:5]))
check("DATA QUALITY!B11 corrected", wb["DATA QUALITY"]["B11"].value ==
      '=COUNTIF(ENGINE!$AI$6:$AI$1005,"TRANSITION UNCERTAIN")', wb["DATA QUALITY"]["B11"].value)

# ---------- B. Diff vs v0.5.1 ----------
diff_lines, diff_counts, formula_diffs = [], {}, []
for name in wb_v051.sheetnames:
    ws_old, ws_new = wb_v051[name], wb[name]
    mr = max(ws_old.max_row, ws_new.max_row); mc = max(ws_old.max_column, ws_new.max_column)
    cnt = 0
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            vo = normval(ws_old.cell(row=r, column=c).value)
            vn = normval(ws_new.cell(row=r, column=c).value)
            if vo != vn:
                cnt += 1
                diff_lines.append(f"{name}\t{ws_new.cell(row=r, column=c).coordinate}\t{repr(vo)}\t{repr(vn)}")
                if isinstance(vo, tuple) or (isinstance(vo, str) and vo.startswith("=")):
                    formula_diffs.append((name, ws_new.cell(row=r, column=c).coordinate))
    if cnt: diff_counts[name] = cnt
with open("v0.5.2_vs_v0.5.1_full_diff.tsv", "w") as f:
    f.write("sheet\tcell\tv0.5.1_value\tv0.5.2_value\n" + "\n".join(diff_lines) + "\n")
print("\nDiff vs v0.5.1:", diff_counts)

check("Diff vs v0.5.1 touches only ENGINE, DATA QUALITY, START HERE, CHANGELOG",
      set(diff_counts) == {"ENGINE", "DATA QUALITY", "START HERE", "CHANGELOG"}, str(diff_counts))
check("ENGINE diff vs v0.5.1 == 1000 (AI only)", diff_counts.get("ENGINE") == 1000, str(diff_counts.get("ENGINE")))
check("DATA QUALITY diff vs v0.5.1 == 1 (B11 only)", diff_counts.get("DATA QUALITY") == 1, str(diff_counts.get("DATA QUALITY")))
check("START HERE diff vs v0.5.1 == 1 (banner only)", diff_counts.get("START HERE") == 1, str(diff_counts.get("START HERE")))
check("Formula diffs vs v0.5.1 == 1001 (1000 ENGINE!AI + 1 DATA QUALITY!B11)",
      len(formula_diffs) == 1001, str(len(formula_diffs)))
non_expected_formula = [d for d in formula_diffs if d[0] not in ("ENGINE", "DATA QUALITY")]
check("All formula diffs vs v0.5.1 confined to ENGINE!AI and DATA QUALITY!B11", not non_expected_formula, str(non_expected_formula[:5]))

# ---------- C. Diff vs v0.4.2 (lineage) ----------
diff_counts_042 = {}
formula_diffs_042 = []
for name in wb_v042.sheetnames:
    ws_old, ws_new = wb_v042[name], wb[name]
    mr = max(ws_old.max_row, ws_new.max_row); mc = max(ws_old.max_column, ws_new.max_column)
    cnt = 0
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            vo = normval(ws_old.cell(row=r, column=c).value)
            vn = normval(ws_new.cell(row=r, column=c).value)
            if vo != vn:
                cnt += 1
                if isinstance(vo, tuple) or (isinstance(vo, str) and vo.startswith("=")):
                    formula_diffs_042.append((name, ws_new.cell(row=r, column=c).coordinate))
    if cnt: diff_counts_042[name] = cnt
print("Diff vs v0.4.2 (lineage):", diff_counts_042)
check("Diff vs v0.4.2 touches the same 8 sheets as v0.5.1 did (no new sheet touched)",
      set(diff_counts_042) == {"START HERE", "ENGINE", "DATA QUALITY", "SETTINGS", "FCS TIERS", "CLEAN", "DICTIONARY", "CHANGELOG"},
      str(set(diff_counts_042)))
check("Formula diffs vs v0.4.2 == 4001 (v0.5.1's 4000 + DATA QUALITY!B11, "
      "which v0.5.1 left unchanged vs v0.4.2 and this build now corrects)",
      len(formula_diffs_042) == 4001, str(len(formula_diffs_042)))

# ---------- D. Zero remaining stale strings anywhere ----------
EXPECTED_HISTORICAL_MENTION = {("CHANGELOG", "C43")}  # the v0.5.2 correction
# entry itself, which must name the old string to explain what changed
stale = []
for name in wb.sheetnames:
    for row in wb[name].iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, ArrayFormula): v = v.text
            if isinstance(v, str) and re.search(r'FCS[/-]TRANSITION UNCERTAIN', v):
                if (name, cell.coordinate) not in EXPECTED_HISTORICAL_MENTION:
                    stale.append((name, cell.coordinate))
check("Zero remaining stale FCS/TRANSITION or FCS-TRANSITION UNCERTAIN references "
      "(excluding the v0.5.2 correction entry's deliberate historical mention)",
      not stale, str(stale[:10]))

# ---------- E. Standing checks re-run ----------
ps = wb["PRESEASON"]
for label, cols in {"TTW (L,N,O)": [12, 14, 15], "VSiN (U,V,W)": [21, 22, 23]}.items():
    nb = [(r, c) for r in range(6, 144) for c in cols if ps.cell(row=r, column=c).value is not None]
    check(f"PRESEASON {label} fully blank", not nb, str(nb[:5]))
qb = wb["QB VALUES"]
nb = [(r, c) for r in range(6, 144) for c in [3,4,5,6,8,9,10,11,12] if qb.cell(row=r, column=c).value is not None]
check("QB VALUES manual-input columns fully blank (0 of 138 loaded)", not nb, str(nb[:5]))
ml = wb["MARKET LINES"]
manual_cols = [1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
nb_ml = [(r, c) for r in range(6, 1006) for c in manual_cols if ml.cell(row=r, column=c).value is not None]
check("MARKET LINES manual-input columns fully blank", not nb_ml, str(nb_ml[:5]))
ft = wb["FCS TIERS"]
nonblank = [(r, c) for r in range(6, 206) for c in range(1, 14) if c != 1 and ft.cell(row=r, column=c).value is not None]
check("FCS TIERS still carries zero rating/tier data", not nonblank, str(nonblank[:5]))
tm = wb["TEAM MAP"]
check("SAC still FBS-RECLASSIFYING", tm.cell(row=114, column=4).value == "FBS-RECLASSIFYING", "")
check("NDSU still FBS-RECLASSIFYING", tm.cell(row=122, column=4).value == "FBS-RECLASSIFYING", "")
tr = wb["TEAM RATINGS"]
check("TEAM RATINGS!X114 (SAC review-cleared) still blank", tr.cell(row=114, column=24).value is None, "")
check("TEAM RATINGS!X122 (NDSU review-cleared) still blank", tr.cell(row=122, column=24).value is None, "")
sched = wb["IMPORT SCHEDULE"]
check("IMPORT SCHEDULE 888 game ids intact", sum(1 for r in range(6,894) if sched.cell(row=r, column=1).value) == 888, "")
cl = wb["CHANGELOG"]
r = 7; counts = {}
while cl.cell(row=r, column=1).value is not None:
    counts[cl.cell(row=r, column=1).value] = counts.get(cl.cell(row=r, column=1).value, 0) + 1
    r += 1
check("CHANGELOG has exactly 1 v0.5.2 entry", counts.get("v0.5.2") == 1, str(counts.get("v0.5.2")))
check("CHANGELOG v0.5.1 entry count unchanged at 9", counts.get("v0.5.1") == 9, str(counts.get("v0.5.1")))

print()
if failures:
    print(f"*** {len(failures)} FAILURES ***")
    for f_ in failures: print("   ", f_)
    raise SystemExit(1)
print("ALL CHECKS PASSED")

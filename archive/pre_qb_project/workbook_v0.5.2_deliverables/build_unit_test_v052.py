"""
v0.5.2 live-calc re-verification: same 5 scenarios as v0.5.1's unit test,
using the CORRECTED formula text extracted byte-for-byte from
v0.5.2_working.xlsx, confirming the only change is the returned string
("TRANSITION UNCERTAIN" instead of "FCS/TRANSITION UNCERTAIN") and every
other branch/priority is identical.
"""
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

src = openpyxl.load_workbook("../workbook/v0.5.2_working.xlsx", data_only=False)
engine, dq = src["ENGINE"], src["DATA QUALITY"]

def fv(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else v

AI6 = fv(engine.cell(row=6, column=35))
print("Extracted AI6:", AI6)
assert "TRANSITION UNCERTAIN" in AI6
assert "FCS/TRANSITION UNCERTAIN" not in AI6
assert AI6.count('"FCS — NO PLAY"') == 1  # untouched

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TEST"
calc_ws = wb.create_sheet("CALC")
settings_ws = wb.create_sheet("SETTINGS")
settings_ws["B24"] = 3

scenarios = {
    6: dict(desc="FBS v FBS, no block, no market line -> PENDING LINE",
            AH="", AE="OK", AF="", AG="", B="1", C="2026-09-05",
            CALC_S=1, CALC_Q=0),
    7: dict(desc="FBS v FBS, has market line -> READY",
            AH="", AE="OK", AF="", AG="", B="1", C="2026-09-05",
            CALC_S=0, CALC_Q=0),
    8: dict(desc="FBS v FCS, no market line -> FCS - NO PLAY (unaffected by this fix)",
            AH="", AE="OK", AF="", AG="FCS OPP", B="1", C="2026-09-05",
            CALC_S=1, CALC_Q=0),
    9: dict(desc="FBS v FCS AND blocked (dup id) -> BLOCKED outranks FCS-NO-PLAY (unaffected)",
            AH="DUP GAMEID; ", AE="OK", AF="", AG="FCS OPP", B="1", C="2026-09-05",
            CALC_S=1, CALC_Q=0),
    10: dict(desc="FBS v FBS, home transitional reclassifier not cleared -> TRANSITION UNCERTAIN (corrected string)",
             AH="", AE="OK", AF="TRANSITIONAL", AG="", B="1", C="2026-09-05",
             CALC_S=0, CALC_Q=0),
}
for r, s in scenarios.items():
    ws.cell(row=r, column=1, value=f"GAME{r}")
    ws.cell(row=r, column=2, value=s["B"])
    ws.cell(row=r, column=3, value=s["C"])
    ws.cell(row=r, column=31, value=s["AE"])
    ws.cell(row=r, column=32, value=s["AF"])
    ws.cell(row=r, column=33, value=s["AG"])
    ws.cell(row=r, column=34, value=s["AH"])
    ws.cell(row=r, column=35, value=AI6.replace("$A6", f"$A{r}").replace("$AH6", f"$AH{r}")
            .replace("$AG6", f"$AG{r}").replace("CALC!$S6", f"CALC!$S{r}")
            .replace("CALC!$Q6", f"CALC!$Q{r}").replace("$AE6", f"$AE{r}")
            .replace("$AF6", f"$AF{r}").replace("$B6", f"$B{r}").replace("$C6", f"$C{r}"))
    calc_ws.cell(row=r, column=17, value=s["CALC_Q"])
    calc_ws.cell(row=r, column=19, value=s["CALC_S"])

wb.save("unit_test_model_v052.xlsx")
print("Saved unit_test_model_v052.xlsx")

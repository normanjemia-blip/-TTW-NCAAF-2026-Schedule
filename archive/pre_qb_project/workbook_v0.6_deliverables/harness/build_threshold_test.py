"""
Isolated threshold-boundary test: exact real X6 (Spread label) formula
text, fed synthetic edge values (V) directly, with AI="READY" and
toggle=Y so the BET path is reachable, testing exact boundary values
0.99/1.00/1.49/1.50/2.99/3.00 per SETTINGS!B8=1, B9=1.5, B10=3.
"""
import openpyxl

X_FORMULA = ('=IF(OR($V{r}="",$AI{r}="BLOCKED",$AI{r}="PENDING LINE",$AI{r}="STALE LINE"),"",'
             'IF(ABS($V{r})<SETTINGS!$B$8,"",IF(ABS($V{r})<SETTINGS!$B$9,"LEAN",'
             'IF(OR(ABS($V{r})<SETTINGS!$B$10,SETTINGS!$B$11<>"Y",$AI{r}<>"READY",'
             '$AF{r}<>"",$AG{r}<>""),"INVESTIGATE","BET"))))')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TEST"
st = wb.create_sheet("SETTINGS")
st["B8"] = 1
st["B9"] = 1.5
st["B10"] = 3
st["B11"] = "Y"

edges = [0.99, 1.00, 1.49, 1.50, 2.99, 3.00, 3.01]
for i, e in enumerate(edges):
    r = 6 + i
    ws.cell(row=r, column=22, value=e)          # V
    ws.cell(row=r, column=35, value="READY")    # AI
    ws.cell(row=r, column=32, value="")          # AF
    ws.cell(row=r, column=33, value="")          # AG
    ws.cell(row=r, column=24, value=X_FORMULA.format(r=r))  # X

wb.save("threshold_test.xlsx")
print("saved, edges:", edges)

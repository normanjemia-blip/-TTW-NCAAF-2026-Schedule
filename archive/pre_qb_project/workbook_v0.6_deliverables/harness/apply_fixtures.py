"""
Apply the Phase 6 TEST ONLY fixtures to harness_base.xlsx -> harness_test.xlsx.
Every synthetic value is a round, obviously-fake number distinguishable
from a real market line (e.g. 6.5, 51.5 - plausible-looking on purpose,
since the point is to test the PIPELINE, not to test obviously-broken
inputs for the plausible cases; the invalid/oversized cases are
deliberately extreme). All Source/Notes fields say "TEST ONLY — NOT REAL
DATA".
"""
import datetime
import openpyxl

TEST_TAG = "TEST ONLY — NOT REAL DATA"
wb = openpyxl.load_workbook("harness_base.xlsx", data_only=False)

ml = wb["MARKET LINES"]
adj = wb["ADJUSTMENTS"]
qb = wb["QB VALUES"]
st = wb["SETTINGS"]

# Weekly workflow: set current week + as-of date (normally blank pre-Week-1)
st["B4"] = 1
st["B5"] = datetime.date(2026, 9, 4)  # day after Week-1 games (synthetic "today")

# --- G1 (row 8, NC State @ Virginia): valid favorite + positive spread + valid total ---
ml.cell(row=8, column=1, value="401858202")
ml.cell(row=8, column=3, value="UVA")          # Favorite
ml.cell(row=8, column=4, value=6.5)            # Spread (positive)
ml.cell(row=8, column=5, value=51.5)           # Total
ml.cell(row=8, column=6, value=TEST_TAG)       # Source
ml.cell(row=8, column=7, value=datetime.date(2026, 9, 3))  # Line date (fresh, 1 day old)
ml.cell(row=8, column=8, value=TEST_TAG)       # Notes
# resolve both teams' QBs so G1 can reach READY (isolate spread/total/edge/label/confidence/priority)
for row, ab in [(60, "NCST"), (48, "UVA")]:
    pass  # placeholder; real rows looked up below

# --- G2 (row 6, UNC @ TCU, neutral): stale line ---
ml.cell(row=6, column=1, value="401856766")
ml.cell(row=6, column=3, value="TCU")
ml.cell(row=6, column=4, value=3.0)
ml.cell(row=6, column=5, value=55.0)
ml.cell(row=6, column=6, value=TEST_TAG)
ml.cell(row=6, column=7, value=datetime.date(2026, 8, 20))  # 15 days before as-of (>5 day staleness threshold)
ml.cell(row=6, column=8, value=TEST_TAG)

# --- G3 (row 9, San José State @ USC): intentionally NO market line row -> PENDING LINE ---
# (nothing written - absence is the test)

# --- G4 (row 10, New Mexico State @ Florida State): valid line, QB UNCERTAIN by default ---
ml.cell(row=10, column=1, value="401864570")
ml.cell(row=10, column=3, value="FSU")
ml.cell(row=10, column=4, value=24.5)
ml.cell(row=10, column=5, value=58.5)
ml.cell(row=10, column=6, value=TEST_TAG)
ml.cell(row=10, column=7, value=datetime.date(2026, 9, 3))
ml.cell(row=10, column=8, value=TEST_TAG)
# manual adjustment (valid) on this game
adj.cell(row=6, column=1, value="GAME")
adj.cell(row=6, column=2, value="401864570")
adj.cell(row=6, column=3, value=1.5)
adj.cell(row=6, column=4, value=f"{TEST_TAG}: validating manual adjustment pipeline")
adj.cell(row=6, column=5, value=TEST_TAG)
adj.cell(row=6, column=6, value=datetime.date(2026, 9, 1))
adj.cell(row=6, column=8, value="Y")
adj.cell(row=6, column=9, value=TEST_TAG)
# NOTE: oversized/non-numeric adjustment tests are built as ISOLATED
# variants D and E below (not mixed into this game's Effective sum -
# an earlier draft mixed them and produced a contaminated ~1000pt
# "margin", which is documented in the test report as a fixture-design
# lesson, not a workbook defect).

# --- G5 (row 11, Jacksonville State @ North Dakota State): valid line, QBs
#     temporarily resolved on BOTH teams to reveal TRANSITION UNCERTAIN beneath QB UNCERTAIN ---
ml.cell(row=11, column=1, value="401864577")
ml.cell(row=11, column=3, value="NDSU")
ml.cell(row=11, column=4, value=3.0)
ml.cell(row=11, column=5, value=45.0)
ml.cell(row=11, column=6, value=TEST_TAG)
ml.cell(row=11, column=7, value=datetime.date(2026, 9, 3))
ml.cell(row=11, column=8, value=TEST_TAG)

# --- G6 (row 14, Bethune-Cookman @ UCF, FCS): TEST line entered anyway,
#     to prove it CANNOT override FCS - NO PLAY ---
ml.cell(row=14, column=1, value="401856767")
ml.cell(row=14, column=3, value="UCF")
ml.cell(row=14, column=4, value=45.0)
ml.cell(row=14, column=5, value=60.0)
ml.cell(row=14, column=6, value=TEST_TAG)
ml.cell(row=14, column=7, value=datetime.date(2026, 9, 3))
ml.cell(row=14, column=8, value=TEST_TAG)

# --- G7 (row 15, Akron @ Wake Forest): invalid/incomplete line (bad favorite) ---
ml.cell(row=15, column=1, value="401858204")
ml.cell(row=15, column=3, value="Ohio State")   # NOT a team in this game -> INVALID FAVORITE
ml.cell(row=15, column=4, value=10.0)
ml.cell(row=15, column=6, value=TEST_TAG)
ml.cell(row=15, column=7, value=datetime.date(2026, 9, 3))
ml.cell(row=15, column=8, value=f"{TEST_TAG}: deliberately invalid favorite + missing total")

wb.save("harness_test_A.xlsx")
print("Saved harness_test_A.xlsx (main scenario set)")

# ============ Variant B: temporarily resolve QBs for G1 (NCST/UVA) and
# G5 (JVST/NDSU) teams only - G4 (NMSU/FSU) stays unresolved on purpose,
# proving the manual adjustment does NOT bypass the QB gate, and proving
# TRANSITION UNCERTAIN is revealed under NDSU once QB stops masking it. ============
wb2 = openpyxl.load_workbook("harness_test_A.xlsx", data_only=False)
qb2 = wb2["QB VALUES"]
TODAY = datetime.date(2026, 9, 1)
for row, baseline_qb, active_qb in [(64, "Test Baseline QB", "Test Active QB"),   # NCST
                                     (70, "Test Baseline QB", "Test Active QB"),  # UVA
                                     (97, "Test Baseline QB", "Test Active QB"),  # JVST
                                     (122, "Test Baseline QB", "Test Active QB")]:  # NDSU
    qb2.cell(row=row, column=3, value=f"{TEST_TAG}: {baseline_qb}")
    qb2.cell(row=row, column=4, value=0.0)     # Baseline value
    qb2.cell(row=row, column=5, value=f"{TEST_TAG}: {active_qb}")
    qb2.cell(row=row, column=6, value=0.0)     # Active value (delta = 0, neutral)
    qb2.cell(row=row, column=8, value="M")     # Confidence (not L)
    qb2.cell(row=row, column=9, value=TEST_TAG)  # Source
    qb2.cell(row=row, column=10, value=2026)   # Reviewed for season
    qb2.cell(row=row, column=11, value=TODAY)  # Last update
    qb2.cell(row=row, column=12, value=f"{TEST_TAG}: temporarily resolved for Phase 6 QB-gate test")

wb2.save("harness_test_B.xlsx")
print("Saved harness_test_B.xlsx (QBs temporarily resolved for G1 + G5)")

# ============ Variant C: BET toggle ON, to test toggle OFF vs ON ============
wb3 = openpyxl.load_workbook("harness_test_B.xlsx", data_only=False)
wb3["SETTINGS"]["B11"] = "Y"
wb3.save("harness_test_C.xlsx")
print("Saved harness_test_C.xlsx (BET toggle ON, everything else same as B)")

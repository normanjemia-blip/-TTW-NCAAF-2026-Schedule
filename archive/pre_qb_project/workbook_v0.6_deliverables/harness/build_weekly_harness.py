"""
Weekly-rating-workflow harness (Phase 6 item 2): extends harness_base.xlsx
with (a) 2 completed FBS-vs-FBS games + 1 completed FBS-vs-FCS game
(synthetic TEST scores) to test effective-game accrual, (b) a small
IMPORT STATS fixture (2 matched teams + 1 deliberately unmatched team
name), (c) a HISTORY prior-week snapshot for one team, set far enough
from its new blended rating to test the +/-2.5 movement cap.
All real formulas, byte-exact; only cell VALUES are synthetic.
"""
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

TEST_TAG = "TEST ONLY — NOT REAL DATA"
src = openpyxl.load_workbook("../workbook/v0.6_working.xlsx", data_only=False)

def fv(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else v

wb = openpyxl.load_workbook("harness_base.xlsx", data_only=False)

# --- (a) Mark G1 (row 8, NCST@UVA) and a second FBS-FBS game as completed
#     with scores; mark G5's away side... use G1 + G3 (SJSU@USC) as FBS-FBS
#     completed; use G6 (row 14, Bethune-Cookman@UCF, FCS) as completed too,
#     to prove the FCS result contributes zero effective games/movement. ---
sched = wb["IMPORT SCHEDULE"]
clean = wb["CLEAN"]
for row, away_pts, home_pts in [(8, 17, 24), (9, 10, 41), (14, 6, 45)]:
    sched.cell(row=row, column=10, value=away_pts)   # J away_points
    sched.cell(row=row, column=11, value=home_pts)   # K home_points
    sched.cell(row=row, column=12, value=True)       # L completed
    sched.cell(row=row, column=14, value=f"{TEST_TAG}: synthetic completed-game score for Phase 6 weekly-workflow test")
    # CLEAN J/K/L/M/T/U/V/W/X/Y/Z/AA/AB/AC/AD/AE need the real formulas for these rows
    for c, letter in [(10,'J'),(11,'K'),(12,'L'),(20,'T'),(21,'U'),(22,'V'),(23,'W'),(24,'X'),(25,'Y'),(26,'Z'),(27,'AA'),(28,'AB'),(29,'AC'),(30,'AD'),(31,'AE')]:
        f = fv(src["CLEAN"].cell(row=row, column=c))
        if f is not None:
            clean.cell(row=row, column=c, value=f)

# --- (b) IMPORT STATS fixture: 3 rows, 2 matched (NCST, UVA display names
#     as they'd appear in a CFBD-style export) + 1 deliberately unmatched ---
ist = wb["IMPORT STATS"]  # already exists (empty placeholder) in harness_base.xlsx
ist_src = src["IMPORT STATS"]
for c in range(1, 12):
    f = fv(ist_src.cell(row=6, column=c))
    if f is not None:
        ist.cell(row=6, column=c, value=f)
        ist.cell(row=7, column=c, value=f.replace("6", "7") if isinstance(f, str) and "6" in f else f)
        ist.cell(row=8, column=c, value=f.replace("6", "8") if isinstance(f, str) and "6" in f else f)
# row 6: NC State (matches TEAM MAP alias "NC State Wolfpack" style)
ist.cell(row=6, column=1, value="NC State")
ist.cell(row=6, column=2, value=1)
ist.cell(row=6, column=3, value=0.15)
ist.cell(row=6, column=4, value=-0.05)
ist.cell(row=6, column=5, value=0.44)
ist.cell(row=6, column=6, value=0.38)
ist.cell(row=6, column=7, value=0.35)
ist.cell(row=6, column=8, value=0.28)
ist.cell(row=6, column=9, value=68.0)
# row 7: Virginia
ist.cell(row=7, column=1, value="Virginia")
ist.cell(row=7, column=2, value=1)
ist.cell(row=7, column=3, value=0.08)
ist.cell(row=7, column=4, value=-0.02)
ist.cell(row=7, column=5, value=0.41)
ist.cell(row=7, column=6, value=0.40)
ist.cell(row=7, column=7, value=0.30)
ist.cell(row=7, column=8, value=0.29)
ist.cell(row=7, column=9, value=70.0)
# row 8: deliberately unmatched team name (typo/nonexistent)
ist.cell(row=8, column=1, value=f"{TEST_TAG}: Nonexistent Fictional University")
ist.cell(row=8, column=2, value=1)
ist.cell(row=8, column=3, value=0.10)
ist.cell(row=8, column=4, value=-0.10)
ist.cell(row=8, column=5, value=0.40)
ist.cell(row=8, column=6, value=0.40)
ist.cell(row=8, column=7, value=0.30)
ist.cell(row=8, column=8, value=0.30)
ist.cell(row=8, column=9, value=69.0)
# fix K formula for row 6/7/8 (auto-resolve column) - use exact real template
K_TEMPLATE = fv(ist_src.cell(row=6, column=11))
for r in (6, 7, 8):
    ist.cell(row=r, column=11, value=K_TEMPLATE.replace("$A6", f"$A{r}"))

# --- (c) HISTORY prior-week snapshot for NC State (NCST, row 64), set far
#     from its new blended rating to exercise the +/-2.5 cap ---
hist = wb["HISTORY"]
hist.cell(row=6, column=1, value=1)          # week
hist.cell(row=6, column=2, value="NCST")     # abbrev
hist.cell(row=6, column=3, value=-20.0)      # prior rating (deliberately far off)
wb["SETTINGS"]["B4"] = 2  # current week = 2, so HISTORY week 1 lookup applies

wb.save("weekly_harness.xlsx")
print("saved weekly_harness.xlsx")

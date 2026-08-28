"""
Phase 6 operational-test harness: a compact workbook built from byte-exact
real formulas extracted from v0.6_working.xlsx, covering the full
pipeline (IMPORT SCHEDULE -> CLEAN -> ENGINE -> CALC -> TEAM RATINGS ->
DASHBOARD -> DATA QUALITY -> MARKET LINES -> QB VALUES -> ADJUSTMENTS ->
PRESEASON -> TEAM MAP -> SETTINGS) for:
  - All 138 real team rows (6:143) in TEAM MAP/PRESEASON/TEAM RATINGS/
    QB VALUES/CALC-team-block, so INDEX/MATCH lookups behave exactly as
    in production.
  - 7 real game rows (the G1..G7 fixture games) in CLEAN/ENGINE/
    CALC-game-block/MARKET LINES/DASHBOARD, at their ORIGINAL row
    numbers (6,8,9,10,11,14,15) - not renumbered - so every formula's
    row-relative references are unchanged from production.
No formula text is invented; every cell is copied verbatim from
v0.6_working.xlsx.
"""
import copy
import json
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter

SRC = "../workbook/v0.6_working.xlsx"
GAME_ROWS = [6, 8, 9, 10, 11, 14, 15]

def fv(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else v

src = openpyxl.load_workbook(SRC, data_only=False)
out = openpyxl.Workbook()
out.remove(out.active)

def copy_sheet_rows(sheet_name, rows, max_col, out_name=None):
    ws_src = src[sheet_name]
    ws_out = out.create_sheet(out_name or sheet_name)
    for r in rows:
        for c in range(1, max_col + 1):
            v = fv(ws_src.cell(row=r, column=c))
            if v is not None:
                ws_out.cell(row=r, column=c, value=v)
    return ws_out

# --- Full 138-team sheets ---
TEAM_ROWS = list(range(6, 144))
copy_sheet_rows("TEAM MAP", TEAM_ROWS, 6)
copy_sheet_rows("PRESEASON", TEAM_ROWS, 31)
copy_sheet_rows("TEAM RATINGS", TEAM_ROWS, 24)
copy_sheet_rows("QB VALUES", TEAM_ROWS, 13)
# CALC team block (A:J) - separate copy pass below merges with game block

# --- TEAM MAP extra ranges: K:L alias table (6:605), N conference list (6:40) ---
tm_src, tm_out = src["TEAM MAP"], out["TEAM MAP"]
for r in range(6, 606):
    for c in (11, 12):  # K, L
        v = tm_src.cell(row=r, column=c).value
        if v is not None:
            tm_out.cell(row=r, column=c, value=v)
for r in range(6, 41):
    v = tm_src.cell(row=r, column=14).value  # N
    if v is not None:
        tm_out.cell(row=r, column=14, value=v)

# --- Game-row sheets (7 real games, original row numbers) ---
copy_sheet_rows("IMPORT SCHEDULE", GAME_ROWS, 14)
copy_sheet_rows("CLEAN", GAME_ROWS, 33)
copy_sheet_rows("ENGINE", GAME_ROWS, 38)
copy_sheet_rows("DASHBOARD", GAME_ROWS, 26)
copy_sheet_rows("MARKET LINES", GAME_ROWS, 18)

# --- CALC: merge team block (A:J, 138 rows) + game block (K:Z, 7 rows) ---
calc_src, calc_out = src["CALC"], out.create_sheet("CALC")
for r in TEAM_ROWS:
    for c in range(1, 11):  # A:J
        v = fv(calc_src.cell(row=r, column=c))
        if v is not None:
            calc_out.cell(row=r, column=c, value=v)
for r in GAME_ROWS:
    for c in range(11, 27):  # K:Z
        v = fv(calc_src.cell(row=r, column=c))
        if v is not None:
            calc_out.cell(row=r, column=c, value=v)

# --- ADJUSTMENTS (rows will be added by the fixture-application script;
#     copy the auto-formula columns J/K for a generous row range now) ---
adj_src, adj_out = src["ADJUSTMENTS"], out.create_sheet("ADJUSTMENTS")
for r in range(6, 16):
    for c in (10, 11):  # J, K
        v = fv(adj_src.cell(row=r, column=c))
        if v is not None:
            adj_out.cell(row=r, column=c, value=v)

# --- DATA QUALITY: the count/check formulas ---
dq_src, dq_out = src["DATA QUALITY"], out.create_sheet("DATA QUALITY")
for r in range(6, 22):
    v = fv(dq_src.cell(row=r, column=2))
    if v is not None:
        dq_out.cell(row=r, column=1, value=dq_src.cell(row=r, column=1).value)
        dq_out.cell(row=r, column=2, value=v)

# --- SETTINGS: copy every constant verbatim ---
st_src, st_out = src["SETTINGS"], out.create_sheet("SETTINGS")
for r in list(range(3, 33)) + list(range(36, 47)):
    for c in (1, 2, 3):
        v = st_src.cell(row=r, column=c).value
        if v is not None:
            st_out.cell(row=r, column=c, value=v)

out.save("harness_base.xlsx")
print("Saved harness_base.xlsx")
print("Sheets:", out.sheetnames)

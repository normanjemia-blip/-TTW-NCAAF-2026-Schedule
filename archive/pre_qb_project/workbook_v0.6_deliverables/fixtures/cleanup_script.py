"""
Phase 6 cleanup script: reverses every synthetic fixture applied by
apply_to_real_workbook.py, producing v0.6_test_cleaned.xlsx. Proof step:
this file must be cell-value-identical to the untouched v0.6_working.xlsx
/ approved v0.5.2_working.xlsx (verified by a full cell-by-cell diff, not
just a file hash, since re-saving through openpyxl changes irrelevant
zip/XML metadata even when logical content is identical).

IMPORTANT: only MANUAL-INPUT columns are cleared on each sheet - formula
("auto") columns are never touched, matching exactly the columns the
fixture-application script wrote to. An earlier draft of this script
naively cleared whole column ranges including formula columns (MARKET
LINES!B, ADJUSTMENTS!J:K, QB VALUES!G) and blanked IMPORT SCHEDULE cells
that had genuine non-blank pre-existing values (L=FALSE, N=an ESPN
source note) instead of restoring them - both caught and fixed by the
cell-by-cell diff proof below, before this file was ever presented as
the "cleaned" deliverable.
"""
import openpyxl

SRC = "v0.6_test_applied.xlsx"
OUT = "v0.6_test_cleaned.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=False)

# --- SETTINGS: revert week/as-of/toggle to blank/default ---
st = wb["SETTINGS"]
st["B4"] = None
st["B5"] = None
st["B11"] = "N"

# --- MARKET LINES: clear only the manual-input columns (A,C,D,E,F,G,H) -
#     column B (Matchup, auto) was never written to and must survive ---
ml = wb["MARKET LINES"]
for row in (6, 8, 9, 10, 11, 14, 15):
    for c in (1, 3, 4, 5, 6, 7, 8):
        ml.cell(row=row, column=c).value = None

# --- ADJUSTMENTS: clear only A:I (1-9) - J,K (Effective/Flags) are
#     formulas and were never written to ---
adj = wb["ADJUSTMENTS"]
for c in range(1, 10):
    adj.cell(row=6, column=c).value = None

# --- QB VALUES: clear only C,D,E,F,H,I,J,K,L - A,B (formulas from TEAM
#     RATINGS), G (delta, formula), M (status, formula) were never written to ---
qb = wb["QB VALUES"]
for row in (64, 70, 97, 122):
    for c in (3, 4, 5, 6, 8, 9, 10, 11, 12):
        qb.cell(row=row, column=c).value = None

# --- IMPORT SCHEDULE: restore each cell's TRUE original value (not
#     blindly None - L was FALSE, not blank, before the fixture; N carries
#     a genuine ESPN source note on rows 8/9, blank on row 14) ---
sched = wb["IMPORT SCHEDULE"]
ORIGINAL = {
    8: {10: None, 11: None, 12: False,
        14: "ESPN source: season_type=2 (regular-season), week=1. Normalized "
            "to project Week 0 (opening slate before the 2026-09-03 Week 1 kickoff)."},
    9: {10: None, 11: None, 12: False,
        14: "ESPN source: season_type=2 (regular-season), week=1. Normalized "
            "to project Week 0 (opening slate before the 2026-09-03 Week 1 kickoff)."},
    14: {10: None, 11: None, 12: False, 14: None},
}
for row, cols in ORIGINAL.items():
    for c, v in cols.items():
        sched.cell(row=row, column=c).value = v

# --- IMPORT STATS: clear the 3 test rows (all columns were genuinely blank before) ---
ist = wb["IMPORT STATS"]
for row in (6, 7, 8):
    for c in range(1, 10):
        ist.cell(row=row, column=c).value = None

# --- HISTORY: clear the 1 test row (genuinely blank before - HISTORY had
#     no snapshots yet, preseason) ---
hist = wb["HISTORY"]
for c in range(1, 5):
    hist.cell(row=6, column=c).value = None

# --- START HERE: restore banner to the approved v0.5.2 text ---
sh = wb["START HERE"]
sh["A1"] = "TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.5.2 CORRECTION BUILD — PENDING APPROVAL)"

wb.save(OUT)
print("Saved", OUT)

# --- Self-check: zero TEST_TAG cells remain ---
TEST_TAG = "TEST ONLY — NOT REAL DATA"
tagged = []
for name in wb.sheetnames:
    for row in wb[name].iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and TEST_TAG in cell.value:
                tagged.append((name, cell.coordinate))
print(f"Remaining TEST_TAG cells: {len(tagged)}", tagged[:5])

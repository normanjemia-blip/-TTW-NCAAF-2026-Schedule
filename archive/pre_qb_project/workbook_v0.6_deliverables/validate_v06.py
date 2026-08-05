"""
v0.6 final validation battery vs approved v0.5.2.
"""
import json
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

failures = []
def check(label, ok, detail=""):
    print(("PASS  " if ok else "FAIL  ") + label + (f"  {detail}" if detail else ""))
    if not ok:
        failures.append((label, detail))

def normval(v):
    if isinstance(v, ArrayFormula):
        return ("ARRAYFORMULA", v.ref, v.text)
    return v

wb = openpyxl.load_workbook("v0.6_working.xlsx", data_only=False)
wb_base = openpyxl.load_workbook("v0.5.2_working.xlsx", data_only=False)

# ---------- A. Full diff vs v0.5.2: ZERO formula changes, only banner+CHANGELOG ----------
diff_lines, diff_counts, formula_diffs = [], {}, []
for name in wb_base.sheetnames:
    ws1, ws2 = wb_base[name], wb[name]
    mr = max(ws1.max_row, ws2.max_row); mc = max(ws1.max_column, ws2.max_column)
    cnt = 0
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            v1 = normval(ws1.cell(row=r, column=c).value)
            v2 = normval(ws2.cell(row=r, column=c).value)
            if v1 != v2:
                cnt += 1
                diff_lines.append(f"{name}\t{ws2.cell(row=r, column=c).coordinate}\t{repr(v1)}\t{repr(v2)}")
                if isinstance(v1, tuple) or (isinstance(v1, str) and v1.startswith("=")):
                    formula_diffs.append((name, ws2.cell(row=r, column=c).coordinate))
    if cnt: diff_counts[name] = cnt
with open("v0.6_vs_v0.5.2_full_diff.tsv", "w") as f:
    f.write("sheet\tcell\tv0.5.2_value\tv0.6_value\n" + "\n".join(diff_lines) + "\n")
print("Diff vs v0.5.2:", diff_counts)
check("Diff vs v0.5.2 touches ONLY START HERE + CHANGELOG", set(diff_counts) == {"START HERE", "CHANGELOG"}, str(diff_counts))
check("ZERO formula cells changed anywhere vs v0.5.2", not formula_diffs, str(formula_diffs[:5]))
check("START HERE diff == 1 (banner only)", diff_counts.get("START HERE") == 1, str(diff_counts.get("START HERE")))
check("CHANGELOG diff == 28 (7 new rows x 4)", diff_counts.get("CHANGELOG") == 28, str(diff_counts.get("CHANGELOG")))

# ---------- B. Target counts (schedule composition) ----------
cls = json.load(open("fcs_count_harness.xlsx".replace(".xlsx", "") + ".json") if False else open("fcs_classification_888.json"))
fbs_fbs = sum(1 for o, q in cls.values() if o == "FBS" and q == "FBS")
fcs_involved = sum(1 for o, q in cls.values() if o == "FCS" or q == "FCS")
check("888 total classified games", len(cls) == 888, str(len(cls)))
check("761 FBS-vs-FBS games", fbs_fbs == 761, str(fbs_fbs))
check("127 FCS-involved games", fcs_involved == 127, str(fcs_involved))

sched = wb["IMPORT SCHEDULE"]
check("888 unique schedule games in v0.6", sum(1 for r in range(6, 894) if sched.cell(row=r, column=1).value) == 888, "")
ids = [sched.cell(row=r, column=1).value for r in range(6, 894)]
check("All 888 GameIDs unique", len(set(ids)) == 888, str(len(set(ids))))

# ---------- C. Zero test data / clean state ----------
qb = wb["QB VALUES"]
nb = [(r, c) for r in range(6, 144) for c in [3,4,5,6,8,9,10,11,12] if qb.cell(row=r, column=c).value is not None]
check("138 QB rows fully blank (all UNCERTAIN)", not nb, str(nb[:5]))

ml = wb["MARKET LINES"]
manual_cols = [1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
nb_ml = [(r, c) for r in range(6, 1006) for c in manual_cols if ml.cell(row=r, column=c).value is not None]
check("0 market lines (MARKET LINES manual columns fully blank)", not nb_ml, str(nb_ml[:5]))

adj = wb["ADJUSTMENTS"]
nb_adj = [(r, c) for r in range(6, 256) for c in range(1, 10) if adj.cell(row=r, column=c).value is not None]
check("0 manual adjustments", not nb_adj, str(nb_adj[:5]))

ist = wb["IMPORT STATS"]
nb_ist = [(r, c) for r in range(6, 206) for c in range(1, 10) if ist.cell(row=r, column=c).value is not None]
check("0 imported in-season statistics", not nb_ist, str(nb_ist[:5]))

st = wb["SETTINGS"]
check("BET toggle OFF (SETTINGS!B11 = N)", st["B11"].value == "N", str(st["B11"].value))
check("SETTINGS!B4 (current week) blank (preseason)", st["B4"].value is None, "")
check("SETTINGS!B5 (as-of date) blank (preseason)", st["B5"].value is None, "")

hist = wb["HISTORY"]
nb_hist = [(r, c) for r in range(6, 3006) for c in range(1, 5) if hist.cell(row=r, column=c).value is not None]
check("0 HISTORY snapshots", not nb_hist, str(nb_hist[:5]))

for row, ab in [(8, "G1"), (9, "G3"), (14, "G6")]:
    v = sched.cell(row=row, column=10).value
    check(f"IMPORT SCHEDULE row {row} away_points blank (no synthetic score)", v is None, str(v))

# ---------- D. No remaining TEST/SYNTHETIC/fixture data anywhere ----------
import re
stale = []
for name in wb.sheetnames:
    for row in wb[name].iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, ArrayFormula): v = v.text
            if isinstance(v, str) and re.search(r'TEST ONLY|SYNTHETIC|Nonexistent Fictional', v, re.I):
                # allow the CHANGELOG's own description of the test process
                if name == "CHANGELOG":
                    continue
                stale.append((name, cell.coordinate))
check("No remaining TEST/SYNTHETIC data outside CHANGELOG's own description", not stale, str(stale[:10]))

# ---------- E. NDSU/Sacramento State unaffected ----------
tm = wb["TEAM MAP"]
check("SAC still FBS-RECLASSIFYING", tm.cell(row=114, column=4).value == "FBS-RECLASSIFYING", "")
check("NDSU still FBS-RECLASSIFYING", tm.cell(row=122, column=4).value == "FBS-RECLASSIFYING", "")
tr = wb["TEAM RATINGS"]
check("TEAM RATINGS!X114 (SAC review-cleared) blank", tr.cell(row=114, column=24).value is None, "")
check("TEAM RATINGS!X122 (NDSU review-cleared) blank", tr.cell(row=122, column=24).value is None, "")

# ---------- F. CHANGELOG structural ----------
cl = wb["CHANGELOG"]
r = 7; counts = {}
while cl.cell(row=r, column=1).value is not None:
    counts[cl.cell(row=r, column=1).value] = counts.get(cl.cell(row=r, column=1).value, 0) + 1
    r += 1
check("CHANGELOG has 7 v0.6 entries", counts.get("v0.6") == 7, str(counts.get("v0.6")))
check("CHANGELOG v0.5.2 entry count unchanged (1)", counts.get("v0.5.2") == 1, str(counts.get("v0.5.2")))

print()
if failures:
    print(f"*** {len(failures)} FAILURES ***")
    for f_ in failures: print("   ", f_)
    raise SystemExit(1)
print("ALL CHECKS PASSED")

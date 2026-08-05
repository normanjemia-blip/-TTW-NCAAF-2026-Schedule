"""
v0.5 validation battery (Phase 5 vs approved v0.4.2):
 A. FCS named-override revalidation: independently re-derive the SRS
    ratings from the raw 2025 game archive and compare to FCS TIERS!J:M.
 B. FCS tier-value revalidation: independently recompute the quartile
    fallback and compare to FCS TIERS!A:D.
 C. No NDSU/Sacramento State rows in the FCS override table.
 D. Full cell/formula diff v0.5 vs v0.3.1's already-approved v0.4.2
    baseline -> TSV; ZERO formula cells changed.
 E. Blank assertions: TTW/VSiN inputs, QB VALUES manual columns, MARKET
    LINES (item: no market lines entered), TEAM RATINGS!X114/X122 blank.
 F. SP+/FPI/TR (v0.4.2 baseline) unchanged.
"""
import datetime
import json
import statistics

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

P4, P5 = "../phase4", "../phase5"
failures = []

def check(label, ok, detail=""):
    print(("PASS  " if ok else "FAIL  ") + label + (f"  {detail}" if detail else ""))
    if not ok:
        failures.append((label, detail))

wb = openpyxl.load_workbook("v0.5_working.xlsx", data_only=False)
ft = wb["FCS TIERS"]

# ---------- A. Independent re-derivation of the SRS ratings ----------
FBS = {"1", "4", "5", "8", "9", "12", "15", "17", "18", "151", "37"}
HFA, CAP = 2.5, 28
games = json.load(open(f"{P5}/espn2025_all_games.json"))

def is_fbs(t): return t.get("conferenceId") in FBS

edges, gc = {}, {}
for g in games.values():
    if not g["completed"]: continue
    t = g["teams"]
    if "home" not in t or "away" not in t: continue
    h, a = t["home"], t["away"]
    if h.get("score") is None or a.get("score") is None: continue
    hm = int(h["score"]) - int(a["score"])
    if not g["neutral"]: hm -= HFA
    hm = max(-CAP, min(CAP, hm))
    edges.setdefault(h["name"], []).append((a["name"], hm))
    edges.setdefault(a["name"], []).append((h["name"], -hm))
    gc[h["name"]] = gc.get(h["name"], 0) + 1
    gc[a["name"]] = gc.get(a["name"], 0) + 1

fbs_flag = {}
for g in games.values():
    for side in g.get("teams", {}).values():
        fbs_flag[side["name"]] = is_fbs(side)

rating = {t: 0.0 for t in edges}
for _ in range(500):
    new_r = {t: statistics.mean(rating[o] + m for o, m in edges[t]) for t in edges}
    delta = max(abs(new_r[t] - rating[t]) for t in edges)
    rating = new_r
    if delta < 1e-10: break
shift = statistics.mean(r for t, r in rating.items() if fbs_flag.get(t))
rating = {t: r - shift for t, r in rating.items()}

targets = json.load(open(f"{P4}/fcs_opponents.json"))
override_rows = {}
for r in range(6, 106):
    name = ft.cell(row=r, column=10).value
    if name is None: continue
    override_rows[name] = {
        "rating": ft.cell(row=r, column=11).value,
        "source": ft.cell(row=r, column=12).value,
        "date": ft.cell(row=r, column=13).value,
    }
check("FCS TIERS J:M has exactly 100 named-override rows", len(override_rows) == 100, str(len(override_rows)))
check("Named-override set == the 100 2026-schedule FCS opponents",
      set(override_rows) == set(targets), str(set(override_rows) ^ set(targets)))

mismatch = []
for name, rec in override_rows.items():
    recomputed = round(rating[name], 2)
    if abs(recomputed - rec["rating"]) > 0.01:
        mismatch.append((name, rec["rating"], recomputed))
check("FCS named-override ratings match independent re-derivation (tol 0.01)", not mismatch, str(mismatch[:5]))
bad_date = [n for n, r in override_rows.items()
            if not (isinstance(r["date"], (datetime.date, datetime.datetime)) and r["date"].strftime("%Y-%m-%d") == "2026-07-20")]
check("All 100 override dates == 2026-07-20", not bad_date, str(bad_date[:5]))

# ---------- B. Tier fallback revalidation ----------
rated_nonfbs = sorted([(t, rating[t]) for t in rating if not fbs_flag.get(t) and gc.get(t, 0) >= 6],
                       key=lambda kv: -kv[1])
n = len(rated_nonfbs); q = n // 4
quart = {"Elite FCS": rated_nonfbs[0:q], "Strong FCS": rated_nonfbs[q:2*q],
         "Average FCS": rated_nonfbs[2*q:3*q], "Weak FCS": rated_nonfbs[3*q:]}
recomputed_tiers = {k: round(statistics.mean(v for _, v in members), 1) for k, members in quart.items()}

tier_order = ["Elite FCS", "Strong FCS", "Average FCS", "Weak FCS", "Transitional/uncertain"]
labels_ok = [ft.cell(row=6+i, column=1).value for i in range(5)] == tier_order
check("FCS TIERS!A6:A10 labels unchanged", labels_ok, str([ft.cell(row=6+i, column=1).value for i in range(5)]))
tier_mismatch = []
for i, tname in enumerate(tier_order[:4]):
    v = ft.cell(row=6+i, column=2).value
    if abs(v - recomputed_tiers[tname]) > 0.01:
        tier_mismatch.append((tname, v, recomputed_tiers[tname]))
check("FCS tier fallback values match independent re-derivation", not tier_mismatch, str(tier_mismatch))
check("Transitional/uncertain tier value left blank", ft.cell(row=10, column=2).value is None, "")

# ---------- C. NDSU/Sac State excluded from FCS framework ----------
excluded_ok = not any(("North Dakota State" in n or "Sacramento" in n) for n in override_rows)
check("NDSU/Sacramento State NOT present in FCS named-override table", excluded_ok, "")

# ---------- D. Full diff vs v0.4.2 (which itself sits on the approved v0.3.1 base) ----------
def normval(v):
    if isinstance(v, ArrayFormula): return ("ARRAYFORMULA", v.ref, v.text)
    return v

wb_base = openpyxl.load_workbook("v0.4.2_working.xlsx", data_only=False)
diff_lines, diff_counts, formula_diffs = [], {}, []
for name in wb_base.sheetnames:
    ws_old, ws_new = wb_base[name], wb[name]
    mr = max(ws_old.max_row, ws_new.max_row); mc = max(ws_old.max_column, ws_new.max_column)
    cnt = 0
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            vo = normval(ws_old.cell(row=r, column=c).value)
            vn = normval(ws_new.cell(row=r, column=c).value)
            if vo != vn:
                cnt += 1
                diff_lines.append(f"{name}\t{ws_new.cell(row=r, column=c).coordinate}\t{repr(vo)}\t{repr(vn)}")
                if isinstance(vo, tuple) or (isinstance(vo, str) and vo.startswith("=")):
                    formula_diffs.append((name, ws_new.cell(row=r, column=c).coordinate))
    if cnt: diff_counts[name] = cnt
with open("v0.5_vs_v0.4.2_full_diff.tsv", "w") as f:
    f.write("sheet\tcell\tv0.4.2_value\tv0.5_value\n" + "\n".join(diff_lines) + "\n")
print("\nDiff counts (v0.5 vs v0.4.2):", diff_counts)
check("ZERO formula cells changed vs v0.4.2", not formula_diffs, str(formula_diffs[:8]))
expected_sheets = {"START HERE", "FCS TIERS", "DICTIONARY", "CHANGELOG"}
check("Diff touches only expected sheets", set(diff_counts) == expected_sheets, str(set(diff_counts) ^ expected_sheets))
check("FCS TIERS diff == 413 (400 override cells [100x4] + 12 tier value/source/date [4x3] + A2 text)",
      diff_counts.get("FCS TIERS") == 413, str(diff_counts.get("FCS TIERS")))
check("START HERE diff == 2 (banner + A3)", diff_counts.get("START HERE") == 2, str(diff_counts.get("START HERE")))
check("DICTIONARY diff == 2 (A21, B21)", diff_counts.get("DICTIONARY") == 2, str(diff_counts.get("DICTIONARY")))
check("CHANGELOG diff == 20 (5 new rows x 4)", diff_counts.get("CHANGELOG") == 20, str(diff_counts.get("CHANGELOG")))

# also diff vs the ORIGINAL approved v0.3.1 to reconfirm v0.4/v0.4.1/v0.4.2 layers untouched
wb_v031 = openpyxl.load_workbook("v0.3.1_with_schedule.xlsx", data_only=False)
formula_diffs_v031 = []
for name in wb_v031.sheetnames:
    ws_old, ws_new = wb_v031[name], wb[name]
    mr = max(ws_old.max_row, ws_new.max_row); mc = max(ws_old.max_column, ws_new.max_column)
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            vo = normval(ws_old.cell(row=r, column=c).value)
            if isinstance(vo, tuple) or (isinstance(vo, str) and vo.startswith("=")):
                vn = normval(ws_new.cell(row=r, column=c).value)
                if vo != vn:
                    formula_diffs_v031.append((name, ws_new.cell(row=r, column=c).coordinate))
check("ZERO formula cells changed vs original approved v0.3.1 (full chain)", not formula_diffs_v031, str(formula_diffs_v031[:8]))

# ---------- E. Blank assertions ----------
ps = wb["PRESEASON"]
for label, cols in {"TTW raw/date/cite (L,N,O)": [12, 14, 15], "VSiN raw/date/cite (U,V,W)": [21, 22, 23]}.items():
    nb = [(r, c) for r in range(6, 144) for c in cols if ps.cell(row=r, column=c).value is not None]
    check(f"PRESEASON {label} fully blank", not nb, str(nb[:5]))
qb = wb["QB VALUES"]
nb = [(r, c) for r in range(6, 144) for c in [3,4,5,6,8,9,10,11,12] if qb.cell(row=r, column=c).value is not None]
check("QB VALUES manual-input columns fully blank (all 138 UNCERTAIN)", not nb, str(nb[:5]))
ml = wb["MARKET LINES"]
manual_cols = [1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]  # excludes B (formula auto-lookup)
nb_ml = [(r, c) for r in range(6, 1006) for c in manual_cols if ml.cell(row=r, column=c).value is not None]
check("MARKET LINES manual-input columns fully blank (no lines entered, per Phase 5 scope limit)", not nb_ml, str(nb_ml[:5]))
tr = wb["TEAM RATINGS"]
check("TEAM RATINGS!X114 (SAC transition review cleared) still blank", tr.cell(row=114, column=24).value is None, "")
check("TEAM RATINGS!X122 (NDSU transition review cleared) still blank", tr.cell(row=122, column=24).value is None, "")

# ---------- F. v0.4.2 sources unchanged ----------
bad_d = [(r,) for r in range(6,144) if ps.cell(row=r, column=4).value != wb_base["PRESEASON"].cell(row=r, column=4).value]
bad_h = [(r,) for r in range(6,144) if ps.cell(row=r, column=8).value != wb_base["PRESEASON"].cell(row=r, column=8).value]
bad_q = [(r,) for r in range(6,144) if ps.cell(row=r, column=17).value != wb_base["PRESEASON"].cell(row=r, column=17).value]
check("SP+ (D) unchanged from v0.4.2", not bad_d, str(bad_d[:5]))
check("FPI (H) unchanged from v0.4.2", not bad_h, str(bad_h[:5]))
check("TeamRankings (Q) unchanged from v0.4.2", not bad_q, str(bad_q[:5]))

# ---------- structural ----------
sched = wb["IMPORT SCHEDULE"]
check("IMPORT SCHEDULE 888 game ids intact", sum(1 for r in range(6,894) if sched.cell(row=r, column=1).value) == 888, "")
cl = wb["CHANGELOG"]
r = 7; counts = {}
while cl.cell(row=r, column=1).value is not None:
    counts[cl.cell(row=r, column=1).value] = counts.get(cl.cell(row=r, column=1).value, 0) + 1
    r += 1
check("CHANGELOG has 5 v0.5 entries", counts.get("v0.5") == 5, str(counts.get("v0.5")))

print()
if failures:
    print(f"*** {len(failures)} FAILURES ***")
    for f_ in failures: print("   ", f_)
    raise SystemExit(1)
print("ALL CHECKS PASSED")

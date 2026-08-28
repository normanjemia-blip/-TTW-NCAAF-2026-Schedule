"""
v0.6.1 correction build from v0.6 (v0.5.2 and v0.6 untouched).

Fixes:
 1. ADJUSTMENTS!K6:K255 - the user's exact corrected Flags formula
    (fixes the #VALUE! propagation defect found in Phase 6).
 2. ADJUSTMENTS!J6:J255 - Effective now requires K to be blank in
    addition to Active=Y and not-expired, so a flagged row (missing
    reason, non-numeric value, or oversized non-override adjustment)
    never flows into ENGINE calculations.
Every row's exact pre-edit formula is asserted before being overwritten.
"""
import openpyxl

FILE = "v0.6.1_working.xlsx"
wb = openpyxl.load_workbook(FILE, data_only=False)
adj = wb["ADJUSTMENTS"]

K_OLD = ('=IF($A{r}="","",IF($D{r}="","REASON MISSING; ","")&'
         'IF(NOT(ISNUMBER($C{r})),"VALUE NOT NUMERIC; ","")&'
         'IF(AND(ISNUMBER($C{r}),ABS($C{r})>4,$A{r}<>"MARGIN OVERRIDE",$A{r}<>"TOTAL OVERRIDE"),'
         '"LARGE ADJ (>4); ",""))')
K_NEW = ('=IF($A{r}="","",IF($D{r}="","REASON MISSING; ","")&'
         'IF(NOT(ISNUMBER($C{r})),"VALUE NOT NUMERIC; ","")&'
         'IF(ISNUMBER($C{r}),IF(AND(ABS($C{r})>4,$A{r}<>"MARGIN OVERRIDE",$A{r}<>"TOTAL OVERRIDE"),'
         '"LARGE ADJ (>4); ",""),""))')

J_OLD = ('=IF($A{r}="","",IF(AND($H{r}="Y",OR($G{r}="",SETTINGS!$B$5="",$G{r}>=SETTINGS!$B$5)),1,0))')
J_NEW = ('=IF($A{r}="","",IF(AND($H{r}="Y",OR($G{r}="",SETTINGS!$B$5="",$G{r}>=SETTINGS!$B$5),'
         '$K{r}=""),1,0))')

k_count, j_count, mismatches = 0, 0, []
for r in range(6, 256):
    ck = adj.cell(row=r, column=11)
    expected_k = K_OLD.format(r=r)
    if ck.value != expected_k:
        mismatches.append(("K", r, ck.value)); continue
    ck.value = K_NEW.format(r=r)
    k_count += 1

    cj = adj.cell(row=r, column=10)
    expected_j = J_OLD.format(r=r)
    if cj.value != expected_j:
        mismatches.append(("J", r, cj.value)); continue
    cj.value = J_NEW.format(r=r)
    j_count += 1

print(f"K corrected: {k_count} rows (6-255)")
print(f"J corrected: {j_count} rows (6-255)")
if mismatches:
    print(f"*** {len(mismatches)} MISMATCHES (aborting) ***")
    for m in mismatches[:5]:
        print("  ", m)
    raise SystemExit(1)
assert k_count == 250 and j_count == 250

wb.save(FILE)
print("Saved", FILE)

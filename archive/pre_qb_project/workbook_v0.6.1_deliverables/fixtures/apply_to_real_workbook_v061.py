"""
Apply the consolidated Phase 6.1 test-fixture set to a REAL, full-scale
copy of the workbook (v0.6.1_test_applied.xlsx). Combines the original
v0.6 line-entry fixtures (kept for continuity) with the new v0.6.1
additions: a genuine BET-toggle fixture (G1 driven to true READY with
|edge|>=3.0) and NDSU's 5 simulated completed FBS games (functional
transitional-restriction proof). The isolated DATA INCOMPLETE
remove/restore test is demonstrated only in the harness (build_harness.py
output) - by definition a temporary single-field removal that gets
restored, it doesn't fit into a static "applied" snapshot alongside
everything else.

Every synthetic value is tagged TEST ONLY — NOT REAL DATA.
"""
import datetime
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

TEST_TAG = "TEST ONLY — NOT REAL DATA"
FILE = "v0.6.1_test_applied.xlsx"
wb = openpyxl.load_workbook(FILE, data_only=False)

st = wb["SETTINGS"]
st["B4"] = 2
st["B5"] = datetime.date(2026, 9, 4)
st["B11"] = "N"  # BET toggle OFF for the main applied snapshot

ml = wb["MARKET LINES"]
def set_line(row, gid, favorite, spread, total, line_date, notes_extra=""):
    ml.cell(row=row, column=1, value=gid)
    ml.cell(row=row, column=3, value=favorite)
    ml.cell(row=row, column=4, value=spread)
    if total is not None:
        ml.cell(row=row, column=5, value=total)
    ml.cell(row=row, column=6, value=TEST_TAG)
    ml.cell(row=row, column=7, value=line_date)
    ml.cell(row=row, column=8, value=f"{TEST_TAG}{(': ' + notes_extra) if notes_extra else ''}")

set_line(8, "401858202", "UVA", 6.5, 51.5, datetime.date(2026, 9, 3))
set_line(6, "401856766", "TCU", 3.0, 55.0, datetime.date(2026, 8, 20), "deliberately stale line date")
set_line(10, "401864570", "FSU", 24.5, 58.5, datetime.date(2026, 9, 3))
set_line(11, "401864577", "NDSU", 3.0, 45.0, datetime.date(2026, 9, 3))
set_line(14, "401856767", "UCF", 45.0, 60.0, datetime.date(2026, 9, 3), "deliberately entered on an FCS game to test that it CANNOT override FCS — NO PLAY")
set_line(15, "401858204", "Ohio State", 10.0, None, datetime.date(2026, 9, 3), "deliberately invalid favorite (not a team in this game) + missing total")

adj = wb["ADJUSTMENTS"]
adj.cell(row=6, column=1, value="GAME")
adj.cell(row=6, column=2, value="401864570")
adj.cell(row=6, column=3, value=1.5)
adj.cell(row=6, column=4, value=f"{TEST_TAG}: validating manual adjustment pipeline")
adj.cell(row=6, column=5, value=TEST_TAG)
adj.cell(row=6, column=6, value=datetime.date(2026, 9, 1))
adj.cell(row=6, column=8, value="Y")
adj.cell(row=6, column=9, value=TEST_TAG)
# NEW: isolated oversized adjustment on a second row, to show the v0.6.1
# fix live in the real workbook (Effective should compute 0)
adj.cell(row=7, column=1, value="GAME")
adj.cell(row=7, column=2, value="401856766")  # G2, not used by any other adjustment
adj.cell(row=7, column=3, value=999)
adj.cell(row=7, column=4, value=f"{TEST_TAG}: oversized adjustment - v0.6.1 fix should force Effective=0")
adj.cell(row=7, column=5, value=TEST_TAG)
adj.cell(row=7, column=6, value=datetime.date(2026, 9, 1))
adj.cell(row=7, column=8, value="Y")
adj.cell(row=7, column=9, value=TEST_TAG)

qb = wb["QB VALUES"]
for row in (64, 70, 97, 122):  # NCST, UVA, JVST, NDSU
    qb.cell(row=row, column=3, value=f"{TEST_TAG}: Test Baseline QB")
    qb.cell(row=row, column=4, value=0.0)
    qb.cell(row=row, column=5, value=f"{TEST_TAG}: Test Active QB")
    qb.cell(row=row, column=6, value=0.0)
    qb.cell(row=row, column=8, value="M")
    qb.cell(row=row, column=9, value=TEST_TAG)
    qb.cell(row=row, column=10, value=2026)
    qb.cell(row=row, column=11, value=datetime.date(2026, 9, 1))
    qb.cell(row=row, column=12, value=f"{TEST_TAG}: temporarily resolved for Phase 6.1 QB-gate/transitional tests")

sched = wb["IMPORT SCHEDULE"]
for row, away_pts, home_pts in [(8, 17, 24), (9, 10, 41), (14, 6, 45)]:
    sched.cell(row=row, column=10, value=away_pts)
    sched.cell(row=row, column=11, value=home_pts)
    sched.cell(row=row, column=12, value=True)
    sched.cell(row=row, column=14, value=f"{TEST_TAG}: synthetic completed-game score for Phase 6 weekly-workflow test")

# NEW (v0.6.1): NDSU's 5 simulated completed FBS games this season, on
# previously-unused rows 20-24
opponents = ["Air Force Falcons", "UNLV Rebels", "New Mexico Lobos", "San José State Spartans", "UTEP Miners"]
for i, opp in enumerate(opponents):
    r = 20 + i
    gid = f"90000000{i}"
    sched.cell(row=r, column=1, value=gid)
    sched.cell(row=r, column=2, value=2026)
    sched.cell(row=r, column=3, value=1 + i)
    sched.cell(row=r, column=4, value=datetime.date(2026, 9, 5 + i))
    sched.cell(row=r, column=5, value=False)
    sched.cell(row=r, column=6, value="North Dakota State Bison")
    sched.cell(row=r, column=7, value="Mountain West")
    sched.cell(row=r, column=8, value=opp)
    sched.cell(row=r, column=9, value="American Conference" if "Air Force" in opp else "Mountain West Conference")
    sched.cell(row=r, column=10, value=17)
    sched.cell(row=r, column=11, value=24)
    sched.cell(row=r, column=12, value=True)
    sched.cell(row=r, column=14, value=f"{TEST_TAG}: synthetic completed game for NDSU games-played functional test")

ist = wb["IMPORT STATS"]
ist_rows = [
    ("NC State", 1, 0.15, -0.05, 0.44, 0.38, 0.35, 0.28, 68.0),
    ("Virginia", 1, 0.08, -0.02, 0.41, 0.40, 0.30, 0.29, 70.0),
    (f"{TEST_TAG}: Nonexistent Fictional University", 1, 0.10, -0.10, 0.40, 0.40, 0.30, 0.30, 69.0),
]
for i, vals in enumerate(ist_rows):
    r = 6 + i
    for c, v in enumerate(vals, start=1):
        ist.cell(row=r, column=c, value=v)

hist = wb["HISTORY"]
hist.cell(row=6, column=1, value=1)
hist.cell(row=6, column=2, value="NCST")
hist.cell(row=6, column=3, value=-20.0)
hist.cell(row=6, column=4, value=TEST_TAG)

sh = wb["START HERE"]
sh["A1"] = "TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.6.1 PHASE 6 — TEST FIXTURES APPLIED — DO NOT USE FOR REAL PICKS)"

wb.save(FILE)
print("Saved", FILE)

tagged = 0
for name in wb.sheetnames:
    for row in wb[name].iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and TEST_TAG in cell.value:
                tagged += 1
print(f"Cells carrying the TEST_TAG marker: {tagged}")

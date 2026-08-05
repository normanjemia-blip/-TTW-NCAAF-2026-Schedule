import copy
import openpyxl

def copy_style(src_cell, dst_cell):
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.number_format = copy.copy(src_cell.number_format)
    dst_cell.protection = copy.copy(src_cell.protection)
    dst_cell.alignment = copy.copy(src_cell.alignment)

FILE = "v0.6.1_working.xlsx"
wb = openpyxl.load_workbook(FILE, data_only=False)
cl = wb["CHANGELOG"]

# --- Correct the v0.6 row-45 overclaim (DATA INCOMPLETE + genuine BET-toggle proof) ---
old_45 = cl.cell(row=45, column=3).value
assert "Confirmed BET labels require toggle=Y AND status=READY together" in old_45
new_45 = old_45.replace(
    "17-scenario line-level test matrix executed against 7 real schedule games "
    "spanning every branch of the status-priority chain (BLOCKED, FCS — NO PLAY, "
    "PENDING LINE, STALE LINE, QB UNCERTAIN, TRANSITION UNCERTAIN, DATA INCOMPLETE, "
    "READY) via a real-Excel-formula engine (not manual trace).",
    "17-scenario line-level test matrix executed against 7 real schedule games "
    "covering 7 of the 8 status-priority branches (BLOCKED, FCS — NO PLAY, PENDING "
    "LINE, STALE LINE, QB UNCERTAIN, TRANSITION UNCERTAIN, READY) via a "
    "real-Excel-formula engine (not manual trace). [CORRECTED in v0.6.1: DATA "
    "INCOMPLETE was NOT genuinely isolated in v0.6 - see v0.6.1's own entries "
    "below, which found it is actually unreachable dead code, a second genuine "
    "defect.]"
).replace(
    "Confirmed BET labels require toggle=Y AND status=READY together (edge >=3.0 "
    "shows INVESTIGATE with toggle=N, BET with toggle=Y).",
    "[CORRECTED in v0.6.1: the v0.6 BET-toggle test used a game that never reached "
    "READY status in either toggle state, so it did NOT actually demonstrate "
    "toggle=Y producing a BET label - it only showed both toggle states producing "
    "INVESTIGATE for an unrelated reason (status). A genuine toggle=N-vs-Y proof on "
    "a truly READY, edge>=3.0 game was completed in v0.6.1 - see below.]"
)
assert new_45 != old_45
cl.cell(row=45, column=3).value = new_45
print("CHANGELOG row 45 (v0.6 overclaim) corrected")

# --- New v0.6.1 CHANGELOG entries ---
next_row = 7
while cl.cell(row=next_row, column=1).value is not None:
    next_row += 1
TODAY = "2026-07-21"

entries = [
    ("v0.6.1", TODAY,
     "CORRECTION (user-authorized): ADJUSTMENTS!K6:K255 (Flags) replaced with "
     "the corrected formula - "
     "=IF($A6=\"\",\"\",IF($D6=\"\",\"REASON MISSING; \",\"\")&"
     "IF(NOT(ISNUMBER($C6)),\"VALUE NOT NUMERIC; \",\"\")&"
     "IF(ISNUMBER($C6),IF(AND(ABS($C6)>4,$A6<>\"MARGIN OVERRIDE\","
     "$A6<>\"TOTAL OVERRIDE\"),\"LARGE ADJ (>4); \",\"\"),\"\")) - so ABS() is "
     "never evaluated on a non-numeric value. Live-calc re-verified: a "
     "non-numeric adjustment value now shows a clean \"VALUE NOT NUMERIC; \" "
     "flag instead of a raw #VALUE! error.",
     "v0.6.1 - defect fix 1 of 2 (ADJUSTMENTS!K)"),
    ("v0.6.1", TODAY,
     "SAFETY CORRECTION (user-authorized): ADJUSTMENTS!J6:J255 (Effective) "
     "changed from =IF($A6=\"\",\"\",IF(AND($H6=\"Y\",OR($G6=\"\","
     "SETTINGS!$B$5=\"\",$G6>=SETTINGS!$B$5)),1,0)) to add a third AND "
     "condition, $K6=\"\" - so a flagged row (missing reason, non-numeric "
     "value, or oversized non-override adjustment) is never Effective, "
     "regardless of Active/expiry. Live-calc re-verified: an isolated +999 "
     "adjustment now shows Effective=0 and contributes $0.0 to ENGINE!O "
     "(was $999.0, fully applied, before this fix); an isolated non-numeric "
     "adjustment likewise shows Effective=0. Confirmed MARGIN OVERRIDE / "
     "TOTAL OVERRIDE rows remain exempt from the LARGE ADJ flag and stay "
     "Effective=1 even at large magnitudes, as designed - live-calc tested "
     "with a 15.0-point MARGIN OVERRIDE row (K blank, J=1, ENGINE!R correctly "
     "adopts the override value).",
     "v0.6.1 - safety fix 2 of 2 (ADJUSTMENTS!J)"),
    ("v0.6.1", TODAY,
     "SECOND GENUINE FORMULA DEFECT FOUND AND DOCUMENTED, NOT REPAIRED (found "
     "while completing the isolated DATA INCOMPLETE test this build requested; "
     "not authorized for repair this round): CLEAN!C6:C1005 (week) and "
     "CLEAN!D6:D1005 (date), pattern =IF($A6=\"\",\"\",'IMPORT SCHEDULE'!C6) - "
     "when the source cell is genuinely blank, a bare cell reference evaluates "
     "to 0 in Excel/Sheets (confirmed with isolated formula probes matching "
     "standard spreadsheet semantics, not a verification-tool artifact), not "
     "\"\". Since ENGINE!AI's DATA INCOMPLETE check is OR($B6=\"\",$C6=\"\"), and "
     "$B6/$C6 trace back through this pass-through, DATA INCOMPLETE can NEVER "
     "fire for a row with a valid GameID - it is unreachable dead code. A row "
     "with a genuinely missing week silently displays as Week 0 (indistinguishable "
     "from a real Week 0 game); a missing date silently displays the spreadsheet "
     "epoch date. Verified directly: temporarily blanking IMPORT SCHEDULE!C8 "
     "(NC State @ Virginia's week field) still produced STATUS=READY, not DATA "
     "INCOMPLETE; the field was restored to its exact original value (0) "
     "immediately after. Proposed correction (documented only, not applied): "
     "nest an explicit blank check, e.g. CLEAN!C6 -> "
     "=IF($A6=\"\",\"\",IF('IMPORT SCHEDULE'!C6=\"\",\"\",'IMPORT SCHEDULE'!C6)), "
     "same pattern for CLEAN!D6 - this would let a genuinely blank source "
     "propagate as \"\", which is all ENGINE!AI's existing DATA INCOMPLETE check "
     "needs. No v0.6.2 was created for this; only the two ADJUSTMENTS fixes "
     "explicitly authorized this round were applied.",
     "v0.6.1 - DEFECT FOUND: DATA INCOMPLETE unreachable (CLEAN!C/D blank-coercion)"),
    ("v0.6.1", TODAY,
     "Status-priority chain re-verified against this build's formulas: BLOCKED, "
     "FCS — NO PLAY, PENDING LINE, STALE LINE, QB UNCERTAIN, TRANSITION "
     "UNCERTAIN, and READY all confirmed correct (7 of 8; DATA INCOMPLETE could "
     "not be positively demonstrated - see the defect entry above). A genuine "
     "BET-toggle proof was completed on a single fixture (NC State @ Virginia) "
     "deliberately driven to true READY status with a calculated |edge|=4.33 "
     "(>=3.0): toggle=N produced INVESTIGATE, toggle=Y produced BET, on the "
     "identical fixture with only the toggle changed - unlike the v0.6 attempt "
     "(corrected above), this is a genuine demonstration.",
     "v0.6.1 - status chain + genuine BET-toggle proof (items 3-4)"),
    ("v0.6.1", TODAY,
     "Weekly-workflow validation strengthened: all 10 populated prior-fade rows "
     "(SETTINGS!C37:C46, effective games 0 through 9) verified via live "
     "calculation of the real TEAM RATINGS!G formula with synthetic effective-"
     "game inputs, not just the 0- and 1-game rows tested in v0.6 - all 10 "
     "match exactly, plus confirmed F=10 correctly clamps to the F=9 floor "
     "weight (0.1). NDSU's transitional restriction re-tested functionally, "
     "not just by label: simulated 5 completed FBS games this season for NDSU "
     "(exceeding the 4-game manual-review minimum, SETTINGS!B25), confirmed "
     "TEAM RATINGS!X122 (review-cleared) remains genuinely unset regardless "
     "(no formula anywhere auto-populates it), and confirmed that even with a "
     "fresh valid line, resolved QBs, BET toggle=Y, and a very large calculated "
     "edge (11.46), NDSU's game stays STATUS=TRANSITION UNCERTAIN and its label "
     "stays capped at INVESTIGATE, never BET - the substantive restriction "
     "holds regardless of games played, edge size, or toggle state, not merely "
     "the D122=\"Y\" label.",
     "v0.6.1 - weekly-workflow strengthened (item 5)"),
    ("v0.6.1", TODAY,
     "iPad checklist corrected: QB VALUES step now lists columns I (Source) "
     "and K (Last update) alongside C/D/E/F/H/J; added a note that DATA "
     "QUALITY's QB UNCERTAIN count can read 0 while games are still PENDING "
     "LINE, because PENDING LINE outranks QB UNCERTAIN and masks it from that "
     "count (the count only reflects ENGINE!AI values that actually resolved "
     "to QB UNCERTAIN); replaced the phrase \"home underdog favorite\" with a "
     "direct instruction that the spread is always entered as a positive "
     "number regardless of whether the favorite is home or away.",
     "v0.6.1 - iPad checklist corrections (item 6)"),
]
for i, (ver, date, change, reason) in enumerate(entries):
    r = next_row + i
    cl.cell(row=r, column=1, value=ver)
    cl.cell(row=r, column=2, value=date)
    cl.cell(row=r, column=3, value=change)
    cl.cell(row=r, column=4, value=reason)
    for c in range(1, 5):
        copy_style(cl.cell(row=next_row - 1, column=c), cl.cell(row=r, column=c))
print(f"CHANGELOG: added {len(entries)} v0.6.1 entries starting at row {next_row}")

sh = wb["START HERE"]
sh["A1"] = "TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.6.1 PHASE 6 CORRECTION + TEST-COMPLETION BUILD — PENDING APPROVAL)"

wb.save(FILE)
print("Saved", FILE)

"""
Full prior-fade table test: exact real TEAM RATINGS!G formula, fed
synthetic F (effective games) values 0 through 10, verified against
SETTINGS!C37:C46 directly.
"""
import openpyxl

G_FORMULA = ('=IF(OR($A{r}="",$F{r}=""),"",IF(MIN($F{r},9)>=9,SETTINGS!$C$46,'
             'INDEX(SETTINGS!$C$37:$C$46,INT(MIN($F{r},9))+1)+(MIN($F{r},9)-INT(MIN($F{r},9)))*'
             '(INDEX(SETTINGS!$C$37:$C$46,INT(MIN($F{r},9))+2)-INDEX(SETTINGS!$C$37:$C$46,INT(MIN($F{r},9))+1))))')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TEST"
st = wb.create_sheet("SETTINGS")
fade = {0: 1, 1: 0.8, 2: 0.65, 3: 0.5, 4: 0.4, 5: 0.3, 6: 0.225, 7: 0.175, 8: 0.125, 9: 0.1}
for i, (eg, w) in enumerate(fade.items()):
    st.cell(row=37 + i, column=3, value=w)

for i, f_val in enumerate(range(0, 11)):
    r = 6 + i
    ws.cell(row=r, column=1, value=f"T{r}")
    ws.cell(row=r, column=6, value=f_val)
    ws.cell(row=r, column=7, value=G_FORMULA.format(r=r))

wb.save("fade_table_test.xlsx")
print("saved, testing F=0..10")

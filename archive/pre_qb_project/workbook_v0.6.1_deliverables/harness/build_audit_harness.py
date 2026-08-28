"""
Live-calc proof for the 9 "cheap" AUDIT invariants (rows 6,7,8,9,12,13,14,
15,16,17,18) that don't depend on the full CLEAN/ENGINE 1000-row pipeline -
built the same way as every other harness this project uses: real formula
text extracted byte-exact from the pristine v0.6.1_working.xlsx at its
original row/column position, run against the REAL full-scale source data
(TEAM MAP 481 rows, PRESEASON 143 rows, SETTINGS, MARKET LINES 1000 rows)
copied verbatim, then evaluated by the `formulas` engine.

Rows 10 and 19 need the full CLEAN/ENGINE chain (confirmed too costly for
full-scale live recalc in this sandbox, per Phase 3/6 findings) and are
instead proven by direct structural/data argument in validate_v061.py.
"""
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

SRC = "v0.6.1_working.xlsx"
OUT = "audit_harness.xlsx"
src = openpyxl.load_workbook(SRC, data_only=False)

wb = openpyxl.Workbook()
wb.remove(wb.active)

def copy_range(sheet_name, col_letters, max_row, header_row=None):
    ws_src = src[sheet_name]
    ws = wb.create_sheet(sheet_name) if sheet_name not in wb.sheetnames else wb[sheet_name]
    rows = list(range(1, max_row + 1))
    if header_row:
        rows = [header_row] + [r for r in rows if r != header_row]
    for col in col_letters:
        c_idx = openpyxl.utils.column_index_from_string(col)
        for r in rows:
            v = ws_src.cell(row=r, column=c_idx).value
            ws.cell(row=r, column=c_idx).value = v

copy_range("TEAM MAP", ["A", "H", "J", "K"], 481)
copy_range("PRESEASON", ["Z"], 143)
copy_range("SETTINGS", ["B"], 12)
copy_range("MARKET LINES", ["A"], 1005)

au_src = src["AUDIT"]
au = wb.create_sheet("AUDIT")
for r in (6, 7, 8, 9, 12, 13, 14, 15, 16, 17, 18):
    v = au_src.cell(row=r, column=2).value
    if isinstance(v, ArrayFormula):
        au.cell(row=r, column=2).value = ArrayFormula(ref=f"B{r}", text=v.text)
    else:
        au.cell(row=r, column=2).value = v

wb.save(OUT)
print("Saved", OUT)

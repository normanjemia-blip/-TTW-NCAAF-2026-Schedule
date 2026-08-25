#!/usr/bin/env python3
"""v0.8.6 promotion certificate — READ-ONLY. Writes nothing.

v0.8.6 = v0.8.5 + supplemental packet items 1-3 ONLY: 18 cells, 4 of them the
approved zeros. Zero formula changes. Zero model-output changes.

Exit code 0 iff every check passes.
"""
import collections, hashlib, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
V085 = os.path.join(ROOT, "promotion_v0.8.5",
                    "TTW_College_Football_Power_Ratings_v0.8.5_AUTHORITATIVE.xlsx")
V086 = os.path.join(ROOT, "promotion_v0.8.6",
                    "TTW_College_Football_Power_Ratings_v0.8.6_AUTHORITATIVE.xlsx")
FROZEN_V085_SHA = "0676aa1a05d661ca0d99c917c8dc471c0030128cc42ea8fd1bd2f17dcea767be"

ACTIVATED = {35: "RUTG", 80: "WSU"}
CORRECTED = {74: "CSU"}
MUST_NOT_MOVE = ("TTU", "STAN", "NIU", "TULN", "FRES", "SYR", "ALA", "TENN", "GASO")

PASS, FAIL = [], []


def chk(msg, ok, detail=""):
    (PASS if ok else FAIL).append(msg)
    print(f"  [{'PASS' if ok else 'FAIL'}] {msg}" + (f" - {detail}" if detail else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def mean(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def main():
    print("=" * 78)
    print("v0.8.6 PROMOTION CERTIFICATE — supplemental packet items 1-3")
    print("=" * 78)

    print("\n1. PREDECESSOR AND IDENTITY")
    h85, h86 = sha256(V085), sha256(V086)
    print(f"  v0.8.5 SHA-256: {h85}")
    print(f"  v0.8.6 SHA-256: {h86}")
    chk("1.1 v0.8.5 retains its frozen SHA-256", h85 == FROZEN_V085_SHA, h85[:16])
    chk("1.2 v0.8.6 differs from v0.8.5", h86 != h85)

    a = openpyxl.load_workbook(V085)
    b = openpyxl.load_workbook(V086)
    chk("1.3 21 sheets, identical names and order",
        a.sheetnames == b.sheetnames and len(b.sheetnames) == 21, f"{len(b.sheetnames)}")
    chk("1.4 sheet visibility preserved",
        {s: a[s].sheet_state for s in a.sheetnames} == {s: b[s].sheet_state for s in b.sheetnames})
    chk("1.5 named ranges preserved",
        sorted(a.defined_names.keys()) == sorted(b.defined_names.keys()))

    changed, formula_changed = [], []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                x, y = norm(wa.cell(row=r, column=c).value), norm(wbk.cell(row=r, column=c).value)
                if x != y:
                    changed.append((s, r, wbk.cell(row=r, column=c).coordinate))
                    if x[0] == "F" or y[0] == "F":
                        formula_changed.append((s, wbk.cell(row=r, column=c).coordinate))
    chk("1.6 ZERO formula differences", not formula_changed, str(formula_changed))
    bysheet = collections.Counter(s for s, _, _ in changed)
    chk("1.7 only START HERE (1) and QB VALUES (17) changed",
        dict(bysheet) == {"START HERE": 1, "QB VALUES": 17}, str(dict(bysheet)))
    chk("1.8 total changed cells = 18", len(changed) == 18, str(len(changed)))
    touched = {r for s, r, _ in changed if s == "QB VALUES"}
    expect = set(ACTIVATED) | set(CORRECTED)
    chk("1.9 exactly the 3 approved QB rows were touched",
        touched == expect, f"touched={sorted(touched)} expected={sorted(expect)}")

    tm, qb, st, ml, sh = b["TEAM MAP"], b["QB VALUES"], b["SETTINGS"], b["MARKET LINES"], b["START HERE"]
    season = st["B3"].value

    print("\n2. THE FOUR APPROVED ZEROS — AND NO OTHERS")
    zeros, nonzero = [], []
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F = qb.cell(row=r, column=4).value, qb.cell(row=r, column=6).value
        for col, v in (("D", D), ("F", F)):
            if v == 0:
                zeros.append((ab, col))
            elif v is not None:
                nonzero.append((ab, col, v))
    new_zeros = [(ab, c) for ab, c in zeros if ab in ACTIVATED.values()]
    chk("2.1 exactly 4 zeros written on the 2 activated rows",
        len(new_zeros) == 4, f"{len(new_zeros)} {new_zeros}")
    chk("2.2 ZERO nonzero QB values anywhere in the workbook", not nonzero, str(nonzero))
    # Every activated row carries exactly one 0/0 pair and nothing else is numeric,
    # so the workbook-wide zero count must be exactly twice the OK-row count.
    ok_rows = 0
    for r in range(6, 144):
        if not tm.cell(row=r, column=1).value:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        G = "" if (D is None or F is None) else F - D
        if not (G == "" or H == "L" or J != season):
            ok_rows += 1
    chk("2.3 total zeros = 220 = 2 x 110 OK rows (was 216 = 2 x 108 in v0.8.5, so +4)",
        len(zeros) == 220 and len(zeros) == 2 * ok_rows,
        f"zeros={len(zeros)} ok_rows={ok_rows}")
    for row, ab in ACTIVATED.items():
        D, F = qb.cell(row=row, column=4).value, qb.cell(row=row, column=6).value
        chk(f"2.x {ab} r{row}: D=0 and F=0, so G = F - D = 0 exactly",
            D == 0 and F == 0, f"D={D} F={F}")

    print("\n3. PROTECTED CELLS")
    for row, ab in {**ACTIVATED, **CORRECTED}.items():
        g_ok = isf(qb.cell(row=row, column=7).value)
        m_ok = isf(qb.cell(row=row, column=13).value)
        j = qb.cell(row=row, column=10).value
        chk(f"3.x {ab} r{row}: G and M still formulas, J still 2026",
            g_ok and m_ok and j == 2026, f"G={g_ok} M={m_ok} J={j}")

    print("\n4. CENSUSES")
    codes, sts = collections.Counter(), collections.Counter()
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        G = "" if (D is None or F is None) else F - D
        codes[H] += 1
        sts["UNCERTAIN" if (G == "" or H == "L" or J != season) else "OK"] += 1
    chk("4.1 QB status census = 110 OK / 28 UNCERTAIN",
        (sts["OK"], sts["UNCERTAIN"]) == (110, 28), str(dict(sts)))
    chk("4.2 confidence census = 73 H / 40 M / 25 L",
        (codes["H"], codes["M"], codes["L"]) == (73, 40, 25), str(dict(codes)))
    chk("4.3 census totals 138", sum(codes.values()) == 138, str(sum(codes.values())))
    banner = sh["A1"].value
    chk("4.4 banner states v0.8.6 AUTHORITATIVE", "v0.8.6 AUTHORITATIVE" in banner)
    chk("4.5 banner census reads 73 H / 40 M / 25 L", "73 H / 40 M / 25 L" in banner)
    chk("4.6 banner carries no stale v0.8.5 identifier", "v0.8.5" not in banner)

    print("\n5. PER-ROW OUTCOMES")
    def rec(ab):
        r = [x for x in range(6, 144) if tm.cell(row=x, column=1).value == ab][0]
        C, D, E, F, H = (qb.cell(row=r, column=c).value for c in (3, 4, 5, 6, 8))
        G = "" if (D is None or F is None) else F - D
        M = "UNCERTAIN" if (G == "" or H == "L" or qb.cell(row=r, column=10).value != season) else "OK"
        return dict(row=r, C=C, D=D, E=E, F=F, H=H, G=G, M=M)

    for ab, want_h, want_qb in (("RUTG", "M", "Dylan Lonergan"), ("WSU", "H", "Caden Pinnick")):
        x = rec(ab)
        chk(f"5.x {ab} activated: {want_h}, {want_qb}, G=0, status OK",
            x["H"] == want_h and x["C"] == want_qb and x["E"] == want_qb
            and x["G"] == 0 and x["M"] == "OK",
            f"H={x['H']} C={x['C']} E={x['E']} G={x['G']} status={x['M']}")

    # the zero-deviation claim for WSU rests on C80 predating this promotion
    c80_before = a["QB VALUES"].cell(row=80, column=3).value
    c80_after = qb.cell(row=80, column=3).value
    chk("5.3 C80 was ALREADY 'Caden Pinnick' in v0.8.5 and is unchanged here",
        c80_before == "Caden Pinnick" and c80_after == "Caden Pinnick",
        f"before={c80_before!r} after={c80_after!r}")
    chk("5.4 Rutgers baseline QB C35 was blank in v0.8.5 and is now populated",
        a["QB VALUES"].cell(row=35, column=3).value is None and rec("RUTG")["C"] == "Dylan Lonergan")

    c = rec("CSU")
    chk("5.5 Colorado State field standardised, still L / UNCERTAIN, values blank",
        c["E"] == "Hauss Hejny vs. K'saan Farrar" and c["H"] == "L"
        and c["D"] is None and c["F"] is None and c["C"] is None and c["M"] == "UNCERTAIN",
        f"E={c['E']} H={c['H']} status={c['M']}")
    prior_l74 = a["QB VALUES"].cell(row=74, column=12).value
    chk("5.6 the prior Colorado State note is retained verbatim, only appended to",
        str(qb.cell(row=74, column=12).value).startswith(str(prior_l74)))

    print("\n5b. DARIUS CURRY MUST APPEAR NOWHERE")
    curry = []
    for ws in b.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "Curry" in cell.value:
                    curry.append(f"{ws.title}!{cell.coordinate}")
    chk("5.7 Darius Curry appears nowhere in the workbook", not curry, str(curry))

    print("\n5c. BASELINE-QB INVARIANT")
    ok_blank_c = []
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        G = "" if (D is None or F is None) else F - D
        if not (G == "" or H == "L" or J != season) and not qb.cell(row=r, column=3).value:
            ok_blank_c.append(ab)
    chk("5.8 all 110 OK rows carry a populated baseline QB in column C",
        not ok_blank_c, str(ok_blank_c))

    print("\n6. ROWS THAT MUST NOT HAVE MOVED")
    for ab in MUST_NOT_MOVE:
        r = [x for x in range(6, 144) if tm.cell(row=x, column=1).value == ab][0]
        same = all(norm(a["QB VALUES"].cell(row=r, column=c).value) ==
                   norm(qb.cell(row=r, column=c).value) for c in range(1, 14))
        chk(f"6.x {ab} r{r} completely unchanged from v0.8.5", same)
    tt = rec("TTU")
    chk("6.10 Texas Tech medical gate still held: H, D/F blank, UNCERTAIN",
        tt["H"] == "H" and tt["D"] is None and tt["F"] is None and tt["M"] == "UNCERTAIN")

    print("\n7. MODEL OUTPUTS UNCHANGED")
    ps = b["PRESEASON"]
    S = {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 33)}
    raw = {c: [ps.cell(row=r, column=c).value for r in range(6, 144)] for c in (4, 8, 17)}
    mu = {c: mean(v) for c, v in raw.items()}
    W = {4: S["B28"], 8: S["B29"], 17: S["B31"]}
    prior = {}
    for i, r in enumerate(range(6, 144)):
        ab = tm.cell(row=r, column=1).value
        num = den = 0.0
        for c in (4, 8, 17):
            v = raw[c][i]
            if isinstance(v, (int, float)):
                num += W[c] * (v - mu[c]); den += W[c]
        prior[ab] = (num / den) if den else ""
    alias = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()
    delta = {}
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F = qb.cell(row=r, column=4).value, qb.cell(row=r, column=6).value
        delta[ab] = "" if (D is None or F is None) else F - D
    z = lambda v: 0 if v == "" else v
    sch = b["IMPORT SCHEDULE"]
    got, n_games, fcs = {}, 0, 0
    for r in range(6, 1006):
        gid = sch.cell(row=r, column=1).value
        if not gid:
            continue
        n_games += 1
        A = alias.get(str(sch.cell(row=r, column=6).value).strip(), "")
        H = alias.get(str(sch.cell(row=r, column=8).value).strip(), "")
        if not A or not H:
            fcs += 1
            continue
        neutral = bool(sch.cell(row=r, column=5).value)
        got[(A, H)] = (prior[H] - prior[A]) + (S["B7"] if neutral else S["B6"]) \
                      + (z(delta.get(H, "")) - z(delta.get(A, "")))
    chk("7.1 888 games / 761 FBS-v-FBS / 127 FCS-involved / 0 BLOCK",
        n_games == 888 and fcs == 127 and n_games - fcs == 761, f"{n_games}/{n_games-fcs}/{fcs}")
    for (A, H), want in ((("MEM", "UNLV"), "UNLV -5.6"), (("UNC", "TCU"), "TCU -4.2"),
                         (("NMSU", "FSU"), "FSU -27.7"), (("SJSU", "USC"), "USC -35.2"),
                         (("HAW", "STAN"), "STAN -3.7")):
        m = got[(A, H)]
        lab = f"{H} -{abs(m):.1f}" if m > 0 else f"{A} -{abs(m):.1f}"
        chk(f"7.x {A} at {H} model spread unchanged at {want}", lab == want, lab)
    chk("7.2 every QB delta is 0 or blank, so ENGINE!M contributes nothing",
        all(d in ("", 0) for d in delta.values()),
        str([(k, v) for k, v in delta.items() if v not in ("", 0)][:5]))
    # the two newly activated teams specifically must contribute nothing
    for ab in ("RUTG", "WSU"):
        chk(f"7.3 {ab} QB delta is exactly 0 - contributes nothing to any game",
            delta[ab] == 0, repr(delta[ab]))

    print("\n8. SETTINGS AND SAFEGUARDS")
    lines = sum(1 for r in range(6, 1006)
                if ml.cell(row=r, column=1).value is not None
                or ml.cell(row=r, column=4).value is not None)
    chk("8.1 repository artifact ships with MARKET LINES blank", lines == 0, str(lines))
    chk("8.2 BET toggle remains N", st["B11"].value == "N", repr(st["B11"].value))
    chk("8.3 totals remain unavailable (B22/B23 blank)",
        st["B22"].value is None and st["B23"].value is None)
    for t_ in ("NDSU", "SAC"):
        rr = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == t_][0]
        chk(f"8.4 {t_} remains transitional", tm.cell(row=rr, column=5).value == "Y", f"row {rr}")
    au = b["AUDIT"]
    chk("8.5 AUDIT market-line invariant unchanged from v0.8.3 design",
        "SUMPRODUCT" in str(au["B16"].value) and "REMOVE TEST LINES" not in str(au["B16"].value))

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"v0.8.5 SHA-256 (frozen): {h85}")
    print(f"v0.8.6 SHA-256:          {h86}")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

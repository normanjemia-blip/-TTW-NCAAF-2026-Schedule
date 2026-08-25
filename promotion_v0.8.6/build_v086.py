#!/usr/bin/env python3
"""Build v0.8.6 from the frozen v0.8.5 workbook.

APPROVED SCOPE — supplemental packet items 1-3 ONLY. 18 cells.

  ACTIVATIONS (2 rows, 14 cells, 4 of them the approved zeros)
    Rutgers          r35  M -> M (unchanged)  Dylan Lonergan       (ESPN/Thamel 2026-08-24)
    Washington State r80  L -> H              Caden Pinnick        (team announcement 2026-08-24)

  RECORD CORRECTION (1 row, 3 cells, no numerics)
    Colorado State   r74  active field standardised; stays L / UNCERTAIN

  BANNER (1 cell) version + confidence census

EXPLICIT OWNER CONSTRAINTS HONOURED BY THIS BUILD:
  * C80 is NOT written. Caden Pinnick was ALREADY the recorded baseline quarterback
    before this build, so the announced starter and the preseason assumption are the
    same player and the deviation is literally zero. The build asserts C80 is
    unchanged rather than rewriting it.
  * C35 IS written. Every one of the 108 OK rows in v0.8.5 carries a populated
    baseline quarterback in column C (108/108, no exceptions). Activating Rutgers
    with a blank C would have created the first OK row with no baseline QB.
  * H35 is NOT written - Rutgers already carries M and the confidence is unchanged.
  * Darius Curry is NOT added anywhere. The build asserts his absence afterwards.
  * Colorado State receives no numerical entry and keeps L / UNCERTAIN.

Target census after this build: 110 OK / 28 UNCERTAIN and 73 H / 40 M / 25 L.

Run:  python3 promotion_v0.8.6/build_v086.py
"""
import datetime, hashlib, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "promotion_v0.8.5",
                   "TTW_College_Football_Power_Ratings_v0.8.5_AUTHORITATIVE.xlsx")
OUT = os.path.join(ROOT, "promotion_v0.8.6",
                   "TTW_College_Football_Power_Ratings_v0.8.6_AUTHORITATIVE.xlsx")
FROZEN_V085_SHA = "0676aa1a05d661ca0d99c917c8dc471c0030128cc42ea8fd1bd2f17dcea767be"
D24 = datetime.datetime(2026, 8, 24)

ZERO_JUSTIFICATION = (
    "Baseline and active values are 0/0 under the deviation-only convention: "
    "QB VALUES!G = F - D, so 0 - 0 = 0 and ENGINE!M contributes exactly nothing "
    "to any game. The zeros do not rate the quarterback - they record that the "
    "active starter IS the quarterback the preseason rating already assumed, so "
    "no deviation applies. No nonzero QB adjustment and no model change."
)

ACTIVATIONS = {
 "RUTG": dict(row=35, C="Dylan Lonergan", E="Dylan Lonergan", H=None,
   I=("ESPN / Pete Thamel via On3 The Knight Report + 247Sports 2026-08-24 "
      "(https://www.on3.com/sites/the-knight-report/news/"
      "rutgers-names-dylan-lonergan-starting-quarterback-for-2026/)"),
   L=("2026-08-24 ACTIVATED, confidence M (UNCHANGED from M). ESPN's Pete Thamel reported that "
      "Rutgers named Dylan Lonergan its starting quarterback, won over AJ Surace after spring and "
      "fall camp; carried by On3/The Knight Report, 247Sports, Yahoo Sports and On the Banks. "
      "M rather than H because this is a reporter-sourced claim rather than a team or coach "
      "announcement, matching the precedent set when North Carolina and Georgia Southern were "
      "activated at M on Thamel sources reports. INDEPENDENTLY CORROBORATED BY THE WORKBOOK: "
      "reporting that Lonergan leads Rutgers into Week 1 against UMass matches IMPORT SCHEDULE "
      "exactly (wk1 2026-09-03 Massachusetts Minutemen @ Rutgers Scarlet Knights). Supersedes the "
      "2026-08-03 SI projection entry, which recorded only 'likely starter'. BASELINE QB RECORDED "
      "IN COLUMN C AT ACTIVATION: the field was blank, and every OK row in the workbook carries a "
      "populated baseline quarterback, so Lonergan - named as the projected starter by the SI "
      "source the prior entry cited - is recorded as the baseline the preseason blend assumed. "
      + ZERO_JUSTIFICATION)),
 "WSU": dict(row=80, C=None, E="Caden Pinnick", H="H",
   I=("Washington State official team announcement (@WSUCougarFB) via Spokesman-Review 2026-08-24 "
      "(https://www.spokesman.com/stories/2026/aug/24/"
      "uc-davis-transfer-caden-pinnick-to-start-at-qb-for/)"),
   L=("2026-08-24 ACTIVATED, confidence L -> H. Washington State announced Caden Pinnick as QB1; "
      "the program posted the announcement from @WSUCougarFB on the afternoon of 2026-08-24, "
      "reported by the Spokesman-Review: 'The Cougars will start UC Davis transfer Caden Pinnick, "
      "the program announced on social media Monday afternoon.' H rather than M because this is a "
      "FIRST-PARTY TEAM ANNOUNCEMENT, not a reporter-sourced claim - a tier above the Georgia "
      "Southern / North Carolina / Rutgers precedent. Resolves the three-way competition with Owen "
      "Eshelman and Julian Dugger and satisfies the deadline HC Kirby Moore set publicly on "
      "2026-08-06 ('by Aug. 24') and restated 2026-08-21 ('by Monday'). Supersedes the 2026-08-04 "
      "defect-corrected entry, which recorded a wide-open camp battle with no leader. "
      "INDEPENDENTLY CORROBORATED BY THE WORKBOOK: the report that WSU opens Sept. 6 at Washington "
      "matches IMPORT SCHEDULE exactly (wk1 2026-09-06 Washington State Cougars @ Washington "
      "Huskies). DEVIATION IS EXACTLY ZERO AND VERIFIED: Pinnick was ALREADY the recorded baseline "
      "quarterback in column C before this activation, so the announced starter and the preseason "
      "assumption are the same player and no deviation applies. " + ZERO_JUSTIFICATION)),
}

# Colorado State: the prior 2026-08-03 note is RETAINED VERBATIM and the ruling appended.
CSU_RULING = (
    " RULING 2026-08-24: active competition field standardised to \"Hauss Hejny vs. K'saan "
    "Farrar\". Confidence L and status UNCERTAIN are UNCHANGED - this is a record correction "
    "only, with no numerical entry and no model effect. The active competition field is "
    "limited to these two quarterbacks; no third quarterback is carried in it."
)
CORRECTION = {
 "CSU": dict(row=74, E="Hauss Hejny vs. K'saan Farrar", append_L=CSU_RULING),
}

COLS = dict(C=3, D=4, E=5, F=6, H=8, I=9, K=11, L=12)


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def main():
    got = sha256(SRC)
    assert got == FROZEN_V085_SHA, f"v0.8.5 is not the expected artifact: {got}"
    print(f"source v0.8.5 SHA-256 verified: {got}")

    wb = openpyxl.load_workbook(SRC)
    tm, qb, sh = wb["TEAM MAP"], wb["QB VALUES"], wb["START HERE"]

    idx = {}
    for ab in list(ACTIVATIONS) + list(CORRECTION):
        rows = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == ab]
        assert len(rows) == 1, f"{ab} must resolve to exactly one row, got {rows}"
        idx[ab] = rows[0]
    for ab, spec in {**ACTIVATIONS, **CORRECTION}.items():
        assert idx[ab] == spec["row"], f"{ab} expected row {spec['row']}, got {idx[ab]}"
    print("all affected teams located by abbreviation; rows match the approved packet")

    # C80 must already hold Caden Pinnick -- the whole zero-deviation claim rests on it
    c80_before = qb.cell(row=80, column=COLS["C"]).value
    assert c80_before == "Caden Pinnick", f"C80 expected 'Caden Pinnick', got {c80_before!r}"
    print("C80 pre-verified as 'Caden Pinnick' (baseline == announced starter, deviation is zero)")

    # every OK row carries a populated baseline QB -- the invariant C35 protects
    def status(r):
        D = qb.cell(row=r, column=COLS["D"]).value
        F = qb.cell(row=r, column=COLS["F"]).value
        H = qb.cell(row=r, column=COLS["H"]).value
        J = qb.cell(row=r, column=10).value
        blank = D in (None, "") or F in (None, "")
        return "UNCERTAIN" if (blank or H == "L" or J != 2026) else "OK"

    ok_blank_c = [r for r in range(6, 144)
                  if status(r) == "OK" and not qb.cell(row=r, column=COLS["C"]).value]
    assert not ok_blank_c, f"v0.8.5 already violates the baseline-QB invariant at {ok_blank_c}"
    print("invariant confirmed in v0.8.5: every OK row has a populated baseline QB in column C")

    # rows that must NOT move
    untouched = {}
    for ab in ("TTU", "STAN", "NIU", "TULN", "FRES", "SYR", "ALA", "TENN", "GASO"):
        r = [x for x in range(6, 144) if tm.cell(row=x, column=1).value == ab][0]
        untouched[ab] = (r, [qb.cell(row=r, column=c).value for c in range(1, 14)])

    def guard(r):
        for name, col in (("A", 1), ("B", 2), ("G", 7), ("M", 13)):
            assert isf(qb.cell(row=r, column=col).value), f"{name}{r} must be a formula"
        assert qb.cell(row=r, column=10).value == 2026, f"J{r} must be 2026"

    cells = 0
    zeros_written = 0
    for ab, s in ACTIVATIONS.items():
        r = s["row"]; guard(r)
        if s["C"] is not None:
            qb.cell(row=r, column=COLS["C"]).value = s["C"]; cells += 1
        qb.cell(row=r, column=COLS["D"]).value = 0; zeros_written += 1; cells += 1
        qb.cell(row=r, column=COLS["E"]).value = s["E"]; cells += 1
        qb.cell(row=r, column=COLS["F"]).value = 0; zeros_written += 1; cells += 1
        if s["H"] is not None:
            qb.cell(row=r, column=COLS["H"]).value = s["H"]; cells += 1
        qb.cell(row=r, column=COLS["I"]).value = s["I"]; cells += 1
        qb.cell(row=r, column=COLS["K"]).value = D24; cells += 1
        qb.cell(row=r, column=COLS["L"]).value = s["L"]; cells += 1
    assert zeros_written == 4, f"expected exactly 4 zeros, wrote {zeros_written}"
    assert qb.cell(row=80, column=COLS["C"]).value == "Caden Pinnick", "C80 must be unchanged"
    assert qb.cell(row=35, column=COLS["H"]).value == "M", "H35 must remain M"
    print(f"activations applied: {len(ACTIVATIONS)} rows, {zeros_written} zeros (exactly the 4 approved)")

    for ab, s in CORRECTION.items():
        r = s["row"]; guard(r)
        prior = qb.cell(row=r, column=COLS["L"]).value
        qb.cell(row=r, column=COLS["E"]).value = s["E"]; cells += 1
        qb.cell(row=r, column=COLS["K"]).value = D24; cells += 1
        qb.cell(row=r, column=COLS["L"]).value = prior + s["append_L"]; cells += 1
        assert qb.cell(row=r, column=COLS["L"]).value.startswith(prior), \
            "the prior Colorado State note must be retained verbatim"
        assert qb.cell(row=r, column=COLS["D"]).value is None, "D74 must stay blank"
        assert qb.cell(row=r, column=COLS["F"]).value is None, "F74 must stay blank"
        assert qb.cell(row=r, column=COLS["H"]).value == "L", "H74 must stay L"
        assert qb.cell(row=r, column=COLS["C"]).value is None, "C74 must stay blank"
    print(f"correction applied: {len(CORRECTION)} row, no numerics, confidence unchanged")

    # Darius Curry must appear nowhere in the workbook
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "Curry" in c.value:
                    raise AssertionError(f"Darius Curry must not be added: {ws.title}!{c.coordinate}")
    print("confirmed: Darius Curry appears nowhere in the workbook")

    for ab, (r, before) in untouched.items():
        assert [qb.cell(row=r, column=c).value for c in range(1, 14)] == before, \
            f"{ab} row {r} was modified but must not be"
    print(f"confirmed untouched: {', '.join(untouched)}")

    banner = sh["A1"].value
    assert "v0.8.5 AUTHORITATIVE" in banner
    assert "72 H / 40 M / 26 L" in banner, f"unexpected banner census: {banner[:200]}"
    sh["A1"].value = (banner.replace("v0.8.5 AUTHORITATIVE", "v0.8.6 AUTHORITATIVE")
                            .replace("72 H / 40 M / 26 L", "73 H / 40 M / 25 L"))
    assert "v0.8.5" not in sh["A1"].value
    cells += 1
    print("banner updated: version + confidence census")

    assert cells == 18, f"expected exactly 18 cells, wrote {cells}"
    print(f"total cells written: {cells}")

    tmp = OUT + ".building.xlsx"
    wb.save(tmp)
    os.replace(tmp, OUT)
    print(f"written: {OUT}")
    print(f"v0.8.6 SHA-256: {sha256(OUT)}")
    assert sha256(SRC) == FROZEN_V085_SHA, "v0.8.5 was modified by the build"
    print("v0.8.5 confirmed still frozen")


if __name__ == "__main__":
    sys.exit(main())

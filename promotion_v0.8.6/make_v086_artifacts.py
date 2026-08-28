#!/usr/bin/env python3
"""Generate the v0.8.6 promotion artifacts EXPLICITLY.

Kept separate from verify_v086.py on purpose: a certificate must be read-only,
so artifact generation never happens as a verification side effect.

Writes:
  promotion_v0.8.6/diff_v085_to_v086.csv
  promotion_v0.8.6/regression_log_v086.txt
"""
import csv, os, subprocess, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HERE = os.path.join(ROOT, "promotion_v0.8.6")
V085 = os.path.join(ROOT, "promotion_v0.8.5",
                    "TTW_College_Football_Power_Ratings_v0.8.5_AUTHORITATIVE.xlsx")
V086 = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.6_AUTHORITATIVE.xlsx")

CLASS = {35: "ACTIVATION", 80: "ACTIVATION", 74: "RECORD CORRECTION"}
TEAM = {35: "Rutgers", 80: "Washington State", 74: "Colorado State"}


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def main():
    a = openpyxl.load_workbook(V085)
    b = openpyxl.load_workbook(V086)
    rows = []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                x, y = wa.cell(row=r, column=c), wbk.cell(row=r, column=c)
                if norm(x.value) != norm(y.value):
                    rows.append(dict(
                        sheet=s, cell=y.coordinate, row=r,
                        team=TEAM.get(r, "") if s == "QB VALUES" else "",
                        change_class=CLASS.get(r, "BANNER") if s == "QB VALUES" else "BANNER",
                        old=str(x.value) if x.value is not None else "",
                        new=str(y.value) if y.value is not None else "",
                        is_numeric_zero=(y.value == 0)))
    out = os.path.join(HERE, "diff_v085_to_v086.csv")
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0]))
        w.writeheader()
        w.writerows(rows)
    zeros = sum(1 for r in rows if r["is_numeric_zero"])
    print(f"wrote {out}: {len(rows)} changed cells, {zeros} of them the integer 0")
    assert len(rows) == 18, f"expected 18 changed cells, got {len(rows)}"
    assert zeros == 4, f"expected 4 zeros, got {zeros}"

    log = os.path.join(HERE, "regression_log_v086.txt")
    with open(log, "w", encoding="utf-8") as f:
        for script in ("promotion_v0.8.6/verify_v086.py",
                       "phase11_week0_dryrun/week0_dryrun.py"):
            f.write("=" * 78 + f"\n{script}\n" + "=" * 78 + "\n")
            p = subprocess.run([sys.executable, os.path.join(ROOT, script)],
                               capture_output=True, text=True, cwd=ROOT)
            f.write(p.stdout + p.stderr + f"\n[exit {p.returncode}]\n\n")
    print(f"wrote {log}")


if __name__ == "__main__":
    sys.exit(main())

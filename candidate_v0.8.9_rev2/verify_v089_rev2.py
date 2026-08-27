#!/usr/bin/env python3
"""v0.8.9 REV 2 certificate — READ-ONLY. Writes nothing.

Every fixture is evaluated AT EACH BUILD'S OWN PRODUCTION SETTINGS. That is the
specific gap that let REV 1 ship a totals regression: its check 9.1 pinned the
toggle to the v0.8.8 value on both sides, isolating the thresholds and hiding
the shipped-configuration difference.
"""
import collections, hashlib, os, re, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
V088 = os.path.join(ROOT, "promotion_v0.8.8",
                    "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
REV2 = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.9_REV2_CANDIDATE.xlsx")
SRC_SHA = "b2a920feddc0f49f0647957334db0ecd0e922fe6a3933fc6a11af31587b56450"

LIVE_LINES = {
    "401856766": ("TCU", 7.5), "401858201": ("STAN", 3.5), "401858202": ("UVA", 5.5),
    "401864494": ("USC", 38.5), "401864570": ("FSU", 30.5), "401864577": ("NDSU", 7.5),
    "401866408": ("EMU", 9.5), "401862693": ("UNLV", 4.5),
}
EXPECTED_CHANGES = {"401866408", "401856766", "401864494", "401864570"}

PASS, FAIL = [], []


def chk(m, ok, d=""):
    (PASS if ok else FAIL).append(m)
    print(f"  [{'PASS' if ok else 'FAIL'}] {m}" + (f" - {d}" if d else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", (v.text or "").strip())
    if isinstance(v, str) and v.startswith("="):
        return ("F", v.strip())
    return ("V", v)


def txt(v):
    return v.text if isinstance(v, ArrayFormula) else v


def mean(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def spread_label(V, S, AI, AF, AG):
    """ENGINE!X - identical in both builds; reads B8/B9/B10/B11."""
    if V == "" or AI in ("BLOCKED", "PENDING LINE", "STALE LINE"):
        return ""
    if abs(V) < S["B8"]:
        return ""
    if abs(V) < S["B9"]:
        return "LEAN"
    if (abs(V) < S["B10"] or S["B11"] != "Y" or AI != "READY" or AF != "" or AG != ""):
        return "INVESTIGATE"
    return "BET"


def total_label_088(AA, S, AI, AF, AG):
    """v0.8.8 ENGINE!AB - coupled to B8/B9/B10/B11."""
    if AA == "":
        return ""
    if abs(AA) < S["B8"] * 2:
        return ""
    if abs(AA) < S["B9"] * 2:
        return "LEAN"
    if (S["B11"] != "Y" or AI != "READY" or abs(AA) < S["B10"] * 2 or AF != "" or AG != ""):
        return "INVESTIGATE"
    return "BET"


def total_label_rev2(AA, S, AI, AF, AG):
    """REV 2 ENGINE!AB - dedicated B49/B50/B51 thresholds and B52 toggle."""
    if AA == "":
        return ""
    if abs(AA) < S["B49"]:
        return ""
    if abs(AA) < S["B50"]:
        return "LEAN"
    if (S["B52"] != "Y" or AI != "READY" or abs(AA) < S["B51"] or AF != "" or AG != ""):
        return "INVESTIGATE"
    return "BET"


def audit_b12(S, xf):
    return "OK" if (S.get("B10") == 1.5 and S.get("B11") == "Y"
                    and "SETTINGS!$B$10" in xf and "SETTINGS!$B$11" in xf
                    and "SETTINGS!$B$51" not in xf and "SETTINGS!$B$52" not in xf) else "CHECK"


def audit_b13(S, abf):
    return "OK" if (S.get("B49") == 2 and S.get("B50") == 3 and S.get("B51") == 6
                    and S.get("B52") == "N"
                    and all(t in abf for t in ("SETTINGS!$B$49", "SETTINGS!$B$50",
                                               "SETTINGS!$B$51", "SETTINGS!$B$52"))
                    and "SETTINGS!$B$10" not in abf
                    and "SETTINGS!$B$11" not in abf) else "CHECK"


def engine(wb, lines):
    tm, qb, st, ps = wb["TEAM MAP"], wb["QB VALUES"], wb["SETTINGS"], wb["PRESEASON"]
    S = {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 55)}
    raw = {c: [ps.cell(row=r, column=c).value for r in range(6, 144)] for c in (4, 8, 17)}
    mu = {c: mean(v) for c, v in raw.items()}
    W = {4: S["B28"], 8: S["B29"], 17: S["B31"]}
    prior = {}
    for i, r in enumerate(range(6, 144)):
        ab = tm.cell(row=r, column=1).value
        num = den = 0.0
        for c in (4, 8, 17):
            v = raw[c][i]
            if isinstance(v, (int, float)):
                num += W[c] * (v - mu[c]); den += W[c]
        prior[ab] = (num / den) if den else ""
    alias = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()
    delta, status = {}, {}
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        g = "" if (D is None or F is None) else F - D
        delta[ab] = g
        status[ab] = "UNCERTAIN" if (g == "" or H == "L" or J != S["B3"]) else "OK"
    z = lambda v: 0 if v == "" else v
    sch = wb["IMPORT SCHEDULE"]
    out = {}
    for r in range(6, 1006):
        gid = sch.cell(row=r, column=1).value
        if not gid:
            continue
        A = alias.get(str(sch.cell(row=r, column=6).value).strip(), "")
        H = alias.get(str(sch.cell(row=r, column=8).value).strip(), "")
        if not A or not H:
            continue
        R = (prior[H] - prior[A]) + (S["B7"] if bool(sch.cell(row=r, column=5).value) else S["B6"]) \
            + (z(delta.get(H, "")) - z(delta.get(A, "")))
        ln = lines.get(str(gid))
        edge = side = ""
        gate = "QB UNCERTAIN" if (status.get(A) == "UNCERTAIN" or status.get(H) == "UNCERTAIN") else "READY"
        if ln:
            fav, spread = ln
            T = -abs(spread) if fav == H else (abs(spread) if fav == A else "")
            if T != "":
                edge = R + T
                side = H if edge > 0 else A
        out[str(gid)] = dict(away=A, home=H, model=R, edge=edge, side=side, gate=gate, S=S)
    return out


def main():
    print("=" * 78)
    print("v0.8.9 REV 2 CERTIFICATE — spread rule with totals fully separated")
    print("=" * 78)
    h8, h2 = sha256(V088), sha256(REV2)
    print(f"  v0.8.8 source SHA-256 : {h8}")
    print(f"  REV 2 candidate SHA   : {h2}")

    print("\n1. SOURCE")
    chk("1.1 v0.8.8 remains the byte-exact source", h8 == SRC_SHA, h8[:16])

    a = openpyxl.load_workbook(V088)
    b = openpyxl.load_workbook(REV2)
    Sa = {f"B{r}": a["SETTINGS"].cell(row=r, column=2).value for r in range(3, 55)}
    Sb = {f"B{r}": b["SETTINGS"].cell(row=r, column=2).value for r in range(3, 55)}
    xf, abf = txt(b["ENGINE"]["X6"].value), txt(b["ENGINE"]["AB6"].value)

    print("\n2. EVERY CHANGED CELL")
    changed = []
    for s in a.sheetnames:
        wa, wb_ = a[s], b[s]
        R = max(wa.max_row, wb_.max_row); C = max(wa.max_column, wb_.max_column)
        for r in range(1, R + 1):
            for c in range(1, C + 1):
                if norm(wa.cell(row=r, column=c).value) != norm(wb_.cell(row=r, column=c).value):
                    changed.append((s, wb_.cell(row=r, column=c).coordinate))
    by = collections.Counter(s for s, _ in changed)
    chk("2.1 exactly 1023 cells changed", len(changed) == 1023, str(len(changed)))
    chk("2.2 sheets: SETTINGS 15, ENGINE 1000, AUDIT 4, CHANGELOG 4",
        dict(by) == {"SETTINGS": 15, "ENGINE": 1000, "AUDIT": 4, "CHANGELOG": 4}, str(dict(by)))
    setc = sorted(c for s, c in changed if s == "SETTINGS")
    chk("2.3 SETTINGS cells enumerated exactly",
        setc == ["A10", "A11", "A48", "A49", "A50", "A51", "A52", "A8", "A9",
                 "B10", "B11", "B49", "B50", "B51", "B52"], str(setc))
    chk("2.4 AUDIT cells are A12,A13,B12,B13",
        sorted(c for s, c in changed if s == "AUDIT") == ["A12", "A13", "B12", "B13"])
    chk("2.5 CHANGELOG cells are A87..D87",
        sorted(c for s, c in changed if s == "CHANGELOG") == ["A87", "B87", "C87", "D87"])
    engc = {c for s, c in changed if s == "ENGINE"}
    chk("2.6 all 1000 ENGINE cells are AB6:AB1005",
        len(engc) == 1000 and all(re.fullmatch(r"AB\d+", c) and 6 <= int(c[2:]) <= 1005
                                  for c in engc))

    print("\n3-4. SPREAD BOUNDARY AT REV 2 PRODUCTION SETTINGS")
    for v, want in ((1.49, False), (-1.49, False), (1.50, True), (-1.50, True),
                    (1.51, True), (-1.51, True)):
        lab = spread_label(v, Sb, "READY", "", "")
        chk(f"{'4' if want else '3'}.x edge {v:+.2f} -> {lab or '(blank)'} "
            f"({'BET' if want else 'not BET'})",
            (lab == "BET") == want, lab)

    print("\n5-6. SETTINGS!B11 AFFECTS SPREADS ONLY")
    chk("5.1 ENGINE!X references B10 and B11", "SETTINGS!$B$10" in xf and "SETTINGS!$B$11" in xf)
    chk("5.2 ENGINE!AB references NEITHER B10 nor B11",
        "SETTINGS!$B$10" not in abf and "SETTINGS!$B$11" not in abf)
    fixtures = [1.99, -1.99, 2.00, -2.00, 2.99, -2.99, 3.00, -3.00,
                5.99, -5.99, 6.00, -6.00, 6.01, -6.01, ""]
    flip_b11 = dict(Sb); flip_b11["B11"] = "N"
    tdiff = [aa for aa in fixtures
             if total_label_rev2(aa, Sb, "READY", "", "")
             != total_label_rev2(aa, flip_b11, "READY", "", "")]
    chk("6.1 flipping B11 produces ZERO totals-label changes across all 15 fixtures",
        not tdiff, str(tdiff))
    sdiff = [v for v in (0.5, 1.49, 1.5, 3.0, 6.0, 12.0)
             if spread_label(v, Sb, "READY", "", "") == spread_label(v, flip_b11, "READY", "", "")]
    chk("6.2 flipping B11 DOES change spread labels (it is a live spread control)",
        len(sdiff) < 6, f"{6-len(sdiff)} of 6 spread fixtures moved")

    print("\n7-8. THE DEDICATED TOTALS TOGGLE AFFECTS TOTALS ONLY")
    chk("7.1 ENGINE!AB references the dedicated toggle B52", "SETTINGS!$B$52" in abf)
    chk("7.2 ENGINE!X does NOT reference B52 or B51",
        "SETTINGS!$B$52" not in xf and "SETTINGS!$B$51" not in xf)
    flip_b52 = dict(Sb); flip_b52["B52"] = "Y"
    sp_diff = [v for v in (0.5, 1.49, 1.5, 3.0, 6.0, 12.0, -1.5, -6.0)
               if spread_label(v, Sb, "READY", "", "")
               != spread_label(v, flip_b52, "READY", "", "")]
    chk("8.1 flipping the totals toggle produces ZERO spread-label changes",
        not sp_diff, str(sp_diff))

    print("\n9. TOTALS AT PRODUCTION SETTINGS — v0.8.8 vs REV 2, EACH AT ITS OWN CONFIG")
    print(f"     {'AA':>8}  {'v0.8.8':<14}{'REV 2':<14}")
    mism = []
    for aa in fixtures:
        l8 = total_label_088(aa, Sa, "READY", "", "")
        l2 = total_label_rev2(aa, Sb, "READY", "", "")
        if l8 != l2:
            mism.append((aa, l8, l2))
        lbl = "blank" if aa == "" else f"{aa:+.2f}"
        print(f"     {lbl:>8}  {l8 or '(blank)':<14}{l2 or '(blank)':<14}"
              f"{'' if l8 == l2 else '  <-- DIFFERS'}")
    chk("9.1 totals identical at every fixture, each build at its OWN production config",
        not mism, str(mism))
    chk("9.2 dedicated thresholds are exactly 2.0 / 3.0 / 6.0",
        (Sb["B49"], Sb["B50"], Sb["B51"]) == (2, 3, 6))
    chk("9.3 totals BET toggle ships N", Sb["B52"] == "N", repr(Sb["B52"]))

    print("\n10. HARNESS-ONLY: totals toggle Y makes BET reachable at ±6 without touching spreads")
    chk("10.1 with B52='Y' (harness only) |AA|=6.00 -> BET",
        total_label_rev2(6.00, flip_b52, "READY", "", "") == "BET",
        total_label_rev2(6.00, flip_b52, "READY", "", ""))
    chk("10.2 with B52='Y' |AA|=5.99 -> INVESTIGATE (boundary intact)",
        total_label_rev2(5.99, flip_b52, "READY", "", "") == "INVESTIGATE")
    chk("10.3 spreads unaffected by that harness change", not sp_diff)

    print("\n11. TOTALS REMAIN INERT IN THE ACTUAL WORKBOOK")
    chk("11.1 SETTINGS!B22/B23 still blank", Sb["B22"] is None and Sb["B23"] is None)
    y6 = txt(b["ENGINE"]["Y6"].value)
    chk("11.2 ENGINE!Y still gated on B22/B23", "SETTINGS!$B$22" in y6 and "SETTINGS!$B$23" in y6)
    chk("11.3 every totals label is blank in production", total_label_rev2("", Sb, "READY", "", "") == "")

    print("\n12-13. BOTH AUDIT ROWS")
    chk("12.1 AUDIT!B12 (spread) = OK in the approved configuration",
        audit_b12(Sb, xf) == "OK", audit_b12(Sb, xf))
    chk("12.2 AUDIT!B13 (totals) = OK in the approved configuration",
        audit_b13(Sb, abf) == "OK", audit_b13(Sb, abf))
    for name, S_, f_, fn, who in (
            ("spread BET threshold -> 3", dict(Sb, B10=3), xf, audit_b12, "B12"),
            ("spread toggle -> N", dict(Sb, B11="N"), xf, audit_b12, "B12"),
            ("spread formula loses B11", Sb, xf.replace("SETTINGS!$B$11", "TRUE"), audit_b12, "B12"),
            ("totals LEAN -> 2.5", dict(Sb, B49=2.5), abf, audit_b13, "B13"),
            ("totals INVESTIGATE -> 4", dict(Sb, B50=4), abf, audit_b13, "B13"),
            ("totals BET -> 3", dict(Sb, B51=3), abf, audit_b13, "B13"),
            ("totals toggle -> Y", dict(Sb, B52="Y"), abf, audit_b13, "B13"),
            ("totals formula re-coupled to B11", Sb,
             abf.replace("SETTINGS!$B$52", "SETTINGS!$B$11"), audit_b13, "B13")):
        chk(f"13.x AUDIT!{who} = CHECK when {name}", fn(S_, f_) == "CHECK")
    chk("13.9 AUDIT!B12 fails safe to CHECK if FORMULATEXT is unavailable",
        audit_b12(Sb, "") == "CHECK")
    chk("13.10 AUDIT!B13 fails safe to CHECK if FORMULATEXT is unavailable",
        audit_b13(Sb, "") == "CHECK")

    print("\n14-15. THE EIGHT LINED GAMES")
    E8, E2 = engine(a, LIVE_LINES), engine(b, LIVE_LINES)
    print(f"     {'game':<14}{'edge':>8}  {'v0.8.8':<14}{'REV 2':<14}{'gate'}")
    moved = set()
    for g in LIVE_LINES:
        r8, r2 = E8[g], E2[g]
        l8 = spread_label(r8["edge"], Sa, r8["gate"], "", "")
        l2 = spread_label(r2["edge"], Sb, r2["gate"], "", "")
        if l8 != l2:
            moved.add(g)
        print(f"     {r2['away']+'@'+r2['home']:<14}{r2['edge']:+8.2f}  "
              f"{l8 or '(blank)':<14}{l2 or '(blank)':<14}{r2['gate']}")
    chk("14.1 exactly four lined spread games change", len(moved) == 4, str(len(moved)))
    chk("14.2 they are SAC@EMU, UNC@TCU, SJSU@USC, NMSU@FSU", moved == EXPECTED_CHANGES,
        str(sorted(moved)))
    mem = E2["401862693"]
    chk("15.1 Memphis at UNLV remains QB UNCERTAIN", mem["gate"] == "QB UNCERTAIN")
    chk("15.2 Memphis at UNLV remains LEAN",
        spread_label(mem["edge"], Sb, mem["gate"], "", "") == "LEAN",
        spread_label(mem["edge"], Sb, mem["gate"], "", ""))

    print("\n16. NOTHING ELSE MOVED")
    for sheet in ("QB VALUES", "IMPORT SCHEDULE", "TEAM MAP", "PRESEASON", "TEAM RATINGS",
                  "MARKET LINES", "ADJUSTMENTS", "CLEAN", "CALC", "DASHBOARD", "START HERE"):
        chk(f"16.x {sheet} byte-identical", by.get(sheet, 0) == 0, str(by.get(sheet, 0)))
    proj = [g for g in E8 if (round(E8[g]["model"], 9), E8[g]["edge"], E8[g]["side"], E8[g]["gate"])
            != (round(E2[g]["model"], 9), E2[g]["edge"], E2[g]["side"], E2[g]["gate"])]
    chk("16.12 model spread, edge, side and gate identical across all 761 games", not proj,
        str(proj[:5]))
    chk("16.13 rating inputs untouched",
        (Sb["B6"], Sb["B7"], Sb["B12"], Sb["B8"], Sb["B9"]) == (2.5, 0, 2.5, 1, 1.5))

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"v0.8.8 source : {h8}")
    print(f"REV 2 SHA-256 : {h2}")
    print("STATUS: CANDIDATE ONLY — NOT PROMOTED")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

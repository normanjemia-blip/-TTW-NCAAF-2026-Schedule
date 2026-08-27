#!/usr/bin/env python3
"""Build v0.8.9 REV 2 from AUTHORITATIVE v0.8.8 — not from REV 1.

REV 1 (commit 2963afe) is preserved as superseded evidence. Its defect: the
BET toggle SETTINGS!B11 stayed shared with totals, so enabling it for spreads
also enabled totals BET at |total edge| >= 6.

REV 2 gives totals their OWN toggle and their OWN thresholds, in a clean
contiguous block on unused, unreferenced SETTINGS rows 48-52.

  SETTINGS!A8/A9/A10/A11   labels made explicit: these control SPREADS only
  SETTINGS!B10             3 -> 1.5    spread BET threshold
  SETTINGS!B11             N -> Y      spread BET toggle (SPREADS ONLY)
  SETTINGS!A48             section header for the totals block
  SETTINGS!A49/B49         TOTALS LEAN         = 2
  SETTINGS!A50/B50         TOTALS INVESTIGATE  = 3
  SETTINGS!A51/B51         TOTALS BET          = 6
  SETTINGS!A52/B52         Enable totals BET labels? (Y/N) = N
  ENGINE!AB6:AB1005        totals classification -> B49/B50/B51/B52,
                           with NO reference to B10 or B11
  AUDIT!A12/B12            spread configuration guard
  AUDIT!A13/B13            totals configuration guard
  CHANGELOG row 87         v0.8.9 entry

Totals are NOT enabled: B22/B23 stay blank and the totals toggle ships N.
ENGINE!X is not edited at all.

Run:  python3 candidate_v0.8.9_rev2/build_v089_rev2.py
"""
import datetime, hashlib, os, shutil, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, "promotion_v0.8.8",
                   "TTW_College_Football_Power_Ratings_v0.8.8_AUTHORITATIVE.xlsx")
OUT = os.path.join(HERE, "TTW_College_Football_Power_Ratings_v0.8.9_REV2_CANDIDATE.xlsx")
SRC_SHA = "b2a920feddc0f49f0647957334db0ecd0e922fe6a3933fc6a11af31587b56450"

SPREAD_LABELS = {
    "A8":  "SPREAD LEAN threshold (abs edge)",
    "A9":  "SPREAD INVESTIGATE threshold (abs edge)",
    "A10": "SPREAD BET threshold (abs edge)",
    "A11": "Enable SPREAD BET labels? (Y/N) — spreads only",
}
TOTALS_BLOCK = {
    "A48": ("TOTALS MARKET CONTROLS — independent of the spread controls above", None),
    "A49": ("TOTALS LEAN threshold (abs edge)", 2),
    "A50": ("TOTALS INVESTIGATE threshold (abs edge)", 3),
    "A51": ("TOTALS BET threshold (abs edge)", 6),
    "A52": ("Enable totals BET labels? (Y/N)", "N"),
}
REPOINT = (
    ("SETTINGS!$B$8*2", "SETTINGS!$B$49"),
    ("SETTINGS!$B$9*2", "SETTINGS!$B$50"),
    ("SETTINGS!$B$10*2", "SETTINGS!$B$51"),
    ('SETTINGS!$B$11<>"Y"', 'SETTINGS!$B$52<>"Y"'),
)

AUDIT_A12 = "SPREAD config: BET threshold 1.5, spread toggle Y, formula independent of totals"
AUDIT_B12 = (
    '=IF(AND(SETTINGS!$B$10=1.5,SETTINGS!$B$11="Y",'
    'IFERROR(ISNUMBER(SEARCH("SETTINGS!$B$10",FORMULATEXT(ENGINE!$X$6))),FALSE),'
    'IFERROR(ISNUMBER(SEARCH("SETTINGS!$B$11",FORMULATEXT(ENGINE!$X$6))),FALSE),'
    'IFERROR(NOT(ISNUMBER(SEARCH("SETTINGS!$B$51",FORMULATEXT(ENGINE!$X$6)))),FALSE),'
    'IFERROR(NOT(ISNUMBER(SEARCH("SETTINGS!$B$52",FORMULATEXT(ENGINE!$X$6)))),FALSE)),'
    '"OK — spread BET 1.5, spread toggle Y, independent of totals",'
    '"CHECK — "'
    '&IF(SETTINGS!$B$10<>1.5,"spread BET threshold is "&SETTINGS!$B$10&"; ","")'
    '&IF(SETTINGS!$B$11<>"Y","spread BET toggle is "&SETTINGS!$B$11&"; ","")'
    '&IF(IFERROR(AND(ISNUMBER(SEARCH("SETTINGS!$B$10",FORMULATEXT(ENGINE!$X$6))),'
    'ISNUMBER(SEARCH("SETTINGS!$B$11",FORMULATEXT(ENGINE!$X$6))),'
    'NOT(ISNUMBER(SEARCH("SETTINGS!$B$52",FORMULATEXT(ENGINE!$X$6))))),FALSE),"",'
    '"spread formula does not use its own independent controls; "))')

AUDIT_A13 = "TOTALS config: thresholds 2.0/3.0/6.0, totals toggle N, formula independent of spreads"
AUDIT_B13 = (
    '=IF(AND(SETTINGS!$B$49=2,SETTINGS!$B$50=3,SETTINGS!$B$51=6,SETTINGS!$B$52="N",'
    'IFERROR(ISNUMBER(SEARCH("SETTINGS!$B$49",FORMULATEXT(ENGINE!$AB$6))),FALSE),'
    'IFERROR(ISNUMBER(SEARCH("SETTINGS!$B$50",FORMULATEXT(ENGINE!$AB$6))),FALSE),'
    'IFERROR(ISNUMBER(SEARCH("SETTINGS!$B$51",FORMULATEXT(ENGINE!$AB$6))),FALSE),'
    'IFERROR(ISNUMBER(SEARCH("SETTINGS!$B$52",FORMULATEXT(ENGINE!$AB$6))),FALSE),'
    'IFERROR(NOT(ISNUMBER(SEARCH("SETTINGS!$B$10",FORMULATEXT(ENGINE!$AB$6)))),FALSE),'
    'IFERROR(NOT(ISNUMBER(SEARCH("SETTINGS!$B$11",FORMULATEXT(ENGINE!$AB$6)))),FALSE)),'
    '"OK — totals 2.0/3.0/6.0, totals toggle N, independent of spreads",'
    '"CHECK — "'
    '&IF(OR(SETTINGS!$B$49<>2,SETTINGS!$B$50<>3,SETTINGS!$B$51<>6),'
    '"totals thresholds not 2.0/3.0/6.0; ","")'
    '&IF(SETTINGS!$B$52<>"N","totals BET toggle is "&SETTINGS!$B$52&"; ","")'
    '&IF(IFERROR(AND(ISNUMBER(SEARCH("SETTINGS!$B$52",FORMULATEXT(ENGINE!$AB$6))),'
    'NOT(ISNUMBER(SEARCH("SETTINGS!$B$11",FORMULATEXT(ENGINE!$AB$6)))),'
    'NOT(ISNUMBER(SEARCH("SETTINGS!$B$10",FORMULATEXT(ENGINE!$AB$6))))),FALSE),"",'
    '"totals formula still references the spread controls; "))')

CHANGELOG = (
    "v0.8.9",
    datetime.datetime(2026, 8, 26),
    ("SPREAD BET RULE + TOTALS SEPARATION. (1) Spread BET threshold SETTINGS!B10 changed from 3.0 "
     "to 1.5, so an absolute ATS edge of 1.50 or greater is now BET; 1.49 and below remain under the "
     "existing lower classifications. (2) Spread BET labels enabled - SETTINGS!B11 N -> Y; B11 now "
     "controls SPREADS ONLY. (3) Totals thresholds separated onto dedicated cells B49/B50/B51 and "
     "PRESERVED at exactly 2.0 / 3.0 / 6.0, the values previously produced by B8*2 / B9*2 / B10*2. "
     "(4) Totals BET toggle separated onto dedicated cell B52 and RETAINED at N, so the spread "
     "toggle can no longer reach the totals market. ENGINE!AB no longer references B10 or B11. "
     "(5) No projection, rating, edge, side or totals output changed: totals remain disabled "
     "(B22/B23 blank) and every totals label stays blank. Four lined games move INVESTIGATE -> BET "
     "on spreads: Sacramento State at Eastern Michigan, North Carolina at TCU, San Jose State at "
     "USC, New Mexico State at Florida State. Memphis at UNLV remains LEAN and QB UNCERTAIN. "
     "SUPERSEDES v0.8.9 REV 1, which left the BET toggle shared with totals."),
    "Approved spread-threshold rule; REV 1 superseded",
)


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def txt(v):
    if isinstance(v, ArrayFormula):
        return v.text
    return v if isinstance(v, str) and v.startswith("=") else None


def main():
    got = sha256(SRC)
    assert got == SRC_SHA, f"source v0.8.8 is not the expected artifact: {got}"
    print(f"source v0.8.8 SHA-256 verified: {got}")

    shutil.copyfile(SRC, OUT + ".building.xlsx")
    wb = openpyxl.load_workbook(OUT + ".building.xlsx")
    st, en, au, cl = wb["SETTINGS"], wb["ENGINE"], wb["AUDIT"], wb["CHANGELOG"]

    assert st["B10"].value == 3 and st["B11"].value == "N"
    assert st["B8"].value == 1 and st["B9"].value == 1.5
    assert st["B22"].value is None and st["B23"].value is None
    for r in range(47, 53):
        for c in range(1, 6):
            assert st.cell(row=r, column=c).value is None, f"SETTINGS row {r} must be empty"
    assert cl.cell(row=87, column=1).value is None, "CHANGELOG row 87 must be free"
    print("preconditions verified: B10=3, B11='N', rows 47-52 empty, CHANGELOG 87 free, totals inert")

    written = []

    # 7. disambiguate the spread labels
    for coord, text in SPREAD_LABELS.items():
        st[coord].value = text; written.append(f"SETTINGS!{coord}")

    # 1-3. dedicated totals block, production toggle N
    for coord, (label, val) in TOTALS_BLOCK.items():
        st[coord].value = label; written.append(f"SETTINGS!{coord}")
        if val is not None:
            bc = "B" + coord[1:]
            st[bc].value = val; written.append(f"SETTINGS!{bc}")

    # 4-5. spread controls
    st["B10"].value = 1.5; written.append("SETTINGS!B10")
    st["B11"].value = "Y"; written.append("SETTINGS!B11")

    # 6. totals formula -> dedicated thresholds AND dedicated toggle
    n = 0
    for r in range(6, 1006):
        f = txt(en.cell(row=r, column=28).value)
        if not f:
            continue
        new = f
        for old, rep in REPOINT:
            new = new.replace(old, rep)
        for bad in ("SETTINGS!$B$8", "SETTINGS!$B$9", "SETTINGS!$B$10", "SETTINGS!$B$11"):
            assert bad not in new, f"AB{r} still references {bad}"
        back = new
        for old, rep in REPOINT:
            back = back.replace(rep, old)
        assert back == f, f"AB{r}: more than the four control tokens changed"
        en.cell(row=r, column=28).value = new
        written.append(f"ENGINE!AB{r}")
        n += 1
    assert n == 1000, f"expected 1000 totals formulas, got {n}"
    print(f"totals classification repointed: {n} cells -> B49/B50/B51 + toggle B52")

    # ENGINE!X must be untouched and still use the spread controls
    x = txt(en["X6"].value)
    assert "SETTINGS!$B$10" in x and "SETTINGS!$B$11" in x
    assert "SETTINGS!$B$51" not in x and "SETTINGS!$B$52" not in x
    print("ENGINE!X confirmed untouched and still bound to the spread controls")

    # audit repair - BOTH rows
    au["A12"].value = AUDIT_A12; written.append("AUDIT!A12")
    au["B12"].value = AUDIT_B12; written.append("AUDIT!B12")
    au["A13"].value = AUDIT_A13; written.append("AUDIT!A13")
    au["B13"].value = AUDIT_B13; written.append("AUDIT!B13")

    # changelog
    for i, val in enumerate(CHANGELOG, start=1):
        cl.cell(row=87, column=i).value = val
        written.append(f"CHANGELOG!{chr(64+i)}87")

    assert st["B22"].value is None and st["B23"].value is None, "totals must stay disabled"
    assert st["B52"].value == "N", "totals BET toggle must ship N"
    assert st["B8"].value == 1 and st["B9"].value == 1.5

    wb.save(OUT + ".building.xlsx")
    os.replace(OUT + ".building.xlsx", OUT)
    print(f"\ncells written: {len(written)}")
    print(f"written: {OUT}")
    print(f"REV 2 SHA-256: {sha256(OUT)}")
    assert sha256(SRC) == SRC_SHA, "authoritative v0.8.8 was modified by the build"
    print("authoritative v0.8.8 confirmed unmodified")
    return 0


if __name__ == "__main__":
    sys.exit(main())

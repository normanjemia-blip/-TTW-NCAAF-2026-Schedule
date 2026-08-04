import shutil, os, datetime, openpyxl
from openpyxl.worksheet.formula import ArrayFormula
ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
SRC=f"{ROOT}/workbook_v0.7.6_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.6_CANDIDATE.xlsx"
OUT=f"{ROOT}/workbook_v0.7.7_candidate"; DST=f"{OUT}/TTW_NCAAF_Power_Ratings_2026_v0.7.7_CANDIDATE.xlsx"
D8=datetime.datetime(2026,8,3)
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
os.makedirs(OUT,exist_ok=True); shutil.copyfile(SRC,DST)
wb=openpyxl.load_workbook(DST); qb=wb["QB VALUES"]; tm=wb["TEAM MAP"]
R={tm.cell(row=r,column=1).value:r for r in range(6,144)}
AQB,BV,AV,CONF,UPD,NOTE=5,4,6,8,11,12
def guard(r):
    for c in (AQB,BV,AV,CONF,UPD,NOTE): assert not isf(qb.cell(row=r,column=c).value), f"formula row{r}"
def setrec(ab, aqb=None, code=None, clear_vals=False, note=None, expect_code=None):
    r=R[ab]; guard(r)
    if expect_code: assert qb.cell(row=r,column=CONF).value==expect_code, f"{ab} expected {expect_code}"
    if aqb is not None: qb.cell(row=r,column=AQB).value=aqb
    if code is not None: qb.cell(row=r,column=CONF).value=code
    if clear_vals:
        assert qb.cell(row=r,column=BV).value==0
        qb.cell(row=r,column=BV).value=None; qb.cell(row=r,column=AV).value=None
    qb.cell(row=r,column=UPD).value=D8
    qb.cell(row=r,column=NOTE).value=note
    print(f"  {ab:6} row{r:>4} code={qb.cell(row=r,column=CONF).value} D={qb.cell(row=r,column=BV).value!r}")

print("=== MAC ===")
setrec("M-OH", aqb="Open (David McComb / Thomas Gotkowski / Caleb Heavner)", expect_code="L",
 note=("7D.2 CONFLICT RESOLVED 2026-08-03: both prior claims reconciled. Thomas Gotkowski IS on the official Miami roster "
  "(RS-So; 5 games, 3 starts in 2025 incl. the MAC title-clinching win). The 'Kansas transfer' is DAVID McCOMB (RS-Fr), "
  "NOT Isaiah Marshall - Marshall remains on Kansas's roster. Three-way spring competition: Gotkowski, McComb, and senior "
  "transfer Caleb Heavner; entering MAC Media Days the Kansas transfer appeared to hold the projected starting role. "
  "Genuine open competition -> L retained; values blank; UNCERTAIN. RECHECK: official depth chart / staff naming."))
setrec("BUFF", aqb="Open (Jason Wright / Mason Cumbie / Elijah Holmes)", expect_code="L",
 note=("DEFECT CORRECTED 2026-08-03: prior entry listed CJ Ogbonna, who was Buffalo's 2024 SENIOR starter (13 starts, "
  "2,381 yds, 19 TD) and is no longer on the roster - a stale record. HC Pete Lembo's current room is Jason Wright, Mason "
  "Cumbie and Elijah Holmes (Wingate transfer, three seasons there). Buffalo News reported UB 'still searching for a "
  "starting quarterback' after the spring game. L retained; values blank; UNCERTAIN."))
setrec("SAC", expect_code="M",
 note=("VERIFIED 2026-08-03: Carson Conklin CONFIRMED returning to Sacramento State for 2026 - listed on the official "
  "hornetsports.com 2026 roster and announced by Sac State Football ('HE'S BACK') after playing 2025 at Fresno State. He "
  "was the Hornets' 2024 starter and was named Fresno State's starter for the stretch run in Oct 2025. Experienced "
  "returning-to-program starter, no competition reported -> M retained (no formal 2026 naming located). NOTE: Sacramento "
  "State is a transitional program; workbook transitional rules apply separately."))

print("=== CONFERENCE USA ===")
setrec("FIU", expect_code="M",
 note="VERIFIED 2026-08-03: JJ Kohl (Appalachian State transfer) confirmed as FIU's new starting quarterback in current CUSA QB coverage. Transfer expected to start -> M retained; no official naming located.")
setrec("KENN", aqb="Rickie Collins (leader; Landon Varnes competing)", expect_code="M",
 note="VERIFIED 2026-08-03: Rickie Collins (Syracuse transfer) is Kennesaw State's new QB, but must HOLD OFF juco transfer Landon Varnes in fall practice. Clear leader with real competition -> M retained (not H); entry refined to name the competitor.")
setrec("MOST", aqb="Open competition (new HC Casey Woods)", expect_code="L",
 note="VERIFIED 2026-08-03: Missouri State welcomes new head coach Casey Woods and has a quarterback battle; NO starting quarterback has been definitively named. Genuine open competition -> L retained; values blank; UNCERTAIN.")
setrec("NMSU", expect_code="L",
 note="VERIFIED 2026-08-03: New Mexico State remains in a rebuild with a likely quarterback competition; NO starter named. L retained; values blank; UNCERTAIN.")

print("=== AMERICAN ===")
setrec("ECU", expect_code="M",
 note="VERIFIED 2026-08-03: Chaston Ditta likely takes the reins for East Carolina after starting the 2025 Military Bowl. Projected starter awaiting formal announcement -> M retained.")
setrec("NAVY", expect_code="M",
 note="VERIFIED 2026-08-03: Braxton Woodson confirmed as Navy's projected starter in current American QB coverage. M retained; no official naming located.")
setrec("RICE", expect_code="M",
 note="VERIFIED 2026-08-03: Jacurri Brown confirmed as Rice's quarterback in current American QB coverage. M retained; no official naming located.")
setrec("TEM", aqb="Jaxon Smolik (inside track; Ajani Sheppard competing)", expect_code="M",
 note="VERIFIED 2026-08-03: Temple's spring battle was between transfers Jaxon Smolik (Penn State) and Ajani Sheppard (Washington State), with SMOLIK OWNING THE INSIDE TRACK. Clear leader with competition -> M retained; entry refined.")
setrec("TLSA", expect_code="M",
 note="VERIFIED 2026-08-03: Baylor Hayes confirmed as Tulsa's quarterback in current American QB coverage. M retained; no official naming located.")

print("=== MOUNTAIN WEST ===")
setrec("NDSU", expect_code="M",
 note="VERIFIED 2026-08-03: Nathan Hayes is North Dakota State's projected starter and a FIRST-TIME starter (senior); optimism cited given NDSU's QB-development record. Projected starter, unproven -> M retained. NOTE: NDSU is transitional; workbook transitional rules apply separately.")
setrec("NEV", aqb="Open competition (fall camp)", expect_code="L",
 note="VERIFIED 2026-08-03: Nevada is going through a quarterback COMPETITION in fall camp; no leader established in current coverage. L retained; entry generalized from a single name; values blank; UNCERTAIN.")
setrec("NIU", aqb="Taron Dickens (Western Carolina transfer)", code="M", expect_code="L",
 note=("DEFECT CORRECTED 2026-08-03: prior entry listed 'Open (Davidson / Macon / Hamric)' - none appear in current NIU "
  "coverage. Verified: TARON DICKENS, Western Carolina transfer and FCS WALTER PAYTON AWARD RUNNER-UP in 2025, is NIU's "
  "quarterback, paired with new OC Tony Petersen (ex-Illinois State, hired March 2026). High-profile transfer widely "
  "expected to start -> L upgraded to M. Per the L->M rule numerical values REMAIN BLANK (valuation unresolved), so "
  "status correctly stays UNCERTAIN."))
setrec("SJSU", aqb="Open (three-way; starter not selected)", expect_code="L",
 note="VERIFIED 2026-08-03: HC Ken Niumatalolo enters year three with a TOTAL RESET and must pick a quarterback among THREE separate candidates. Genuine open competition -> L retained; entry generalized from a single name; values blank; UNCERTAIN.")
setrec("UTEP", code="H", expect_code="M",
 note=("VERIFIED 2026-08-03: EJ Colson (Incarnate Word transfer) was NAMED UTEP's starting quarterback in the SPRING. An "
  "official naming with no competition reported meets the H standard -> M upgraded to H. Zero initialization RETAINED "
  "(correct for H under the preseason baseline-delta method); status remains OK. RECHECK: any camp change to the spring "
  "depth chart."))
setrec("WYO", expect_code="M",
 note="VERIFIED 2026-08-03: Tyler Hughes (William & Mary transfer) confirmed as Wyoming's incoming QB - 2,330 pass yds, 20 TD, 670 rush yds, 11 rush TD in 2025; previously worked with the Wyoming staff at W&M in 2022-23. Transfer widely expected to start -> M retained.")

print("=== PAC-12 ===")
setrec("CSU", aqb="Hauss Hejny (K'saan Farrar competing)", expect_code="L",
 note="VERIFIED 2026-08-03: Hauss Hejny is Colorado State's projected starter but is LOCKED IN A BATTLE with UConn transfer K'saan Farrar that is expected to LAST INTO FALL CAMP. Genuine unresolved competition -> L retained; entry refined. (Context: CSU's 2026 HC is Jim Mora, who left UConn - see the 7D.3A UConn finding.)")
setrec("FRES", aqb="Open (three-way battle into August)", expect_code="L",
 note="VERIFIED 2026-08-03: Fresno State has a THREE-WAY quarterback battle 'likely to run deep into August' involving a Western Michigan transfer and Mercer transfer Braden Atkinson (2025 Jerry Rice Award winner, FCS freshman of the year). Prior single-name entry could not be corroborated. L retained; entry generalized; values blank; UNCERTAIN.")
setrec("ORST", aqb="Maalik Murphy (leader; Braden Atkinson pushing)", code="M", expect_code="L",
 note=("VERIFIED 2026-08-03 (with documented tension): SI headline reports ex-Texas QB Maalik Murphy NAMED STARTER at "
  "Oregon State, while SI's Pac-12 QB tiers describe Murphy with 'a slight upper hand' and a young FCS transfer 'right on "
  "his tail in fall camp.' Both agree Murphy LEADS. Clear leader -> L upgraded to M (NOT H, because the naming claim is "
  "contradicted by active-competition language). Per the L->M rule values REMAIN BLANK; status stays UNCERTAIN. "
  "RECHECK: official Oregon State depth chart."))
setrec("USU", expect_code="M",
 note="VERIFIED 2026-08-03: Utah State reunites former Virginia OC Robert Anae with RETURNING quarterback McCae Hillstead. Returning starter with a new coordinator -> M retained; no formal naming located.")

print("=== SUN BELT ===")
setrec("APP", aqb="Open (Malachi Singleton / Hasselbeck)", code="L", clear_vals=True, expect_code="M",
 note=("DEFECT CORRECTED 2026-08-03: prior entry asserted Malachi Singleton as the likely starter (M). Current coverage "
  "describes Singleton AND Hasselbeck as 'the main competitors for the starting gig' with NO leader identified. Per the "
  "standard, M requires a clear likely starter and must not be assigned merely because one player is the most recognizable "
  "name - Singleton is an Arkansas/Purdue transfer, which is name recognition, not evidence of a lead. -> M downgraded to "
  "L. Numerical zeros CLEARED to blank per the approved Akron consistency methodology; status becomes UNCERTAIN."))
setrec("ARST", expect_code="L",
 note="VERIFIED 2026-08-03: HC Butch Jones at Sun Belt Media Days said the starting quarterback race CONTINUES with 'four real capable players' at the position. Genuine open competition -> L retained; values blank; UNCERTAIN.")
setrec("CCU", aqb="Open (Deuce Bailey among candidates)", expect_code="L",
 note="VERIFIED 2026-08-03: Coastal Carolina's incoming transfers include POTENTIAL starting quarterback Deuce Bailey, who followed HC Beard from Missouri State. 'Potential' starter, room unsettled -> L retained; entry refined.")
setrec("JMU", expect_code="M",
 note="VERIFIED 2026-08-03 (with tension): James Madison 'reloads at QB' with Memphis transfer Arrington Maiden, though the same coverage says he is 'expected to COMPETE for the starting role.' Transfer expected to start but not confirmed as the settled leader -> M retained (not H). RECHECK: JMU depth chart or staff naming.")

# --- CHANGELOG + banner ---
cl=wb["CHANGELOG"]; last=0
for rr in range(1,cl.max_row+1):
    if any(cl.cell(row=rr,column=c).value not in (None,"") for c in range(1,7)): last=rr
rows=[("v0.7.7","2026-08-03",
 "Phase 7D.4 FINAL G5 BATCH: 31 teams in scope; 26 VERIFIED, 5 not verified (North Texas, Texas State, Washington State, "
 "Georgia Southern, Old Dominion - dates NOT refreshed). TWO STALE-ENTRY DEFECTS: Buffalo listed CJ Ogbonna, its 2024 "
 "SENIOR starter no longer on the roster (corrected to the Wright/Cumbie/Holmes room), and Northern Illinois listed "
 "'Davidson/Macon/Hamric' none of whom appear in current coverage (corrected to Taron Dickens, Western Carolina transfer "
 "and FCS Walter Payton Award runner-up; L->M). ONE CLASSIFICATION DEFECT: Appalachian State M->L - Singleton and "
 "Hasselbeck are 'main competitors' with no leader; M was resting on name recognition. Zeros CLEARED to blank.",
 "Final G5 verification batch"),
 ("v0.7.7","2026-08-03",
 "Phase 7D.4 MIAMI (OH) CONFLICT RESOLVED: both 7D.2 claims reconciled. Thomas Gotkowski IS on the official roster (3 "
 "starts in 2025 incl. the MAC title-clinching win); the 'Kansas transfer' is DAVID McCOMB, not Isaiah Marshall (who "
 "remains at Kansas). Three-way race with senior transfer Caleb Heavner; L retained. OTHER CODE CHANGES: UTEP M->H (EJ "
 "Colson officially NAMED starter in spring; zeros retained) and Oregon State L->M (Maalik Murphy leads; a 'named "
 "starter' headline is in tension with active-competition language, so M not H; values stay blank).",
 "Final G5 verification batch"),
 ("v0.7.7","2026-08-03",
 "Phase 7D.4 RESULT: backlog 31 -> 5. Counts 64 H / 43 M / 31 L; status 101 OK / 37 UNCERTAIN; 0 nonzero QB values; "
 "blank 37 / zero 101. Five teams remain unverified and are NOT date-refreshed. Promotion recommendation: OPTION B - "
 "DEFER AND REPAIR, because the backlog has not reached zero.","Final G5 batch result")]
for i,(v,d,c,rn) in enumerate(rows):
    rr=last+1+i
    cl.cell(row=rr,column=1).value=v; cl.cell(row=rr,column=2).value=d
    cl.cell(row=rr,column=3).value=c; cl.cell(row=rr,column=4).value=rn
wb["START HERE"]["A1"].value=("TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.7.7 CANDIDATE — final G5 batch: 26 of 31 verified, "
 "3 defects corrected (Buffalo, Northern Illinois, Appalachian State), Miami OH conflict resolved; backlog 5; "
 "NOT AUTHORITATIVE, NOT PROMOTED)")
wb.save(DST)
print(f"\nCHANGELOG rows {last+1}-{last+len(rows)}; saved {DST}")

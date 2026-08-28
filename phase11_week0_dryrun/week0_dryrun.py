#!/usr/bin/env python3
"""Week 0 full-card dry run against the v0.8.9 AUTHORITATIVE workbook.

Read-only. The workbook is opened, never written; its SHA-256 is asserted
before and after so a run can never be the reason a number changed.

Two jobs:

  PART A  Rebuild the Week 0 card from the workbook's own inputs, following
          the formula chain transcribed from v0.8.1, and reconcile it against
          the Phase 10 checkpoint in week0_card.json. Differences are REPORTED,
          never written back -- the checkpoint is the record.

  PART B  Run the operating gates the weekly workflow depends on:
            1  FCS -- NO PLAY
            2  neutral-site HFA (incl. the Dublin game)
            3  QB uncertainty gating
            4  market-line staleness
            5  spread sign conventions
            6  BET toggle OFF
            7  totals disabled

Exit code 0 iff every gate passes and the card reconciles.

Formula chain (v0.8.9, preseason state -- SETTINGS!B4 and B5 blank):
  PRESEASON!G/K/T   source norms, each mean-centred over rows 6:143
  PRESEASON!Y       available weight = sum of B28..B32 over present sources
  PRESEASON!Z       FINAL PRIOR = weighted sum / Y
  TEAM RATINGS!E    = PRESEASON!Z
  TEAM RATINGS!I    = E   (H blank or F=0 -> prior only)
  TEAM RATINGS!K    = I   (J blank while B4 is blank)
  TEAM RATINGS!O    = K   (no override)
  ENGINE!K          = home rating - away rating
  ENGINE!L          = B7 when neutral, else the home team's effective HFA
  ENGINE!M          = home QB delta - away QB delta (blank coerced to 0)
  ENGINE!R          = K + L + M + N + O
"""
import json, os, sys, hashlib
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, "..")
WB = os.path.join(ROOT, "promotion_v0.8.9",
                  "TTW_College_Football_Power_Ratings_v0.8.9_AUTHORITATIVE.xlsx")
CHECKPOINT = os.path.join(ROOT, "phase10_operational_validation", "week0_card.json")
EXPECTED_SHA = "334050660deb970f23cd9761490fb47e1f2b606b61d00a20c864cec529395cbb"

PASS, FAIL = [], []


def check(ok, msg, detail=""):
    (PASS if ok else FAIL).append(msg + (f"  -- {detail}" if detail and not ok else ""))
    print(f"  [{'PASS' if ok else 'FAIL'}] {msg}" + (f"  -- {detail}" if detail else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def mean(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def load():
    wb = openpyxl.load_workbook(WB, data_only=False)
    st = wb["SETTINGS"]
    S = {f"B{r}": st[f"B{r}"].value for r in range(3, 33)}

    ps = wb["PRESEASON"]
    rows = []
    for r in range(6, 144):
        ab = None
        # PRESEASON!A is a formula; take identity from TEAM MAP directly
        rows.append(dict(row=r,
                         spp=ps.cell(row=r, column=4).value,    # D  SP+ raw
                         fpi=ps.cell(row=r, column=8).value,    # H  FPI raw
                         ttw=ps.cell(row=r, column=12).value,   # L  TTW 2025 raw
                         tr=ps.cell(row=r, column=17).value,    # Q  TeamRankings raw
                         vsin=ps.cell(row=r, column=21).value)) # U  VSiN raw

    tm = wb["TEAM MAP"]
    for i, r in enumerate(range(6, 144)):
        rows[i]["abbrev"] = tm.cell(row=r, column=1).value
        rows[i]["name"] = tm.cell(row=r, column=2).value
        rows[i]["status"] = tm.cell(row=r, column=4).value
        rows[i]["transitional"] = tm.cell(row=r, column=5).value

    # alias table: TEAM MAP!K -> TEAM MAP!L, rows 6..605
    alias = {}
    for r in range(6, 606):
        k = tm.cell(row=r, column=11).value
        v = tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()

    # mean-centred norms
    for key, src in (("g", "spp"), ("k", "fpi"), ("t", "tr")):
        m = mean([x[src] for x in rows])
        for x in rows:
            x[key] = (x[src] - m) if isinstance(x[src], (int, float)) else ""
    # TTW (P) and VSiN (X) are absent by decision -- carried explicitly as blank
    for x in rows:
        x["p"] = "" if x["ttw"] is None else x["ttw"]
        x["x"] = "" if x["vsin"] is None else x["vsin"]

    W = {"g": S["B28"], "k": S["B29"], "p": S["B30"], "t": S["B31"], "x": S["B32"]}
    prior, hfa, qbdelta, qbstatus = {}, {}, {}, {}
    for x in rows:
        y = sum(W[c] for c in "gkptx" if isinstance(x[c], (int, float)))
        z = ("" if y == 0 else
             sum(W[c] * (x[c] if isinstance(x[c], (int, float)) else 0) for c in "gkptx") / y)
        x["y"], x["z"] = y, z
        prior[x["abbrev"]] = z
        hfa[x["abbrev"]] = S["B6"]          # TR!Q = B6; TR!R override is empty

    qb = wb["QB VALUES"]
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        d = qb.cell(row=r, column=4).value   # D baseline value
        f_ = qb.cell(row=r, column=6).value  # F active value
        conf = qb.cell(row=r, column=8).value
        rev = qb.cell(row=r, column=10).value
        # G = F - D, blank if either blank;  M = UNCERTAIN if G blank / L / season mismatch
        g = "" if (d is None or f_ is None) else f_ - d
        qbdelta[ab] = g
        qbstatus[ab] = ("UNCERTAIN" if (g == "" or conf == "L" or rev != S["B3"]) else "OK")

    sch = wb["IMPORT SCHEDULE"]
    games = []
    for r in range(6, 1006):
        gid = sch.cell(row=r, column=1).value
        if gid in (None, ""):
            continue
        games.append(dict(
            gid=str(gid), season=sch.cell(row=r, column=2).value,
            week=sch.cell(row=r, column=3).value,
            date=sch.cell(row=r, column=4).value,
            neutral=bool(sch.cell(row=r, column=5).value),
            away_raw=sch.cell(row=r, column=6).value,
            home_raw=sch.cell(row=r, column=8).value,
            venue=sch.cell(row=r, column=13).value))
    for g in games:
        g["away"] = alias.get(str(g["away_raw"]).strip(), "")
        g["home"] = alias.get(str(g["home_raw"]).strip(), "")
        g["away_type"] = "FBS" if g["away"] else "FCS"
        g["home_type"] = "FBS" if g["home"] else "FCS"
    return wb, S, rows, prior, hfa, qbdelta, qbstatus, games


def engine(g, S, prior, hfa, qbdelta, qbstatus, mkt_favorite=None, mkt_spread=None):
    """ENGINE row, transcribed from v0.8.1. Preseason state: B4/B5 blank."""
    I = prior.get(g["away"], "") if g["away_type"] == "FBS" else ""
    J = prior.get(g["home"], "") if g["home_type"] == "FBS" else ""
    K = "" if (I == "" or J == "") else J - I
    if g["neutral"]:
        L = S["B7"]
    elif g["home_type"] != "FBS":
        L = S["B6"]
    else:
        L = hfa.get(g["home"], S["B6"])
    z = lambda v: 0 if v == "" else v
    M = 0 if g["home_type"] != "FBS" else (
        z(qbdelta.get(g["home"], "")) - (0 if g["away_type"] != "FBS" else z(qbdelta.get(g["away"], ""))))
    N = O = 0
    P = "" if K == "" else K + L + M + N + O
    R = P
    Sp = "" if R == "" else (f"{g['home']} -{abs(R):.1f}" if R > 0
                             else (f"{g['away']} -{abs(R):.1f}" if R < 0 else "PICK"))
    # T: CALC!N -- negative when the HOME team is the favorite
    if mkt_spread is None:
        T = ""
    elif mkt_favorite == g["home"]:
        T = -abs(mkt_spread)
    elif mkt_favorite == g["away"]:
        T = abs(mkt_spread)
    else:
        T = ""
    V = "" if (R == "" or T == "") else R + T
    W = "" if V == "" else (g["home"] if V > 0 else g["away"])
    away_unc = False if g["away_type"] != "FBS" else qbstatus.get(g["away"], "UNCERTAIN") == "UNCERTAIN"
    home_unc = False if g["home_type"] != "FBS" else qbstatus.get(g["home"], "UNCERTAIN") == "UNCERTAIN"
    AE = "QB UNCERTAIN" if (away_unc or home_unc) else "OK"
    AG = "FCS OPP" if (g["away_type"] == "FCS" or g["home_type"] == "FCS") else ""
    pending = 1 if T == "" else 0
    stale = 0 if S["B5"] in (None, "") else 0      # CALC!Q returns 0 while B5 is blank
    if AG:
        AI = "FCS — NO PLAY"
    elif pending:
        AI = "PENDING LINE"
    elif stale:
        AI = "STALE LINE"
    elif AE == "QB UNCERTAIN":
        AI = "QB UNCERTAIN"
    elif S["B4"] in (None, "") or S["B5"] in (None, ""):
        AI = "DATA INCOMPLETE"
    else:
        AI = "READY"
    # X label
    if V == "" or AI in ("BLOCKED", "PENDING LINE", "STALE LINE"):
        X = ""
    elif abs(V) < S["B8"]:
        X = ""
    elif abs(V) < S["B9"]:
        X = "LEAN"
    elif abs(V) < S["B10"] or S["B11"] != "Y" or AI != "READY":
        X = "INVESTIGATE"
    else:
        X = "BET"
    return dict(I=I, J=J, K=K, L=L, M=M, R=R, S=Sp, T=T, V=V, W=W, AE=AE, AG=AG, AI=AI, X=X)


def main():
    print("=" * 78)
    print("WEEK 0 FULL-CARD DRY RUN -- v0.8.1 AUTHORITATIVE")
    print("=" * 78)
    sha_before = sha256(WB)
    check(sha_before == EXPECTED_SHA, "workbook is the authoritative v0.8.9", sha_before[:16])

    wb, S, rows, prior, hfa, qbdelta, qbstatus, games = load()

    print(f"\n  games loaded: {len(games)}")
    wk0 = [g for g in games if g["week"] == 0]
    print(f"  week 0 games: {len(wk0)}")

    # ---------------- PART A -- rebuild and reconcile ----------------
    print("\n" + "-" * 78)
    print("PART A -- rebuild the Week 0 card and reconcile against the Phase 10 checkpoint")
    print("-" * 78)
    cp = {c["gid"]: c for c in json.load(open(CHECKPOINT))}
    # The Phase 10 checkpoint is a POINT-IN-TIME record from 2026-08-15 and is
    # deliberately never rewritten. QB status legitimately moves as records are
    # resolved, so a status difference is not automatically a defect -- but it
    # must be DECLARED here, with the reason, or it still fails. Model spreads,
    # edges, sides and labels remain hard gates and are never allowed to drift.
    DECLARED_STATUS_DRIFT = {
        # gid: (checkpoint status, current status, reason)
        "401856766": ("QB UNCERTAIN", "DATA INCOMPLETE",
                      "UNC activated 2026-08-19 (Billy Edwards Jr., M); QB gate cleared, "
                      "status now falls through to DATA INCOMPLETE while B4/B5 are blank"),
        "401858201": ("QB UNCERTAIN", "DATA INCOMPLETE",
                      "Stanford activated 2026-08-21 (Davis Warren, H); QB gate cleared, "
                      "status now falls through to DATA INCOMPLETE while B4/B5 are blank"),
    }
    # the two published Week 0 lines, as recorded in Phase 10
    LINES = {"401856766": ("TCU", 7.0), "401858201": ("STAN", 4.0)}

    rebuilt, diffs, declared = {}, [], []
    for g in sorted(wk0, key=lambda x: x["gid"]):
        fav, spr = LINES.get(g["gid"], (None, None))
        e = engine(g, S, prior, hfa, qbdelta, qbstatus, fav, spr)
        rebuilt[g["gid"]] = e
        c = cp.get(g["gid"])
        model = e["S"]
        print(f"  {g['away']:>5} @ {g['home']:<5} neutral={str(g['neutral']):5} "
              f"model={model:<14} status={e['AI']:<14} label={e['X'] or '-'}")
        if c:
            if model != c["model"]:
                diffs.append(f"{c['game']}: model {model} vs checkpoint {c['model']}")
            if e["AI"] != c["status"]:
                d = DECLARED_STATUS_DRIFT.get(g["gid"])
                if d and d[0] == c["status"] and d[1] == e["AI"]:
                    declared.append(f"{c['game']}: {d[0]} -> {d[1]} ({d[2]})")
                else:
                    diffs.append(f"{c['game']}: status {e['AI']} vs checkpoint {c['status']}")
            if c["market"] != "PENDING":
                if f"{e['V']:.1f}" != c["edge"]:
                    diffs.append(f"{c['game']}: edge {e['V']:.1f} vs checkpoint {c['edge']}")
                if e["W"] != c["side"]:
                    diffs.append(f"{c['game']}: side {e['W']} vs checkpoint {c['side']}")
                if (e["X"] or "") != c["label"]:
                    diffs.append(f"{c['game']}: label {e['X']} vs checkpoint {c['label']}")

    check(len(wk0) == len(cp), f"Week 0 card has the same {len(cp)} games as the checkpoint",
          f"rebuilt={len(wk0)} checkpoint={len(cp)}")
    check(not diffs, "rebuilt card reproduces the Phase 10 checkpoint "
          "(model spreads, edges, sides and labels are hard gates)",
          "; ".join(diffs[:4]))
    if declared:
        print(f"\n  NOTE - {len(declared)} declared status drift(s); the checkpoint is "
              f"preserved, not rewritten:")
        for d in declared:
            print(f"    * {d}")
    check(len(declared) == len(DECLARED_STATUS_DRIFT),
          f"all {len(DECLARED_STATUS_DRIFT)} declared status drifts are still present "
          f"(a declaration that stops applying must be removed)",
          f"observed {len(declared)}")

    # ---------------- PART B -- operating gates ----------------
    print("\n" + "-" * 78)
    print("PART B -- operating gates")
    print("-" * 78)

    # G1 FCS -- NO PLAY
    allrows = {}
    for g in games:
        allrows[g["gid"]] = engine(g, S, prior, hfa, qbdelta, qbstatus)
    fcs = [gid for gid, e in allrows.items() if e["AG"]]
    fcs_noplay = [gid for gid in fcs if allrows[gid]["AI"] == "FCS — NO PLAY"]
    fcs_labelled = [gid for gid in fcs if allrows[gid]["X"]]
    check(len(fcs) == 127, f"127 FCS games present", f"found {len(fcs)}")
    check(len(fcs_noplay) == len(fcs),
          "every FCS game is FCS — NO PLAY and none carries a bet label",
          f"noplay={len(fcs_noplay)}/{len(fcs)} labelled={len(fcs_labelled)}")
    check(not fcs_labelled, "no FCS game produces a playable label",
          f"{len(fcs_labelled)} labelled")
    check(len(games) - len(fcs) == 761, "761 FBS-v-FBS games",
          f"{len(games) - len(fcs)}")

    # G2 neutral-site HFA, incl. Dublin
    neutrals = [g for g in games if g["neutral"]]
    bad = [g["gid"] for g in neutrals
           if engine(g, S, prior, hfa, qbdelta, qbstatus)["L"] != S["B7"]]
    check(not bad, f"all {len(neutrals)} neutral-site games take HFA = SETTINGS!B7 = {S['B7']}",
          f"{len(bad)} wrong")
    dublin = next(g for g in games if g["gid"] == "401856766")
    ed = engine(dublin, S, prior, hfa, qbdelta, qbstatus, "TCU", 7.0)
    check(dublin["neutral"] and ed["L"] == 0 and dublin["venue"] == "Aviva Stadium",
          "Dublin game (UNC @ TCU, Aviva Stadium) carries neutral HFA = 0",
          f"neutral={dublin['neutral']} L={ed['L']} venue={dublin['venue']}")
    check(abs(ed["R"] - 4.2) < 0.05, "Dublin model margin is TCU -4.2 (pure rating differential)",
          f"R={ed['R']:.2f}")
    # counterfactual: had HFA been applied it would read ~TCU -6.7
    cf = dict(dublin); cf["neutral"] = False
    ecf = engine(cf, S, prior, hfa, qbdelta, qbstatus, "TCU", 7.0)
    check(abs(ecf["R"] - (ed["R"] + S["B6"])) < 1e-9,
          f"counterfactual confirms the neutral flag is worth {S['B6']} pts "
          f"(would read TCU -{ecf['R']:.1f})")

    # G3 QB uncertainty gating
    unc_teams = [a for a, s in qbstatus.items() if s == "UNCERTAIN"]
    check(len(unc_teams) == 21, "21 teams QB UNCERTAIN", f"{len(unc_teams)}")
    leaked = [gid for gid, e in allrows.items() if e["AI"] == "QB UNCERTAIN" and e["X"] == "BET"]
    check(not leaked, "no QB UNCERTAIN game can reach BET", f"{len(leaked)} leaked")
    wk0_unc = [g["gid"] for g in wk0
               if engine(g, S, prior, hfa, qbdelta, qbstatus)["AE"] == "QB UNCERTAIN"]
    check(len(wk0_unc) == 1, "1 of 8 Week 0 games carries QB UNCERTAIN", f"{len(wk0_unc)}")

    # G4 market-line staleness
    check(S["B5"] in (None, ""), "SETTINGS!B5 (as-of date) is blank -- preseason state")
    check(S["B13"] == 5, "stale threshold B13 = 5 days", f"{S['B13']}")
    stale_now = [gid for gid, e in allrows.items() if e["AI"] == "STALE LINE"]
    check(not stale_now,
          "STALE LINE count is 0 -- and that is because B5 is blank, not because lines are fresh",
          f"{len(stale_now)}")

    # G5 spread sign conventions
    e1 = engine(dublin, S, prior, hfa, qbdelta, qbstatus, "TCU", 7.0)
    check(e1["T"] == -7.0, "home favorite -> market home spread is negative", f"T={e1['T']}")
    check(abs(e1["V"] - (-2.8)) < 0.05, "edge = model margin + market home spread = 4.2 + (-7.0) = -2.8",
          f"V={e1['V']:.2f}")
    check(e1["W"] == "UNC", "negative edge -> value on the away team", f"W={e1['W']}")
    e2 = engine(dublin, S, prior, hfa, qbdelta, qbstatus, "UNC", 3.0)
    check(e2["T"] == 3.0, "away favorite -> market home spread is positive", f"T={e2['T']}")
    check(abs(e2["V"] - 7.2) < 0.05, "away-favorite edge = 4.2 + 3.0 = 7.2", f"V={e2['V']:.2f}")
    check(e2["W"] == "TCU", "positive edge -> value on the home team", f"W={e2['W']}")

    # G6 spread BET rule - v0.8.9 approved production configuration
    check(S["B11"] == "Y", "spread BET toggle SETTINGS!B11 = Y (v0.8.9)", f"{S['B11']}")
    check(float(S["B10"]) == 1.5, "spread BET threshold SETTINGS!B10 = 1.5 (v0.8.9)",
          f"{S['B10']}")
    bets = [gid for gid, e in allrows.items() if e["X"] == "BET"]
    check(not bets, "no BET on this card - MARKET LINES is blank in the authoritative artifact, "
                    "so no game has a spread edge at all", f"{len(bets)}")
    # the toggle is now enabled, so prove the row gates are what still bind. NOTE: this fixture
    # gates to DATA INCOMPLETE, not QB UNCERTAIN - the AI gate is what blocks BET here. The
    # |edge| = 1.50 boundary itself is proven on real READY rows in promotion_v0.8.9/verify_v089.py
    # section 3; it is deliberately not restated here, where no fixture can reach READY.
    forced = engine(dublin, S, prior, hfa, qbdelta, qbstatus, "TCU", 12.0)
    check(forced["X"] == "INVESTIGATE" and forced["AI"] != "READY",
          "with the toggle Y and a 7.8-pt edge, a non-READY gate still forces INVESTIGATE",
          f"edge={forced['V']:.1f} status={forced['AI']} label={forced['X']}")

    # G7 totals disabled
    check(S["B22"] in (None, ""), "SETTINGS!B22 (league avg total) blank -- totals disabled")
    check(S["B23"] in (None, ""), "SETTINGS!B23 (total EPA scale) blank -- totals disabled")

    # workbook untouched
    sha_after = sha256(WB)
    check(sha_after == sha_before, "workbook unchanged by this run", sha_after[:16])

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

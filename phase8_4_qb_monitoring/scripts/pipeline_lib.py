#!/usr/bin/env python3
"""Shared helpers for the Phase 8.4 QB-resolution -> workbook-candidate pipeline.

Deviation-only methodology (binding):
  - QB VALUES columns: A Abbrev(f), B Team(f), C Baseline QB, D Baseline value,
    E Active QB, F Active value, G QB delta(f), H Confidence, I Source,
    J Reviewed for season, K Last update, L Notes, M QB status(f).
  - Team rows are 6..143 (138 teams). Abbrev constant lives in TEAM MAP!A.
  - Writable QB-input columns: C,D,E,F,H,I,J,K,L  (3,4,5,6,8,9,10,11,12).
  - Formula columns (never written): A,B,G,M  (1,2,7,13).
  - Baseline value is ALWAYS 0. Active value = 0 (zero-init) or an approved
    quarter-point deviation in [-4.00,+2.00] (baseline mismatch). Delta = F - D.
  - Nothing nonzero is written without a matching APPROVED review-log record.
"""
import os, csv, json, datetime, hashlib
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

# ---- workbook geometry -----------------------------------------------------
FIRST_ROW, LAST_ROW = 6, 143            # 138 team rows
EXPECTED_TEAMS = 138
EXPECTED_SHEETS = 21
EXPECTED_FORMULAS = 123011
SEASON_CELL = ("SETTINGS", "B3")
BANNER_CELL = ("START HERE", "A1")

COL = {  # QB VALUES column letters -> 1-based index
    "abbrev": 1, "team": 2, "baseline_qb": 3, "baseline_value": 4,
    "active_qb": 5, "active_value": 6, "qb_delta": 7, "confidence": 8,
    "source": 9, "reviewed_season": 10, "last_update": 11, "notes": 12,
    "qb_status": 13,
}
FORMULA_COLS = {1, 2, 7, 13}            # A,B,G,M
WRITABLE_COLS = {3, 4, 5, 6, 8, 9, 10, 11, 12}
# field -> writable column mapping used by the apply step
WRITE_FIELDS = {
    "baseline_qb": 3, "baseline_value": 4, "active_qb": 5, "active_value": 6,
    "confidence": 8, "source": 9, "reviewed_season": 10, "last_update": 11,
    "notes": 12,
}

APPROVED_STATUS = "APPROVED"            # pending resolution_status that is actionable
DEV_MIN, DEV_MAX = -4.00, 2.00


# ---- basic helpers ---------------------------------------------------------
def is_formula(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))

def norm(v):
    """Normalize a cell value for comparison (ArrayFormula -> its text)."""
    if isinstance(v, ArrayFormula):
        return ("ARR", v.text)
    return v

def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()

def is_iso_date(s):
    if not isinstance(s, str):
        return False
    try:
        datetime.date.fromisoformat(s)
        return True
    except ValueError:
        return False

def is_quarter_point(x):
    """True if x is a multiple of 0.25 (within float tolerance)."""
    try:
        f = float(x)
    except (TypeError, ValueError):
        return False
    return abs(round(f / 0.25) - (f / 0.25)) < 1e-9

def in_dev_bounds(x):
    try:
        f = float(x)
    except (TypeError, ValueError):
        return False
    return DEV_MIN - 1e-9 <= f <= DEV_MAX + 1e-9

def blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


# ---- workbook loading ------------------------------------------------------
def load_wb(path):
    return openpyxl.load_workbook(path)

def team_row_map(wb):
    """abbrev -> row (6..143), from TEAM MAP!A (constant abbrevs)."""
    tm = wb["TEAM MAP"]
    m = {}
    for r in range(FIRST_ROW, LAST_ROW + 1):
        ab = tm.cell(row=r, column=1).value
        if ab not in (None, ""):
            m[str(ab)] = r
    return m

def formula_count(wb):
    return sum(1 for s in wb.sheetnames for row in wb[s].iter_rows()
               for c in row if is_formula(c.value))

def season_of(wb):
    return wb[SEASON_CELL[0]][SEASON_CELL[1]].value


# ---- pending-resolution ledger --------------------------------------------
# Canonical apply-input fields (extra columns are ignored):
PENDING_FIELDS = [
    "abbrev", "team", "resolution_status", "confidence",
    "baseline_qb", "baseline_value", "active_qb", "active_value",
    "source", "source_date", "last_update", "reviewed_season", "notes",
]

def _coerce_num(v):
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str) and v.strip() != "":
        try:
            return float(v) if ("." in v or "e" in v.lower()) else int(v)
        except ValueError:
            return v
    return v

def load_pending(path):
    """Load pending resolutions from CSV or JSON -> list[dict] (lower-cased keys).

    Numeric-looking baseline_value/active_value/reviewed_season are coerced.
    """
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext == ".json":
        data = json.load(open(path))
        recs = data.get("records", data) if isinstance(data, dict) else data
        for r in recs:
            rows.append({str(k).strip().lower(): v for k, v in r.items()})
    elif ext in (".csv", ".tsv"):
        delim = "\t" if ext == ".tsv" else ","
        with open(path, newline="") as f:
            for r in csv.DictReader(f, delimiter=delim):
                rows.append({(k or "").strip().lower(): v for k, v in r.items()})
    else:
        raise ValueError(f"unsupported pending format: {ext}")
    for r in rows:
        for k in ("baseline_value", "active_value", "reviewed_season"):
            if k in r and not blank(r[k]):
                r[k] = _coerce_num(r[k])
    return rows

def approved_pending(rows):
    """Subset whose resolution_status == APPROVED (case-insensitive)."""
    return [r for r in rows
            if str(r.get("resolution_status", "")).strip().upper() == APPROVED_STATUS]


# ---- deviation review log --------------------------------------------------
def load_review_log(path):
    """Load the review log (CSV or JSON) -> list[dict] with original headers."""
    if not path or not os.path.exists(path):
        return []
    ext = os.path.splitext(path)[1].lower()
    if ext == ".json":
        data = json.load(open(path))
        return data.get("records", []) if isinstance(data, dict) else data
    with open(path, newline="") as f:
        return list(csv.DictReader(f))

def _rl_get(rec, *names):
    for n in names:
        if n in rec and rec[n] not in (None, ""):
            return rec[n]
        # case-insensitive fallback
        for k, v in rec.items():
            if str(k).strip().lower() == n.strip().lower() and v not in (None, ""):
                return v
    return ""

def find_approved_deviation(review_log, abbrev, active_qb, value):
    """Return (ok, reason). A review-log record approves `value` for (abbrev,
    active_qb) when Status == 'active', Review action == 'entered', the
    Applied/Revised deviation equals `value` (quarter-point, in bounds), and a
    Reviewer + Review date are present. 'proposed'/'under review'/blank fail.
    """
    want = float(value)
    for rec in review_log:
        if str(_rl_get(rec, "Abbrev")).strip().upper() != abbrev.upper():
            continue
        if str(_rl_get(rec, "Active QB")).strip() != str(active_qb).strip():
            continue
        status = str(_rl_get(rec, "Status")).strip().lower()
        action = str(_rl_get(rec, "Review action")).strip().lower()
        dev_raw = _rl_get(rec, "Revised deviation", "Applied deviation")
        reviewer = _rl_get(rec, "Reviewer")
        review_date = _rl_get(rec, "Review date")
        try:
            dev = float(dev_raw)
        except (TypeError, ValueError):
            continue
        if status != "active":
            continue
        if action != "entered":
            continue
        if not is_quarter_point(dev) or not in_dev_bounds(dev):
            continue
        if abs(dev - want) > 1e-9:
            continue
        if blank(reviewer) or not is_iso_date(str(review_date)):
            continue
        return True, f"approved review-log record (dev={dev:+.2f}, reviewer={reviewer}, {review_date})"
    return False, ("no APPROVED review-log record for "
                   f"{abbrev}/{active_qb} dev={want:+.2f} "
                   "(need Status=active, Review action=entered, matching quarter-point deviation, reviewer+ISO review date)")


# ---- focused QB status evaluation (faithful G/M, self-contained) -----------
def eval_status_counts(wb):
    """Return (ok, uncertain, nonzero, per_row) by faithfully evaluating the
    actual QB VALUES G and M formulas from real D/F/H/J and SETTINGS!B3."""
    qb = wb["QB VALUES"]
    season = season_of(wb)
    ok = unc = nonzero = 0
    per_row = {}
    for r in range(FIRST_ROW, LAST_ROW + 1):
        D = qb.cell(row=r, column=COL["baseline_value"]).value
        F = qb.cell(row=r, column=COL["active_value"]).value
        H = qb.cell(row=r, column=COL["confidence"]).value
        J = qb.cell(row=r, column=COL["reviewed_season"]).value
        if blank(D) or blank(F):
            G = ""
        else:
            G = F - D
            if G != 0:
                nonzero += 1
        status = "UNCERTAIN" if (G == "" or H == "L" or J != season) else "OK"
        if status == "OK":
            ok += 1
        else:
            unc += 1
        per_row[r] = {"D": D, "F": F, "H": H, "J": J, "G": G, "M": status}
    return ok, unc, nonzero, per_row

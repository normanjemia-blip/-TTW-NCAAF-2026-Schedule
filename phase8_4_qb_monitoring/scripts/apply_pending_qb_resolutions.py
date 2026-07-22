#!/usr/bin/env python3
"""Apply APPROVED pending QB resolutions to a copy of the source workbook.

Writes ONLY the allowed QB VALUES input columns (C,D,E,F,H,I,J,K,L) for approved
teams, plus the START HERE!A1 banner and new CHANGELOG rows. Never writes the
calculated QB delta (G) or QB status (M), and refuses if any target cell is a
formula. Validates everything up front and, in apply mode, aborts atomically
(writes no output) if any approved row is invalid.

CLI:
  apply_pending_qb_resolutions.py \
      --source SRC.xlsx --pending pend.csv --review-log log.csv \
      --output OUT.xlsx --version-label v0.7.3 [--apply] [--build-date YYYY-MM-DD]

Default mode is DRY-RUN (no file written). Pass --apply to write the output.
"""
import argparse, os, sys, datetime, shutil
import openpyxl
import pipeline_lib as L


# ---- validation + planning -------------------------------------------------
def validate_and_plan(wb, pending_all, review_log):
    """Return (plans, errors, info).

    plans: list of dicts {abbrev, row, writes:{col:value}, kind, expected_delta,
                          expected_status}
    errors: list of 'ABBR: message' strings (any error blocks apply)
    info: notes (non-blocking)
    """
    errors, info, plans = [], [], []
    rowmap = L.team_row_map(wb)

    # 138 unique teams present
    if len(rowmap) != L.EXPECTED_TEAMS:
        errors.append(f"WORKBOOK: expected {L.EXPECTED_TEAMS} teams in TEAM MAP, found {len(rowmap)}")

    # no team appears more than once in the ENTIRE pending file
    seen = {}
    for r in pending_all:
        ab = str(r.get("abbrev", "")).strip().upper()
        if not ab:
            errors.append("PENDING: a row is missing 'abbrev'")
            continue
        seen[ab] = seen.get(ab, 0) + 1
    for ab, n in seen.items():
        if n > 1:
            errors.append(f"{ab}: appears {n} times in the pending file (must be unique)")

    approved = L.approved_pending(pending_all)
    if not approved:
        info.append("no APPROVED rows in the pending ledger")

    season = L.season_of(wb)
    qb = wb["QB VALUES"]

    for r in approved:
        ab = str(r.get("abbrev", "")).strip().upper()
        errs = []
        # 1) abbrev matches exactly one workbook team
        row = rowmap.get(ab)
        if row is None:
            errs.append(f"unknown abbreviation (not a workbook team)")
            errors += [f"{ab or '?'}: {e}" for e in errs]
            continue
        # 2) confidence H or M
        conf = str(r.get("confidence", "")).strip().upper()
        if conf not in ("H", "M"):
            errs.append(f"confidence must be H or M (got {r.get('confidence')!r})")
        # 3) baseline/active QB populated
        bqb = r.get("baseline_qb"); aqb = r.get("active_qb")
        if L.blank(bqb):
            errs.append("baseline_qb is blank")
        if L.blank(aqb):
            errs.append("active_qb is blank")
        # 4) source, source_date, last_update, notes populated
        for fld in ("source", "source_date", "last_update", "notes"):
            if L.blank(r.get(fld)):
                errs.append(f"{fld} is blank")
        # 5) reviewed season == 2026 (== workbook season)
        try:
            rs = int(r.get("reviewed_season"))
        except (TypeError, ValueError):
            rs = None
        if rs != season:
            errs.append(f"reviewed_season must be {season} (got {r.get('reviewed_season')!r})")
        # 6) ISO dates
        for fld in ("source_date", "last_update"):
            if not L.blank(r.get(fld)) and not L.is_iso_date(str(r.get(fld))):
                errs.append(f"{fld} must be ISO YYYY-MM-DD (got {r.get(fld)!r})")
        # 7) baseline_value must be 0 (deviation-only)
        bv = r.get("baseline_value")
        try:
            bvf = float(bv)
        except (TypeError, ValueError):
            bvf = None
        if bvf != 0:
            errs.append(f"baseline_value must be 0 (got {bv!r})")
        # 8) active_value / zero-init vs baseline-mismatch
        av = r.get("active_value")
        try:
            avf = float(av)
        except (TypeError, ValueError):
            avf = None
        same_qb = (not L.blank(bqb) and not L.blank(aqb)
                   and str(bqb).strip() == str(aqb).strip())
        kind = "zero-init" if same_qb else "baseline-mismatch"
        if avf is None:
            errs.append(f"active_value must be numeric (got {av!r})")
        elif same_qb:
            if avf != 0:
                errs.append(f"zero-init (same QB) requires active_value 0 (got {av!r})")
        else:
            # baseline mismatch -> require approved review-log record
            if not L.is_quarter_point(avf) or not L.in_dev_bounds(avf):
                errs.append(f"deviation {av!r} must be a quarter-point in [{L.DEV_MIN},{L.DEV_MAX}]")
            else:
                ok_rl, why = L.find_approved_deviation(review_log, ab, aqb, avf)
                if not ok_rl:
                    errs.append(why)
        # 9) no formula cell will be overwritten
        if row is not None:
            for fld, col in L.WRITE_FIELDS.items():
                if col in L.FORMULA_COLS:
                    errs.append(f"internal: field {fld} maps to a formula column")
                cur = qb.cell(row=row, column=col).value
                if L.is_formula(cur):
                    errs.append(f"target cell {openpyxl.utils.get_column_letter(col)}{row} "
                                f"({fld}) is a formula - refusing to overwrite")

        if errs:
            errors += [f"{ab}: {e}" for e in errs]
            continue

        # build the write plan (only writable columns)
        writes = {
            L.WRITE_FIELDS["baseline_qb"]: str(bqb),
            L.WRITE_FIELDS["baseline_value"]: 0,
            L.WRITE_FIELDS["active_qb"]: str(aqb),
            L.WRITE_FIELDS["active_value"]: avf,
            L.WRITE_FIELDS["confidence"]: conf,
            L.WRITE_FIELDS["source"]: str(r.get("source")),
            L.WRITE_FIELDS["reviewed_season"]: season,
            L.WRITE_FIELDS["last_update"]: str(r.get("last_update")),
            L.WRITE_FIELDS["notes"]: str(r.get("notes")),
        }
        expected_delta = avf - 0
        expected_status = "OK" if (conf in ("H", "M") and rs == season) else "UNCERTAIN"
        plans.append({
            "abbrev": ab, "row": row, "writes": writes, "kind": kind,
            "active_qb": str(aqb), "active_value": avf,
            "expected_delta": expected_delta, "expected_status": expected_status,
        })

    return plans, errors, info


# ---- application -----------------------------------------------------------
def _last_changelog_row(cl):
    last = 0
    for r in range(1, cl.max_row + 1):
        if any(cl.cell(row=r, column=c).value not in (None, "") for c in range(1, 7)):
            last = r
    return last

def apply_plans(wb, plans, version_label, build_date):
    """Mutate wb in place: QB rows + banner + CHANGELOG. Returns changelog rows added."""
    qb = wb["QB VALUES"]
    # guard again at write time: never touch formula columns / formula cells
    for p in plans:
        for col, val in p["writes"].items():
            assert col in L.WRITABLE_COLS, f"refuse to write non-writable column {col}"
            cell = qb.cell(row=p["row"], column=col)
            assert not L.is_formula(cell.value), f"refuse to overwrite formula at {cell.coordinate}"
            cell.value = val

    # recompute counts on the modified workbook
    ok, unc, nonzero, _ = L.eval_status_counts(wb)

    # banner
    n = len(plans)
    zi = sum(1 for p in plans if p["kind"] == "zero-init")
    dv = n - zi
    dev_note = "no nonzero deviation" if nonzero == 0 else f"{nonzero} approved nonzero deviation(s)"
    banner = (f"TO THE WINDOW — NCAAF POWER RATINGS 2026 ({version_label} QB VALUES CANDIDATE "
              f"— DEVIATION-ONLY; {ok} QB OK / {unc} UNCERTAIN; {n} resolution(s) applied "
              f"({zi} zero-init, {dv} approved-deviation); {dev_note} — NOT AUTHORITATIVE)")
    wb[L.BANNER_CELL[0]][L.BANNER_CELL[1]].value = banner

    # CHANGELOG
    cl = wb["CHANGELOG"]
    start = _last_changelog_row(cl) + 1
    abbrevs = ", ".join(p["abbrev"] for p in plans)
    rows = [(
        version_label, build_date,
        f"Phase 8.4 (CANDIDATE — not authoritative): applied {n} approved QB resolution(s) "
        f"[{abbrevs}] — {zi} zero-init, {dv} approved-deviation. QB status now {ok} OK / "
        f"{unc} UNCERTAIN; {dev_note}. Only QB VALUES input cells for these teams, this banner, "
        f"and these CHANGELOG rows changed; formula count unchanged (123,011).",
        "Phase 8.4 approved batch (pending_qb_resolutions + qb_deviation_review_log)",
    )]
    for p in plans:
        rows.append((
            version_label, build_date,
            f"{p['abbrev']}: {p['kind']} — Active QB '{p['active_qb']}', Active value "
            f"{p['active_value']:+.2f} (delta {p['expected_delta']:+.2f}), status {p['expected_status']}.",
            "approved resolution",
        ))
    for i, (v, d, ch, rn) in enumerate(rows):
        rr = start + i
        cl.cell(row=rr, column=1).value = v
        cl.cell(row=rr, column=2).value = d
        cl.cell(row=rr, column=3).value = ch
        cl.cell(row=rr, column=4).value = rn
    return {"banner": banner, "changelog_rows": list(range(start, start + len(rows))),
            "ok": ok, "uncertain": unc, "nonzero": nonzero}


# ---- CLI -------------------------------------------------------------------
def run(source, pending, review_log, output, version_label, apply_mode, build_date):
    wb = L.load_wb(source)
    pend = L.load_pending(pending)
    rlog = L.load_review_log(review_log) if review_log else []
    plans, errors, info = validate_and_plan(wb, pend, rlog)

    print(f"Source     : {source}")
    print(f"Pending    : {pending}  ({len(pend)} rows, {len(L.approved_pending(pend))} APPROVED)")
    print(f"Review log : {review_log or '(none)'}  ({len(rlog)} records)")
    print(f"Mode       : {'APPLY' if apply_mode else 'DRY-RUN'}")
    print(f"Version    : {version_label}")
    for m in info:
        print(f"  info: {m}")
    print(f"\nPlanned changes: {len(plans)} team(s)")
    for p in plans:
        cells = ", ".join(f"{openpyxl.utils.get_column_letter(c)}{p['row']}" for c in sorted(p["writes"]))
        print(f"  {p['abbrev']:5s} row {p['row']:>3} [{p['kind']:16s}] "
              f"active={p['active_value']:+.2f} delta={p['expected_delta']:+.2f} "
              f"status={p['expected_status']}  -> {cells}")
    if errors:
        print(f"\nVALIDATION ERRORS ({len(errors)}):")
        for e in errors:
            print(f"  FAIL {e}")

    if not apply_mode:
        print("\nDRY-RUN: no output written.")
        return {"ok": not errors, "errors": errors, "plans": plans, "written": False}

    # APPLY mode
    if errors:
        print("\nAPPLY ABORTED: validation errors present; no output written; source untouched.")
        return {"ok": False, "errors": errors, "plans": plans, "written": False}
    if not plans:
        print("\nAPPLY ABORTED: no approved resolutions to apply; no output written.")
        return {"ok": False, "errors": ["no approved resolutions"], "plans": [], "written": False}

    # write to a temp copy first, then move into place (atomic-ish).
    # openpyxl validates by extension, so the temp must end in .xlsx.
    tmp = output + ".building.xlsx"
    shutil.copyfile(source, tmp)
    wb2 = L.load_wb(tmp)
    result = apply_plans(wb2, plans, version_label, build_date)
    wb2.save(tmp)
    os.replace(tmp, output)
    print(f"\nAPPLIED -> {output}")
    print(f"  banner updated; CHANGELOG rows {result['changelog_rows']}")
    print(f"  live counts: {result['ok']} OK / {result['uncertain']} UNCERTAIN / {result['nonzero']} nonzero")
    return {"ok": True, "errors": [], "plans": plans, "written": True, "result": result}


def main(argv=None):
    ap = argparse.ArgumentParser(description="Apply approved pending QB resolutions to a workbook copy.")
    ap.add_argument("--source", required=True)
    ap.add_argument("--pending", required=True)
    ap.add_argument("--review-log", default="")
    ap.add_argument("--output", required=True)
    ap.add_argument("--version-label", required=True)
    ap.add_argument("--apply", action="store_true", help="write output (default is dry-run)")
    ap.add_argument("--build-date", default=datetime.date.today().isoformat())
    a = ap.parse_args(argv)
    res = run(a.source, a.pending, a.review_log, a.output, a.version_label, a.apply, a.build_date)
    return 0 if res["ok"] else 1


if __name__ == "__main__":
    sys.exit(main())

#!/usr/bin/env python3
"""Verify a generated QB candidate against its source workbook.

All checks must pass. Emits a machine-readable JSON report and a readable text
report. Uses the pending ledger to know which QB rows are approved-to-change and
to derive the expected OK/UNCERTAIN counts; uses the review log to validate any
nonzero Active value. Does NOT run the full 123,011-formula engine.

CLI:
  verify_qb_candidate.py --source SRC.xlsx --candidate CAND.xlsx \
      --pending pend.csv --review-log log.csv \
      --report-json out.json --report-md out.md
"""
import argparse, os, sys, json, datetime
import openpyxl
from openpyxl.utils import get_column_letter
import pipeline_lib as L

QB_SHEET, BANNER_SHEET, CL_SHEET = "QB VALUES", "START HERE", "CHANGELOG"


def _last_changelog_row(cl):
    last = 0
    for r in range(1, cl.max_row + 1):
        if any(cl.cell(row=r, column=c).value not in (None, "") for c in range(1, 7)):
            last = r
    return last


def diff_cells(a_ws, b_ws):
    """Yield (row, col, old, new) for every differing cell (ArrayFormula-aware)."""
    mr = max(a_ws.max_row, b_ws.max_row)
    mc = max(a_ws.max_column, b_ws.max_column)
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            va = L.norm(a_ws.cell(row=r, column=c).value)
            vb = L.norm(b_ws.cell(row=r, column=c).value)
            if va != vb:
                yield r, c, va, vb


def verify(source, candidate, pending, review_log):
    a = L.load_wb(source)
    b = L.load_wb(candidate)
    pend = L.load_pending(pending) if pending else []
    approved = L.approved_pending(pend)
    rlog = L.load_review_log(review_log) if review_log else []

    checks = []
    def chk(name, ok, detail=""):
        checks.append({"check": name, "pass": bool(ok), "detail": str(detail)})

    # ---- approved rows + allowed changed-cell set --------------------------
    rowmap = L.team_row_map(b)
    approved_rows = {}
    for r in approved:
        ab = str(r.get("abbrev", "")).strip().upper()
        if ab in rowmap:
            approved_rows[rowmap[ab]] = ab
    src_cl_last = _last_changelog_row(a["CHANGELOG"])

    def allowed(sheet, r, c):
        if sheet == BANNER_SHEET:
            return (r, c) == (1, 1)
        if sheet == CL_SHEET:
            return r > src_cl_last          # only appended rows
        if sheet == QB_SHEET:
            return (r in approved_rows) and (c in L.WRITABLE_COLS)
        return False

    # ---- 1-3 sheets / order / visibility -----------------------------------
    chk("21 sheets", len(b.sheetnames) == L.EXPECTED_SHEETS, len(b.sheetnames))
    chk("identical sheet order", a.sheetnames == b.sheetnames)
    chk("identical hidden/visible states",
        [a[s].sheet_state for s in a.sheetnames] == [b[s].sheet_state for s in b.sheetnames])

    # ---- 4 formula count ----------------------------------------------------
    fa, fb = L.formula_count(a), L.formula_count(b)
    chk("exactly 123,011 formulas (candidate)", fb == L.EXPECTED_FORMULAS, fb)
    chk("formula count unchanged vs source", fa == fb, f"{fa} -> {fb}")

    # ---- 5 formula coords + text identical (no formula overwritten) --------
    formula_changed = []
    for s in a.sheetnames:
        wa, wb_ = a[s], b[s]
        mr = max(wa.max_row, wb_.max_row); mc = max(wa.max_column, wb_.max_column)
        for r in range(1, mr + 1):
            for c in range(1, mc + 1):
                va = wa.cell(row=r, column=c).value
                vb = wb_.cell(row=r, column=c).value
                if (L.is_formula(va) or L.is_formula(vb)) and L.norm(va) != L.norm(vb):
                    formula_changed.append((s, get_column_letter(c) + str(r)))
    chk("identical formula coordinates + text (no formula overwritten/added/removed)",
        not formula_changed, formula_changed[:8])

    # ---- 6-7-12 changed cells only in allowed locations --------------------
    all_changes = {}
    disallowed = []
    for s in a.sheetnames:
        for r, c, old, new in diff_cells(a[s], b[s]):
            all_changes.setdefault(s, []).append((r, c))
            if not allowed(s, r, c):
                disallowed.append((s, get_column_letter(c) + str(r)))
    other_sheets_changed = sorted(set(all_changes) - {QB_SHEET, BANNER_SHEET, CL_SHEET})
    chk("only QB VALUES + banner + CHANGELOG changed",
        not other_sheets_changed, other_sheets_changed)
    chk("no unrelated / disallowed cells changed", not disallowed, disallowed[:8])
    chk("no MARKET LINES / ADJUSTMENTS / TEAM RATINGS / HFA / SETTINGS / IMPORT / SCHEDULE change",
        not other_sheets_changed, other_sheets_changed)
    # every QB VALUES change is in an approved row + writable column
    qb_bad = [get_column_letter(c) + str(r) for (r, c) in all_changes.get(QB_SHEET, [])
              if not (r in approved_rows and c in L.WRITABLE_COLS)]
    chk("QB VALUES changes only in approved rows + input columns", not qb_bad, qb_bad[:8])
    chk("banner change is exactly START HERE!A1",
        all_changes.get(BANNER_SHEET, []) in ([], [(1, 1)]),
        all_changes.get(BANNER_SHEET, []))
    cl_bad = [(r, c) for (r, c) in all_changes.get(CL_SHEET, []) if r <= src_cl_last]
    chk("CHANGELOG changes only in appended rows", not cl_bad, cl_bad[:8])

    # ---- 8 all 138 teams present + unique ----------------------------------
    chk("138 teams present and unique", len(rowmap) == L.EXPECTED_TEAMS, len(rowmap))
    chk("A/B/G/M still formulas on all team rows",
        all(L.is_formula(b[QB_SHEET].cell(row=r, column=c).value)
            for r in range(L.FIRST_ROW, L.LAST_ROW + 1) for c in L.FORMULA_COLS))

    # ---- 9-10 D/F values only blank/0/approved-deviation -------------------
    qb = b[QB_SHEET]
    bad_val, nonzero_ok, nonzero_bad = [], [], []
    for r in range(L.FIRST_ROW, L.LAST_ROW + 1):
        D = qb.cell(row=r, column=L.COL["baseline_value"]).value
        F = qb.cell(row=r, column=L.COL["active_value"]).value
        aqb = qb.cell(row=r, column=L.COL["active_qb"]).value
        # baseline value must be blank or 0
        if not (L.blank(D) or D == 0):
            bad_val.append(f"D{r}={D!r}")
        # active value must be blank, 0, or an approved deviation
        if L.blank(F) or F == 0:
            pass
        else:
            abbr = next((ab for rr, ab in approved_rows.items() if rr == r), None) \
                   or next((k for k, v in rowmap.items() if v == r), "?")
            ok_rl, why = L.find_approved_deviation(rlog, abbr, aqb, F)
            if ok_rl and L.is_quarter_point(F) and L.in_dev_bounds(F):
                nonzero_ok.append(f"{abbr} F{r}={F:+.2f}")
            else:
                nonzero_bad.append(f"{abbr} F{r}={F!r}: {why}")
    chk("Baseline/Active values are only blank, 0, or approved deviation",
        not bad_val and not nonzero_bad, (bad_val + nonzero_bad)[:8])
    chk("every nonzero Active value has an approved review-log record",
        not nonzero_bad, nonzero_bad[:8] or f"{len(nonzero_ok)} approved nonzero value(s)")

    # ---- 11 focused QB status counts ---------------------------------------
    s_ok, s_unc, _, s_rows = L.eval_status_counts(a)
    c_ok, c_unc, c_nz, _ = L.eval_status_counts(b)
    # expected: each approved team that was UNCERTAIN in source becomes OK
    became_ok = sum(1 for row, ab in approved_rows.items() if s_rows[row]["M"] == "UNCERTAIN")
    exp_ok, exp_unc = s_ok + became_ok, s_unc - became_ok
    chk("live OK count matches expected", c_ok == exp_ok, f"candidate={c_ok} expected={exp_ok}")
    chk("live UNCERTAIN count matches expected", c_unc == exp_unc, f"candidate={c_unc} expected={exp_unc}")
    chk("OK + UNCERTAIN == 138", c_ok + c_unc == 138, c_ok + c_unc)

    passed = all(c["pass"] for c in checks)
    report = {
        "verified_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "source": os.path.basename(source),
        "candidate": os.path.basename(candidate),
        "source_sha256": L.sha256(source),
        "candidate_sha256": L.sha256(candidate),
        "approved_rows": {get_column_letter(1) + str(r): ab for r, ab in sorted(approved_rows.items())},
        "counts": {"candidate_ok": c_ok, "candidate_uncertain": c_unc,
                   "candidate_nonzero_delta": c_nz,
                   "expected_ok": exp_ok, "expected_uncertain": exp_unc,
                   "source_ok": s_ok, "source_uncertain": s_unc},
        "changed_sheets": {s: len(v) for s, v in sorted(all_changes.items())},
        "checks": checks,
        "pass": passed,
    }
    return report


def write_reports(report, json_path, md_path):
    if json_path:
        json.dump(report, open(json_path, "w"), indent=1, ensure_ascii=False)
    if md_path:
        lines = [f"# QB Candidate Verification — {'PASS' if report['pass'] else 'FAIL'}", ""]
        lines.append(f"- Source: `{report['source']}`  (`{report['source_sha256'][:16]}…`)")
        lines.append(f"- Candidate: `{report['candidate']}`  (`{report['candidate_sha256'][:16]}…`)")
        c = report["counts"]
        lines.append(f"- Counts: **{c['candidate_ok']} OK / {c['candidate_uncertain']} UNCERTAIN** "
                     f"(expected {c['expected_ok']}/{c['expected_uncertain']}); "
                     f"nonzero deltas: {c['candidate_nonzero_delta']}")
        lines.append(f"- Approved rows: {', '.join(report['approved_rows'].values()) or '(none)'}")
        lines.append(f"- Changed sheets: {report['changed_sheets'] or '(none)'}")
        lines.append("")
        lines.append("| Check | Result | Detail |")
        lines.append("|---|---|---|")
        for ch in report["checks"]:
            lines.append(f"| {ch['check']} | {'PASS' if ch['pass'] else 'FAIL'} | {ch['detail']} |")
        lines.append("")
        lines.append(f"**Overall: {'PASS' if report['pass'] else 'FAIL'}**")
        open(md_path, "w").write("\n".join(lines) + "\n")


def main(argv=None):
    ap = argparse.ArgumentParser(description="Verify a QB candidate vs its source workbook.")
    ap.add_argument("--source", required=True)
    ap.add_argument("--candidate", required=True)
    ap.add_argument("--pending", default="")
    ap.add_argument("--review-log", default="")
    ap.add_argument("--report-json", default="")
    ap.add_argument("--report-md", default="")
    a = ap.parse_args(argv)
    report = verify(a.source, a.candidate, a.pending, a.review_log)
    write_reports(report, a.report_json, a.report_md)
    for ch in report["checks"]:
        print(("PASS " if ch["pass"] else "FAIL ") + ch["check"]
              + (f"  [{ch['detail']}]" if ch["detail"] else ""))
    print("\nOVERALL:", "PASS" if report["pass"] else "FAIL")
    return 0 if report["pass"] else 1


if __name__ == "__main__":
    sys.exit(main())

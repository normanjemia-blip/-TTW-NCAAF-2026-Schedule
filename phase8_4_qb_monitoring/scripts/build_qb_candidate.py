#!/usr/bin/env python3
"""One-command orchestrator: pending ledger -> verified QB candidate.

Steps:
  1. Validate the pending ledger (apply dry-run).
  2. Show the intended changes (dry-run output).
  3. Create the candidate ONLY when >=1 approved resolution exists.
  4. Apply approved changes (via apply_pending_qb_resolutions).
  5. Update banner + CHANGELOG (done by apply).
  6. Run verify_qb_candidate automatically.
  7. Generate changed-cell CSV + JSON.
  8. Refuse to retain the candidate if verification fails (delete it).

Empty / no-approved ledger => NO candidate is created (exit 0, nothing written).
Validation error on an approved row => abort, nothing written (exit 2).
Verification failure => candidate deleted (exit 3).

CLI:
  build_qb_candidate.py --source SRC.xlsx --pending pend.csv \
      --review-log log.csv --output OUT.xlsx --version-label v0.7.3 \
      [--report-dir DIR] [--build-date YYYY-MM-DD]
"""
import argparse, os, sys, csv, json, datetime
from openpyxl.utils import get_column_letter
import pipeline_lib as L
import apply_pending_qb_resolutions as APP
import verify_qb_candidate as VER


def changed_cells_audit(source, candidate, out_dir, version_label):
    a = L.load_wb(source); b = L.load_wb(candidate)
    changes = []
    for s in a.sheetnames:
        for r, c, old, new in VER.diff_cells(a[s], b[s]):
            changes.append({
                "sheet": s, "cell": f"{get_column_letter(c)}{r}",
                "old_value": "" if old is None else (old[1] if isinstance(old, tuple) else str(old)),
                "new_value": "" if new is None else (new[1] if isinstance(new, tuple) else str(new)),
            })
    csv_path = os.path.join(out_dir, f"changed_cells_{version_label}.csv")
    json_path = os.path.join(out_dir, f"changed_cells_{version_label}.json")
    with open(csv_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["sheet", "cell", "old_value", "new_value"])
        w.writeheader(); w.writerows(changes)
    json.dump({"from": os.path.basename(source), "to": os.path.basename(candidate),
               "version": version_label, "total_changed_cells": len(changes),
               "by_sheet": {s: sum(1 for ch in changes if ch["sheet"] == s)
                            for s in sorted({ch["sheet"] for ch in changes})},
               "changes": changes},
              open(json_path, "w"), indent=1, ensure_ascii=False)
    return csv_path, json_path, len(changes)


def run(source, pending, review_log, output, version_label, report_dir, build_date):
    src_sha = L.sha256(source)
    out_dir = report_dir or os.path.dirname(os.path.abspath(output)) or "."
    os.makedirs(out_dir, exist_ok=True)

    print("=" * 66)
    print("STEP 1-2: validate ledger + dry-run")
    print("=" * 66)
    dry = APP.run(source, pending, review_log, output, version_label,
                  apply_mode=False, build_date=build_date)

    if dry["errors"]:
        print("\nRESULT: validation errors -> NO candidate created.")
        _assert_source_unchanged(source, src_sha)
        return {"status": "validation_error", "exit": 2, "errors": dry["errors"], "created": False}

    if not dry["plans"]:
        print("\nSTEP 3: no approved resolutions in the ledger -> NO candidate created "
              "(this is expected for an empty/unapproved ledger).")
        _assert_source_unchanged(source, src_sha)
        return {"status": "empty_ledger", "exit": 0, "errors": [], "created": False}

    print("\n" + "=" * 66)
    print(f"STEP 3-5: apply {len(dry['plans'])} approved resolution(s) -> {output}")
    print("=" * 66)
    app = APP.run(source, pending, review_log, output, version_label,
                  apply_mode=True, build_date=build_date)
    if not app["written"]:
        print("\nRESULT: apply did not write output -> NO candidate.")
        _assert_source_unchanged(source, src_sha)
        return {"status": "apply_failed", "exit": 2, "errors": app["errors"], "created": False}

    print("\n" + "=" * 66)
    print("STEP 6: verify candidate")
    print("=" * 66)
    report = VER.verify(source, output, pending, review_log)
    rj = os.path.join(out_dir, f"verification_report_{version_label}.json")
    rm = os.path.join(out_dir, f"verification_report_{version_label}.md")
    VER.write_reports(report, rj, rm)
    for ch in report["checks"]:
        print(("PASS " if ch["pass"] else "FAIL ") + ch["check"]
              + (f"  [{ch['detail']}]" if ch["detail"] else ""))

    print("\n" + "=" * 66)
    print("STEP 7: changed-cell audit")
    print("=" * 66)
    ccsv, cjson, ncells = changed_cells_audit(source, output, out_dir, version_label)
    print(f"  {ncells} changed cells -> {os.path.basename(ccsv)}, {os.path.basename(cjson)}")

    print("\n" + "=" * 66)
    print("STEP 8: retention decision")
    print("=" * 66)
    _assert_source_unchanged(source, src_sha)
    if not report["pass"]:
        os.remove(output)
        print(f"  VERIFICATION FAILED -> candidate REMOVED ({os.path.basename(output)}).")
        return {"status": "verify_failed", "exit": 3, "created": False,
                "report_json": rj, "report_md": rm}
    print(f"  VERIFICATION PASSED -> candidate RETAINED ({os.path.basename(output)}).")
    print(f"  counts: {report['counts']['candidate_ok']} OK / "
          f"{report['counts']['candidate_uncertain']} UNCERTAIN / "
          f"{report['counts']['candidate_nonzero_delta']} nonzero")
    return {"status": "ok", "exit": 0, "created": True, "output": output,
            "report_json": rj, "report_md": rm, "changed_cells_csv": ccsv,
            "changed_cells_json": cjson, "counts": report["counts"]}


def _assert_source_unchanged(source, expected_sha):
    now = L.sha256(source)
    if now != expected_sha:
        raise SystemExit(f"FATAL: source workbook changed during build "
                         f"({expected_sha[:12]} -> {now[:12]}); aborting.")


def main(argv=None):
    ap = argparse.ArgumentParser(description="Build a verified QB candidate from the pending ledger.")
    ap.add_argument("--source", required=True)
    ap.add_argument("--pending", required=True)
    ap.add_argument("--review-log", default="")
    ap.add_argument("--output", required=True)
    ap.add_argument("--version-label", required=True)
    ap.add_argument("--report-dir", default="")
    ap.add_argument("--build-date", default=datetime.date.today().isoformat())
    a = ap.parse_args(argv)
    res = run(a.source, a.pending, a.review_log, a.output, a.version_label,
              a.report_dir, a.build_date)
    print(f"\nSTATUS: {res['status']}  (created={res['created']})")
    return res["exit"]


if __name__ == "__main__":
    sys.exit(main())

"""Phase 7D.4A verification: v0.7.7 -> v0.7.8 diff, regression battery, inventory, manifest."""
import openpyxl, hashlib, json, csv, datetime, collections, os
from openpyxl.worksheet.formula import ArrayFormula

ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
OLD=f"{ROOT}/workbook_v0.7.7_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.7_CANDIDATE.xlsx"
NEW=f"{ROOT}/workbook_v0.7.8_candidate/TTW_NCAAF_Power_Ratings_2026_v0.7.8_CANDIDATE.xlsx"
AUTH=f"{ROOT}/workbook_v0.6.2_deliverables/TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"
OUT=f"{ROOT}/workbook_v0.7.8_candidate"

def sha(p):
    h=hashlib.sha256()
    with open(p,"rb") as f:
        for b in iter(lambda:f.read(1<<20),b""): h.update(b)
    return h.hexdigest()

def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
def norm(v): return v.text if isinstance(v,ArrayFormula) else v

wo=openpyxl.load_workbook(OLD); wn=openpyxl.load_workbook(NEW)
R=[]
def say(s): print(s); R.append(s)

# ---------- 1. structure ----------
say("="*78); say("1. STRUCTURE")
say(f"  sheets old/new: {len(wo.sheetnames)} / {len(wn.sheetnames)}  identical order: {wo.sheetnames==wn.sheetnames}")
vis_o={s.title:s.sheet_state for s in wo.worksheets}; vis_n={s.title:s.sheet_state for s in wn.worksheets}
say(f"  visibility identical: {vis_o==vis_n}")
fo=sum(1 for s in wo.worksheets for row in s.iter_rows() for c in row if isf(c.value))
fn=sum(1 for s in wn.worksheets for row in s.iter_rows() for c in row if isf(c.value))
say(f"  formula cells: {fo} -> {fn}  (delta {fn-fo})")

# ---------- 2. cell diff ----------
say("="*78); say("2. CELL DIFF v0.7.7 -> v0.7.8")
diffs=[]; formula_diffs=[]
for name in wn.sheetnames:
    so,sn=wo[name],wn[name]
    mr=max(so.max_row,sn.max_row); mc=max(so.max_column,sn.max_column)
    for r in range(1,mr+1):
        for c in range(1,mc+1):
            a=norm(so.cell(row=r,column=c).value); b=norm(sn.cell(row=r,column=c).value)
            if a!=b:
                addr=f"{openpyxl.utils.get_column_letter(c)}{r}"
                diffs.append((name,addr,a,b))
                if isf(so.cell(row=r,column=c).value) or isf(sn.cell(row=r,column=c).value):
                    formula_diffs.append((name,addr))
say(f"  changed cells: {len(diffs)}   FORMULA changes: {len(formula_diffs)}")
by_sheet=collections.Counter(d[0] for d in diffs)
for k,v in by_sheet.items(): say(f"    {k}: {v}")

FIVE={78:"TXST",80:"WSU",87:"UNT",131:"GASO",137:"ODU"}
CONF_LABEL={"TXST":"VERIFIED PAC-12 QB UPDATE","WSU":"VERIFIED PAC-12 QB UPDATE",
            "UNT":"VERIFIED AMERICAN QB UPDATE","GASO":"VERIFIED SUN BELT QB UPDATE",
            "ODU":"VERIFIED SUN BELT QB UPDATE"}
classified=[]
for name,addr,a,b in diffs:
    col=''.join(ch for ch in addr if ch.isalpha()); row=int(''.join(ch for ch in addr if ch.isdigit()))
    if name=="QB VALUES" and row in FIVE:
        ab=FIVE[row]
        lab="NUMERICAL-CELL CONSISTENCY REPAIR" if col in ("D","F") else CONF_LABEL[ab]
    elif name in ("CHANGELOG","START HERE"):
        lab="VERSION OR CHANGELOG"
    else:
        lab="UNRELATED / UNAUTHORIZED / UNKNOWN"
    classified.append((name,addr,lab,a,b))
cnt=collections.Counter(x[2] for x in classified)
say("  classification:")
for k,v in sorted(cnt.items(), key=lambda kv:-kv[1]): say(f"    {k}: {v}")
unknown=[x for x in classified if x[2]=="UNRELATED / UNAUTHORIZED / UNKNOWN"]
say(f"  UNRELATED/UNAUTHORIZED/UNKNOWN: {len(unknown)}")
for x in unknown[:20]: say(f"      !! {x[0]}!{x[1]}  {x[3]!r} -> {x[4]!r}")

# ---------- 3. QB VALUES inventory ----------
say("="*78); say("3. QB VALUES INVENTORY (v0.7.8)")
qb=wn["QB VALUES"]; tm=wn["TEAM MAP"]
inv=[]
for r in range(6,144):
    ab=tm.cell(row=r,column=1).value; team=tm.cell(row=r,column=2).value
    inv.append(dict(row=r,abbrev=ab,team=team,
        baseline_qb=qb.cell(row=r,column=3).value, baseline_value=qb.cell(row=r,column=4).value,
        active_qb=qb.cell(row=r,column=5).value, active_value=qb.cell(row=r,column=6).value,
        confidence=qb.cell(row=r,column=8).value, source=qb.cell(row=r,column=9).value,
        reviewed_for_season=qb.cell(row=r,column=10).value,
        last_update=(qb.cell(row=r,column=11).value.date().isoformat()
                     if isinstance(qb.cell(row=r,column=11).value,datetime.datetime) else qb.cell(row=r,column=11).value),
        note=qb.cell(row=r,column=12).value))
say(f"  rows: {len(inv)}   unique abbrevs: {len(set(x['abbrev'] for x in inv))}   unique teams: {len(set(x['team'] for x in inv))}")
codes=collections.Counter(x["confidence"] for x in inv)
say(f"  codes: {dict(codes)}")
say(f"  invalid codes: {[x['abbrev'] for x in inv if x['confidence'] not in ('H','M','L')]}")
blank=[x for x in inv if x["baseline_value"] is None and x["active_value"] is None]
zero =[x for x in inv if x["baseline_value"]==0 and x["active_value"]==0]
other=[x for x in inv if x not in blank and x not in zero]
say(f"  blank numerical: {len(blank)}   zero numerical: {len(zero)}   other: {len(other)}")
nonzero=[x for x in inv if (x["baseline_value"] not in (None,0)) or (x["active_value"] not in (None,0))]
say(f"  NONZERO QB VALUES: {len(nonzero)}")
lviol=[x["abbrev"] for x in inv if x["confidence"]=="L" and not (x["baseline_value"] is None and x["active_value"] is None)]
say(f"  L-coded teams with non-blank numerical cells: {len(lviol)} {lviol}")

# faithful direct evaluation of G and M
SEASON=wn["SETTINGS"]["B3"].value
def G(x):
    return "" if (x["baseline_value"] in (None,"") or x["active_value"] in (None,"")) else x["active_value"]-x["baseline_value"]
def M(x):
    g=G(x)
    return "UNCERTAIN" if (g=="" or x["confidence"]=="L" or x["reviewed_for_season"]!=SEASON) else "OK"
st=collections.Counter(M(x) for x in inv)
say(f"  SETTINGS!B3 = {SEASON}")
say(f"  QB status: {dict(st)}")
say(f"  nonzero deltas G: {sum(1 for x in inv if G(x) not in ('',0))}")

# ---------- 4. backlog (carried-forward ledger, same definition as 7D.1-7D.4) ----------
say("="*78); say("4. VERIFICATION BACKLOG (carried-forward ledger)")
prev=json.load(open(f"{ROOT}/workbook_v0.7.7_candidate/qb_inventory_v077.json"))
LEDGER={r["abbrev"]:r["verification_status"] for r in prev["records"]}
say(f"  v0.7.7 declared backlog_remaining = {prev['backlog_remaining']} "
    f"({[b['abbrev'] for b in prev['backlog']]})")
for ab in FIVE.values(): LEDGER[ab]="VERIFIED 7D.4A"
for x in inv: x["verification_status"]=LEDGER[x["abbrev"]]
tier1=[x for x in inv if x["confidence"] in ("M","L")]
backlog=[x for x in inv if x["verification_status"].startswith("NOT VERIFIED")]
say(f"  Tier 1 population (M+L): {len(tier1)}")
say(f"  ledger status counts: {dict(collections.Counter(x['verification_status'] for x in inv))}")
say(f"  >>> REMAINING QB VERIFICATION BACKLOG: {len(backlog)} {[x['abbrev'] for x in backlog]}")

say("")
say("  AUDIT-TRAIL GAP (separate finding, NOT the backlog):")
stale=[x for x in inv if x["last_update"]=="2026-07-21"]
st_t1=[x for x in stale if x["confidence"] in ("M","L")]
st_h =[x for x in stale if x["confidence"]=="H"]
say(f"    rows still stamped 2026-07-21: {len(stale)}  (Tier 1 M/L: {len(st_t1)} | H-coded tier 2: {len(st_h)})")
by_status=collections.Counter(x["verification_status"] for x in st_t1)
say(f"    of the {len(st_t1)} Tier 1 rows, ledger says: {dict(by_status)}")
say(f"    -> {len(st_t1)} Tier 1 teams are CREDITED as verified but carry no in-workbook")
say(f"       verification stamp (original 2026-07-21 build date and build note).")
say(f"    Tier 1 unstamped: {[x['abbrev'] for x in st_t1]}")
say("    NOTE: PHASE_7D4_REPORT section 9 attributed all 82 non-backlog stale rows to")
say("    H-coded tier-2 teams. That was wrong: 61 are H-coded and 21 are Tier 1 M/L.")

# ---------- 5. regression battery ----------
say("="*78); say("5. REGRESSION BATTERY")
def t(n,ok,detail=""): say(f"  [{'PASS' if ok else 'FAIL'}] {n}{(' - '+detail) if detail else ''}")
t("1 no formula changes", len(formula_diffs)==0, f"{len(formula_diffs)}")
t("2 formula count unchanged", fo==fn, f"{fo} -> {fn}")
t("3 sheet count/order/visibility unchanged", wo.sheetnames==wn.sheetnames and vis_o==vis_n)
t("4 138 unique teams, no shifted rows", len(inv)==138 and len(set(x['team'] for x in inv))==138)
t("5 no invalid H/M/L codes", all(x["confidence"] in ("H","M","L") for x in inv))
t("6 zero nonzero QB values", len(nonzero)==0)
t("7 zero nonzero QB deltas", sum(1 for x in inv if G(x) not in ("",0))==0)
t("8 every L-coded team blank-gated", len(lviol)==0)
t("9 every L-coded team UNCERTAIN", all(M(x)=="UNCERTAIN" for x in inv if x["confidence"]=="L"))
t("10 TEAM RATINGS untouched", "TEAM RATINGS" not in by_sheet)
t("11 ENGINE untouched (no spread movement)", "ENGINE" not in by_sheet)
t("12 SETTINGS untouched (BET toggle/HFA/season)", "SETTINGS" not in by_sheet)
t("13 ADJUSTMENTS untouched", "ADJUSTMENTS" not in by_sheet)
t("14 MARKET LINES untouched", "MARKET LINES" not in by_sheet)
t("15 PRESEASON untouched (source weights)", "PRESEASON" not in by_sheet)
t("16 IMPORT SCHEDULE untouched", "IMPORT SCHEDULE" not in by_sheet)
t("17 only QB VALUES + CHANGELOG + START HERE in diff",
  set(by_sheet)<= {"QB VALUES","CHANGELOG","START HERE"}, str(sorted(by_sheet)))
t("18 no unrelated/unauthorized/unknown cells", len(unknown)==0)
t("19 QB VALUES diff confined to the five rows",
  all(int(''.join(ch for ch in a if ch.isdigit())) in FIVE for s,a,_,_ in diffs if s=="QB VALUES"))
t("20 QB VALUES formula columns A,B,G,M intact on all 138 rows",
  all(isf(qb.cell(row=r,column=c).value) for r in range(6,144) for c in (1,2,7,13)))
t("21 QB VALUES input columns hold constants only",
  all(not isf(qb.cell(row=r,column=c).value) for r in range(6,144) for c in (3,4,5,6,8,9,10,11,12)))
t("22 SETTINGS!B3 season = 2026", SEASON==2026)
t("23 SETTINGS!B6 HFA = 2.5", wn["SETTINGS"]["B6"].value==2.5)
t("24 SETTINGS!B11 BET toggle = N", wn["SETTINGS"]["B11"].value=="N")
t("25 ledger backlog = 0", len(backlog)==0, f"{len(backlog)}")
t("25b AUDIT-TRAIL GAP (finding, out of 7D.4A scope): Tier 1 rows credited as "
  "verified but never stamped in-workbook", False,
  f"{len(st_t1)} teams: {[x['abbrev'] for x in st_t1]}")
# WSU-specific
wsu=[x for x in inv if x["abbrev"]=="WSU"][0]
t("26 WSU M->L applied, zeros cleared",
  wsu["confidence"]=="L" and wsu["baseline_value"] is None and wsu["active_value"] is None)
t("27 WSU status now UNCERTAIN", M(wsu)=="UNCERTAIN")
gaso=[x for x in inv if x["abbrev"]=="GASO"][0]
t("28 GASO L->M applied, values still blank",
  gaso["confidence"]=="M" and gaso["baseline_value"] is None and gaso["active_value"] is None)
t("29 GASO remains UNCERTAIN (blank-gated, no numerical adjustment created)", M(gaso)=="UNCERTAIN")
for ab in ("TXST","UNT","ODU"):
    x=[y for y in inv if y["abbrev"]==ab][0]
    t(f"30 {ab} M retained with zeros, status OK",
      x["confidence"]=="M" and x["baseline_value"]==0 and x["active_value"]==0 and M(x)=="OK")
t("31 all five dated 2026-08-04",
  all([y for y in inv if y["abbrev"]==ab][0]["last_update"]=="2026-08-04" for ab in FIVE.values()))
# ENGINE status precedence, unchanged structurally
eng=wn["ENGINE"]
t("32 ENGINE!AI status-precedence formula unchanged",
  norm(wo["ENGINE"]["AI7"].value)==norm(eng["AI7"].value))
t("33 ENGINE!M (QB adj) formula unchanged", norm(wo["ENGINE"]["M7"].value)==norm(eng["M7"].value))
t("34 ENGINE!AE (QB status) formula unchanged", norm(wo["ENGINE"]["AE7"].value)==norm(eng["AE7"].value))

# ---------- 6. zero-change verification ----------
say("="*78); say("6. ZERO-CHANGE VERIFICATION / MANIFEST")
h_auth=sha(AUTH); h_old=sha(OLD); h_new=sha(NEW)
EXP_AUTH="bbb17b50fbfb728bea2a23d3d20771935cc61e238313a054473aafe1ca838efd"
EXP_OLD ="3da33d0c10a375c6bd3e43c06f1119b1a6a72cfb49d16abff65ed9c670d02a73"
say(f"  v0.6.2 AUTHORITATIVE {h_auth}  UNCHANGED={h_auth==EXP_AUTH}")
say(f"  v0.7.7 CANDIDATE     {h_old}  UNCHANGED={h_old==EXP_OLD}")
say(f"  v0.7.8 CANDIDATE     {h_new}  (new)")
pm=json.load(open(f"{ROOT}/PROJECT_MANIFEST.json"))
say(f"  PROJECT_MANIFEST.json match for v0.6.2: {h_auth in json.dumps(pm)}")

# ---------- write artifacts ----------
summary=dict(generated="2026-08-04", counts=dict(codes), status=dict(st),
  blank_D=len(blank), zero_D=len(zero), nonzero_qb_values=len(nonzero),
  backlog_remaining=len(backlog), backlog=[dict(abbrev=x["abbrev"],team=x["team"],code=x["confidence"]) for x in backlog],
  audit_trail_gap_tier1=[dict(abbrev=x["abbrev"],team=x["team"],code=x["confidence"],
                              ledger=x["verification_status"]) for x in st_t1],
  audit_trail_gap_h_coded=len(st_h), records=inv)
with open(f"{OUT}/qb_inventory_v078.json","w") as f: json.dump(summary,f,indent=1,default=str)
with open(f"{OUT}/qb_inventory_v078.csv","w",newline="") as f:
    w=csv.DictWriter(f,fieldnames=list(inv[0].keys())); w.writeheader()
    for x in inv: w.writerow(x)
with open(f"{OUT}/diff_v077_to_v078.csv","w",newline="") as f:
    w=csv.writer(f); w.writerow(["sheet","cell","classification","old","new"])
    for s,a,l,o,n in classified: w.writerow([s,a,l,o,n])
with open(f"{OUT}/verification_log_v078.txt","w") as f: f.write("\n".join(R)+"\n")
say(f"\n  artifacts written to {OUT}")

"""Isolated prior-fade table test: feed synthetic F (effective games) =
0..10 into the real TEAM RATINGS!G formula, with SETTINGS!C37:C46 copied
verbatim from the authoritative v0.6.2 file. Confirms every populated
prior-weight value and the F>=9 floor/clamp. Does not alter the table."""
import openpyxl
SRC = "TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"
src = openpyxl.load_workbook(SRC, data_only=False)
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "TEST"
st = wb.create_sheet("SETTINGS")
for r in range(37, 47):
    st.cell(row=r, column=3, value=src["SETTINGS"].cell(row=r, column=3).value)
G = ('=IF(OR($A{r}="",$F{r}=""),"",IF(MIN($F{r},9)>=9,SETTINGS!$C$46,'
     'INDEX(SETTINGS!$C$37:$C$46,INT(MIN($F{r},9))+1)+(MIN($F{r},9)-INT(MIN($F{r},9)))*'
     '(INDEX(SETTINGS!$C$37:$C$46,INT(MIN($F{r},9))+2)-INDEX(SETTINGS!$C$37:$C$46,INT(MIN($F{r},9))+1))))')
for f in range(0, 11):
    r = 6 + f
    ws.cell(row=r, column=1, value=f"T{f}")   # A non-blank
    ws.cell(row=r, column=6, value=f)          # F = effective games
    ws.cell(row=r, column=7, value=G.format(r=r))
wb.save("h62_fade_table.xlsx")
print("Saved h62_fade_table.xlsx")

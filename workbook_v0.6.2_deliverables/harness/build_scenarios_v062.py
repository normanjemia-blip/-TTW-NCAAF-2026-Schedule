"""
Build every Phase 6.2 regression scenario as a compact harness variant
from h62_base.xlsx (which itself carries v0.6.2's byte-exact formulas).
Each variant only sets manual-input cells; all formulas come from the real
workbook. Every synthetic value carries the TEST_TAG marker.
"""
import copy, datetime, openpyxl

BASE = "h62_base.xlsx"
TAG = "TEST ONLY — NOT REAL DATA"

G = {"G1": 8, "G2": 6, "G3": 9, "G4": 10, "G5": 11, "G6": 14, "G7": 15}
GID = {"G1": "401858202", "G2": "401856766", "G3": "401864494",
       "G4": "401864570", "G5": "401864577", "G6": "401856767", "G7": "401858204"}
QROW = {"NCST": 64, "UVA": 70, "JVST": 97, "NDSU": 122, "FSU": 60, "NMSU": 102}

def load():
    return openpyxl.load_workbook(BASE, data_only=False)

def settings(wb, week=2, asof=datetime.date(2026, 9, 4), toggle="N"):
    st = wb["SETTINGS"]
    st["B4"] = week
    st["B5"] = asof
    st["B11"] = toggle

def line(wb, row, gid, fav, spread, total, ldate):
    ml = wb["MARKET LINES"]
    ml.cell(row=row, column=1, value=gid)
    ml.cell(row=row, column=3, value=fav)
    ml.cell(row=row, column=4, value=spread)
    ml.cell(row=row, column=5, value=total)
    ml.cell(row=row, column=6, value=TAG)
    ml.cell(row=row, column=7, value=ldate)
    ml.cell(row=row, column=8, value=TAG)

def resolve_qb(wb, trow):
    qb = wb["QB VALUES"]
    qb.cell(row=trow, column=3, value=f"{TAG}: base QB")
    qb.cell(row=trow, column=4, value=0.0)   # D baseline value
    qb.cell(row=trow, column=5, value=f"{TAG}: active QB")
    qb.cell(row=trow, column=6, value=0.0)   # F active value  -> delta 0 (non-blank)
    qb.cell(row=trow, column=8, value="M")   # H confidence not L
    qb.cell(row=trow, column=9, value=TAG)
    qb.cell(row=trow, column=10, value=2026) # J reviewed for season
    qb.cell(row=trow, column=11, value=datetime.date(2026, 9, 1))
    qb.cell(row=trow, column=12, value=TAG)

def adj(wb, row, typ, target, value, reason, active="Y", expires=None):
    a = wb["ADJUSTMENTS"]
    a.cell(row=row, column=1, value=typ)
    a.cell(row=row, column=2, value=target)
    a.cell(row=row, column=3, value=value)
    a.cell(row=row, column=4, value=reason if reason is not None else None)
    a.cell(row=row, column=5, value=TAG)
    a.cell(row=row, column=6, value=datetime.date(2026, 9, 1))
    if expires is not None:
        a.cell(row=row, column=7, value=expires)
    a.cell(row=row, column=8, value=active)
    a.cell(row=row, column=9, value=TAG)

def completed(wb, srow, apts, hpts):
    s = wb["IMPORT SCHEDULE"]
    s.cell(row=srow, column=10, value=apts)
    s.cell(row=srow, column=11, value=hpts)
    s.cell(row=srow, column=12, value=True)
    s.cell(row=srow, column=14, value=f"{TAG}: synthetic completed game")

def ready_G1(wb):
    """G1 driven to a genuine READY with a qualifying edge (small spread)."""
    settings(wb)
    line(wb, G["G1"], GID["G1"], "UVA", 1.0, 51.5, datetime.date(2026, 9, 3))
    resolve_qb(wb, QROW["NCST"]); resolve_qb(wb, QROW["UVA"])

saved = []
def save(wb, name):
    wb.save(name); saved.append(name)

# ---- 1. DATA INCOMPLETE repair sequence (G1) ----
wb = load(); ready_G1(wb); save(wb, "h62_datainc_control.xlsx")       # -> READY

wb = load(); ready_G1(wb)
wb["IMPORT SCHEDULE"].cell(row=8, column=3).value = None               # blank WEEK
save(wb, "h62_datainc_week_blank.xlsx")                                # -> DATA INCOMPLETE

wb = load(); ready_G1(wb)
wb["IMPORT SCHEDULE"].cell(row=8, column=4).value = None               # blank DATE
save(wb, "h62_datainc_date_blank.xlsx")                                # -> DATA INCOMPLETE

wb = load(); ready_G1(wb)
wb["IMPORT SCHEDULE"].cell(row=8, column=3).value = None
wb["IMPORT SCHEDULE"].cell(row=8, column=3).value = 0                  # blank then restore to real Week 0
save(wb, "h62_datainc_restored.xlsx")                                  # -> READY

# ---- 2. Stacked-priority proofs: DATA INCOMPLETE must never override a higher status ----
# G5 (NDSU transitional) + resolved QBs + line + BLANK WEEK -> TRANSITION UNCERTAIN (not DATA INCOMPLETE)
wb = load(); settings(wb)
line(wb, G["G5"], GID["G5"], "NDSU", 3.0, 45.0, datetime.date(2026, 9, 3))
resolve_qb(wb, QROW["JVST"]); resolve_qb(wb, QROW["NDSU"])
wb["IMPORT SCHEDULE"].cell(row=11, column=3).value = None             # blank week on G5
save(wb, "h62_stack_transitional_over_datainc.xlsx")

# G1 QB-unresolved + line + BLANK WEEK -> QB UNCERTAIN (not DATA INCOMPLETE)
wb = load(); settings(wb)
line(wb, G["G1"], GID["G1"], "UVA", 6.5, 51.5, datetime.date(2026, 9, 3))
wb["IMPORT SCHEDULE"].cell(row=8, column=3).value = None
save(wb, "h62_stack_qbunc_over_datainc.xlsx")

# G3 no line + BLANK WEEK -> PENDING LINE (not DATA INCOMPLETE)
wb = load(); settings(wb)
wb["IMPORT SCHEDULE"].cell(row=9, column=3).value = None
save(wb, "h62_stack_pending_over_datainc.xlsx")

# G7 invalid favorite (BLOCK) + BLANK WEEK -> BLOCKED (not DATA INCOMPLETE)
wb = load(); settings(wb)
line(wb, G["G7"], GID["G7"], "Ohio State", 10.0, 44.0, datetime.date(2026, 9, 3))  # not in this game
wb["IMPORT SCHEDULE"].cell(row=15, column=3).value = None
save(wb, "h62_stack_blocked_over_datainc.xlsx")

# ---- 3. Full single-condition status chain ----
wb = load(); settings(wb)
line(wb, G["G7"], GID["G7"], "Ohio State", 10.0, 44.0, datetime.date(2026, 9, 3))  # BLOCKED
line(wb, G["G6"], GID["G6"], "UCF", 45.0, 60.0, datetime.date(2026, 9, 3))         # FCS + line
line(wb, G["G2"], GID["G2"], "TCU", 3.0, 55.0, datetime.date(2026, 8, 20))         # STALE
line(wb, G["G1"], GID["G1"], "UVA", 6.5, 51.5, datetime.date(2026, 9, 3))          # QB UNCERTAIN (no QB)
line(wb, G["G5"], GID["G5"], "NDSU", 3.0, 45.0, datetime.date(2026, 9, 3))         # transitional
resolve_qb(wb, QROW["JVST"]); resolve_qb(wb, QROW["NDSU"])                         # G5 -> TRANSITION UNCERTAIN
# G3 left with no line -> PENDING LINE
save(wb, "h62_status_chain.xlsx")

# ---- 4. ADJUSTMENTS reconfirm (G4) ----
wb = load(); settings(wb)
line(wb, G["G4"], GID["G4"], "FSU", 24.5, 58.5, datetime.date(2026, 9, 3))
resolve_qb(wb, QROW["NMSU"]); resolve_qb(wb, QROW["FSU"])
adj(wb, 6, "GAME", GID["G4"], 1.5, f"{TAG}: valid adjustment")        # valid -> J=1,K blank
save(wb, "h62_adj_valid.xlsx")

wb = load(); settings(wb)
line(wb, G["G4"], GID["G4"], "FSU", 24.5, 58.5, datetime.date(2026, 9, 3))
adj(wb, 6, "GAME", GID["G4"], 999, f"{TAG}: oversized")               # oversized -> J=0, K LARGE ADJ
save(wb, "h62_adj_oversized.xlsx")

wb = load(); settings(wb)
line(wb, G["G4"], GID["G4"], "FSU", 24.5, 58.5, datetime.date(2026, 9, 3))
adj(wb, 6, "GAME", GID["G4"], "abc", f"{TAG}: nonnumeric")            # nonnumeric -> J=0, K VALUE NOT NUMERIC
save(wb, "h62_adj_nonnumeric.xlsx")

wb = load(); settings(wb)
line(wb, G["G4"], GID["G4"], "FSU", 24.5, 58.5, datetime.date(2026, 9, 3))
adj(wb, 6, "MARGIN OVERRIDE", GID["G4"], 15, f"{TAG}: override exempt")  # override -> J=1,K blank, exempt
save(wb, "h62_adj_override.xlsx")

wb = load(); settings(wb)
line(wb, G["G4"], GID["G4"], "FSU", 24.5, 58.5, datetime.date(2026, 9, 3))
adj(wb, 6, "GAME", GID["G4"], 2.0, None)                              # missing reason -> J=0, K REASON MISSING
save(wb, "h62_adj_noreason.xlsx")

# ---- 5. BET toggle (G1 READY, qualifying edge) ----
wb = load(); ready_G1(wb); settings(wb, toggle="N"); save(wb, "h62_bet_N.xlsx")
wb = load(); ready_G1(wb); settings(wb, toggle="Y"); save(wb, "h62_bet_Y.xlsx")

# ---- 6. Transitional restriction (NDSU): games+edge+line+toggle=Y still cannot BET ----
wb = load(); settings(wb, toggle="Y")
# 5 synthetic completed FBS games for NDSU at harness rows 20-24 (unused in the compact harness)
sched = wb["IMPORT SCHEDULE"]; cl = wb["CLEAN"]; en = wb["ENGINE"]; calc = wb["CALC"]
src = openpyxl.load_workbook("TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx", data_only=False)
from openpyxl.worksheet.formula import ArrayFormula
def fv(c):
    v=c.value; return v.text if isinstance(v,ArrayFormula) else v
opps = [("Air Force Falcons","American Conference"),("UNLV Rebels","Mountain West Conference"),
        ("New Mexico Lobos","Mountain West Conference"),("San José State Spartans","Mountain West Conference"),
        ("UTEP Miners","Conference USA")]
for i,(opp,conf) in enumerate(opps):
    r = 20+i
    sched.cell(row=r, column=1, value=f"9900000{i}")
    sched.cell(row=r, column=2, value=2026)
    sched.cell(row=r, column=3, value=1+i)
    sched.cell(row=r, column=4, value=datetime.date(2026,9,5+i))
    sched.cell(row=r, column=5, value=False)
    sched.cell(row=r, column=6, value="North Dakota State Bison")
    sched.cell(row=r, column=7, value="Missouri Valley Football Conference")
    sched.cell(row=r, column=8, value=opp)
    sched.cell(row=r, column=9, value=conf)
    sched.cell(row=r, column=10, value=17)
    sched.cell(row=r, column=11, value=24)
    sched.cell(row=r, column=12, value=True)
    sched.cell(row=r, column=14, value=f"{TAG}: synthetic NDSU completed game")
    # copy the real per-row formulas for CLEAN(A:AG), ENGINE(A:AL), CALC(K:Z) from the authoritative file
    for c in range(1,34):
        v=fv(src["CLEAN"].cell(row=r,column=c))
        if v is not None: cl.cell(row=r,column=c,value=v)
    for c in range(1,39):
        v=fv(src["ENGINE"].cell(row=r,column=c))
        if v is not None: en.cell(row=r,column=c,value=v)
    for c in range(11,27):
        v=fv(src["CALC"].cell(row=r,column=c))
        if v is not None: calc.cell(row=r,column=c,value=v)
line(wb, G["G5"], GID["G5"], "NDSU", 0.5, 45.0, datetime.date(2026, 9, 3))
resolve_qb(wb, QROW["JVST"]); resolve_qb(wb, QROW["NDSU"])
save(wb, "h62_ndsu_functional.xlsx")

# ---- 7. FCS protection (G6 with a market line) ----
wb = load(); settings(wb)
line(wb, G["G6"], GID["G6"], "UCF", 45.0, 60.0, datetime.date(2026, 9, 3))
save(wb, "h62_fcs_protection.xlsx")

print("Built", len(saved), "scenario files:")
for s in saved: print("  ", s)

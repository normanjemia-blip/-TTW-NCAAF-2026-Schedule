"""
Phase 6.2 build: repair the DATA INCOMPLETE pathway defect on the
AUTHORITATIVE v0.6.1 baseline (the native Google Sheet exported to xlsx as
v0.6.1_authoritative_export.xlsx).

Only three authorized change-sets are applied:
  1. CLEAN!C6:C1005  (week)  - wrap the bare source reference in an
     explicit blank guard so a genuinely blank Week input stays blank.
  2. CLEAN!D6:D1005  (date)  - same repair for Date.
  3. START HERE!A1 banner -> v0.6.2, and one new CHANGELOG row.

Every CLEAN!C/D cell is verified byte-exact against the expected OLD
formula BEFORE being overwritten; a single mismatch aborts the whole run
(assert-before-overwrite). No other cell, style, sheet state, defined
name, or calc setting is touched.
"""
import json
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

SRC = "v0.6.1_authoritative_export.xlsx"
OUT = "TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"

def old_C(r): return f'=IF($A{r}="","",\'IMPORT SCHEDULE\'!C{r})'
def old_D(r): return f'=IF($A{r}="","",\'IMPORT SCHEDULE\'!D{r})'
def new_C(r): return f'=IF($A{r}="","",IF(\'IMPORT SCHEDULE\'!C{r}="","",\'IMPORT SCHEDULE\'!C{r}))'
def new_D(r): return f'=IF($A{r}="","",IF(\'IMPORT SCHEDULE\'!D{r}="","",\'IMPORT SCHEDULE\'!D{r}))'

wb = openpyxl.load_workbook(SRC, data_only=False)
cl = wb["CLEAN"]

# --- capture before-state + assert, then overwrite ---
before = {"CLEAN_C": {}, "CLEAN_D": {}}
for r in range(6, 1006):
    vc = cl.cell(row=r, column=3).value
    vd = cl.cell(row=r, column=4).value
    assert not isinstance(vc, ArrayFormula), f"C{r} unexpected ArrayFormula"
    assert not isinstance(vd, ArrayFormula), f"D{r} unexpected ArrayFormula"
    assert vc == old_C(r), f"C{r} old-formula mismatch: {vc!r}"
    assert vd == old_D(r), f"D{r} old-formula mismatch: {vd!r}"
    before["CLEAN_C"][r] = vc
    before["CLEAN_D"][r] = vd

for r in range(6, 1006):
    cl.cell(row=r, column=3).value = new_C(r)
    cl.cell(row=r, column=4).value = new_D(r)

# --- banner ---
sh = wb["START HERE"]
old_banner = sh["A1"].value
assert old_banner == ("TO THE WINDOW — NCAAF POWER RATINGS 2026 "
                      "(v0.6.1 PHASE 6 CORRECTION + TEST-COMPLETION BUILD — PENDING APPROVAL)"), \
    f"banner mismatch: {old_banner!r}"
new_banner = ("TO THE WINDOW — NCAAF POWER RATINGS 2026 "
              "(v0.6.2 PHASE 6.2 DATA-INCOMPLETE PATHWAY REPAIR — PENDING APPROVAL)")
sh["A1"].value = new_banner

# --- CHANGELOG: append one v0.6.2 row at the first empty row (57) ---
clog = wb["CHANGELOG"]
row = 6
while clog.cell(row=row, column=1).value not in (None, ""):
    row += 1
assert row == 57, f"expected first empty CHANGELOG row 57, got {row}"
clog.cell(row=row, column=1).value = "v0.6.2"
clog.cell(row=row, column=2).value = "2026-07-21"
clog.cell(row=row, column=3).value = (
    "CORRECTION (user-authorized, Phase 6.2): repaired the documented DATA INCOMPLETE "
    "pathway defect. CLEAN!C6:C1005 (week) and CLEAN!D6:D1005 (date) previously fell "
    "through to a bare source reference, so a genuinely blank Week or Date was read as "
    "numeric 0 / epoch date and ENGINE!AI's OR($B6=\"\",$C6=\"\") check could never "
    "fire, making DATA INCOMPLETE unreachable. Both ranges now wrap the source in an "
    "explicit blank guard: =IF($A6=\"\",\"\",IF('IMPORT SCHEDULE'!C6=\"\",\"\",'IMPORT "
    "SCHEDULE'!C6)) (and D-equivalent). A blank Week/Date now propagates as blank and "
    "the game reaches DATA INCOMPLETE; once restored it returns to its correct "
    "lower-priority status. Valid Week values (including Week 0) and Date values pass "
    "through unchanged. Only these 2000 formula cells, this banner, and this CHANGELOG "
    "row changed vs the authoritative v0.6.1."
)
clog.cell(row=row, column=4).value = "v0.6.2 - DATA INCOMPLETE pathway repair (CLEAN!C/D)"

# match the plain-text style of neighbouring changelog cells (General format)
for c in range(1, 5):
    clog.cell(row=row, column=c).number_format = clog.cell(row=56, column=c).number_format

wb.save(OUT)

before["banner_old"] = old_banner
before["banner_new"] = new_banner
json.dump(before, open("before_state_clean_cd.json", "w"), indent=0)
print("Saved", OUT)
print("CLEAN!C/D rows repaired:", 1000, "each (6-1005); asserts all passed")
print("Banner updated; CHANGELOG row", row, "added (v0.6.2)")

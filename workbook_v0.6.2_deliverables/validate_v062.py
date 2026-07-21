"""
Phase 6.2 final validation battery for the authoritative v0.6.2 workbook.
Verifies the repair, structural preservation, clean production state,
target counts, and that the change set vs v0.6.1 is exactly the intended
2005 cells.
"""
import json, openpyxl
from openpyxl.worksheet.formula import ArrayFormula

V062 = "TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"
V061 = "v0.6.1_authoritative_export.xlsx"
V061_WORK = "../workbook/v0.6.1_working.xlsx"   # prior-round working copy (classification source)

fails = []
def chk(label, ok, detail=""):
    print(("PASS " if ok else "FAIL ") + label + (f"  {detail}" if detail else ""))
    if not ok: fails.append(label)

def norm(v):
    if isinstance(v, ArrayFormula): return ("ARR", v.ref, v.text)
    return v

wb = openpyxl.load_workbook(V062, data_only=False)
base = openpyxl.load_workbook(V061, data_only=False)

# ---------- A. Structure preserved ----------
chk("21 sheets", len(wb.sheetnames) == 21, str(len(wb.sheetnames)))
EXPECT_STATE = {'START HERE':'visible','DASHBOARD':'visible','ENGINE':'visible','MARKET LINES':'visible',
 'ADJUSTMENTS':'visible','QB VALUES':'visible','TEAM RATINGS':'visible','DATA QUALITY':'visible',
 'SETTINGS':'visible','IMPORT SCHEDULE':'visible','IMPORT STATS':'visible','PRESEASON':'hidden',
 'FCS TIERS':'hidden','TEAM MAP':'hidden','CLEAN':'hidden','CALC':'hidden','HISTORY':'hidden',
 'BACKTEST':'hidden','AUDIT':'hidden','DICTIONARY':'hidden','CHANGELOG':'hidden'}
chk("sheet hidden/visible states exactly as approved",
    all(wb[s].sheet_state == EXPECT_STATE[s] for s in wb.sheetnames), "")
chk("hidden/visible states unchanged vs v0.6.1",
    all(wb[s].sheet_state == base[s].sheet_state for s in wb.sheetnames))
chk("fullCalcOnLoad = True (recalc on open enabled)", wb.calculation.fullCalcOnLoad == True,
    str(wb.calculation.fullCalcOnLoad))

# ---------- B. The repair itself ----------
cl = wb["CLEAN"]
def newC(r): return f'=IF($A{r}="","",IF(\'IMPORT SCHEDULE\'!C{r}="","",\'IMPORT SCHEDULE\'!C{r}))'
def newD(r): return f'=IF($A{r}="","",IF(\'IMPORT SCHEDULE\'!D{r}="","",\'IMPORT SCHEDULE\'!D{r}))'
badC = [r for r in range(6, 1006) if cl.cell(row=r, column=3).value != newC(r)]
badD = [r for r in range(6, 1006) if cl.cell(row=r, column=4).value != newD(r)]
chk("CLEAN!C6:C1005 all carry the repaired blank-guarded formula", not badC, str(badC[:3]))
chk("CLEAN!D6:D1005 all carry the repaired blank-guarded formula", not badD, str(badD[:3]))

# ---------- C. Change set vs v0.6.1 is EXACTLY the intended 2005 cells ----------
diffs = {}
for s in base.sheetnames:
    wsa, wsb = base[s], wb[s]
    mr = max(wsa.max_row, wsb.max_row); mc = max(wsa.max_column, wsb.max_column)
    n = 0
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            if norm(wsa.cell(row=r, column=c).value) != norm(wsb.cell(row=r, column=c).value):
                n += 1
    if n: diffs[s] = n
chk("only CLEAN + START HERE + CHANGELOG differ vs v0.6.1", set(diffs) == {"CLEAN", "START HERE", "CHANGELOG"}, str(diffs))
chk("CLEAN diff == 2000 (C+D)", diffs.get("CLEAN") == 2000, str(diffs.get("CLEAN")))
chk("START HERE diff == 1 (banner)", diffs.get("START HERE") == 1, str(diffs.get("START HERE")))
chk("CHANGELOG diff == 4 (one new v0.6.2 row)", diffs.get("CHANGELOG") == 4, str(diffs.get("CHANGELOG")))

# ---------- D. ADJUSTMENTS J/K unchanged (v0.6.1 safety logic intact) ----------
adj = wb["ADJUSTMENTS"]; adjb = base["ADJUSTMENTS"]
jk_changed = [(r, c) for r in range(6, 256) for c in (10, 11)
              if norm(adj.cell(row=r, column=c).value) != norm(adjb.cell(row=r, column=c).value)]
chk("ADJUSTMENTS!J/K 250 rows byte-identical to v0.6.1 (no unrelated adj change)", not jk_changed, str(jk_changed[:3]))
chk("ADJUSTMENTS!J6 requires K blank (safety logic present)",
    '$K6=""' in adj.cell(row=6, column=10).value, "")
chk("ADJUSTMENTS!K6 has ISNUMBER guard (no raw #VALUE!)",
    'IF(ISNUMBER($C6)' in adj.cell(row=6, column=11).value, "")

# ---------- E. Clean production state ----------
qb = wb["QB VALUES"]
nb_qb = [(r, c) for r in range(6, 144) for c in [3,4,5,6,8,9,10,11,12] if qb.cell(row=r, column=c).value is not None]
chk("138 QB rows fully blank (all UNCERTAIN)", not nb_qb, str(nb_qb[:5]))
chk("138 QB VALUES rows present", sum(1 for r in range(6,144) if qb.cell(row=r,column=1).value not in (None,"")) == 138, "")
ml = wb["MARKET LINES"]
nb_ml = [(r, c) for r in range(6, 1006) for c in [1,3,4,5,6,7,8] if ml.cell(row=r, column=c).value is not None]
chk("0 market lines", not nb_ml, str(nb_ml[:5]))
nb_adj = [(r, c) for r in range(6, 256) for c in range(1, 10) if adj.cell(row=r, column=c).value is not None]
chk("0 manual adjustments", not nb_adj, str(nb_adj[:5]))
ist = wb["IMPORT STATS"]
nb_ist = [(r, c) for r in range(6, 206) for c in range(1, 10) if ist.cell(row=r, column=c).value is not None]
chk("0 imported statistics", not nb_ist, str(nb_ist[:5]))
st = wb["SETTINGS"]
chk("BET toggle = N (approved default)", st["B11"].value == "N", str(st["B11"].value))
chk("SETTINGS!B4 (week) blank", st["B4"].value is None, "")
chk("SETTINGS!B5 (as-of) blank", st["B5"].value is None, "")
hist = wb["HISTORY"]
nb_hist = [(r, c) for r in range(6, 3006) for c in range(1, 5) if hist.cell(row=r, column=c).value is not None]
chk("0 HISTORY snapshots", not nb_hist, str(nb_hist[:5]))

# ---------- F. Counts: 888 / 761 / 127, all unique, classification transfers ----------
sched = wb["IMPORT SCHEDULE"]
ids = [sched.cell(row=r, column=1).value for r in range(6, 894) if sched.cell(row=r, column=1).value not in (None, "")]
chk("888 scheduled games", len(ids) == 888, str(len(ids)))
chk("all 888 GameIDs unique", len(set(ids)) == 888, str(len(set(ids))))
chk("rows 894-1005 blank (no phantom games)",
    all(sched.cell(row=r, column=1).value in (None, "") for r in range(894, 1006)), "")

# classification transfer: identical schedule inputs + identical CLEAN!O/Q/M => identical split.
# The classification JSON is keyed by CLEAN/schedule ROW number (6..893), one per populated game.
cls = json.load(open("fcs_classification_888.json"))
populated_rows = set(str(r) for r in range(6, 894) if sched.cell(row=r, column=1).value not in (None, ""))
chk("classification covers exactly the 888 populated schedule rows",
    populated_rows == set(cls.keys()), f"{len(populated_rows ^ set(cls.keys()))} mismatched")
fbs_fbs = sum(1 for o, q in cls.values() if o == "FBS" and q == "FBS")
fcs_inv = sum(1 for o, q in cls.values() if o == "FCS" or q == "FCS")
blocked = sum(1 for o, q in cls.values() if o == "BLOCK" or q == "BLOCK")
chk("761 FBS-vs-FBS games", fbs_fbs == 761, str(fbs_fbs))
chk("127 FCS-involved / FCS — NO PLAY games", fcs_inv == 127, str(fcs_inv))
chk("0 BLOCK-classified games in pristine schedule", blocked == 0, str(blocked))
# prove the classification inputs are byte-identical to the v0.6.1 source it was computed from
try:
    work = openpyxl.load_workbook(V061_WORK, data_only=False)
    sched_id = all(sched.cell(row=r, column=c).value == work["IMPORT SCHEDULE"].cell(row=r, column=c).value
                   for r in range(6, 894) for c in range(1, 15))
    tm_id = all(norm(wb["TEAM MAP"].cell(row=r, column=c).value) == norm(work["TEAM MAP"].cell(row=r, column=c).value)
                for r in range(6, 606) for c in range(1, 13))
    chk("IMPORT SCHEDULE 888 rows byte-identical to classification source (v0.6.1 working)", sched_id, "")
    chk("TEAM MAP byte-identical to classification source (v0.6.1 working)", tm_id, "")
except FileNotFoundError:
    print("NOTE  v0.6.1 working copy not present for input-identity cross-check")

# ---------- G. NDSU / Sacramento State transitional intact ----------
tm = wb["TEAM MAP"]
chk("SAC still FBS-RECLASSIFYING", tm.cell(row=114, column=4).value == "FBS-RECLASSIFYING", "")
chk("NDSU still FBS-RECLASSIFYING", tm.cell(row=122, column=4).value == "FBS-RECLASSIFYING", "")
tr = wb["TEAM RATINGS"]
chk("TEAM RATINGS!X114 (SAC review-cleared) blank", tr.cell(row=114, column=24).value is None, "")
chk("TEAM RATINGS!X122 (NDSU review-cleared) blank", tr.cell(row=122, column=24).value is None, "")

# ---------- H. No stored formula-error cells introduced ----------
errs = ("#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!")
lit = []
for s in wb.sheetnames:
    if s == "CHANGELOG": continue
    for row in wb[s].iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, ArrayFormula): v = v.text
            if isinstance(v, str) and not v.startswith("=") and any(e in v for e in errs):
                lit.append((s, cell.coordinate))
chk("0 stored formula-error cells (excl. CHANGELOG prose)", not lit, str(lit[:5]))

# ---------- I. Banner / CHANGELOG ----------
chk("banner shows v0.6.2",
    wb["START HERE"]["A1"].value == "TO THE WINDOW — NCAAF POWER RATINGS 2026 "
    "(v0.6.2 PHASE 6.2 DATA-INCOMPLETE PATHWAY REPAIR — PENDING APPROVAL)", "")
clog = wb["CHANGELOG"]
v062_rows = [r for r in range(6, 60) if clog.cell(row=r, column=1).value == "v0.6.2"]
chk("exactly one v0.6.2 CHANGELOG row", len(v062_rows) == 1, str(v062_rows))
chk("prior CHANGELOG history unchanged (v0.6.1 still 6 rows)",
    sum(1 for r in range(6, 60) if clog.cell(row=r, column=1).value == "v0.6.1") == 6, "")

print()
if fails:
    print(f"*** {len(fails)} FAILURES ***")
    for f_ in fails: print("  ", f_)
    raise SystemExit(1)
print("ALL VALIDATION CHECKS PASSED")

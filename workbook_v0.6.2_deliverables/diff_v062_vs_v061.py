"""
Cell/formula-level diff: v0.6.2 vs the authoritative v0.6.1 baseline.
Proves the change set is limited to CLEAN!C6:C1005, CLEAN!D6:D1005, the
START HERE banner, and one CHANGELOG row - and that sheet count, hidden/
visible states, fullCalcOnLoad, and every unchanged cell's number_format
are preserved.
"""
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

A = "v0.6.1_authoritative_export.xlsx"
B = "TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"

def norm(v):
    if isinstance(v, ArrayFormula):
        return ("ARR", v.ref, v.text)
    return v

wa = openpyxl.load_workbook(A, data_only=False)
wb = openpyxl.load_workbook(B, data_only=False)

fails = []
def chk(label, ok, detail=""):
    print(("PASS " if ok else "FAIL ") + label + (f"  {detail}" if detail else ""))
    if not ok: fails.append(label)

chk("sheet count == 21 both", len(wa.sheetnames) == 21 and len(wb.sheetnames) == 21,
    f"{len(wa.sheetnames)}/{len(wb.sheetnames)}")
chk("sheet names identical & in order", wa.sheetnames == wb.sheetnames)
chk("sheet states (hidden/visible) unchanged",
    all(wa[s].sheet_state == wb[s].sheet_state for s in wa.sheetnames))
chk("fullCalcOnLoad preserved (recalc on open)",
    wa.calculation.fullCalcOnLoad == wb.calculation.fullCalcOnLoad == True,
    f"{wa.calculation.fullCalcOnLoad}/{wb.calculation.fullCalcOnLoad}")

val_diffs = []
fmt_diffs = []
for s in wa.sheetnames:
    wsa, wsb = wa[s], wb[s]
    mr = max(wsa.max_row, wsb.max_row); mc = max(wsa.max_column, wsb.max_column)
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            ca, cb = wsa.cell(row=r, column=c), wsb.cell(row=r, column=c)
            if norm(ca.value) != norm(cb.value):
                val_diffs.append((s, ca.coordinate, norm(ca.value), norm(cb.value)))
            if ca.number_format != cb.number_format:
                fmt_diffs.append((s, ca.coordinate, ca.number_format, cb.number_format))

by_sheet = {}
for s, coord, _, _ in val_diffs:
    by_sheet[s] = by_sheet.get(s, 0) + 1
print("\nValue-diff counts by sheet:", by_sheet)

clean_c = [d for d in val_diffs if d[0] == "CLEAN" and d[1][0] == "C"]
clean_d = [d for d in val_diffs if d[0] == "CLEAN" and d[1][0] == "D"]
sh_diffs = [d for d in val_diffs if d[0] == "START HERE"]
cl_diffs = [d for d in val_diffs if d[0] == "CHANGELOG"]

chk("only CLEAN, START HERE, CHANGELOG changed", set(by_sheet) == {"CLEAN", "START HERE", "CHANGELOG"}, str(set(by_sheet)))
chk("CLEAN!C changed cells == 1000 (C6:C1005)", len(clean_c) == 1000, str(len(clean_c)))
chk("CLEAN!D changed cells == 1000 (D6:D1005)", len(clean_d) == 1000, str(len(clean_d)))
chk("CLEAN changed cells are EXACTLY C+D (no other CLEAN column)", by_sheet.get("CLEAN") == 2000, str(by_sheet.get("CLEAN")))
chk("all CLEAN!C changed rows are 6..1005", sorted(int(d[1][1:]) for d in clean_c) == list(range(6, 1006)))
chk("all CLEAN!D changed rows are 6..1005", sorted(int(d[1][1:]) for d in clean_d) == list(range(6, 1006)))
chk("START HERE change is exactly A1 (banner)", len(sh_diffs) == 1 and sh_diffs[0][1] == "A1", str([d[1] for d in sh_diffs]))
chk("CHANGELOG changes are exactly row 57 A:D", sorted(d[1] for d in cl_diffs) == ["A57", "B57", "C57", "D57"], str(sorted(d[1] for d in cl_diffs)))
chk("total value diffs == 2005", len(val_diffs) == 2005, str(len(val_diffs)))
chk("number_format diffs == 0 (no formatting changed anywhere)", len(fmt_diffs) == 0, str(fmt_diffs[:5]))

# Verify the NEW CLEAN formulas are exactly the intended repair
def new_C(r): return f'=IF($A{r}="","",IF(\'IMPORT SCHEDULE\'!C{r}="","",\'IMPORT SCHEDULE\'!C{r}))'
def new_D(r): return f'=IF($A{r}="","",IF(\'IMPORT SCHEDULE\'!D{r}="","",\'IMPORT SCHEDULE\'!D{r}))'
clB = wb["CLEAN"]
badC = [r for r in range(6, 1006) if clB.cell(row=r, column=3).value != new_C(r)]
badD = [r for r in range(6, 1006) if clB.cell(row=r, column=4).value != new_D(r)]
chk("every CLEAN!C6:C1005 == intended repaired formula", not badC, str(badC[:3]))
chk("every CLEAN!D6:D1005 == intended repaired formula", not badD, str(badD[:3]))

# Write the machine-readable diff
with open("v0.6.2_vs_v0.6.1_cell_diff.tsv", "w") as f:
    f.write("sheet\tcell\tv0.6.1_value\tv0.6.2_value\n")
    for s, coord, va, vb in val_diffs:
        f.write(f"{s}\t{coord}\t{va!r}\t{vb!r}\n")

print("\n" + ("*** %d FAILURES ***" % len(fails) if fails else "ALL DIFF CHECKS PASSED"))
if fails:
    for f_ in fails: print("  ", f_)
    raise SystemExit(1)

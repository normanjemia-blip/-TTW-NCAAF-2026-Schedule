"""
Zero-difference cleanup proof: cleaned test-copy vs pristine authoritative
v0.6.2. Machine-readable output to cleanup_zero_diff_proof_v062.tsv.
"""
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

A = "TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"
B = "v0.6.2_test_cleaned.xlsx"

def norm(v):
    if isinstance(v, ArrayFormula): return ("ARR", v.ref, v.text)
    return v

wa = openpyxl.load_workbook(A, data_only=False)
wb = openpyxl.load_workbook(B, data_only=False)

diffs = []
fmt = 0
for s in wa.sheetnames:
    wsa, wsb = wa[s], wb[s]
    mr = max(wsa.max_row, wsb.max_row); mc = max(wsa.max_column, wsb.max_column)
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            ca, cb = wsa.cell(row=r, column=c), wsb.cell(row=r, column=c)
            if norm(ca.value) != norm(cb.value):
                diffs.append((s, ca.coordinate, norm(ca.value), norm(cb.value)))
            if ca.number_format != cb.number_format:
                fmt += 1

TAG = "TEST ONLY — NOT REAL DATA"
tags = sum(1 for s in wb.sheetnames for row in wb[s].iter_rows() for cell in row
           if isinstance(cell.value, str) and TAG in cell.value)

# residual test-data zone checks
ml = wb["MARKET LINES"]; adjv = wb["ADJUSTMENTS"]; ist = wb["IMPORT STATS"]
test_lines = sum(1 for r in range(6, 1006) if ml.cell(row=r, column=1).value not in (None, ""))
test_adj = sum(1 for r in range(6, 256) if adjv.cell(row=r, column=1).value not in (None, ""))
test_stats = sum(1 for r in range(6, 206) if ist.cell(row=r, column=1).value not in (None, ""))

with open("cleanup_zero_diff_proof_v062.tsv", "w") as f:
    f.write("sheet\tcell\tpristine_value\tcleaned_value\n")
    for s, coord, va, vb in diffs:
        f.write(f"{s}\t{coord}\t{va!r}\t{vb!r}\n")

print("cell-level differences (cleaned vs pristine v0.6.2):", len(diffs))
if diffs:
    for d in diffs[:20]: print("  DIFF", d)
print("number_format differences:", fmt)
print("remaining TEST_TAG cells:", tags)
print("residual market lines:", test_lines)
print("residual adjustments:", test_adj)
print("residual import stats:", test_stats)
ok = (len(diffs) == 0 and fmt == 0 and tags == 0 and test_lines == 0 and test_adj == 0 and test_stats == 0)
print("\n" + ("ZERO-DIFFERENCE CLEANUP PROVEN" if ok else "*** CLEANUP NOT CLEAN ***"))
if not ok: raise SystemExit(1)

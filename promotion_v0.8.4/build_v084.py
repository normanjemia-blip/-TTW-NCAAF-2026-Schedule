#!/usr/bin/env python3
"""Build v0.8.4 from the frozen v0.8.3 workbook.

TWO COMPONENTS, kept deliberately separate:

  A. RESYNC (5 rows) - the repository artifact had drifted behind the live
     production master by five QB records (UNC, STAN, MOST, SJSU promoted;
     TTU raised to H). Every value here is TRANSCRIBED VERBATIM from the live
     sheet read on 2026-08-21. Nothing is authored.

  B. CORRECTIONS (2 rows) - the approved Northern Illinois and Tulane packet.
     NIU confidence M -> L. Tulane candidate list corrected. Colorado State is
     deliberately NOT touched. Texas Tech's medical gate is preserved exactly:
     H confidence, blank D/F, status stays UNCERTAIN.

No nonzero QB value anywhere. G and M stay formulas. J stays 2026.

Run:  python3 promotion_v0.8.4/build_v084.py
"""
import datetime, hashlib, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "promotion_v0.8.3",
                   "TTW_College_Football_Power_Ratings_v0.8.3_AUTHORITATIVE.xlsx")
OUT = os.path.join(ROOT, "promotion_v0.8.4",
                   "TTW_College_Football_Power_Ratings_v0.8.4_AUTHORITATIVE.xlsx")
FROZEN_V083_SHA = "ff55782586ef1adb662eba59710e824dc382769a24579e48917b101fbcdd96b8"

D = datetime.datetime
DATE_0819 = D(2026, 8, 19)
DATE_0821 = D(2026, 8, 21)

# ---------------------------------------------------------------- component A
# Transcribed verbatim from the live production master, 2026-08-21.
RESYNC = {
 "UNC": dict(row=65, C="Billy Edwards Jr.", Dv=0, E="Billy Edwards Jr.", Fv=0, H="M",
   I=("ESPN (Pete Thamel), 2026-08-18: https://www.espn.com/college-football/story/_/"
      "id/49651014/sources-billy-edwards-jr-start-qb-north-carolina ; corroborated by "
      "Inside Carolina/On3: https://www.on3.com/teams/north-carolina-tar-heels/news/"
      "billy-edwards-jr-named-unc-starting-quarterback/"), K=DATE_0819,
   L=("VERIFIED 2026-08-19: ESPN's Pete Thamel reported on 2026-08-18 that North Carolina "
      "named veteran transfer Billy Edwards Jr. the starter for the 2026-08-29 opener vs TCU "
      "in Dublin after a close competition with Miles O'Neill. Inside Carolina/On3 "
      "independently confirmed Edwards had been informed of the decision. No public "
      "GoHeels/UNC Football announcement was located at review time, so confidence is M "
      "rather than H. Recent local reporting indicates Edwards is healthy, with no current "
      "availability restriction reported. Baseline and active values set to 0/0 under the "
      "deviation-only convention; no nonzero QB adjustment or model change. RECHECK: official "
      "UNC game notes, depth chart, or direct team announcement.")),
 "STAN": dict(row=68, C="Davis Warren", Dv=0, E="Davis Warren", Fv=0, H="H",
   I=("Press Democrat / ESPN, 2026-08-18 (https://www.pressdemocrat.com/2026/08/18/"
      "stanford-makes-it-official-qb-davis-warren-will-start-season-opener-against-hawaii-on-aug-29/"
      " ; https://www.espn.com/college-football/story/_/id/49654443/"
      "stanford-names-ex-michigan-qb-davis-warren-starter)"), K=DATE_0821,
   L=("2026-08-21 FINAL QB VERIFICATION: Stanford formally named Davis Warren the starter for "
      "the Aug. 29 opener against Hawaii; head coach Tavita Pritchard publicly confirmed the "
      "decision. Baseline and active values 0/0 under the deviation-only convention; no "
      "numerical adjustment proposed. RECHECK: Week 0 official depth chart or game notes.")),
 "MOST": dict(row=101, C="Skyler Locklear", Dv=0, E="Skyler Locklear", Fv=0, H="H",
   I=("Missouri State Football official / ESPN Pete Thamel, 2026-08-19 "
      "(https://x.com/MoStateFootball/status/2090208075986698243 ; "
      "https://x.com/PeteThamel/status/2090185825598226471)"), K=DATE_0821,
   L=("2026-08-21 FINAL QB VERIFICATION: Missouri State officially named veteran UTEP transfer "
      "Skyler Locklear its starting quarterback; Pete Thamel reported that Locklear won the job "
      "over Henry Belin IV. Baseline and active values 0/0 under the deviation-only convention; "
      "no numerical adjustment proposed. RECHECK: Official Week 1 depth chart or game notes.")),
 "SJSU": dict(row=124, C="Luke Weaver", Dv=0, E="Luke Weaver", Fv=0, H="H",
   I=("San José State Football official / ESPN Pete Thamel, 2026-08-19 "
      "(https://x.com/SanJoseStateFB/status/2090177593983328666 ; "
      "https://x.com/PeteThamel/status/2090096585140994125)"), K=DATE_0821,
   L=("2026-08-21 FINAL QB VERIFICATION: San José State officially confirmed Luke Weaver as its "
      "starting quarterback after he won the fall-camp competition. Baseline and active values "
      "0/0 under the deviation-only convention; no numerical adjustment proposed. RECHECK: "
      "Week 0 official depth chart or game notes.")),
 # TTU: identity only. D/F stay BLANK so the medical gate holds.
 "TTU": dict(row=52, C="Will Hammond", Dv=None, E="Will Hammond", Fv=None, H="H",
   I=("Texas Tech Football official, 2026-07-24 / ESPN, 2026-07-07 "
      "(https://x.com/TexasTechFB/status/2080671454715154707 ; "
      "https://www.espn.com/college-football/story/_/id/49301606/"
      "texas-tech-hammond-acl-week-1-starter-cleared)"), K=DATE_0821,
   L=("2026-08-21 CORRECTION — IDENTITY CONFIRMED, MEDICALLY GATED: Texas Tech publicly "
      "identified Will Hammond as QB1, and head coach Joey McGuire confirmed the staff is moving "
      "forward with Hammond as its starting quarterback. Identity confidence H. Week 1 "
      "availability remains conditional on final team medical clearance for all football-related "
      "activity; that clearance was not independently verified as of this review. Baseline and "
      "active values remain blank pending availability clearance, so QB status remains UNCERTAIN "
      "and no numerical adjustment is applied. RECHECK: Texas Tech team medical release or "
      "official Week 1 depth chart/game notes.")),
}

# ---------------------------------------------------------------- component B
CORRECTIONS = {
 "NIU": dict(row=123, E="Open (Davidson / Macon / Hamric / Dickens)", H="L",
   I=("HERO Sports NIU quarterback report (https://herosports.com/"
      "fbs-prolific-fcs-quarterback-taron-tyger-dickens-win-starting-job-northern-illinois-rcrc/)"),
   K=DATE_0821,
   L=("2026-08-21 REVERSES THE 2026-08-03 UPGRADE. That review raised L to M on the reasoning "
      "that the prior candidates 'Davidson / Macon / Hamric' did not appear in current NIU "
      "coverage and that Taron Dickens was widely expected to start. Current reporting "
      "contradicts that premise directly: coaches have NOT named a starter among returners Brady "
      "Davidson and Jalen Macon and transfers Ean Hamric and Taron Dickens — all four are named "
      "in current coverage. 'Widely expected to start' is a projection, not a naming, so M is not "
      "supported. M downgraded to L and the four-way candidate list restored. Also corrects a "
      "citation defect: the prior source was a MOUNTAIN WEST conference preview cited for a MAC "
      "team. Context: HC Thomas Hammock departed for the NFL; Rob Harley is interim head coach. "
      "Numerical values remain blank; status stays UNCERTAIN either way. RECHECK: Week 1 depth "
      "chart, opener 2026-09-05 at Iowa.")),
 "TULN": dict(row=91, E="Open (Semonza / Chriss-Gremillion / Johnson / Bruno)", H=None,
   I=("FOX 8 New Orleans, 2026-07-24 (https://www.fox8live.com/2026/07/24/"
      "four-qbs-battling-starting-spot-tulane/)"), K=DATE_0821,
   L=("2026-08-21 CANDIDATE RECORD CORRECTED: the active field named Kadin Semonza alone, which "
      "overstated a four-way competition the record's own note already described. Current "
      "reporting confirms four quarterbacks taking fall-camp reps — Semonza, Zeon "
      "Chriss-Gremillion, Trace Johnson and Dagan Bruno — with head coach Will Hall yet to "
      "select one, and reporting that the job defaults to Chriss-Gremillion if no one separates. "
      "Candidate list corrected to match. Confidence L is UNCHANGED and correct. Numerical values "
      "remain blank; status stays UNCERTAIN. RECHECK: Will Hall naming or Week 1 depth chart.")),
}

COLS = dict(C=3, D=4, E=5, F=6, H=8, I=9, K=11, L=12)


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def ftext(v):
    return v.text if isinstance(v, ArrayFormula) else v


def main():
    got = sha256(SRC)
    assert got == FROZEN_V083_SHA, f"v0.8.3 is not the expected artifact: {got}"
    print(f"source v0.8.3 SHA-256 verified: {got}")

    wb = openpyxl.load_workbook(SRC)
    tm, qb, sh = wb["TEAM MAP"], wb["QB VALUES"], wb["START HERE"]

    # locate every affected team by abbreviation
    idx = {}
    for ab in list(RESYNC) + list(CORRECTIONS):
        rows = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == ab]
        assert len(rows) == 1, f"{ab} must resolve to exactly one row, got {rows}"
        idx[ab] = rows[0]
    for ab, spec in {**RESYNC, **CORRECTIONS}.items():
        assert idx[ab] == spec["row"], f"{ab} expected row {spec['row']}, got {idx[ab]}"
    print("all affected teams located by abbreviation; rows match")

    # Colorado State must NOT be touched - capture it to prove that afterwards
    csu_row = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == "CSU"][0]
    csu_before = [qb.cell(row=csu_row, column=c).value for c in range(1, 14)]

    def guard(r):
        for name, col in (("A", 1), ("B", 2), ("G", 7), ("M", 13)):
            assert isf(qb.cell(row=r, column=col).value), f"{name}{r} must be a formula"
        assert qb.cell(row=r, column=10).value == 2026, f"J{r} must be 2026"

    # ---- component A: resync ------------------------------------------------
    for ab, s in RESYNC.items():
        r = s["row"]; guard(r)
        qb.cell(row=r, column=COLS["C"]).value = s["C"]
        qb.cell(row=r, column=COLS["D"]).value = s["Dv"]
        qb.cell(row=r, column=COLS["E"]).value = s["E"]
        qb.cell(row=r, column=COLS["F"]).value = s["Fv"]
        qb.cell(row=r, column=COLS["H"]).value = s["H"]
        qb.cell(row=r, column=COLS["I"]).value = s["I"]
        qb.cell(row=r, column=COLS["K"]).value = s["K"]
        qb.cell(row=r, column=COLS["L"]).value = s["L"]
    print(f"component A: {len(RESYNC)} rows resynced from the live master")

    # ---- component B: corrections ------------------------------------------
    for ab, s in CORRECTIONS.items():
        r = s["row"]; guard(r)
        qb.cell(row=r, column=COLS["E"]).value = s["E"]
        if s["H"] is not None:
            qb.cell(row=r, column=COLS["H"]).value = s["H"]
        qb.cell(row=r, column=COLS["I"]).value = s["I"]
        qb.cell(row=r, column=COLS["K"]).value = s["K"]
        qb.cell(row=r, column=COLS["L"]).value = s["L"]
        # corrections never touch the numeric cells
        assert qb.cell(row=r, column=COLS["D"]).value is None
        assert qb.cell(row=r, column=COLS["F"]).value is None
    print(f"component B: {len(CORRECTIONS)} rows corrected")

    # ---- invariants ---------------------------------------------------------
    assert [qb.cell(row=csu_row, column=c).value for c in range(1, 14)] == csu_before, \
        "Colorado State was modified"
    print("Colorado State confirmed untouched")
    ttu = RESYNC["TTU"]["row"]
    assert qb.cell(row=ttu, column=COLS["D"]).value is None
    assert qb.cell(row=ttu, column=COLS["F"]).value is None
    assert qb.cell(row=ttu, column=COLS["H"]).value == "H"
    print("Texas Tech medical gate preserved: H confidence, D/F blank")

    # ---- banner -------------------------------------------------------------
    banner = sh["A1"].value
    assert "v0.8.3 AUTHORITATIVE" in banner
    assert "65 H / 41 M / 32 L" in banner, f"unexpected banner census: {banner[:200]}"
    # Live sheet measured at 69 H / 41 M / 28 L (the owner-quoted 68/42/28 predates
    # Texas Tech's own 2026-08-21 M->H identity correction). NIU M->L then gives:
    sh["A1"].value = (banner.replace("v0.8.3 AUTHORITATIVE", "v0.8.4 AUTHORITATIVE")
                            .replace("65 H / 41 M / 32 L", "69 H / 40 M / 29 L"))
    assert "v0.8.3" not in sh["A1"].value
    print("banner updated: version identifier + confidence census")

    tmp = OUT + ".building.xlsx"
    wb.save(tmp)
    os.replace(tmp, OUT)
    print(f"written: {OUT}")
    print(f"v0.8.4 SHA-256: {sha256(OUT)}")
    assert sha256(SRC) == FROZEN_V083_SHA, "v0.8.3 was modified by the build"
    print("v0.8.3 confirmed still frozen")


if __name__ == "__main__":
    sys.exit(main())

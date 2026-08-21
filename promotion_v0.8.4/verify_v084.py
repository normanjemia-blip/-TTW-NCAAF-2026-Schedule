#!/usr/bin/env python3
"""v0.8.4 promotion certificate — READ-ONLY. Writes nothing.

v0.8.4 = v0.8.3 + 45 input/note cells:
    A. RESYNC   5 QB rows transcribed verbatim from the live production master
                (UNC, STAN, MOST, SJSU promoted; TTU raised to H, gate held)
    B. CORRECT  2 QB rows (NIU M->L + candidate list; TULN candidate list)
    plus START HERE!A1 version + confidence census.

Zero formula changes. Colorado State untouched. Texas Tech medical gate held.

Exit code 0 iff every check passes.
"""
import collections, hashlib, os, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
V083 = os.path.join(ROOT, "promotion_v0.8.3",
                    "TTW_College_Football_Power_Ratings_v0.8.3_AUTHORITATIVE.xlsx")
V084 = os.path.join(ROOT, "promotion_v0.8.4",
                    "TTW_College_Football_Power_Ratings_v0.8.4_AUTHORITATIVE.xlsx")
FROZEN_V083_SHA = "ff55782586ef1adb662eba59710e824dc382769a24579e48917b101fbcdd96b8"

RESYNC_ROWS = {52: "TTU", 65: "UNC", 68: "STAN", 101: "MOST", 124: "SJSU"}
CORRECT_ROWS = {91: "TULN", 123: "NIU"}
ALL_ROWS = {**RESYNC_ROWS, **CORRECT_ROWS}

PASS, FAIL = [], []


def chk(msg, ok, detail=""):
    (PASS if ok else FAIL).append(msg)
    print(f"  [{'PASS' if ok else 'FAIL'}] {msg}" + (f" - {detail}" if detail else ""))


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


def isf(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def norm(v):
    if isinstance(v, ArrayFormula):
        return ("F", v.text)
    if isinstance(v, str) and v.startswith("="):
        return ("F", v)
    return ("V", v)


def mean(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def main():
    print("=" * 78)
    print("v0.8.4 PROMOTION CERTIFICATE — NIU/Tulane packet + live resync")
    print("=" * 78)

    print("\n1. PREDECESSOR AND IDENTITY")
    h83, h84 = sha256(V083), sha256(V084)
    print(f"  v0.8.3 SHA-256: {h83}")
    print(f"  v0.8.4 SHA-256: {h84}")
    chk("1.1 v0.8.3 retains its frozen SHA-256", h83 == FROZEN_V083_SHA, h83[:16])
    chk("1.2 v0.8.4 differs from v0.8.3", h84 != h83)

    a = openpyxl.load_workbook(V083)
    b = openpyxl.load_workbook(V084)
    chk("1.3 21 sheets, identical names and order",
        a.sheetnames == b.sheetnames and len(b.sheetnames) == 21, f"{len(b.sheetnames)}")
    chk("1.4 sheet visibility preserved",
        {s: a[s].sheet_state for s in a.sheetnames} == {s: b[s].sheet_state for s in b.sheetnames})
    chk("1.5 named ranges preserved",
        sorted(a.defined_names.keys()) == sorted(b.defined_names.keys()))

    changed, formula_changed = [], []
    for s in a.sheetnames:
        wa, wbk = a[s], b[s]
        for r in range(1, max(wa.max_row, wbk.max_row) + 1):
            for c in range(1, max(wa.max_column, wbk.max_column) + 1):
                x, y = norm(wa.cell(row=r, column=c).value), norm(wbk.cell(row=r, column=c).value)
                if x != y:
                    changed.append((s, r, wbk.cell(row=r, column=c).coordinate))
                    if x[0] == "F" or y[0] == "F":
                        formula_changed.append((s, wbk.cell(row=r, column=c).coordinate))
    chk("1.6 ZERO formula differences", not formula_changed, str(formula_changed))
    bysheet = collections.Counter(s for s, _, _ in changed)
    chk("1.7 only START HERE (1) and QB VALUES (44) changed",
        dict(bysheet) == {"START HERE": 1, "QB VALUES": 44}, str(dict(bysheet)))
    touched = {r for s, r, _ in changed if s == "QB VALUES"}
    chk("1.8 exactly the 7 authorized QB rows were touched",
        touched == set(ALL_ROWS), f"touched={sorted(touched)} expected={sorted(ALL_ROWS)}")
    chk("1.9 total changed cells = 45", len(changed) == 45, str(len(changed)))

    tm, qb, st, ml, sh = b["TEAM MAP"], b["QB VALUES"], b["SETTINGS"], b["MARKET LINES"], b["START HERE"]
    season = st["B3"].value

    print("\n2. ROW IDENTITY AND PROTECTED CELLS")
    for row, ab in sorted(ALL_ROWS.items()):
        found = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == ab]
        ok = found == [row]
        prot = isf(qb.cell(row=row, column=7).value) and isf(qb.cell(row=row, column=13).value)
        j = qb.cell(row=row, column=10).value
        chk(f"2.x {ab} at row {row}; G/M formulas intact; J=2026",
            ok and prot and j == 2026, f"rows={found} G/M={prot} J={j}")

    print("\n3. CENSUSES")
    codes, sts, nonzero = collections.Counter(), collections.Counter(), []
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F, H, J = (qb.cell(row=r, column=c).value for c in (4, 6, 8, 10))
        G = "" if (D is None or F is None) else F - D
        codes[H] += 1
        sts["UNCERTAIN" if (G == "" or H == "L" or J != season) else "OK"] += 1
        if (D not in (None, 0)) or (F not in (None, 0)):
            nonzero.append(ab)
    chk("3.1 QB status census = 104 OK / 34 UNCERTAIN",
        (sts["OK"], sts["UNCERTAIN"]) == (104, 34), str(dict(sts)))
    chk("3.2 confidence census = 69 H / 40 M / 29 L",
        (codes["H"], codes["M"], codes["L"]) == (69, 40, 29), str(dict(codes)))
    chk("3.3 nonzero QB values = 0", not nonzero, str(nonzero))
    banner = sh["A1"].value
    chk("3.4 banner states v0.8.4 AUTHORITATIVE", "v0.8.4 AUTHORITATIVE" in banner)
    chk("3.5 banner census matches 69 H / 40 M / 29 L", "69 H / 40 M / 29 L" in banner)
    chk("3.6 banner carries no stale v0.8.3 identifier", "v0.8.3" not in banner)

    print("\n4. PACKET-SPECIFIC ASSERTIONS")
    def rec(ab):
        r = [x for x in range(6, 144) if tm.cell(row=x, column=1).value == ab][0]
        C, D, E, F, H = (qb.cell(row=r, column=c).value for c in (3, 4, 5, 6, 8))
        G = "" if (D is None or F is None) else F - D
        M = "UNCERTAIN" if (G == "" or H == "L" or qb.cell(row=r, column=10).value != season) else "OK"
        return dict(row=r, C=C, D=D, E=E, F=F, H=H, G=G, M=M)

    n = rec("NIU")
    chk("4.1 NIU confidence downgraded M -> L", n["H"] == "L", repr(n["H"]))
    chk("4.2 NIU candidate list restored to the four-way",
        n["E"] == "Open (Davidson / Macon / Hamric / Dickens)", repr(n["E"]))
    chk("4.3 NIU values stay blank; status UNCERTAIN",
        n["D"] is None and n["F"] is None and n["M"] == "UNCERTAIN")
    t = rec("TULN")
    chk("4.4 Tulane candidate list corrected to the four-way",
        t["E"] == "Open (Semonza / Chriss-Gremillion / Johnson / Bruno)", repr(t["E"]))
    chk("4.5 Tulane confidence UNCHANGED at L", t["H"] == "L", repr(t["H"]))
    chk("4.6 Tulane values stay blank; status UNCERTAIN",
        t["D"] is None and t["F"] is None and t["M"] == "UNCERTAIN")
    tt = rec("TTU")
    chk("4.7 Texas Tech identity confidence H", tt["H"] == "H", repr(tt["H"]))
    chk("4.8 Texas Tech QB1 named in baseline AND active",
        tt["C"] == "Will Hammond" and tt["E"] == "Will Hammond")
    chk("4.9 Texas Tech MEDICAL GATE HELD: D/F blank, status UNCERTAIN",
        tt["D"] is None and tt["F"] is None and tt["M"] == "UNCERTAIN")
    # Colorado State must be byte-identical to v0.8.3
    csu = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == "CSU"][0]
    same = all(norm(a["QB VALUES"].cell(row=csu, column=c).value) ==
               norm(qb.cell(row=csu, column=c).value) for c in range(1, 14))
    chk("4.10 Colorado State row is completely unchanged", same, f"row {csu}")
    for ab, want in (("UNC", "M"), ("STAN", "H"), ("MOST", "H"), ("SJSU", "H")):
        r = rec(ab)
        chk(f"4.x {ab} resynced: {want}, 0/0, status OK",
            r["H"] == want and r["D"] == 0 and r["F"] == 0 and r["G"] == 0 and r["M"] == "OK",
            f"H={r['H']} D={r['D']} F={r['F']} status={r['M']}")

    print("\n5. MODEL OUTPUTS AND MARKET ROWS")
    ps = b["PRESEASON"]
    S = {f"B{r}": st.cell(row=r, column=2).value for r in range(3, 33)}
    raw = {c: [ps.cell(row=r, column=c).value for r in range(6, 144)] for c in (4, 8, 17)}
    mu = {c: mean(v) for c, v in raw.items()}
    W = {4: S["B28"], 8: S["B29"], 17: S["B31"]}
    prior = {}
    for i, r in enumerate(range(6, 144)):
        ab = tm.cell(row=r, column=1).value
        num = den = 0.0
        for c in (4, 8, 17):
            v = raw[c][i]
            if isinstance(v, (int, float)):
                num += W[c] * (v - mu[c]); den += W[c]
        prior[ab] = (num / den) if den else ""
    alias = {}
    for r in range(6, 606):
        k, v = tm.cell(row=r, column=11).value, tm.cell(row=r, column=12).value
        if k and v:
            alias[str(k).strip()] = str(v).strip()
    delta = {}
    for r in range(6, 144):
        ab = tm.cell(row=r, column=1).value
        if not ab:
            continue
        D, F = qb.cell(row=r, column=4).value, qb.cell(row=r, column=6).value
        delta[ab] = "" if (D is None or F is None) else F - D
    z = lambda v: 0 if v == "" else v
    sch = b["IMPORT SCHEDULE"]
    got = {}
    n_games = fcs = 0
    for r in range(6, 1006):
        gid = sch.cell(row=r, column=1).value
        if not gid:
            continue
        n_games += 1
        A = alias.get(str(sch.cell(row=r, column=6).value).strip(), "")
        H = alias.get(str(sch.cell(row=r, column=8).value).strip(), "")
        if not A or not H:
            fcs += 1
            continue
        neutral = bool(sch.cell(row=r, column=5).value)
        m = (prior[H] - prior[A]) + (S["B7"] if neutral else S["B6"]) \
            + (z(delta.get(H, "")) - z(delta.get(A, "")))
        got[(A, H)] = m
    chk("5.1 888 games / 761 FBS-v-FBS / 127 FCS",
        n_games == 888 and fcs == 127 and n_games - fcs == 761, f"{n_games}/{n_games-fcs}/{fcs}")
    for (A, H), want in ((("MEM", "UNLV"), "UNLV -5.6"), (("UNC", "TCU"), "TCU -4.2"),
                         (("NMSU", "FSU"), "FSU -27.7"), (("SJSU", "USC"), "USC -35.2"),
                         (("HAW", "STAN"), "STAN -3.7")):
        m = got[(A, H)]
        lab = f"{H} -{abs(m):.1f}" if m > 0 else f"{A} -{abs(m):.1f}"
        chk(f"5.x {A} at {H} model spread unchanged at {want}", lab == want, lab)
    lines = sum(1 for r in range(6, 1006)
                if ml.cell(row=r, column=1).value is not None
                or ml.cell(row=r, column=4).value is not None)
    chk("5.2 repository artifact ships with MARKET LINES blank", lines == 0, str(lines))

    print("\n6. SETTINGS, SAFEGUARDS AND AUDIT")
    chk("6.1 BET toggle remains N", st["B11"].value == "N", repr(st["B11"].value))
    chk("6.2 totals remain unavailable (B22/B23 blank)",
        st["B22"].value is None and st["B23"].value is None)
    for t_, row in (("NDSU", 122), ("SAC", 114)):
        rr = [r for r in range(6, 144) if tm.cell(row=r, column=1).value == t_][0]
        chk(f"6.3 {t_} remains transitional", tm.cell(row=rr, column=5).value == "Y", f"row {rr}")
    au = b["AUDIT"]
    chk("6.4 AUDIT market-line invariant is the v0.8.3 operational formula",
        "SUMPRODUCT" in str(au["B16"].value) and "REMOVE TEST LINES" not in str(au["B16"].value))
    abbrevs = [tm.cell(row=r, column=1).value for r in range(6, 144)]
    filled = [x for x in abbrevs if x]

    def dupes(seq):
        vals = [str(x).strip().lower() for x in seq if x not in (None, "")]
        cnt = collections.Counter(vals)
        return sum(1 for v in cnt.values() if v > 1)

    manual = [tm.cell(row=r, column=8).value for r in range(6, 482)]
    imports = [tm.cell(row=r, column=11).value for r in range(6, 482)]
    at = collections.defaultdict(set)
    for r in range(6, 482):
        h, i = tm.cell(row=r, column=8).value, tm.cell(row=r, column=9).value
        if h not in (None, ""):
            at[str(h).strip().lower()].add(str(i).strip() if i else "")
    inv = {
        "138 teams in master list": len(filled) == 138,
        "No duplicate abbreviations": dupes(filled) == 0,
        "No duplicate manual aliases": dupes(manual) == 0,
        "No ambiguous manual alias": len({k: v for k, v in at.items() if len(v) > 1}) == 0,
        "No duplicate import aliases": dupes(imports) == 0,
        "BET toggle default OFF": st["B11"].value == "N",
        "Thresholds 1.0 / 1.5 / 3.0": (S["B8"], S["B9"], S["B10"]) == (1, 1.5, 3),
        "Default HFA 2.5 / Neutral 0": (S["B6"], S["B7"]) == (2.5, 0),
        "Movement cap 2.5": S["B12"] == 2.5,
        "No market lines entered": lines == 0,
    }
    failing = [k for k, v in inv.items() if not v]
    for k, v in inv.items():
        print(f"      {'OK  ' if v else 'FAIL'}  {k}")
    chk("6.5 AUDIT structural invariants: zero failing", not failing, str(failing))

    print("\n" + "=" * 78)
    print(f"RESULT: {len(PASS)} passed, {len(FAIL)} failed")
    print(f"v0.8.3 SHA-256 (frozen): {h83}")
    print(f"v0.8.4 SHA-256:          {h84}")
    print("=" * 78)
    for f in FAIL:
        print("  FAIL " + f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())

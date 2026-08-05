"""Phase 7E.1 regression — v0.8.0 -> v0.8.1 documentation maintenance patch.

Asserts the patch touched exactly one cell and nothing else, and that engine
behaviour and workbook structure are unchanged.
"""
import openpyxl, hashlib, zipfile, csv, collections, re, sys
from openpyxl.worksheet.formula import ArrayFormula

ROOT="/home/user/-TTW-NCAAF-2026-Schedule"
V80=f"{ROOT}/promotion_v0.8.0/TTW_College_Football_Power_Ratings_v0.8.0_AUTHORITATIVE.xlsx"
V81=f"{ROOT}/promotion_v0.8.1/TTW_College_Football_Power_Ratings_v0.8.1_AUTHORITATIVE.xlsx"
AUTH=f"{ROOT}/workbook_v0.6.2_deliverables/TTW_NCAAF_Power_Ratings_2026_v0.6.2_AUTHORITATIVE.xlsx"
OUT=[]; FAIL=[]
def say(s): print(s); OUT.append(s)
def chk(n,ok,d=""):
    say(f"  [{'PASS' if ok else 'FAIL'}] {n}{(' - '+d) if d else ''}")
    if not ok: FAIL.append(n)
def isf(v): return isinstance(v,ArrayFormula) or (isinstance(v,str) and v.startswith("="))
def ftext(v): return v.text if isinstance(v,ArrayFormula) else v
def sha(p): return hashlib.sha256(open(p,"rb").read()).hexdigest()

say("="*78); say("PHASE 7E.1 REGRESSION — v0.8.0 -> v0.8.1 (documentation maintenance)")
say("="*78)

w0=openpyxl.load_workbook(V80); w1=openpyxl.load_workbook(V81)
wa=openpyxl.load_workbook(AUTH)

# ---- 1. formula count ----
say("\n1. FORMULA COUNT")
def fcount(wb): return {s.title: sum(1 for r in s.iter_rows() for c in r if isf(c.value)) for s in wb.worksheets}
f0,f1,fa = fcount(w0), fcount(w1), fcount(wa)
t0,t1,ta = sum(f0.values()), sum(f1.values()), sum(fa.values())
say(f"  v0.6.2 {ta} | v0.8.0 {t0} | v0.8.1 {t1}")
chk("1.1 total formula count unchanged v0.8.0 -> v0.8.1", t0==t1, f"{t0} vs {t1}")
chk("1.2 total is the invariant 123011", t1==123011, str(t1))
chk("1.3 unchanged vs v0.6.2 AUTHORITATIVE", t1==ta, f"{ta} vs {t1}")
chk("1.4 per-sheet formula counts unchanged", f0==f1,
    str({k:(f0.get(k),f1.get(k)) for k in f1 if f0.get(k)!=f1.get(k)}))

# ---- 2. only START HERE changed ----
say("\n2. SCOPE OF CHANGE")
chk("2.1 sheet names and order identical", w0.sheetnames==w1.sheetnames)
vis0={s.title:s.sheet_state for s in w0.worksheets}; vis1={s.title:s.sheet_state for s in w1.worksheets}
chk("2.2 worksheet visibility identical", vis0==vis1)
changed=[]
for n in w1.sheetnames:
    s0,s1=w0[n],w1[n]
    for r in range(1, max(s0.max_row,s1.max_row)+1):
        for c in range(1, max(s0.max_column,s1.max_column)+1):
            a=ftext(s0.cell(row=r,column=c).value); b=ftext(s1.cell(row=r,column=c).value)
            if a!=b: changed.append((n, s1.cell(row=r,column=c).coordinate, a, b))
say(f"  changed cells: {len(changed)}")
for n,addr,a,b in changed:
    say(f"    {n}!{addr}")
    say(f"      OLD: {a}")
    say(f"      NEW: {b}")
bysheet=collections.Counter(c[0] for c in changed)
chk("2.3 EXACTLY ONE cell changed in the entire workbook", len(changed)==1, str(len(changed)))
chk("2.4 the only sheet changed is START HERE", set(bysheet)=={"START HERE"}, str(sorted(bysheet)))
chk("2.5 the only cell changed is START HERE!A1",
    len(changed)==1 and changed[0][0]=="START HERE" and changed[0][1]=="A1",
    str([(c[0],c[1]) for c in changed]))
chk("2.6 no formula cell changed", not any(isf(w0[n][a].value) or isf(w1[n][a].value) for n,a,_,_ in changed))
# raw XML parts
def parts(p):
    with zipfile.ZipFile(p) as z: return {n: hashlib.sha256(z.read(n)).hexdigest() for n in sorted(z.namelist())}
p0,p1=parts(V80),parts(V81)
dp=[n for n in p1 if p0.get(n)!=p1[n]]
ws_dp=[n for n in dp if n.startswith("xl/worksheets/")]
say(f"  differing zip parts: {dp}")
chk("2.7 exactly ONE worksheet XML part differs", len(ws_dp)==1, str(ws_dp))
chk("2.8 20 of 21 worksheet XML parts byte-identical",
    len([n for n in p1 if n.startswith('xl/worksheets/')])-len(ws_dp)==20,
    str(len(ws_dp)))

# ---- 3. banner content ----
say("\n3. BANNER CONTENT")
banner=w1["START HERE"]["A1"].value
say(f"  {banner}")
for bad in ("NOT PROMOTED","NOT AUTHORITATIVE","awaiting owner approval","NOT AUTHORITATIVE, NOT PROMOTED"):
    chk(f"3.x banner no longer contains {bad!r}", bad not in banner)
chk("3.1 banner contains 'v0.8.1 AUTHORITATIVE'", "v0.8.1 AUTHORITATIVE" in banner)
chk("3.2 banner states promotion complete", "promotion complete" in banner)
chk("3.3 banner states 73 Tier-1 records", "73 Tier-1" in banner)
chk("3.4 banner no longer says 74 Tier-1", "74 Tier-1" not in banner)
chk("3.5 banner no longer references v0.7.9", "v0.7.9" not in banner)
chk("3.6 banner is text, not a formula", not isf(w1["START HERE"]["A1"].value))

# ---- 4. engine + data unchanged ----
say("\n4. ENGINE OUTPUTS AND WORKBOOK BEHAVIOUR")
for sheet,cells in (("ENGINE",("M7","AE7","AI7")),("QB VALUES",("G6","M6")),
                    ("TEAM RATINGS",("C6","D6")),("CALC",("Q7","S7"))):
    for c in cells:
        try:
            chk(f"4.x {sheet}!{c} formula identical to v0.6.2",
                ftext(wa[sheet][c].value)==ftext(w1[sheet][c].value))
        except Exception as e:
            chk(f"4.x {sheet}!{c} comparable", False, str(e))
s=w1["SETTINGS"]
chk("4.1 SETTINGS!B3 season = 2026", s["B3"].value==2026)
chk("4.2 SETTINGS!B6 HFA = 2.5", s["B6"].value==2.5)
chk("4.3 SETTINGS!B11 BET toggle = N", s["B11"].value=="N")
qb=w1["QB VALUES"]; tm=w1["TEAM MAP"]
codes=collections.Counter(qb.cell(row=r,column=8).value for r in range(6,144))
D=[qb.cell(row=r,column=4).value for r in range(6,144)]
F=[qb.cell(row=r,column=6).value for r in range(6,144)]
H=[qb.cell(row=r,column=8).value for r in range(6,144)]
J=[qb.cell(row=r,column=10).value for r in range(6,144)]
G=["" if (d is None or f is None) else f-d for d,f in zip(D,F)]
st=collections.Counter("UNCERTAIN" if (G[i]=="" or H[i]=="L" or J[i]!=2026) else "OK" for i in range(138))
say(f"  QB codes {dict(codes)} | status {dict(st)} | Tier-1 {sum(1 for h in H if h in ('M','L'))}")
chk("4.4 QB codes unchanged (65 H / 40 M / 33 L)",
    codes.get("H")==65 and codes.get("M")==40 and codes.get("L")==33, str(dict(codes)))
chk("4.5 QB status unchanged (99 OK / 39 UNCERTAIN)",
    st.get("OK")==99 and st.get("UNCERTAIN")==39, str(dict(st)))
chk("4.6 Tier-1 population is 73 (matches the new banner)",
    sum(1 for h in H if h in ("M","L"))==73)
chk("4.7 zero nonzero QB values",
    sum(1 for d,f in zip(D,F) if (d not in (None,0)) or (f not in (None,0)))==0)
chk("4.8 every L-coded team blank-gated",
    all(D[i] is None for i in range(138) if H[i]=="L"))
chk("4.9 138 unique teams", len({tm.cell(row=r,column=1).value for r in range(6,144)})==138)
ml=w1["MARKET LINES"]
nums=sum(1 for r in range(1,ml.max_row+1) for c in range(1,ml.max_column+1)
         if isinstance(ml.cell(row=r,column=c).value,(int,float)))
chk("4.10 still zero market spreads loaded (pristine preseason)", nums==0, str(nums))
d0={n:str(v.value) for n,v in w0.defined_names.items()}
d1={n:str(v.value) for n,v in w1.defined_names.items()}
chk("4.11 defined names identical", d0==d1)

# ---- 5. manifest ----
say("\n5. MANIFEST")
h80,h81=sha(V80),sha(V81)
say(f"  v0.8.0 SHA-256: {h80}")
say(f"  v0.8.1 SHA-256: {h81}")
chk("5.1 v0.8.0 unchanged by this patch",
    h80=="661f8ab0e6120290d4ffd8d4ddac738d7e19d7bd0bbcf69bc9df51fb3cef97c7", h80[:16])
chk("5.2 v0.8.1 differs from v0.8.0 (the patch landed)", h80!=h81)
chk("5.3 v0.6.2 rollback target untouched",
    sha(AUTH)=="bbb17b50fbfb728bea2a23d3d20771935cc61e238313a054473aafe1ca838efd")

with open(f"{ROOT}/promotion_v0.8.1/diff_v080_to_v081.csv","w",newline="") as fh:
    w=csv.writer(fh); w.writerow(["sheet","cell","classification","old","new"])
    for n,a,o,nv in changed: w.writerow([n,a,"DOCUMENTATION — BANNER TEXT",o,nv])

say("\n"+"#"*78)
say(f"# RESULT: {len(FAIL)} FAILURES")
for f in FAIL: say(f"#   - {f}")
say(f"# v0.8.1 SHA-256: {h81}")
say("#"*78)
open(f"{ROOT}/promotion_v0.8.1/regression_log_v081.txt","w").write("\n".join(OUT)+"\n")
sys.exit(1 if FAIL else 0)

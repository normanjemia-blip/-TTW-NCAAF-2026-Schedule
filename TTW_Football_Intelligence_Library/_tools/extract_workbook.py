#!/usr/bin/env python3
"""
TTW Football Intelligence Library — Phase 6 workbook read
==========================================================

Reads the TTW College Football Power Ratings Workbook v0.8.1 AUTHORITATIVE
**read-only**, and writes what it found to _source/verified/. The workbook
itself is never opened for writing, never re-saved, and never copied into
the library's tracked tree.

Two facts about the file shape the whole of Phase 6 and are recorded here
rather than discovered later:

  * The workbook stores **no cached formula results.** It was written
    programmatically and has not been recalculated by a spreadsheet
    application, so `TEAM RATINGS!EFFECTIVE RATING` and every other
    computed cell reads as empty. No TTW rating can therefore be *read*
    out of the file. Anything numeric on the TTW side of a comparison has
    to be *derived* by reimplementing the workbook's own printed formulas,
    and must be labelled as derived wherever it appears.

  * The PRESEASON sheet does store literal source inputs with dates and
    citation URLs. Those are read verbatim.

Source of the file: git blob 06d817cd on branch
claude/2026-ncaaf-schedule-build-by6j5n, path
promotion_v0.8.1/TTW_College_Football_Power_Ratings_v0.8.1_AUTHORITATIVE.xlsx

Usage:  python3 _tools/extract_workbook.py <path-to-xlsx>
"""

import hashlib
import json
import sys

import openpyxl

FIRST, LAST = 6, 143            # data rows on TEAM MAP / PRESEASON
COLS = {"sp_raw": 4, "sp_date": 5, "sp_cite": 6,
        "fpi_raw": 8, "fpi_date": 9, "fpi_cite": 10,
        "ttw25_raw": 12, "ttw_date": 14, "ttw_cite": 15,
        "tr_raw": 17, "tr_date": 18, "tr_cite": 19,
        "vsin_raw": 21, "vsin_date": 22, "vsin_cite": 23}


def cell(ws, row, col):
    v = ws.cell(row, col).value
    if isinstance(v, str) and v.startswith("="):
        return None                      # a formula, not a stored value
    return v


def main():
    path = sys.argv[1]
    digest = hashlib.sha256(open(path, "rb").read()).hexdigest()
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)

    tm, ps, st = wb["TEAM MAP"], wb["PRESEASON"], wb["SETTINGS"]

    settings = {}
    for r in range(1, 60):
        k, v = st.cell(r, 1).value, st.cell(r, 2).value
        if isinstance(k, str) and v is not None:
            settings[k.strip()] = v

    rows = []
    for i in range(FIRST, LAST + 1):
        rec = {"abbrev": tm.cell(i, 1).value, "team": tm.cell(i, 2).value,
               "conference": tm.cell(i, 3).value, "status": tm.cell(i, 4).value}
        for name, col in COLS.items():
            v = cell(ps, i, col)
            rec[name] = v.isoformat()[:10] if hasattr(v, "isoformat") else v
        rows.append(rec)

    # What the file does and does not hold, stated rather than assumed.
    tr = wb["TEAM RATINGS"]
    cached = sum(1 for i in range(FIRST, LAST + 1)
                 if cell(tr, i, 15) is not None)     # EFFECTIVE RATING
    out = {
        "source_file": "TTW_College_Football_Power_Ratings_v0.8.1_AUTHORITATIVE.xlsx",
        "git_blob": "06d817cdaa2814aa71630c5637d90af978c17b98",
        "git_branch": "claude/2026-ncaaf-schedule-build-by6j5n",
        "git_path": "promotion_v0.8.1/",
        "sha256": digest,
        "read_only": True,
        "sheets": wb.sheetnames,
        "team_rows": len(rows),
        "cached_effective_ratings": cached,
        "cached_values_present": cached > 0,
        "preseason_source_weights": {
            k: v for k, v in settings.items()
            if k.startswith(("SP+ 2026", "FPI 2026", "TTW independent",
                             "TeamRankings", "VSiN"))},
        "coverage": {name: sum(1 for r in rows
                               if isinstance(r[name], (int, float)))
                     for name in ("sp_raw", "fpi_raw", "ttw25_raw", "tr_raw",
                                  "vsin_raw")},
        "settings": {k: (v.isoformat()[:10] if hasattr(v, "isoformat") else v)
                     for k, v in settings.items()},
        "rows": rows,
    }
    with open("_source/verified/workbook_preseason_v081.json", "w") as fh:
        json.dump(out, fh, indent=1, ensure_ascii=False)

    print(f"workbook sha256           {digest}")
    print(f"team rows                 {len(rows)}")
    print(f"cached effective ratings  {cached}  "
          f"({'values present' if cached else 'NONE — formulas only'})")
    for k, v in out["coverage"].items():
        print(f"  {k:<12} {v}/138 numeric")
    print("preseason weights:", out["preseason_source_weights"])


if __name__ == "__main__":
    main()

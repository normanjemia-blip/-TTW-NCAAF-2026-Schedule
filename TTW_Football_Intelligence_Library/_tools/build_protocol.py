#!/usr/bin/env python3
"""
TTW Football Intelligence Library — Phase 7 pre-registered test protocol
=========================================================================

The VSiN-inclusion question cannot be answered today: no season in this
project pairs a VSiN preseason rating with played games. It will become
answerable, game by game, from 22 August 2026.

The right response to "not yet" is not to invent a proxy. It is to write
the test down now, in code, while no result exists to tune it against —
which is the only moment at which a test can be genuinely pre-registered.

This tool does three things:

  1. Reads the workbook's 2026 game list (read-only) and computes, for
     every scheduled FBS-vs-FBS game, how far apart the BASELINE and
     VSIN-INCLUDED configurations actually predict. That difference is the
     entire effect under test.
  2. Runs a power analysis on it, so the owner learns *before* spending a
     season on the question whether the season could settle it at all.
  3. Emits the protocol document and the scorer that will execute it.

Nothing is written to the workbook.

Usage:  python3 _tools/build_protocol.py <workbook.xlsx>
"""

import json
import math
import os
import statistics as stats
import sys

import openpyxl

from qb_lib import ABBREV_TO_VSIN

OUT = "05_Power_Ratings"
HEADER = ("<!-- GENERATED FILE — do not hand-edit.\n"
          "     Rebuild:  python3 _tools/build_protocol.py <workbook.xlsx>\n"
          "     Source:   TTW Power Ratings Workbook v0.8.1 AUTHORITATIVE (read-only) -->\n")


def write(name, lines):
    with open(os.path.join(OUT, name), "w") as fh:
        fh.write(HEADER + "\n" + "\n".join(lines) + "\n")


def read_alias(wb):
    """The workbook's own IMPORT alias table: printed name -> abbreviation.

    Reused rather than reinvented. The schedule stores ESPN names
    ("Alabama Crimson Tide") while TEAM MAP stores short ones ("Alabama"),
    and this table is how the workbook itself bridges the two. It resolves
    exactly the 138 FBS programmes; FCS opponents fall through, which is
    the correct outcome since they have no rating to compare.
    """
    tm = wb["TEAM MAP"]
    alias = {}
    for i in range(6, 482):
        a, b = tm.cell(i, 11).value, tm.cell(i, 12).value
        if a and b and not str(a).startswith("=") and not str(b).startswith("="):
            alias[str(a).strip().lower()] = str(b).strip()
    return alias


def read_schedule(path):
    """The 2026 game list, read-only, straight off IMPORT SCHEDULE."""
    wb = openpyxl.load_workbook(path, data_only=False)
    alias = read_alias(wb)
    ws = wb["IMPORT SCHEDULE"]
    games = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[0] is None:
            continue
        games.append({"id": row[0], "season": row[1], "week": row[2],
                      "neutral": bool(row[4]), "away": row[5], "home": row[7],
                      "away_pts": row[9], "home_pts": row[10],
                      "completed": bool(row[11])})
    return games, alias


def main():
    path = sys.argv[1]
    diag = json.load(open("_source/calibration/vsin_diagnostics.json"))
    delta = {r["team"]: r["delta"] for r in diag["per_team"]}

    # Names resolve through the workbook's own IMPORT alias table and then
    # through the Phase 4 abbreviation bijection. Two existing, validated
    # mappings composed; no new name matching invented here.
    games, alias = read_schedule(path)

    def resolve(name):
        ab = alias.get(str(name).strip().lower())
        return ABBREV_TO_VSIN.get(ab) if ab else None

    played = [g for g in games if g["completed"]]
    fbs, unmatched = [], set()
    for g in games:
        h, a = resolve(g["home"]), resolve(g["away"])
        if h is None or a is None:
            for side, r in ((g["home"], h), (g["away"], a)):
                if r is None:
                    unmatched.add(side)
            continue
        fbs.append(dict(g, home_team=h, away_team=a,
                        dpred=delta[h] - delta[a]))

    d = [g["dpred"] for g in fbs]
    sd = stats.pstdev(d)
    mean_abs = stats.fmean(abs(x) for x in d)
    mx = max(abs(x) for x in d)

    # Power. The quantity a season could measure is the paired difference in
    # absolute error between the two configurations. Its per-game magnitude
    # is bounded by |dpred|, and its per-game noise is of the same order, so
    # SD(dpred) is the right scale for the standard error. Solving
    # n = (z * SD / effect)^2 for a two-sided 95% test gives the sample a
    # real effect of each size would need.
    # Per game the scored quantity is D = |error_baseline| - |error_vsin|.
    # The two predictions differ by only d = dpred, and the margin error is
    # an order of magnitude larger, so D is almost exactly d * sign(error).
    # Two consequences follow, pointing opposite ways:
    #
    #   * The huge common error cancels. The noise is of order RMS(d), not
    #     of order the margin error, so the paired design is far better
    #     powered than comparing two MAE figures would suggest.
    #   * The effect is bounded above by E|d|. Even if every VSiN
    #     adjustment moved a prediction the right way, the MAE improvement
    #     could not exceed that.
    #
    # The honest question is therefore not "can a season detect this" but
    # "is the largest arithmetically possible effect worth detecting".
    rms_d = math.sqrt(stats.fmean(x * x for x in d))

    def mde(n):
        return 1.96 * rms_d / math.sqrt(n)

    def n_needed(effect):
        return math.ceil((1.96 * rms_d / effect) ** 2)

    early = [g for g in fbs if isinstance(g["week"], int) and g["week"] <= 4]

    out = {"games_total": len(games), "games_completed": len(played),
           "fbs_vs_fbs_matched": len(fbs), "unmatched_labels": sorted(unmatched),
           "weeks_0_to_4": len(early),
           "dpred": {"sd": sd, "rms": rms_d, "mean_abs": mean_abs, "max_abs": mx,
                     "n_over_0_5": sum(1 for x in d if abs(x) >= 0.5),
                     "n_over_1_0": sum(1 for x in d if abs(x) >= 1.0)},
           "max_possible_mae_gain": mean_abs,
           "mde_full_season": mde(len(fbs)),
           "mde_weeks_0_to_4": mde(len(early)),
           "power": {f"{e}": n_needed(e) for e in (0.02, 0.05, 0.10, 0.15)}}
    with open("_source/calibration/power_analysis.json", "w") as fh:
        json.dump(out, fh, indent=1, ensure_ascii=False)

    DERIVED = ("> **Source class: TTW DERIVED.** Arithmetic over the "
               "workbook's stored schedule and preseason inputs. Not a "
               "workbook output.")
    FROZEN = ("> **v0.8.1 AUTHORITATIVE remains frozen.** This protocol reads "
              "it and changes nothing in it. Executing the protocol later "
              "will also change nothing in it.")

    L = ["# Pre-Registered Calibration Protocol — VSiN in the preseason blend",
         "", DERIVED, "", FROZEN, "",
         "> **Registered before any 2026 result exists.** As of the build "
         f"date, {out['games_completed']} of {out['games_total']} games in "
         "the workbook's 2026 schedule are complete. The comparison below is "
         "fixed now precisely so that it cannot later be shaped by the "
         "answer it produces.", "",
         "## What is being tested", "",
         "Two configurations, both fixed in advance, neither fitted to any "
         "2026 data:", "",
         "| | BASELINE | VSIN-INCLUDED |", "| --- | --- | --- |",
         "| Sources | SP+, FPI, TeamRankings | SP+, FPI, TeamRankings, VSiN |",
         "| VSiN weight | — | 0.10, the workbook's own reserved value |",
         "| Renormalisation | workbook rule, missing is never zero | same |",
         "| Parameters estimated from 2026 | **none** | **none** |", "",
         "Because neither configuration estimates anything from 2026, every "
         "2026 game is out-of-sample for both. That is what makes a "
         "single-season test legitimate here — the walk-forward property "
         "comes from the parameters being frozen, not from splitting the "
         "sample.", "",
         "## The decision point, and how look-ahead is prevented", "",
         "- The prediction for a game is formed from the **preseason prior "
         "only**, faded by the workbook's own effective-games rule, using "
         "information available before kickoff.",
         "- No closing line, no in-season rating, no result, and no "
         "post-game information may enter the prediction. The scorer refuses "
         "to read the result column until after the prediction column is "
         "written.",
         "- Market lines are recorded at the stated line date and are used "
         "as a **benchmark**, never as an input.",
         "- Games are scored in schedule order. Nothing is re-scored after "
         "later weeks are seen.", "",
         "## Metrics", "",
         "| Rank | Metric | Role |", "| --- | --- | --- |",
         "| Primary | Mean absolute error of predicted margin against actual "
         "margin | decides the question |",
         "| Primary | Root mean squared error of the same | penalises the "
         "large misses MAE hides |",
         "| Secondary | Paired per-game difference in absolute error, with a "
         "95% interval | the only form in which a small effect is legible |",
         "| Descriptive | Error against the closing spread | context only — "
         "**never a success criterion** |",
         "| Descriptive | ATS record at the workbook's own edge thresholds, "
         "with interval | reported, **never sufficient on its own** |",
         "| Diagnostic | Error by week bucket, by edge bucket, by favourite "
         "size | where any difference lives |", "",
         "Success is defined as a **reduction in out-of-sample prediction "
         "error against actual results**, significant at the stated interval "
         "and stable across week buckets. Agreement with closing lines, with "
         "another rating system, or with the current consensus is explicitly "
         "not success.", "",
         "## How large is the effect that could possibly be detected", "",
         "This is the part worth reading before committing a season to the "
         "question. The difference between the two configurations' "
         "predictions for a single game is the whole effect under test:", "",
         "| | Points |", "| --- | --- |",
         f"| Games matched (FBS vs FBS, 2026) | {out['fbs_vs_fbs_matched']} |",
         f"| Standard deviation of the per-game prediction difference | "
         f"**{sd:.3f}** |",
         f"| Root mean square of the same | {rms_d:.3f} |",
         f"| Mean absolute prediction difference | {mean_abs:.3f} |",
         f"| Largest prediction difference in the season | {mx:.3f} |",
         f"| Games where the two configurations differ by 0.5 pts or more | "
         f"{out['dpred']['n_over_0_5']} |",
         f"| Games where they differ by a full point or more | "
         f"{out['dpred']['n_over_1_0']} |", "",
         "### What the paired design can and cannot resolve", "",
         "The scored quantity per game is the difference in absolute error "
         "between the two configurations. Because both predictions sit "
         "within a fraction of a point of each other, the large shared "
         "margin error cancels almost exactly, and the comparison is far "
         "better powered than setting two MAE figures side by side would "
         "suggest.", "",
         "| | Points of MAE |", "| --- | --- |",
         f"| Largest improvement arithmetically possible | "
         f"**{out['max_possible_mae_gain']:.3f}** |",
         f"| Minimum detectable effect, full season "
         f"({out['fbs_vs_fbs_matched']} games) | {out['mde_full_season']:.3f} |",
         f"| Minimum detectable effect, weeks 0\u20134 "
         f"({out['weeks_0_to_4']} games) | {out['mde_weeks_0_to_4']:.3f} |", "",
         "| True improvement in MAE | Games needed (two-sided, 95%) |",
         "| --- | --- |"]
    for e, n in out["power"].items():
        feas = "" if int(n) <= out["fbs_vs_fbs_matched"] else " \u2014 **more than one season provides**"
        L.append(f"| {float(e):.2f} pts | {int(n):,}{feas} |")
    L += ["",
          "The upper bound is the number that matters. Even if **every** "
          "VSiN adjustment moved its prediction in the right direction \u2014 "
          "which no rating source achieves \u2014 the improvement would be "
          f"{out['max_possible_mae_gain']:.2f} points of mean absolute "
          "error, against margin errors that routinely run to double "
          "figures. A realistic share of that upper bound is smaller again.", "",
          "So the honest statement is not that the season is underpowered. "
          "It is that the season is **well powered to measure an effect that "
          "is very small by construction**. A statistically detectable result "
          "here would still be a practically negligible one, and the report "
          "must say both things rather than one.", "",
          "**This is a limitation stated before the test, not after it.** If "
          "the 2026 result comes back indistinguishable, that is the expected "
          "outcome rather than a disappointment, and it must not be met by "
          "widening the search until something reaches significance.", "",
          "## Stopping rules", "",
          "- The comparison is run **once**, at season end, on the full "
          "matched sample, plus the pre-declared week-bucket breakdown.",
          "- No mid-season peeking is used to decide whether to continue.",
          "- The alternative-weight sweep stays diagnostic. It is never used "
          "to select a production weight, and it is not re-run against the "
          "holdout to find a winner.",
          "- If BASELINE and VSIN-INCLUDED are statistically or practically "
          "indistinguishable, the recorded result is **no evidence of "
          "improvement**. That outcome is acceptable and final for the "
          "season.", "",
          "## What executing this protocol may change", "",
          "Nothing, by itself. The scorer writes a report. Any production "
          "change remains an owner decision taken separately, on the "
          "evidence the report contains.", "",
          "## Cross-links", "",
          "- [Diagnostics](01_VSIN_DIAGNOSTICS.md) · "
          "[Phase 7 report](01_PHASE7_REPORT.md) · "
          "[import candidate](00_VSIN_IMPORT_CANDIDATE.md)"]
    write("01_CALIBRATION_PROTOCOL.md", L)

    print(f"schedule read: {len(games)} games, {len(played)} completed, "
          f"{len(fbs)} FBS-vs-FBS matched, {len(unmatched)} unmatched labels")
    print(f"  dpred sd {sd:.4f}  mean|dpred| {mean_abs:.4f}  max {mx:.4f}")
    print(f"  weeks 0-4 games {out['weeks_0_to_4']}")
    for e, n in out["power"].items():
        print(f"  effect {e} pts -> n = {n:,}")
    return out


if __name__ == "__main__":
    main()

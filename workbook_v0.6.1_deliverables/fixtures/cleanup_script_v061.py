"""
Phase 6.1 cleanup script: reverses every synthetic fixture applied by
apply_to_real_workbook_v061.py, producing v0.6.1_test_cleaned.xlsx. Proof
step: this file must be cell-value-identical to the untouched
v0.6.1_working.xlsx (verified by a full cell-by-cell diff, not just a file
hash, since re-saving through openpyxl changes irrelevant zip/XML metadata
even when logical content is identical).

Reuses the proven-correct pattern from Phase 6's cleanup_script.py: only
MANUAL-INPUT columns are cleared (never a whole column range, never a
formula column); .value = None is used everywhere (NOT
.cell(value=None), which openpyxl treats as a no-op); and every
IMPORT SCHEDULE cell that had a genuine non-blank original value is
restored to that exact original value, not blanked.

IMPORTANT CORRECTION caught while building this script: the v0.6.1
fixture-application script's docstring described IMPORT SCHEDULE rows
20-24 as "previously-unused rows." That was wrong. Direct inspection of
the pristine v0.6.1_working.xlsx shows those 5 rows hold REAL production
schedule games (Arkansas-Pine Bluff@Missouri, Idaho@Utah,
Colorado@Georgia Tech, Eastern Illinois@Minnesota, UAB@Illinois) that the
NDSU synthetic-history fixture overwrote. This script restores those 5
rows to their exact original values (captured directly from
v0.6.1_working.xlsx below), not merely blanks them - a blank restoration
would have left 5 real games permanently missing from the "cleaned" file
and would have failed the 0-diff proof.
"""
import datetime
import openpyxl

SRC = "v0.6.1_test_applied.xlsx"
OUT = "v0.6.1_test_cleaned.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=False)

# --- SETTINGS: revert week/as-of/toggle to blank/default ---
st = wb["SETTINGS"]
st["B4"] = None
st["B5"] = None
st["B11"] = "N"

# --- MARKET LINES: clear only the manual-input columns (A,C,D,E,F,G,H) -
#     column B (Matchup, auto) was never written to and must survive ---
ml = wb["MARKET LINES"]
for row in (6, 8, 9, 10, 11, 14, 15):
    for c in (1, 3, 4, 5, 6, 7, 8):
        ml.cell(row=row, column=c).value = None

# --- ADJUSTMENTS: clear only A:I (1-9) on rows 6 AND 7 (v0.6.1 added a
#     second isolated-oversized-adjustment row) - J,K (Effective/Flags)
#     are formulas and were never written to ---
adj = wb["ADJUSTMENTS"]
for row in (6, 7):
    for c in range(1, 10):
        adj.cell(row=row, column=c).value = None

# --- QB VALUES: clear only C,D,E,F,H,I,J,K,L - A,B (formulas from TEAM
#     RATINGS), G (delta, formula), M (status, formula) were never written to ---
qb = wb["QB VALUES"]
for row in (64, 70, 97, 122):
    for c in (3, 4, 5, 6, 8, 9, 10, 11, 12):
        qb.cell(row=row, column=c).value = None

# --- IMPORT SCHEDULE: restore each cell's TRUE original value ---
sched = wb["IMPORT SCHEDULE"]

# Rows 8, 9, 14: only columns J,K,L,N (10,11,12,14) were written to by the
# completed-score fixture; L was FALSE (not blank) before, N carries a
# genuine ESPN source note on rows 8/9, blank on row 14.
ORIGINAL_SCORES = {
    8: {10: None, 11: None, 12: False,
        14: "ESPN source: season_type=2 (regular-season), week=1. Normalized "
            "to project Week 0 (opening slate before the 2026-09-03 Week 1 kickoff)."},
    9: {10: None, 11: None, 12: False,
        14: "ESPN source: season_type=2 (regular-season), week=1. Normalized "
            "to project Week 0 (opening slate before the 2026-09-03 Week 1 kickoff)."},
    14: {10: None, 11: None, 12: False, 14: None},
}
for row, cols in ORIGINAL_SCORES.items():
    for c, v in cols.items():
        sched.cell(row=row, column=c).value = v

# Rows 20-24: the NDSU synthetic-history fixture overwrote all 14 columns
# of 5 REAL production schedule games. Restore full original rows,
# captured directly from the pristine v0.6.1_working.xlsx.
ORIGINAL_ROWS_20_24 = {
    20: ['401856663', 2026, 1, datetime.datetime(2026, 9, 4, 0, 0), False,
         'Arkansas-Pine Bluff Golden Lions', 'Southwestern Athletic Conference',
         'Missouri Tigers', 'Southeastern Conference', None, None, False,
         'Memorial Stadium', None],
    21: ['401856768', 2026, 1, datetime.datetime(2026, 9, 4, 0, 0), False,
         'Idaho Vandals', 'Big Sky Conference',
         'Utah Utes', 'Big 12 Conference', None, None, False,
         'Rice-Eccles Stadium', None],
    22: ['401856776', 2026, 1, datetime.datetime(2026, 9, 4, 0, 0), False,
         'Colorado Buffaloes', 'Big 12 Conference',
         'Georgia Tech Yellow Jackets', 'Atlantic Coast Conference', None, None, False,
         'Bobby Dodd Stadium', None],
    23: ['401858422', 2026, 1, datetime.datetime(2026, 9, 4, 0, 0), False,
         'Eastern Illinois Panthers', 'OVC-Big South Association',
         'Minnesota Golden Gophers', 'Big Ten Conference', None, None, False,
         'Huntington Bank Stadium', None],
    24: ['401858424', 2026, 1, datetime.datetime(2026, 9, 4, 0, 0), False,
         'UAB Blazers', 'American Conference',
         'Illinois Fighting Illini', 'Big Ten Conference', None, None, False,
         'Gies Memorial Stadium', None],
}
for row, vals in ORIGINAL_ROWS_20_24.items():
    for i, v in enumerate(vals, start=1):
        sched.cell(row=row, column=i).value = v

# --- IMPORT STATS: clear the 3 test rows (all columns were genuinely blank before) ---
ist = wb["IMPORT STATS"]
for row in (6, 7, 8):
    for c in range(1, 10):
        ist.cell(row=row, column=c).value = None

# --- HISTORY: clear the 1 test row (genuinely blank before - HISTORY had
#     no snapshots yet, preseason) ---
hist = wb["HISTORY"]
for c in range(1, 5):
    hist.cell(row=6, column=c).value = None

# --- START HERE: restore banner to the approved v0.6.1 (pending-approval) text ---
sh = wb["START HERE"]
sh["A1"] = "TO THE WINDOW — NCAAF POWER RATINGS 2026 (v0.6.1 PHASE 6 CORRECTION + TEST-COMPLETION BUILD — PENDING APPROVAL)"

wb.save(OUT)
print("Saved", OUT)

# --- Self-check: zero TEST_TAG cells remain ---
TEST_TAG = "TEST ONLY — NOT REAL DATA"
tagged = []
for name in wb.sheetnames:
    for row in wb[name].iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and TEST_TAG in cell.value:
                tagged.append((name, cell.coordinate))
print(f"Remaining TEST_TAG cells: {len(tagged)}", tagged[:5])
